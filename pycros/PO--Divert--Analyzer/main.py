from __future__ import annotations

import os
import re
import threading
from collections import OrderedDict, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from typing import Callable, DefaultDict, Dict, Iterable, List, Optional, Tuple

try:
    from PySide6.QtCore import Qt, Signal
    from PySide6.QtWidgets import (
        QFileDialog,
        QHBoxLayout,
        QLabel,
        QSizePolicy,
        QTextEdit,
        QVBoxLayout,
        QWidget,
    )
    from qfluentwidgets import MessageBox, PrimaryPushButton
except Exception:  # pragma: no cover - allows processor-only CLI/test imports without PySide6
    class _DummySignal:
        def __init__(self, *_args, **_kwargs):
            pass

        def connect(self, *_args, **_kwargs):
            pass

        def emit(self, *_args, **_kwargs):
            pass

    class _DummyWidget:
        def __init__(self, *_args, **_kwargs):
            pass

    class _DummyQt:
        AlignLeft = 0
        AlignTop = 0
        AlignVCenter = 0
        TextSelectableByMouse = 0

    Qt = _DummyQt()
    Signal = _DummySignal
    QWidget = _DummyWidget
    QFileDialog = QHBoxLayout = QLabel = QSizePolicy = QTextEdit = QVBoxLayout = _DummyWidget
    MessageBox = PrimaryPushButton = _DummyWidget

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import pdfplumber
except Exception:  # pragma: no cover - handled at runtime with clear message
    pdfplumber = None


# -----------------------------------------------------------------------------
# Pycro UI
# -----------------------------------------------------------------------------


class MainWidget(QWidget):
    log_message = Signal(str)
    processing_done = Signal(int, int, str)

    def __init__(self):
        super().__init__()
        self.setObjectName("po_divert_analyzer_widget")
        self._build_ui()
        self._connect_signals()

    def _build_ui(self):
        self.desc_label = QLabel("", self)
        self.desc_label.setWordWrap(True)
        self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.desc_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.desc_label.setStyleSheet(
            "color: #dcdcdc; background: transparent; padding: 6px; "
            "border: 1px solid #3a3a3a; border-radius: 6px;"
        )
        self.set_long_description("")

        self.select_search_btn = PrimaryPushButton("Select PO Search Results Excel", self)
        self.select_pdf_btn = PrimaryPushButton("Select Original PO PDF(s)", self)
        self.run_btn = PrimaryPushButton("Run", self)

        self.files_label = QLabel("Selected files", self)
        self.files_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.files_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        self.logs_label = QLabel("Process logs", self)
        self.logs_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.logs_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        self.files_box = QTextEdit(self)
        self.files_box.setReadOnly(True)
        self.files_box.setPlaceholderText("Selected Excel and PDF files will appear here")
        self.files_box.setStyleSheet(
            "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText("Live process log will appear here")
        self.log_box.setStyleSheet(
            "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)

        main_layout.addWidget(self.desc_label, 1)

        button_row_1 = QHBoxLayout()
        button_row_1.addStretch(1)
        button_row_1.addWidget(self.select_search_btn, 1)
        button_row_1.addWidget(self.select_pdf_btn, 1)
        button_row_1.addStretch(1)
        main_layout.addLayout(button_row_1, 0)

        button_row_2 = QHBoxLayout()
        button_row_2.addStretch(1)
        button_row_2.addWidget(self.run_btn, 1)
        button_row_2.addStretch(1)
        main_layout.addLayout(button_row_2, 0)

        label_row = QHBoxLayout()
        label_row.addWidget(self.files_label, 1)
        label_row.addWidget(self.logs_label, 1)
        main_layout.addLayout(label_row, 0)

        body_row = QHBoxLayout()
        body_row.addWidget(self.files_box, 1)
        body_row.addWidget(self.log_box, 1)
        main_layout.addLayout(body_row, 4)

        self.search_results_path = ""
        self.pdf_paths: List[str] = []

    def set_long_description(self, text: str):
        clean = (text or "").strip()
        if clean:
            self.desc_label.setText(clean)
            self.desc_label.show()
        else:
            self.desc_label.clear()
            self.desc_label.hide()

    def _connect_signals(self):
        self.select_search_btn.clicked.connect(self.select_search_results)
        self.select_pdf_btn.clicked.connect(self.select_pdfs)
        self.run_btn.clicked.connect(self.run_process)
        self.log_message.connect(self.append_log)
        self.processing_done.connect(self.on_processing_done)

    def select_search_results(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select PO Search Results Excel",
            "",
            "Excel files (*.xlsx *.xlsm);;All files (*.*)",
        )
        if path:
            self.search_results_path = path
            self._refresh_files_box()

    def select_pdfs(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select Original PO PDF(s)",
            "",
            "PDF files (*.pdf);;All files (*.*)",
        )
        if files:
            self.pdf_paths = list(files)
            self._refresh_files_box()

    def _refresh_files_box(self):
        lines: List[str] = []
        if self.search_results_path:
            lines.append("[PO Search Results Excel]")
            lines.append(self.search_results_path)
        if self.pdf_paths:
            lines.append("")
            lines.append("[Original PO PDF(s)]")
            lines.extend(self.pdf_paths)
        self.files_box.setPlainText("\n".join(lines))

    def run_process(self):
        if not self.search_results_path:
            MessageBox("Warning", "Please select the PO search-results Excel file.", self).exec()
            return
        if not self.pdf_paths:
            MessageBox("Warning", "Please select at least one original PO PDF.", self).exec()
            return

        self.log_box.clear()
        self.log_message.emit("Process starts")
        self.run_btn.setEnabled(False)
        self.select_search_btn.setEnabled(False)
        self.select_pdf_btn.setEnabled(False)

        def worker():
            ok, fail, out_path = 0, 0, ""
            try:
                out_path = process_files(self.search_results_path, self.pdf_paths, self.log_message.emit)
                ok = 1
            except Exception as exc:
                fail = 1
                self.log_message.emit(f"ERROR: {exc}")
            self.processing_done.emit(ok, fail, out_path)

        threading.Thread(target=worker, daemon=True).start()

    def append_log(self, text: str):
        self.log_box.append(text)
        self.log_box.ensureCursorVisible()

    def on_processing_done(self, ok: int, fail: int, out_path: str):
        if out_path:
            self.log_message.emit(f"Output workbook saved to: {out_path}")
        self.log_message.emit(f"Completed: {ok} success, {fail} failed.")
        self.run_btn.setEnabled(True)
        self.select_search_btn.setEnabled(True)
        self.select_pdf_btn.setEnabled(True)

        title = "Processing complete" if fail == 0 else "Processing finished with issues"
        lines = [f"Success: {ok}", f"Failed: {fail}"]
        if out_path:
            lines.append(f"Output: {os.path.basename(out_path)}")
        else:
            lines.append("Output: (see logs)")
        msg = MessageBox(title, "\n".join(lines), self)
        msg.yesButton.setText("OK")
        msg.cancelButton.hide()
        msg.exec()


def get_widget():
    return MainWidget()


# -----------------------------------------------------------------------------
# Processing model
# -----------------------------------------------------------------------------


SIZE_HEADERS: List[object] = [
    28, 30, 32, 34, 36, 38, 40, 42,
    "CUST1", "CUST2", "CUST3", "CUST4", "CUST5", "CUST6", "CUST7", "CUSTM",
    "S+", "M+", "L+", "XL+", "0X", "1X", "2X", "3X", "4X",
    "1SIZE", "2SIZE", "3SIZE", "2XSS", "XSS", "SS", "MS", "LS", "XLS", "2XLS", "3XLS", "4XLS",
    "2XS", "XXS", "XS", "S", "M", "L", "XL", "2XL", "3XL", "4XL", "5XL",
    "2XS-T", "XS-T", "S-T", "M-T", "L-T", "XL-T", "2XL-T", "3XL-T", "4XL-T", "5XL-T",
    "XSTT", "STT", "MTT", "LTT", "XLTT", "2XLTT", "3XLTT", "4XLTT", "5XLTT",
]

NEW_START_COL = 1
ORI_START_COL = 72
NOW_START_COL = 143
SECTION_WIDTH = 70

REQUIRED_SEARCH_HEADERS = [
    "Purchase Order Number",
    "PO Line Item Number",
    "Size Description",
    "Size Quantity",
    "Item Text",
]

DIVERT_TO_RE = re.compile(
    r"Quantity\s+([\d,]+(?:\.\d+)?)\s+Diverted\s+to\s+Purchase\s+Order\s+(\d+)\s+Line\s+Item\s+(\d+)",
    re.IGNORECASE,
)

DIVERT_FROM_RE = re.compile(
    r"Quantity\s+([\d,]+(?:\.\d+)?)\s+Diverted\s+from\s+Purchase\s+Order\s+(\d+)\s+on\s+(\d+)",
    re.IGNORECASE,
)

PDF_SIZE_RE = re.compile(
    r"^EA\s+[\d,]+(?:\.\d+)?\s+USD\s+(?P<size>[A-Za-z0-9+\-]+)\s+(?P<qty>[\d,]+)\s+",
    re.IGNORECASE,
)

PDF_TOTAL_RE = re.compile(r"ITEM\s+TOTAL\s+IN\s+USD\s+([\d,]+)", re.IGNORECASE)
PDF_PO_RE = re.compile(r"\b(\d{10})\s*/\s*(\d{10})\b")


@dataclass
class POGroup:
    po: int
    line: int
    sizes: "OrderedDict[str, int]" = field(default_factory=OrderedDict)
    total_item_quantity: int = 0
    item_text: str = ""
    moved_to: str = ""
    first_row_index: int = 0

    def add_size(self, size: object, qty: object):
        size_text = normalize_size(size)
        if not size_text:
            return
        qty_int = to_int(qty)
        if size_text not in self.sizes:
            self.sizes[size_text] = 0
        self.sizes[size_text] += qty_int

    @property
    def total_from_sizes(self) -> int:
        return sum(self.sizes.values())


@dataclass
class DivertEvent:
    source_po: int
    source_line: int
    target_po: int
    target_line: int
    target_ordinal: int
    qty: int
    source_row_order: int
    raw_target_code: str


@dataclass
class DivertFromEvent:
    source_po: int
    source_line: int
    source_ordinal: int
    qty: int
    event_order: int
    raw_source_code: str


@dataclass
class TargetResult:
    target_key: Tuple[int, int]
    source_keys: List[Tuple[int, int]]
    allocations: Dict[Tuple[int, int], Dict[str, int]]
    residuals: Dict[str, int]


class LogCollector:
    def __init__(self, emit: Optional[Callable[[str], None]] = None):
        self.emit = emit or (lambda _msg: None)
        self.lines: List[str] = []

    def __call__(self, message: str):
        self.lines.append(message)
        try:
            self.emit(message)
        except Exception:
            pass


# -----------------------------------------------------------------------------
# Public process entry point
# -----------------------------------------------------------------------------


def process_files(search_results_path: str, pdf_paths: List[str], log_emit: Optional[Callable[[str], None]] = None) -> str:
    logger = LogCollector(log_emit)
    processor = PODivertProcessor(logger)
    return processor.process(search_results_path, pdf_paths)


class PODivertProcessor:
    def __init__(self, log: LogCollector):
        self.log = log
        self.warnings: List[str] = []

    def process(self, search_results_path: str, pdf_paths: List[str]) -> str:
        self._require_files(search_results_path, pdf_paths)

        self.log("Reading PO search-results workbook...")
        search_groups, events = read_search_results(search_results_path, self.log)
        self.log(f"Loaded {len(search_groups)} PO/line groups from search results.")
        self.log(f"Parsed {len(events)} diverted-to item-text events from search results.")

        self.log("Reading original PO PDF(s)...")
        original_groups: Dict[Tuple[int, int], OrderedDict[str, int]] = {}
        for pdf_path in pdf_paths:
            pdf_groups = read_original_pdf(pdf_path, self.log)
            original_groups.update(pdf_groups)
            self.log(f"Loaded {len(pdf_groups)} original PO line groups from {os.path.basename(pdf_path)}.")

        target_results = build_target_results(search_groups, original_groups, events, self.log)
        if not target_results:
            raise ValueError("No diverted target PO lines were found. Check the Item Text column and selected files.")
        self.log(f"Built {len(target_results)} diverted target block(s).")

        out_path = proposed_output_path(search_results_path)
        self.log("Writing output workbook...")
        write_output_workbook(out_path, target_results, search_groups, original_groups, self.log.lines)
        self.log(f"Saved output workbook: {out_path}")
        return out_path

    @staticmethod
    def _require_files(search_results_path: str, pdf_paths: List[str]):
        if not search_results_path or not os.path.isfile(search_results_path):
            raise FileNotFoundError("PO search-results Excel file was not found.")
        if not pdf_paths:
            raise FileNotFoundError("No original PO PDF was selected.")
        missing = [p for p in pdf_paths if not os.path.isfile(p)]
        if missing:
            raise FileNotFoundError("Missing PDF file(s): " + ", ".join(missing))


# -----------------------------------------------------------------------------
# Search-results reader
# -----------------------------------------------------------------------------


def read_search_results(path: str, log: Optional[Callable[[str], None]] = None) -> Tuple[Dict[Tuple[int, int], POGroup], List[DivertEvent]]:
    logger = log or (lambda _msg: None)
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        groups: Dict[Tuple[int, int], POGroup] = {}
        row_order_by_key: Dict[Tuple[int, int], int] = {}
        processed_sheets: List[str] = []
        skipped_sheets: List[str] = []
        global_row_order = 0

        for ws in wb.worksheets:
            try:
                header_row = next(ws.iter_rows(min_row=1, max_row=1))
            except StopIteration:
                skipped_sheets.append(f"{ws.title} (empty)")
                continue

            headers = [cell.value for cell in header_row]
            header_map = build_header_map(headers)
            missing = [h for h in REQUIRED_SEARCH_HEADERS if canonical_header(h) not in header_map]
            if missing:
                skipped_sheets.append(f"{ws.title} (missing: {', '.join(missing)})")
                continue

            processed_sheets.append(ws.title)
            sheet_data_rows = 0

            for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                po = to_optional_int(get_by_header(row, header_map, "Purchase Order Number"))
                line = to_optional_int(get_by_header(row, header_map, "PO Line Item Number"))
                if po is None or line is None:
                    continue

                global_row_order += 1
                sheet_data_rows += 1

                key = (po, line)
                if key not in groups:
                    groups[key] = POGroup(po=po, line=line, first_row_index=global_row_order)
                    row_order_by_key[key] = global_row_order

                group = groups[key]
                group.add_size(
                    get_by_header(row, header_map, "Size Description"),
                    get_by_header(row, header_map, "Size Quantity"),
                )
                total = to_optional_int(get_by_header(row, header_map, "Total Item Quantity"))
                if total is not None:
                    group.total_item_quantity = total
                item_text = clean_text(get_by_header(row, header_map, "Item Text"))
                if item_text and not group.item_text:
                    group.item_text = item_text
                moved_to = clean_text(get_by_header(row, header_map, "Moved To"))
                if moved_to and not group.moved_to:
                    group.moved_to = moved_to

            logger(f"Read worksheet '{ws.title}': {sheet_data_rows} PO search-result row(s).")

        if not processed_sheets:
            detail = "; ".join(skipped_sheets) if skipped_sheets else "No worksheets were found."
            raise ValueError(
                "No worksheet with the required PO search-results headers was found. "
                f"Checked: {detail}"
            )

        if skipped_sheets:
            logger("Skipped worksheet(s) without the required PO search-results headers: " + "; ".join(skipped_sheets))
        logger("Processed PO search-results worksheet(s): " + ", ".join(processed_sheets))

        events: List[DivertEvent] = []
        for key in sorted(groups, key=lambda k: row_order_by_key.get(k, 10**9)):
            group = groups[key]
            for qty, target_po, target_code in parse_divert_to_events(group.item_text):
                target_line, target_ordinal = split_line_item_code(target_code)
                events.append(
                    DivertEvent(
                        source_po=group.po,
                        source_line=group.line,
                        target_po=target_po,
                        target_line=target_line,
                        target_ordinal=target_ordinal,
                        qty=qty,
                        source_row_order=group.first_row_index,
                        raw_target_code=target_code,
                    )
                )
        return groups, events
    finally:
        wb.close()

def build_header_map(headers: List[object]) -> Dict[str, int]:
    result: Dict[str, int] = {}
    for idx, header in enumerate(headers):
        canon = canonical_header(header)
        if canon and canon not in result:
            result[canon] = idx
    return result


def canonical_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def get_by_header(row: Tuple[object, ...], header_map: Dict[str, int], name: str) -> object:
    idx = header_map.get(canonical_header(name))
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def parse_divert_to_events(text: str) -> Iterable[Tuple[int, int, str]]:
    if not text:
        return []
    results = []
    for match in DIVERT_TO_RE.finditer(text):
        qty = to_int(match.group(1))
        po = int(match.group(2))
        code = match.group(3)
        results.append((qty, po, code))
    return results


def parse_divert_from_events(text: str) -> Iterable[Tuple[int, int, str]]:
    if not text:
        return []
    results = []
    for match in DIVERT_FROM_RE.finditer(text):
        qty = to_int(match.group(1))
        source_po = int(match.group(2))
        source_code = match.group(3)
        results.append((qty, source_po, source_code))
    return results


def parse_divert_from_records(text: str) -> List[DivertFromEvent]:
    if not text:
        return []
    results: List[DivertFromEvent] = []
    for event_order, match in enumerate(DIVERT_FROM_RE.finditer(text)):
        qty = to_int(match.group(1))
        source_po = int(match.group(2))
        source_code = match.group(3)
        source_line, source_ordinal = split_line_item_code(source_code)
        results.append(
            DivertFromEvent(
                source_po=source_po,
                source_line=source_line,
                source_ordinal=source_ordinal,
                qty=qty,
                event_order=event_order,
                raw_source_code=source_code,
            )
        )
    return results


# -----------------------------------------------------------------------------
# PDF reader
# -----------------------------------------------------------------------------


def read_original_pdf(path: str, log: Optional[Callable[[str], None]] = None) -> Dict[Tuple[int, int], OrderedDict[str, int]]:
    if pdfplumber is None:
        raise RuntimeError("pdfplumber is required to read original PO PDFs. Install the pycro requirements first.")

    logger = log or (lambda _msg: None)
    groups: Dict[Tuple[int, int], OrderedDict[str, int]] = {}

    with pdfplumber.open(path) as pdf:
        current_po: Optional[int] = None
        current_line: Optional[int] = None
        current_sizes: Optional[OrderedDict[str, int]] = None

        for page in pdf.pages:
            text = page.extract_text(x_tolerance=2, y_tolerance=3) or ""
            page_po = extract_pdf_po_number(text)
            if page_po is not None:
                current_po = page_po

            for raw_line in text.splitlines():
                line = normalize_pdf_line(raw_line)
                item_info = parse_pdf_item_line(line)
                if item_info is not None:
                    if current_po is None:
                        raise ValueError(f"Could not determine PO number before item line in {os.path.basename(path)}.")
                    item_line = item_info
                    current_line = item_line
                    current_sizes = OrderedDict()
                    groups[(current_po, current_line)] = current_sizes
                    continue

                size_info = parse_pdf_size_line(line)
                if size_info is not None and current_po is not None and current_line is not None and current_sizes is not None:
                    size, qty = size_info
                    current_sizes[size] = current_sizes.get(size, 0) + qty
                    continue

                total_info = parse_pdf_total_line(line)
                if total_info is not None and current_line is not None and current_sizes is not None:
                    actual_total = sum(current_sizes.values())
                    if actual_total != total_info:
                        logger(
                            f"Warning: PDF item {current_line} total mismatch on page {page.page_number}: "
                            f"sizes sum {actual_total}, PDF total {total_info}."
                        )
                    current_line = None
                    current_sizes = None

    if not groups:
        raise ValueError(f"No item size rows were extracted from original PDF: {os.path.basename(path)}")
    return groups


def extract_pdf_po_number(text: str) -> Optional[int]:
    match = PDF_PO_RE.search(text or "")
    if not match:
        return None
    # Nike PDF header is commonly "Trading Co PO / Purchase Order".
    return int(match.group(2))


def normalize_pdf_line(line: str) -> str:
    return re.sub(r"\s+", " ", (line or "").strip())


def parse_pdf_item_line(line: str) -> Optional[int]:
    if not line:
        return None
    parts = line.split()
    if not parts:
        return None

    item_token = ""
    if re.fullmatch(r"\d{5}", parts[0]):
        item_token = parts[0]
    elif len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
        joined = parts[0] + parts[1]
        if re.fullmatch(r"\d{5}", joined):
            item_token = joined

    if not item_token:
        return None

    # Require a material-like token after the item number to avoid false positives in terms pages.
    remaining = " ".join(parts[1:])
    if not re.search(r"\b[A-Z0-9]{2,}\d{2,}[-A-Z0-9]*\b", remaining):
        return None
    return int(item_token)


def parse_pdf_size_line(line: str) -> Optional[Tuple[str, int]]:
    match = PDF_SIZE_RE.search(line or "")
    if not match:
        return None
    return normalize_size(match.group("size")), to_int(match.group("qty"))


def parse_pdf_total_line(line: str) -> Optional[int]:
    match = PDF_TOTAL_RE.search(line or "")
    if not match:
        return None
    return to_int(match.group(1))


# -----------------------------------------------------------------------------
# Reconciliation
# -----------------------------------------------------------------------------


def build_allocations_from_divert_from_records(
    target_group: POGroup,
    records: List[DivertFromEvent],
    log: Optional[Callable[[str], None]] = None,
) -> Optional[TargetResult]:
    """
    Some SAP search-result rows only show the complete source breakdown in the
    target row's "Diverted from" text.  Those records do not reliably expose the
    target size ordinal, so we assign the quantities by solving them against the
    target line's final size quantities.

    Example:
        target XS/S/M/L/XL = 48/36/168/444/504
        records = 15,20,13,36,117,51,137,136,171,91,102,311

    The solver groups those records into exact target-size totals:
        XS = 15+20+13, S = 36, M = 117+51, L = 137+136+171, XL = 91+102+311

    If the records cannot exactly fill the target sizes, return None so the
    older direct "Diverted to" logic can still run as a fallback.
    """
    logger = log or (lambda _msg: None)
    usable_records = [record for record in records if record.qty > 0]
    if not usable_records:
        return None

    target_sizes = OrderedDict((size, qty) for size, qty in target_group.sizes.items() if qty)
    if not target_sizes:
        return None

    assignments = solve_event_partition(usable_records, target_sizes)
    if assignments is None:
        return None

    allocations: DefaultDict[Tuple[int, int], DefaultDict[str, int]] = defaultdict(lambda: defaultdict(int))
    used_indices: set[int] = set()
    for size, record_indices in assignments.items():
        for record_index in record_indices:
            record = usable_records[record_index]
            source_key = (record.source_po, record.source_line)
            allocations[source_key][size] += record.qty
            used_indices.add(record_index)

    unused_records = [record for idx, record in enumerate(usable_records) if idx not in used_indices]
    for record in unused_records:
        logger(
            f"Skipped unused diverted-from event: {record.qty} from "
            f"{record.source_po}-{record.source_line} on {record.raw_source_code} "
            f"for target {target_group.po}-{target_group.line}."
        )

    source_keys = sorted(allocations.keys(), key=lambda k: (k[0], k[1]))
    return TargetResult(
        target_key=(target_group.po, target_group.line),
        source_keys=source_keys,
        allocations={key: dict(value) for key, value in allocations.items()},
        residuals={},
    )


def solve_event_partition(
    records: List[DivertFromEvent],
    target_sizes: "OrderedDict[str, int]",
) -> Optional["OrderedDict[str, Tuple[int, ...]]"]:
    size_targets = [(size, qty) for size, qty in target_sizes.items() if qty > 0]
    total_target = sum(qty for _size, qty in size_targets)
    total_records = sum(record.qty for record in records)
    if total_records < total_target:
        return None

    all_indices = tuple(range(len(records)))
    failed_states: set[Tuple[int, Tuple[int, ...]]] = set()

    def solve_size(size_index: int, remaining: Tuple[int, ...]) -> Optional[List[Tuple[str, Tuple[int, ...]]]]:
        state = (size_index, remaining)
        if state in failed_states:
            return None

        if size_index >= len(size_targets):
            return []

        size, required_qty = size_targets[size_index]
        for subset in candidate_record_subsets(records, remaining, required_qty):
            subset_set = set(subset)
            next_remaining = tuple(idx for idx in remaining if idx not in subset_set)
            tail = solve_size(size_index + 1, next_remaining)
            if tail is not None:
                return [(size, subset)] + tail

        failed_states.add(state)
        return None

    solved = solve_size(0, all_indices)
    if solved is None:
        return None
    return OrderedDict(solved)


def candidate_record_subsets(
    records: List[DivertFromEvent],
    remaining: Tuple[int, ...],
    target_qty: int,
    max_candidates: int = 2500,
) -> List[Tuple[int, ...]]:
    if target_qty <= 0:
        return [tuple()]

    ordered_indices = sorted(remaining, key=lambda idx: (-records[idx].qty, records[idx].event_order, idx))
    suffix_sums = [0] * (len(ordered_indices) + 1)
    for pos in range(len(ordered_indices) - 1, -1, -1):
        suffix_sums[pos] = suffix_sums[pos + 1] + records[ordered_indices[pos]].qty

    candidates: List[Tuple[int, ...]] = []

    def dfs(pos: int, current_qty: int, chosen: List[int]) -> bool:
        if current_qty == target_qty:
            candidates.append(tuple(sorted(chosen)))
            return len(candidates) >= max_candidates
        if current_qty > target_qty:
            return False
        if pos >= len(ordered_indices):
            return False
        if current_qty + suffix_sums[pos] < target_qty:
            return False

        previous_qty_at_level: Optional[int] = None
        for next_pos in range(pos, len(ordered_indices)):
            record_index = ordered_indices[next_pos]
            qty = records[record_index].qty
            if previous_qty_at_level == qty:
                continue
            previous_qty_at_level = qty
            if current_qty + qty > target_qty:
                continue
            chosen.append(record_index)
            should_stop = dfs(next_pos + 1, current_qty + qty, chosen)
            chosen.pop()
            if should_stop:
                return True
        return False

    dfs(0, 0, [])
    candidates.sort(key=lambda subset: (len(subset), [records[idx].event_order for idx in subset]))
    return candidates


def build_target_results(
    search_groups: Dict[Tuple[int, int], POGroup],
    original_groups: Dict[Tuple[int, int], OrderedDict[str, int]],
    events: List[DivertEvent],
    log: Optional[Callable[[str], None]] = None,
) -> List[TargetResult]:
    logger = log or (lambda _msg: None)

    target_keys = sorted({(event.target_po, event.target_line) for event in events if (event.target_po, event.target_line) in search_groups})

    # Fallback: include lines whose own Item Text says "Diverted from" even if source rows did not parse.
    for key, group in search_groups.items():
        if any(True for _ in parse_divert_from_events(group.item_text)) and key not in target_keys:
            target_keys.append(key)
    target_keys = sorted(set(target_keys), key=lambda k: (k[0], k[1]))

    source_delta_remaining: DefaultDict[Tuple[int, int], DefaultDict[str, int]] = defaultdict(lambda: defaultdict(int))

    for key in sorted(search_groups):
        ori_sizes = original_groups.get(key, OrderedDict())
        now_sizes = search_groups[key].sizes
        for size in SIZE_HEADERS_AS_TEXT:
            delta = ori_sizes.get(size, 0) - now_sizes.get(size, 0)
            if delta > 0:
                source_delta_remaining[key][size] = delta

    events_by_target: DefaultDict[Tuple[int, int], List[DivertEvent]] = defaultdict(list)
    for event in sorted(events, key=lambda e: (e.source_row_order, e.source_line, e.target_line, e.target_ordinal)):
        events_by_target[(event.target_po, event.target_line)].append(event)

    results: List[TargetResult] = []
    for target_key in target_keys:
        target_group = search_groups.get(target_key)
        if target_group is None:
            continue

        divert_from_records = parse_divert_from_records(target_group.item_text)
        from_result = build_allocations_from_divert_from_records(target_group, divert_from_records, logger)
        if from_result is not None:
            results.append(from_result)
            continue
        if divert_from_records:
            logger(
                f"Warning: target {target_key[0]}-{target_key[1]} diverted-from records could not be "
                "matched exactly to target sizes. Falling back to diverted-to/source-delta matching."
            )

        target_size_order = list(target_group.sizes.keys())
        target_remaining: Dict[str, int] = dict(target_group.sizes)
        allocations: DefaultDict[Tuple[int, int], DefaultDict[str, int]] = defaultdict(lambda: defaultdict(int))
        referenced_sources: List[Tuple[int, int]] = []

        for event in events_by_target.get(target_key, []):
            source_key = (event.source_po, event.source_line)
            if source_key not in referenced_sources:
                referenced_sources.append(source_key)

            if event.target_ordinal < 1 or event.target_ordinal > len(target_size_order):
                logger(
                    f"Warning: target line item {event.raw_target_code} has size ordinal {event.target_ordinal}, "
                    f"but target {target_key[0]}-{target_key[1]} has only {len(target_size_order)} size row(s)."
                )
                continue

            size = target_size_order[event.target_ordinal - 1]
            by_target_remaining = max(0, target_remaining.get(size, 0))
            by_source_remaining = max(0, source_delta_remaining[source_key].get(size, 0))
            qty_to_allocate = min(event.qty, by_target_remaining, by_source_remaining)

            if qty_to_allocate <= 0:
                if event.qty > 0:
                    logger(
                        f"Skipped stale or capped event: {event.qty} from {event.source_po}-{event.source_line} "
                        f"to {event.target_po}-{event.target_line} size {size}."
                    )
                continue

            allocations[source_key][size] += qty_to_allocate
            target_remaining[size] = by_target_remaining - qty_to_allocate
            source_delta_remaining[source_key][size] = by_source_remaining - qty_to_allocate

        # If source rows were not parsed from "Diverted to", fall back to target "Diverted from" source keys.
        if not referenced_sources:
            for _qty, source_po, source_code in parse_divert_from_events(target_group.item_text):
                source_line, _source_ordinal = split_line_item_code(source_code)
                source_key = (source_po, source_line)
                if source_key not in referenced_sources:
                    referenced_sources.append(source_key)

        source_keys = [key for key in sorted(referenced_sources, key=lambda k: (k[0], k[1])) if key in search_groups or key in original_groups]
        residuals = {size: qty for size, qty in target_remaining.items() if qty}
        if residuals:
            logger(
                f"Warning: target {target_key[0]}-{target_key[1]} has unallocated residuals: "
                + ", ".join(f"{size}={qty}" for size, qty in residuals.items())
            )

        results.append(
            TargetResult(
                target_key=target_key,
                source_keys=source_keys,
                allocations={k: dict(v) for k, v in allocations.items()},
                residuals=residuals,
            )
        )

    return results


# The header values are mixed strings/integers for Excel, but all data keys use strings.
SIZE_HEADERS_AS_TEXT = [str(h).strip().upper() for h in SIZE_HEADERS]


# -----------------------------------------------------------------------------
# Workbook writer
# -----------------------------------------------------------------------------


def write_output_workbook(
    out_path: str,
    target_results: List[TargetResult],
    search_groups: Dict[Tuple[int, int], POGroup],
    original_groups: Dict[Tuple[int, int], OrderedDict[str, int]],
    log_lines: List[str],
):
    report_size_headers = determine_report_size_headers(target_results, search_groups, original_groups)
    report_size_keys = [str(header).strip().upper() for header in report_size_headers]
    section_starts, section_width = build_report_layout(report_size_headers)

    wb = Workbook()
    ws = wb.active
    ws.title = "Chart (2)"

    write_report_headers(ws, section_starts, report_size_headers)
    current_row = 3
    for result in target_results:
        target_group = search_groups[result.target_key]
        target_row = current_row
        write_new_section_row(ws, target_row, result.target_key, target_group.sizes, section_starts, report_size_keys)

        first_source_row = current_row + 1
        for offset, source_key in enumerate(result.source_keys):
            row = first_source_row + offset
            write_new_section_row(ws, row, source_key, result.allocations.get(source_key, {}), section_starts, report_size_keys)
            write_ori_now_rows(ws, row, source_key, original_groups, search_groups, section_starts, report_size_keys)

        check_row = first_source_row + len(result.source_keys)
        write_check_row(ws, check_row, target_row, first_source_row, check_row - 1, section_starts, len(report_size_headers))
        current_row = check_row + 2

    apply_report_formatting(ws, max_row=max(2, current_row - 1), section_starts=section_starts, section_width=section_width)
    write_log_sheet(wb, log_lines)
    wb.save(out_path)


def determine_report_size_headers(
    target_results: List[TargetResult],
    search_groups: Dict[Tuple[int, int], POGroup],
    original_groups: Dict[Tuple[int, int], OrderedDict[str, int]],
) -> List[object]:
    used_sizes: set[str] = set()

    def mark(size_values: Dict[str, int]):
        for size, qty in (size_values or {}).items():
            if to_int(qty) != 0:
                used_sizes.add(normalize_size(size))

    for result in target_results:
        target_group = search_groups.get(result.target_key)
        if target_group is not None:
            mark(target_group.sizes)
        mark(result.residuals)
        for source_key in result.source_keys:
            mark(result.allocations.get(source_key, {}))
            mark(original_groups.get(source_key, {}))
            now_group = search_groups.get(source_key)
            if now_group is not None:
                mark(now_group.sizes)

    headers = [header for header in SIZE_HEADERS if str(header).strip().upper() in used_sizes]
    extra_sizes = sorted(size for size in used_sizes if size not in {str(header).strip().upper() for header in SIZE_HEADERS})
    headers.extend(extra_sizes)
    return headers or ["1SIZE"]


def build_report_layout(size_headers: List[object]) -> Tuple[Dict[str, int], int]:
    section_width = 2 + len(size_headers) + 1
    section_starts = {
        "NEW": 1,
        "ORI": 1 + section_width + 1,
        "NOW": 1 + (section_width + 1) * 2,
    }
    return section_starts, section_width


def write_report_headers(ws, section_starts: Dict[str, int], size_headers: List[object]):
    sections = [
        (section_starts["NEW"], "Size Breakdown (NEW)"),
        (section_starts["ORI"], "Size Breakdown (ORI)"),
        (section_starts["NOW"], "Size Breakdown (NOW)"),
    ]
    for start_col, title in sections:
        ws.cell(row=1, column=start_col + 2, value=title)
        ws.cell(row=2, column=start_col, value="Divert PO#")
        ws.cell(row=2, column=start_col + 1, value="PO line")
        for idx, header in enumerate(size_headers, start=start_col + 2):
            ws.cell(row=2, column=idx, value=header)
        ws.cell(row=2, column=start_col + 2 + len(size_headers), value="Total")


def write_new_section_row(
    ws,
    row: int,
    key: Tuple[int, int],
    size_values: Dict[str, int],
    section_starts: Dict[str, int],
    size_headers_as_text: List[str],
):
    write_section_row(ws, row, section_starts["NEW"], key, size_values, size_headers_as_text)


def write_ori_now_rows(
    ws,
    row: int,
    source_key: Tuple[int, int],
    original_groups: Dict[Tuple[int, int], OrderedDict[str, int]],
    search_groups: Dict[Tuple[int, int], POGroup],
    section_starts: Dict[str, int],
    size_headers_as_text: List[str],
):
    write_section_row(ws, row, section_starts["ORI"], source_key, original_groups.get(source_key, {}), size_headers_as_text)
    now_group = search_groups.get(source_key)
    now_values = now_group.sizes if now_group else {}
    write_section_row(ws, row, section_starts["NOW"], source_key, now_values, size_headers_as_text)


def write_section_row(
    ws,
    row: int,
    start_col: int,
    key: Tuple[int, int],
    size_values: Dict[str, int],
    size_headers_as_text: List[str],
):
    ws.cell(row=row, column=start_col, value=key[0])
    ws.cell(row=row, column=start_col + 1, value=key[1])
    size_start = start_col + 2
    total_col = start_col + 2 + len(size_headers_as_text)

    normalized_values = {normalize_size(size): to_int(qty) for size, qty in (size_values or {}).items()}
    for offset, header in enumerate(size_headers_as_text):
        value = normalized_values.get(header, 0)
        if value != 0:
            ws.cell(row=row, column=size_start + offset, value=value)
        elif header in normalized_values:
            ws.cell(row=row, column=size_start + offset, value=0)

    first_size_letter = get_column_letter(size_start)
    last_size_letter = get_column_letter(total_col - 1)
    ws.cell(row=row, column=total_col, value=f"=SUM({first_size_letter}{row}:{last_size_letter}{row})")


def write_check_row(
    ws,
    row: int,
    target_row: int,
    first_source_row: int,
    last_source_row: int,
    section_starts: Dict[str, int],
    size_count: int,
):
    size_start = section_starts["NEW"] + 2
    total_col = section_starts["NEW"] + 2 + size_count
    if first_source_row > last_source_row:
        for col in range(size_start, total_col + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=row, column=col, value=f"={col_letter}{target_row}")
        return

    for col in range(size_start, total_col + 1):
        col_letter = get_column_letter(col)
        source_refs = "-".join(f"{col_letter}{source_row}" for source_row in range(first_source_row, last_source_row + 1))
        ws.cell(row=row, column=col, value=f"={col_letter}{target_row}-{source_refs}")


def apply_report_formatting(ws, max_row: int, section_starts: Dict[str, int], section_width: int):
    thin = Side(style="thin", color="FFB7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill("solid", fgColor="FF333333")
    header_fill = PatternFill("solid", fgColor="FFBFBFBF")
    check_fill = PatternFill("solid", fgColor="FFFFF2CC")
    title_font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    header_font = Font(name="Calibri", size=10, bold=True, color="FF000000")
    body_font = Font(name="Calibri", size=10, color="FF000000")

    last_col = section_starts["NOW"] + section_width - 1
    for col in range(1, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 8
    for start_col in section_starts.values():
        ws.column_dimensions[get_column_letter(start_col)].width = 14
        ws.column_dimensions[get_column_letter(start_col + 1)].width = 10

    for start_col in section_starts.values():
        title_cell = ws.cell(row=1, column=start_col + 2)
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.border = border
        for col in range(start_col, start_col + section_width):
            cell = ws.cell(row=2, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    for row in range(3, max_row + 1):
        is_check_row = not ws.cell(row=row, column=section_starts["NEW"]).value and not ws.cell(row=row, column=section_starts["NEW"] + 1).value
        for section_name, start_col in section_starts.items():
            for col in range(start_col, start_col + section_width):
                cell = ws.cell(row=row, column=col)
                cell.font = body_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
                if is_check_row and section_name == "NEW":
                    cell.fill = check_fill

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"


def write_log_sheet(wb: Workbook, log_lines: List[str]):
    ws = wb.create_sheet("LOG")
    ws.cell(row=1, column=1, value="Timestamp")
    ws.cell(row=1, column=2, value="Message")
    now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for idx, line in enumerate(log_lines, start=2):
        ws.cell(row=idx, column=1, value=now_text)
        ws.cell(row=idx, column=2, value=line)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 140
    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="FF333333")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------


def split_line_item_code(code: object) -> Tuple[int, int]:
    digits = re.sub(r"\D", "", str(code or ""))
    if len(digits) < 3:
        raise ValueError(f"Unexpected line item code: {code}")
    line_prefix = digits[:-2]
    ordinal_text = digits[-2:]
    return int(line_prefix) * 100, int(ordinal_text)


def normalize_size(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return text.upper()


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def to_optional_int(value: object) -> Optional[int]:
    if value is None or value == "":
        return None
    return to_int(value)


def to_int(value: object) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(round(value))
    text = str(value).strip().replace(",", "")
    if not text:
        return 0
    return int(round(float(text)))


def proposed_output_path(search_results_path: str) -> str:
    base_dir = os.path.dirname(os.path.abspath(search_results_path))
    stem = os.path.splitext(os.path.basename(search_results_path))[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(base_dir, f"{stem}_divert_analysis_{timestamp}.xlsx")
