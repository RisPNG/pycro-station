"""
VTEC Monitoring Charter - Python Port v1.8.9

Converts the VBA MasterStart macro to Python with openpyxl.
Key fixes:
- v1.8.0: Updates existing groups when new/modified data is encountered.
- v1.8.1: Fixes zero-balance detection/copying and improves update/highlight behavior.
- v1.8.2: Adds verbose progress logging for long runs.
- v1.8.3: Fixes duplicate group blocks by improving existing-group detection and auto-removing duplicates.
- v1.8.4: Formats invoice dates as d/m/yyyy (no time).
- v1.8.5: Fixes payment parsing for numeric sheet names with whitespace (e.g. '1025 ').
- v1.8.6: When a monitoring workbook is provided, saves output as a timestamped copy instead of overwriting.
- v1.8.7: Fixes Total formulas after row deletions by rebuilding them from the final sheet layout.
- v1.8.8: Re-compresses consecutive blank rows between groups (without breaking Total formulas).
- v1.8.9: Ignores Balance columns when detecting last data row (prevents huge gaps from pre-filled formulas).

Performance optimizations:
- Uses read_only=True for source files (much faster for large files)
- Bulk reads with iter_rows(values_only=True)
- Minimizes cell-by-cell operations
- Caches styles to avoid recreation
"""

import os
import re
import threading
import time
from copy import copy
from datetime import date, datetime
from typing import Dict, List, Any, Optional, Tuple

# GUI Imports
from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QVBoxLayout,
    QLabel,
    QTextEdit,
    QWidget,
    QSizePolicy
)
from qfluentwidgets import PrimaryPushButton, MessageBox

# Excel Imports
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

# Pre-create styles (reusing is faster than creating new ones each time)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
GREY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)

INVOICE_DATE_NUMBER_FORMAT = "d/m/yyyy"
MONITOR_DATA_MAX_COL = 9  # Only columns A-I determine the "last data row" (ignore Balance J/K which may be pre-filled)

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

BOLD_FONT = Font(bold=True)
ITALIC_FONT = Font(italic=True)

# Regex for extracting prefix (e.g., BA985283 from BA985283MJ)
PREFIX_PATTERN = re.compile(r'^([A-Za-z]+\d+)', re.IGNORECASE)


def extract_prefix(value: str) -> Optional[str]:
    """Extract prefix like 'BA985283' from 'BA985283MJ'."""
    if not value:
        return None
    match = PREFIX_PATTERN.match(str(value).strip())
    return match.group(1) if match else None


def safe_float(value) -> float:
    """Safely convert to float."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(",", ""))
    except ValueError:
        return 0.0


def normalize_excel_date(value) -> Optional[date]:
    """Normalize various Excel date representations to a Python date."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel
            dt = from_excel(value)
            if isinstance(dt, datetime):
                return dt.date()
            if isinstance(dt, date):
                return dt
        except Exception:
            return None
    s = str(value).strip()
    if not s:
        return None
    # Strip trailing time if present, e.g. "05/09/2025 00:00:00"
    date_part = s.split()[0]
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(date_part, fmt).date()
        except ValueError:
            continue
    return None


def month_num_to_name(num_str: str) -> str:
    """Convert month number to abbreviation."""
    mapping = {
        "01": "JAN", "02": "FEB", "03": "MAR", "04": "APR",
        "05": "MAY", "06": "JUN", "07": "JUL", "08": "AUG",
        "09": "SEP", "10": "OCT", "11": "NOV", "12": "DEC"
    }
    return mapping.get(num_str, num_str)


def normalize_group(grp: str) -> str:
    """Normalize group string for comparison."""
    return str(grp).strip().upper().replace(" ", "").replace("-", "")


def _has_meaningful_value(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and not value.strip():
        return False
    return True


def _row_has_any_values(row: tuple) -> bool:
    return any(_has_meaningful_value(v) for v in row)


def _ensure_unique_path(path: str) -> str:
    """Append (n) to path if needed to avoid overwriting existing files."""
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    n = 1
    while True:
        candidate = f"{base} ({n}){ext}"
        if not os.path.exists(candidate):
            return candidate
        n += 1


class VTECProcessor:
    """Main processing logic for VTEC Monitoring Chart."""

    def __init__(self, log_emit):
        self.log = log_emit

        # Dictionaries to store parsed data
        self.dict_payment = {}        # prefix -> {job_no -> (amount, inv_date, inv_qty)}
        self.dict_duplicate_prefixes = set()
        self.dict_pi = {}             # group|pi_full -> (group, pi_amt, pi_full, pi_month, pi_qty, prefix)
        self.dict_pi_job_nos = {}     # group|pi_full -> job_no
        self.dict_group_pi = {}       # group -> total pi amount
        self.dict_group_qty = {}      # group -> total pi qty
        self.dict_group_prefixes = {} # group -> [pi_keys...]
        self.dict_debit = {}          # group -> (debit_full, debit_month, debit_amount)
        self.dict_charge = {}         # normalized_group -> (credit, debit)

    def parse_payment_file(self, filepath: str):
        """Parse payment file - looks for numeric sheet names, data starts at row 9."""
        t0 = time.perf_counter()
        self.log(f"Parsing Payment file: {os.path.basename(filepath)}")

        try:
            # Use read_only=True for MUCH faster reading of large files
            wb = load_workbook(filepath, data_only=True, read_only=True)

            # Some payment workbooks use numeric sheet names with trailing spaces (e.g. "1025 ").
            numeric_sheets = [s for s in wb.sheetnames if str(s).strip().isdigit()]
            self.log(f"  Payment workbook sheets: {len(wb.sheetnames)} (numeric sheets: {len(numeric_sheets)})")

            rows_total = 0
            rows_used = 0
            last_progress = time.monotonic()

            for sheet_idx, sheet_name in enumerate(numeric_sheets, 1):
                if not str(sheet_name).strip().isdigit():
                    continue

                ws = wb[sheet_name]
                display_name = str(sheet_name).strip()
                if display_name != str(sheet_name):
                    self.log(
                        f"  [Payment] Sheet {sheet_idx}/{len(numeric_sheets)}: {display_name} "
                        f"(note: sheet name has whitespace)"
                    )
                else:
                    self.log(f"  [Payment] Sheet {sheet_idx}/{len(numeric_sheets)}: {display_name}")

                # Bulk read all rows at once - much faster than cell-by-cell
                # Columns: B=2, C=3, D=4, E=5, X=24
                for row in ws.iter_rows(min_row=9, values_only=True):
                    rows_total += 1
                    if not row or len(row) < 24:
                        continue

                    col_d = str(row[3] or "").strip()  # D - PI No (index 3)

                    if not col_d:
                        continue

                    prefix = extract_prefix(col_d)
                    if not prefix:
                        continue

                    rows_used += 1
                    job_no = str(row[2] or "").strip()  # C - Job No (index 2)
                    vx = safe_float(row[23])  # X - Amount (index 23)
                    inv_qty = safe_float(row[4])  # E - Qty (index 4)
                    inv_date = normalize_excel_date(row[1])  # B - Invoice Date (index 1)

                    if prefix not in self.dict_payment:
                        self.dict_payment[prefix] = {}

                    if job_no not in self.dict_payment[prefix]:
                        self.dict_payment[prefix][job_no] = [vx, inv_date, inv_qty]
                    else:
                        # Accumulate
                        self.dict_payment[prefix][job_no][0] += vx
                        self.dict_payment[prefix][job_no][2] += inv_qty

                    now = time.monotonic()
                    if now - last_progress >= 5.0:
                        self.log(
                            f"    ... scanned {rows_total:,} rows (matched {rows_used:,}), "
                            f"prefixes so far: {len(self.dict_payment):,}"
                        )
                        last_progress = now

            wb.close()

            # Mark duplicate prefixes (prefixes with multiple jobs)
            for prefix, jobs in self.dict_payment.items():
                if len(jobs) > 1:
                    self.dict_duplicate_prefixes.add(prefix)

            elapsed = time.perf_counter() - t0
            self.log(
                f"  Found {len(self.dict_payment):,} prefixes in payment data "
                f"(duplicates: {len(self.dict_duplicate_prefixes):,})"
            )
            self.log(f"  Payment parse finished in {elapsed:.1f}s (rows scanned: {rows_total:,})")

        except Exception as e:
            self.log(f"  Error parsing payment file: {e}")

    def parse_pi_file(self, filepath: str):
        """Parse PI file - expects 'NK' sheet, data starts at row 10."""
        t0 = time.perf_counter()
        self.log(f"Parsing PI file: {os.path.basename(filepath)}")

        try:
            # Use read_only=True for faster reading
            wb = load_workbook(filepath, data_only=True, read_only=True)

            if "NK" not in wb.sheetnames:
                self.log("  ERROR: Sheet 'NK' not found in PI file")
                wb.close()
                return

            ws = wb["NK"]

            # Bulk read - Columns: B=2, C=3, D=4, H=8, J=10, L=12
            rows_total = 0
            rows_used = 0
            groups_seen = set()
            last_progress = time.monotonic()
            for row in ws.iter_rows(min_row=10, values_only=True):
                rows_total += 1
                if not row or len(row) < 12:
                    continue

                col_c = str(row[2] or "").strip()  # C - PI No (index 2)

                if not col_c:
                    continue

                prefix = extract_prefix(col_c)
                if not prefix:
                    continue

                rows_used += 1
                group = str(row[9] or "").strip()  # J - Group (index 9)
                if group:
                    groups_seen.add(group)
                pi_amt = safe_float(row[7])  # H - PI Amount (index 7)
                pi_full = col_c
                pi_month = row[3]  # D - PI Month (index 3)
                pi_qty = safe_float(row[11])  # L - PI Qty (index 11)
                pi_job_no = str(row[1] or "").strip()  # B - Job No (index 1)

                pi_key = f"{group}|{pi_full}"

                if pi_key not in self.dict_pi_job_nos:
                    self.dict_pi_job_nos[pi_key] = pi_job_no

                if pi_key not in self.dict_pi:
                    self.dict_pi[pi_key] = (group, pi_amt, pi_full, pi_month, pi_qty, prefix)

                now = time.monotonic()
                if now - last_progress >= 5.0:
                    self.log(
                        f"    ... scanned {rows_total:,} rows (matched {rows_used:,}), "
                        f"groups so far: {len(groups_seen):,}, PI keys: {len(self.dict_pi):,}"
                    )
                    last_progress = now

            wb.close()

            # Build group-level sums
            for pi_key, rec in self.dict_pi.items():
                group, pi_amt, _, _, pi_qty, prefix = rec

                self.dict_group_pi[group] = self.dict_group_pi.get(group, 0) + pi_amt
                self.dict_group_qty[group] = self.dict_group_qty.get(group, 0) + pi_qty

                if group not in self.dict_group_prefixes:
                    self.dict_group_prefixes[group] = []
                self.dict_group_prefixes[group].append(pi_key)

            elapsed = time.perf_counter() - t0
            self.log(f"  Found {len(self.dict_group_prefixes):,} groups in PI data (PI keys: {len(self.dict_pi):,})")
            self.log(f"  PI parse finished in {elapsed:.1f}s (rows scanned: {rows_total:,})")

        except Exception as e:
            self.log(f"  Error parsing PI file: {e}")

    def parse_debit_file(self, filepath: str):
        """Parse Debit Note file - data starts at row 2."""
        t0 = time.perf_counter()
        self.log(f"Parsing Debit Note file: {os.path.basename(filepath)}")

        try:
            # Use read_only=True for faster reading
            wb = load_workbook(filepath, data_only=True, read_only=True)
            ws = wb.active

            # Bulk read - Columns: A=1, B=2, C=3
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 3:
                    continue

                group = str(row[0] or "").strip()  # A - Group (index 0)

                if not group:
                    continue

                debit_full = str(row[1] or "")  # B - Debit Note No (index 1)
                debit_amt = safe_float(row[2])  # C - Amount (index 2)

                # Extract month number from debit_full
                debit_num = ""
                for char in debit_full:
                    if char.isdigit():
                        debit_num += char
                        if len(debit_num) == 2:
                            break

                if group not in self.dict_debit:
                    self.dict_debit[group] = (debit_full, debit_num, debit_amt)

            wb.close()
            self.log(f"  Found {len(self.dict_debit):,} debit note entries")
            self.log(f"  Debit Note parse finished in {time.perf_counter() - t0:.1f}s")

        except Exception as e:
            self.log(f"  Error parsing debit file: {e}")

    def parse_charge_file(self, filepath: str):
        """Parse Charge to Purchase file - data starts at row 2."""
        t0 = time.perf_counter()
        self.log(f"Parsing Charge to Purchase file: {os.path.basename(filepath)}")

        try:
            # Use read_only=True for faster reading
            wb = load_workbook(filepath, data_only=True, read_only=True)
            ws = wb.active

            # Bulk read - Columns: A=1, B=2, C=3
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 3:
                    continue

                pi_no = str(row[0] or "").strip()  # A - PI No (index 0)

                if not pi_no:
                    continue

                normalized = normalize_group(pi_no)
                credit = safe_float(row[1])  # B - Credit (index 1)
                debit = safe_float(row[2])   # C - Debit (index 2)

                self.dict_charge[normalized] = (credit, debit)

            wb.close()
            self.log(f"  Found {len(self.dict_charge):,} charge entries")
            self.log(f"  Charge parse finished in {time.perf_counter() - t0:.1f}s")

        except Exception as e:
            self.log(f"  Error parsing charge file: {e}")

    def _find_last_data_row(self, ws, start_row: int = 5, max_col: int = 11) -> int:
        """Find the last row containing any value (ignores style-only cells)."""
        last_data_row = start_row - 1
        for r, row in enumerate(
            ws.iter_rows(min_row=start_row, max_col=max_col, values_only=True),
            start=start_row,
        ):
            if row and _row_has_any_values(row):
                last_data_row = r
        return max(last_data_row, 4)

    def _rebuild_total_row_formulas(self, ws):
        """Rebuild per-group Total row formulas based on current row positions."""
        last_row = self._find_last_data_row(ws, start_row=4, max_col=MONITOR_DATA_MAX_COL)
        rows = list(ws.iter_rows(min_row=5, max_row=last_row, max_col=7, values_only=True))
        row_offset = 5

        rebuilt = 0
        idx = 0
        while idx < len(rows):
            col_a, col_b, col_c, col_d, _, col_f, col_g = rows[idx]
            group = str(col_a or "").strip()
            b_str = str(col_b or "").strip()
            c_str = str(col_c or "").strip()
            header_marker = (
                _has_meaningful_value(col_d)
                or _has_meaningful_value(col_f)
                or _has_meaningful_value(col_g)
            )

            # Header row: group in A, blank B/C, and some header fields filled (PI Month / PI totals).
            # (Avoids confusing the debit placeholder row with the group header.)
            if group and (not b_str) and (not c_str) and header_marker:
                header_row = idx + row_offset
                first_job_row = header_row + 1
                last_job_row = first_job_row - 1
                exch_row = None

                scan = idx + 1
                while scan < len(rows):
                    sa, sb, sc, sd, _, sf, sg = rows[scan]
                    sa_str = str(sa or "").strip()
                    sb_str = str(sb or "").strip()
                    sc_str = str(sc or "").strip()
                    sb_u = sb_str.upper()

                    # If we hit another header row before Total, stop this block.
                    next_header_marker = (
                        _has_meaningful_value(sd)
                        or _has_meaningful_value(sf)
                        or _has_meaningful_value(sg)
                    )
                    if sa_str and sa_str != group and (not sb_str) and (not sc_str) and next_header_marker:
                        break

                    if sb_u == "EXCHANGE GAIN OR LOSS":
                        exch_row = scan + row_offset
                    elif sb_u == "TOTAL":
                        total_row = scan + row_offset
                        exch_for_formula = exch_row or (total_row - 1)
                        if last_job_row < first_job_row:
                            last_job_row = first_job_row

                        ws.cell(row=total_row, column=6).value = f"=F{header_row}"
                        ws.cell(row=total_row, column=6).number_format = "#,##0"
                        ws.cell(row=total_row, column=7).value = f"=ROUND(SUM(G{header_row}:G{exch_for_formula}),2)"
                        ws.cell(row=total_row, column=7).number_format = "#,##0.00"
                        ws.cell(row=total_row, column=8).value = f"=SUM(H{first_job_row}:H{last_job_row})"
                        ws.cell(row=total_row, column=8).number_format = "#,##0"
                        ws.cell(row=total_row, column=9).value = f"=ROUND(SUM(I{first_job_row}:I{exch_for_formula}),2)"
                        ws.cell(row=total_row, column=9).number_format = "#,##0.00"
                        ws.cell(row=total_row, column=10).value = f"=ROUND(G{total_row}-I{total_row},2)"
                        ws.cell(row=total_row, column=10).number_format = "#,##0.00"
                        ws.cell(row=total_row, column=11).value = f"=ROUND(F{total_row}-H{total_row},2)"
                        ws.cell(row=total_row, column=11).number_format = "#,##0"
                        for col in range(1, 12):
                            ws.cell(row=total_row, column=col).font = ITALIC_FONT

                        rebuilt += 1
                        idx = scan + 1
                        break

                    if sc_str:
                        last_job_row = scan + row_offset
                    scan += 1
                else:
                    idx += 1
            else:
                idx += 1

        self.log(f"Rebuilt Total formulas for {rebuilt:,} group(s)")

    def process_monitoring_workbook(self, wb: Workbook) -> Workbook:
        """
        Process the monitoring workbook - create/update VTEC Monitoring Chart.
        This is the main logic that includes the fix for updating existing groups.
        """
        # Check if sheet exists
        fresh_sheet = "VTEC Monitoring Chart" not in wb.sheetnames

        t0 = time.perf_counter()
        if fresh_sheet:
            ws = wb.create_sheet("VTEC Monitoring Chart")
            self.log("Created new 'VTEC Monitoring Chart' sheet")
        else:
            ws = wb["VTEC Monitoring Chart"]
            self.log("Updating existing 'VTEC Monitoring Chart' sheet")

        # Track existing groups and their row ranges
        existing_groups = {}  # group -> {'header': row, 'first_job': row, 'last_job': row, 'debit': row, 'charge': row, 'exch': row, 'total': row}
        existing_pi_rows = {}  # pi_full -> row

        if fresh_sheet:
            # Add headers at row 4
            headers = [
                "PI No.", "VTEC", "Job No.", "PI Month", "Invoice Date",
                "PI Qty", "PI Amount / Credit (USD)", "Invoice Qty",
                "Material Amount / Debit (USD)", "Balance (USD)", "Balance (Qty)"
            ]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_idx)
                cell.value = header
                cell.fill = GREY_FILL
                cell.font = BOLD_FONT
            out_row = 6
        else:
            last_row = ws.max_row

            # OPTIMIZATION: Read all data in bulk first (much faster than cell-by-cell)
            # This creates a list of tuples for all rows
            self.log("Caching existing sheet values for faster scanning...")
            cache_t0 = time.perf_counter()
            all_rows_data = list(ws.iter_rows(min_row=5, max_row=last_row, max_col=MONITOR_DATA_MAX_COL, values_only=True))
            row_offset = 5  # Starting row number
            last_data_row = row_offset - 1
            for i in range(len(all_rows_data) - 1, -1, -1):
                if all_rows_data[i] and _row_has_any_values(all_rows_data[i]):
                    last_data_row = i + row_offset
                    all_rows_data = all_rows_data[: i + 1]
                    break

            self.log(
                f"Cached {len(all_rows_data):,} rows in {time.perf_counter() - cache_t0:.1f}s "
                f"(last data row: {last_data_row:,})"
            )

            # Clear previous-run green highlights (match VBA behavior).
            self.log(f"Clearing previous-run green highlights (rows 5-{last_data_row})...")
            clear_t0 = time.perf_counter()
            last_progress = time.monotonic()
            for row in range(5, last_data_row + 1):
                now = time.monotonic()
                if now - last_progress >= 5.0:
                    self.log(f"  ... cleared green up to row {row:,}/{last_data_row:,}")
                    last_progress = now
                for col in range(1, 12):
                    cell = ws.cell(row=row, column=col)
                    try:
                        rgb = str(getattr(getattr(cell.fill, "start_color", None), "rgb", "") or "").upper()
                        if rgb.endswith("C6EFCE"):
                            cell.fill = NO_FILL
                    except Exception:
                        pass

            self.log(f"Cleared green highlights in {time.perf_counter() - clear_t0:.1f}s")

            # Build existing group info by scanning the cached data
            row_idx = 1  # Index in all_rows_data (row 6 = index 1)
            self.log("Scanning existing group blocks in monitoring sheet...")
            scan_t0 = time.perf_counter()
            last_progress = time.monotonic()
            group_headers = set(self.dict_group_prefixes.keys())
            occurrences_by_group: dict[str, list[dict]] = {}

            while row_idx < len(all_rows_data):
                now = time.monotonic()
                if now - last_progress >= 5.0:
                    self.log(f"  ... scanned {row_idx:,}/{len(all_rows_data):,} cached rows (groups found: {len(existing_groups):,})")
                    last_progress = now
                row_data = all_rows_data[row_idx]
                actual_row = row_idx + row_offset

                if not row_data:
                    row_idx += 1
                    continue

                cell_a = str(row_data[0] or "").strip()
                cell_b = str(row_data[1] or "").strip()
                cell_c = str(row_data[2] or "").strip() if len(row_data) > 2 else ""
                header_marker = (
                    (len(row_data) > 3 and _has_meaningful_value(row_data[3]))
                    or (len(row_data) > 5 and _has_meaningful_value(row_data[5]))
                    or (len(row_data) > 6 and _has_meaningful_value(row_data[6]))
                )

                # Check if this is a group header (has group ID in A)
                if (
                    cell_a
                    and cell_a in group_headers
                    and (not cell_b)
                    and (not cell_c)
                    and header_marker
                ):
                    group = cell_a
                    header_row = actual_row

                    # Find first job row
                    first_job_row = actual_row + 1

                    # Find last job row, debit, charge, exch, total rows
                    scan_idx = row_idx + 1
                    last_job_row = first_job_row - 1
                    debit_row = None
                    charge_row = None
                    exch_row = None
                    total_row = None

                    while scan_idx < len(all_rows_data):
                        scan_data = all_rows_data[scan_idx]
                        scan_actual_row = scan_idx + row_offset

                        if not scan_data:
                            scan_idx += 1
                            continue

                        scan_a = str(scan_data[0] or "").strip()
                        scan_b = str(scan_data[1] or "").strip()
                        scan_b_u = scan_b.upper()

                        # Stop at the next group block (including groups not in the current PI file).
                        if scan_a and scan_a != group:
                            break

                        if scan_b_u == "CHARGE TO PURCHASE":
                            charge_row = scan_actual_row
                        elif scan_b_u == "EXCHANGE GAIN OR LOSS":
                            exch_row = scan_actual_row
                        elif scan_b_u == "TOTAL":
                            total_row = scan_actual_row
                            break
                        else:
                            pi_full_cell = scan_data[2] if len(scan_data) > 2 else None
                            pi_full = str(pi_full_cell).strip() if pi_full_cell is not None else ""
                            if pi_full:
                                last_job_row = scan_actual_row
                                existing_pi_rows[f"{group}|{pi_full}"] = scan_actual_row

                        scan_idx += 1

                    if total_row:
                        if charge_row and debit_row is None:
                            # Debit placeholder row is always immediately before charge row.
                            candidate = charge_row - 1
                            if candidate > header_row:
                                debit_row = candidate
                        if last_job_row < first_job_row:
                            last_job_row = first_job_row

                        existing_groups[group] = {
                            'header': header_row,
                            'first_job': first_job_row,
                            'last_job': last_job_row,
                            'debit': debit_row,
                            'charge': charge_row,
                            'exch': exch_row,
                            'total': total_row
                        }
                        occurrences_by_group.setdefault(group, []).append(existing_groups[group])
                        row_idx = total_row - row_offset + 1
                    else:
                        row_idx += 1
                else:
                    row_idx += 1

            self.log(f"Existing group scan finished in {time.perf_counter() - scan_t0:.1f}s")

            # De-duplicate group blocks if the same group appears multiple times.
            duplicate_groups = {g: occs for g, occs in occurrences_by_group.items() if len(occs) > 1}
            if duplicate_groups:
                dedup_t0 = time.perf_counter()
                self.log(
                    f"WARNING: Found {len(duplicate_groups):,} duplicate group(s) in the monitoring sheet. "
                    "Keeping the most recent block and moving older duplicates to 'Duplicate Groups (Review)'."
                )

                dup_ws_name = "Duplicate Groups (Review)"
                if dup_ws_name in wb.sheetnames:
                    dup_ws = wb[dup_ws_name]
                    dup_ws.delete_rows(1, dup_ws.max_row)
                else:
                    dup_ws = wb.create_sheet(dup_ws_name)

                dup_ws.cell(row=1, column=1).value = "Duplicate Groups (Review) - removed blocks"
                # Copy headers from monitoring chart (row 4)
                for col in range(1, 12):
                    src = ws.cell(row=4, column=col)
                    dst = dup_ws.cell(row=2, column=col)
                    dst.value = src.value
                    dst.fill = copy(src.fill)
                    dst.font = copy(src.font)
                    dst.border = copy(src.border)
                    dst.alignment = copy(src.alignment)
                    dst.number_format = src.number_format

                paste_row = 3
                ranges_to_delete: list[tuple[int, int]] = []

                for group, occs in sorted(duplicate_groups.items(), key=lambda x: x[0]):
                    occs_sorted = sorted(occs, key=lambda o: o.get("header", 0))
                    keep = occs_sorted[-1]
                    keep_header = int(keep.get("header") or 0)

                    for occ in occs_sorted[:-1]:
                        start_row = int(occ.get("header") or 0)
                        end_row = int(occ.get("total") or 0)
                        if start_row <= 0 or end_row <= 0 or end_row < start_row:
                            continue

                        dup_ws.cell(row=paste_row, column=1).value = (
                            f"REMOVED DUPLICATE: {group} (rows {start_row}-{end_row}) "
                            f"kept header row {keep_header}"
                        )
                        dup_ws.cell(row=paste_row, column=1).font = BOLD_FONT
                        paste_row += 1

                        for src_r in range(start_row, end_row + 1):
                            for col in range(1, 12):
                                src_cell = ws.cell(row=src_r, column=col)
                                dst_cell = dup_ws.cell(row=paste_row + (src_r - start_row), column=col)
                                dst_cell.value = src_cell.value
                                dst_cell.fill = copy(src_cell.fill)
                                dst_cell.font = copy(src_cell.font)
                                dst_cell.border = copy(src_cell.border)
                                dst_cell.alignment = copy(src_cell.alignment)
                                dst_cell.number_format = src_cell.number_format
                        paste_row += (end_row - start_row + 1) + 1

                        ranges_to_delete.append((start_row, end_row))

                # Delete from monitoring sheet (reverse order to avoid row shift issues)
                ranges_to_delete.sort(reverse=True, key=lambda x: x[0])
                for start_row, end_row in ranges_to_delete:
                    ws.delete_rows(start_row, end_row - start_row + 1)

                for col in range(1, 12):
                    dup_ws.column_dimensions[get_column_letter(col)].auto_size = True

                self.log(
                    f"Duplicate cleanup finished in {time.perf_counter() - dedup_t0:.1f}s "
                    f"(removed blocks: {len(ranges_to_delete):,})"
                )

                # Re-scan after deletions (row indices changed)
                last_row = ws.max_row
                self.log("Re-scanning monitoring sheet after duplicate cleanup...")
                cache_t0 = time.perf_counter()
                all_rows_data = list(ws.iter_rows(min_row=5, max_row=last_row, max_col=MONITOR_DATA_MAX_COL, values_only=True))
                last_data_row = row_offset - 1
                for i in range(len(all_rows_data) - 1, -1, -1):
                    if all_rows_data[i] and _row_has_any_values(all_rows_data[i]):
                        last_data_row = i + row_offset
                        all_rows_data = all_rows_data[: i + 1]
                        break
                self.log(
                    f"Re-cached {len(all_rows_data):,} rows in {time.perf_counter() - cache_t0:.1f}s "
                    f"(last data row: {last_data_row:,})"
                )

                existing_groups.clear()
                existing_pi_rows.clear()
                row_idx = 1
                while row_idx < len(all_rows_data):
                    row_data = all_rows_data[row_idx]
                    actual_row = row_idx + row_offset

                    cell_a = str(row_data[0] or "").strip()
                    cell_b = str(row_data[1] or "").strip()
                    cell_c = str(row_data[2] or "").strip() if len(row_data) > 2 else ""
                    header_marker = (
                        (len(row_data) > 3 and _has_meaningful_value(row_data[3]))
                        or (len(row_data) > 5 and _has_meaningful_value(row_data[5]))
                        or (len(row_data) > 6 and _has_meaningful_value(row_data[6]))
                    )
                    if (
                        cell_a
                        and cell_a in group_headers
                        and (not cell_b)
                        and (not cell_c)
                        and header_marker
                    ):
                        group = cell_a
                        header_row = actual_row
                        first_job_row = actual_row + 1
                        scan_idx = row_idx + 1
                        last_job_row = first_job_row - 1
                        charge_row = None
                        exch_row = None
                        total_row = None

                        while scan_idx < len(all_rows_data):
                            scan_data = all_rows_data[scan_idx]
                            scan_actual_row = scan_idx + row_offset
                            scan_a = str(scan_data[0] or "").strip()
                            scan_b = str(scan_data[1] or "").strip()
                            scan_b_u = scan_b.upper()

                            if scan_a and scan_a != group:
                                break
                            if scan_b_u == "CHARGE TO PURCHASE":
                                charge_row = scan_actual_row
                            elif scan_b_u == "EXCHANGE GAIN OR LOSS":
                                exch_row = scan_actual_row
                            elif scan_b_u == "TOTAL":
                                total_row = scan_actual_row
                                break
                            else:
                                pi_full_cell = scan_data[2] if len(scan_data) > 2 else None
                                pi_full = str(pi_full_cell).strip() if pi_full_cell is not None else ""
                                if pi_full:
                                    last_job_row = scan_actual_row
                                    existing_pi_rows[f"{group}|{pi_full}"] = scan_actual_row
                            scan_idx += 1

                        if total_row:
                            debit_row = None
                            if charge_row:
                                candidate = charge_row - 1
                                if candidate > header_row:
                                    debit_row = candidate
                            if last_job_row < first_job_row:
                                last_job_row = first_job_row

                            existing_groups[group] = {
                                'header': header_row,
                                'first_job': first_job_row,
                                'last_job': last_job_row,
                                'debit': debit_row,
                                'charge': charge_row,
                                'exch': exch_row,
                                'total': total_row
                            }
                            row_idx = total_row - row_offset + 1
                        else:
                            row_idx += 1
                    else:
                        row_idx += 1

            # Find next output row (after any de-duplication)
            out_row = last_data_row + 3

        self.log(f"Found {len(existing_groups)} existing groups to potentially update")

        # Track rows to highlight
        rows_to_highlight = set()

        # Dict to store new group ranges
        group_ranges = {}  # group -> (header_row, first_job, last_job, debit_row, total_row)

        groups_in_data = list(self.dict_group_prefixes.keys())
        existing_sorted = []
        if not fresh_sheet and existing_groups:
            existing_sorted = sorted(
                [g for g in groups_in_data if g in existing_groups],
                key=lambda g: existing_groups[g]["header"],
            )
        groups_to_process = existing_sorted + [g for g in groups_in_data if g not in existing_groups]
        existing_index = {g: i for i, g in enumerate(existing_sorted)}
        out_row_rebased = False
        total_groups = len(groups_to_process)
        new_groups_count = len([g for g in groups_in_data if g not in existing_groups]) if not fresh_sheet else len(groups_in_data)
        self.log(
            f"Processing {total_groups:,} group(s) "
            f"(existing: {len(existing_sorted):,}, new: {new_groups_count:,})..."
        )

        added_groups = 0
        updated_groups = 0
        total_rows_inserted = 0
        last_progress = time.monotonic()

        # Process each group
        for idx, group in enumerate(groups_to_process, 1):
            pi_keys = self.dict_group_prefixes[group]

            # If we inserted rows while updating existing groups, recompute the append position
            # right before adding the first new group.
            if not fresh_sheet and (not out_row_rebased) and group not in existing_groups:
                out_row = self._find_last_data_row(ws, start_row=5, max_col=MONITOR_DATA_MAX_COL) + 3
                out_row_rebased = True

            now = time.monotonic()
            if idx == 1 or idx == total_groups or (now - last_progress) >= 5.0 or (idx % 25 == 0):
                action = "Updating" if (not fresh_sheet and group in existing_groups) else "Adding"
                self.log(
                    f"  [{idx:,}/{total_groups:,}] {action} group: {group} "
                    f"(updated: {updated_groups:,}, added: {added_groups:,}, inserted rows: {total_rows_inserted:,})"
                )
                last_progress = now

            if fresh_sheet or group not in existing_groups:
                # === ADD NEW GROUP ===
                header_row = out_row

                # Header row
                ws.cell(row=header_row, column=1).value = group
                ws.cell(row=header_row, column=4).value = self.dict_pi[pi_keys[0]][3]  # PI Month
                ws.cell(row=header_row, column=6).value = self.dict_group_qty[group]
                ws.cell(row=header_row, column=6).number_format = "#,##0"
                ws.cell(row=header_row, column=7).value = round(self.dict_group_pi[group], 2)
                ws.cell(row=header_row, column=7).number_format = "#,##0.00"

                out_row += 1
                first_job_row = out_row

                # Job rows
                for pi_key in pi_keys:
                    rec = self.dict_pi[pi_key]
                    group_val, pi_amt, pi_full, pi_month, pi_qty, prefix = rec

                    has_payment = prefix in self.dict_payment

                    if has_payment:
                        for job_no, pay_data in self.dict_payment[prefix].items():
                            pay_amt, inv_date, inv_qty = pay_data

                            ws.cell(row=out_row, column=1).value = group
                            ws.cell(row=out_row, column=2).value = job_no
                            ws.cell(row=out_row, column=3).value = pi_full
                            ws.cell(row=out_row, column=4).value = pi_month
                            date_cell = ws.cell(row=out_row, column=5)
                            date_cell.value = inv_date
                            date_cell.number_format = INVOICE_DATE_NUMBER_FORMAT
                            ws.cell(row=out_row, column=6).value = pi_qty
                            ws.cell(row=out_row, column=6).number_format = "#,##0"
                            ws.cell(row=out_row, column=8).value = inv_qty
                            ws.cell(row=out_row, column=8).number_format = "#,##0"
                            ws.cell(row=out_row, column=9).value = round(pay_amt, 2)
                            ws.cell(row=out_row, column=9).number_format = "#,##0.00"

                            rows_to_highlight.add(out_row)
                            out_row += 1
                    else:
                        # No payment match - yellow highlight
                        pi_job_no = self.dict_pi_job_nos.get(pi_key, "")

                        ws.cell(row=out_row, column=1).value = group
                        ws.cell(row=out_row, column=2).value = pi_job_no
                        ws.cell(row=out_row, column=3).value = pi_full
                        ws.cell(row=out_row, column=4).value = pi_month
                        ws.cell(row=out_row, column=6).value = pi_qty
                        ws.cell(row=out_row, column=6).number_format = "#,##0"
                        ws.cell(row=out_row, column=6).fill = YELLOW_FILL

                        rows_to_highlight.add(out_row)
                        out_row += 1

                last_job_row = out_row - 1

                # Debit row
                debit_row = out_row
                ws.cell(row=debit_row, column=1).value = group
                if group in self.dict_debit:
                    debit_full, debit_num, debit_amt = self.dict_debit[group]
                    ws.cell(row=debit_row, column=2).value = debit_full

                    if debit_amt < 0:
                        ws.cell(row=debit_row, column=7).value = round(abs(debit_amt), 2)
                        ws.cell(row=debit_row, column=7).number_format = "#,##0.00"
                    else:
                        ws.cell(row=debit_row, column=9).value = round(debit_amt, 2)
                        ws.cell(row=debit_row, column=9).number_format = "#,##0.00"

                    rows_to_highlight.add(debit_row)
                out_row += 1

                # Charge row
                charge_row = out_row
                ws.cell(row=charge_row, column=1).value = group
                ws.cell(row=charge_row, column=2).value = "CHARGE TO PURCHASE"

                normalized_grp = normalize_group(group)
                if normalized_grp in self.dict_charge:
                    credit, debit = self.dict_charge[normalized_grp]
                    if credit != 0:
                        ws.cell(row=charge_row, column=7).value = credit
                        ws.cell(row=charge_row, column=7).number_format = "#,##0.00"
                    if debit != 0:
                        ws.cell(row=charge_row, column=9).value = debit
                        ws.cell(row=charge_row, column=9).number_format = "#,##0.00"
                out_row += 1

                # Exchange row
                exch_row = out_row
                ws.cell(row=exch_row, column=1).value = group
                ws.cell(row=exch_row, column=2).value = "EXCHANGE GAIN OR LOSS"
                out_row += 1

                # Total row
                total_row = out_row
                ws.cell(row=total_row, column=1).value = group
                ws.cell(row=total_row, column=2).value = "Total"
                ws.cell(row=total_row, column=6).value = f"=F{header_row}"
                ws.cell(row=total_row, column=6).number_format = "#,##0"
                ws.cell(row=total_row, column=7).value = f"=ROUND(SUM(G{header_row}:G{exch_row}),2)"
                ws.cell(row=total_row, column=7).number_format = "#,##0.00"
                ws.cell(row=total_row, column=8).value = f"=SUM(H{first_job_row}:H{last_job_row})"
                ws.cell(row=total_row, column=8).number_format = "#,##0"
                ws.cell(row=total_row, column=9).value = f"=ROUND(SUM(I{first_job_row}:I{exch_row}),2)"
                ws.cell(row=total_row, column=9).number_format = "#,##0.00"
                ws.cell(row=total_row, column=10).value = f"=ROUND(G{total_row}-I{total_row},2)"
                ws.cell(row=total_row, column=10).number_format = "#,##0.00"
                ws.cell(row=total_row, column=11).value = f"=ROUND(F{total_row}-H{total_row},2)"
                ws.cell(row=total_row, column=11).number_format = "#,##0"

                # Italic for total row
                for col in range(1, 12):
                    ws.cell(row=total_row, column=col).font = ITALIC_FONT

                # Green highlight for new group
                for row in range(header_row, total_row + 1):
                    for col in range(1, 12):
                        cell = ws.cell(row=row, column=col)
                        # Preserve yellow if present
                        if cell.fill.start_color.rgb != "00FFFF00" and cell.fill.start_color.rgb != "FFFF00":
                            cell.fill = GREEN_FILL

                group_ranges[group] = (header_row, first_job_row, last_job_row, debit_row, total_row)
                out_row = total_row + 3
                added_groups += 1

            else:
                # === UPDATE EXISTING GROUP ===
                # This is the key fix - update existing groups instead of skipping them

                grp_info = existing_groups[group]
                header_row = grp_info['header']
                first_job_row = grp_info['first_job']
                last_job_row = grp_info['last_job']
                debit_row = grp_info.get('debit')
                charge_row = grp_info.get('charge')
                exch_row = grp_info.get('exch')
                total_row = grp_info['total']

                inserted_rows = 0
                group_changed_rows = 0

                # Update header row data (PI month / totals)
                header_changed = False
                header_month = self.dict_pi[pi_keys[0]][3] if pi_keys else None
                if ws.cell(row=header_row, column=4).value != header_month:
                    ws.cell(row=header_row, column=4).value = header_month
                    header_changed = True

                new_qty_total = self.dict_group_qty.get(group, 0)
                new_amt_total = round(self.dict_group_pi.get(group, 0), 2)

                if abs(safe_float(ws.cell(row=header_row, column=6).value) - new_qty_total) > 0.01:
                    ws.cell(row=header_row, column=6).value = new_qty_total
                    ws.cell(row=header_row, column=6).number_format = "#,##0"
                    header_changed = True
                if abs(safe_float(ws.cell(row=header_row, column=7).value) - new_amt_total) > 0.01:
                    ws.cell(row=header_row, column=7).value = new_amt_total
                    ws.cell(row=header_row, column=7).number_format = "#,##0.00"
                    header_changed = True

                if header_changed:
                    rows_to_highlight.add(header_row)
                    group_changed_rows += 1

                # Build existing job mapping (pi_full -> set(job_no))
                existing_jobs_by_pi = {}
                existing_pi_fulls = set()
                for r in range(first_job_row, last_job_row + 1):
                    pi_full_val = str(ws.cell(row=r, column=3).value or "").strip()
                    if not pi_full_val:
                        continue
                    existing_pi_fulls.add(pi_full_val)
                    job_val = str(ws.cell(row=r, column=2).value or "").strip()
                    if pi_full_val not in existing_jobs_by_pi:
                        existing_jobs_by_pi[pi_full_val] = set()
                    if job_val:
                        existing_jobs_by_pi[pi_full_val].add(job_val)

                # Update existing job rows
                for r in range(first_job_row, last_job_row + 1):
                    pi_full_val = str(ws.cell(row=r, column=3).value or "").strip()
                    if not pi_full_val:
                        continue
                    pi_key = f"{group}|{pi_full_val}"
                    if pi_key not in self.dict_pi:
                        continue

                    _, _, _, pi_month, pi_qty, prefix = self.dict_pi[pi_key]
                    row_changed = False

                    if ws.cell(row=r, column=1).value != group:
                        ws.cell(row=r, column=1).value = group
                        row_changed = True
                    if ws.cell(row=r, column=4).value != pi_month:
                        ws.cell(row=r, column=4).value = pi_month
                        row_changed = True
                    if abs(safe_float(ws.cell(row=r, column=6).value) - pi_qty) > 0.01:
                        ws.cell(row=r, column=6).value = pi_qty
                        row_changed = True
                    ws.cell(row=r, column=6).number_format = "#,##0"

                    if prefix in self.dict_payment:
                        pay_jobs = self.dict_payment[prefix]
                        current_job = str(ws.cell(row=r, column=2).value or "").strip()
                        if current_job in pay_jobs:
                            job_no, (pay_amt, inv_date, inv_qty) = current_job, pay_jobs[current_job]
                        else:
                            job_no, (pay_amt, inv_date, inv_qty) = next(iter(pay_jobs.items()))
                            if current_job != job_no:
                                ws.cell(row=r, column=2).value = job_no
                                row_changed = True

                        new_amt = round(pay_amt, 2)
                        new_qty = inv_qty

                        existing_inv_date = normalize_excel_date(ws.cell(row=r, column=5).value)
                        if existing_inv_date != inv_date:
                            date_cell = ws.cell(row=r, column=5)
                            date_cell.value = inv_date
                            date_cell.number_format = INVOICE_DATE_NUMBER_FORMAT
                            row_changed = True
                        if abs(safe_float(ws.cell(row=r, column=8).value) - new_qty) > 0.01:
                            ws.cell(row=r, column=8).value = new_qty
                            row_changed = True
                        if abs(safe_float(ws.cell(row=r, column=9).value) - new_amt) > 0.01:
                            ws.cell(row=r, column=9).value = new_amt
                            row_changed = True

                        ws.cell(row=r, column=8).number_format = "#,##0"
                        ws.cell(row=r, column=9).number_format = "#,##0.00"
                    else:
                        pi_job_no = self.dict_pi_job_nos.get(pi_key, "")
                        if str(ws.cell(row=r, column=2).value or "").strip() != pi_job_no:
                            ws.cell(row=r, column=2).value = pi_job_no
                            row_changed = True
                        if ws.cell(row=r, column=5).value is not None:
                            ws.cell(row=r, column=5).value = None
                            row_changed = True
                        if ws.cell(row=r, column=8).value is not None:
                            ws.cell(row=r, column=8).value = None
                            row_changed = True
                        if ws.cell(row=r, column=9).value is not None:
                            ws.cell(row=r, column=9).value = None
                            row_changed = True

                    if row_changed:
                        rows_to_highlight.add(r)
                        group_changed_rows += 1

                # Rebuild mapping after potential job_no updates
                existing_jobs_by_pi = {}
                existing_pi_fulls = set()
                for r in range(first_job_row, last_job_row + 1):
                    pi_full_val = str(ws.cell(row=r, column=3).value or "").strip()
                    if not pi_full_val:
                        continue
                    existing_pi_fulls.add(pi_full_val)
                    job_val = str(ws.cell(row=r, column=2).value or "").strip()
                    if pi_full_val not in existing_jobs_by_pi:
                        existing_jobs_by_pi[pi_full_val] = set()
                    if job_val:
                        existing_jobs_by_pi[pi_full_val].add(job_val)

                # Insert missing rows for new PI entries / new jobs
                insert_pos = debit_row or charge_row or exch_row or total_row or (last_job_row + 1)
                for pi_key in pi_keys:
                    if pi_key not in self.dict_pi:
                        continue
                    _, _, pi_full, pi_month, pi_qty, prefix = self.dict_pi[pi_key]

                    if prefix in self.dict_payment:
                        for job_no, pay_data in self.dict_payment[prefix].items():
                            if job_no in existing_jobs_by_pi.get(pi_full, set()):
                                continue
                            pay_amt, inv_date, inv_qty = pay_data

                            ws.insert_rows(insert_pos)
                            inserted_rows += 1

                            ws.cell(row=insert_pos, column=1).value = group
                            ws.cell(row=insert_pos, column=2).value = job_no
                            ws.cell(row=insert_pos, column=3).value = pi_full
                            ws.cell(row=insert_pos, column=4).value = pi_month
                            date_cell = ws.cell(row=insert_pos, column=5)
                            date_cell.value = inv_date
                            date_cell.number_format = INVOICE_DATE_NUMBER_FORMAT
                            ws.cell(row=insert_pos, column=6).value = pi_qty
                            ws.cell(row=insert_pos, column=6).number_format = "#,##0"
                            ws.cell(row=insert_pos, column=8).value = inv_qty
                            ws.cell(row=insert_pos, column=8).number_format = "#,##0"
                            ws.cell(row=insert_pos, column=9).value = round(pay_amt, 2)
                            ws.cell(row=insert_pos, column=9).number_format = "#,##0.00"

                            rows_to_highlight.add(insert_pos)
                            group_changed_rows += 1
                            existing_pi_fulls.add(pi_full)
                            if pi_full not in existing_jobs_by_pi:
                                existing_jobs_by_pi[pi_full] = set()
                            existing_jobs_by_pi[pi_full].add(job_no)

                            insert_pos += 1
                            last_job_row += 1
                            if debit_row:
                                debit_row += 1
                            if charge_row:
                                charge_row += 1
                            if exch_row:
                                exch_row += 1
                            total_row += 1
                    else:
                        if pi_full in existing_pi_fulls:
                            continue

                        pi_job_no = self.dict_pi_job_nos.get(pi_key, "")
                        ws.insert_rows(insert_pos)
                        inserted_rows += 1

                        ws.cell(row=insert_pos, column=1).value = group
                        ws.cell(row=insert_pos, column=2).value = pi_job_no
                        ws.cell(row=insert_pos, column=3).value = pi_full
                        ws.cell(row=insert_pos, column=4).value = pi_month
                        ws.cell(row=insert_pos, column=6).value = pi_qty
                        ws.cell(row=insert_pos, column=6).number_format = "#,##0"
                        ws.cell(row=insert_pos, column=6).fill = YELLOW_FILL

                        rows_to_highlight.add(insert_pos)
                        group_changed_rows += 1
                        existing_pi_fulls.add(pi_full)
                        insert_pos += 1
                        last_job_row += 1
                        if debit_row:
                            debit_row += 1
                        if charge_row:
                            charge_row += 1
                        if exch_row:
                            exch_row += 1
                        total_row += 1

                # Ensure a debit row exists if we have debit data but no debit row in the sheet
                if group in self.dict_debit and debit_row is None:
                    debit_insert_pos = charge_row or exch_row or total_row
                    if debit_insert_pos:
                        ws.insert_rows(debit_insert_pos)
                        inserted_rows += 1
                        debit_row = debit_insert_pos
                        if charge_row:
                            charge_row += 1
                        if exch_row:
                            exch_row += 1
                        total_row += 1

                # Update debit note row if changed
                if group in self.dict_debit and debit_row:
                    debit_full, debit_num, debit_amt = self.dict_debit[group]
                    old_credit = safe_float(ws.cell(row=debit_row, column=7).value)
                    old_debit = safe_float(ws.cell(row=debit_row, column=9).value)
                    old_full = ws.cell(row=debit_row, column=2).value

                    ws.cell(row=debit_row, column=1).value = group
                    ws.cell(row=debit_row, column=2).value = debit_full
                    if debit_amt < 0:
                        ws.cell(row=debit_row, column=7).value = round(abs(debit_amt), 2)
                        ws.cell(row=debit_row, column=7).number_format = "#,##0.00"
                        ws.cell(row=debit_row, column=9).value = None
                    else:
                        ws.cell(row=debit_row, column=9).value = round(debit_amt, 2)
                        ws.cell(row=debit_row, column=9).number_format = "#,##0.00"
                        ws.cell(row=debit_row, column=7).value = None

                    if old_full != debit_full or abs(old_credit - safe_float(ws.cell(row=debit_row, column=7).value)) > 0.01 or abs(old_debit - safe_float(ws.cell(row=debit_row, column=9).value)) > 0.01:
                        rows_to_highlight.add(debit_row)
                        group_changed_rows += 1

                # Update charge to purchase if changed
                if charge_row:
                    ws.cell(row=charge_row, column=1).value = group
                    ws.cell(row=charge_row, column=2).value = "CHARGE TO PURCHASE"
                    normalized_grp = normalize_group(group)
                    if normalized_grp in self.dict_charge:
                        credit, debit = self.dict_charge[normalized_grp]
                        old_credit = safe_float(ws.cell(row=charge_row, column=7).value)
                        old_debit = safe_float(ws.cell(row=charge_row, column=9).value)

                        ws.cell(row=charge_row, column=7).value = credit if credit != 0 else None
                        ws.cell(row=charge_row, column=9).value = debit if debit != 0 else None
                        if credit != 0:
                            ws.cell(row=charge_row, column=7).number_format = "#,##0.00"
                        if debit != 0:
                            ws.cell(row=charge_row, column=9).number_format = "#,##0.00"

                        if abs(old_credit - credit) > 0.01 or abs(old_debit - debit) > 0.01:
                            rows_to_highlight.add(charge_row)
                            group_changed_rows += 1

                # Recalculate total row formulas (row insertions won't reliably adjust them)
                exch_for_formula = exch_row or (total_row - 1)
                ws.cell(row=total_row, column=6).value = f"=F{header_row}"
                ws.cell(row=total_row, column=6).number_format = "#,##0"
                ws.cell(row=total_row, column=7).value = f"=ROUND(SUM(G{header_row}:G{exch_for_formula}),2)"
                ws.cell(row=total_row, column=7).number_format = "#,##0.00"
                ws.cell(row=total_row, column=8).value = f"=SUM(H{first_job_row}:H{last_job_row})"
                ws.cell(row=total_row, column=8).number_format = "#,##0"
                ws.cell(row=total_row, column=9).value = f"=ROUND(SUM(I{first_job_row}:I{exch_for_formula}),2)"
                ws.cell(row=total_row, column=9).number_format = "#,##0.00"
                ws.cell(row=total_row, column=10).value = f"=ROUND(G{total_row}-I{total_row},2)"
                ws.cell(row=total_row, column=10).number_format = "#,##0.00"
                ws.cell(row=total_row, column=11).value = f"=ROUND(F{total_row}-H{total_row},2)"
                ws.cell(row=total_row, column=11).number_format = "#,##0"
                for col in range(1, 12):
                    ws.cell(row=total_row, column=col).font = ITALIC_FONT

                group_ranges[group] = (header_row, first_job_row, last_job_row, debit_row if debit_row else first_job_row, total_row)

                # Shift cached row indices for groups below this one.
                if inserted_rows and group in existing_index:
                    idx0 = existing_index[group]
                    for g2 in existing_sorted[idx0 + 1:]:
                        info2 = existing_groups[g2]
                        for key in ("header", "first_job", "last_job", "debit", "charge", "exch", "total"):
                            if info2.get(key) is not None:
                                info2[key] += inserted_rows

                if inserted_rows or group_changed_rows:
                    self.log(
                        f"    -> {group}: inserted {inserted_rows:,} row(s), updated {group_changed_rows:,} row(s) "
                        f"(job rows now {first_job_row:,}-{last_job_row:,})"
                    )
                updated_groups += 1
                total_rows_inserted += inserted_rows

        # Apply yellow highlighting (duplicates / unmatched) across all job rows
        self.log("Applying yellow attention highlights (duplicates/unmatched)...")
        yellow_t0 = time.perf_counter()
        groups_total = len(group_ranges)
        last_progress = time.monotonic()
        for idx, (group, (_, first_job_row, last_job_row, _, _)) in enumerate(group_ranges.items(), 1):
            now = time.monotonic()
            if idx == 1 or idx == groups_total or (now - last_progress) >= 5.0 or (idx % 50 == 0):
                self.log(f"  ... yellow pass {idx:,}/{groups_total:,}")
                last_progress = now
            for r in range(first_job_row, last_job_row + 1):
                pi_full_val = str(ws.cell(row=r, column=3).value or "").strip()
                if not pi_full_val:
                    continue
                prefix = extract_prefix(pi_full_val)
                needs_attention = (not prefix) or (prefix not in self.dict_payment) or (prefix in self.dict_duplicate_prefixes)

                qty_cell = ws.cell(row=r, column=6)
                try:
                    rgb = str(getattr(getattr(qty_cell.fill, "start_color", None), "rgb", "") or "").upper()
                    is_yellow = rgb.endswith("FFFF00")
                except Exception:
                    is_yellow = False

                if needs_attention:
                    qty_cell.fill = YELLOW_FILL
                else:
                    if is_yellow:
                        qty_cell.fill = NO_FILL
        self.log(f"Yellow highlights applied in {time.perf_counter() - yellow_t0:.1f}s")

        # Apply green highlighting to updated rows
        self.log(f"Applying green highlight to {len(rows_to_highlight):,} updated/new row(s)...")
        green_t0 = time.perf_counter()
        last_progress = time.monotonic()
        rows_total = len(rows_to_highlight)
        for idx, row in enumerate(rows_to_highlight, 1):
            now = time.monotonic()
            if rows_total >= 2000 and ((now - last_progress) >= 5.0 or idx == rows_total):
                self.log(f"  ... green applied to {idx:,}/{rows_total:,} row(s)")
                last_progress = now
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                # Don't override yellow
                try:
                    rgb = str(getattr(getattr(cell.fill, "start_color", None), "rgb", "") or "").upper()
                    if rgb.endswith("FFFF00"):
                        continue
                except Exception:
                    pass
                cell.fill = GREEN_FILL
        self.log(f"Green highlights applied in {time.perf_counter() - green_t0:.1f}s")

        # Apply borders
        last_data_row = self._find_last_data_row(ws, start_row=4, max_col=MONITOR_DATA_MAX_COL)
        self.log(f"Applying borders to monitoring sheet (rows 4-{last_data_row})...")
        borders_t0 = time.perf_counter()
        last_progress = time.monotonic()
        for row in range(4, last_data_row + 1):
            now = time.monotonic()
            if now - last_progress >= 5.0:
                self.log(f"  ... borders applied up to row {row:,}/{last_data_row:,}")
                last_progress = now
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.border = THIN_BORDER
                if col == 5:
                    cell.number_format = INVOICE_DATE_NUMBER_FORMAT
        self.log(f"Borders applied in {time.perf_counter() - borders_t0:.1f}s")

        # Auto-fit columns
        self.log("Auto-sizing columns A-K...")
        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].auto_size = True

        # Process Zero Balance Groups
        self.log("Processing Zero Balance Groups...")
        self._process_zero_balance_groups(wb, ws, group_ranges)

        # Compress blank rows (keep only one consecutive blank row between groups)
        self.log("Compressing consecutive blank rows...")
        self._compress_blank_rows(ws)

        # Rebuild Total row formulas after any row deletions (openpyxl does not auto-adjust formulas).
        self.log("Rebuilding Total formulas after row moves/deletions...")
        self._rebuild_total_row_formulas(ws)

        # Trim trailing blank rows (safe: does not move data rows)
        self.log("Trimming trailing blank rows...")
        self._trim_trailing_blank_rows(ws)

        self.log(f"Monitoring workbook update finished in {time.perf_counter() - t0:.1f}s")
        return wb

    def _process_zero_balance_groups(self, wb: Workbook, monitor_ws, group_ranges: dict):
        """Move groups with zero balance to separate sheet."""

        # Create or clear Zero Balance Groups sheet
        if "Zero Balance Groups" in wb.sheetnames:
            zero_ws = wb["Zero Balance Groups"]
            zero_ws.delete_rows(1, zero_ws.max_row)
        else:
            zero_ws = wb.create_sheet("Zero Balance Groups")

        # Add header
        zero_ws.cell(row=1, column=1).value = "Groups with Zero Balance"

        # Copy headers
        for col in range(1, 12):
            zero_ws.cell(row=2, column=col).value = monitor_ws.cell(row=4, column=col).value
            zero_ws.cell(row=2, column=col).fill = GREY_FILL
            zero_ws.cell(row=2, column=col).font = BOLD_FONT

        paste_row = 3
        ranges_to_delete = []
        moved_groups = 0
        last_progress = time.monotonic()
        groups_total = len(group_ranges)
        self.log(f"  Checking {groups_total:,} group(s) for zero balance...")

        for idx, (group, (header_row, first_job, last_job, debit_row, total_row)) in enumerate(group_ranges.items(), 1):
            now = time.monotonic()
            if idx == 1 or idx == groups_total or (now - last_progress) >= 5.0 or (idx % 50 == 0):
                self.log(f"    [{idx:,}/{groups_total:,}] zero-balance scan... (moved so far: {moved_groups:,})")
                last_progress = now

            # openpyxl does not calculate formulas, so compute balance from the underlying
            # credit/debit numbers in the group block (header..exchange row).
            credit_total = 0.0
            debit_total = 0.0
            for r in range(header_row, total_row):
                credit_total += safe_float(monitor_ws.cell(row=r, column=7).value)
                debit_total += safe_float(monitor_ws.cell(row=r, column=9).value)
            balance = round(credit_total - debit_total, 2)

            if abs(balance) < 0.01:
                # Copy group to zero balance sheet
                row_delta = paste_row - header_row
                for src_row in range(header_row, total_row + 1):
                    for col in range(1, 12):
                        src_cell = monitor_ws.cell(row=src_row, column=col)
                        dst_cell = zero_ws.cell(row=paste_row, column=col)
                        value = src_cell.value
                        if isinstance(value, str) and value.startswith("=") and row_delta:
                            def _shift_ref(m):
                                return f"{m.group(1)}{int(m.group(2)) + row_delta}"
                            value = re.sub(r"(\$?[A-Z]{1,3}\$?)(\d+)", _shift_ref, value)
                        dst_cell.value = value
                        dst_cell.fill = copy(src_cell.fill)
                        dst_cell.font = copy(src_cell.font)
                        dst_cell.border = copy(src_cell.border)
                        dst_cell.alignment = copy(src_cell.alignment)
                        dst_cell.number_format = src_cell.number_format
                    paste_row += 1
                paste_row += 1  # Spacing

                ranges_to_delete.append((header_row, total_row))
                moved_groups += 1

        # Delete from monitor sheet (in reverse order to avoid index issues)
        ranges_to_delete.sort(reverse=True, key=lambda x: x[0])
        for start_row, end_row in ranges_to_delete:
            monitor_ws.delete_rows(start_row, end_row - start_row + 1)

        self.log(f"  Moved {moved_groups:,} group(s) to 'Zero Balance Groups'")

        # Auto-fit zero balance sheet
        for col in range(1, 12):
            zero_ws.column_dimensions[get_column_letter(col)].auto_size = True

    def _trim_trailing_blank_rows(self, ws):
        """Remove trailing blank rows at the bottom, leaving at most one."""
        last_data_row = self._find_last_data_row(ws, start_row=4, max_col=MONITOR_DATA_MAX_COL)
        keep_until = last_data_row + 1  # keep a single blank row for readability
        if ws.max_row > keep_until:
            ws.delete_rows(keep_until + 1, ws.max_row - keep_until)

    def _compress_blank_rows(self, ws):
        """Remove consecutive blank rows, leaving only one (between groups)."""
        # Use last real data row (avoid ws.max_row inflation from formatting).
        last_data_row = self._find_last_data_row(ws, start_row=4, max_col=MONITOR_DATA_MAX_COL)
        last_row = max(last_data_row + 1, 4)  # allow one blank row after last data row

        blank_count = 0
        for row in range(last_row, 4, -1):
            is_blank = True
            for col in range(1, MONITOR_DATA_MAX_COL + 1):
                if _has_meaningful_value(ws.cell(row=row, column=col).value):
                    is_blank = False
                    break

            if is_blank:
                blank_count += 1
                if blank_count > 1:
                    ws.delete_rows(row)
            else:
                blank_count = 0


def run_process(payment_path: str, pi_path: str, debit_path: str, charge_path: str,
                monitor_path: Optional[str], log_emit) -> Tuple[str, bool]:
    """
    Main processing function.

    Args:
        payment_path: Path to Payment file
        pi_path: Path to PI file (must have 'NK' sheet)
        debit_path: Path to Debit Note file
        charge_path: Path to Charge to Purchase file
        monitor_path: Path to existing monitoring workbook (optional)
        log_emit: Function to emit log messages

    Returns:
        Tuple of (output_path, success)
    """
    try:
        t0 = time.perf_counter()
        try:
            log_emit(
                "Inputs: "
                f"Payment='{os.path.basename(payment_path)}', "
                f"PI='{os.path.basename(pi_path)}', "
                f"Debit='{os.path.basename(debit_path)}', "
                f"Charge='{os.path.basename(charge_path)}', "
                f"Workbook={'(new)' if not monitor_path else os.path.basename(monitor_path)}"
            )
        except Exception:
            pass

        processor = VTECProcessor(log_emit)

        # Parse source files
        processor.parse_payment_file(payment_path)
        processor.parse_pi_file(pi_path)
        processor.parse_debit_file(debit_path)
        processor.parse_charge_file(charge_path)

        # Open or create monitoring workbook
        if monitor_path and os.path.exists(monitor_path):
            log_emit(f"Opening existing workbook: {os.path.basename(monitor_path)}")
            ext = os.path.splitext(monitor_path)[1].lower()
            keep_vba = ext in (".xlsm", ".xltm", ".xlam")
            wb = load_workbook(monitor_path, data_only=False, keep_vba=keep_vba)
        else:
            log_emit("Creating new monitoring workbook")
            wb = Workbook()
            # Remove default sheet if we're creating fresh
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # Process
        wb = processor.process_monitoring_workbook(wb)

        # Determine output path
        if monitor_path:
            # Save beside the original monitoring workbook, but never overwrite it.
            base_dir = os.path.dirname(monitor_path) or os.getcwd()
            base_name = os.path.basename(monitor_path)
            stem, ext = os.path.splitext(base_name)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            candidate = os.path.join(base_dir, f"{stem}_UPDATED_{timestamp}{ext}")
            output_path = _ensure_unique_path(candidate)
        else:
            # Generate new file in same directory as payment file
            base_dir = os.path.dirname(payment_path)
            output_path = os.path.join(base_dir, "VTEC_Monitoring_Chart.xlsx")

        # Save
        wb.save(output_path)
        wb.close()

        log_emit(f"Saved to: {output_path}")
        log_emit(f"Total runtime: {time.perf_counter() - t0:.1f}s")
        return output_path, True

    except Exception as e:
        log_emit(f"ERROR: {e}")
        import traceback
        log_emit(traceback.format_exc())
        return "", False


# --- UI Class ---

class MainWidget(QWidget):
    log_message = Signal(str)
    processing_done = Signal(str, bool)

    def __init__(self):
        super().__init__()
        self.setObjectName("vtec_widget")
        self._build_ui()
        self._connect_signals()

        # File paths
        self.payment_path = ""
        self.pi_path = ""
        self.debit_path = ""
        self.charge_path = ""
        self.monitor_path = ""

    def _build_ui(self):
        self.desc_label = QLabel("", self)
        self.desc_label.setWordWrap(True)
        self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.desc_label.setStyleSheet(
            "color: #dcdcdc; background: transparent; padding: 6px; "
            "border: 1px solid #3a3a3a; border-radius: 6px;"
        )
        self.set_long_description("")

        # Buttons for source files
        self.select_payment_btn = PrimaryPushButton("Select Payment File", self)
        self.select_pi_btn = PrimaryPushButton("Select PI File", self)
        self.select_debit_btn = PrimaryPushButton("Select Debit Note File", self)
        self.select_charge_btn = PrimaryPushButton("Select Charge to Purchase File", self)

        # Button for monitoring workbook (optional)
        self.select_monitor_btn = PrimaryPushButton("Select Monitoring Workbook (Optional)", self)

        # Run button
        self.run_btn = PrimaryPushButton("Run", self)

        # Labels
        self.payment_label = QLabel("Payment file: (not selected)", self)
        self.pi_label = QLabel("PI file: (not selected)", self)
        self.debit_label = QLabel("Debit Note file: (not selected)", self)
        self.charge_label = QLabel("Charge to Purchase file: (not selected)", self)
        self.monitor_label = QLabel("Monitoring workbook: (will create new)", self)

        for lbl in [self.payment_label, self.pi_label, self.debit_label,
                    self.charge_label, self.monitor_label]:
            lbl.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")
            lbl.setWordWrap(True)

        # Log box
        self.log_label = QLabel("Process log:", self)
        self.log_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setStyleSheet(
            "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        # Layout
        main_layout = QVBoxLayout(self)
        main_layout.addWidget(self.desc_label)

        # Source file buttons row 1
        row1 = QHBoxLayout()
        row1.addWidget(self.select_payment_btn)
        row1.addWidget(self.select_pi_btn)
        main_layout.addLayout(row1)

        # Source file buttons row 2
        row2 = QHBoxLayout()
        row2.addWidget(self.select_debit_btn)
        row2.addWidget(self.select_charge_btn)
        main_layout.addLayout(row2)

        # Monitoring workbook button
        row3 = QHBoxLayout()
        row3.addWidget(self.select_monitor_btn)
        main_layout.addLayout(row3)

        # File labels
        main_layout.addWidget(self.payment_label)
        main_layout.addWidget(self.pi_label)
        main_layout.addWidget(self.debit_label)
        main_layout.addWidget(self.charge_label)
        main_layout.addWidget(self.monitor_label)

        # Run button
        run_row = QHBoxLayout()
        run_row.addStretch()
        run_row.addWidget(self.run_btn)
        run_row.addStretch()
        main_layout.addLayout(run_row)

        # Log
        main_layout.addWidget(self.log_label)
        main_layout.addWidget(self.log_box, 2)

    def set_long_description(self, text: str):
        clean = (text or "").strip()
        if clean:
            self.desc_label.setText(clean)
            self.desc_label.show()
        else:
            self.desc_label.hide()

    def _connect_signals(self):
        self.select_payment_btn.clicked.connect(self.select_payment_file)
        self.select_pi_btn.clicked.connect(self.select_pi_file)
        self.select_debit_btn.clicked.connect(self.select_debit_file)
        self.select_charge_btn.clicked.connect(self.select_charge_file)
        self.select_monitor_btn.clicked.connect(self.select_monitor_file)
        self.run_btn.clicked.connect(self.run_process)
        self.log_message.connect(self.append_log)
        self.processing_done.connect(self.on_processing_done)

    def select_payment_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Payment File", "",
            "Excel Files (*.xlsx *.xlsm *.xls)"
        )
        if path:
            self.payment_path = path
            self.payment_label.setText(f"Payment file: {os.path.basename(path)}")

    def select_pi_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select PI File (must have 'NK' sheet)", "",
            "Excel Files (*.xlsx *.xlsm *.xls)"
        )
        if path:
            self.pi_path = path
            self.pi_label.setText(f"PI file: {os.path.basename(path)}")

    def select_debit_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Debit Note File", "",
            "Excel Files (*.xlsx *.xlsm *.xls)"
        )
        if path:
            self.debit_path = path
            self.debit_label.setText(f"Debit Note file: {os.path.basename(path)}")

    def select_charge_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Charge to Purchase File", "",
            "Excel Files (*.xlsx *.xlsm *.xls)"
        )
        if path:
            self.charge_path = path
            self.charge_label.setText(f"Charge to Purchase file: {os.path.basename(path)}")

    def select_monitor_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Monitoring Workbook (contains 'VTEC Monitoring Chart' and 'Zero Balance Groups' sheets)", "",
            "Excel Files (*.xlsx *.xlsm)"
        )
        if path:
            self.monitor_path = path
            self.monitor_label.setText(f"Monitoring workbook: {os.path.basename(path)}")
        else:
            self.monitor_path = ""
            self.monitor_label.setText("Monitoring workbook: (will create new)")

    def run_process(self):
        # Validate required files
        if not self.payment_path:
            MessageBox("Warning", "Please select a Payment file.", self).exec()
            return
        if not self.pi_path:
            MessageBox("Warning", "Please select a PI file.", self).exec()
            return
        if not self.debit_path:
            MessageBox("Warning", "Please select a Debit Note file.", self).exec()
            return
        if not self.charge_path:
            MessageBox("Warning", "Please select a Charge to Purchase file.", self).exec()
            return

        self.log_box.clear()
        self.log_message.emit("Process Started...")

        # Disable buttons
        self.run_btn.setEnabled(False)
        self.select_payment_btn.setEnabled(False)
        self.select_pi_btn.setEnabled(False)
        self.select_debit_btn.setEnabled(False)
        self.select_charge_btn.setEnabled(False)
        self.select_monitor_btn.setEnabled(False)

        def worker():
            try:
                output_path, success = run_process(
                    self.payment_path,
                    self.pi_path,
                    self.debit_path,
                    self.charge_path,
                    self.monitor_path if self.monitor_path else None,
                    self.log_message.emit
                )
                self.processing_done.emit(output_path, success)
            except Exception as e:
                self.log_message.emit(f"CRITICAL ERROR: {e}")
                self.processing_done.emit("", False)

        threading.Thread(target=worker, daemon=True).start()

    def append_log(self, text):
        self.log_box.append(text)

    def on_processing_done(self, output_path: str, success: bool):
        # Re-enable buttons
        self.run_btn.setEnabled(True)
        self.select_payment_btn.setEnabled(True)
        self.select_pi_btn.setEnabled(True)
        self.select_debit_btn.setEnabled(True)
        self.select_charge_btn.setEnabled(True)
        self.select_monitor_btn.setEnabled(True)

        if success:
            self.log_message.emit("Process completed successfully!")
            msg = MessageBox(
                "Success",
                f"VTEC Monitoring Chart updated.\n\nSaved to:\n{output_path}",
                self
            )
            msg.yesButton.setText("OK")
            msg.cancelButton.hide()
            msg.exec()
        else:
            self.log_message.emit("Process failed. See log for details.")
            msg = MessageBox("Error", "Process failed. Check the log for details.", self)
            msg.yesButton.setText("OK")
            msg.cancelButton.hide()
            msg.exec()


def get_widget():
    return MainWidget()
