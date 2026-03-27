#!/usr/bin/env python3
from __future__ import annotations

import argparse
import os
import re
import threading
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import quote_sheetname
from openpyxl.utils.datetime import from_excel

try:
    from PySide6.QtCore import Qt, Signal
    from PySide6.QtWidgets import (
        QFileDialog,
        QHBoxLayout,
        QLabel,
        QSizePolicy,
        QTextEdit,
        QVBoxLayout,
        QWidget,
    )
    from qfluentwidgets import MessageBox, PrimaryPushButton

    GUI_AVAILABLE = True
except Exception:
    GUI_AVAILABLE = False


SCAN_MAX_COL = 64
HEADER_SCAN_ROWS = 10
VND_NUMBER_FORMAT = "#,##0"
USD_NUMBER_FORMAT = "#,###.00"
GENERAL_RATE_FORMAT = "#,##0.00"
DATE_NUMBER_FORMAT = "yyyy/mm/dd"
SUMMARY_DATE_DISPLAY_FORMAT = "%-m/%-d/%Y"
SUMMARY_DATE_FALLBACK_FORMAT = "%m/%d/%Y"
DATA_ROW_HEIGHT = 18
SUMMARY_DATA_ROW_HEIGHT = 24
MONTH_ABBR = {
    1: "JAN",
    2: "FEB",
    3: "MAR",
    4: "APR",
    5: "MAY",
    6: "JUN",
    7: "JUL",
    8: "AUG",
    9: "SEP",
    10: "OCT",
    11: "NOV",
    12: "DEC",
}


@dataclass(frozen=True)
class ColumnSpec:
    key: str
    title: str
    required_header: bool
    matcher: Callable[[str], bool]


@dataclass
class SourceRow:
    source_file: Path
    sheet_name: str
    row_number: int
    payment_to_supplier_date: date
    vat_invoice_date: date
    values: Dict[str, object]


@dataclass
class AuditLog:
    folders_without_xlsx: List[str] = field(default_factory=list)
    multi_sheet_files: List[str] = field(default_factory=list)
    source_file_errors: List[str] = field(default_factory=list)
    missing_required_headers: List[str] = field(default_factory=list)
    file_level_row_issues: List[str] = field(default_factory=list)
    invalid_payment_to_supplier_rows: List[str] = field(default_factory=list)
    vcb_file_errors: List[str] = field(default_factory=list)
    duplicate_vcb_rates: List[str] = field(default_factory=list)
    missing_payment_rates: Dict[str, List[str]] = field(default_factory=lambda: defaultdict(list))
    missing_vat_rates: Dict[str, List[str]] = field(default_factory=lambda: defaultdict(list))


@dataclass
class ProcessResult:
    workbook_path: Path
    log_path: Path
    source_files_found: int
    source_files_imported: int
    source_files_skipped: int
    rows_imported: int
    rows_ignored_non_date_payment_to_supplier: int
    month_sheet_count: int


def _emit(log_emit: Callable[[str], None] | None, text: str):
    if callable(log_emit):
        try:
            log_emit(text)
            return
        except Exception:
            pass
    print(text)


def ensure_unique_path(path: Path) -> Path:
    if not path.exists():
        return path

    counter = 1
    while True:
        candidate = path.with_name(f"{path.stem} ({counter}){path.suffix}")
        if not candidate.exists():
            return candidate
        counter += 1


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text.upper()


def starts_with_header(actual: str, expected: str) -> bool:
    if not actual or not expected:
        return False
    if actual == expected:
        return True
    if not actual.startswith(expected):
        return False
    if len(actual) == len(expected):
        return True
    next_char = actual[len(expected)]
    return not next_char.isalnum()


def match_po_amount(norm_header: str, currency_suffix: str) -> bool:
    return norm_header.startswith("PO AMOUNT") and norm_header.endswith(currency_suffix)


SOURCE_COLUMNS: List[ColumnSpec] = [
    ColumnSpec("no", "No", False, lambda s: starts_with_header(s, "NO")),
    ColumnSpec("supplier_name", "Supplier Name", False, lambda s: starts_with_header(s, "SUPPLIER NAME")),
    ColumnSpec("payment_term", "Payment Term", False, lambda s: starts_with_header(s, "PAYMENT TERM")),
    ColumnSpec("fabric_accessories", "Fabric / Accessories", False, lambda s: starts_with_header(s, "FABRIC / ACCESSORIES")),
    ColumnSpec("order_type", "Order Type", False, lambda s: starts_with_header(s, "ORDER TYPE")),
    ColumnSpec("vat_invoice_date", "VAT Invoice Date", True, lambda s: starts_with_header(s, "VAT INVOICE DATE")),
    ColumnSpec("vat_invoice_number", "VAT Invoice Number", False, lambda s: starts_with_header(s, "VAT INVOICE NUMBER")),
    ColumnSpec("other_references_number", "Other References Number", False, lambda s: starts_with_header(s, "OTHER REFERENCES NUMBER")),
    ColumnSpec("po_amount_before_vat_vnd", "PO Amount BEFORE VAT (VND)", True, lambda s: match_po_amount(s, "(VND)")),
    ColumnSpec("surcharge_other_vnd", "Surcharge / Other (VND)", False, lambda s: starts_with_header(s, "SURCHARGE / OTHER (VND)")),
    ColumnSpec("total_vnd", "Total (VND)", False, lambda s: starts_with_header(s, "TOTAL (VND)")),
    ColumnSpec("currency_rate", "Currency Rate", False, lambda s: starts_with_header(s, "CURRENCY RATE")),
    ColumnSpec("po_amount_before_vat_usd", "PO Amount BEFORE VAT (USD)", True, lambda s: match_po_amount(s, "(USD)")),
    ColumnSpec("surcharge_other_usd", "Surcharge / Other (USD)", False, lambda s: starts_with_header(s, "SURCHARGE / OTHER (USD)")),
    ColumnSpec("total_usd", "Total (USD)", False, lambda s: starts_with_header(s, "TOTAL (USD)")),
    ColumnSpec("purchaser", "Purchaser", False, lambda s: starts_with_header(s, "PURCHASER")),
    ColumnSpec("from", "From", False, lambda s: starts_with_header(s, "FROM")),
    ColumnSpec("to", "To", False, lambda s: starts_with_header(s, "TO")),
    ColumnSpec("sig_due_date", "SIG Due Date", False, lambda s: starts_with_header(s, "SIG DUE DATE") or starts_with_header(s, "DUE DATE")),
    ColumnSpec("payment_to_vtec", "Payment to VTEC", False, lambda s: starts_with_header(s, "PAYMENT TO VTEC")),
    ColumnSpec("payment_to_supplier", "Payment to Supplier", True, lambda s: starts_with_header(s, "PAYMENT TO SUPPLIER")),
]
SOURCE_COLUMN_BY_KEY = {spec.key: spec for spec in SOURCE_COLUMNS}

ORIGINAL_OUTPUT_HEADERS = [spec.title for spec in SOURCE_COLUMNS]
CALC_HEADERS_ROW1 = [
    "VTEC have to sell the USD to VCB bank at VCB's buying rate and then pay to suppliers in VND",
    "Red Invoice amount Before VAT",
    "Gain/(Loss) in Forex",
    "VCB's Buying Rate at VAT Invoice Date",
    "Gain/(Loss) in Forex",
    "Gain/(Loss) in Forex",
]
CALC_HEADERS_ROW2 = [
    "USD",
    "Rate (VND/USD)",
    "VND",
    "VND",
    "VND",
    "VND/USD",
    "VND",
    "USD",
]

OUTPUT_COL_USD = len(ORIGINAL_OUTPUT_HEADERS) + 1
OUTPUT_COL_PAYMENT_RATE = OUTPUT_COL_USD + 1
OUTPUT_COL_PAYMENT_VND = OUTPUT_COL_USD + 2
OUTPUT_COL_RED_INVOICE_VND = OUTPUT_COL_USD + 3
OUTPUT_COL_FOREX_PAYMENT_VND = OUTPUT_COL_USD + 4
OUTPUT_COL_VAT_RATE = OUTPUT_COL_USD + 5
OUTPUT_COL_FOREX_VAT_VND = OUTPUT_COL_USD + 6
OUTPUT_COL_FOREX_VAT_USD = OUTPUT_COL_USD + 7
TOTAL_OUTPUT_COLUMNS = OUTPUT_COL_FOREX_VAT_USD

SUMMARY_COL_MONTH = 1
SUMMARY_COL_SIG_USD = 2
SUMMARY_COL_SIG_VND = 3
SUMMARY_COL_SIG_AVG_RATE = 4
SUMMARY_COL_PAYMENT_DATE = 5
SUMMARY_COL_BANK_USD = 6
SUMMARY_COL_BANK_RATE = 7
SUMMARY_COL_BANK_VND = 8
SUMMARY_COL_PAYMENT_LOSS_VND = 9
SUMMARY_COL_PAYMENT_LOSS_USD = 10
SUMMARY_COL_VAT_LOSS_VND = 11
SUMMARY_COL_VAT_LOSS_USD = 12

MONTHLY_HEADER_FILL = "FFF200"
MONTHLY_GROUP_FILL = "D9E7F5"
MONTHLY_SUBGROUP_FILL = "EAF2FB"
MONTHLY_GREEN_FILL = "E2F0D9"
MONTHLY_RATE_FILL = "FCE4D6"
MONTHLY_WHITE_FILL = "FFFFFF"
MONTHLY_CALC_FILL = "DCE6F1"
MONTHLY_GREEN_DATA_FILL = "EEF6E7"
SUMMARY_YELLOW_FILL = "FFF200"
SUMMARY_BLUE_FILL = "D9E7F5"
SUMMARY_GREEN_FILL = "E2F0D9"
SUMMARY_TOTAL_FILL = "F2F2F2"
BLACK = "000000"
RED = "FF0000"

VND_KEYS = {"po_amount_before_vat_vnd", "surcharge_other_vnd", "total_vnd"}
USD_KEYS = {"po_amount_before_vat_usd", "surcharge_other_usd", "total_usd"}
DATE_KEYS = {"vat_invoice_date", "sig_due_date", "payment_to_vtec", "payment_to_supplier"}


def is_blank(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def trim_trailing_blank(values: Iterable[object]) -> List[object]:
    trimmed = list(values)
    while trimmed and is_blank(trimmed[-1]):
        trimmed.pop()
    return trimmed


def read_row_values(ws, row_index: int, max_col: int = SCAN_MAX_COL) -> List[object]:
    row = next(
        ws.iter_rows(
            min_row=row_index,
            max_row=row_index,
            min_col=1,
            max_col=max_col,
            values_only=True,
        ),
        (),
    )
    return trim_trailing_blank(row)


def find_header_row_from_rows(scanned_rows: List[List[object]]) -> tuple[int, Dict[str, int]] | tuple[None, None]:
    best_row = None
    best_map = None
    best_score = -1

    for row_index, row_values in enumerate(scanned_rows, start=1):
        if not row_values:
            continue
        mapping = build_header_map(row_values)
        if "no" not in mapping:
            continue
        if "supplier_name" not in mapping and "vat_invoice_date" not in mapping:
            continue
        score = len(mapping)
        if score > best_score:
            best_score = score
            best_row = row_index
            best_map = mapping

    if best_row is None or best_map is None:
        return None, None
    return best_row, best_map


def build_header_map(row_values: List[object]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for col_index, cell_value in enumerate(row_values, start=1):
        norm = normalize_header(cell_value)
        if not norm:
            continue
        for spec in SOURCE_COLUMNS:
            if spec.key in mapping:
                continue
            if spec.matcher(norm):
                mapping[spec.key] = col_index
                break
    return mapping


def find_header_row(ws) -> tuple[int, Dict[str, int]] | tuple[None, None]:
    scanned_rows: List[List[object]] = []
    for row_values in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row, HEADER_SCAN_ROWS),
        min_col=1,
        max_col=SCAN_MAX_COL,
        values_only=True,
    ):
        scanned_rows.append(trim_trailing_blank(row_values))
    return find_header_row_from_rows(scanned_rows)


def row_value(row_values: List[object], one_based_index: Optional[int]) -> object:
    if not one_based_index:
        return None
    zero_based_index = one_based_index - 1
    if zero_based_index < 0 or zero_based_index >= len(row_values):
        return None
    return row_values[zero_based_index]


def parse_number(value: object) -> Optional[float]:
    if is_blank(value):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"
    try:
        return float(text)
    except Exception:
        return None


def parse_date_value(value: object, excel_epoch) -> Optional[date]:
    if is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            excel_dt = from_excel(value, excel_epoch)
            if isinstance(excel_dt, datetime):
                return excel_dt.date()
            if isinstance(excel_dt, date):
                return excel_dt
        except Exception:
            return None

    text = str(value).strip()
    if not text:
        return None

    text = re.sub(r"\s+", " ", text)
    formats = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d.%m.%Y",
        "%d.%m.%y",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%d %b %Y",
        "%d %B %Y",
        "%Y.%m.%d",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def normalize_output_value(key: str, value: object, excel_epoch) -> object:
    if key in DATE_KEYS:
        parsed = parse_date_value(value, excel_epoch)
        return parsed if parsed else value
    if key in VND_KEYS or key in USD_KEYS or key == "currency_rate":
        numeric = parse_number(value)
        return numeric if numeric is not None else value
    return value


def month_sheet_name(value: date) -> str:
    return f"{MONTH_ABBR[value.month]}'{value.year % 100:02d}"


def choose_numeric(primary: object, fallback: object) -> Optional[float]:
    if not is_blank(primary):
        primary_number = parse_number(primary)
        if primary_number is not None:
            return primary_number
    if not is_blank(fallback):
        return parse_number(fallback)
    return None


def build_output_paths(output_dir: Path) -> tuple[Path, Path]:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    workbook_path = ensure_unique_path(output_dir / f"VTEC_Payment_Monthly_Combined_{timestamp}.xlsx")
    log_path = workbook_path.with_name(f"{workbook_path.stem}_issues.txt")
    return workbook_path, log_path


def collect_source_files(input_folders: List[Path], audit_log: AuditLog) -> List[Path]:
    collected: List[Path] = []
    seen: set[Path] = set()

    for folder in input_folders:
        if not folder.is_dir():
            audit_log.source_file_errors.append(f"{folder} | folder does not exist or is not accessible")
            continue

        folder_files: List[Path] = []

        def handle_walk_error(exc: OSError):
            error_path = getattr(exc, "filename", None) or folder
            audit_log.source_file_errors.append(f"{error_path} | could not scan subfolder: {exc.strerror or exc}")

        for root, _dirs, files in os.walk(folder, onerror=handle_walk_error):
            root_path = Path(root)
            for file_name in sorted(files, key=str.lower):
                if file_name.startswith("~$") or not file_name.lower().endswith(".xlsx"):
                    continue
                file_path = root_path / file_name
                if file_path.is_file():
                    folder_files.append(file_path)

        folder_files.sort(key=lambda path: str(path.relative_to(folder)).lower())
        if not folder_files:
            audit_log.folders_without_xlsx.append(str(folder))
            continue

        for path in folder_files:
            resolved = path.resolve()
            if resolved not in seen:
                collected.append(path)
                seen.add(resolved)

    return collected


def aggregate_missing_rate(
    bucket: Dict[str, List[str]],
    missing_date: date,
    source_file: Path,
    row_number: int,
):
    key = missing_date.isoformat()
    if len(bucket[key]) < 5:
        bucket[key].append(f"{source_file.name} row {row_number}")


def load_vcb_usd_rates(vcb_files: List[Path], audit_log: AuditLog, log_emit: Callable[[str], None] | None = None) -> Dict[date, float]:
    rates: Dict[date, float] = {}

    for vcb_file in vcb_files:
        _emit(log_emit, f"Loading VCB rates: {vcb_file}")
        try:
            wb = load_workbook(vcb_file, read_only=True, data_only=True)
        except Exception as exc:
            audit_log.vcb_file_errors.append(f"{vcb_file} | could not open workbook: {exc}")
            continue

        try:
            matched_sheet = False
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_iter = ws.iter_rows(min_row=1, min_col=1, max_col=12, values_only=True)
                header_values = trim_trailing_blank(next(rows_iter, ()))
                header_map = {normalize_header(value): idx for idx, value in enumerate(header_values, start=1) if not is_blank(value)}
                required_headers = {
                    "DATE": None,
                    "CURRENCY CODE": None,
                    "TELEGRAPHIC BUYING": None,
                }
                for expected in list(required_headers.keys()):
                    for actual, idx in header_map.items():
                        if starts_with_header(actual, expected):
                            required_headers[expected] = idx
                            break
                if any(idx is None for idx in required_headers.values()):
                    continue

                matched_sheet = True
                for row_index, raw_values in enumerate(rows_iter, start=2):
                    row_values = trim_trailing_blank(raw_values)
                    if not row_values:
                        continue

                    currency_code = row_value(row_values, required_headers["CURRENCY CODE"])
                    if normalize_header(currency_code) != "USD":
                        continue

                    rate_date = parse_date_value(row_value(row_values, required_headers["DATE"]), wb.epoch)
                    telegraphic_buying = parse_number(row_value(row_values, required_headers["TELEGRAPHIC BUYING"]))
                    if rate_date is None or telegraphic_buying is None:
                        continue

                    if rate_date in rates and rates[rate_date] != telegraphic_buying:
                        audit_log.duplicate_vcb_rates.append(
                            f"{vcb_file} | duplicate USD rate for {rate_date.isoformat()} "
                            f"({rates[rate_date]} kept, {telegraphic_buying} ignored)"
                        )
                        continue

                    rates.setdefault(rate_date, telegraphic_buying)
                break

            if not matched_sheet:
                audit_log.vcb_file_errors.append(
                    f"{vcb_file} | no sheet with headers Date / Currency Code / Telegraphic Buying was found"
                )
        finally:
            wb.close()

    return rates


def parse_source_file(
    source_file: Path,
    audit_log: AuditLog,
    log_emit: Callable[[str], None] | None = None,
) -> tuple[List[SourceRow], int]:
    _emit(log_emit, f"Reading source workbook: {source_file}")

    try:
        wb = load_workbook(source_file, read_only=True, data_only=True)
    except Exception as exc:
        audit_log.source_file_errors.append(f"{source_file} | could not open workbook: {exc}")
        return [], 0

    try:
        if len(wb.sheetnames) != 1:
            audit_log.multi_sheet_files.append(f"{source_file} | sheets: {', '.join(wb.sheetnames)}")
            return [], 0

        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        scanned_rows: List[List[object]] = []
        for raw_values in ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row, HEADER_SCAN_ROWS),
            min_col=1,
            max_col=SCAN_MAX_COL,
            values_only=True,
        ):
            scanned_rows.append(trim_trailing_blank(raw_values))

        header_row, header_map = find_header_row_from_rows(scanned_rows)
        if header_row is None or header_map is None:
            audit_log.missing_required_headers.append(
                f"{source_file} | no recognizable header row was found within the first {HEADER_SCAN_ROWS} rows"
            )
            return [], 0

        missing_headers = [
            SOURCE_COLUMN_BY_KEY[key].title
            for key in ("vat_invoice_date", "po_amount_before_vat_vnd", "po_amount_before_vat_usd", "payment_to_supplier")
            if key not in header_map
        ]
        if missing_headers:
            audit_log.missing_required_headers.append(
                f"{source_file} | missing required header(s): {', '.join(missing_headers)}"
            )
            return [], 0

        rows: List[SourceRow] = []
        missing_vat_rows: List[int] = []
        missing_payment_rows: List[int] = []
        invalid_payment_rows: List[int] = []

        for row_index, raw_values in enumerate(
            ws.iter_rows(
                min_row=header_row + 1,
                min_col=1,
                max_col=SCAN_MAX_COL,
                values_only=True,
            ),
            start=header_row + 1,
        ):
            row_values = trim_trailing_blank(raw_values)
            no_value = row_value(row_values, header_map.get("no"))
            if is_blank(no_value):
                continue

            payment_to_supplier_raw = row_value(row_values, header_map.get("payment_to_supplier"))
            if is_blank(payment_to_supplier_raw):
                missing_payment_rows.append(row_index)
                continue

            payment_to_supplier_date = parse_date_value(payment_to_supplier_raw, wb.epoch)
            if payment_to_supplier_date is None:
                invalid_payment_rows.append(row_index)
                continue

            vat_invoice_raw = row_value(row_values, header_map.get("vat_invoice_date"))
            vat_invoice_date = parse_date_value(vat_invoice_raw, wb.epoch)
            if vat_invoice_date is None:
                missing_vat_rows.append(row_index)
                continue

            normalized_values: Dict[str, object] = {}
            for spec in SOURCE_COLUMNS:
                normalized_values[spec.key] = normalize_output_value(
                    spec.key,
                    row_value(row_values, header_map.get(spec.key)),
                    wb.epoch,
                )

            normalized_values["vat_invoice_date"] = vat_invoice_date
            normalized_values["payment_to_supplier"] = payment_to_supplier_date

            rows.append(
                SourceRow(
                    source_file=source_file,
                    sheet_name=sheet_name,
                    row_number=row_index,
                    payment_to_supplier_date=payment_to_supplier_date,
                    vat_invoice_date=vat_invoice_date,
                    values=normalized_values,
                )
            )

        if missing_payment_rows or missing_vat_rows:
            parts = [f"{source_file} | skipped whole file"]
            if missing_payment_rows:
                parts.append(
                    "Payment to Supplier missing/invalid at row(s): "
                    + ", ".join(str(row) for row in missing_payment_rows)
                )
            if missing_vat_rows:
                parts.append(
                    "VAT Invoice Date missing/invalid at row(s): "
                    + ", ".join(str(row) for row in missing_vat_rows)
                )
            audit_log.file_level_row_issues.append(" | ".join(parts))
            return [], 0

        if invalid_payment_rows:
            audit_log.invalid_payment_to_supplier_rows.append(
                f"{source_file} | row(s) ignored because Payment to Supplier was not a date: "
                + ", ".join(str(row) for row in invalid_payment_rows)
            )

        _emit(log_emit, f"Imported {len(rows)} row(s) from {source_file.name}")
        return rows, len(invalid_payment_rows)
    finally:
        wb.close()


def apply_header_style(cell, fill_color: str, font_color: str = BLACK, font_size: int = 10):
    cell.font = Font(bold=True, color=font_color, name="Arial", size=font_size)
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color=BLACK)
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def apply_data_style(cell, fill_color: str = MONTHLY_WHITE_FILL, font_color: str = BLACK, bold: bool = False):
    thin = Side(style="thin", color="BFBFBF")
    cell.font = Font(name="Arial", size=10)
    if font_color != BLACK or bold:
        cell.font = Font(name="Arial", size=10, color=font_color, bold=bold)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.fill = PatternFill("solid", fgColor=fill_color)


def apply_total_style(cell):
    thin = Side(style="thin", color=BLACK)
    cell.font = Font(name="Arial", size=11, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.fill = PatternFill("solid", fgColor=SUMMARY_TOTAL_FILL)


def summary_display_date(value: date) -> str:
    try:
        return value.strftime(SUMMARY_DATE_DISPLAY_FORMAT)
    except ValueError:
        return value.strftime(SUMMARY_DATE_FALLBACK_FORMAT)


def month_range(sheet_name: str, column: str, start_row: int, end_row: int) -> str:
    return f"{quote_sheetname(sheet_name)}!${column}${start_row}:${column}${end_row}"


def summary_date_label(rows: List[SourceRow]) -> str:
    dates = sorted({row.payment_to_supplier_date for row in rows})
    if not dates:
        return ""
    if len(dates) == 1:
        return summary_display_date(dates[0])
    return f"{summary_display_date(dates[0])} - {summary_display_date(dates[-1])}"


def setup_month_sheet(ws):
    for col_index, header in enumerate(ORIGINAL_OUTPUT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        apply_header_style(cell, MONTHLY_HEADER_FILL)
        apply_header_style(ws.cell(row=2, column=col_index, value=""), MONTHLY_HEADER_FILL)

    ws.merge_cells(start_row=1, start_column=OUTPUT_COL_USD, end_row=1, end_column=OUTPUT_COL_PAYMENT_VND)
    cell = ws.cell(
        row=1,
        column=OUTPUT_COL_USD,
        value="VTEC have to sell the USD to VCB bank at VCB's buying rate and then pay to suppliers in VND",
    )
    apply_header_style(cell, MONTHLY_GROUP_FILL, font_size=12)

    monthly_row1_headers = {
        OUTPUT_COL_RED_INVOICE_VND: ("Red Invoice amount Before VAT", MONTHLY_RATE_FILL, RED),
        OUTPUT_COL_FOREX_PAYMENT_VND: ("Gain/(Loss) in Forex", MONTHLY_WHITE_FILL, RED),
        OUTPUT_COL_VAT_RATE: ("VCB's Buying Rate at VAT Invoice Date", MONTHLY_GREEN_FILL, RED),
        OUTPUT_COL_FOREX_VAT_VND: ("Gain/(Loss) in Forex", MONTHLY_GREEN_FILL, RED),
        OUTPUT_COL_FOREX_VAT_USD: ("Gain/(Loss) in Forex", MONTHLY_GREEN_FILL, RED),
    }
    for col_index, (header, fill_color, font_color) in monthly_row1_headers.items():
        cell = ws.cell(row=1, column=col_index, value=header)
        apply_header_style(cell, fill_color, font_color=font_color)

    row2_headers = {
        OUTPUT_COL_USD: ("USD", MONTHLY_SUBGROUP_FILL, BLACK),
        OUTPUT_COL_PAYMENT_RATE: ("Rate (VND/USD)", MONTHLY_RATE_FILL, RED),
        OUTPUT_COL_PAYMENT_VND: ("VND", MONTHLY_SUBGROUP_FILL, BLACK),
        OUTPUT_COL_RED_INVOICE_VND: ("VND", MONTHLY_RATE_FILL, RED),
        OUTPUT_COL_FOREX_PAYMENT_VND: ("VND", MONTHLY_WHITE_FILL, RED),
        OUTPUT_COL_VAT_RATE: ("VND/USD", MONTHLY_GREEN_FILL, RED),
        OUTPUT_COL_FOREX_VAT_VND: ("VND", MONTHLY_GREEN_FILL, RED),
        OUTPUT_COL_FOREX_VAT_USD: ("USD", MONTHLY_GREEN_FILL, RED),
    }
    for col_index, (header, fill_color, font_color) in row2_headers.items():
        cell = ws.cell(row=2, column=col_index, value=header)
        apply_header_style(cell, fill_color, font_color=font_color)

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 38
    ws.row_dimensions[2].height = 22

    widths = {
        1: 8,
        2: 34,
        3: 20,
        4: 10,
        5: 14,
        6: 14,
        7: 18,
        8: 18,
        9: 18,
        10: 10,
        11: 12,
        12: 14,
        13: 18,
        14: 10,
        15: 18,
        16: 14,
        17: 8,
        18: 8,
        19: 14,
        20: 14,
        21: 17,
        22: 14,
        23: 12,
        24: 14,
        25: 14,
        26: 14,
        27: 12,
        28: 14,
        29: 14,
    }
    for col_index, width in widths.items():
        ws.column_dimensions[get_column_letter(col_index)].width = width


def build_summary_sheet(
    ws,
    year: int,
    month_entries: List[tuple[int, List[SourceRow], str, int]],
):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells(start_row=1, start_column=SUMMARY_COL_MONTH, end_row=1, end_column=SUMMARY_COL_SIG_AVG_RATE)
    apply_header_style(
        ws.cell(row=1, column=SUMMARY_COL_MONTH, value="SIG payment to Vtec"),
        SUMMARY_YELLOW_FILL,
        font_size=12,
    )
    ws.merge_cells(start_row=1, start_column=SUMMARY_COL_PAYMENT_DATE, end_row=1, end_column=SUMMARY_COL_BANK_VND)
    apply_header_style(
        ws.cell(
            row=1,
            column=SUMMARY_COL_PAYMENT_DATE,
            value="VTEC have to sell the USD to VCB bank at VCB's buying rate and then pay to suppliers in VND",
        ),
        SUMMARY_BLUE_FILL,
        font_size=12,
    )
    ws.merge_cells(start_row=1, start_column=SUMMARY_COL_PAYMENT_LOSS_VND, end_row=1, end_column=SUMMARY_COL_PAYMENT_LOSS_USD)
    apply_header_style(
        ws.cell(
            row=1,
            column=SUMMARY_COL_PAYMENT_LOSS_VND,
            value="Gain/(Loss) in Forex\n(if VCB's buying rate at VTEC's date payment)",
        ),
        SUMMARY_BLUE_FILL,
        font_size=12,
    )
    ws.merge_cells(start_row=1, start_column=SUMMARY_COL_VAT_LOSS_VND, end_row=1, end_column=SUMMARY_COL_VAT_LOSS_USD)
    apply_header_style(
        ws.cell(
            row=1,
            column=SUMMARY_COL_VAT_LOSS_VND,
            value="Gain/(Loss) in Forex\n(if VCB's buying rate at VAT invoice's date)",
        ),
        SUMMARY_GREEN_FILL,
        font_size=12,
    )

    row2 = {
        SUMMARY_COL_MONTH: ("Month", SUMMARY_YELLOW_FILL, BLACK),
        SUMMARY_COL_SIG_USD: ("USD", SUMMARY_YELLOW_FILL, RED),
        SUMMARY_COL_SIG_VND: ("VND", SUMMARY_YELLOW_FILL, RED),
        SUMMARY_COL_SIG_AVG_RATE: ("Average\nRate", SUMMARY_YELLOW_FILL, RED),
        SUMMARY_COL_PAYMENT_DATE: ("Date", SUMMARY_BLUE_FILL, BLACK),
        SUMMARY_COL_BANK_USD: ("USD", SUMMARY_BLUE_FILL, BLACK),
        SUMMARY_COL_BANK_RATE: ("(VCB' buying rate at\nVTEC's date payment)", SUMMARY_BLUE_FILL, RED),
        SUMMARY_COL_BANK_VND: ("VND", SUMMARY_BLUE_FILL, BLACK),
        SUMMARY_COL_PAYMENT_LOSS_VND: ("VND", SUMMARY_BLUE_FILL, BLACK),
        SUMMARY_COL_PAYMENT_LOSS_USD: ("USD", SUMMARY_BLUE_FILL, BLACK),
        SUMMARY_COL_VAT_LOSS_VND: ("VND", SUMMARY_GREEN_FILL, BLACK),
        SUMMARY_COL_VAT_LOSS_USD: ("USD", SUMMARY_GREEN_FILL, BLACK),
    }
    for col_index, (label, fill, font_color) in row2.items():
        apply_header_style(ws.cell(row=2, column=col_index, value=label), fill, font_color=font_color, font_size=11)

    ws.row_dimensions[1].height = 48
    ws.row_dimensions[2].height = 34

    widths = {
        SUMMARY_COL_MONTH: 12,
        SUMMARY_COL_SIG_USD: 16,
        SUMMARY_COL_SIG_VND: 18,
        SUMMARY_COL_SIG_AVG_RATE: 14,
        SUMMARY_COL_PAYMENT_DATE: 16,
        SUMMARY_COL_BANK_USD: 16,
        SUMMARY_COL_BANK_RATE: 18,
        SUMMARY_COL_BANK_VND: 18,
        SUMMARY_COL_PAYMENT_LOSS_VND: 18,
        SUMMARY_COL_PAYMENT_LOSS_USD: 14,
        SUMMARY_COL_VAT_LOSS_VND: 18,
        SUMMARY_COL_VAT_LOSS_USD: 14,
    }
    for col_index, width in widths.items():
        ws.column_dimensions[get_column_letter(col_index)].width = width

    for row_offset, (month, rows, month_sheet, end_row) in enumerate(month_entries, start=3):
        sig_usd_range = month_range(month_sheet, get_column_letter(OUTPUT_COL_USD), 3, end_row)
        red_invoice_range = month_range(month_sheet, get_column_letter(OUTPUT_COL_RED_INVOICE_VND), 3, end_row)
        payment_vnd_range = month_range(month_sheet, get_column_letter(OUTPUT_COL_PAYMENT_VND), 3, end_row)
        payment_loss_vnd_range = month_range(month_sheet, get_column_letter(OUTPUT_COL_FOREX_PAYMENT_VND), 3, end_row)
        vat_loss_vnd_range = month_range(month_sheet, get_column_letter(OUTPUT_COL_FOREX_VAT_VND), 3, end_row)
        vat_loss_usd_range = month_range(month_sheet, get_column_letter(OUTPUT_COL_FOREX_VAT_USD), 3, end_row)

        ws.cell(row=row_offset, column=SUMMARY_COL_MONTH, value=MONTH_ABBR[month])
        ws.cell(row=row_offset, column=SUMMARY_COL_SIG_USD, value=f"=SUM({sig_usd_range})")
        ws.cell(row=row_offset, column=SUMMARY_COL_SIG_VND, value=f"=SUM({red_invoice_range})")
        ws.cell(
            row=row_offset,
            column=SUMMARY_COL_SIG_AVG_RATE,
            value=f'=IFERROR({get_column_letter(SUMMARY_COL_SIG_VND)}{row_offset}/{get_column_letter(SUMMARY_COL_SIG_USD)}{row_offset},"")',
        )
        ws.cell(row=row_offset, column=SUMMARY_COL_PAYMENT_DATE, value=summary_date_label(rows))
        ws.cell(row=row_offset, column=SUMMARY_COL_BANK_USD, value=f"={get_column_letter(SUMMARY_COL_SIG_USD)}{row_offset}")
        ws.cell(
            row=row_offset,
            column=SUMMARY_COL_BANK_RATE,
            value=f'=IFERROR({get_column_letter(SUMMARY_COL_BANK_VND)}{row_offset}/{get_column_letter(SUMMARY_COL_BANK_USD)}{row_offset},"")',
        )
        ws.cell(row=row_offset, column=SUMMARY_COL_BANK_VND, value=f"=SUM({payment_vnd_range})")
        ws.cell(row=row_offset, column=SUMMARY_COL_PAYMENT_LOSS_VND, value=f"=SUM({payment_loss_vnd_range})")
        ws.cell(
            row=row_offset,
            column=SUMMARY_COL_PAYMENT_LOSS_USD,
            value=(
                f'=IFERROR({get_column_letter(SUMMARY_COL_PAYMENT_LOSS_VND)}{row_offset}/'
                f'{get_column_letter(SUMMARY_COL_BANK_RATE)}{row_offset},"")'
            ),
        )
        ws.cell(row=row_offset, column=SUMMARY_COL_VAT_LOSS_VND, value=f"=SUM({vat_loss_vnd_range})")
        ws.cell(row=row_offset, column=SUMMARY_COL_VAT_LOSS_USD, value=f"=SUM({vat_loss_usd_range})")

        for col_index in range(SUMMARY_COL_MONTH, SUMMARY_COL_VAT_LOSS_USD + 1):
            fill = MONTHLY_WHITE_FILL
            if col_index <= SUMMARY_COL_SIG_AVG_RATE:
                fill = SUMMARY_YELLOW_FILL
            elif col_index <= SUMMARY_COL_PAYMENT_LOSS_USD:
                fill = SUMMARY_BLUE_FILL
            elif col_index <= SUMMARY_COL_VAT_LOSS_USD:
                fill = SUMMARY_GREEN_FILL
            font_color = BLACK
            if col_index in {SUMMARY_COL_SIG_AVG_RATE, SUMMARY_COL_BANK_RATE}:
                font_color = RED
            if col_index in {SUMMARY_COL_PAYMENT_LOSS_USD, SUMMARY_COL_VAT_LOSS_USD}:
                apply_data_style(ws.cell(row=row_offset, column=col_index), fill_color=fill, font_color=BLACK, bold=True)
            else:
                apply_data_style(ws.cell(row=row_offset, column=col_index), fill_color=fill, font_color=font_color)

        ws.row_dimensions[row_offset].height = SUMMARY_DATA_ROW_HEIGHT
        ws.cell(row=row_offset, column=SUMMARY_COL_SIG_USD).number_format = USD_NUMBER_FORMAT
        ws.cell(row=row_offset, column=SUMMARY_COL_SIG_VND).number_format = VND_NUMBER_FORMAT
        ws.cell(row=row_offset, column=SUMMARY_COL_SIG_AVG_RATE).number_format = VND_NUMBER_FORMAT
        ws.cell(row=row_offset, column=SUMMARY_COL_BANK_USD).number_format = USD_NUMBER_FORMAT
        ws.cell(row=row_offset, column=SUMMARY_COL_BANK_RATE).number_format = VND_NUMBER_FORMAT
        ws.cell(row=row_offset, column=SUMMARY_COL_BANK_VND).number_format = VND_NUMBER_FORMAT
        ws.cell(row=row_offset, column=SUMMARY_COL_PAYMENT_LOSS_VND).number_format = VND_NUMBER_FORMAT
        ws.cell(row=row_offset, column=SUMMARY_COL_PAYMENT_LOSS_USD).number_format = USD_NUMBER_FORMAT
        ws.cell(row=row_offset, column=SUMMARY_COL_VAT_LOSS_VND).number_format = VND_NUMBER_FORMAT
        ws.cell(row=row_offset, column=SUMMARY_COL_VAT_LOSS_USD).number_format = USD_NUMBER_FORMAT

    total_row = len(month_entries) + 3
    apply_total_style(ws.cell(row=total_row, column=SUMMARY_COL_MONTH, value="TOTAL:"))
    for col_index in range(SUMMARY_COL_SIG_USD, SUMMARY_COL_VAT_LOSS_USD + 1):
        apply_total_style(ws.cell(row=total_row, column=col_index))

    ws.cell(row=total_row, column=SUMMARY_COL_PAYMENT_LOSS_VND, value=f"=SUM(I3:I{total_row - 1})")
    ws.cell(row=total_row, column=SUMMARY_COL_PAYMENT_LOSS_USD, value=f"=SUM(J3:J{total_row - 1})")
    ws.cell(row=total_row, column=SUMMARY_COL_VAT_LOSS_VND, value=f"=SUM(K3:K{total_row - 1})")
    ws.cell(row=total_row, column=SUMMARY_COL_VAT_LOSS_USD, value=f"=SUM(L3:L{total_row - 1})")
    ws.cell(row=total_row, column=SUMMARY_COL_PAYMENT_LOSS_VND).number_format = VND_NUMBER_FORMAT
    ws.cell(row=total_row, column=SUMMARY_COL_PAYMENT_LOSS_USD).number_format = USD_NUMBER_FORMAT
    ws.cell(row=total_row, column=SUMMARY_COL_VAT_LOSS_VND).number_format = VND_NUMBER_FORMAT
    ws.cell(row=total_row, column=SUMMARY_COL_VAT_LOSS_USD).number_format = USD_NUMBER_FORMAT
    ws.row_dimensions[total_row].height = SUMMARY_DATA_ROW_HEIGHT


def set_number_format(cell, key: str):
    if key in DATE_KEYS and isinstance(cell.value, (date, datetime)):
        cell.number_format = DATE_NUMBER_FORMAT
    elif key in VND_KEYS:
        cell.number_format = VND_NUMBER_FORMAT
    elif key in USD_KEYS:
        cell.number_format = USD_NUMBER_FORMAT
    elif key == "currency_rate" and isinstance(cell.value, (int, float)):
        cell.number_format = GENERAL_RATE_FORMAT


def write_data_row(
    ws,
    output_row_index: int,
    source_row: SourceRow,
    usd_rates: Dict[date, float],
    audit_log: AuditLog,
):
    for col_index, spec in enumerate(SOURCE_COLUMNS, start=1):
        value = source_row.values.get(spec.key)
        cell = ws.cell(row=output_row_index, column=col_index, value=value)
        apply_data_style(cell)
        set_number_format(cell, spec.key)
        if spec.key in {"supplier_name", "payment_term", "other_references_number"}:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    usd_value = choose_numeric(
        source_row.values.get("total_usd"),
        source_row.values.get("po_amount_before_vat_usd"),
    )
    red_invoice_vnd_value = choose_numeric(
        source_row.values.get("total_vnd"),
        source_row.values.get("po_amount_before_vat_vnd"),
    )
    payment_rate = usd_rates.get(source_row.payment_to_supplier_date)
    vat_rate = usd_rates.get(source_row.vat_invoice_date)

    if payment_rate is None:
        aggregate_missing_rate(
            audit_log.missing_payment_rates,
            source_row.payment_to_supplier_date,
            source_row.source_file,
            source_row.row_number,
        )
    if vat_rate is None:
        aggregate_missing_rate(
            audit_log.missing_vat_rates,
            source_row.vat_invoice_date,
            source_row.source_file,
            source_row.row_number,
        )

    ws.cell(row=output_row_index, column=OUTPUT_COL_USD, value=usd_value)
    ws.cell(row=output_row_index, column=OUTPUT_COL_PAYMENT_RATE, value=payment_rate)
    ws.cell(row=output_row_index, column=OUTPUT_COL_RED_INVOICE_VND, value=red_invoice_vnd_value)
    ws.cell(row=output_row_index, column=OUTPUT_COL_VAT_RATE, value=vat_rate)

    usd_ref = f"{get_column_letter(OUTPUT_COL_USD)}{output_row_index}"
    payment_rate_ref = f"{get_column_letter(OUTPUT_COL_PAYMENT_RATE)}{output_row_index}"
    payment_vnd_ref = f"{get_column_letter(OUTPUT_COL_PAYMENT_VND)}{output_row_index}"
    red_invoice_ref = f"{get_column_letter(OUTPUT_COL_RED_INVOICE_VND)}{output_row_index}"
    vat_rate_ref = f"{get_column_letter(OUTPUT_COL_VAT_RATE)}{output_row_index}"
    forex_vat_vnd_ref = f"{get_column_letter(OUTPUT_COL_FOREX_VAT_VND)}{output_row_index}"

    ws.cell(
        row=output_row_index,
        column=OUTPUT_COL_PAYMENT_VND,
        value=f'=IF(OR({usd_ref}="",{payment_rate_ref}=""),"",{usd_ref}*{payment_rate_ref})',
    )
    ws.cell(
        row=output_row_index,
        column=OUTPUT_COL_FOREX_PAYMENT_VND,
        value=f'=IF(OR({payment_vnd_ref}="",{red_invoice_ref}=""),"",{payment_vnd_ref}-{red_invoice_ref})',
    )
    ws.cell(
        row=output_row_index,
        column=OUTPUT_COL_FOREX_VAT_VND,
        value=f'=IF(OR({usd_ref}="",{vat_rate_ref}="",{red_invoice_ref}=""),"",{usd_ref}*{vat_rate_ref}-{red_invoice_ref})',
    )
    ws.cell(
        row=output_row_index,
        column=OUTPUT_COL_FOREX_VAT_USD,
        value=f'=IF(OR({forex_vat_vnd_ref}="",{vat_rate_ref}=""),"",{forex_vat_vnd_ref}/{vat_rate_ref})',
    )

    for col_index in range(OUTPUT_COL_USD, TOTAL_OUTPUT_COLUMNS + 1):
        cell = ws.cell(row=output_row_index, column=col_index)
        fill = MONTHLY_WHITE_FILL
        font_color = BLACK
        if col_index in {OUTPUT_COL_USD, OUTPUT_COL_PAYMENT_VND}:
            fill = MONTHLY_CALC_FILL
        elif col_index in {OUTPUT_COL_PAYMENT_RATE, OUTPUT_COL_RED_INVOICE_VND}:
            fill = MONTHLY_RATE_FILL
            font_color = RED
        elif col_index in {OUTPUT_COL_VAT_RATE, OUTPUT_COL_FOREX_VAT_VND, OUTPUT_COL_FOREX_VAT_USD}:
            fill = MONTHLY_GREEN_DATA_FILL
            if col_index == OUTPUT_COL_VAT_RATE:
                font_color = RED
        apply_data_style(cell, fill_color=fill, font_color=font_color)
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

    ws.cell(row=output_row_index, column=OUTPUT_COL_USD).number_format = USD_NUMBER_FORMAT
    ws.cell(row=output_row_index, column=OUTPUT_COL_PAYMENT_RATE).number_format = VND_NUMBER_FORMAT
    ws.cell(row=output_row_index, column=OUTPUT_COL_PAYMENT_VND).number_format = VND_NUMBER_FORMAT
    ws.cell(row=output_row_index, column=OUTPUT_COL_RED_INVOICE_VND).number_format = VND_NUMBER_FORMAT
    ws.cell(row=output_row_index, column=OUTPUT_COL_FOREX_PAYMENT_VND).number_format = VND_NUMBER_FORMAT
    ws.cell(row=output_row_index, column=OUTPUT_COL_VAT_RATE).number_format = VND_NUMBER_FORMAT
    ws.cell(row=output_row_index, column=OUTPUT_COL_FOREX_VAT_VND).number_format = VND_NUMBER_FORMAT
    ws.cell(row=output_row_index, column=OUTPUT_COL_FOREX_VAT_USD).number_format = USD_NUMBER_FORMAT
    ws.row_dimensions[output_row_index].height = DATA_ROW_HEIGHT


def render_audit_log(
    audit_log: AuditLog,
    input_folders: List[Path],
    vcb_files: List[Path],
    result: ProcessResult,
) -> str:
    lines = [
        "VTEC Payment Monthly Combiner - Audit Log",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "Selections",
        "----------",
        f"Input folders: {len(input_folders)}",
    ]
    lines.extend(f"- {folder}" for folder in input_folders)
    lines.append(f"VCB files: {len(vcb_files)}")
    lines.extend(f"- {path}" for path in vcb_files)
    lines.extend(
        [
            "",
            "Summary",
            "-------",
            f"Source files found: {result.source_files_found}",
            f"Source files imported: {result.source_files_imported}",
            f"Source files skipped: {result.source_files_skipped}",
            f"Rows imported: {result.rows_imported}",
            f"Rows ignored because Payment to Supplier was not a date: {result.rows_ignored_non_date_payment_to_supplier}",
            f"Monthly sheets created: {result.month_sheet_count}",
            f"Workbook output: {result.workbook_path}",
            f"Audit log: {result.log_path}",
        ]
    )

    sections = [
        ("Folders with no .xlsx files in them or their subfolders", audit_log.folders_without_xlsx),
        ("Source files skipped because they have more than 1 sheet", audit_log.multi_sheet_files),
        ("Source file read/open errors", audit_log.source_file_errors),
        ("Source files skipped because required headers were missing", audit_log.missing_required_headers),
        ("Source files skipped because mandatory date cells were missing or invalid", audit_log.file_level_row_issues),
        ("Rows ignored because Payment to Supplier was not a date", audit_log.invalid_payment_to_supplier_rows),
        ("VCB file errors", audit_log.vcb_file_errors),
        ("Duplicate USD VCB rates ignored", audit_log.duplicate_vcb_rates),
    ]

    for title, items in sections:
        lines.extend(["", title, "-" * len(title)])
        if items:
            lines.extend(f"- {item}" for item in items)
        else:
            lines.append("- None")

    def append_missing_rate_section(title: str, mapping: Dict[str, List[str]]):
        lines.extend(["", title, "-" * len(title)])
        if not mapping:
            lines.append("- None")
            return
        for missing_date in sorted(mapping.keys()):
            examples = "; ".join(mapping[missing_date])
            lines.append(
                f"- {missing_date}: {len(mapping[missing_date])} example(s) with no USD Telegraphic Buying rate"
                + (f" | {examples}" if examples else "")
            )

    append_missing_rate_section("Missing Payment to Supplier USD rates", audit_log.missing_payment_rates)
    append_missing_rate_section("Missing VAT Invoice Date USD rates", audit_log.missing_vat_rates)
    return "\n".join(lines) + "\n"


def process_files(
    input_folders: List[Path],
    vcb_files: List[Path],
    output_dir: Optional[Path] = None,
    log_emit: Callable[[str], None] | None = None,
) -> ProcessResult:
    if not input_folders:
        raise ValueError("Please select at least one input folder.")
    if not vcb_files:
        raise ValueError("Please select at least one VCB exchange-rate workbook.")

    audit_log = AuditLog()
    output_root = output_dir if output_dir else input_folders[0].parent
    output_root.mkdir(parents=True, exist_ok=True)
    workbook_path, log_path = build_output_paths(output_root)

    _emit(log_emit, "Scanning selected input folders recursively for .xlsx files...")
    source_files = collect_source_files(input_folders, audit_log)
    if not source_files:
        raise ValueError("No .xlsx files were found in the selected input folders or their subfolders.")

    usd_rates = load_vcb_usd_rates(vcb_files, audit_log, log_emit=log_emit)
    if not usd_rates:
        raise ValueError("No USD Telegraphic Buying rates could be read from the selected VCB file(s).")

    all_rows: List[SourceRow] = []
    imported_files = 0
    skipped_files = 0
    rows_ignored_non_date_payment_to_supplier = 0

    for source_file in source_files:
        parsed_rows, ignored_rows = parse_source_file(source_file, audit_log, log_emit=log_emit)
        rows_ignored_non_date_payment_to_supplier += ignored_rows

        if parsed_rows:
            imported_files += 1
            all_rows.extend(parsed_rows)
        else:
            skipped_files += 1

    if not all_rows:
        skipped_files = max(skipped_files, len(source_files))
        result = ProcessResult(
            workbook_path=workbook_path,
            log_path=log_path,
            source_files_found=len(source_files),
            source_files_imported=imported_files,
            source_files_skipped=skipped_files,
            rows_imported=0,
            rows_ignored_non_date_payment_to_supplier=rows_ignored_non_date_payment_to_supplier,
            month_sheet_count=0,
        )
        log_path.write_text(render_audit_log(audit_log, input_folders, vcb_files, result), encoding="utf-8")
        raise ValueError(f"No valid rows were imported. Audit log written to: {log_path}")

    grouped_rows: Dict[tuple[int, int], List[SourceRow]] = defaultdict(list)
    for source_row in all_rows:
        grouped_rows[(source_row.payment_to_supplier_date.year, source_row.payment_to_supplier_date.month)].append(source_row)

    wb = Workbook()
    sorted_group_keys = sorted(grouped_rows.keys())
    sorted_years = sorted({year for year, _month in sorted_group_keys})

    first_sheet = wb.active
    summary_sheets: Dict[int, object] = {}
    for summary_index, year in enumerate(sorted_years):
        ws = first_sheet if summary_index == 0 else wb.create_sheet()
        ws.title = f"SUM'{year % 100:02d}"
        summary_sheets[year] = ws

    month_sheet_meta: Dict[tuple[int, int], tuple[str, int]] = {}
    for group_key in sorted_group_keys:
        year, month = group_key
        ws = wb.create_sheet()
        ws.title = month_sheet_name(date(year, month, 1))
        setup_month_sheet(ws)

        sorted_rows = sorted(
            grouped_rows[group_key],
            key=lambda item: (
                item.payment_to_supplier_date,
                item.vat_invoice_date,
                str(item.values.get("supplier_name") or "").upper(),
                item.source_file.name.upper(),
                item.row_number,
            ),
        )
        for output_row_index, source_row in enumerate(sorted_rows, start=3):
            write_data_row(ws, output_row_index, source_row, usd_rates, audit_log)
        month_sheet_meta[group_key] = (ws.title, ws.max_row)

    for year in sorted_years:
        month_entries = [
            (month, grouped_rows[(year, month)], month_sheet_meta[(year, month)][0], month_sheet_meta[(year, month)][1])
            for (_year, month) in sorted_group_keys
            if _year == year
        ]
        build_summary_sheet(summary_sheets[year], year, month_entries)

    wb.save(workbook_path)

    result = ProcessResult(
        workbook_path=workbook_path,
        log_path=log_path,
        source_files_found=len(source_files),
        source_files_imported=imported_files,
        source_files_skipped=skipped_files,
        rows_imported=len(all_rows),
        rows_ignored_non_date_payment_to_supplier=rows_ignored_non_date_payment_to_supplier,
        month_sheet_count=len(sorted_group_keys),
    )
    log_path.write_text(render_audit_log(audit_log, input_folders, vcb_files, result), encoding="utf-8")
    return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Combine VTEC payment summary workbooks into monthly worksheets using VCB USD rates."
    )
    parser.add_argument(
        "folders",
        nargs="+",
        help="Input folders containing .xlsx payment-summary files. Subfolders are scanned automatically.",
    )
    parser.add_argument(
        "--vcb",
        nargs="+",
        required=True,
        help="One or more VCB exchange-rate workbook paths.",
    )
    parser.add_argument(
        "--output-dir",
        help="Output folder. Defaults to the first selected input folder's parent.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    result = process_files(
        input_folders=[Path(path) for path in args.folders],
        vcb_files=[Path(path) for path in args.vcb],
        output_dir=Path(args.output_dir) if args.output_dir else None,
        log_emit=print,
    )
    print(f"Workbook: {result.workbook_path}")
    print(f"Audit log: {result.log_path}")
    print(f"Rows imported: {result.rows_imported}")
    print(f"Monthly sheets: {result.month_sheet_count}")


if __name__ == "__main__":
    main()


if GUI_AVAILABLE:
    class MainWidget(QWidget):
        log_message = Signal(str)
        processing_done = Signal(bool, str, str, int, int, int, int)

        def __init__(self):
            super().__init__()
            self.setObjectName("vtec_payment_monthly_combiner_widget")
            self.input_folders: List[Path] = []
            self.vcb_files: List[Path] = []
            self.output_dir: Path = Path.home()
            self.output_dir_locked = False

            self._build_ui()
            self._connect_signals()
            self._refresh_summary()

        def _build_ui(self):
            self.desc_label = QLabel("", self)
            self.desc_label.setWordWrap(True)
            self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
            self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
            self.desc_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
            self.desc_label.setStyleSheet(
                "color: #dcdcdc; background: transparent; padding: 6px; "
                "border: 1px solid #3a3a3a; border-radius: 6px;"
            )
            self.set_long_description("")

            self.select_folders_btn = PrimaryPushButton("Add Input Folder", self)
            self.clear_folders_btn = PrimaryPushButton("Clear Input Folders", self)
            self.select_vcb_btn = PrimaryPushButton("Select VCB File(s)", self)
            self.select_output_btn = PrimaryPushButton("Select Output Folder", self)
            self.run_btn = PrimaryPushButton("Generate Monthly Workbook", self)

            self.summary_label = QLabel("Selections", self)
            self.summary_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

            self.logs_label = QLabel("Process logs", self)
            self.logs_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

            box_style = (
                "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
                "border: 1px solid #3a3a3a; border-radius: 6px;}"
            )

            self.summary_box = QTextEdit(self)
            self.summary_box.setReadOnly(True)
            self.summary_box.setPlaceholderText("Input folders, VCB files, and output details will appear here")
            self.summary_box.setStyleSheet(box_style)

            self.log_box = QTextEdit(self)
            self.log_box.setReadOnly(True)
            self.log_box.setPlaceholderText("Live process log will appear here")
            self.log_box.setStyleSheet(box_style)

            layout = QVBoxLayout(self)
            layout.setContentsMargins(16, 16, 16, 16)
            layout.setSpacing(12)
            layout.addWidget(self.desc_label)

            row1 = QHBoxLayout()
            row1.addWidget(self.select_folders_btn)
            row1.addWidget(self.clear_folders_btn)
            row1.addWidget(self.select_vcb_btn)
            row1.addWidget(self.select_output_btn)
            layout.addLayout(row1)

            row2 = QHBoxLayout()
            row2.addStretch(1)
            row2.addWidget(self.run_btn)
            row2.addStretch(1)
            layout.addLayout(row2)

            row3 = QHBoxLayout()
            row3.addWidget(self.summary_label, 1)
            row3.addWidget(self.logs_label, 1)
            layout.addLayout(row3)

            row4 = QHBoxLayout()
            row4.addWidget(self.summary_box, 1)
            row4.addWidget(self.log_box, 1)
            layout.addLayout(row4, 3)

        def _connect_signals(self):
            self.select_folders_btn.clicked.connect(self.select_input_folders)
            self.clear_folders_btn.clicked.connect(self.clear_input_folders)
            self.select_vcb_btn.clicked.connect(self.select_vcb_files)
            self.select_output_btn.clicked.connect(self.select_output_folder)
            self.run_btn.clicked.connect(self.run_process)
            self.log_message.connect(self.append_log)
            self.processing_done.connect(self.on_processing_done)

        def set_long_description(self, text: str):
            clean = (text or "").strip()
            if clean:
                self.desc_label.setText(clean)
                self.desc_label.show()
            else:
                self.desc_label.clear()
                self.desc_label.hide()

        def _dedupe_paths(self, paths: Iterable[Path]) -> List[Path]:
            deduped: List[Path] = []
            seen: set[Path] = set()
            for path in paths:
                resolved = path.resolve()
                if resolved not in seen:
                    deduped.append(path)
                    seen.add(resolved)
            return deduped

        def _refresh_default_output_dir(self):
            if not self.output_dir_locked and self.input_folders:
                self.output_dir = self.input_folders[0].parent

        def _refresh_summary(self):
            self._refresh_default_output_dir()

            lines = [
                f"Input folders selected: {len(self.input_folders)}",
                "Input scan: recursive (includes subfolders)",
            ]
            if self.input_folders:
                lines.extend(f"- {path}" for path in self.input_folders)
            else:
                lines.append("- None")

            lines.append("")
            lines.append(f"VCB files selected: {len(self.vcb_files)}")
            if self.vcb_files:
                lines.extend(f"- {path}" for path in self.vcb_files)
            else:
                lines.append("- None")

            lines.extend(
                [
                    "",
                    f"Output folder: {self.output_dir}",
                    "Output workbook name: auto-generated with timestamp",
                    "Audit log name: matches workbook name with _issues.txt suffix",
                ]
            )

            self.summary_box.setPlainText("\n".join(lines))

        def _select_input_folder(self) -> Optional[Path]:
            start_dir = self.input_folders[-1] if self.input_folders else self.output_dir
            selected = QFileDialog.getExistingDirectory(self, "Select Input Folder", str(start_dir))
            if selected:
                return Path(selected)
            return None

        def select_input_folders(self):
            selected = self._select_input_folder()
            if not selected:
                return
            self.input_folders = self._dedupe_paths([*self.input_folders, selected])
            self._refresh_summary()

        def clear_input_folders(self):
            self.input_folders = []
            if not self.output_dir_locked:
                self.output_dir = Path.home()
            self._refresh_summary()

        def select_vcb_files(self):
            files, _ = QFileDialog.getOpenFileNames(
                self,
                "Select VCB Exchange Rate Files",
                "",
                "Excel Files (*.xlsx)",
            )
            if files:
                self.vcb_files = self._dedupe_paths(Path(path) for path in files)
            else:
                self.vcb_files = []
            self._refresh_summary()

        def select_output_folder(self):
            selected = QFileDialog.getExistingDirectory(self, "Select Output Folder", str(self.output_dir))
            if selected:
                self.output_dir = Path(selected)
                self.output_dir_locked = True
                self._refresh_summary()

        def _set_running_state(self, running: bool):
            enabled = not running
            self.select_folders_btn.setEnabled(enabled)
            self.clear_folders_btn.setEnabled(enabled)
            self.select_vcb_btn.setEnabled(enabled)
            self.select_output_btn.setEnabled(enabled)
            self.run_btn.setEnabled(enabled)

        def run_process(self):
            if not self.input_folders:
                MessageBox("Warning", "Please add at least one input folder.", self).exec()
                return
            if not self.vcb_files:
                MessageBox("Warning", "Please select at least one VCB exchange-rate workbook.", self).exec()
                return

            self.log_box.clear()
            self.log_message.emit("Process started...")
            self.log_message.emit(f"Input folders: {len(self.input_folders)}")
            self.log_message.emit(f"VCB files: {len(self.vcb_files)}")
            self.log_message.emit(f"Output folder: {self.output_dir}")
            self.log_message.emit("")
            self._set_running_state(True)

            def worker():
                try:
                    result = process_files(
                        input_folders=self.input_folders,
                        vcb_files=self.vcb_files,
                        output_dir=self.output_dir,
                        log_emit=self.log_message.emit,
                    )
                    self.processing_done.emit(
                        True,
                        str(result.workbook_path),
                        str(result.log_path),
                        result.source_files_imported,
                        result.source_files_skipped,
                        result.rows_imported,
                        result.month_sheet_count,
                    )
                except Exception as exc:
                    self.log_message.emit(f"CRITICAL ERROR: {exc}")
                    self.processing_done.emit(False, "", "", 0, 0, 0, 0)

            threading.Thread(target=worker, daemon=True).start()

        def append_log(self, text: str):
            self.log_box.append(text)
            self.log_box.ensureCursorVisible()

        def on_processing_done(
            self,
            success: bool,
            workbook_path: str,
            log_path: str,
            imported_files: int,
            skipped_files: int,
            rows_imported: int,
            month_sheet_count: int,
        ):
            self._set_running_state(False)

            if success:
                self.log_message.emit("")
                self.log_message.emit(f"Workbook saved: {workbook_path}")
                self.log_message.emit(f"Audit log saved: {log_path}")
                self.log_message.emit(
                    f"Completed: {imported_files} imported file(s), {skipped_files} skipped file(s), "
                    f"{rows_imported} row(s), {month_sheet_count} monthly sheet(s)"
                )

                msg = MessageBox(
                    "Monthly workbook created",
                    "\n".join(
                        [
                            f"Imported files: {imported_files}",
                            f"Skipped files: {skipped_files}",
                            f"Rows imported: {rows_imported}",
                            f"Monthly sheets: {month_sheet_count}",
                            f"Workbook: {Path(workbook_path).name}",
                            f"Audit log: {Path(log_path).name}",
                        ]
                    ),
                    self,
                )
            else:
                msg = MessageBox(
                    "Monthly workbook finished with issues",
                    "No workbook was created. Check the process log for details.",
                    self,
                )

            msg.yesButton.setText("OK")
            msg.cancelButton.hide()
            msg.exec()


    def get_widget():
        return MainWidget()


else:
    def get_widget():
        raise RuntimeError("PySide6 and qfluentwidgets are required to use this pycro inside Pycro Station.")
