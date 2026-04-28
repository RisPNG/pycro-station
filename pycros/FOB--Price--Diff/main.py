import os
import csv
import re
import threading
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from typing import List, Tuple, Any, Dict, Optional

# GUI Imports
from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QVBoxLayout,
    QLabel,
    QCheckBox,
    QTextEdit,
    QWidget,
    QSizePolicy
)
from qfluentwidgets import PrimaryPushButton, MessageBox

# Excel Imports
from openpyxl import Workbook, load_workbook


try:
    import xlrd  # for legacy .xls support
except Exception:
    xlrd = None

# Excel Automation for Formula Calculation
try:
    import xlwings as xw
    HAS_XLWINGS = True
except ImportError:
    HAS_XLWINGS = False

try:
    from PackagesPage import CheckIconButton  # type: ignore
except Exception:
    CheckIconButton = None

# --- Logic Implementation ---

# Extended Size Order for comparison logic
SIZE_ORDER = [
    "0", "2", "4", "6", "8", "10", "12", "14", "16", "18", "20", "22", "24", "26", "28", "30", "32", "34", "36", "38", "40", "42",
    "2XS", "XS", "S", "M", "L", "XL", "2XL", "3XL", "4XL", "5XL",
    "2XSS", "XSS", "SS", "MS", "LS", "XLS", "2XLS", "3XLS", "4XLS", "5XLS",
    "2XSL", "XSL", "SL", "ML", "XLL", "2XLL", "3XLL", "4XLL", "5XLL",
    "2XST", "XST", "ST", "MT", "LT", "XLT", "2XLT", "3XLT", "4XLT", "5XLT",
    "2XSTT", "XSTT", "STT", "MTT", "LTT", "XLTT", "2XLTT", "3XLTT", "4XLTT", "5XLTT",
    "X", "0X", "1X", "2X", "3X", "4X", "5X",
    "XT", "0XT", "1XT", "2XT", "3XT", "4XT", "5XT",
    "XTT", "0XTT", "1XTT", "2XTT", "3XTT", "4XTT", "5XTT",
    "CUST2XS", "CUSTXS", "CUSTS", "CUSTM", "CUSTL", "CUSTXL", "CUST2XL", "CUST3XL", "CUST4XL", "CUST5XL",
    "CUST", "CUST0", "CUST1", "CUST2", "CUST3", "CUST4", "CUST5"
]

def normalize_header(header_text):
    """Normalize header text for comparison (remove newlines, extra spaces, uppercase)."""
    if not header_text:
        return ""
    return str(header_text).replace("\n", " ").replace("\r", "").strip().upper()

def get_col_index(headers, target_names):
    """Find index of a header that matches one of the target names."""
    if isinstance(target_names, str):
        target_names = [target_names]

    target_names = [normalize_header(t) for t in target_names]

    for idx, h in enumerate(headers):
        if normalize_header(h) in target_names:
            return idx
    return -1

MONEY_ZERO = Decimal("0")
MONEY_CENT = Decimal("0.01")

def safe_decimal(value: Any) -> Decimal:
    """Safely convert value to Decimal, handling currency strings."""
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return Decimal("0")

    s_val = str(value).strip().replace(" ", "").replace("$", "").replace(",", "")
    if s_val == "-" or s_val == "":
        return Decimal("0")
    try:
        return Decimal(s_val)
    except InvalidOperation:
        return Decimal("0")

def money_abs_diff(a: Any, b: Any) -> Decimal:
    """Absolute difference between 2 money values, rounded to cents."""
    return (safe_decimal(a) - safe_decimal(b)).copy_abs().quantize(MONEY_CENT, rounding=ROUND_HALF_UP)

def format_money_trace(value: Any) -> str:
    """Format money values consistently for trace output."""
    return f"{safe_decimal(value).quantize(MONEY_CENT, rounding=ROUND_HALF_UP):.2f}"

def money_sum(values: List[Any]) -> Decimal:
    """Add money values using Decimal end-to-end."""
    total = MONEY_ZERO
    for value in values:
        total += safe_decimal(value)
    return total

def money_average(values: List[Any]) -> Decimal:
    """Average money values using Decimal end-to-end."""
    if not values:
        return MONEY_ZERO
    return money_sum(values) / Decimal(len(values))

def normalize_line_item(value: Any) -> str:
    """Normalize PO line item values without float conversion."""
    if value is None:
        return ""
    raw = str(value).strip()
    if not raw:
        return ""
    try:
        dec = Decimal(raw.replace(",", ""))
        if dec == dec.to_integral_value():
            return str(dec.to_integral_value())
    except Exception:
        pass
    return raw

def get_row_value(row: Any, idx: int, default: Any = None) -> Any:
    """Safely read a row value by zero-based index."""
    if idx < 0 or row is None:
        return default
    try:
        return row[idx] if idx < len(row) else default
    except TypeError:
        return default

def normalize_date_str(date_val):
    """
    Convert various date formats (datetime obj, 'MM/DD/YYYY', 'YYYY-MM-DD')
    to standard 'MM/DD/YYYY' string for comparison.
    """
    if not date_val:
        return ""

    if isinstance(date_val, datetime):
        return date_val.strftime("%m/%d/%Y")

    s_val = str(date_val).strip()

    # Try parsing common formats
    formats = ["%m/%d/%Y", "%Y-%m-%d", "%d-%b-%y", "%m-%d-%Y"]
    for fmt in formats:
        try:
            dt = datetime.strptime(s_val, fmt)
            return dt.strftime("%m/%d/%Y")
        except ValueError:
            continue

    return s_val # Return as is if parsing fails (fallback)

def calculate_target_effective_date(buy_mth_str):
    """
    Converts OCCC 'BUY MTH' to PPS Effective Date string.

    Supported formats:
    - 'YY-M' or 'YY-MM' with optional suffix letters (e.g., '25-1E', '26-10M')
    - 'MMYYYY' or 'MYYYY' with optional suffix letters (e.g., '042025E', '072025M')

    Logic:
    M=1,2,12 -> Dec 1st (Prev year for 1,2; Curr year for 12)
    M=3,4,5 -> Mar 1st
    M=6,7,8 -> Jun 1st
    M=9,10,11 -> Sep 1st
    """
    s_val = str(buy_mth_str).strip().upper()
    if not s_val:
        return None

    # Format: YY-M (optionally followed by letters, e.g. 25-4E)
    match = re.match(r"^(\d{2})-(\d{1,2})", s_val)
    if match:
        yy = int(match.group(1))
        m = int(match.group(2))
        year = 2000 + yy
    else:
        # Format: MMYYYY / MYYYY (optionally followed by letters, e.g. 042025E)
        match = re.match(r"^(\d{1,2})(\d{4})", s_val)
        if not match:
            return None
        m = int(match.group(1))
        year = int(match.group(2))

    if m < 1 or m > 12:
        return None

    target_month = 1
    target_year = year

    if m in [12, 1, 2]:
        target_month = 12
        if m in [1, 2]:
            target_year = year - 1
        else:
            target_year = year
    elif m in [3, 4, 5]:
        target_month = 3
    elif m in [6, 7, 8]:
        target_month = 6
    elif m in [9, 10, 11]:
        target_month = 9

    return f"{target_month:02d}/01/{target_year}"

def normalize_pps_season_year(value: Any) -> str:
    """Normalize PPS/OCCC season-year values like 'SP26' / \"SP'26\" / 'SP 26'."""
    if value is None:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(value).strip().upper())

def build_target_pps_season_year(season_val: Any, season_year_val: Any) -> str:
    """Build the PPS SEASON_YEAR key from OCCC SEASON + SEASON YEAR."""
    season = re.sub(r"[^A-Z]", "", str(season_val).strip().upper()) if season_val is not None else ""
    year_digits = re.sub(r"\D", "", str(season_year_val).strip()) if season_year_val is not None else ""
    if not season or len(year_digits) < 2:
        return ""
    return normalize_pps_season_year(f"{season}{year_digits[-2:]}")


def build_timestamped_copy_path(input_path: str, timestamp: str, label: str = "updated", output_ext: Optional[str] = None) -> str:
    """Generate output path beside input file with timestamp prefix (no overwrite)."""
    base_dir = os.path.dirname(input_path) or os.getcwd()
    base_name = os.path.basename(input_path)
    if output_ext:
        base_name = f"{os.path.splitext(base_name)[0]}{output_ext}"
    out_name = f"{timestamp}_{label}_{base_name}"
    out_path = os.path.join(base_dir, out_name)

    if not os.path.exists(out_path):
        return out_path

    name, ext = os.path.splitext(out_name)
    counter = 1
    while True:
        candidate = os.path.join(base_dir, f"{name} ({counter}){ext}")
        if not os.path.exists(candidate):
            return candidate
        counter += 1

EXT_SIZE_EMPTY = {"-", "", "NONE", "NA", "N/A"}

def normalize_size_code(size_val: Any) -> str:
    """Normalize size strings for SIZE_ORDER lookup (e.g., 'XL-T' -> 'XLT')."""
    if size_val is None:
        return ""
    s_val = str(size_val).strip().upper()
    if not s_val:
        return ""
    return re.sub(r"[^A-Z0-9]", "", s_val)

def is_tall_size(size_code: Any) -> bool:
    """Best-effort detection for Tall sizes (e.g., ST, MTT, XLTT, 2XLT)."""
    s = normalize_size_code(size_code)
    if not s or s.startswith("CUST"):
        return False
    return s.endswith("TT") or s.endswith("T")

def extract_ext_threshold_size_code(ext_def_val: Any) -> str:
    """
    Extract a SIZE_ORDER code from OCCC 'Extended Sizes' values.

    Examples:
    - '3XL&Abv' -> '3XL'
    - '38 ONWARDS' -> '38'
    - \"3XL.Follow to FA'26\" -> '3XL'
    - \"Follow to SP'26\" -> '' (no usable size code)
    """
    raw = str(ext_def_val).strip().upper() if ext_def_val is not None else ""
    if not raw or raw in EXT_SIZE_EMPTY:
        return ""

    # Avoid accidentally treating season/year notes (e.g. "SP'26") as numeric sizes like "26"
    cleaned = re.sub(r"\b[A-Z]{2,3}'\d{2}\b", " ", raw)
    cleaned = re.sub(r"'\d{2}\b", " ", cleaned)
    cleaned = re.sub(r"\b20\d{2}\b", " ", cleaned)

    tokens = [t for t in re.split(r"[\s&./,;()]+", cleaned) if t]
    for t in tokens:
        code = normalize_size_code(t)
        if code in SIZE_ORDER:
            return code
    return ""

def is_extended_size(ppm_size_str, occc_threshold_str):
    """Determine if a size is extended based on the OCCC threshold or TALL logic."""
    ppm_size = normalize_size_code(ppm_size_str)
    threshold_raw = str(occc_threshold_str).strip().upper() if occc_threshold_str is not None else ""

    if not ppm_size:
        return False

    if threshold_raw in EXT_SIZE_EMPTY:
        return False

    if "TALL" in threshold_raw:
        return is_tall_size(ppm_size)

    threshold = extract_ext_threshold_size_code(threshold_raw)
    if not threshold:
        return False

    try:
        idx_ppm = SIZE_ORDER.index(ppm_size)
    except ValueError:
        return False

    try:
        idx_threshold = SIZE_ORDER.index(threshold)
    except ValueError:
        return False

    return idx_ppm >= idx_threshold

def refresh_excel_formulas(filepath, log_emit):
    """
    Uses xlwings to open, calculate, and save the file.
    This ensures openpyxl reads the calculated formula results instead of None/0.0.
    """
    if not HAS_XLWINGS:
        log_emit("Warning: xlwings not installed. Formulas might read as 0.0.")
        return False

    try:
        log_emit(f"Auto-calculating formulas for {os.path.basename(filepath)}... (This may take a moment)")
        with xw.App(visible=False, add_book=False) as app:
            app.display_alerts = False
            try:
                wb = app.books.open(filepath)
                wb.save()
                wb.close()
                log_emit("Formulas calculated and file saved.")
            except Exception as e:
                log_emit(f"Excel Automation Error: {e}")
    except Exception as e:
        log_emit(f"Could not launch Excel: {e}")

def refine_remarks(remarks_list, trace_steps: Optional[List[str]] = None):
    """Post-process remarks to consolidate messages."""
    if not remarks_list:
        return []

    s_pps_match_reg = "PPS OFOB match for regular sizes"
    s_pps_match_ext = "PPS OFOB match for extended sizes"
    s_pps_miss_reg = "PPS OFOB doesn't match for regular sizes"
    s_pps_miss_ext = "PPS OFOB doesn't match for extended sizes"
    s_ofob_miss_reg = "OFOB (Regular sizes) doesn't match with PPM"
    s_ofob_miss_ext = "OFOB (Extended sizes) doesn't match with PPM"
    s_final_miss_reg = "FINAL FOB (Regular sizes) doesn't match with PPM"
    s_final_miss_ext = "FINAL FOB (Extended sizes) doesn't match with PPM"
    s_nike_final_all = "NIKE FINAL FOB issue for all sizes"
    surcharge_remarks = {
        "S/C MIN PRODUCTION (ZPMX) doesn't match",
        "S/C Min Material (ZMMX) doesn't match",
        "S/C Misc (ZMSX) doesn't match",
        "S/C VAS Manual (ZVAX) doesn't match",
    }

    r_set = set(remarks_list)
    targets = {
        s_pps_match_reg, s_pps_match_ext,
        s_pps_miss_reg, s_pps_miss_ext,
        s_ofob_miss_reg, s_ofob_miss_ext,
        s_final_miss_reg, s_final_miss_ext,
    }

    final_list = []

    # 1. Keep unrelated remarks
    for r in remarks_list:
        if r not in targets:
            final_list.append(r)

    # 2. OFOB / PPS regular logic
    added_pps_issue_reg = False
    added_nike_issue_reg = False
    keep_ofob_miss_reg = False

    if s_pps_miss_reg in r_set and s_ofob_miss_reg in r_set:
        added_pps_issue_reg = True
        if trace_steps is not None:
            trace_steps.append(
                "Consolidation: PPS OFOB regular mismatch + OFOB regular mismatch -> PPS OFOB issue for regular sizes."
            )
    elif s_pps_match_reg in r_set and s_ofob_miss_reg in r_set:
        added_nike_issue_reg = True
        if trace_steps is not None:
            trace_steps.append(
                "Consolidation: PPS OFOB regular match + OFOB regular mismatch -> NIKE OFOB issue for regular sizes."
            )
    elif s_pps_miss_reg in r_set:
        added_pps_issue_reg = True
        if trace_steps is not None:
            trace_steps.append(
                "Consolidation: PPS OFOB regular mismatch -> PPS OFOB issue for regular sizes."
            )
    elif s_ofob_miss_reg in r_set:
        keep_ofob_miss_reg = True
        if trace_steps is not None:
            trace_steps.append(
                "Consolidation: keeping OFOB regular mismatch because PPS OFOB regular data did not resolve it into PPS OFOB or NIKE OFOB issue."
            )

    # 3. OFOB / PPS extended logic
    added_pps_issue_ext = False
    added_nike_issue_ext = False
    keep_ofob_miss_ext = False

    if s_pps_miss_ext in r_set and s_ofob_miss_ext in r_set:
        added_pps_issue_ext = True
        if trace_steps is not None:
            trace_steps.append(
                "Consolidation: PPS OFOB extended mismatch + OFOB extended mismatch -> PPS OFOB issue for extended sizes."
            )
    elif s_pps_match_ext in r_set and s_ofob_miss_ext in r_set:
        added_nike_issue_ext = True
        if trace_steps is not None:
            trace_steps.append(
                "Consolidation: PPS OFOB extended match + OFOB extended mismatch -> NIKE OFOB issue for extended sizes."
            )
    elif s_pps_miss_ext in r_set:
        added_pps_issue_ext = True
        if trace_steps is not None:
            trace_steps.append(
                "Consolidation: PPS OFOB extended mismatch -> PPS OFOB issue for extended sizes."
            )
    elif s_ofob_miss_ext in r_set:
        keep_ofob_miss_ext = True
        if trace_steps is not None:
            trace_steps.append(
                "Consolidation: keeping OFOB extended mismatch because PPS OFOB extended data did not resolve it into PPS OFOB or NIKE OFOB issue."
            )

    # 4. Keep raw OFOB mismatches that were not consumed by issue logic
    if keep_ofob_miss_reg:
        final_list.append(s_ofob_miss_reg)
    if keep_ofob_miss_ext:
        final_list.append(s_ofob_miss_ext)

    # 5. FINAL FOB logic
    has_ppm_surcharge_issue = any(r in r_set for r in surcharge_remarks)
    has_ofob_ppm_issue = s_ofob_miss_reg in r_set or s_ofob_miss_ext in r_set
    has_nike_ofob_issue = added_nike_issue_reg or added_nike_issue_ext
    suppress_final_fob_outputs = has_ppm_surcharge_issue or has_ofob_ppm_issue or has_nike_ofob_issue

    if suppress_final_fob_outputs:
        if (s_final_miss_reg in r_set or s_final_miss_ext in r_set) and trace_steps is not None:
            reasons = []
            if has_ppm_surcharge_issue:
                reasons.append("PPM surcharge mismatch exists")
            if has_ofob_ppm_issue:
                reasons.append("OFOB vs PPM mismatch exists")
            if has_nike_ofob_issue:
                reasons.append("NIKE OFOB issue exists")
            trace_steps.append(
                f"Consolidation: suppressing FINAL FOB remarks because {'; '.join(reasons)}."
            )
    else:
        if s_final_miss_reg in r_set and s_final_miss_ext in r_set:
            final_list.append(s_nike_final_all)
            if trace_steps is not None:
                trace_steps.append(
                    "Consolidation: FINAL FOB regular mismatch + FINAL FOB extended mismatch -> NIKE FINAL FOB issue for all sizes."
                )
        else:
            if s_final_miss_reg in r_set:
                final_list.append(s_final_miss_reg)
                if trace_steps is not None:
                    trace_steps.append(
                        "Consolidation: keeping FINAL FOB regular mismatch because only one size bucket is present."
                    )
            if s_final_miss_ext in r_set:
                final_list.append(s_final_miss_ext)
                if trace_steps is not None:
                    trace_steps.append(
                        "Consolidation: keeping FINAL FOB extended mismatch because only one size bucket is present."
                    )

    # 6. Consolidate PPS OFOB issues
    if added_pps_issue_reg and added_pps_issue_ext:
        final_list.append("PPS OFOB issue for all sizes")
        if trace_steps is not None:
            trace_steps.append(
                "Consolidation: regular + extended PPS OFOB issues collapsed into PPS OFOB issue for all sizes."
            )
    else:
        if added_pps_issue_reg: final_list.append("PPS OFOB issue for regular sizes")
        if added_pps_issue_ext: final_list.append("PPS OFOB issue for extended sizes")

    # 7. Consolidate NIKE OFOB issues
    if added_nike_issue_reg and added_nike_issue_ext:
        final_list.append("NIKE OFOB issue for all sizes")
        if trace_steps is not None:
            trace_steps.append(
                "Consolidation: regular + extended NIKE OFOB issues collapsed into NIKE OFOB issue for all sizes."
            )
    else:
        if added_nike_issue_reg: final_list.append("NIKE OFOB issue for regular sizes")
        if added_nike_issue_ext: final_list.append("NIKE OFOB issue for extended sizes")

    return final_list

def build_trace_output_path(output_path: str) -> str:
    """Place a companion trace file beside the generated output file."""
    base, _ = os.path.splitext(output_path)
    return f"{base}_trace.txt"

def write_trace_report(trace_path: str, source_path: str, output_path: str, row_traces: List[Dict[str, Any]]) -> None:
    """Write a human-readable trace report for generated remarks."""
    with open(trace_path, mode='w', encoding='utf-8') as f:
        f.write("FOB Price Diff Trace Report\n")
        f.write(f"Source master: {source_path}\n")
        f.write(f"Generated output: {output_path}\n")
        f.write(f"Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("\n")

        if not row_traces:
            f.write("No non-CORRECT remarks were generated.\n")
            return

        for row_trace in row_traces:
            f.write("=" * 80 + "\n")
            f.write(f"Row {row_trace['row_number']}\n")
            f.write(f"Master source row: {source_path} :: row {row_trace['row_number']}\n")
            f.write(f"NK SAP PO: {row_trace['po'] or '(blank)'}\n")
            f.write(f"PO LINE ITEM: {row_trace['line'] or '(blank)'}\n")
            f.write(f"STYLE: {row_trace['style'] or '(blank)'}\n")
            f.write(f"BUY MTH: {row_trace['buy_mth'] or '(blank)'}\n")
            f.write(f"SEASON: {row_trace['season'] or '(blank)'}\n")
            f.write(f"SEASON YEAR: {row_trace['season_year'] or '(blank)'}\n")
            f.write(f"Target PPS SEASON_YEAR: {row_trace['target_pps_season_year'] or '(blank)'}\n")
            f.write(f"CW: {row_trace['cw'] or '(blank)'}\n")
            f.write(f"PRICE DIFF REMARKS: {row_trace['final_remark']}\n")
            f.write(f"DPOM - Incorrect FOB: {row_trace['final_dpom']}\n")
            f.write("Trace:\n")
            for step in row_trace["steps"]:
                f.write(f"- {step}\n")
            f.write("\n")

def format_ppm_entry_trace(entry: Dict[str, Any]) -> str:
    """Render a compact trace line for a matched PPM row."""
    return (
        f"{entry.get('source_path', '(unknown file)')} :: row {entry.get('source_row', '?')}"
        f" | Size Description {entry.get('size') or '(blank)'}"
        f" | Surcharge Min Mat Main Body {format_money_trace(entry.get('ag', 0))}"
        f" | Surcharge Min Material Trim {format_money_trace(entry.get('ai', 0))}"
        f" | Surcharge Min Productivity {format_money_trace(entry.get('ak', 0))}"
        f" | Surcharge Misc {format_money_trace(entry.get('am', 0))}"
        f" | Surcharge VAS {format_money_trace(entry.get('ao', 0))}"
        f" | Gross Price/FOB {format_money_trace(entry.get('aq', 0))}"
    )

def format_pps_entry_trace(entry: Dict[str, Any]) -> str:
    """Render a compact trace line for a matched PPS row."""
    return (
        f"{entry.get('source_path', '(unknown file)')} :: row {entry.get('source_row', '?')}"
        f" | SEASON_YEAR {entry.get('season_year') or '(blank)'}"
        f" | COLOR {entry.get('color') or '(blank)'}"
        f" | SIZE_DATA {entry.get('size_data') or '(blank)'}"
        f" | LOCAL_QUOTE_AMOUNT {format_money_trace(entry.get('quote', 0))}"
    )

def format_money_list_trace(values: List[Any]) -> str:
    """Format a list of money values for trace output."""
    if not values:
        return "(none)"
    return ", ".join(format_money_trace(v) for v in values)

def format_ppm_total_breakdown(entry: Dict[str, Any]) -> str:
    """Explain how a PPM total was calculated for a single source row."""
    return (
        f"Surcharge Min Mat Main Body {format_money_trace(entry.get('ag', 0))}"
        f" + Surcharge Min Material Trim {format_money_trace(entry.get('ai', 0))}"
        f" + Surcharge Min Productivity {format_money_trace(entry.get('ak', 0))}"
        f" + Surcharge Misc {format_money_trace(entry.get('am', 0))}"
        f" + Surcharge VAS {format_money_trace(entry.get('ao', 0))}"
        f" + Gross Price/FOB {format_money_trace(entry.get('aq', 0))}"
    )

def _xlrd_cell_value(book, cell):
    """Convert an xlrd cell into a normal Python value."""
    from xlrd import XL_CELL_BOOLEAN, XL_CELL_DATE, xldate_as_datetime

    if cell.ctype == XL_CELL_DATE:
        try:
            return xldate_as_datetime(cell.value, book.datemode)
        except Exception:
            return cell.value
    if cell.ctype == XL_CELL_BOOLEAN:
        return bool(cell.value)
    return cell.value

def _pick_xlrd_sheet(book, preferred_names: Optional[List[str]] = None):
    """Pick an xlrd sheet by preferred name, falling back to the first sheet."""
    if preferred_names:
        normalized_targets = {normalize_header(name) for name in preferred_names}
        for sheet in book.sheets():
            if normalize_header(sheet.name) in normalized_targets:
                return sheet
    return book.sheet_by_index(0)

def load_xls_rows(path: str, preferred_names: Optional[List[str]] = None) -> List[List[Any]]:
    """Load .xls rows with xlrd."""
    if xlrd is None:
        raise RuntimeError("xlrd is required to read .xls files. Please install xlrd.")

    book = xlrd.open_workbook(path, formatting_info=False)
    sheet = _pick_xlrd_sheet(book, preferred_names)
    rows: List[List[Any]] = []
    for r_idx in range(sheet.nrows):
        row_values = []
        for c_idx in range(sheet.ncols):
            row_values.append(_xlrd_cell_value(book, sheet.cell(r_idx, c_idx)))
        rows.append(row_values)
    return rows

def load_xls_as_workbook(path: str, preferred_names: Optional[List[str]] = None) -> Workbook:
    """Load a legacy .xls sheet into an openpyxl Workbook for output as .xlsx."""
    rows = load_xls_rows(path, preferred_names)
    wb = Workbook()
    ws = wb.active
    ws.title = preferred_names[0] if preferred_names else "Sheet1"
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    return wb

def load_file_data(path, log_emit) -> Tuple[List[Any], List[List[Any]], Any]:
    """Load data from Excel or CSV."""
    ext = os.path.splitext(path)[1].lower()

    if ext == '.csv':
        try:
            with open(path, mode='r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                rows = list(reader)
                if not rows:
                    return [], [], None
                return rows, rows, None
        except Exception as e:
            log_emit(f"Error reading CSV {path}: {e}")
            raise e
    elif ext in ['.xlsx', '.xlsm']:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        return rows, rows, wb
    elif ext == '.xls':
        rows = load_xls_rows(path)
        return rows, rows, None
    else:
        raise ValueError("Unsupported file format")

def pick_worksheet(wb, preferred_names: List[str]):
    """Pick worksheet by name (case/whitespace insensitive), fallback to active."""
    if not wb:
        return None
    for name in preferred_names:
        if name in wb.sheetnames:
            return wb[name]
    normalized_targets = {normalize_header(n) for n in preferred_names}
    for name in wb.sheetnames:
        if normalize_header(name) in normalized_targets:
            return wb[name]
    return wb.active

def find_header_row_idx(rows: List[Any], marker: str = "NK SAP PO") -> int:
    """Find the OCCC header row by the unique NK SAP PO header cell."""
    targets = {normalize_header(marker), "NK SAP PO (45/35)"}
    for row_idx, row in enumerate(rows):
        if not row:
            continue
        for cell in row:
            if normalize_header(cell) in targets:
                return row_idx
    raise ValueError("Could not find OCCC header row. Expected a unique 'NK SAP PO' header cell.")

ColumnSpec = Tuple[str, List[str]]

def _dedupe_column_specs(specs: List[ColumnSpec]) -> List[ColumnSpec]:
    """Preserve order while removing duplicate display labels."""
    seen = set()
    unique: List[ColumnSpec] = []
    for label, names in specs:
        if label in seen:
            continue
        seen.add(label)
        unique.append((label, names))
    return unique

def missing_required_columns(headers: List[Any], specs: List[ColumnSpec]) -> List[str]:
    """Return required display labels that are not present in headers."""
    missing = []
    for label, names in _dedupe_column_specs(specs):
        if get_col_index(headers, names) == -1:
            missing.append(label)
    return missing

def require_columns(headers: List[Any], specs: List[ColumnSpec], context: str) -> None:
    """Raise when required columns are missing, so output is not generated from incomplete data."""
    missing = missing_required_columns(headers, specs)
    if missing:
        joined = ", ".join(missing)
        raise ValueError(f"{context} is missing required column(s): {joined}")

PPM_REQUIRED_COLUMNS: List[ColumnSpec] = [
    ("Purchase Order Number / TC PO (85/58)", ["Purchase Order Number", "TC PO (85/58)"]),
    ("PO Line Item Number / PO LINE ITEM", ["PO Line Item Number", "PO LINE ITEM"]),
    ("Size Description", ["Size Description"]),
    ("Surcharge Min Mat Main Body", ["Surcharge Min Mat Main Body"]),
    ("Surcharge Min Material Trim", ["Surcharge Min Material Trim"]),
    ("Surcharge Min Productivity", ["Surcharge Min Productivity"]),
    ("Surcharge Misc", ["Surcharge Misc"]),
    ("Surcharge VAS", ["Surcharge VAS"]),
    ("Gross Price/FOB", ["Gross Price/FOB"]),
]

PPS_REQUIRED_COLUMNS: List[ColumnSpec] = [
    ("STYLE", ["STYLE"]),
    ("EFFECTIVE_DATE", ["EFFECTIVE_DATE"]),
    ("SEASON_YEAR / SEASON YEAR", ["SEASON_YEAR", "SEASON YEAR"]),
    ("COLOR", ["COLOR"]),
    ("SIZE_DATA", ["SIZE_DATA"]),
    ("LOCAL_QUOTE_AMOUNT", ["LOCAL_QUOTE_AMOUNT"]),
]

OCCC_PPM_REQUIRED_COLUMNS: List[ColumnSpec] = [
    ("NK SAP PO / NK SAP PO (45/35)", ["NK SAP PO (45/35)", "NK SAP PO"]),
    ("PO LINE ITEM", ["PO LINE ITEM"]),
    ("S/C Min Production (ZPMX)", ["S/C Min Production (ZPMX)"]),
    ("S/C Min Material (ZMMX)", ["S/C Min Material (ZMMX)"]),
    ("S/C Min Material (ZMMX) Comment", ["S/C Min Material (ZMMX) Comment"]),
    ("S/C Misc (ZMSX)", ["S/C Misc (ZMSX)"]),
    ("S/C Misc (ZMSX) Comment", ["S/C Misc (ZMSX) Comment"]),
    ("S/C VAS Manual (ZVAX)", ["S/C VAS Manual (ZVAX)"]),
    ("OFOB (Regular sizes)", ["OFOB (Regular sizes)"]),
    ("OFOB (Extended sizes)", ["OFOB (Extended sizes)"]),
    ("FINAL FOB (Regular sizes)", ["FINAL FOB (Regular sizes)"]),
    ("FINAL FOB (Extended sizes)", ["FINAL FOB (Extended sizes)", "FINAL FOB (Extended sizes) (2)"]),
    ("Extended Sizes", ["Extended Sizes", "Extended Sizes (2)", "EXT SIZE", "EXT SIZES", "EXTENDED SIZE"]),
]

OCCC_PPS_REQUIRED_COLUMNS: List[ColumnSpec] = [
    ("STYLE", ["STYLE"]),
    ("BUY MTH", ["BUY MTH"]),
    ("SEASON", ["SEASON"]),
    ("SEASON YEAR / SEASON_YEAR", ["SEASON YEAR", "SEASON_YEAR"]),
    ("CW", ["CW"]),
    ("OFOB (Regular sizes)", ["OFOB (Regular sizes)"]),
    ("OFOB (Extended sizes)", ["OFOB (Extended sizes)"]),
    ("Extended Sizes", ["Extended Sizes", "Extended Sizes (2)", "EXT SIZE", "EXT SIZES", "EXTENDED SIZE"]),
]

def required_occc_columns_for_run(has_ppm: bool, has_pps: bool) -> List[ColumnSpec]:
    """Only require OCCC columns needed by the selected validation path(s)."""
    specs: List[ColumnSpec] = []
    if has_ppm:
        specs.extend(OCCC_PPM_REQUIRED_COLUMNS)
    if has_pps:
        specs.extend(OCCC_PPS_REQUIRED_COLUMNS)
    return _dedupe_column_specs(specs)

def process_logic(master_files, ppm_files, pps_files, log_emit, report_emit, debug_mode: bool = False) -> Tuple[str, int, int]:
    success_count = 0
    fail_count = 0
    last_output = ""
    row_emit = report_emit if callable(report_emit) else log_emit
    run_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    support_errors: List[str] = []

    # --- 1. Parse PPM Files ---
    ppm_lookup = {}
    log_emit("Parsing PPM Files...")
    for ppm_path in ppm_files:
        try:
            rows, _, _ = load_file_data(ppm_path, log_emit)
            if not rows: continue

            header_row_idx = 0
            headers = [str(c) for c in rows[header_row_idx]]
            require_columns(headers, PPM_REQUIRED_COLUMNS, f"PPM file '{os.path.basename(ppm_path)}'")

            col_po = get_col_index(headers, ["Purchase Order Number", "TC PO (85/58)"])
            col_line = get_col_index(headers, ["PO Line Item Number", "PO LINE ITEM"])
            col_size = get_col_index(headers, ["Size Description"])

            # Costs
            col_ag = get_col_index(headers, ["Surcharge Min Mat Main Body"])
            col_ai = get_col_index(headers, ["Surcharge Min Material Trim"])
            col_ak = get_col_index(headers, ["Surcharge Min Productivity"])
            col_am = get_col_index(headers, ["Surcharge Misc"])
            col_ao = get_col_index(headers, ["Surcharge VAS"])
            col_aq = get_col_index(headers, ["Gross Price/FOB"])

            if col_po == -1 or col_line == -1:
                continue

            for r_idx in range(header_row_idx + 1, len(rows)):
                row = rows[r_idx]
                if not row: continue

                po_num = str(row[col_po]).strip()
                line_item = normalize_line_item(get_row_value(row, col_line, ""))
                key = (po_num, line_item)

                costs = {
                    'ag': safe_decimal(get_row_value(row, col_ag, 0)),
                    'ai': safe_decimal(get_row_value(row, col_ai, 0)),
                    'ak': safe_decimal(get_row_value(row, col_ak, 0)),
                    'am': safe_decimal(get_row_value(row, col_am, 0)),
                    'ao': safe_decimal(get_row_value(row, col_ao, 0)),
                    'aq': safe_decimal(get_row_value(row, col_aq, 0)),
                    'size': str(get_row_value(row, col_size, '') or '').strip(),
                    'source_path': ppm_path,
                    'source_row': r_idx + 1,
                }

                if key not in ppm_lookup: ppm_lookup[key] = []
                ppm_lookup[key].append(costs)
        except Exception as e:
            msg = f"Error parsing PPM {os.path.basename(ppm_path)}: {e}"
            log_emit(msg)
            support_errors.append(msg)

    # --- 2. Parse PPS Files ---
    pps_lookup = {}
    log_emit("Parsing PPS Files...")
    for pps_path in pps_files:
        try:
            rows, _, _ = load_file_data(pps_path, log_emit)
            if not rows: continue

            headers = [str(c) for c in rows[0]]
            require_columns(headers, PPS_REQUIRED_COLUMNS, f"PPS file '{os.path.basename(pps_path)}'")

            col_style = get_col_index(headers, ["STYLE"])
            col_eff_date = get_col_index(headers, ["EFFECTIVE_DATE"])
            col_season_year = get_col_index(headers, ["SEASON_YEAR", "SEASON YEAR"])
            col_color = get_col_index(headers, ["COLOR"])
            col_size_data = get_col_index(headers, ["SIZE_DATA"])
            col_quote = get_col_index(headers, ["LOCAL_QUOTE_AMOUNT"])

            if col_style == -1 or col_eff_date == -1:
                continue

            for r_idx in range(1, len(rows)):
                row = rows[r_idx]
                if not row: continue

                style = str(row[col_style]).strip()
                eff_date = normalize_date_str(row[col_eff_date])

                if not style or not eff_date:
                    continue

                color = str(row[col_color]).strip() if col_color != -1 and row[col_color] is not None else ""
                season_year = normalize_pps_season_year(row[col_season_year]) if col_season_year != -1 and row[col_season_year] is not None else ""
                size_data = str(row[col_size_data]).strip() if col_size_data != -1 and row[col_size_data] is not None else ""
                quote = safe_decimal(get_row_value(row, col_quote, 0))

                key = (style, eff_date)
                entry = {
                    'season_year': season_year,
                    'color': color,
                    'size_data': size_data,
                    'quote': quote,
                    'source_path': pps_path,
                    'source_row': r_idx + 1,
                }

                if key not in pps_lookup: pps_lookup[key] = []
                pps_lookup[key].append(entry)

        except Exception as e:
            msg = f"Error parsing PPS {os.path.basename(pps_path)}: {e}"
            log_emit(msg)
            support_errors.append(msg)

    log_emit(f"PPS Data Loaded. Found {len(pps_lookup)} Style/Date keys.")

    if support_errors:
        log_emit("Validation aborted: one or more selected report files are missing required columns or could not be parsed.")
        return "", 0, len(master_files)

    # --- 3. Process OCCC Files ---
    for occc_path in master_files:
        try:
            # === xlwings Magic: Calculate Formulas ===
            if occc_path.lower().endswith(('.xlsx', '.xlsm')):
                refresh_excel_formulas(occc_path, log_emit)

            log_emit(f"Processing Master: {os.path.basename(occc_path)}")
            ext = os.path.splitext(occc_path)[1].lower()
            is_excel = ext in ('.xlsx', '.xlsm', '.xls')
            is_xls = ext == '.xls'

            if ext in ('.xlsx', '.xlsm'):
                keep_vba = ext == '.xlsm'
                wb_read = load_workbook(occc_path, data_only=True, keep_vba=keep_vba)
                ws_read = pick_worksheet(wb_read, ["OCCC"])
                rows_read = list(ws_read.values)
                wb_write = load_workbook(occc_path, data_only=False, keep_vba=keep_vba)
                ws_write = pick_worksheet(wb_write, ["OCCC"])
            elif is_xls:
                wb_write = load_xls_as_workbook(occc_path, ["OCCC"])
                ws_write = wb_write.active
                rows_read = list(ws_write.values)
            else:
                rows_read, _, _ = load_file_data(occc_path, log_emit)
                output_csv_data = [list(r) for r in rows_read]
                ws_write = None

            if not rows_read:
                log_emit("Master file is empty.")
                fail_count += 1
                continue

            header_idx = find_header_row_idx(rows_read)
            headers = [str(x) for x in rows_read[header_idx]]
            occc_required_columns = required_occc_columns_for_run(bool(ppm_files), bool(pps_files))
            require_columns(headers, occc_required_columns, f"OCCC master '{os.path.basename(occc_path)}'")

            # Map Columns
            idx_nk_po = get_col_index(headers, ["NK SAP PO (45/35)", "NK SAP PO"])
            idx_line = get_col_index(headers, ["PO LINE ITEM"])
            idx_sc_min_prod = get_col_index(headers, ["S/C Min Production (ZPMX)"])
            idx_sc_min_mat = get_col_index(headers, ["S/C Min Material (ZMMX)"])
            idx_sc_min_mat_comment = get_col_index(headers, ["S/C Min Material (ZMMX) Comment"])
            idx_sc_misc = get_col_index(headers, ["S/C Misc (ZMSX)"])
            idx_sc_misc_comment = get_col_index(headers, ["S/C Misc (ZMSX) Comment"])
            idx_sc_vas = get_col_index(headers, ["S/C VAS Manual (ZVAX)"])

            idx_style = get_col_index(headers, ["STYLE"])
            idx_buy_mth = get_col_index(headers, ["BUY MTH"])
            idx_season = get_col_index(headers, ["SEASON"])
            idx_season_year = get_col_index(headers, ["SEASON YEAR", "SEASON_YEAR"])
            idx_cw = get_col_index(headers, ["CW"])

            idx_ofob_reg = get_col_index(headers, ["OFOB (Regular sizes)"])
            idx_ofob_ext = get_col_index(headers, ["OFOB (Extended sizes)"])
            idx_final_reg = get_col_index(headers, ["FINAL FOB (Regular sizes)"])
            idx_final_ext = get_col_index(headers, ["FINAL FOB (Extended sizes)", "FINAL FOB (Extended sizes) (2)"])
            idx_ext_sizes_def = get_col_index(headers, [
                "Extended Sizes",
                "Extended Sizes (2)",
                "EXT SIZE",
                "EXT SIZES",
                "EXTENDED SIZE",
            ])

            idx_remarks = get_col_index(headers, ["PRICE DIFF REMARKS"])
            idx_dpom_fob = get_col_index(headers, ["DPOM - Incorrect FOB"])

            # Init Header for Remarks
            if idx_remarks == -1:
                # Put it at the end when missing
                insert_pos = len(headers)
                if is_excel:
                    ws_write.insert_cols(insert_pos + 1)
                    ws_write.cell(row=header_idx+1, column=insert_pos+1).value = "PRICE DIFF REMARKS"
                    idx_remarks = insert_pos
                else:
                    output_csv_data[header_idx].append("PRICE DIFF REMARKS")
                    idx_remarks = len(headers)

            # Init Header for DPOM - Incorrect FOB (re-check in case it shifted due to insert)
            if is_excel:
                # Re-fetch headers if we inserted a column to ensure we don't mess up indexing
                headers = [str(cell.value) for cell in ws_write[header_idx+1]]
                idx_dpom_fob = get_col_index(headers, ["DPOM - Incorrect FOB"])

            if idx_dpom_fob == -1:
                # Place it after Remarks, or at the end
                insert_pos = idx_remarks + 1
                if is_excel:
                    ws_write.insert_cols(insert_pos + 1)
                    ws_write.cell(row=header_idx+1, column=insert_pos+1).value = "DPOM - Incorrect FOB"
                    idx_dpom_fob = insert_pos
                else:
                    output_csv_data[header_idx].append("DPOM - Incorrect FOB")
                    idx_dpom_fob = len(output_csv_data[header_idx]) - 1

            trace_rows = []

            for r_i in range(header_idx + 1, len(rows_read)):
                row_vals = rows_read[r_i]
                if not row_vals:
                    continue

                remarks = []
                dpom_errors = [] # Store "Size Price" mismatches
                row_trace_steps = []

                # Common row context
                po_val = str(get_row_value(row_vals, idx_nk_po, "") or "").strip()
                line_val = normalize_line_item(get_row_value(row_vals, idx_line, ""))
                style_val = str(get_row_value(row_vals, idx_style, "") or "").strip()
                buy_mth_val = str(get_row_value(row_vals, idx_buy_mth, "") or "").strip()
                season_val = str(get_row_value(row_vals, idx_season, "") or "").strip().upper()
                season_year_val = str(get_row_value(row_vals, idx_season_year, "") or "").strip()
                target_pps_season_year = build_target_pps_season_year(season_val, season_year_val)
                cw_val = str(get_row_value(row_vals, idx_cw, "") or "").strip()
                ext_threshold = str(get_row_value(row_vals, idx_ext_sizes_def, "") or "").strip()

                # --- PPM Comparison ---
                if po_val and line_val:
                    ppm_entries = ppm_lookup.get((po_val, line_val))
                    if ppm_entries:
                        row_trace_steps.append(
                            f"PPM lookup matched {len(ppm_entries)} row(s) for NK SAP PO {po_val} / PO LINE ITEM {line_val}."
                        )
                        for idx_ppm, entry in enumerate(ppm_entries, start=1):
                            row_trace_steps.append(
                                f"PPM matched row {idx_ppm}: {format_ppm_entry_trace(entry)}."
                            )

                        # Avg calc for surcharges
                        count = len(ppm_entries)
                        ppm_ag_values = [entry['ag'] for entry in ppm_entries]
                        ppm_ai_values = [entry['ai'] for entry in ppm_entries]
                        ppm_am_values = [entry['am'] for entry in ppm_entries]
                        ppm_ao_values = [entry['ao'] for entry in ppm_entries]

                        ave_ppm_ag = money_average(ppm_ag_values)
                        ave_ppm_ai = money_average(ppm_ai_values)
                        ave_ppm_am = money_average(ppm_am_values)
                        ave_ppm_ao = money_average(ppm_ao_values)

                        # Surcharge Checks - treat >= $0.01 as mismatch
                        if idx_sc_min_prod != -1 and money_abs_diff(row_vals[idx_sc_min_prod], ave_ppm_ag) >= MONEY_CENT:
                            remarks.append("S/C MIN PRODUCTION (ZPMX) doesn't match")
                            row_trace_steps.append(
                                f"S/C Min Production mismatch: OCCC {format_money_trace(row_vals[idx_sc_min_prod])} from column S/C Min Production (ZPMX) vs PPM avg {format_money_trace(ave_ppm_ag)} from Surcharge Min Mat Main Body across {count} matched PPM row(s) [{format_money_list_trace(ppm_ag_values)}]."
                            )

                        # Min Mat
                        occc_zmmx = safe_decimal(get_row_value(row_vals, idx_sc_min_mat, 0))
                        zmmx_cmt = str(get_row_value(row_vals, idx_sc_min_mat_comment, "") or "").strip().upper()
                        if idx_sc_min_mat != -1 and zmmx_cmt != "DN" and money_abs_diff(occc_zmmx, ave_ppm_ai) >= MONEY_CENT:
                            remarks.append("S/C Min Material (ZMMX) doesn't match")
                            row_trace_steps.append(
                                f"S/C Min Material mismatch: OCCC {format_money_trace(occc_zmmx)} from column S/C Min Material (ZMMX) vs PPM avg {format_money_trace(ave_ppm_ai)} from Surcharge Min Material Trim across {count} matched PPM row(s) [{format_money_list_trace(ppm_ai_values)}]."
                            )

                        # Misc
                        occc_zmsx = safe_decimal(get_row_value(row_vals, idx_sc_misc, 0))
                        zmsx_cmt = str(get_row_value(row_vals, idx_sc_misc_comment, "") or "").strip().upper()
                        if idx_sc_misc != -1 and zmsx_cmt != "DN" and money_abs_diff(occc_zmsx, ave_ppm_am) >= MONEY_CENT:
                            remarks.append("S/C Misc (ZMSX) doesn't match")
                            row_trace_steps.append(
                                f"S/C Misc mismatch: OCCC {format_money_trace(occc_zmsx)} from column S/C Misc (ZMSX) vs PPM avg {format_money_trace(ave_ppm_am)} from Surcharge Misc across {count} matched PPM row(s) [{format_money_list_trace(ppm_am_values)}]."
                            )

                        if idx_sc_vas != -1 and money_abs_diff(row_vals[idx_sc_vas], ave_ppm_ao) >= MONEY_CENT:
                            remarks.append("S/C VAS Manual (ZVAX) doesn't match")
                            row_trace_steps.append(
                                f"S/C VAS mismatch: OCCC {format_money_trace(row_vals[idx_sc_vas])} from column S/C VAS Manual (ZVAX) vs PPM avg {format_money_trace(ave_ppm_ao)} from Surcharge VAS across {count} matched PPM row(s) [{format_money_list_trace(ppm_ao_values)}]."
                            )

                        # OFOB / Final FOB Checks
                        occc_ofob_reg = safe_decimal(get_row_value(row_vals, idx_ofob_reg, 0))
                        occc_ofob_ext = safe_decimal(get_row_value(row_vals, idx_ofob_ext, 0))
                        occc_final_reg = safe_decimal(get_row_value(row_vals, idx_final_reg, 0))
                        occc_final_ext = safe_decimal(get_row_value(row_vals, idx_final_ext, 0))

                        ofob_mismatch_found_reg = False
                        ofob_mismatch_found_ext = False
                        fob_mismatch_found_reg = False
                        fob_mismatch_found_ext = False

                        for entry in ppm_entries:
                            ppm_gross_fob = safe_decimal(entry['aq']).quantize(MONEY_CENT, rounding=ROUND_HALF_UP)
                            ppm_total = (
                                safe_decimal(entry['ag'])
                                + safe_decimal(entry['ai'])
                                + safe_decimal(entry['ak'])
                                + safe_decimal(entry['am'])
                                + safe_decimal(entry['ao'])
                                + safe_decimal(entry['aq'])
                            ).quantize(MONEY_CENT, rounding=ROUND_HALF_UP)
                            is_ext = is_extended_size(entry['size'], ext_threshold)
                            target_ofob = safe_decimal(occc_ofob_ext if is_ext else occc_ofob_reg).quantize(MONEY_CENT, rounding=ROUND_HALF_UP)
                            target_fob = safe_decimal(occc_final_ext if is_ext else occc_final_reg).quantize(MONEY_CENT, rounding=ROUND_HALF_UP)

                            # OFOB vs PPM Gross Price/FOB - treat >= $0.01 as mismatch
                            if ppm_gross_fob > MONEY_ZERO and money_abs_diff(target_ofob, ppm_gross_fob) >= MONEY_CENT:
                                lbl = "Extended" if is_ext else "Regular"
                                row_trace_steps.append(
                                    f"OFOB {lbl.lower()} mismatch for size {entry['size'] or '(blank)'} using PPM row {entry.get('source_path', '(unknown file)')} :: row {entry.get('source_row', '?')}: OCCC {format_money_trace(target_ofob)} from column OFOB ({lbl} sizes) vs PPM Gross Price/FOB {format_money_trace(ppm_gross_fob)}."
                                )

                                label_seen = ofob_mismatch_found_ext if is_ext else ofob_mismatch_found_reg
                                if not label_seen:
                                    remarks.append(f"OFOB ({lbl} sizes) doesn't match with PPM")
                                    if is_ext:
                                        ofob_mismatch_found_ext = True
                                    else:
                                        ofob_mismatch_found_reg = True

                            # treat >= $0.01 as mismatch
                            if ppm_total > MONEY_ZERO and money_abs_diff(target_fob, ppm_total) >= MONEY_CENT:
                                lbl = "Extended" if is_ext else "Regular"

                                # Add to DPOM Error List: "Size Price"
                                dpom_errors.append(f"{entry['size']} {format_money_trace(ppm_total)}")
                                row_trace_steps.append(
                                    f"FINAL FOB {lbl.lower()} mismatch for size {entry['size'] or '(blank)'} using PPM row {entry.get('source_path', '(unknown file)')} :: row {entry.get('source_row', '?')}: OCCC {format_money_trace(target_fob)} from column FINAL FOB ({lbl} sizes) vs PPM total {format_money_trace(ppm_total)} = {format_ppm_total_breakdown(entry)}."
                                )

                                # Add one FINAL FOB remark per size bucket while still keeping all DPOM size mismatches.
                                label_seen = fob_mismatch_found_ext if is_ext else fob_mismatch_found_reg
                                if not label_seen:
                                    row_emit(f"Mismatch Row {r_i+1} PO {po_val}: {lbl} Size - OCCC {format_money_trace(target_fob)} vs PPM {format_money_trace(ppm_total)}")
                                    remarks.append(f"FINAL FOB ({lbl} sizes) doesn't match with PPM")
                                    if is_ext:
                                        fob_mismatch_found_ext = True
                                    else:
                                        fob_mismatch_found_reg = True

                                # Do NOT break here. Continue checking other sizes for DPOM column.
                    else:
                        row_trace_steps.append(
                            f"PPM lookup found no rows for NK SAP PO {po_val} / PO LINE ITEM {line_val}."
                        )
                else:
                    row_trace_steps.append("PPM lookup skipped because NK SAP PO or PO LINE ITEM is blank.")

                # --- PPS Comparison ---
                if style_val and buy_mth_val:
                    target_date = calculate_target_effective_date(buy_mth_val)
                    if target_date:
                        row_trace_steps.append(
                            f"PPS lookup key: STYLE {style_val}, BUY MTH {buy_mth_val} -> EFFECTIVE_DATE {target_date}, SEASON {season_val or '(blank)'}, SEASON YEAR {season_year_val or '(blank)'} -> PPS SEASON_YEAR {target_pps_season_year or '(blank)'}, CW {cw_val or '(blank)'}."
                        )
                        pps_candidates = pps_lookup.get((style_val, target_date))
                        if pps_candidates:
                            row_trace_steps.append(
                                f"PPS style/date pool matched {len(pps_candidates)} row(s) for STYLE {style_val} and EFFECTIVE_DATE {target_date}."
                            )
                            season_filtered_rows = pps_candidates
                            if target_pps_season_year:
                                season_filtered_rows = [r for r in pps_candidates if r['season_year'] == target_pps_season_year]
                                row_trace_steps.append(
                                    f"PPS SEASON_YEAR matches for {target_pps_season_year}: {len(season_filtered_rows)} row(s)."
                                )
                                other_season_rows = [r for r in pps_candidates if r['season_year'] != target_pps_season_year]
                                if other_season_rows:
                                    other_season_values = sorted({r['season_year'] or '(blank)' for r in other_season_rows})
                                    sample_seasons = ", ".join(other_season_values[:10])
                                    suffix = "" if len(other_season_values) <= 10 else ", ..."
                                    row_trace_steps.append(
                                        f"PPS rows excluded by SEASON_YEAR filter: {len(other_season_rows)} row(s) with other SEASON_YEAR values ({sample_seasons}{suffix})."
                                    )
                            else:
                                row_trace_steps.append(
                                    "PPS SEASON_YEAR filter skipped because OCCC SEASON or SEASON YEAR is blank/invalid."
                                )

                            if not season_filtered_rows:
                                remarks.append("No matching PPS found")
                                row_trace_steps.append(
                                    f"No PPS row matched SEASON_YEAR {target_pps_season_year or '(blank)'} after STYLE {style_val} and EFFECTIVE_DATE {target_date}."
                                )
                            else:
                                matched_rows = [r for r in season_filtered_rows if r['color'] == cw_val]
                                blank_color_rows = [r for r in season_filtered_rows if not r['color']]
                                other_color_rows = [r for r in season_filtered_rows if r['color'] and r['color'] != cw_val]

                                row_trace_steps.append(
                                    f"PPS exact COLOR matches for {cw_val or '(blank)'}: {len(matched_rows)} row(s)."
                                )
                                row_trace_steps.append(
                                    f"PPS blank COLOR rows available for fallback: {len(blank_color_rows)} row(s)."
                                )
                                if other_color_rows:
                                    other_colors = sorted({r['color'] for r in other_color_rows})
                                    sample_colors = ", ".join(other_colors[:10])
                                    suffix = "" if len(other_colors) <= 10 else ", ..."
                                    row_trace_steps.append(
                                        f"PPS rows excluded by color filter: {len(other_color_rows)} row(s) with other COLOR values ({sample_colors}{suffix})."
                                    )
                                if matched_rows:
                                    row_trace_steps.append(
                                        f"Using exact COLOR match row(s) for COLOR {cw_val or '(blank)' }."
                                    )
                                    for idx_pps, entry in enumerate(matched_rows, start=1):
                                        row_trace_steps.append(
                                            f"PPS color-matched row {idx_pps}: {format_pps_entry_trace(entry)}."
                                        )
                                else:
                                    matched_rows = blank_color_rows
                                    if matched_rows:
                                        row_trace_steps.append(
                                            f"No PPS row matched COLOR {cw_val or '(blank)'}; using {len(matched_rows)} blank COLOR fallback row(s)."
                                        )
                                        for idx_pps, entry in enumerate(matched_rows, start=1):
                                            row_trace_steps.append(
                                                f"PPS blank-color fallback row {idx_pps}: {format_pps_entry_trace(entry)}."
                                            )

                                if not matched_rows:
                                    remarks.append("No matching PPS found")
                                    row_trace_steps.append(
                                        f"No PPS row matched COLOR {cw_val or '(blank)'} and no blank COLOR fallback row was available."
                                    )
                                else:
                                    # Regular - THRESHOLD 0.01
                                    reg_match = next((r for r in matched_rows if not r['size_data']), None)
                                    occc_ofob_reg = safe_decimal(get_row_value(row_vals, idx_ofob_reg, 0))
                                    if reg_match:
                                        row_trace_steps.append(
                                            f"Selected PPS regular row: {format_pps_entry_trace(reg_match)}."
                                        )
                                        if money_abs_diff(reg_match['quote'], occc_ofob_reg) == 0:
                                            remarks.append("PPS OFOB match for regular sizes")
                                            row_trace_steps.append(
                                                f"Regular PPS comparison matched: PPS LOCAL_QUOTE_AMOUNT {format_money_trace(reg_match['quote'])} vs OCCC OFOB (Regular sizes) {format_money_trace(occc_ofob_reg)}."
                                            )
                                        else:
                                            remarks.append("PPS OFOB doesn't match for regular sizes")
                                            row_trace_steps.append(
                                                f"Regular PPS comparison mismatched: PPS LOCAL_QUOTE_AMOUNT {format_money_trace(reg_match['quote'])} vs OCCC OFOB (Regular sizes) {format_money_trace(occc_ofob_reg)}."
                                            )
                                    elif occc_ofob_reg > MONEY_ZERO:
                                        remarks.append("PPS OFOB missing regular size entry")
                                        row_trace_steps.append(
                                            f"No PPS regular row with blank SIZE_DATA was found while OCCC OFOB (Regular sizes) is {format_money_trace(occc_ofob_reg)}."
                                        )

                                    # Extended - THRESHOLD 0.01
                                    occc_ofob_ext = safe_decimal(get_row_value(row_vals, idx_ofob_ext, 0))
                                    if ext_threshold not in ["-", "", "NONE", "NA"] or occc_ofob_ext > MONEY_ZERO:
                                        ext_match = next((r for r in matched_rows if is_extended_size(r['size_data'], ext_threshold)), None)
                                        if ext_match:
                                            row_trace_steps.append(
                                                f"Selected PPS extended row with Extended Sizes threshold {ext_threshold or '(blank)'}: {format_pps_entry_trace(ext_match)}."
                                            )
                                            if money_abs_diff(ext_match['quote'], occc_ofob_ext) == 0:
                                                remarks.append("PPS OFOB match for extended sizes")
                                                row_trace_steps.append(
                                                    f"Extended PPS comparison matched: PPS LOCAL_QUOTE_AMOUNT {format_money_trace(ext_match['quote'])} vs OCCC OFOB (Extended sizes) {format_money_trace(occc_ofob_ext)}."
                                                )
                                            else:
                                                remarks.append("PPS OFOB doesn't match for extended sizes")
                                                row_trace_steps.append(
                                                    f"Extended PPS comparison mismatched: PPS LOCAL_QUOTE_AMOUNT {format_money_trace(ext_match['quote'])} vs OCCC OFOB (Extended sizes) {format_money_trace(occc_ofob_ext)}."
                                                )
                                        elif occc_ofob_ext > MONEY_ZERO:
                                            remarks.append("PPS OFOB missing extended size entry")
                                            row_trace_steps.append(
                                                f"No PPS extended row met Extended Sizes threshold {ext_threshold or '(blank)'} while OCCC OFOB (Extended sizes) is {format_money_trace(occc_ofob_ext)}."
                                            )
                        else:
                            remarks.append("No matching PPS found")
                            row_trace_steps.append(
                                f"No PPS rows matched STYLE {style_val} and EFFECTIVE_DATE {target_date}."
                            )
                    else:
                        remarks.append("Invalid BUY MTH format")
                        row_trace_steps.append(f"BUY MTH {buy_mth_val} could not be converted into a PPS EFFECTIVE_DATE.")
                else:
                    row_trace_steps.append("PPS lookup skipped because STYLE or BUY MTH is blank.")

                # --- Post-Processing ---
                if remarks:
                    row_trace_steps.append(f"Raw remarks before consolidation: {'; '.join(remarks)}.")
                    remarks = refine_remarks(remarks, row_trace_steps)

                # --- Write Output ---

                # 1. PRICE DIFF REMARKS
                final_remark = "; ".join(remarks) if remarks else "CORRECT"

                # 2. DPOM - Incorrect FOB
                final_dpom_val = " / ".join(dpom_errors) if dpom_errors else "CORRECT"

                if is_excel:
                    # Write Remarks
                    target_col_idx = idx_remarks + 1
                    if ws_write.cell(row=header_idx+1, column=target_col_idx).value != "PRICE DIFF REMARKS":
                        ws_write.cell(row=header_idx+1, column=target_col_idx).value = "PRICE DIFF REMARKS"
                    ws_write.cell(row=r_i+1, column=target_col_idx).value = final_remark

                    # Write DPOM
                    target_dpom_idx = idx_dpom_fob + 1
                    if ws_write.cell(row=header_idx+1, column=target_dpom_idx).value != "DPOM - Incorrect FOB":
                        ws_write.cell(row=header_idx+1, column=target_dpom_idx).value = "DPOM - Incorrect FOB"
                    ws_write.cell(row=r_i+1, column=target_dpom_idx).value = final_dpom_val

                else:
                    # CSV Handling

                    # Ensure list is long enough for Remarks
                    while len(output_csv_data[r_i]) <= idx_remarks:
                        output_csv_data[r_i].append("")
                    output_csv_data[r_i][idx_remarks] = final_remark

                    # Ensure list is long enough for DPOM
                    while len(output_csv_data[r_i]) <= idx_dpom_fob:
                        output_csv_data[r_i].append("")
                    output_csv_data[r_i][idx_dpom_fob] = final_dpom_val

                if final_remark != "CORRECT" or final_dpom_val != "CORRECT":
                    row_trace_steps.append(f"Final PRICE DIFF REMARKS: {final_remark}.")
                    row_trace_steps.append(f"Final DPOM - Incorrect FOB: {final_dpom_val}.")
                    trace_rows.append({
                        "row_number": r_i + 1,
                        "po": po_val,
                        "line": line_val,
                        "style": style_val,
                        "buy_mth": buy_mth_val,
                        "season": season_val,
                        "season_year": season_year_val,
                        "target_pps_season_year": target_pps_season_year,
                        "cw": cw_val,
                        "final_remark": final_remark,
                        "final_dpom": final_dpom_val,
                        "steps": row_trace_steps,
                    })

            out_path = build_timestamped_copy_path(occc_path, run_timestamp, label="updated", output_ext=".xlsx" if is_xls else None)
            if is_excel:
                wb_write.save(out_path)
            else:
                with open(out_path, mode='w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    writer.writerows(output_csv_data)

            success_count += 1
            last_output = out_path
            log_emit(f"Output saved: {out_path}")
            if debug_mode:
                trace_path = build_trace_output_path(out_path)
                write_trace_report(trace_path, occc_path, out_path, trace_rows)
                log_emit(f"Trace saved: {trace_path}")

        except Exception as e:
            log_emit(f"Failed to process {os.path.basename(occc_path)}: {e}")
            fail_count += 1

    return last_output, success_count, fail_count

# --- UI Class ---

class MainWidget(QWidget):
    log_message = Signal(str)
    report_message = Signal(str)
    processing_done = Signal(int, int, str)

    def __init__(self):
        super().__init__()
        self.setObjectName("mmu_widget")
        self._build_ui()
        self._connect_signals()

    def _build_ui(self):
        self.desc_label = QLabel("", self)
        self.desc_label.setWordWrap(True)
        self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.desc_label.setStyleSheet(
            "color: #dcdcdc; background: transparent; padding: 6px; "
            "border: 1px solid #3a3a3a; border-radius: 6px;"
        )
        self.set_long_description("")

        self.debug_mode_toggle = None
        self.debug_mode_checkbox = None
        self.debug_mode_label = QLabel("Debug mode (generate trace text file)", self)
        self.debug_mode_label.setStyleSheet("color: #dcdcdc; background: transparent;")
        if CheckIconButton is not None:
            self.debug_mode_toggle = CheckIconButton(self, initially_checked=False)
        else:
            self.debug_mode_checkbox = QCheckBox(self.debug_mode_label.text(), self)
            self.debug_mode_checkbox.setChecked(False)
            self.debug_mode_checkbox.setStyleSheet("color: #dcdcdc; background: transparent;")

        self.select_master_btn = PrimaryPushButton("Select Master (OCCC)", self)
        self.select_ppm_btn = PrimaryPushButton("Select PPM Reports", self)
        self.select_pps_btn = PrimaryPushButton("Select PPS Reports", self)
        self.run_btn = PrimaryPushButton("Run Validation", self)

        self.master_files_label = QLabel("Master file(s)", self)
        self.ppm_files_label = QLabel("PPM report file(s)", self)
        self.pps_files_label = QLabel("PPS report file(s)", self)
        self.logs_label = QLabel("Process logs", self)
        self.reports_label = QLabel("Report output", self)

        for lbl in [self.master_files_label, self.ppm_files_label, self.pps_files_label, self.logs_label, self.reports_label]:
            lbl.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        shared_style = (
            "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        self.master_files_box = QTextEdit(self)
        self.master_files_box.setReadOnly(True)
        self.master_files_box.setStyleSheet(shared_style)

        self.ppm_files_box = QTextEdit(self)
        self.ppm_files_box.setReadOnly(True)
        self.ppm_files_box.setStyleSheet(shared_style)

        self.pps_files_box = QTextEdit(self)
        self.pps_files_box.setReadOnly(True)
        self.pps_files_box.setStyleSheet(shared_style)

        self.reports_box = QTextEdit(self)
        self.reports_box.setReadOnly(True)
        self.reports_box.setStyleSheet(shared_style)

        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setStyleSheet(shared_style)

        main_layout = QVBoxLayout(self)
        main_layout.addWidget(self.desc_label)

        debug_layout = QHBoxLayout()
        if self.debug_mode_toggle is not None:
            debug_layout.addWidget(self.debug_mode_toggle, 0, Qt.AlignVCenter)
            debug_layout.addWidget(self.debug_mode_label, 0, Qt.AlignVCenter)
            debug_layout.addStretch(1)
            main_layout.addLayout(debug_layout)
        elif self.debug_mode_checkbox is not None:
            main_layout.addWidget(self.debug_mode_checkbox, 0)
        else:
            main_layout.addWidget(self.debug_mode_label, 0)

        row1 = QHBoxLayout()
        row1.addWidget(self.select_master_btn)
        row1.addWidget(self.select_ppm_btn)
        row1.addWidget(self.select_pps_btn)
        main_layout.addLayout(row1)

        row_btn = QHBoxLayout()
        row_btn.addStretch()
        row_btn.addWidget(self.run_btn)
        row_btn.addStretch()
        main_layout.addLayout(row_btn)

        row_labels = QHBoxLayout()
        row_labels.addWidget(self.master_files_label)
        row_labels.addWidget(self.ppm_files_label)
        row_labels.addWidget(self.pps_files_label)
        main_layout.addLayout(row_labels)

        row_boxes = QHBoxLayout()
        row_boxes.addWidget(self.master_files_box)
        row_boxes.addWidget(self.ppm_files_box)
        row_boxes.addWidget(self.pps_files_box)
        main_layout.addLayout(row_boxes, 2)

        row_log_lbl = QHBoxLayout()
        row_log_lbl.addWidget(self.reports_label)
        row_log_lbl.addWidget(self.logs_label)
        main_layout.addLayout(row_log_lbl)

        row_logs = QHBoxLayout()
        row_logs.addWidget(self.reports_box)
        row_logs.addWidget(self.log_box)
        main_layout.addLayout(row_logs, 2)

    def set_long_description(self, text: str):
        clean = (text or "").strip()
        if clean:
            self.desc_label.setText(clean)
            self.desc_label.show()
        else:
            self.desc_label.hide()

    def _connect_signals(self):
        self.select_master_btn.clicked.connect(lambda: self.select_files(self.master_files_box))
        self.select_ppm_btn.clicked.connect(lambda: self.select_files(self.ppm_files_box))
        self.select_pps_btn.clicked.connect(lambda: self.select_files(self.pps_files_box))
        self.run_btn.clicked.connect(self.run_process)
        self.log_message.connect(self.append_log)
        self.report_message.connect(self.append_report)
        self.processing_done.connect(self.on_processing_done)

    def select_files(self, text_box):
        files, _ = QFileDialog.getOpenFileNames(self, "Select Files", "", "Excel/CSV Files (*.xlsx *.xlsm *.xls *.csv)")
        if files:
            text_box.setPlainText("\n".join(files))
        else:
            text_box.clear()

    def get_files_from_box(self, text_box):
        text = text_box.toPlainText().strip()
        return [line.strip() for line in text.split("\n") if line.strip()]

    def _debug_mode_enabled(self) -> bool:
        if self.debug_mode_toggle is not None:
            return bool(self.debug_mode_toggle.isChecked())
        if self.debug_mode_checkbox is not None:
            return bool(self.debug_mode_checkbox.isChecked())
        return False

    def run_process(self):
        master_files = self.get_files_from_box(self.master_files_box)
        ppm_files = self.get_files_from_box(self.ppm_files_box)
        pps_files = self.get_files_from_box(self.pps_files_box)
        debug_mode = self._debug_mode_enabled()

        if not master_files:
            MessageBox("Warning", "Please select OCCC Master file.", self).exec()
            return

        if not ppm_files and not pps_files:
             MessageBox("Warning", "Please select at least one report file (PPM or PPS).", self).exec()
             return

        self.log_box.clear()
        self.reports_box.clear()
        self.log_message.emit("Process Started...")

        self.run_btn.setEnabled(False)
        self.select_master_btn.setEnabled(False)
        self.select_ppm_btn.setEnabled(False)
        self.select_pps_btn.setEnabled(False)

        def worker():
            try:
                last_file, ok, fail = process_logic(
                    master_files,
                    ppm_files,
                    pps_files,
                    self.log_message.emit,
                    self.report_message.emit,
                    debug_mode=debug_mode,
                )
                self.processing_done.emit(ok, fail, last_file)
            except Exception as e:
                self.log_message.emit(f"CRITICAL ERROR: {e}")
                self.processing_done.emit(0, 0, "")

        threading.Thread(target=worker, daemon=True).start()

    def append_log(self, text):
        self.log_box.append(text)

    def append_report(self, text):
        self.reports_box.append(text)

    def on_processing_done(self, ok, fail, last_file):
        self.log_message.emit(f"Done. Success: {ok}, Failed: {fail}")
        if last_file:
            self.log_message.emit(f"Last processed: {last_file}")
        self.run_btn.setEnabled(True)
        self.select_master_btn.setEnabled(True)
        self.select_ppm_btn.setEnabled(True)
        self.select_pps_btn.setEnabled(True)

        title = "Process complete" if fail == 0 else "Process finished with issues"
        lines = [f"Success: {ok}", f"Failed: {fail}"]
        if last_file:
            lines.append(f"Last processed: {last_file}")
        msg = MessageBox(title, "\n".join(lines), self)
        msg.yesButton.setText("OK")
        msg.cancelButton.hide()
        msg.exec()

def get_widget():
    return MainWidget()

# - If value contains "PPS OFOB match for extended sizes" but not accompanied with "FINAL FOB (Extended sizes) doesn't match", replace the "PPS OFOB match for extended sizes" with ""
# - If value contains "PPS OFOB match for regular sizes" but not accompanied with "FINAL FOB (Regular sizes) doesn't match", replace the "PPS OFOB match for regular sizes" with ""
# - If value contains "PPS OFOB match for extended sizes" but not accompanied with "FINAL FOB (Extended sizes) doesn't match", replace the "PPS OFOB match for extended sizes" with ""
# - If value contains "PPS OFOB doesn't match for regular sizes" and accompanied with "FINAL FOB (Regular sizes) doesn't match", replace them both with a single "PPS OFOB issue for regular sizes"
# - If value contains "PPS OFOB doesn't match for extended sizes" and accompanied with "FINAL FOB (Extended sizes) doesn't match", replace them both with a single "PPS OFOB issue for extended sizes"
# - If value contains "PPS OFOB match for regular sizes" and accompanied with "FINAL FOB (Regular sizes) doesn't match", replace them both with a single "NIKE OFOB issue for regular sizes"
# - If value contains "PPS OFOB match for extended sizes" and accompanied with "FINAL FOB (Extended sizes) doesn't match", replace them both with a single "NIKE OFOB issue for extended sizes"
# - If value contains "PPS OFOB doesn't match for regular sizes" and not accompanied with "FINAL FOB (Regular sizes) doesn't match", replace it with a single "PPS OFOB issue for regular sizes"
# - If value contains "PPS OFOB doesn't match for extended sizes" and not accompanied with "FINAL FOB (Extended sizes) doesn't match", replace it with a single "PPS OFOB issue for extended sizes"

# once all that is process, do another pass for the following:
# - If value contains "PPS OFOB issue for regular sizes" and accompanied with "PPS OFOB issue for extended sizes", replace them both with a single "PPS OFOB issue for all sizes"
# - If value contains "NIKE OFOB issue for regular sizes" and accompanied with "NIKE OFOB issue for extended sizes", replace them both with a single "NIKE OFOB issue for all sizes"
