#!/usr/bin/env python3
from __future__ import annotations

import argparse
import math
import re
import threading
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# GUI Imports (for Pycro Station)
from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QVBoxLayout,
    QLabel,
    QTextEdit,
    QWidget,
    QSizePolicy,
)
from qfluentwidgets import PrimaryPushButton, MessageBox


# ===== Output headers (Row 1) =====
# NOTE: Column C is intentionally blank ("EMPTY HEADER")
OUTPUT_HEADERS: List[str] = [
    "DATE RECEIVED",
    "MAT NO",
    "",  # C) EMPTY HEADER
    "MATERIAL OETC",
    "CATEGORY",
    "SEASON",
    "STYLE",
    "SEQ",
    "SUPPLIER",
    "ITEM",
    "ITEM DESCRIPTION",
    "CW",
    "COLOR",
    "SIZE",
    "QUALITY",
    "SIZE MATRIX/COO",
    "QTY",
    "UOM",
    "REMARKS",
    "MOQ",
    "LT",
    "PRICE",
    "PURCHASE DATE",
    "PO",
    "PI DATE",
    "DEL DATE",
    "ETD",
    "ETA",
    "AWB NO",
    "STATUS",
    "CURRENT PI DATE",
    "REASON OF DELAY",
    "DELAY/NEW",
    "F/A",
    "SHOPPING LIST REVISE DATE",
    "REASON OF REVISE 1 - BUYER 2 - MER 3 - SUPPLIER 4 - MAT 5 - PROD 6 - OTHERS (MENTION THE ACTUAL REASON)",
    "SUPPLIER RESPOND (DAY)",
    "RECEIVED PI/CFM DEL DATE FROM SUPPLIER",
    "(0) - MEET OETC (1) - No greige/No Prebook (2) - Fab/Acc Matching Issue (3) - LT and MOQ issue (4) - Added by Buyer (5) - New Develop (6) - Production Schedule (7) - Late approve by buyer (8) - Quality Issue (9) - Added by MER (10) - Supplier missing order /Send out (11) - Internal Issue (12) - ShipDoc Issue",
]

# ===== Normalization helpers =====
_WS_RE = re.compile(r"\s+")
_PREFIX_RE = re.compile(r"^[A-Z]{1,2}\)\s*", flags=re.IGNORECASE)

def _is_nan(x: object) -> bool:
    return isinstance(x, float) and math.isnan(x)

def strip_prefix(s: str) -> str:
    return _PREFIX_RE.sub("", s)

def norm_cell(x: object) -> str:
    """
    Normalize header strings to handle newlines/double spaces/human errors:
    - convert to str
    - replace newlines with spaces
    - collapse whitespace
    - strip
    - remove leading 'A) ' / 'AA) ' style prefix
    - uppercase
    """
    if x is None or _is_nan(x):
        return ""
    s = str(x).replace("\r", " ").replace("\n", " ")
    s = _WS_RE.sub(" ", s).strip()
    s = strip_prefix(s)
    return s.upper().strip()

def pick_engine(path: Path) -> Optional[str]:
    """
    Optional explicit engines to avoid 'format cannot be determined' issues.
    pandas will usually auto-pick correctly, but this is safer.
    """
    ext = path.suffix.lower()
    if ext in {".ods", ".odf", ".odt"}:
        return "odf"
    if ext == ".xls":
        return "xlrd"
    if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return "openpyxl"
    return None


# ===== Core logic =====
def _emit(log_emit, text: str):
    if callable(log_emit):
        try:
            log_emit(text)
            return
        except Exception:
            pass
    print(text)

def iter_input_files(items: Iterable[str], recursive: bool) -> List[Path]:
    exts = {".xlsx", ".xls", ".ods"}
    out: List[Path] = []
    for it in items:
        p = Path(it)
        if p.is_dir():
            pattern = "**/*" if recursive else "*"
            for f in p.glob(pattern):
                if f.is_file() and f.suffix.lower() in exts:
                    out.append(f)
        else:
            if p.is_file() and p.suffix.lower() in exts:
                out.append(p)
    # de-dup while preserving order
    seen = set()
    uniq = []
    for f in out:
        if f not in seen:
            uniq.append(f)
            seen.add(f)
    return uniq

def build_col_map(header_row: pd.Series) -> Dict[str, int]:
    """
    Map normalized header -> column index (first occurrence wins).
    """
    col_map: Dict[str, int] = {}
    for idx, val in enumerate(header_row.tolist()):
        key = norm_cell(val)
        if key and key not in col_map:
            col_map[key] = idx
    return col_map

def find_best_header_row(df: pd.DataFrame) -> Optional[int]:
    """
    1) find all rows containing a cell that normalizes to 'DATE RECEIVED'
    2) among those candidates, choose the row with the highest number of matches
       against our desired headers (A..AM)
    """
    target = "DATE RECEIVED"
    desired_keys = [norm_cell(h) for h in OUTPUT_HEADERS if h]  # skip blank C

    candidates: List[int] = []
    for r in range(df.shape[0]):
        row = df.iloc[r]
        if any(norm_cell(v) == target for v in row.tolist()):
            candidates.append(r)

    if not candidates:
        return None

    best_row = None
    best_score = -1
    for r in candidates:
        col_map = build_col_map(df.iloc[r])
        score = sum(1 for k in desired_keys if k in col_map)
        if score > best_score:
            best_score = score
            best_row = r

    return best_row

def append_file(ws, path: Path, log_emit=None) -> Optional[int]:
    engine = pick_engine(path)
    df = pd.read_excel(path, sheet_name=0, header=None, dtype=object, engine=engine)

    header_r = find_best_header_row(df)
    if header_r is None:
        _emit(log_emit, f"[SKIP] No 'DATE RECEIVED' header found in: {path}")
        return None

    col_map = build_col_map(df.iloc[header_r])

    # For each output column, find matching source column (or None)
    desired_src_cols: List[Optional[int]] = []
    for h in OUTPUT_HEADERS:
        if not h:  # blank C
            desired_src_cols.append(None)
            continue
        key = norm_cell(h)
        desired_src_cols.append(col_map.get(key))

    appended = 0
    for r in range(header_r + 1, df.shape[0]):
        out_row: List[object] = []
        any_val = False

        for src in desired_src_cols:
            if src is None:
                out_row.append(None)
                continue

            val = df.iat[r, src] if src < df.shape[1] else None
            if pd.isna(val):
                val = None

            if val not in (None, ""):
                any_val = True
            out_row.append(val)

        if any_val:
            ws.append(out_row)
            appended += 1

    _emit(log_emit, f"[OK] {path.name}: appended {appended} row(s) (header row = {header_r + 1})")
    return appended


def proposed_output_path(files: List[Path]) -> Path:
    """Generate output path in first input folder with timestamp suffix."""
    base_dir = files[0].parent if files else Path.cwd()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = f"MonitoringChart-Consolidated-{timestamp}"
    candidate = base_dir / f"{stem}.xlsx"
    if not candidate.exists():
        return candidate
    n = 1
    while True:
        candidate = base_dir / f"{stem} ({n}).xlsx"
        if not candidate.exists():
            return candidate
        n += 1


def process_files(
    items: List[str],
    recursive: bool,
    log_emit=None,
    output: Optional[Path] = None,
) -> Tuple[str, int, int]:
    files = iter_input_files(items, recursive=recursive)
    if not files:
        raise ValueError("No input files found (.xlsx/.xls/.ods).")

    out_path = output if output else proposed_output_path(files)

    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated"

    ws.append(OUTPUT_HEADERS)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(OUTPUT_HEADERS))}1"

    total_rows = 0
    ok_files = 0
    fail_files = 0
    for f in files:
        res = append_file(ws, f, log_emit=log_emit)
        if res is None:
            fail_files += 1
        else:
            ok_files += 1
            total_rows += res

    # Sort consolidated rows by column B (index 1) ascending
    data_count = ws.max_row - 1
    if data_count > 1:
        rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))

        def sort_key(row):
            val = row[1] if len(row) > 1 else None
            if val is None:
                return (2, "")
            if isinstance(val, (int, float)):
                return (0, float(val))
            s = str(val).strip()
            if not s:
                return (2, "")
            try:
                return (0, float(s))
            except Exception:
                return (1, s.upper())

        rows.sort(key=sort_key)
        ws.delete_rows(2, data_count)
        for r in rows:
            ws.append(list(r))

    # Auto-fit column widths, capped at 5.00 inches (~68 Excel width units)
    max_width_chars = int(round((5.0 * 96 - 5) / 7))  # Excel approx conversion
    dims: Dict[int, int] = {}
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    for r in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for c_idx, cell in enumerate(r, start=1):
            v = cell.value
            if v is None:
                l = 0
            else:
                s = str(v)
                l = len(s)
            if l:
                dims[c_idx] = max(dims.get(c_idx, 0), l)

    for c_idx, w in dims.items():
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(w + 2, 8), max_width_chars)

    # Fixed row height at 0.20 inches for all rows
    row_height_points = 0.20 * 72  # Excel uses points
    for row_idx in range(1, (ws.max_row or 0) + 1):
        ws.row_dimensions[row_idx].height = row_height_points

    wb.save(out_path)
    _emit(log_emit, f"\nDone. Wrote {total_rows} total row(s) to: {out_path}")
    return str(out_path), ok_files, fail_files


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Consolidate rows from multiple xlsx/xls/ods into one template with fixed headers."
    )
    ap.add_argument(
        "paths",
        nargs="+",
        help="Input files/folders. Legacy: you may provide output path first.",
    )
    ap.add_argument(
        "--output",
        "-o",
        help="Output .xlsx file path. If omitted, auto-generated beside first input.",
    )
    ap.add_argument("--recursive", action="store_true", help="Recurse into folders")
    args = ap.parse_args()

    output_path: Optional[Path] = Path(args.output) if args.output else None
    inputs = list(args.paths)
    if output_path is None and len(inputs) >= 2:
        first = Path(inputs[0])
        if (not first.exists()) and first.suffix.lower() in {".xlsx", ".xlsm"}:
            output_path = first
            inputs = inputs[1:]

    out_str, ok, fail = process_files(
        [str(p) for p in inputs],
        recursive=args.recursive,
        output=output_path,
    )
    print(f"\nDone. Success: {ok}, Failed: {fail}")
    print(f"Wrote consolidated file to: {out_str}")


if __name__ == "__main__":
    main()


# --- UI Class (Pycro Station) ---

class MainWidget(QWidget):
    log_message = Signal(str)
    processing_done = Signal(int, int, str)

    def __init__(self):
        super().__init__()
        self.setObjectName("mc_append_widget")
        self._build_ui()
        self._connect_signals()

    def _build_ui(self):
        self.desc_label = QLabel("", self)
        self.desc_label.setWordWrap(True)
        self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.desc_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.desc_label.setStyleSheet(
            "color: #dcdcdc; background: transparent; padding: 6px; "
            "border: 1px solid #3a3a3a; border-radius: 6px;"
        )
        self.set_long_description("")

        self.select_btn = PrimaryPushButton("Select Monitoring Chart Files", self)
        self.run_btn = PrimaryPushButton("Consolidate", self)

        self.files_label = QLabel("Selected files", self)
        self.files_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.files_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        self.logs_label = QLabel("Process logs", self)
        self.logs_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.logs_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        shared_style = (
            "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        self.files_box = QTextEdit(self)
        self.files_box.setReadOnly(True)
        self.files_box.setPlaceholderText("Selected Excel files will appear here")
        self.files_box.setStyleSheet(shared_style)

        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText("Live process log will appear here")
        self.log_box.setStyleSheet(shared_style)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)

        main_layout.addWidget(self.desc_label, 1)

        row1 = QHBoxLayout()
        row1.addStretch(1)
        row1.addWidget(self.select_btn, 1)
        row1.addStretch(1)
        main_layout.addLayout(row1, 0)

        row2 = QHBoxLayout()
        row2.addStretch(1)
        row2.addWidget(self.run_btn, 1)
        row2.addStretch(1)
        main_layout.addLayout(row2, 0)

        row3 = QHBoxLayout()
        row3.addWidget(self.files_label, 1)
        row3.addWidget(self.logs_label, 1)
        main_layout.addLayout(row3, 0)

        row4 = QHBoxLayout()
        row4.addWidget(self.files_box, 1)
        row4.addWidget(self.log_box, 1)
        main_layout.addLayout(row4, 4)

    def set_long_description(self, text: str):
        clean = (text or "").strip()
        if clean:
            self.desc_label.setText(clean)
            self.desc_label.show()
        else:
            self.desc_label.clear()
            self.desc_label.hide()

    def _connect_signals(self):
        self.select_btn.clicked.connect(self.select_files)
        self.run_btn.clicked.connect(self.run_process)
        self.log_message.connect(self.append_log)
        self.processing_done.connect(self.on_done)

    def select_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select Monitoring Chart Files",
            "",
            "Excel Files (*.xlsx *.xlsm *.xls *.ods)",
        )
        if files:
            self.files_box.setPlainText("\n".join(files))
        else:
            self.files_box.clear()

    def _selected_files(self) -> List[str]:
        text = self.files_box.toPlainText().strip()
        if not text:
            return []
        return [line.strip() for line in text.split("\n") if line.strip()]

    def run_process(self):
        inputs = self._selected_files()
        if not inputs:
            MessageBox("Warning", "Please select Excel files to consolidate.", self).exec()
            return

        self.log_box.clear()
        self.log_message.emit("Consolidation process started...")
        self.log_message.emit(f"Files to process: {len(inputs)}")
        self.log_message.emit("Output will be saved beside first input with timestamp suffix.")
        self.log_message.emit("")

        self.run_btn.setEnabled(False)
        self.select_btn.setEnabled(False)

        def worker():
            try:
                out_path, ok, fail = process_files(inputs, recursive=False, log_emit=self.log_message.emit)
                self.processing_done.emit(ok, fail, out_path)
            except Exception as e:
                self.log_message.emit(f"CRITICAL ERROR: {e}")
                self.processing_done.emit(0, 0, "")

        threading.Thread(target=worker, daemon=True).start()

    def append_log(self, text: str):
        self.log_box.append(text)
        self.log_box.ensureCursorVisible()

    def on_done(self, ok: int, fail: int, out_path: str):
        self.log_message.emit("")
        self.log_message.emit(f"Consolidation complete: {ok} succeeded, {fail} failed")
        if out_path:
            self.log_message.emit(f"Output: {out_path}")

        self.run_btn.setEnabled(True)
        self.select_btn.setEnabled(True)

        title = "Consolidation complete" if fail == 0 else "Consolidation finished with issues"
        lines = [f"Success: {ok}", f"Failed: {fail}"]
        if out_path:
            lines.append(f"File: {Path(out_path).name}")
        msg = MessageBox(title, "\n".join(lines), self)
        msg.yesButton.setText("OK")
        msg.cancelButton.hide()
        msg.exec()


def get_widget():
    return MainWidget()
