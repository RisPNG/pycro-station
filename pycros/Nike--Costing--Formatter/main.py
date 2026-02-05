import os
import threading
from copy import copy
from datetime import datetime
from typing import Callable, List, Optional, Tuple

from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QSizePolicy,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)
from qfluentwidgets import MessageBox, PrimaryPushButton

from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


GRAY_HEX = "FFB2B2B2"
GRAY_FILL = PatternFill(fill_type="solid", start_color=GRAY_HEX, end_color=GRAY_HEX)

BLACK_HEX = "FF000000"
TOP_THIN = Side(style="thin", color=BLACK_HEX)
BOTTOM_DOUBLE = Side(style="double", color=BLACK_HEX)

HEADER_MARKER = "embellishment cost"
GAIN_VALUES = {"gain", "gain/pc"}

AXB_MARKER = "c=axb"
GAIN_AXB_VALUES = {"gain :", "gain:"}

SUB_TOTAL_NEEDLE = "sub total"
SUB_TOTAL_REPLACEMENT = "SUB TOTAL FOR NIKE (A & B)"

COL_I = 9
COL_B = 2
COL_AE = 31
COL_P = 16
COL_U = 21


class MainWidget(QWidget):
    log_message = Signal(str)
    processing_done = Signal(int, int)

    def __init__(self):
        super().__init__()
        self.setObjectName("nike_costing_formatter_widget")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(12)

        self.desc_label = QLabel("", self)
        self.desc_label.setWordWrap(True)
        self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.desc_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.desc_label.setStyleSheet(
            "color: #dcdcdc; background: transparent; padding: 6px; "
            "border: 1px solid #3a3a3a; border-radius: 6px;"
        )
        self.set_long_description("")

        self.select_btn = PrimaryPushButton("Select Excel Files", self)
        self.run_btn = PrimaryPushButton("Run", self)

        row_btns = QHBoxLayout()
        row_btns.addWidget(self.select_btn, 1)
        row_btns.addWidget(self.run_btn, 1)

        self.files_box = QTextEdit(self)
        self.files_box.setReadOnly(True)
        self.files_box.setPlaceholderText("Selected files will appear here")
        self.files_box.setStyleSheet(
            "QTextEdit{background: #2a2a2a; color: white; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText("Live process log will appear here")
        self.log_box.setStyleSheet(
            "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        layout.addWidget(self.desc_label, 0)
        layout.addLayout(row_btns, 0)
        layout.addWidget(self.files_box, 1)
        layout.addWidget(self.log_box, 2)

        self.select_btn.clicked.connect(self.select_files)
        self.run_btn.clicked.connect(self.run_process)
        self.log_message.connect(self.append_log)
        self.processing_done.connect(self.on_processing_done)

    def set_long_description(self, text: str):
        clean = (text or "").strip()
        if clean:
            self.desc_label.setText(clean)
            self.desc_label.show()
        else:
            self.desc_label.clear()
            self.desc_label.hide()

    def select_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select Excel files",
            "",
            "Excel Workbook (*.xlsx)",
        )
        if files:
            self.files_box.setPlainText("\n".join(files))
        else:
            self.files_box.clear()

    def _selected_files(self) -> List[str]:
        text = self.files_box.toPlainText().strip()
        if not text:
            return []
        return [line for line in text.split("\n") if line.strip()]

    def run_process(self):
        files = self._selected_files()
        if not files:
            MessageBox("No files", "Please select one or more .xlsx files first.", self).exec()
            return

        self.log_box.clear()
        self.log_message.emit(f"Starting processing of {len(files)} file(s)...")
        self.run_btn.setEnabled(False)
        self.select_btn.setEnabled(False)

        def worker():
            ok, fail = 0, 0
            for path in files:
                try:
                    self.log_message.emit(f"Opening: {path}")
                    success, out_path = process_file(path, self.log_message.emit)
                    if success:
                        self.log_message.emit(f"Saved: {out_path}")
                        ok += 1
                    else:
                        fail += 1
                except Exception as e:
                    self.log_message.emit(f"ERROR processing {path}: {e}")
                    fail += 1
            self.processing_done.emit(ok, fail)

        threading.Thread(target=worker, daemon=True).start()

    def append_log(self, text: str):
        self.log_box.append(text)
        self.log_box.ensureCursorVisible()

    def on_processing_done(self, ok: int, fail: int):
        self.log_message.emit(f"Completed: {ok} success, {fail} failed.")
        self.run_btn.setEnabled(True)
        self.select_btn.setEnabled(True)
        title = "Processing complete" if fail == 0 else "Processing finished with issues"
        lines = [f"Success: {ok}", f"Failed: {fail}", "", "Outputs are saved next to the input file(s)."]
        msg = MessageBox(title, "\n".join(lines), self)
        msg.yesButton.setText("OK")
        msg.cancelButton.hide()
        msg.exec()


def get_widget():
    return MainWidget()


def process_file(path: str, log: Optional[Callable[[str], None]] = None) -> Tuple[bool, str]:
    if log is None:
        log = lambda _msg: None

    if not path.lower().endswith(".xlsx"):
        log("Skipped (not .xlsx).")
        return False, ""

    wb = load_workbook(path)
    for ws in wb.worksheets:
        _process_sheet(ws, log)

    out_path = _build_output_path(path)
    wb.save(out_path)
    return True, out_path


def _process_sheet(ws: Worksheet, log: Callable[[str], None]) -> None:
    log(f"  Sheet: {ws.title}")

    header_row = _find_first_row_by_text(ws, HEADER_MARKER)
    if header_row is not None:
        _set_freeze_panes(ws, row=header_row + 1, col=COL_B)
        end_row = _last_data_row(ws, max_col=COL_AE)
        if end_row < 1:
            end_row = 1
        _bold_row_only(ws, row=header_row - 1, end_col=COL_AE)
        _format_header_row(ws, header_row)
        _format_gain_cells(ws, header_row, end_row, end_col=COL_AE)
        _shade_columns_p_to_u(ws, header_row, end_row)
        _format_sub_totals(ws)
        _autofit_columns(ws, start_col=COL_B, end_col=COL_AE, max_row=end_row)
        return

    axb_cell = _find_first_cell_by_text(ws, AXB_MARKER)
    if axb_cell is None:
        _set_freeze_panes(ws, row=2, col=COL_B)
        log('    - Neither "Embellishment Cost" nor "c=axb" found; skipping sheet.')
        return

    axb_row, axb_col = axb_cell
    _set_freeze_panes(ws, row=axb_row + 1, col=COL_B)
    end_row, end_col = _used_extent(ws)
    format_end_col = max(end_col, axb_col + 11, 1)

    _bold_row_only(ws, row=axb_row - 4, end_col=format_end_col)
    _format_rows_gray_bold(ws, start_row=max(1, axb_row - 3), end_row=axb_row, end_col=format_end_col)
    _format_gain_cells_axb(ws, header_row=axb_row, end_row=end_row, gain_col=axb_col - 3, end_col=format_end_col)
    _format_totals(ws, end_col=format_end_col)
    _shade_relative_columns_axb(ws, header_row=axb_row, end_row=end_row, axb_col=axb_col)
    _autofit_columns(ws, start_col=COL_B, end_col=end_col, max_row=end_row)


def _norm(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_first_row_by_text(ws: Worksheet, needle_lower: str) -> Optional[int]:
    found_row: Optional[int] = None
    found_col: Optional[int] = None
    for cell in _iter_value_cells(ws):
        if _norm(cell.value).lower() != needle_lower:
            continue
        if found_row is None or (cell.row < found_row) or (cell.row == found_row and cell.column < (found_col or 0)):
            found_row = cell.row
            found_col = cell.column
    return found_row


def _find_first_cell_by_text(ws: Worksheet, needle_lower: str) -> Optional[Tuple[int, int]]:
    found: Optional[Tuple[int, int]] = None
    for cell in _iter_value_cells(ws):
        if _norm(cell.value).lower() != needle_lower:
            continue
        candidate = (cell.row, cell.column)
        if found is None or candidate < found:
            found = candidate
    return found


def _used_extent(ws: Worksheet) -> Tuple[int, int]:
    max_row, max_col = 1, 1
    any_value = False
    for cell in _iter_value_cells(ws):
        any_value = True
        if cell.row > max_row:
            max_row = cell.row
        if cell.column > max_col:
            max_col = cell.column
    if not any_value:
        return 1, 1
    return max_row, max_col


def _last_data_row(ws: Worksheet, max_col: int) -> int:
    last = 0
    for cell in _iter_value_cells(ws):
        if cell.column > max_col:
            continue
        if cell.row > last:
            last = cell.row
    return last or 1


def _iter_value_cells(ws: Worksheet):
    # Use worksheet's internal cell store to avoid creating huge numbers of blank cells.
    cells = getattr(ws, "_cells", None)
    if not isinstance(cells, dict):
        return iter(())
    return (
        cell
        for cell in cells.values()
        if cell is not None
        and cell.value is not None
        and not (isinstance(cell.value, str) and cell.value.strip() == "")
    )


def _make_bold(cell) -> None:
    try:
        font = copy(cell.font) if cell.font is not None else Font()
        font.bold = True
        cell.font = font
    except Exception:
        pass


def _apply_gray_fill(cell) -> None:
    try:
        cell.fill = GRAY_FILL
    except Exception:
        pass


def _set_top_thin_double_bottom_border(cell) -> None:
    try:
        border = cell.border
        kwargs = {
            "left": border.left,
            "right": border.right,
            "top": TOP_THIN,
            "bottom": BOTTOM_DOUBLE,
            "diagonal": border.diagonal,
            "diagonal_direction": border.diagonal_direction,
            "outline": border.outline,
            "vertical": border.vertical,
            "horizontal": border.horizontal,
        }
        if hasattr(border, "start"):
            kwargs["start"] = border.start
        if hasattr(border, "end"):
            kwargs["end"] = border.end
        cell.border = Border(**kwargs)
    except Exception:
        pass


def _format_header_row(ws: Worksheet, header_row: int) -> None:
    for col in range(1, COL_AE + 1):
        cell = ws.cell(row=header_row, column=col)
        _make_bold(cell)
        _apply_gray_fill(cell)


def _format_gain_cells(ws: Worksheet, header_row: int, end_row: int, end_col: int) -> None:
    # Only touch existing cells in column I with matching values.
    cells = getattr(ws, "_cells", {}) if isinstance(getattr(ws, "_cells", None), dict) else {}
    gain_rows: set[int] = set()
    for (r, c), cell in cells.items():
        if c != COL_I:
            continue
        if r <= header_row or r > end_row:
            continue
        v = _norm(cell.value).lower()
        if v in GAIN_VALUES:
            gain_rows.add(r)

    for r in sorted(gain_rows):
        _format_gain_row(ws, row=r, end_col=end_col)


def _autofit_columns(ws: Worksheet, start_col: int, end_col: int, max_row: int) -> None:
    dims = {c: 0 for c in range(start_col, end_col + 1)}
    for cell in _iter_value_cells(ws):
        if cell.row > max_row:
            continue
        if cell.column < start_col or cell.column > end_col:
            continue
        s = str(cell.value)
        dims[cell.column] = max(dims[cell.column], len(s))

    for col_idx, max_len in dims.items():
        if max_len <= 0:
            continue
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 8), 60)


def _format_sub_totals(ws: Worksheet) -> None:
    subtotal_rows = []
    for cell in _iter_value_cells(ws):
        if _norm(cell.value).lower() == SUB_TOTAL_NEEDLE:
            subtotal_rows.append(cell.row)
            try:
                cell.value = SUB_TOTAL_REPLACEMENT
            except Exception:
                pass

    for base_row in subtotal_rows:
        for r in range(base_row, base_row + 3):
            for c in range(1, COL_AE + 1):
                cell = ws.cell(row=r, column=c)
                _make_bold(cell)
                _apply_gray_fill(cell)


def _format_rows_gray_bold(ws: Worksheet, start_row: int, end_row: int, end_col: int) -> None:
    start_row = max(1, int(start_row))
    end_row = max(1, int(end_row))
    end_col = max(1, int(end_col))
    for r in range(start_row, end_row + 1):
        for c in range(1, end_col + 1):
            cell = ws.cell(row=r, column=c)
            _make_bold(cell)
            _apply_gray_fill(cell)


def _format_gain_cells_axb(ws: Worksheet, header_row: int, end_row: int, gain_col: int, end_col: int) -> None:
    if gain_col < 1:
        return
    cells = getattr(ws, "_cells", {}) if isinstance(getattr(ws, "_cells", None), dict) else {}
    gain_rows: set[int] = set()
    for (r, c), cell in cells.items():
        if c != gain_col:
            continue
        if r <= header_row or r > end_row:
            continue
        v = _norm(cell.value).lower()
        if v in GAIN_AXB_VALUES:
            gain_rows.add(r)

    for r in sorted(gain_rows):
        _format_gain_row(ws, row=r, end_col=end_col)


def _format_totals(ws: Worksheet, end_col: int) -> None:
    total_rows = []
    for cell in _iter_value_cells(ws):
        if _norm(cell.value).lower() == "total":
            total_rows.append(cell.row)

    for base_row in total_rows:
        for r in range(base_row, base_row + 3):
            for c in range(1, end_col + 1):
                cell = ws.cell(row=r, column=c)
                _make_bold(cell)
                _apply_gray_fill(cell)


def _shade_relative_columns_axb(ws: Worksheet, header_row: int, end_row: int, axb_col: int) -> None:
    start_col = axb_col + 6
    end_col = axb_col + 11
    if start_col < 1 or end_col < 1:
        return
    # For AXB sheets, the "header block" includes header_row and 3 rows above it (gray),
    # plus a 4th row above that which is bold-only (no gray). Shade that bold-only row too.
    start_row = max(1, header_row - 4)
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            _apply_gray_fill(ws.cell(row=r, column=c))


def _bold_row_only(ws: Worksheet, row: int, end_col: int) -> None:
    if row < 1:
        return
    end_col = max(1, int(end_col))
    for c in range(1, end_col + 1):
        _make_bold(ws.cell(row=row, column=c))


def _format_gain_row(ws: Worksheet, row: int, end_col: int) -> None:
    if row < 1:
        return
    end_col = max(1, int(end_col))
    for c in range(1, end_col + 1):
        cell = ws.cell(row=row, column=c)
        _make_bold(cell)
        _set_top_thin_double_bottom_border(cell)


def _set_freeze_panes(ws: Worksheet, row: int, col: int) -> None:
    try:
        row = int(row)
        col = int(col)
        if row < 1:
            row = 1
        if col < 1:
            col = 1
        ws.freeze_panes = ws.cell(row=row, column=col).coordinate
    except Exception:
        pass


def _shade_columns_p_to_u(ws: Worksheet, header_row: int, end_row: int) -> None:
    start_row = max(1, header_row - 1)
    for r in range(start_row, end_row + 1):
        for c in range(COL_P, COL_U + 1):
            _apply_gray_fill(ws.cell(row=r, column=c))


def _build_output_path(src_path: str) -> str:
    base_dir = os.path.dirname(src_path) or os.getcwd()
    base_name = os.path.basename(src_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_name = f"{timestamp}_nike_formatted_{base_name}"
    candidate = os.path.join(base_dir, out_name)
    if not os.path.exists(candidate):
        return candidate

    name, ext = os.path.splitext(out_name)
    counter = 1
    while True:
        candidate = os.path.join(base_dir, f"{name} ({counter}){ext}")
        if not os.path.exists(candidate):
            return candidate
        counter += 1
