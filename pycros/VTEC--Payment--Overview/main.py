from __future__ import annotations

import os
import re
import threading
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Callable, Iterable, List, Optional, Sequence, Tuple

from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QSizePolicy,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)
from qfluentwidgets import MessageBox, PrimaryPushButton

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

try:
    import xlrd  # type: ignore
except Exception:  # pragma: no cover - Pycro Station installs requirements before launch
    xlrd = None


OVERVIEW_SHEET_NAME = "VTEC Payment Overview"
DUPLICATES_SHEET_NAME = "VTEC Payment Duplicates"
LOG_SHEET_NAME = "Processing Log"
REJECTED_LOG_SHEET_NAME = "Rejected Log"

DATA_START_ROW = 4
HEADER_ROW = 3
NUM_OUTPUT_COLS = 19
CONDITIONAL_USD_MAP_KEY = 0

HEADERS: Tuple[str, ...] = (
    "Supplier Name",
    "Payment Term",
    "VAT Invoice Number",
    "Other References",
    "PO Amt. Bfr VAT (VND)",
    "Surcharge / Other (VND)",
    "Total (VND)",
    "PO Amt. Bfr VAT (USD)",
    "Surcharge / Other (USD)",
    "Total (USD)",
    "VAT Invoice Date",
    "Currency Rate",
    "Purchaser",
    "Order Type",
    "From",
    "To",
    "SIG Due Date",
    "Payment To VTEC (LSKhor)",
    "Payment To Supplier (VTEC)",
)

DATE_COLUMN_INDICES = {11, 17, 18, 19}


@dataclass
class ProcessingResult:
    output_path: str = ""
    imported_rows: int = 0
    duplicate_rows_moved: int = 0
    processed_sheets: int = 0
    skipped_sheets: int = 0
    failed_files: int = 0
    success: bool = True
    message: str = ""


class MainWidget(QWidget):
    log_message = Signal(str)
    processing_done = Signal(object)

    def __init__(self):
        super().__init__()
        self.setObjectName("vtec_payment_overview_widget")
        self._build_ui()
        self._connect_signals()

    def _build_ui(self):
        self.desc_label = QLabel("", self)
        self.desc_label.setWordWrap(True)
        self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.desc_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.desc_label.setStyleSheet(
            "color: #dcdcdc; background: transparent; padding: 6px; "
            "border: 1px solid #3a3a3a; border-radius: 6px;"
        )
        self.set_long_description("")

        self.select_overview_btn = PrimaryPushButton("Select Existing Overview Workbook (Optional)", self)
        self.select_output_btn = PrimaryPushButton("Select New Output Folder (Optional)", self)
        self.select_files_btn = PrimaryPushButton("Select Payment Excel Files", self)
        self.run_btn = PrimaryPushButton("Run", self)

        self.overview_label = QLabel("Selected overview workbook (optional)", self)
        self.output_label = QLabel("Selected new output folder (optional)", self)
        self.files_label = QLabel("Selected payment files", self)
        self.logs_label = QLabel("Process logs", self)
        for label in (self.overview_label, self.output_label, self.files_label, self.logs_label):
            label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        self.overview_box = QTextEdit(self)
        self.overview_box.setReadOnly(True)
        self.overview_box.setPlaceholderText("Optional. Leave blank to create a new timestamped overview workbook.")
        self.overview_box.setFixedHeight(48)

        self.output_box = QTextEdit(self)
        self.output_box.setReadOnly(True)
        self.output_box.setPlaceholderText("Optional. Used only when no existing overview workbook is selected.")
        self.output_box.setFixedHeight(48)

        self.files_box = QTextEdit(self)
        self.files_box.setReadOnly(True)
        self.files_box.setPlaceholderText("Selected payment files will appear here")

        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText("Live process log will appear here")

        text_box_style = (
            "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )
        self.overview_box.setStyleSheet(text_box_style)
        self.output_box.setStyleSheet(text_box_style)
        self.files_box.setStyleSheet(text_box_style)
        self.log_box.setStyleSheet(text_box_style)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)

        main_layout.addWidget(self.desc_label, 1)

        overview_btn_row = QHBoxLayout()
        overview_btn_row.addStretch(1)
        overview_btn_row.addWidget(self.select_overview_btn, 1)
        overview_btn_row.addStretch(1)
        main_layout.addLayout(overview_btn_row, 0)

        main_layout.addWidget(self.overview_label, 0)
        main_layout.addWidget(self.overview_box, 0)

        output_btn_row = QHBoxLayout()
        output_btn_row.addStretch(1)
        output_btn_row.addWidget(self.select_output_btn, 1)
        output_btn_row.addStretch(1)
        main_layout.addLayout(output_btn_row, 0)

        main_layout.addWidget(self.output_label, 0)
        main_layout.addWidget(self.output_box, 0)

        source_btn_row = QHBoxLayout()
        source_btn_row.addStretch(1)
        source_btn_row.addWidget(self.select_files_btn, 1)
        source_btn_row.addStretch(1)
        main_layout.addLayout(source_btn_row, 0)

        run_btn_row = QHBoxLayout()
        run_btn_row.addStretch(1)
        run_btn_row.addWidget(self.run_btn, 1)
        run_btn_row.addStretch(1)
        main_layout.addLayout(run_btn_row, 0)

        bottom_labels = QHBoxLayout()
        bottom_labels.addWidget(self.files_label, 1)
        bottom_labels.addWidget(self.logs_label, 1)
        main_layout.addLayout(bottom_labels, 0)

        bottom_boxes = QHBoxLayout()
        bottom_boxes.addWidget(self.files_box, 1)
        bottom_boxes.addWidget(self.log_box, 1)
        main_layout.addLayout(bottom_boxes, 4)

    def set_long_description(self, text: str):
        clean = (text or "").strip()
        if clean:
            self.desc_label.setText(clean)
            self.desc_label.show()
        else:
            self.desc_label.clear()
            self.desc_label.hide()

    def _connect_signals(self):
        self.select_overview_btn.clicked.connect(self.select_overview_workbook)
        self.select_output_btn.clicked.connect(self.select_output_folder)
        self.select_files_btn.clicked.connect(self.select_payment_files)
        self.run_btn.clicked.connect(self.run_process)
        self.log_message.connect(self.append_log)
        self.processing_done.connect(self.on_processing_done)

    def select_overview_workbook(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Existing VTEC Payment Overview Workbook",
            "",
            "Excel Workbooks (*.xlsx *.xlsm)",
        )
        if file_path:
            self.overview_box.setPlainText(file_path)
        else:
            self.overview_box.clear()

    def select_output_folder(self):
        folder_path = QFileDialog.getExistingDirectory(
            self,
            "Select New VTEC Payment Overview Output Folder",
            "",
        )
        if folder_path:
            self.output_box.setPlainText(folder_path)
        else:
            self.output_box.clear()

    def select_payment_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select VTEC Payment Excel Files",
            "",
            "Excel Files (*.xlsx *.xlsm *.xls)",
        )
        if files:
            self.files_box.setPlainText("\n".join(files))
        else:
            self.files_box.clear()

    def _selected_overview_workbook(self) -> Optional[str]:
        text = self.overview_box.toPlainText().strip()
        if not text:
            return None
        first_line = next((line.strip() for line in text.splitlines() if line.strip()), "")
        return first_line or None

    def _selected_output_folder(self) -> Optional[str]:
        text = self.output_box.toPlainText().strip()
        if not text:
            return None
        first_line = next((line.strip() for line in text.splitlines() if line.strip()), "")
        return first_line or None

    def _selected_files(self) -> List[str]:
        text = self.files_box.toPlainText().strip()
        if not text:
            return []
        return [line.strip() for line in text.splitlines() if line.strip()]

    def run_process(self):
        source_files = self._selected_files()
        overview_workbook = self._selected_overview_workbook()
        output_folder = self._selected_output_folder()
        if not source_files:
            MessageBox("Warning", "Nothing to process. Please select payment file(s).", self).exec()
            return

        self.log_box.clear()
        self.log_message.emit("Process starts")
        self._set_buttons_enabled(False)

        def worker():
            try:
                result = process_payment_files(
                    source_files,
                    self.log_message.emit,
                    overview_workbook,
                    output_folder,
                )
            except Exception as exc:
                result = ProcessingResult(success=False, failed_files=len(source_files), message=str(exc))
                self.log_message.emit(f"ERROR: {exc}")
            self.processing_done.emit(result)

        threading.Thread(target=worker, daemon=True).start()

    def _set_buttons_enabled(self, enabled: bool):
        self.select_overview_btn.setEnabled(enabled)
        self.select_output_btn.setEnabled(enabled)
        self.select_files_btn.setEnabled(enabled)
        self.run_btn.setEnabled(enabled)

    def append_log(self, text: str):
        self.log_box.append(text)
        self.log_box.ensureCursorVisible()

    def on_processing_done(self, result: ProcessingResult):
        self._set_buttons_enabled(True)
        title = "Processing complete" if result.success and result.failed_files == 0 else "Processing finished with issues"

        if result.output_path:
            self.log_message.emit(f"Output workbook saved to: {result.output_path}")
        self.log_message.emit(
            "Completed: "
            f"{result.imported_rows} rows imported, "
            f"{result.duplicate_rows_moved} duplicates moved, "
            f"{result.processed_sheets} sheets processed, "
            f"{result.skipped_sheets} sheets skipped, "
            f"{result.failed_files} files failed."
        )

        lines = [
            f"Imported rows: {result.imported_rows}",
            f"Duplicate rows moved: {result.duplicate_rows_moved}",
            f"Processed sheets: {result.processed_sheets}",
            f"Skipped sheets: {result.skipped_sheets}",
            f"Failed files: {result.failed_files}",
        ]
        if result.output_path:
            lines.append(f"Output: {os.path.basename(result.output_path)}")
        if result.message:
            lines.append("")
            lines.append(result.message)

        msg = MessageBox(title, "\n".join(lines), self)
        msg.yesButton.setText("OK")
        msg.cancelButton.hide()
        msg.exec()


def get_widget():
    return MainWidget()


# =====================================================================================
# Pycro main process
# =====================================================================================


def process_payment_files(
    source_paths: Sequence[str],
    log_emit: Optional[Callable[[str], None]] = None,
    overview_workbook_path: Optional[str] = None,
    output_folder_path: Optional[str] = None,
) -> ProcessingResult:
    processor = VTECPaymentProcessor(log_emit)
    return processor.process(source_paths, overview_workbook_path, output_folder_path)


class VTECPaymentProcessor:
    def __init__(self, log_emit: Optional[Callable[[str], None]] = None):
        self.log_emit = log_emit or (lambda _msg: None)

    def log(self, message: str):
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            self.log_emit(f"[{stamp}] {message}")
        except Exception:
            pass

    def process(
        self,
        source_paths: Sequence[str],
        overview_workbook_path: Optional[str] = None,
        output_folder_path: Optional[str] = None,
    ) -> ProcessingResult:
        clean_sources = [os.path.abspath(os.fspath(path)) for path in source_paths if os.fspath(path).strip()]
        if not clean_sources:
            raise ValueError("No source payment files were selected.")

        for path in clean_sources:
            if not os.path.isfile(path):
                raise FileNotFoundError(f"Source file not found: {path}")

        output_path, wb_out, created_new = self._open_or_create_output_workbook(
            clean_sources,
            overview_workbook_path,
            output_folder_path,
        )

        result = ProcessingResult(output_path=output_path)
        if created_new:
            self.log(f"Created output workbook: {output_path}")
        else:
            self.log(f"Using existing overview workbook: {output_path}")

        try:
            overview_ws = get_or_create_sheet(wb_out, OVERVIEW_SHEET_NAME)
            dup_ws = get_or_create_sheet(wb_out, DUPLICATES_SHEET_NAME)
            log_ws = get_or_create_sheet(wb_out, LOG_SHEET_NAME)
            rejected_log_ws = get_or_create_sheet_after(wb_out, REJECTED_LOG_SHEET_NAME, LOG_SHEET_NAME)

            ensure_overview_sheet(overview_ws)
            ensure_overview_sheet(dup_ws)
            ensure_log_headers(log_ws)
            ensure_rejected_log_headers(rejected_log_ws)

            processed_log = load_processed_log(log_ws)
            current_summary_row = max(DATA_START_ROW, last_used_row_in_column(overview_ws, 1) + 1)
            current_dup_row = max(DATA_START_ROW, last_used_row_in_column(dup_ws, 1) + 1)

            new_rows: List[List[Any]] = []

            for source_path in clean_sources:
                file_name = os.path.basename(source_path)
                self.log(f"Opening source workbook: {file_name}")

                source_book: Optional[SourceWorkbookReader] = None
                try:
                    source_book = SourceWorkbookReader(source_path)
                    supplier_lookup = create_supplier_lookup_dict(source_book)

                    for sheet in source_book.sheets:
                        log_key = f"{source_book.name}|{sheet.name}"

                        sheet_name_lower = sheet.name.lower()
                        if "payment" not in sheet_name_lower or "(v)" not in sheet_name_lower:
                            continue

                        if log_key.lower() in processed_log:
                            result.skipped_sheets += 1
                            log_rejected_sheet(
                                rejected_log_ws,
                                source_book.name,
                                sheet.name,
                                "Already processed in Processing Log",
                            )
                            self.log(f"Skipped already processed sheet: {source_book.name} -> {sheet.name}")
                            continue

                        rows_from_sheet, rejection_reason = self._collect_rows_from_payment_sheet(
                            source_book,
                            sheet,
                            supplier_lookup,
                        )

                        if not rows_from_sheet:
                            result.skipped_sheets += 1
                            reason = rejection_reason or "No usable payment rows found"
                            log_rejected_sheet(rejected_log_ws, source_book.name, sheet.name, reason)
                            self.log(f"Rejected sheet: {source_book.name} -> {sheet.name} ({reason})")
                            continue

                        new_rows.extend(rows_from_sheet)
                        result.imported_rows += len(rows_from_sheet)
                        result.processed_sheets += 1
                        log_processed_sheet(log_ws, source_book.name, sheet.name)
                        processed_log.add(log_key.lower())
                        self.log(
                            f"Processed sheet: {source_book.name} -> {sheet.name} "
                            f"({len(rows_from_sheet)} rows)"
                        )

                except Exception as exc:
                    result.failed_files += 1
                    result.success = False
                    self.log(f"ERROR processing {file_name}: {exc}")
                finally:
                    if source_book is not None:
                        source_book.close()

            if new_rows:
                write_output_rows(overview_ws, current_summary_row, new_rows)
                self.log(f"Appended {len(new_rows)} row(s) to {OVERVIEW_SHEET_NAME}.")
            else:
                self.log("No new rows to append.")

            result.duplicate_rows_moved = move_later_duplicates(overview_ws, dup_ws, current_dup_row)
            if result.duplicate_rows_moved:
                self.log(f"Moved {result.duplicate_rows_moved} duplicate row(s) to {DUPLICATES_SHEET_NAME}.")

            apply_formatting(overview_ws)
            apply_formatting(dup_ws)
            autofit_log_sheet(log_ws)
            autofit_rejected_log_sheet(rejected_log_ws)

            ensure_parent_folder(output_path)
            wb_out.save(output_path)
            self.log("Workbook saved.")
            return result
        finally:
            try:
                wb_out.close()
            except Exception:
                pass

    def _resolve_output_path(
        self,
        source_paths: Sequence[str],
        output_folder_path: Optional[str] = None,
    ) -> str:
        selected_folder = (output_folder_path or "").strip()
        if selected_folder:
            output_dir = os.path.abspath(selected_folder)
            if not os.path.isdir(output_dir):
                raise NotADirectoryError(f"Output folder not found: {output_dir}")
        else:
            output_dir = os.path.dirname(os.path.abspath(source_paths[0]))

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return os.path.join(output_dir, f"VTEC Payment Overview {timestamp}.xlsx")

    def _open_or_create_output_workbook(
        self,
        source_paths: Sequence[str],
        overview_workbook_path: Optional[str],
        output_folder_path: Optional[str],
    ) -> Tuple[str, Any, bool]:
        selected_path = (overview_workbook_path or "").strip()
        if selected_path:
            if (output_folder_path or "").strip():
                self.log("Selected new output folder ignored because an existing overview workbook was selected.")
            output_path = os.path.abspath(selected_path)
            if not os.path.isfile(output_path):
                raise FileNotFoundError(f"Overview workbook not found: {output_path}")

            ext = os.path.splitext(output_path)[1].lower()
            if ext not in {".xlsx", ".xlsm"}:
                raise ValueError("Existing overview workbook must be an .xlsx or .xlsm file.")

            keep_vba = ext == ".xlsm"
            workbook = load_workbook(output_path, keep_vba=keep_vba)
            return output_path, workbook, False

        output_path = self._resolve_output_path(source_paths, output_folder_path)
        workbook = self._create_output_workbook()
        return output_path, workbook, True

    def _create_output_workbook(self):
        wb = Workbook()
        overview_ws = wb.active
        overview_ws.title = OVERVIEW_SHEET_NAME
        dup_ws = wb.create_sheet(DUPLICATES_SHEET_NAME)
        log_ws = wb.create_sheet(LOG_SHEET_NAME)
        rejected_log_ws = wb.create_sheet(REJECTED_LOG_SHEET_NAME)

        write_headers(overview_ws)
        write_headers(dup_ws)
        ensure_log_headers(log_ws)
        ensure_rejected_log_headers(rejected_log_ws)

        return wb

    def _collect_rows_from_payment_sheet(
        self,
        source_book: "SourceWorkbookReader",
        sheet: "SheetReader",
        supplier_lookup: dict[str, str],
    ) -> Tuple[List[List[Any]], str]:
        header_row = find_header_row(sheet)
        if header_row == 0:
            return [], "No detected header row"

        last_row = sheet.last_row_in_column(1)
        if last_row <= header_row:
            return [], "No data rows below header row"

        last_col = sheet.last_col_in_row(header_row)
        column_map = identify_columns(sheet, header_row, last_col)

        if column_map[3] == 0:
            return [], "Missing VAT Invoice Number column"

        collected: List[List[Any]] = []
        vat_col = column_map[3]

        for row_idx in range(header_row + 1, last_row + 1):
            vat_value = sheet.cell(row_idx, vat_col)
            if is_blank(vat_value):
                continue

            output_row: List[Any] = [None] * NUM_OUTPUT_COLS
            for output_idx in range(1, NUM_OUTPUT_COLS + 1):
                source_col = column_map[output_idx]
                if source_col > 0:
                    output_row[output_idx - 1] = sheet.cell(row_idx, source_col)

            conditional_usd_col = column_map.get(CONDITIONAL_USD_MAP_KEY, 0)
            if conditional_usd_col > 0:
                usd_value = sheet.cell(row_idx, conditional_usd_col)
                if not is_blank(usd_value):
                    usd_output_idx = 10 if not is_blank(output_row[6]) else 8
                    if is_blank(output_row[usd_output_idx - 1]):
                        output_row[usd_output_idx - 1] = usd_value

            if is_blank(output_row[0]):
                output_row[0] = lookup_supplier_name_from_dict(
                    supplier_lookup,
                    output_row[2],
                    output_row[7],
                )

            for date_col in DATE_COLUMN_INDICES:
                output_row[date_col - 1] = normalise_date_value(output_row[date_col - 1])

            collected.append(output_row)

        if not collected:
            return [], "No usable rows with VAT Invoice Number"

        return collected, ""


# =====================================================================================
# Source workbook readers
# =====================================================================================


class SheetReader:
    def __init__(self, sheet: Any, kind: str, datemode: int = 0):
        self._sheet = sheet
        self._kind = kind
        self._datemode = datemode
        self.name = str(getattr(sheet, "title", None) or getattr(sheet, "name", ""))

    @property
    def max_row(self) -> int:
        if self._kind == "xlrd":
            return int(self._sheet.nrows)
        return int(self._sheet.max_row or 0)

    @property
    def max_col(self) -> int:
        if self._kind == "xlrd":
            return int(self._sheet.ncols)
        return int(self._sheet.max_column or 0)

    def cell(self, row: int, col: int) -> Any:
        if row <= 0 or col <= 0:
            return None
        if self._kind == "xlrd":
            if row > self._sheet.nrows or col > self._sheet.ncols:
                return None
            cell_obj = self._sheet.cell(row - 1, col - 1)
            if xlrd is not None and cell_obj.ctype == xlrd.XL_CELL_EMPTY:
                return None
            if xlrd is not None and cell_obj.ctype == xlrd.XL_CELL_DATE:
                try:
                    return xlrd.xldate_as_datetime(cell_obj.value, self._datemode)
                except Exception:
                    return cell_obj.value
            return cell_obj.value
        return self._sheet.cell(row=row, column=col).value

    def last_col_in_row(self, row: int) -> int:
        for col in range(self.max_col, 0, -1):
            if not is_blank(self.cell(row, col)):
                return col
        return 0

    def last_row_in_column(self, col: int) -> int:
        for row in range(self.max_row, 0, -1):
            if not is_blank(self.cell(row, col)):
                return row
        return 0


class SourceWorkbookReader:
    def __init__(self, path: str):
        self.path = os.path.abspath(path)
        self.name = os.path.basename(path)
        self._kind = self._detect_kind(path)
        self._book: Any = None
        self.sheets: List[SheetReader] = []
        self._open()

    @staticmethod
    def _detect_kind(path: str) -> str:
        ext = os.path.splitext(path)[1].lower()
        if ext == ".xls":
            return "xlrd"
        if ext in {".xlsx", ".xlsm"}:
            return "openpyxl"
        raise ValueError(f"Unsupported source workbook type: {ext or '(no extension)'}")

    def _open(self):
        if self._kind == "xlrd":
            if xlrd is None:
                raise RuntimeError("xlrd is required to read legacy .xls files. Please install this Pycro's requirements.")
            self._book = xlrd.open_workbook(self.path, on_demand=True)
            self.sheets = [SheetReader(self._book.sheet_by_index(i), "xlrd", self._book.datemode) for i in range(self._book.nsheets)]
        else:
            self._book = load_workbook(self.path, data_only=True, read_only=False)
            self.sheets = [SheetReader(ws, "openpyxl") for ws in self._book.worksheets]

    def sheet_by_index(self, index: int) -> Optional[SheetReader]:
        if 0 <= index < len(self.sheets):
            return self.sheets[index]
        return None

    def close(self):
        if self._book is None:
            return
        try:
            self._book.close()
        except Exception:
            pass


# =====================================================================================
# Processing helpers
# =====================================================================================


def find_header_row(sheet: SheetReader) -> int:
    scan_rows = min(20, sheet.max_row)

    for row in range(1, scan_rows + 1):
        col_a = value_to_str(sheet.cell(row, 1)).strip()
        col_b = value_to_str(sheet.cell(row, 2))
        if col_a == "No" or "supplier name" in col_b.lower():
            return row

    for row in range(1, scan_rows + 1):
        temp_last_col = sheet.last_col_in_row(row)
        for col in range(1, temp_last_col + 1):
            if "vat invoice number" in value_to_str(sheet.cell(row, col)).lower():
                return row

    return 0


def identify_columns(sheet: SheetReader, header_row: int, last_col: int) -> dict[int, int]:
    column_map = {idx: 0 for idx in range(1, NUM_OUTPUT_COLS + 1)}

    for col in range(1, last_col + 1):
        header_text = value_to_str(sheet.cell(header_row, col)).lower()
        header_trimmed = header_text.strip()

        if "payment to supplier" in header_text:
            column_map[19] = col
        elif "payment to vtec" in header_text:
            column_map[18] = col
        elif "sig due date" in header_text:
            column_map[17] = col
        elif "vat invoice number" in header_text:
            column_map[3] = col
        elif "vat invoice date" in header_text:
            column_map[11] = col
        elif ("po amount" in header_text or "po amt" in header_text) and "vnd" in header_text:
            column_map[5] = col
        elif ("po amount" in header_text or "po amt" in header_text) and "usd" in header_text:
            column_map[8] = col
        elif "surcharge" in header_text and "vnd" in header_text:
            column_map[6] = col
        elif "surcharge" in header_text and "usd" in header_text:
            column_map[9] = col
        elif "total" in header_text and "vnd" in header_text:
            column_map[7] = col
        elif "total" in header_text and "usd" in header_text:
            column_map[10] = col
        elif header_trimmed == "usd" and column_map.get(CONDITIONAL_USD_MAP_KEY, 0) == 0:
            column_map[CONDITIONAL_USD_MAP_KEY] = col
        elif "supplier name" in header_text:
            column_map[1] = col
        elif "payment term" in header_text:
            column_map[2] = col
        elif "other references" in header_text:
            column_map[4] = col
        elif "currency rate" in header_text or "exchange rate" in header_text:
            column_map[12] = col
        elif "purchaser" in header_text:
            column_map[13] = col
        elif "order type" in header_text:
            column_map[14] = col
        elif header_trimmed == "from":
            column_map[15] = col
        elif header_trimmed == "to":
            column_map[16] = col
        elif column_map[17] == 0 and "due date" in header_text:
            column_map[17] = col

    return column_map


def create_supplier_lookup_dict(source_book: SourceWorkbookReader) -> dict[str, str]:
    lookup: dict[str, str] = {}
    sheet = source_book.sheet_by_index(1)
    if sheet is None:
        return lookup

    last_lookup_row = sheet.last_row_in_column(6)  # column F in the original workbook
    if last_lookup_row < 2:
        return lookup

    for row in range(2, last_lookup_row + 1):
        supplier_name = sheet.cell(row, 2)  # B
        vat_invoice = sheet.cell(row, 6)   # F
        po_amount_usd = sheet.cell(row, 12)  # L

        if is_blank(vat_invoice) or is_blank(supplier_name):
            continue

        key = f"{value_to_str(vat_invoice).strip()}|{number_key(po_amount_usd)}"
        if key.lower() not in lookup:
            lookup[key.lower()] = value_to_str(supplier_name)

    return lookup


def lookup_supplier_name_from_dict(lookup: dict[str, str], vat_invoice: Any, po_amount_usd: Any) -> str:
    if not lookup or is_blank(vat_invoice):
        return ""
    key = f"{value_to_str(vat_invoice).strip()}|{number_key(po_amount_usd)}"
    return lookup.get(key.lower(), "")


def normalise_date_value(value: Any) -> Any:
    if value is None:
        return value
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)

    text = value_to_str(value).strip()
    if not text:
        return value

    if "." in text:
        parts = text.split(".")
        if len(parts) >= 3:
            try:
                day = int(parts[0])
                month = int(parts[1])
                year = int(parts[2])
                if year < 100:
                    year += 2000 if year < 50 else 1900
                return datetime(year, month, day)
            except Exception:
                pass

    if is_number_like(text):
        try:
            num = float(clean_number_text(text))
            if 0 < num < 2958466:
                return from_excel(num)
        except Exception:
            pass

    for fmt in (
        "%d/%m/%Y",
        "%d/%m/%y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    return value


def move_later_duplicates(overview_ws, dup_ws, current_dup_row: int) -> int:
    last_row_overview = last_used_row_in_column(overview_ws, 1)
    if last_row_overview < DATA_START_ROW:
        return 0

    data_rows = read_worksheet_rows(overview_ws, DATA_START_ROW, last_row_overview, NUM_OUTPUT_COLS)
    if not data_rows:
        return 0

    earliest: dict[str, Tuple[Any, int]] = {}
    key_counts: dict[str, int] = {}

    for idx, row_data in enumerate(data_rows, start=1):
        vat_invoice = value_to_str(row_data[2]).strip() if len(row_data) >= 3 else ""
        if not vat_invoice:
            continue

        po_amount_vnd = row_data[4] if len(row_data) >= 5 else None
        key = f"{vat_invoice_match_key(vat_invoice)}|{number_key(po_amount_vnd)}"
        key_counts[key] = key_counts.get(key, 0) + 1

        payment_date = row_data[17] if len(row_data) >= 18 else None
        if key not in earliest:
            earliest[key] = (payment_date, idx)
            continue

        existing_date, _existing_idx = earliest[key]
        new_dt = date_for_compare(payment_date)
        old_dt = date_for_compare(existing_date)

        if new_dt is not None:
            if old_dt is None or new_dt < old_dt:
                earliest[key] = (payment_date, idx)

    earliest_indices = {row_idx for _date_value, row_idx in earliest.values()}
    keep_rows: List[List[Any]] = []
    move_rows: List[List[Any]] = []
    rows_to_highlight: List[int] = []

    for idx, row_data in enumerate(data_rows, start=1):
        full_row = list(row_data[:NUM_OUTPUT_COLS])
        while len(full_row) < NUM_OUTPUT_COLS:
            full_row.append(None)

        vat_invoice = value_to_str(full_row[2]).strip()
        po_amount_vnd = full_row[4]
        key = f"{vat_invoice_match_key(vat_invoice)}|{number_key(po_amount_vnd)}"

        if not vat_invoice:
            keep_rows.append(full_row)
        elif idx in earliest_indices:
            if key_counts.get(key, 0) > 1:
                rows_to_highlight.append(DATA_START_ROW + len(keep_rows))
            keep_rows.append(full_row)
        else:
            move_rows.append(full_row)

    clear_data_area(overview_ws, DATA_START_ROW, NUM_OUTPUT_COLS)
    if keep_rows:
        write_output_rows(overview_ws, DATA_START_ROW, keep_rows)

    moved_count = len(move_rows)
    if move_rows:
        if current_dup_row < DATA_START_ROW:
            current_dup_row = DATA_START_ROW
        write_output_rows(dup_ws, current_dup_row, move_rows)

    if rows_to_highlight:
        highlight_fill = PatternFill("solid", fgColor="FFFFFF00")
        for row_num in rows_to_highlight:
            for col in range(1, NUM_OUTPUT_COLS + 1):
                overview_ws.cell(row=row_num, column=col).fill = highlight_fill

    return moved_count


# =====================================================================================
# Workbook writing and formatting helpers
# =====================================================================================


def load_processed_log(log_ws) -> set[str]:
    ensure_log_headers(log_ws)
    processed: set[str] = set()
    last_row = last_used_row_in_column(log_ws, 1)

    for row in range(2, last_row + 1):
        file_name = log_ws.cell(row=row, column=1).value
        sheet_name = log_ws.cell(row=row, column=2).value
        if not is_blank(file_name) and not is_blank(sheet_name):
            processed.add(f"{value_to_str(file_name)}|{value_to_str(sheet_name)}".lower())

    return processed


def ensure_log_headers(log_ws):
    if not is_blank(log_ws.cell(row=1, column=1).value):
        return
    log_ws.cell(row=1, column=1, value="Source FileName")
    log_ws.cell(row=1, column=2, value="Source Worksheet")
    log_ws.cell(row=1, column=3, value="Timestamp")
    for col in range(1, 4):
        log_ws.cell(row=1, column=col).font = Font(bold=True)


def ensure_rejected_log_headers(log_ws):
    if not is_blank(log_ws.cell(row=1, column=1).value):
        return
    log_ws.cell(row=1, column=1, value="Source FileName")
    log_ws.cell(row=1, column=2, value="Source Worksheet")
    log_ws.cell(row=1, column=3, value="Timestamp")
    log_ws.cell(row=1, column=4, value="Rejected Reason")
    for col in range(1, 5):
        log_ws.cell(row=1, column=col).font = Font(bold=True)


def log_processed_sheet(log_ws, file_name: str, worksheet_name: str):
    ensure_log_headers(log_ws)
    next_row = last_used_row_in_column(log_ws, 1) + 1
    log_ws.cell(row=next_row, column=1, value=file_name)
    log_ws.cell(row=next_row, column=2, value=worksheet_name)
    timestamp_cell = log_ws.cell(row=next_row, column=3, value=datetime.now())
    timestamp_cell.number_format = "yyyy-mm-dd hh:mm:ss"


def log_rejected_sheet(log_ws, file_name: str, worksheet_name: str, reason: str):
    ensure_rejected_log_headers(log_ws)
    next_row = last_used_row_in_column(log_ws, 1) + 1
    log_ws.cell(row=next_row, column=1, value=file_name)
    log_ws.cell(row=next_row, column=2, value=worksheet_name)
    timestamp_cell = log_ws.cell(row=next_row, column=3, value=datetime.now())
    timestamp_cell.number_format = "yyyy-mm-dd hh:mm:ss"
    log_ws.cell(row=next_row, column=4, value=reason)


def get_or_create_sheet(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    if len(wb.worksheets) == 1 and wb.active.title == "Sheet" and worksheet_is_empty(wb.active):
        wb.active.title = sheet_name
        return wb.active

    return wb.create_sheet(title=sheet_name)


def get_or_create_sheet_after(wb, sheet_name: str, after_sheet_name: str):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    if after_sheet_name in wb.sheetnames:
        return wb.create_sheet(title=sheet_name, index=wb.sheetnames.index(after_sheet_name) + 1)

    return wb.create_sheet(title=sheet_name)


def worksheet_is_empty(ws) -> bool:
    for row in ws.iter_rows():
        for cell in row:
            if not is_blank(cell.value):
                return False
    return True


def ensure_overview_sheet(ws):
    if is_blank(ws.cell(row=1, column=1).value):
        write_headers(ws)


def write_headers(ws):
    ws.cell(row=1, column=1, value=ws.title)

    try:
        for merged_range in list(ws.merged_cells.ranges):
            if str(merged_range) == "A1:S1":
                ws.unmerge_cells(str(merged_range))
    except Exception:
        pass

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_OUTPUT_COLS)
    title_cell = ws.cell(row=1, column=1)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.font = Font(size=14, bold=True)

    header_fill = PatternFill("solid", fgColor="FFA9D08E")
    for idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=idx, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_output_rows(ws, start_row: int, rows: Sequence[Sequence[Any]]):
    for row_offset, row_data in enumerate(rows):
        excel_row = start_row + row_offset
        for col_idx in range(1, NUM_OUTPUT_COLS + 1):
            value = row_data[col_idx - 1] if col_idx - 1 < len(row_data) else None
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            if col_idx == 3:
                cell.number_format = "@"
            elif col_idx in DATE_COLUMN_INDICES:
                cell.number_format = "dd/mm/yyyy"


def read_worksheet_rows(ws, start_row: int, end_row: int, num_cols: int) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for row_idx in range(start_row, end_row + 1):
        rows.append([ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, num_cols + 1)])
    return rows


def clear_data_area(ws, start_row: int, num_cols: int):
    max_row = ws.max_row or 0
    if max_row < start_row:
        return
    ws.delete_rows(start_row, max_row - start_row + 1)


def apply_formatting(ws):
    last_row = max(HEADER_ROW, last_used_row_in_column(ws, 1))
    thin = Side(style="thin", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in range(HEADER_ROW, last_row + 1):
        for col in range(1, NUM_OUTPUT_COLS + 1):
            ws.cell(row=row, column=col).border = border

    for col in range(5, 8):
        set_column_number_format(ws, col, "#,##0")
    for col in range(8, 11):
        set_column_number_format(ws, col, "#,##0.00")
    set_column_number_format(ws, 3, "@")
    set_column_number_format(ws, 12, "#,##0.0000")
    set_column_number_format(ws, 11, "dd/mm/yyyy")
    for col in range(17, 20):
        set_column_number_format(ws, col, "dd/mm/yyyy")

    for row in range(1, last_row + 1):
        ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")

    auto_fit_columns(ws, 1, NUM_OUTPUT_COLS, last_row)
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions[get_column_letter(6)].width = 20


def set_column_number_format(ws, col_idx: int, number_format: str):
    last_row = max(HEADER_ROW, ws.max_row or 0)
    for row in range(DATA_START_ROW, last_row + 1):
        ws.cell(row=row, column=col_idx).number_format = number_format


def auto_fit_columns(ws, min_col: int, max_col: int, max_row: int):
    for col_idx in range(min_col, max_col + 1):
        max_len = 0
        for row in range(1, max_row + 1):
            value = ws.cell(row=row, column=col_idx).value
            if value is None:
                continue
            length = max(len(part) for part in str(value).splitlines()) if str(value).splitlines() else 0
            max_len = max(max_len, length)
        width = min(max(max_len + 2, 8), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def autofit_log_sheet(log_ws):
    ensure_log_headers(log_ws)
    max_row = max(1, log_ws.max_row or 1)
    auto_fit_columns(log_ws, 1, 3, max_row)


def autofit_rejected_log_sheet(log_ws):
    ensure_rejected_log_headers(log_ws)
    max_row = max(1, log_ws.max_row or 1)
    auto_fit_columns(log_ws, 1, 4, max_row)


def last_used_row_in_column(ws, col_idx: int) -> int:
    max_row = ws.max_row or 0
    for row in range(max_row, 0, -1):
        if not is_blank(ws.cell(row=row, column=col_idx).value):
            return row
    return 0


# =====================================================================================
# Value helpers
# =====================================================================================


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def value_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def vat_invoice_match_key(value: Any) -> str:
    text = value_to_str(value).strip()
    numeric_tokens = re.findall(r"\d+", text)
    if not numeric_tokens:
        return text
    return numeric_tokens[-1].lstrip("0") or "0"


def clean_number_text(value: Any) -> str:
    return value_to_str(value).strip().replace(",", "")


def is_number_like(value: Any) -> bool:
    text = clean_number_text(value)
    if not text:
        return False
    try:
        Decimal(text)
        return True
    except InvalidOperation:
        return False


def number_key(value: Any) -> str:
    if is_blank(value) or not is_number_like(value):
        return "0"
    try:
        dec = Decimal(clean_number_text(value))
    except InvalidOperation:
        return "0"

    normalized = dec.normalize()
    text = format(normalized, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def date_for_compare(value: Any) -> Optional[datetime]:
    normalized = normalise_date_value(value)
    if isinstance(normalized, datetime):
        return normalized
    if isinstance(normalized, date):
        return datetime(normalized.year, normalized.month, normalized.day)
    return None


def ensure_parent_folder(path: str):
    parent = os.path.dirname(os.path.abspath(path))
    if parent:
        os.makedirs(parent, exist_ok=True)
