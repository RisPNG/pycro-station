import os
import re
import threading
import time
from datetime import datetime
from typing import Callable, List, Optional, Tuple

import openpyxl
from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QSizePolicy,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)
from qfluentwidgets import MessageBox, PrimaryPushButton

# TODO: Sort report checking order by "Change Date". This issue could arise one day but currently it is sorted by filename

class App:
    def __init__(self):
        super().__init__()

        self.title("SIG Memo Match Utility v6.9.0")
        self.geometry("1280x720")

        # set grid layout 1x2
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        # load images with light and dark mode image
        # load images with light and dark mode image
        image_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "test_images")
        self.logo_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "CustomTkinter_logo_single.png")), size=(26, 26))
        self.large_test_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "large_test_image.png")), size=(500, 150))
        self.image_icon_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "file-circle-plus-solid-light.png")), size=(20, 20))
        self.home_image = customtkinter.CTkImage(
            light_image=Image.open(os.path.join(image_path, "home_dark.png")),
            dark_image=Image.open(os.path.join(image_path, "home_light.png")),
            size=(20, 20)
        )
        self.chat_image = customtkinter.CTkImage(
            light_image=Image.open(os.path.join(image_path, "file-solid-dark.png")),
            dark_image=Image.open(os.path.join(image_path, "file-solid-light.png")),
            size=(20, 20)
        )
        self.reports_warning = customtkinter.CTkImage(
            light_image=Image.open(os.path.join(image_path, "file-pen-solid-dark.png")),
            dark_image=Image.open(os.path.join(image_path, "file-pen-solid-light.png")),
            size=(20, 20)
        )
        self.reports_error = customtkinter.CTkImage(
            light_image=Image.open(os.path.join(image_path, "file-circle-exclamation-solid-dark.png")),
            dark_image=Image.open(os.path.join(image_path, "file-circle-exclamation-solid-light.png")),
            size=(20, 20)
        )
        self.reports_done = customtkinter.CTkImage(
            light_image=Image.open(os.path.join(image_path, "file-circle-check-solid-dark.png")),
            dark_image=Image.open(os.path.join(image_path, "file-circle-check-solid-light.png")),
            size=(20, 20)
        )
        self.add_user_image = customtkinter.CTkImage(
            light_image=Image.open(os.path.join(image_path, "file-solid-dark.png")),
            dark_image=Image.open(os.path.join(image_path, "file-solid-light.png")),
            size=(20, 20)
        )


        # create navigation frame
        self.navigation_frame = customtkinter.CTkFrame(self, corner_radius=0)
        self.navigation_frame.grid(row=0, column=0, sticky="nsew")
        self.navigation_frame.grid_rowconfigure(4, weight=1)

        self.navigation_frame_label = customtkinter.CTkLabel(self.navigation_frame, text="  Memo Match Utility", image=self.logo_image,
                                                             compound="left", font=customtkinter.CTkFont(size=15, weight="bold"))
        self.navigation_frame_label.grid(row=0, column=0, padx=20, pady=20)

        self.home_button = customtkinter.CTkButton(self.navigation_frame, corner_radius=0, height=40, border_spacing=10, text="Home",
                                                   fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                   image=self.home_image, anchor="w", command=self.home_button_event)
        self.home_button.grid(row=1, column=0, sticky="ew")

        self.frame_2_button = customtkinter.CTkButton(self.navigation_frame, corner_radius=0, height=40, border_spacing=10, text="Reports",
                                                      fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                      image=self.chat_image, anchor="w", command=self.frame_2_button_event)
        self.frame_2_button.grid(row=2, column=0, sticky="ew")

        self.frame_3_button = customtkinter.CTkButton(self.navigation_frame, corner_radius=0, height=40, border_spacing=10, text="Logs",
                                                      fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                      image=self.add_user_image, anchor="w", command=self.frame_3_button_event)
        self.frame_3_button.grid(row=3, column=0, sticky="ew")

        self.appearance_mode_menu = customtkinter.CTkOptionMenu(self.navigation_frame, values=["System", "Light", "Dark"],
                                                                command=self.change_appearance_mode_event)
        self.appearance_mode_menu.grid(row=6, column=0, padx=20, pady=20, sticky="s")

        # create home frame
        self.home_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.home_frame.grid_columnconfigure(0, weight=1)

        self.home_frame_large_image_label = customtkinter.CTkLabel(self.home_frame, text="", image=self.large_test_image)
        self.home_frame_large_image_label.grid(row=0, column=0, padx=20, pady=10)

        self.home_frame_button_1 = customtkinter.CTkButton(self.home_frame, text="Upload Master File", image=self.image_icon_image, width=200, command=self.home_frame_button_1_event)
        self.home_frame_button_1.grid(row=1, column=0, padx=20, pady=10)
        # self.home_frame_button_1.configure(width=200)
        self.home_frame_textbox_1 = customtkinter.CTkTextbox(self.home_frame, state="disabled", width=400, height=100)
        self.home_frame_textbox_1.grid(row=2, column=0, padx=20, pady=10)
        # self.home_frame_entry_1 = customtkinter.CTkEntry(self.home_frame, placeholder_text="File path will display here", state="readonly", width=400)
        # self.home_frame_entry_1.grid(row=2, column=0, padx=20, pady=10)
        # self.home_frame_entry_1.configure(state="disabled", width=400)

        self.home_frame_button_2 = customtkinter.CTkButton(self.home_frame, text="Upload PPM Report(s)", state="disabled", image=self.image_icon_image, width=200, command=self.home_frame_button_2_event)
        self.home_frame_button_2.grid(row=3, column=0, padx=20, pady=10)
        # self.home_frame_button_2.configure(width=200)
        # self.home_frame_entry_2 = customtkinter.CTkEntry(self.home_frame, placeholder_text="File path will display here")
        # self.home_frame_entry_2.grid(row=4, column=0, padx=20, pady=10)
        # self.home_frame_entry_2.configure(state="disabled", width=400)
        self.home_frame_textbox_2 = customtkinter.CTkTextbox(self.home_frame, state="disabled", width=400, height=100)
        self.home_frame_textbox_2.grid(row=4, column=0, padx=20, pady=10)
        # self.home_frame_textbox_2.configure(state="disabled", width=400)

        self.home_frame_button_3 = customtkinter.CTkButton(self.home_frame, text="Analyze & Report", state="disabled", width=200, command=self.home_frame_button_3_event)
        self.home_frame_button_3.grid(row=5, column=0, padx=20, pady=10)
        # self.home_frame_button_3.configure(state="disabled", width=200)

        # create second frame
        self.second_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.second_frame.grid_columnconfigure(0, weight=1)
        self.second_frame_textbox_1 = customtkinter.CTkTextbox(self.second_frame, state="disabled", width=600, height=600)
        self.second_frame_textbox_1.grid(row=1, column=0, padx=20, pady=40)

        # create third frame
        self.third_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.third_frame.grid_columnconfigure(0, weight=1)
        self.third_frame_textbox_1 = customtkinter.CTkTextbox(self.third_frame, state="disabled", width=600, height=600)
        self.third_frame_textbox_1.grid(row=1, column=0, padx=20, pady=40)

        # self.copy_button = customtkinter.CTkButton(self.third_frame, text="Copy", command=self.copy_textbox_content)
        # self.copy_button.grid(row=2, column=0, padx=20, pady=10)

        # select default frame
        self.select_frame_by_name("home")

    # def copy_textbox_content(self):
    #     # Temporarily enable the textbox to copy its contents
    #     self.third_frame_textbox_1.configure(state="normal")
    #     content = self.third_frame_textbox_1.get("1.0", "end-1c")  # Get all text
    #     self.clipboard_clear()  # Clear clipboard
    #     self.clipboard_append(content)  # Copy content to clipboard
    #     self.third_frame_textbox_1.configure(state="disabled")  # Set back to disabled

    #     # Optional: Show a brief message to confirm copying
    #     print("Text copied to clipboard")

    def home_frame_button_1_event(self):
        file_paths = filedialog.askopenfilenames()
        if file_paths:
            self.full_master_file_path = file_paths
            self.home_frame_textbox_1.configure(state="normal")
            self.home_frame_textbox_1.delete("1.0", 'end')
            for path in file_paths:
                file_name = os.path.basename(path)
                self.home_frame_textbox_1.insert("end", file_name + "\n")
            self.home_frame_textbox_1.configure(state="disabled")
            self.home_frame_button_2.configure(state="normal")

    def home_frame_button_2_event(self):
        file_paths = filedialog.askopenfilenames()
        file_paths = sorted(file_paths, key=lambda x: x.split('/')[-1])
        if file_paths:
            self.full_report_file_paths = file_paths
            self.home_frame_textbox_2.configure(state="normal")
            self.home_frame_textbox_2.delete("1.0", "end")
            for path in file_paths:
                file_name = os.path.basename(path)
                self.home_frame_textbox_2.insert("end", file_name + "\n")
            self.home_frame_textbox_2.configure(state="disabled")
            self.home_frame_button_3.configure(state="normal")

    def home_frame_button_3_event(self):
        self.second_frame_textbox_1.configure(state="normal")
        self.second_frame_textbox_1.delete("1.0", "end")
        self.third_frame_textbox_1.configure(state="normal")
        self.third_frame_textbox_1.delete("1.0", "end")
        self.home_frame_button_1.configure(state="disabled")
        self.home_frame_button_2.configure(state="disabled")
        self.home_frame_button_3.configure(state="disabled")
        self.frame_2_button.configure(image=self.reports_warning, text="Reports (Processing...)")

        # Switch to Logs frame before processing
        self.home_button.configure(fg_color="transparent")
        self.frame_2_button.configure(fg_color="transparent")
        self.frame_3_button.configure(fg_color=("gray75", "gray25"))
        self.home_frame.grid_forget()
        self.second_frame.grid_forget()
        self.third_frame.grid(row=0, column=1, sticky="nsew")

        # Start the processing in a new thread
        thread = threading.Thread(target=self.process_files)
        thread.start()

    def process_files(self) -> Tuple[str, int, int]:
        start_time = time.time()
        last_output: str = ""
        total_inputs = len(getattr(self, "full_master_file_path", []) or [])
        # Keys: Master NK SAP PO <-> PPM Purchase Order Number and Master PO LINE ITEM <-> PPM PO Line Item Number
        master_columns_needed = [
            "NK SAP PO",
            "PO LINE ITEM",
            "CM FOB rec. date",
            "FINAL FOB (Regular sizes)",
            "FINAL FOB (Extended sizes)",
            "Extended Sizes",
            "DPOM - Incorrect FOB",
            "Price (Date)",
            "Price (Changes)",
            "Season",
            "Season Year",
            "Season (Date)",
            "Season (Changes)",
            "FG QTY",
            "FG QTY (Date)",
            "FG QTY (Changes)",
            "Doc Type",
            "Doc Type (Date)",
            "Doc Type (Changes)",
            "SHIP MODE",
            "SHIP MODE (Date)",
            "SHIP MODE (Changes)",
            "Plant Code",
            "Plant Code (Date)",
            "Plant Code (Changes)",
            "SHIP-TO",
            "SHIP-TO (Date)",
            "SHIP-TO (Changes)",
            "AFS Cat",
            "AFS Cat (Date)",
            "AFS Cat (Changes)",
            "VAS name",
            "VAS name (Date)",
            "VAS name (Changes)",
            "Hanger size",
            "Hanger size (Date)",
            "Hanger size (Changes)",
            "Ratio Qty",
            "Ratio Qty (Date)",
            "Ratio Qty (Changes)",
            "Customer PO (Deichmann Group only)",
            "Customer PO (Date)",
            "Customer PO (Changes)",
            "Latest CM Change Date",
            "JOB NO"
            ]
        ppm_columns_needed = [
            "Purchase Order Number",
            "PO Line Item Number",
            "Product Code",
            "Gross Price/FOB currency code",
            "Surcharge Min Mat Main Body currency code",
            "Surcharge Min Material Trim currency code",
            "Surcharge Misc currency code",
            "Surcharge VAS currency code",
            "Gross Price/FOB",
            "Surcharge Min Mat Main Body",
            "Surcharge Min Material Trim",
            "Surcharge Misc",
            "Surcharge VAS",
            "Size Description",
            "Planning Season Code",
            "Planning Season Year",
            "Total Item Quantity",
            "Doc Type",
            "Mode of Transportation Code",
            "Plant Code",
            "Ship To Customer Number",
            "Inventory Segment Code",
            "VAS name",
            "Hanger size",
            "Ratio quantity",
            "Customer PO",
            "Change Date",
            "GAC",
            "DPOM Line Item Status",
            "Document Date"
            ]
        output_texts = set()

        #Functions
        def normalize_header_text(value) -> str:
            if value is None:
                return ""
            # Collapse all whitespace (incl. newlines) and normalize case.
            return " ".join(str(value).split()).strip().lower()

        def find_columns_header(sheet, needed_columns):
            needed_norm = [normalize_header_text(c) for c in needed_columns]
            for row in sheet.iter_rows(max_col=sheet.max_column):
                for cell in row:
                    cell_norm = normalize_header_text(cell.value)
                    if not cell_norm:
                        continue
                    for key_norm in needed_norm:
                        if key_norm and key_norm in cell_norm:
                            return cell.row
            return None

        def find_columns_master(sheet, needed_columns, header):
            column_positions = {column: None for column in needed_columns}
            # Prepare to hold multiple columns
            column_positions["FINAL FOB (Extended sizes)"] = []
            column_positions["Season"] = []
            column_positions["FG QTY"] = []
            column_positions["Doc Type"] = []
            column_positions["SHIP MODE"] = []
            column_positions["Plant Code"] = []
            column_positions["SHIP-TO"] = []
            column_positions["AFS Cat"] = []
            column_positions["VAS name"] = []
            column_positions["Hanger size"] = []
            column_positions["Ratio Qty"] = []

            needed_norm = {key: normalize_header_text(key) for key in needed_columns}
            for row in sheet.iter_rows(min_row=header, max_row=header, max_col=sheet.max_column):
                for cell in row:
                    cell_norm = normalize_header_text(cell.value)
                    if not cell_norm:
                        continue
                    for key, key_norm in needed_norm.items():
                        if key_norm and key_norm in cell_norm:
                            # self.third_frame_textbox_1.insert("end", f"rn master {key_norm} in {cell_norm}...\n\n")
                            # Append to list if it's the specific multiple columned key
                            if key == "FINAL FOB (Extended sizes)":
                                column_positions[key].append(cell.column)
                            elif key == "Season":
                                column_positions[key].append(cell.column)
                            elif key == "FG QTY":
                                column_positions[key].append(cell.column)
                            elif key == "Doc Type":
                                column_positions[key].append(cell.column)
                            elif key == "SHIP MODE":
                                column_positions[key].append(cell.column)
                            elif key == "Plant Code":
                                column_positions[key].append(cell.column)
                            elif key == "SHIP-TO":
                                column_positions[key].append(cell.column)
                            elif key == "AFS Cat":
                                column_positions[key].append(cell.column)
                            elif key == "VAS name":
                                column_positions[key].append(cell.column)
                            elif key == "Hanger size":
                                column_positions[key].append(cell.column)
                            elif key == "Ratio Qty":
                                column_positions[key].append(cell.column)
                            else:
                                column_positions[key] = cell.column
            return column_positions

        def find_columns_report(sheet, needed_columns, header):
            column_positions = {column: None for column in needed_columns}
            needed_norm = {key: normalize_header_text(key) for key in needed_columns}
            for row in sheet.iter_rows(min_row=header, max_row=header, max_col=sheet.max_column):
                for cell in row:
                    cell_norm = normalize_header_text(cell.value)
                    if not cell_norm:
                        continue
                    for key, key_norm in needed_norm.items():
                        if key_norm and key_norm in cell_norm:
                            # self.third_frame_textbox_1.insert("end", f"rn report {cell_norm}...\n\n")
                            column_positions[key] = cell.column
            return column_positions

        def safe_float(value):
            try:
                return float(value)
            except (TypeError, ValueError):
                return float(0)

        def letter(n):
            try:
                n = int(n)
            except (TypeError, ValueError):
                return "N/A"
            if n <= 0:
                return "N/A"
            result = ""
            while n > 0:
                n -= 1
                result = chr(n % 26 + 65) + result
                n //= 26
            return result

        def col_letter(col_num, offset: int = 0) -> str:
            try:
                return letter(int(col_num) + int(offset))
            except (TypeError, ValueError):
                return "N/A"

        def row_value(row, col_num, idx_offset: int = 0):
            if col_num is None:
                return None
            try:
                idx = int(col_num) - 1 + int(idx_offset)
            except (TypeError, ValueError):
                return None
            if idx < 0 or idx >= len(row):
                return None
            return row[idx].value

        def report_change_datetime(row, report_cols) -> datetime:
            """Best-effort report change timestamp (Change Date -> Document Date -> now)."""
            raw = row_value(row, report_cols.get("Change Date"))
            if raw is None:
                raw = row_value(row, report_cols.get("Document Date"))

            if raw is None:
                return datetime.now()

            if hasattr(raw, "strftime"):
                return raw

            s_val = str(raw).strip()
            for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d/%m/%y"):
                try:
                    return datetime.strptime(s_val, fmt)
                except ValueError:
                    continue
            return datetime.now()

        def append_date_changes(master_row, master_cols, report_cols, row, change_type, report_value):
            ex_date = str(master_row[master_cols[f"{change_type} (Date)"] - 1].value).strip() if not isEmptyCell(master_row[master_cols[f"{change_type} (Date)"] - 1].value) else ""
            ex_change = str(master_row[master_cols[f"{change_type} (Changes)"] - 1].value).strip() if not isEmptyCell(master_row[master_cols[f"{change_type} (Changes)"] - 1].value) else ""

            dt = report_change_datetime(row, report_cols)
            current_date = str(dt.strftime('%m/%d')).strip()
            current_date_mmddyy = str(dt.strftime('%m/%d/%y')).strip()
            current_date_mdyy = str(dt.strftime('%m/%d/%y')).strip().replace('/0', '/').lstrip('0')
            current_change = str(report_value).strip()

            if ex_date and ex_change:
                # If there are existing dates and changes, check if they contain commas (indicating multiple entries)
                if "," in ex_date and "," in ex_change:
                    # Split the dates and changes into lists
                    ex_dates = [d.strip() for d in ex_date.split(",")]
                    ex_changes = [c.strip() for c in ex_change.split(",")]

                    # Zip the existing dates and changes together to maintain the pairing
                    existing_pairs = list(zip(ex_dates, ex_changes))
                else:
                    # If there's only one date and change, create a single pair
                    existing_pairs = [(ex_date, ex_change)] if ex_date and ex_change else []

                # Create the current pair for comparison
                current_pair = (current_date_mdyy, current_change)

                # Only append the new pair if it doesn't exist already
                if current_pair not in existing_pairs:
                    existing_pairs.append(current_pair)

                # Unzip the pairs back into separate date and change lists
                if existing_pairs:
                    ex_dates, ex_changes = zip(*existing_pairs)
                    ex_date = " , ".join(ex_dates)
                    ex_change = " , ".join(ex_changes)
            else:
                # If there's no existing date or change, simply assign the current ones
                ex_date = current_date_mdyy
                ex_change = current_change

            master_row[master_cols[f"{change_type} (Date)"] - 1].value = ex_date
            master_row[master_cols[f"{change_type} (Changes)"] - 1].value = ex_change

            master_row[master_cols["Latest CM Change Date"] - 1].value = current_date_mmddyy

            # Output to reports
            if output_text not in output_texts:
                output_texts.add(output_text)
                self.second_frame_textbox_1.insert("end", output_text)
                self.second_frame_textbox_1.see("end")

            # Append Date to CM FOB rec. date
            ex_fob_date = str(master_row[master_cols["CM FOB rec. date"] - 1].value).strip()
            if not isEmptyCell(ex_fob_date):
                ex_fob_dates = [d.strip() for d in str(ex_fob_date).split(",")]
                if current_date_mdyy not in ex_fob_dates:
                    ex_fob_dates.append(current_date_mdyy)
                    ex_fob_date = " , ".join(ex_fob_dates)
            else:
                ex_fob_date = current_date_mdyy

            master_row[master_cols["CM FOB rec. date"] - 1].value = ex_fob_date

        def keep_date_changes(master_row, master_cols, report_cols, row, change_type, report_value, key):
            # Retrieve original date and change values from ori_date and ori_change
            ex_date = ori_date[change_type].get(f"{key}", "")
            ex_change = ori_change[change_type].get(f"{key}", "")

            dt = report_change_datetime(row, report_cols)
            current_date = str(dt.strftime('%m/%d')).strip()
            current_date_mmddyy = str(dt.strftime('%m/%d/%y')).strip()
            current_date_mdyy = str(dt.strftime('%m/%d/%y')).strip().replace('/0', '/').lstrip('0')
            current_change = str(report_value).strip()

            if ex_date and ex_change:
                # If there are existing dates and changes, check if they contain commas (indicating multiple entries)
                if "," in ex_date and "," in ex_change:
                    # Split the dates and changes into lists
                    ex_dates = [d.strip() for d in ex_date.split(",")]
                    ex_changes = [c.strip() for c in ex_change.split(",")]

                    # Zip the existing dates and changes together to maintain the pairing
                    existing_pairs = list(zip(ex_dates, ex_changes))
                else:
                    # If there's only one date and change, create a single pair
                    existing_pairs = [(ex_date, ex_change)] if ex_date and ex_change else []

                # Create the current pair for comparison
                current_pair = (current_date_mdyy, current_change)

                # Only append the new pair if it doesn't exist already
                if current_pair in existing_pairs:
                    existing_pairs.remove(current_pair)

                # Unzip the pairs back into separate date and change lists
                if existing_pairs:
                    ex_dates, ex_changes = zip(*existing_pairs)
                    ex_date = " , ".join(ex_dates)
                    ex_change = " , ".join(ex_changes)

            master_row[master_cols[f"{change_type} (Date)"] - 1].value = ex_date
            master_row[master_cols[f"{change_type} (Changes)"] - 1].value = ex_change

        def empty_date_changes(master_row, master_cols, report_cols, row, change_type, key):
            # Retrieve original date and change values from ori_date and ori_change
            original_date = ori_date[change_type].get(f"{key}", "")
            original_change = ori_change[change_type].get(f"{key}", "")

            # Set the master row values back to the original
            master_row[master_cols[f"{change_type} (Date)"] - 1].value = original_date
            master_row[master_cols[f"{change_type} (Changes)"] - 1].value = original_change

        # Function to check if the value is a number (integer or decimal)
        def is_number(value):
            try:
                float(value)  # Try converting to a float
                return True
            except ValueError:
                return False

        def isEmptyCell(cell):
            cell = str(cell).strip()
            if cell == "None" or not cell or cell == "-":
                return True
            else:
                return False

        def isNumberAfterDash(val):
            if '-' in val:
                parts = val.split('-')
                return all(is_number(part.strip()) for part in parts[1:])  # Check if there's a number in the second part
            return False

        def hasComma(val):
            return ',' in str(val).strip()

        # Process
        try:
            report_date_now = 0
            report_date_bfr = 0
            self.third_frame_textbox_1.insert("end", f"Starts processing...\n\n")
            self.third_frame_textbox_1.see("end")
            total_data = 1
            master_current = 0
            total_master = 0
            master_dict = {}
            master_dict_data = {}
            newPO = {}
            existingPO = {}
            work_path = ""
            sizes = ["2XS", "XS", "S", "M", "L", "XL", "2XL", "3XL", "4XL", "5XL", "2XSS", "XSS", "SS", "MS", "LS", "XLS", "2XLS", "3XLS", "4XLS", "5XLS", "2XSL", "XSL", "SL", "ML", "XLL", "2XLL", "3XLL", "4XLL", "5XLL", "2XST", "XST", "ST", "MT", "LT", "XLT", "2XLT", "3XLT", "4XLT", "5XLT", "2XSTT", "XSTT", "STT", "MTT", "LTT", "XLTT", "2XLTT", "3XLTT", "4XLTT", "5XLTT", "0X", "1X", "2X", "3X", "4X", "5X", "0XT", "1XT", "2XT", "3XT", "4XT", "5XT", "0XTT", "1XTT", "2XTT", "3XTT", "4XTT", "5XTT", "CUST0", "CUST1", "CUST2", "CUST3", "CUST4", "CUST5", "CUST6", "CUST7", "CUST2XS", "CUSTXS", "CUSTS", "CUSTM", "CUSTL", "CUSTXL", "CUST2XL", "CUST3XL", "CUST4XL", "CUST5XL"]

            # Preflight: PPM columns are mandatory. Abort early if any report is missing expected columns.
            report_header_indicator = ["Purchase Order Number"]
            for report_path in self.full_report_file_paths:
                report_name = os.path.basename(report_path)
                self.third_frame_textbox_1.insert("end", f"Validating PPM report columns: {report_name}\n\n")
                self.third_frame_textbox_1.see("end")

                report_wb = openpyxl.load_workbook(report_path, read_only=True, data_only=True)
                try:
                    report_sheet = report_wb.active
                    report_header_row = find_columns_header(report_sheet, report_header_indicator)
                    if not report_header_row:
                        err = (
                            f"ERROR: PPM report '{report_name}' header row not found. "
                            f"Expected to find column: '{report_header_indicator[0]}'."
                        )
                        self.third_frame_textbox_1.insert("end", f"{err}\n\n")
                        self.third_frame_textbox_1.see("end")
                        return err, 0, total_inputs

                    report_cols = find_columns_report(report_sheet, ppm_columns_needed, report_header_row)
                    missing_cols = [c for c in ppm_columns_needed if report_cols.get(c) is None]
                    if missing_cols:
                        missing_list = "\n".join(f"- {c}" for c in missing_cols)
                        err = (
                            f"ERROR: PPM report '{report_name}' is missing required column(s):\n"
                            f"{missing_list}\n"
                            f"\nAborted. Please comeplete the PPM template."
                        )
                        self.third_frame_textbox_1.insert("end", f"{err}\n\n")
                        self.third_frame_textbox_1.see("end")
                        return err, 0, total_inputs
                finally:
                    try:
                        report_wb.close()
                    except Exception:
                        pass

            for master_path in self.full_master_file_path:
                master_current +=1
                total_master = master_current
                work_path = os.path.dirname(master_path)
            master_current = 0
            for master_path in self.full_master_file_path:
                master_current +=1
                self.third_frame_textbox_1.insert("end", f"Reading {os.path.basename(master_path)}...\n\n")
                self.third_frame_textbox_1.see("end")
                # Write mode
                master_wb = openpyxl.load_workbook(master_path)
                master_sheet = master_wb[master_wb.sheetnames[0]]
                # Data only mode for comparison
                master_wb_data = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
                master_sheet_data = master_wb_data[master_wb_data.sheetnames[0]]
                master_header_indicator = ["NK SAP PO"]
                self.third_frame_textbox_1.insert("end", f"Finding header row...\n\n")
                self.third_frame_textbox_1.see("end")
                master_header_row = find_columns_header(master_sheet, master_header_indicator)
                self.third_frame_textbox_1.insert("end", f"Header row found at {master_header_row}.\n\n")
                self.third_frame_textbox_1.see("end")
                master_cols = find_columns_master(master_sheet, master_columns_needed, master_header_row)

                report_text = (
                    f"Master Column Detection Report:\n"
                    f"Header Row: {master_header_row}\n"
                    f"NK SAP PO: {letter(master_cols['NK SAP PO'])}\n"
                    f"PO LINE ITEM: {letter(master_cols['PO LINE ITEM'])}\n"
                    f"CM FOB rec. date: {letter(master_cols['CM FOB rec. date'])}\n"
                    f"FINAL FOB (Regular sizes): {letter(master_cols['FINAL FOB (Regular sizes)'])}\n"
                    f"FINAL FOB (Extended sizes 1): {letter(master_cols['FINAL FOB (Extended sizes)'][0])}\n"
                    f"FINAL FOB (Extended sizes 2): {letter(master_cols['FINAL FOB (Extended sizes)'][1])}\n"
                    f"Extended Sizes 1: {letter(master_cols['Extended Sizes']-2)}\n"
                    f"Extended Sizes 2: {letter(master_cols['Extended Sizes'])}\n"
                    f"DPOM - Incorrect FOB: {letter(master_cols['DPOM - Incorrect FOB'])}\n"
                    f"Season: {letter(master_cols['Season'][0])}\n"
                    f"Season Year: {letter(master_cols['Season Year'])}\n"
                    f"Season (Date): {letter(master_cols['Season (Date)'])}\n"
                    f"Season (Changes): {letter(master_cols['Season (Changes)'])}\n"
                    f"FG QTY: {letter(master_cols['FG QTY'][0])}\n"
                    f"FG QTY (Date): {letter(master_cols['FG QTY (Date)'])}\n"
                    f"FG QTY (Changes): {letter(master_cols['FG QTY (Changes)'])}\n"
                    f"Doc Type: {letter(master_cols['Doc Type'][0])}\n"
                    f"Doc Type (Date): {letter(master_cols['Doc Type (Date)'])}\n"
                    f"Doc Type (Changes): {letter(master_cols['Doc Type (Changes)'])}\n"
                    f"SHIP MODE: {letter(master_cols['SHIP MODE'][0])}\n"
                    f"SHIP MODE (Date): {letter(master_cols['SHIP MODE (Date)'])}\n"
                    f"SHIP MODE (Changes): {letter(master_cols['SHIP MODE (Changes)'])}\n"
                    f"Plant Code: {letter(master_cols['Plant Code'][0])}\n"
                    f"Plant Code (Date): {letter(master_cols['Plant Code (Date)'])}\n"
                    f"Plant Code (Changes): {letter(master_cols['Plant Code (Changes)'])}\n"
                    f"SHIP-TO: {letter(master_cols['SHIP-TO'][0])}\n"
                    f"SHIP-TO (Date): {letter(master_cols['SHIP-TO (Date)'])}\n"
                    f"SHIP-TO (Changes): {letter(master_cols['SHIP-TO (Changes)'])}\n"
                    f"AFS Cat: {letter(master_cols['AFS Cat'][0])}\n"
                    f"AFS Cat (Date): {letter(master_cols['AFS Cat (Date)'])}\n"
                    f"AFS Cat (Changes): {letter(master_cols['AFS Cat (Changes)'])}\n"
                    f"VAS name: {letter(master_cols['VAS name'][0])}\n"
                    f"VAS name (Date): {letter(master_cols['VAS name (Date)'])}\n"
                    f"VAS name (Changes): {letter(master_cols['VAS name (Changes)'])}\n"
                    f"Hanger size: {letter(master_cols['Hanger size'][0])}\n"
                    f"Hanger size (Date): {letter(master_cols['Hanger size (Date)'])}\n"
                    f"Hanger size (Changes): {letter(master_cols['Hanger size (Changes)'])}\n"
                    f"Ratio Qty: {letter(master_cols['Ratio Qty'][0])}\n"
                    f"Ratio Qty (Date): {letter(master_cols['Ratio Qty (Date)'])}\n"
                    f"Ratio Qty (Changes): {letter(master_cols['Ratio Qty (Changes)'])}\n"
                    f"Customer PO (Deichmann Group only): {letter(master_cols['Customer PO (Deichmann Group only)'])}\n"
                    f"Customer PO (Date): {letter(master_cols['Customer PO (Date)'])}\n"
                    f"Customer PO (Changes): {letter(master_cols['Customer PO (Changes)'])}\n"
                    f"Latest CM Change Date: {letter(master_cols['Latest CM Change Date'])}\n"
                    f"JOB NO: {letter(master_cols['JOB NO'])}\n"
                    f"\n"
                )

                self.third_frame_textbox_1.insert("end", report_text)
                self.third_frame_textbox_1.see("end")

                # Build a dictionary for the master sheet data and formula for quick look-up
                master_dict[master_current] = {}
                for master_row in master_sheet.iter_rows(min_row=master_header_row + 1, max_row=master_sheet.max_row):
                    master_po_num = str(master_row[master_cols["NK SAP PO"] - 1].value).strip()
                    master_po_line = str(master_row[master_cols["PO LINE ITEM"] - 1].value).strip()

                    # self.third_frame_textbox_1.insert("end", f"{master_row[0].row}. master_po_num: {master_po_num}-{master_po_line}, master_fob: {master_row[master_cols['FINAL FOB (Regular sizes)'] - 1].value}.\n\n")
                    if (master_po_num, master_po_line) not in master_dict[master_current]:
                        master_dict[master_current][(master_po_num, master_po_line)] = []

                    master_dict[master_current][(master_po_num, master_po_line)].append(master_row)


                # Initialize ori_date and ori_change at the start of the process
                ori_date = {
                    "Season": {},
                    "FG QTY": {},
                    "Doc Type": {},
                    "SHIP MODE": {},
                    "Plant Code": {},
                    "SHIP-TO": {},
                    "AFS Cat": {},
                    "VAS name": {},
                    "Hanger size": {},
                    "Ratio Qty": {},
                    "Customer PO": {},
                    "Price": {},
                    "Currency": {},
                }
                ori_change = {
                    "Season": {},
                    "FG QTY": {},
                    "Doc Type": {},
                    "SHIP MODE": {},
                    "Plant Code": {},
                    "SHIP-TO": {},
                    "AFS Cat": {},
                    "VAS name": {},
                    "Hanger size": {},
                    "Ratio Qty": {},
                    "Customer PO": {},
                    "Price": {},
                    "Currency": {},
                }
                ori_cm_date = {}

                # List of base keys (without (Date) or (Changes) suffix)
                keys = ["Season", "FG QTY", "Doc Type", "SHIP MODE", "Plant Code",
                        "SHIP-TO", "AFS Cat", "VAS name", "Hanger size", "Ratio Qty",
                        "Customer PO", "Price"]

                # Build a dictionary for the master sheet literal data for quick look-up
                master_dict_data[master_current] = {}
                for master_row_data in master_sheet_data.iter_rows(min_row=master_header_row + 1, max_row=master_sheet.max_row):
                    master_po_num = str(master_row_data[master_cols["NK SAP PO"] - 1].value).strip()
                    master_po_line = str(master_row_data[master_cols["PO LINE ITEM"] - 1].value).strip()
                    master_po_job = str(master_row_data[master_cols["JOB NO"] - 1].value).strip()

                    if f"{master_po_num}{master_po_line}{master_po_job}" not in ori_date["Season"]:
                        # Only capture original values if they have not been captured before
                        for key in keys:
                            date_key = f"{key} (Date)"
                            change_key = f"{key} (Changes)"

                            ori_date[key][f"{master_po_num}{master_po_line}{master_po_job}"] = (
                                str(master_row_data[master_cols[date_key] - 1].value).strip()
                                if not isEmptyCell(master_row_data[master_cols[date_key] - 1].value)
                                else ""
                            )
                            ori_change[key][f"{master_po_num}{master_po_line}{master_po_job}"] = (
                                str(master_row_data[master_cols[change_key] - 1].value).strip()
                                if not isEmptyCell(master_row_data[master_cols[change_key] - 1].value)
                                else ""
                            )

                        ori_cm_date[f"{master_po_num}{master_po_line}{master_po_job}"] = (
                                str(master_row_data[master_cols["Latest CM Change Date"] - 1].value).strip()
                                if not isEmptyCell(master_row_data[master_cols["Latest CM Change Date"] - 1].value)
                                else ""
                            )

                    # self.third_frame_textbox_1.insert("end", f"{master_row_data[0].row}. master_po_num: {master_po_num}-{master_po_line}, master_fob: {master_row_data[master_cols['FINAL FOB (Regular sizes)'] - 1].value}.\n\n")
                    if (master_po_num, master_po_line) not in master_dict_data[master_current]:
                        master_dict_data[master_current][(master_po_num, master_po_line)] = []

                    master_dict_data[master_current][(master_po_num, master_po_line)].append(master_row_data)

                if master_current > total_master:
                    break
                for report_path in self.full_report_file_paths:
                    matching_rows = {
                        "Season": [],
                        "FG QTY": [],
                        "Doc Type": [],
                        "SHIP MODE": [],
                        "Plant Code": [],
                        "SHIP-TO": [],
                        "AFS Cat": [],
                        "VAS name": [],
                        "Hanger size": [],
                        "Ratio Qty": [],
                        "Customer PO": [],
                        "Price": [],
                        "Currency": [],
                        }
                    self.third_frame_textbox_1.insert("end", f"Reading {os.path.basename(report_path)}...\n\n")
                    self.third_frame_textbox_1.see("end")
                    # Load workbook
                    # xlwings_open = xlwings.App(visible=False)
                    # xlwings_load = xlwings_open.books.open(report_path)
                    # xlwings_load.save()
                    # xlwings_load.close()
                    # xlwings_open.quit()

                    report_wb = openpyxl.load_workbook(report_path, read_only=True, data_only=True)
                    report_sheet = report_wb.active

                    report_header_indicator = ["Purchase Order Number"]
                    self.third_frame_textbox_1.insert("end", f"Finding header row...\n\n")
                    self.third_frame_textbox_1.see("end")
                    report_header_row = find_columns_header(report_sheet, report_header_indicator)
                    self.third_frame_textbox_1.insert("end", f"Report header row found at {report_header_row}.\n\n")
                    self.third_frame_textbox_1.see("end")
                    report_cols = find_columns_report(report_sheet, ppm_columns_needed, report_header_row)

                    report_text = (
                        f"PPM Report Column Detection Report:\n"
                        f"Header Row: {report_header_row}\n"
                        f"Purchase Order Number: {col_letter(report_cols.get('Purchase Order Number'))}\n"
                        f"PO Line Item Number: {col_letter(report_cols.get('PO Line Item Number'))}\n"
                        f"Product Code: {col_letter(report_cols.get('Product Code'))}\n"
                        f"Gross Price/FOB currency code: {col_letter(report_cols.get('Gross Price/FOB currency code'))}\n"
                        f"Surcharge Min Mat Main Body currency code: {col_letter(report_cols.get('Surcharge Min Mat Main Body currency code'))}\n"
                        f"Surcharge Min Material Trim currency code: {col_letter(report_cols.get('Surcharge Min Material Trim currency code'))}\n"
                        f"Surcharge Misc currency code: {col_letter(report_cols.get('Surcharge Misc currency code'))}\n"
                        f"Surcharge VAS currency code: {col_letter(report_cols.get('Surcharge VAS currency code'))}\n"
                        f"Gross Price/FOB: {col_letter(report_cols.get('Gross Price/FOB'), -1)}\n"
                        f"Surcharge Min Mat Main Body: {col_letter(report_cols.get('Surcharge Min Mat Main Body'), -1)}\n"
                        f"Surcharge Min Material Trim: {col_letter(report_cols.get('Surcharge Min Material Trim'), -1)}\n"
                        f"Surcharge Misc: {col_letter(report_cols.get('Surcharge Misc'), -1)}\n"
                        f"Surcharge VAS: {col_letter(report_cols.get('Surcharge VAS'), -1)}\n"
                        f"Size Description: {col_letter(report_cols.get('Size Description'))}\n"
                        f"Planning Season Code: {col_letter(report_cols.get('Planning Season Code'))}\n"
                        f"Planning Season Year: {col_letter(report_cols.get('Planning Season Year'))}\n"
                        f"Total Item Quantity: {col_letter(report_cols.get('Total Item Quantity'))}\n"
                        f"Doc Type: {col_letter(report_cols.get('Doc Type'), -1)}\n"
                        f"Mode of Transportation Code: {col_letter(report_cols.get('Mode of Transportation Code'))}\n"
                        f"Plant Code: {col_letter(report_cols.get('Plant Code'))}\n"
                        f"Ship To Customer Number: {col_letter(report_cols.get('Ship To Customer Number'))}\n"
                        f"Inventory Segment Code: {col_letter(report_cols.get('Inventory Segment Code'))}\n"
                        f"VAS name: {col_letter(report_cols.get('VAS name'))}\n"
                        f"Hanger size: {col_letter(report_cols.get('Hanger size'))}\n"
                        f"Ratio quantity: {col_letter(report_cols.get('Ratio quantity'))}\n"
                        f"Customer PO: {col_letter(report_cols.get('Customer PO'))}\n"
                        f"Change Date: {col_letter(report_cols.get('Change Date'))}\n"
                        f"GAC: {col_letter(report_cols.get('GAC'), 1)}\n"
                        f"DPOM Line Item Status: {col_letter(report_cols.get('DPOM Line Item Status'))}\n"
                        f"Document Date: {col_letter(report_cols.get('Document Date'))}\n"
                        f"\n"
                    )

                    self.third_frame_textbox_1.insert("end", report_text)
                    self.third_frame_textbox_1.see("end")

                    # Process each report row and compare with the master dict
                    # for row in report_sheet.iter_rows(min_row=report_header_row + 1, max_row=report_sheet.max_row):
                    for row_index, row in enumerate(report_sheet.iter_rows(min_row=report_header_row + 1, max_row=report_sheet.max_row), start=report_header_row + 1):
                        if report_date_bfr == 0:
                            report_date_bfr = report_change_datetime(row, report_cols).strftime('%m/%d').strip()
                        if row_index == (report_header_row + 1):
                            report_date_now = report_change_datetime(row, report_cols).strftime('%m/%d').strip()

                        total_data += 1
                        report_po_num = str(row[report_cols["Purchase Order Number"] - 1].value).strip()
                        report_po_line = str(row[report_cols["PO Line Item Number"] - 1].value).strip()

                        if isEmptyCell(row[report_cols["Product Code"] - 1].value):
                            report_po_style = "INVALID"
                        else:
                            report_po_style = (row[report_cols["Product Code"] - 1].value).split('-')[0].strip()
                        key = (report_po_num, report_po_line)

                        if key in master_dict[master_current] and report_po_style != "INVALID":
                            existingPO[(report_po_style, report_po_num, report_po_line)] = row
                            # Loop through a copy of newPO's keys to avoid modifying it while iterating
                            for keyPO in list(newPO.keys()):
                                # Check if the first three elements match the desired values
                                if keyPO[:3] == (report_po_style, report_po_num, report_po_line):
                                    # Delete the matching key from newPO
                                    del newPO[keyPO]
                            # Declare numeric values
                            master_rows = master_dict[master_current][key]
                            master_rows_data = master_dict_data[master_current][key]

                            for i, (master_row, master_row_data) in enumerate(zip(master_rows, master_rows_data)):
                                noDiscrepancy = True
                                master_po_fob = safe_float(master_row_data[master_cols["FINAL FOB (Regular sizes)"] - 1].value)
                                master_po_fob_ex_1 = safe_float(master_row_data[master_cols['FINAL FOB (Extended sizes)'][0] - 1].value)
                                master_po_fob_ex_2 = safe_float(master_row_data[master_cols['FINAL FOB (Extended sizes)'][1] - 1].value)

                                report_po_fob =  safe_float(row[report_cols["Gross Price/FOB"] - 2].value) + safe_float(row[report_cols["Surcharge Min Mat Main Body"] - 2].value) + safe_float(row[report_cols["Surcharge Min Material Trim"] - 2].value) + safe_float(row[report_cols["Surcharge Misc"] - 2].value) + safe_float(row[report_cols["Surcharge VAS"] - 2].value)

                                # Assign Ship Mode Designation
                                if str(row[report_cols["Mode of Transportation Code"] -1].value) == "VL":
                                    report_ship_mode = "SEA"
                                elif str(row[report_cols["Mode of Transportation Code"] -1].value) == "AF":
                                    report_ship_mode = "NAF"
                                elif str(row[report_cols["Mode of Transportation Code"] -1].value) == "TR" or str(row[report_cols["Mode of Transportation Code"] -1].value) == "TRUCK":
                                    report_ship_mode = "TR"
                                else:
                                    report_ship_mode = str(row[report_cols["Mode of Transportation Code"] - 1].value)

                                # Assign AFS CAT Designation
                                if str(row[report_cols["Inventory Segment Code"] -1].value) == "1000":
                                    report_afs_cat = "01000"
                                else:
                                    report_afs_cat = str(row[report_cols["Inventory Segment Code"] - 1].value)

                                # Assign Ratio Qty Designation
                                if isEmptyCell(master_row_data[master_cols["Ratio Qty"][0] - 1].value):
                                    master_ratio_qty = "0"
                                else:
                                    master_ratio_qty = str(master_row_data[master_cols["Ratio Qty"][0] - 1].value)

                                # Assign Ratio Qty Designation
                                ratio_col = report_cols.get("Ratio quantity")
                                if ratio_col is None:
                                    report_ratio_qty = None
                                else:
                                    ratio_val = row_value(row, ratio_col)
                                    if isEmptyCell(ratio_val):
                                        report_ratio_qty = "0"
                                    else:
                                        report_ratio_qty = str(ratio_val)

                                # Assign FG QTY Designation
                                total_master_fg_qty = 0

                                # First loop to calculate total FG QTY
                                for master_row_data in master_rows_data:
                                    if isEmptyCell(master_row_data[master_cols["FG QTY"][0] - 1].value):
                                        row_fg_qty = 0
                                    else:
                                        row_fg_qty = safe_float(master_row_data[master_cols["FG QTY"][0] - 1].value)

                                    total_master_fg_qty += row_fg_qty

                                # Convert total to string for comparison
                                master_fg_qty = str(total_master_fg_qty).strip()

                                master_po_fob = f"{master_po_fob:.2f}"
                                master_po_fob_ex_1 = f"{master_po_fob_ex_1:.2f}"
                                master_po_fob_ex_2 = f"{master_po_fob_ex_2:.2f}"
                                report_po_fob = f"{report_po_fob:.2f}"

                                # Season Check
                                # self.third_frame_textbox_1.insert("end", f"Debug 1\n\n")
                                if (row[report_cols["Planning Season Code"] - 1].value != master_row_data[master_cols["Season"][0] - 1].value) or (row[report_cols["Planning Season Year"] - 1].value != master_row_data[master_cols["Season Year"] - 1].value):
                                    noDiscrepancy = False
                                    output_text = f"Season diff found for PO {report_po_num} Line {report_po_line} Size {row[report_cols['Size Description'] - 1].value}:\nPPM Season {row[report_cols['Planning Season Code'] - 1].value} {row[report_cols['Planning Season Year'] - 1].value} vs OCCC Season {master_row_data[master_cols['Season'][0] - 1].value} {master_row_data[master_cols['Season Year'] - 1].value}\n\n"
                                    append_date_changes(master_row, master_cols, report_cols, row, "Season", f"{str(row[report_cols['Planning Season Code'] - 1].value)}{str(row[report_cols['Planning Season Year'] - 1].value)}")
                                else:
                                    keep_date_changes(master_row, master_cols, report_cols, row, "Season", f"{str(row[report_cols['Planning Season Code'] - 1].value)}{str(row[report_cols['Planning Season Year'] - 1].value)}", f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")
                                    matching_rows["Season"].append(i)
                                    if report_path == self.full_report_file_paths[-1] and i == matching_rows["Season"][-1]:
                                        empty_date_changes(master_row, master_cols, report_cols, row, "Season", f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")

                                # FG QTY Check
                                if str(safe_float(row[report_cols["Total Item Quantity"] -1].value)) != str(master_fg_qty):
                                    noDiscrepancy = False
                                    output_text = f"Total Item Quantity diff found for PO {report_po_num} Line {report_po_line} Size {row[report_cols['Size Description'] - 1].value}:\nPPM Total Item Quantity {str(row[report_cols['Total Item Quantity'] -1].value)} vs OCCC FG QTY {str(master_fg_qty)}\n\n"
                                    append_date_changes(master_row, master_cols, report_cols, row, "FG QTY", row[report_cols["Total Item Quantity"] - 1].value)
                                else:
                                    keep_date_changes(master_row, master_cols, report_cols, row, "FG QTY", row[report_cols["Total Item Quantity"] - 1].value, f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")
                                    matching_rows["FG QTY"].append(i)
                                    if report_path == self.full_report_file_paths[-1] and i == matching_rows["FG QTY"][-1]:
                                        empty_date_changes(master_row, master_cols, report_cols, row, "FG QTY", f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")

                                # Doc Type Check
                                if str(row[report_cols["Doc Type"] - 2].value) != str(master_row_data[master_cols["Doc Type"][0] - 1].value):
                                    noDiscrepancy = False
                                    output_text = f"Document Type diff found for PO {report_po_num} Line {report_po_line} Size {row[report_cols['Size Description'] - 1].value}:\nPPM Doc Type {row[report_cols['Doc Type'] - 2].value} vs OCCC Doc Type {master_row_data[master_cols['Doc Type'][0] - 1].value}\n\n"
                                    append_date_changes(master_row, master_cols, report_cols, row, "Doc Type", row[report_cols["Doc Type"] - 2].value)
                                else:
                                    keep_date_changes(master_row, master_cols, report_cols, row, "Doc Type", row[report_cols["Doc Type"] - 2].value, f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")
                                    matching_rows["Doc Type"].append(i)
                                    if report_path == self.full_report_file_paths[-1] and i == matching_rows["Doc Type"][-1]:
                                        empty_date_changes(master_row, master_cols, report_cols, row, "Doc Type", f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")

                                # Ship Mode Check
                                if str(report_ship_mode) != str(master_row_data[master_cols["SHIP MODE"][0] - 1].value):
                                    noDiscrepancy = False
                                    output_text = f"SHIP MODE diff found for PO {report_po_num} Line {report_po_line} Size {row[report_cols['Size Description'] - 1].value}:\nPPM Mode of Transportation Code {report_ship_mode} vs OCCC SHIP MODE {master_row_data[master_cols['SHIP MODE'][0] - 1].value}\n\n"
                                    append_date_changes(master_row, master_cols, report_cols, row, "SHIP MODE", report_ship_mode)
                                else:
                                    keep_date_changes(master_row, master_cols, report_cols, row, "SHIP MODE", report_ship_mode, f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")
                                    matching_rows["SHIP MODE"].append(i)
                                    if report_path == self.full_report_file_paths[-1] and i == matching_rows["SHIP MODE"][-1]:
                                        empty_date_changes(master_row, master_cols, report_cols, row, "SHIP MODE", f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")

                                # Plant Code Check
                                if str(row[report_cols["Plant Code"] - 1].value) != str(master_row_data[master_cols["Plant Code"][0] - 1].value):
                                    noDiscrepancy = False
                                    output_text = f"Plant Code diff found for PO {report_po_num} Line {report_po_line} Size {row[report_cols['Size Description'] - 1].value}:\nPPM Plant Code {row[report_cols['Plant Code'] - 1].value} vs OCCC Plant Code {master_row_data[master_cols['Plant Code'][0] - 1].value}\n\n"
                                    append_date_changes(master_row, master_cols, report_cols, row, "Plant Code", row[report_cols["Plant Code"] - 1].value)
                                else:
                                    keep_date_changes(master_row, master_cols, report_cols, row, "Plant Code", row[report_cols["Plant Code"] - 1].value, f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")
                                    matching_rows["Plant Code"].append(i)
                                    if report_path == self.full_report_file_paths[-1] and i == matching_rows["Plant Code"][-1]:
                                        empty_date_changes(master_row, master_cols, report_cols, row, "Plant Code", f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")

                                # SHIP-TO Check
                                if str(row[report_cols["Ship To Customer Number"] - 1].value) != str(master_row_data[master_cols["SHIP-TO"][0] - 1].value) and not isEmptyCell(row[report_cols["Ship To Customer Number"] - 1].value) and not isEmptyCell(master_row_data[master_cols["SHIP-TO"][0] - 1].value):
                                    noDiscrepancy = False
                                    output_text = f"SHIP-TO diff found for PO {report_po_num} Line {report_po_line} Size {row[report_cols['Size Description'] - 1].value}:\nPPM Ship To Customer Number {row[report_cols['Ship To Customer Number'] - 1].value} vs OCCC SHIP-TO {master_row_data[master_cols['SHIP-TO'][0] - 1].value}\n\n"
                                    append_date_changes(master_row, master_cols, report_cols, row, "SHIP-TO", row[report_cols["Ship To Customer Number"] - 1].value)
                                else:
                                    keep_date_changes(master_row, master_cols, report_cols, row, "SHIP-TO", row[report_cols["Ship To Customer Number"] - 1].value, f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")
                                    matching_rows["SHIP-TO"].append(i)
                                    if report_path == self.full_report_file_paths[-1] and i == matching_rows["SHIP-TO"][-1]:
                                        empty_date_changes(master_row, master_cols, report_cols, row, "SHIP-TO", f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")

                                # AFS CAT Check
                                if str(report_afs_cat) != str(master_row_data[master_cols["AFS Cat"][0] - 1].value):
                                    noDiscrepancy = False
                                    output_text = f"AFS Cat diff found for PO {report_po_num} Line {report_po_line} Size {row[report_cols['Size Description'] - 1].value}:\nPPM Inventory Segment Code {report_afs_cat} vs OCCC AFS Cat {master_row_data[master_cols['AFS Cat'][0] - 1].value}\n\n"
                                    append_date_changes(master_row, master_cols, report_cols, row, "AFS Cat", report_afs_cat)
                                else:
                                    keep_date_changes(master_row, master_cols, report_cols, row, "AFS Cat", report_afs_cat, f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")
                                    matching_rows["AFS Cat"].append(i)
                                    if report_path == self.full_report_file_paths[-1] and i == matching_rows["AFS Cat"][-1]:
                                        empty_date_changes(master_row, master_cols, report_cols, row, "AFS Cat", f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")

                                # VAS Name Check
                                vas_col = report_cols.get("VAS name")
                                if vas_col is not None:
                                    report_vas = row_value(row, vas_col)
                                    if report_vas != master_row_data[master_cols["VAS name"][0] - 1].value:
                                        noDiscrepancy = False
                                        output_text = f"VAS name diff found for PO {report_po_num} Line {report_po_line} Size {row[report_cols['Size Description'] - 1].value}:\nPPM VAS name {report_vas} vs OCCC VAS name {master_row_data[master_cols['VAS name'][0] - 1].value}\n\n"
                                        append_date_changes(master_row, master_cols, report_cols, row, "VAS name", report_vas)
                                    else:
                                        keep_date_changes(master_row, master_cols, report_cols, row, "VAS name", report_vas, f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")
                                        matching_rows["VAS name"].append(i)
                                        if report_path == self.full_report_file_paths[-1] and i == matching_rows["VAS name"][-1]:
                                            empty_date_changes(master_row, master_cols, report_cols, row, "VAS name", f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")

                                # Hanger size Check
                                hanger_col = report_cols.get("Hanger size")
                                if hanger_col is not None:
                                    report_hanger = row_value(row, hanger_col)
                                    if report_hanger != master_row_data[master_cols["Hanger size"][0] - 1].value:
                                        noDiscrepancy = False
                                        output_text = f"Hanger size diff found for PO {report_po_num} Line {report_po_line} Size {row[report_cols['Size Description'] - 1].value}:\nPPM Hanger size {report_hanger} vs OCCC Hanger size {master_row_data[master_cols['Hanger size'][0] - 1].value}\n\n"
                                        append_date_changes(master_row, master_cols, report_cols, row, "Hanger size", report_hanger)
                                    else:
                                        keep_date_changes(master_row, master_cols, report_cols, row, "Hanger size", report_hanger, f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")
                                        matching_rows["Hanger size"].append(i)
                                        if report_path == self.full_report_file_paths[-1] and i == matching_rows["Hanger size"][-1]:
                                            empty_date_changes(master_row, master_cols, report_cols, row, "Hanger size", f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")

                                # Ratio Qty Check
                                if report_ratio_qty is not None:
                                    if report_ratio_qty != master_ratio_qty:
                                        noDiscrepancy = False
                                        output_text = f"Ratio Qty diff found for PO {report_po_num} Line {report_po_line} Size {row[report_cols['Size Description'] - 1].value}:\nPPM Ratio quantity {report_ratio_qty} vs OCCC Ratio Qty {master_ratio_qty}\n\n"
                                        append_date_changes(master_row, master_cols, report_cols, row, "Ratio Qty", report_ratio_qty)
                                    else:
                                        keep_date_changes(master_row, master_cols, report_cols, row, "Ratio Qty", report_ratio_qty, f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")
                                        matching_rows["Ratio Qty"].append(i)
                                        if report_path == self.full_report_file_paths[-1] and i == matching_rows["Ratio Qty"][-1]:
                                            empty_date_changes(master_row, master_cols, report_cols, row, "Ratio Qty", f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")

                                # Customer PO (Deichmann Group only) Check
                                if row[report_cols["Customer PO"] - 1].value != master_row_data[master_cols["Customer PO (Deichmann Group only)"] - 1].value and any(group in str(master_row_data[master_cols["VAS name"][0] - 1].value).lower() for group in ("deichmann", "dechmann")):
                                    noDiscrepancy = False
                                    output_text = f"Customer PO (Deichmann Group only) diff found for PO {report_po_num} Line {report_po_line} Size {row[report_cols['Size Description'] - 1].value}:\nPPM Customer PO {row[report_cols['Customer PO'] - 1].value} vs OCCC Customer PO (Deichmann Group only) {master_row_data[master_cols['Customer PO (Deichmann Group only)'] - 1].value}\n\n"
                                    append_date_changes(master_row, master_cols, report_cols, row, "Customer PO", row[report_cols["Customer PO"] - 1].value)
                                else:
                                    keep_date_changes(master_row, master_cols, report_cols, row, "Customer PO", row[report_cols["Customer PO"] - 1].value, f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")
                                    matching_rows["Customer PO"].append(i)
                                    if report_path == self.full_report_file_paths[-1] and i == matching_rows["Customer PO"][-1]:
                                        empty_date_changes(master_row, master_cols, report_cols, row, "Customer PO", f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")

                                # Currency Check
                                columns_to_check = [
                                    "Gross Price/FOB currency code",
                                    "Surcharge Min Mat Main Body currency code",
                                    "Surcharge Min Material Trim currency code",
                                    "Surcharge Misc currency code",
                                    "Surcharge VAS currency code"
                                ]

                                if any(str(row[report_cols[col] - 1].value).strip() not in ("USD", "None") for col in columns_to_check):
                                    for col in columns_to_check:
                                        if str(row[report_cols[col] - 1].value).strip() not in ("USD", "None"):
                                            noDiscrepancy = False
                                            output_text = f"Currency diff found for PO {report_po_num} Line {report_po_line} Size {row[report_cols['Size Description'] - 1].value}:\nCurrency is {row[report_cols[col] - 1].value}\n\n"
                                            append_date_changes(master_row, master_cols, report_cols, row, "Price", row[report_cols[col] - 1].value)
                                else:
                                    matching_rows["Currency"].append(i)
                                    for col in columns_to_check:
                                        if str(row[report_cols[col] - 1].value).strip() not in ("USD", "None"):
                                            keep_date_changes(master_row, master_cols, report_cols, row, "Price", row[report_cols[col] - 1].value, f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")

                                # Price Check
                                if (report_po_fob != master_po_fob or (report_po_fob != master_po_fob_ex_1 and master_po_fob_ex_1 != 0.00) or (report_po_fob != master_po_fob_ex_2 and master_po_fob_ex_2 != 0.00)):
                                    sizesE = []
                                    sizesE2 = []

                                    # Handle sizesE for master_row[master_cols["Extended Sizes"] - 3]
                                    extended_size_value_1 = str(master_row[master_cols["Extended Sizes"] - 3].value).strip().replace('-', '')
                                    extended_size_value_1 = "T" if "tall" in extended_size_value_1.lower() else extended_size_value_1 # Tall become "T" so it matches with the first T it encounters in sizes list
                                    extended_size_value_1 = re.sub(r'([A-Za-z0-9]+)[\.\(\&\+].*', r'\1', extended_size_value_1) # ex. 3XL[&|.|+](any) becomes 3XL so it matches with any of the size in the sizes list
                                    if not isEmptyCell(extended_size_value_1):
                                        matching_size_1 = next((size for size in sizes if extended_size_value_1 in size), None)
                                        if not isEmptyCell(matching_size_1):
                                            index_1 = sizes.index(matching_size_1)
                                            sizesE = sizes[index_1:]  # All sizes after the matching size

                                    # Handle sizesE2 for master_row[master_cols["Extended Sizes"] - 1]
                                    extended_size_value_2 = str(master_row[master_cols["Extended Sizes"] - 1].value).strip().replace('-', '')
                                    extended_size_value_2 = "T" if "tall" in extended_size_value_2.lower() else extended_size_value_2 # Tall become "T" so it matches with the first T it encounters in sizes list
                                    extended_size_value_2 = re.sub(r'([A-Za-z0-9]+)[\.\(\&\+].*', r'\1', extended_size_value_2) # ex. 3XL[&|.|+](any) becomes 3XL so it matches with any of the size in the sizes list
                                    if not isEmptyCell(extended_size_value_2):
                                        matching_size_2 = next((size for size in sizes if extended_size_value_2 in size), None)
                                        if not isEmptyCell(matching_size_2):
                                            index_2 = sizes.index(matching_size_2)
                                            sizesE2 = sizes[index_2:]  # All sizes after the matching size

                                    if (not str(row[report_cols['Size Description'] - 1].value).strip().replace('-', '') in sizesE and not str(row[report_cols['Size Description'] - 1].value).strip().replace('-', '') in sizesE2 and report_po_fob != master_po_fob) or (extended_size_value_1 != 0.00 and str(row[report_cols['Size Description'] - 1].value).strip().replace('-', '') in sizesE and report_po_fob != master_po_fob_ex_1) or (extended_size_value_2 != 0.00 and str(row[report_cols['Size Description'] - 1].value).strip().replace('-', '') in sizesE2 and report_po_fob != master_po_fob_ex_2):
                                        noDiscrepancy = False
                                        if "/" in str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip():
                                            sizesInMaster = str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip().split('/')
                                            for sizeInMaster in sizesInMaster:
                                                if str(row[report_cols['Size Description'] - 1].value).strip().replace('-', '') == sizeInMaster.strip().split(" ")[0].replace('-', ''):
                                                    sizesInMaster.remove(sizeInMaster)
                                                    break
                                            master_row[master_cols["DPOM - Incorrect FOB"] - 1].value = '/'.join(sizesInMaster)
                                        elif str(row[report_cols['Size Description'] - 1].value).strip().replace('-', '') == str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip().split(" ")[0].replace('-', '') if 'CORRECT' not in str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip() and not is_number(str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip()) and not isNumberAfterDash(str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip()) and not hasComma(str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip()) else str(row[report_cols['Size Description'] - 1].value).strip().replace('-', ''):
                                            master_row[master_cols["DPOM - Incorrect FOB"] - 1].value = f""

                                        output_text = f"FOB diff found for PO {report_po_num} Line {report_po_line} Size {row[report_cols['Size Description'] - 1].value}:\nPPM PO FOB {report_po_fob} vs OCCC FOB {master_po_fob}"
                                        if safe_float(master_po_fob_ex_1) > 0:
                                            output_text += f" and FOB EXT. {master_po_fob_ex_1}"
                                        if safe_float(master_po_fob_ex_2) > 0:
                                            output_text += f" and FOB EXT. EXT. {master_po_fob_ex_2}"
                                        output_text += f"\n\n"
                                        append_date_changes(master_row, master_cols, report_cols, row, "Price", report_po_fob)

                                        # Update / Append DPOM - Incorrect FOB
                                        dpom_value = str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip() if not isEmptyCell(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value) else ""
                                        size_description_value = str(row[report_cols['Size Description'] - 1].value).strip() if not isEmptyCell(row[report_cols['Size Description'] - 1].value) else ""
                                        po_fob_value = str(safe_float(report_po_fob)).strip()
                                        new_dpom_value = (f"{size_description_value} {po_fob_value}").strip()

                                        if not isEmptyCell(dpom_value):
                                            dpom_values = [v.strip() for v in dpom_value.split("/")]
                                            dpom_values = str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip().split('/')
                                            # Remove any value that contains 'CORRECT', starts with a number, or contains a dash
                                            dpom_values_cleaned = [
                                                value.strip() for value in dpom_values
                                                if 'CORRECT' not in value and not is_number(value) and not isNumberAfterDash(value) and not hasComma(value)
                                            ]

                                            cleaned_dpom_value = ' / '.join(dpom_values_cleaned)

                                            if new_dpom_value not in dpom_values_cleaned:
                                                cleaned_dpom_value = f"{cleaned_dpom_value} / {new_dpom_value}".strip(" / ")
                                        else:
                                            cleaned_dpom_value = new_dpom_value

                                        master_row[master_cols["DPOM - Incorrect FOB"] - 1].value = cleaned_dpom_value if cleaned_dpom_value else ""
                                    # elif isEmptyCell(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value) or is_number(str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip()) or isNumberAfterDash(str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip()) or hasComma(str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip()):
                                    #     master_row[master_cols["DPOM - Incorrect FOB"] - 1].value = f"CORRECT"
                                    else:
                                        if "/" in str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip():
                                            sizesInMaster = str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip().split('/')
                                            for sizeInMaster in sizesInMaster:
                                                if str(row[report_cols['Size Description'] - 1].value).strip().replace('-', '') == sizeInMaster.strip().split(" ")[0].replace('-', ''):
                                                    sizesInMaster.remove(sizeInMaster)
                                                    break
                                            master_row[master_cols["DPOM - Incorrect FOB"] - 1].value = '/'.join(sizesInMaster)
                                        elif str(row[report_cols['Size Description'] - 1].value).strip().replace('-', '') == str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip().split(" ")[0].replace('-', '') if 'CORRECT' not in str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip() and not is_number(str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip()) and not isNumberAfterDash(str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip()) and not hasComma(str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip()) else str(row[report_cols['Size Description'] - 1].value).strip().replace('-', ''):
                                            master_row[master_cols["DPOM - Incorrect FOB"] - 1].value = f""


                                        dpom_value = str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip() if not isEmptyCell(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value) else ""
                                        if not isEmptyCell(dpom_value):
                                            dpom_values = [v.strip() for v in dpom_value.split("/")]
                                            dpom_values = str(master_row[master_cols["DPOM - Incorrect FOB"] - 1].value).strip().split('/')
                                            # Remove any value that contains 'CORRECT', starts with a number, or contains a dash
                                            dpom_values_cleaned = [
                                                value.strip() for value in dpom_values
                                                if 'CORRECT' not in value and not is_number(value) and not isNumberAfterDash(value) and not hasComma(value)
                                            ]

                                            cleaned_dpom_value = ' / '.join(dpom_values_cleaned)
                                        else:
                                            cleaned_dpom_value = f"CORRECT"
                                            matching_rows["Price"].append(i)
                                            if report_path == self.full_report_file_paths[-1] and i == matching_rows["Price"][-1] and i == matching_rows["Currency"][-1]:
                                                empty_date_changes(master_row, master_cols, report_cols, row, "Price", f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")

                                        master_row[master_cols["DPOM - Incorrect FOB"] - 1].value = cleaned_dpom_value
                                else:
                                    keep_date_changes(master_row, master_cols, report_cols, row, "Price", report_po_fob, f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")
                                    matching_rows["Price"].append(i)
                                    if report_path == self.full_report_file_paths[-1] and i == matching_rows["Price"][-1] and i == matching_rows["Currency"][-1]:
                                        empty_date_changes(master_row, master_cols, report_cols, row, "Price", f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}")

                                if noDiscrepancy:
                                    master_row[master_cols["Latest CM Change Date"] - 1].value = ori_cm_date[f"{report_po_num}{report_po_line}{master_row_data[master_cols['JOB NO'] - 1].value}"]

                        elif key not in master_dict[master_current]:
                            new_key = (
                                report_po_style,
                                report_po_num,
                                report_po_line,
                                row_value(row, report_cols.get("GAC"), idx_offset=-2),
                                row_value(row, report_cols.get("DPOM Line Item Status")),
                                row_value(row, report_cols.get("Doc Type"), idx_offset=-1),
                                row_value(row, report_cols.get("Document Date")),
                                row_value(row, report_cols.get("Change Date")),
                            )
                            if new_key not in newPO:
                                newPO[new_key] = row

                        if row_index == report_sheet.max_row:
                            report_date_bfr = report_change_datetime(row, report_cols).strftime('%m/%d').strip()

                self.third_frame_textbox_1.insert("end", f"Saving file... please wait\n\n")
                self.third_frame_textbox_1.see("end")
                timestamp = getattr(self, "timestamp", datetime.now().strftime("%Y%m%d-%H%M%S"))
                base_no_ext = os.path.splitext(master_path)[0]
                updated_master_path = f"{base_no_ext}_UPDATED_{timestamp}.xlsx"
                master_wb.save(updated_master_path)
                last_output = updated_master_path
                self.third_frame_textbox_1.insert("end", f"Saved new file at {updated_master_path}\n\n")
                self.third_frame_textbox_1.see("end")

            # Create a new workbook and a new worksheet
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            new_ws.title = "New PO Data"

            # Define your headers based on the keys of newPO
            headers = ["Report PO Style", "Report PO Num", "Report PO Line", "GAC", "DPOM Line Item Status", "Doc Type", "Document Date", "PPM Report Date"]

            # Write headers to the first row
            new_ws.append(headers)

            # Define a style for short date formatting
            short_date_style = openpyxl.styles.NamedStyle(name="short_date_style", number_format="YYYY/MM/DD")

            # Register the NamedStyle with the workbook
            if "short_date_style" not in new_wb.named_styles:
                new_wb.add_named_style(short_date_style)

            # Style the header row with a fill color and bold font
            header_fill = openpyxl.styles.PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            header_font = openpyxl.styles.Font(bold=True)
            header_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )

            for col_num, col_name in enumerate(headers, 1):
                cell = new_ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = header_border

            # List to keep track of items to remove from newPO
            items_to_remove = []

            for new_key in newPO.keys():
                # Extract the first three elements from new_key for comparison
                new_first_three_keys = new_key[:3]

                for existing_key in existingPO.keys():
                    # Extract the first three elements from existing_key for comparison
                    existing_first_three_keys = existing_key[:3]

                    # Compare the first three keys
                    if new_first_three_keys == existing_first_three_keys:
                        items_to_remove.append(new_key)
                        break  # Stop checking other existing keys if a match is found

            # Remove the matched items from newPO
            for key in items_to_remove:
                del newPO[key]

            # Iterate through the dictionary and write rows to the worksheet
            for row_idx, (key, row) in enumerate(newPO.items(), start=2):
                new_ws.append([
                    key[0],  # Report PO Style
                    key[1],  # Report PO Num
                    key[2],  # Report PO Line
                    key[3],  # GAC
                    key[4],  # DPOM Line Item Status
                    key[5],  # Doc Type
                    key[6],  # Document Date
                    key[7],  # PPM Report Date
                ])

                # Apply short date style to the GAC and Document Date columns
                new_ws.cell(row=row_idx, column=headers.index("GAC") + 1).style = short_date_style
                new_ws.cell(row=row_idx, column=headers.index("Document Date") + 1).style = short_date_style
                new_ws.cell(row=row_idx, column=headers.index("PPM Report Date") + 1).style = short_date_style

                # Apply borders to each cell in the row
                for col_num in range(1, len(headers) + 1):
                    cell = new_ws.cell(row=row_idx, column=col_num)
                    cell.border = header_border

            # Auto-adjust the width of each column to fit the content
            for col in new_ws.columns:
                max_length = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                new_ws.column_dimensions[col_letter].width = adjusted_width

            # Save the workbook to a file (timestamped next to the master)
            timestamp = getattr(self, "timestamp", datetime.now().strftime("%Y%m%d-%H%M%S"))
            new_po_path = os.path.join(work_path, f"New_PO_List_{timestamp}.xlsx")
            new_wb.save(new_po_path)
            last_output = new_po_path
            self.third_frame_textbox_1.insert("end", f"Saved new PO list file at {new_po_path}\n\n")
            self.third_frame_textbox_1.see("end")

            self.second_frame_textbox_1.insert("end", f"Compared a total of {total_data} data in a span of {(time.time() - start_time):.2f} seconds")
            self.second_frame_textbox_1.see("end")
            self.third_frame_textbox_1.insert("end", f"Processed {total_data} row(s) in {(time.time() - start_time):.2f} seconds\n\n")
            self.third_frame_textbox_1.see("end")
            self.third_frame_textbox_1.insert("end", f"Finished, you may exit now\n\n")
            self.third_frame_textbox_1.see("end")
            return last_output, total_master or total_inputs, 0


        except Exception as e:
            self.third_frame_textbox_1.insert("end", f"Error: {str(e)}\n\n")
            self.third_frame_textbox_1.see("end")
            return last_output, 0, total_inputs

    def select_frame_by_name(self, name):
        # set button color for selected button
        self.home_button.configure(fg_color=("gray75", "gray25") if name == "home" else "transparent")
        self.frame_2_button.configure(fg_color=("gray75", "gray25") if name == "frame_2" else "transparent")
        self.frame_3_button.configure(fg_color=("gray75", "gray25") if name == "frame_3" else "transparent")

        # show selected frame
        if name == "home":
            self.home_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.home_frame.grid_forget()
        if name == "frame_2":
            self.second_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.second_frame.grid_forget()
        if name == "frame_3":
            self.third_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.third_frame.grid_forget()

    def home_button_event(self):
        self.select_frame_by_name("home")

    def frame_2_button_event(self):
        self.select_frame_by_name("frame_2")

    def frame_3_button_event(self):
        self.select_frame_by_name("frame_3")

    def change_appearance_mode_event(self, new_appearance_mode):
        customtkinter.set_appearance_mode(new_appearance_mode)


class _TextboxProxy:
    def __init__(self, emit: Callable[[str], None]):
        self._emit = emit

    def insert(self, *_args):
        if len(_args) >= 2:
            self._emit(str(_args[1]))

    def see(self, *_args):
        pass

    def configure(self, **_kwargs):
        pass

    def delete(self, *_args, **_kwargs):
        pass


def process_logic(
    master_paths: List[str],
    report_paths: List[str],
    log_emit: Callable[[str], None],
    report_emit: Callable[[str], None],
) -> Tuple[str, int, int]:
    """Run the Memo Match logic with callbacks for Pycro Station."""
    app = App.__new__(App)
    app.full_master_file_path = master_paths
    app.full_report_file_paths = report_paths
    app.third_frame_textbox_1 = _TextboxProxy(log_emit)
    app.second_frame_textbox_1 = _TextboxProxy(report_emit)
    app.timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return App.process_files(app)


class MainWidget(QWidget):
    log_message = Signal(str)
    report_message = Signal(str)
    processing_done = Signal(int, int, str)

    def __init__(self):
        super().__init__()
        self.setObjectName("memo_match_utility_widget")
        self._build_ui()
        self._connect_signals()

    def _build_ui(self):
        self.desc_label = QLabel(
            "Match PPM memo reports against OCCC master files. "
            "Output files are auto-generated next to each master as "
            "'<master>_UPDATED_yyyymmdd-hhmmss.xlsx' and 'New_PO_List_yyyymmdd-hhmmss.xlsx'.",
            self,
        )
        self.desc_label.setWordWrap(True)
        self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.desc_label.setStyleSheet(
            "color: #dcdcdc; background: transparent; padding: 8px; "
            "border: 1px solid #3a3a3a; border-radius: 6px;"
        )
        self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        label_style = "color: #dcdcdc; background: transparent; padding-left: 2px;"
        field_style = (
            "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        self.select_master_btn = PrimaryPushButton("Select Master (OCCC)", self)
        self.select_reports_btn = PrimaryPushButton("Select PPM Report(s)", self)
        self.run_btn = PrimaryPushButton("Analyze & Report", self)

        self.master_files_label = QLabel("Master file(s)", self)
        self.master_files_label.setStyleSheet(label_style)
        self.master_files_box = QTextEdit(self)
        self.master_files_box.setReadOnly(True)
        self.master_files_box.setStyleSheet(field_style)

        self.report_files_label = QLabel("PPM report file(s)", self)
        self.report_files_label.setStyleSheet(label_style)
        self.report_files_box = QTextEdit(self)
        self.report_files_box.setReadOnly(True)
        self.report_files_box.setStyleSheet(field_style)

        self.reports_label = QLabel("Report output", self)
        self.reports_label.setStyleSheet(label_style)
        self.logs_label = QLabel("Process logs", self)
        self.logs_label.setStyleSheet(label_style)

        self.reports_box = QTextEdit(self)
        self.reports_box.setReadOnly(True)
        self.reports_box.setStyleSheet(field_style)

        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setStyleSheet(field_style)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(12, 12, 12, 12)
        main_layout.setSpacing(12)
        main_layout.addWidget(self.desc_label)

        row_buttons = QHBoxLayout()
        row_buttons.addWidget(self.select_master_btn)
        row_buttons.addWidget(self.select_reports_btn)
        main_layout.addLayout(row_buttons)

        row_run = QHBoxLayout()
        row_run.addStretch()
        row_run.addWidget(self.run_btn)
        row_run.addStretch()
        main_layout.addLayout(row_run)

        row_labels = QHBoxLayout()
        row_labels.addWidget(self.master_files_label)
        row_labels.addWidget(self.report_files_label)
        main_layout.addLayout(row_labels)

        row_boxes = QHBoxLayout()
        row_boxes.addWidget(self.master_files_box)
        row_boxes.addWidget(self.report_files_box)
        main_layout.addLayout(row_boxes, 1)

        row_log_labels = QHBoxLayout()
        row_log_labels.addWidget(self.reports_label)
        row_log_labels.addWidget(self.logs_label)
        main_layout.addLayout(row_log_labels)

        row_logs = QHBoxLayout()
        row_logs.addWidget(self.reports_box)
        row_logs.addWidget(self.log_box)
        main_layout.addLayout(row_logs, 2)

    def _connect_signals(self):
        self.select_master_btn.clicked.connect(self.select_master_files)
        self.select_reports_btn.clicked.connect(self.select_report_files)
        self.run_btn.clicked.connect(self.run_process)
        self.log_message.connect(self.append_log)
        self.report_message.connect(self.append_report)
        self.processing_done.connect(self.on_processing_done)

    def select_master_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "Select Master (OCCC) Files", "", "Excel Files (*.xlsx *.xlsm *.xls)"
        )
        if files:
            self.master_files_box.setPlainText("\n".join(files))
        else:
            self.master_files_box.clear()

    def select_report_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "Select PPM Report Files", "", "Excel Files (*.xlsx *.xlsm *.xls)"
        )
        if files:
            files = sorted(files, key=lambda x: os.path.basename(x))
            self.report_files_box.setPlainText("\n".join(files))
        else:
            self.report_files_box.clear()

    def get_files_from_box(self, text_box: QTextEdit) -> List[str]:
        text = text_box.toPlainText().strip()
        return [line.strip() for line in text.split("\n") if line.strip()]

    def run_process(self):
        master_files = self.get_files_from_box(self.master_files_box)
        report_files = self.get_files_from_box(self.report_files_box)

        if not master_files:
            MessageBox("Warning", "Please select OCCC master file(s).", self).exec()
            return
        if not report_files:
            MessageBox("Warning", "Please select at least one PPM report file.", self).exec()
            return

        self.log_box.clear()
        self.reports_box.clear()
        self.log_message.emit("Process Started...")

        self.run_btn.setEnabled(False)
        self.select_master_btn.setEnabled(False)
        self.select_reports_btn.setEnabled(False)

        def worker():
            try:
                last_file, ok, fail = process_logic(
                    master_files,
                    report_files,
                    self.log_message.emit,
                    self.report_message.emit,
                )
                self.processing_done.emit(ok, fail, last_file)
            except Exception as e:
                self.log_message.emit(f"CRITICAL ERROR: {e}")
                self.processing_done.emit(0, len(master_files), "")

        threading.Thread(target=worker, daemon=True).start()

    def append_log(self, text: str):
        self.log_box.append(text)
        self.log_box.ensureCursorVisible()

    def append_report(self, text: str):
        self.reports_box.append(text)
        self.reports_box.ensureCursorVisible()

    def on_processing_done(self, ok: int, fail: int, last_file: str):
        self.log_message.emit(f"Done. Success: {ok}, Failed: {fail}")
        error_text = ""
        if last_file and last_file.startswith("ERROR:"):
            error_text = last_file
            last_file = ""
            self.log_message.emit(error_text)
        elif last_file:
            self.log_message.emit(f"Last output: {last_file}")

        self.run_btn.setEnabled(True)
        self.select_master_btn.setEnabled(True)
        self.select_reports_btn.setEnabled(True)

        title = "Process complete" if fail == 0 else "Process finished with issues"
        lines = [f"Success: {ok}", f"Failed: {fail}"]
        if error_text:
            lines.append(error_text)
        if last_file:
            lines.append(f"Last output: {last_file}")
        msg = MessageBox(title, "\n".join(lines), self)
        msg.yesButton.setText("OK")
        msg.cancelButton.hide()
        msg.exec()


def get_widget():
    return MainWidget()
