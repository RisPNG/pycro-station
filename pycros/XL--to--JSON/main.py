import os
import json
import threading
from datetime import datetime
from pathlib import Path
from typing import List

from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QVBoxLayout,
    QLabel,
    QTextEdit,
    QWidget,
    QSizePolicy
)
from qfluentwidgets import PrimaryPushButton, MessageBox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    import xlrd
except ImportError:  # pragma: no cover
    xlrd = None


def _xlrd_cell_value(book, cell):
    """Map xlrd cell to Python types that json can encode."""
    if xlrd is None:  # pragma: no cover
        return None

    from xlrd import (
        XL_CELL_DATE,
        XL_CELL_BOOLEAN,
        XL_CELL_EMPTY,
        XL_CELL_BLANK,
        XL_CELL_ERROR,
        error_text_from_code,
        xldate_as_datetime,
    )

    if cell is None:
        return None

    if cell.ctype in (XL_CELL_EMPTY, XL_CELL_BLANK):
        return None

    if cell.ctype == XL_CELL_DATE:
        try:
            return xldate_as_datetime(cell.value, book.datemode)
        except Exception:
            return cell.value

    if cell.ctype == XL_CELL_BOOLEAN:
        return bool(cell.value)

    if cell.ctype == XL_CELL_ERROR:
        return error_text_from_code.get(cell.value, cell.value)

    return cell.value


def _extract_detailed_data_xls(file_path: str, log_emit):
    """
    Extract detailed cell data from legacy .xls files using xlrd.

    Notes:
    - xlrd does not expose formula text for cells; `formula` will be `None`.
    - We still mark merged cells via `is_merged`.
    """
    if xlrd is None:
        raise RuntimeError(
            "xlrd is required to read .xls files. Click 'Install Requirements' for this pycro (or install xlrd)."
        )

    log_emit("Detected legacy .xls format; extracting values via xlrd (formula text not available for .xls).")

    book = xlrd.open_workbook(file_path, formatting_info=True)
    output_data = {}

    for sheet_name in book.sheet_names():
        sh = book.sheet_by_name(sheet_name)
        sheet_data = {}

        # --- MERGED CELL LOGIC START ---
        # xlrd provides merged_cells as tuples of (row_lo, row_hi, col_lo, col_hi),
        # where hi bounds are exclusive and indexes are 0-based.
        merged_lookup = {}
        max_row = sh.nrows
        max_col = sh.ncols
        for row_lo, row_hi, col_lo, col_hi in getattr(sh, "merged_cells", []) or []:
            max_row = max(max_row, row_hi)
            max_col = max(max_col, col_hi)
            master = (row_lo + 1, col_lo + 1)  # 1-based
            for r0 in range(row_lo, row_hi):
                for c0 in range(col_lo, col_hi):
                    merged_lookup[(r0 + 1, c0 + 1)] = master
        # --- MERGED CELL LOGIC END ---

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                coord = f"{get_column_letter(c)}{r}"

                if (r, c) in merged_lookup:
                    mr, mc = merged_lookup[(r, c)]
                    try:
                        master_cell = sh.cell(mr - 1, mc - 1)
                    except IndexError:
                        master_cell = None
                    val = _xlrd_cell_value(book, master_cell)
                    is_merged = True
                else:
                    try:
                        cell = sh.cell(r - 1, c - 1)
                    except IndexError:
                        cell = None
                    val = _xlrd_cell_value(book, cell)
                    is_merged = False

                # xlrd doesn't provide formula text for cells, so we store None.
                has_formula = False
                formula_str = None

                if val is not None or has_formula or is_merged:
                    sheet_data[coord] = {
                        "value": val,
                        "formula": formula_str,
                        "type": "formula" if has_formula else "static",
                        "is_merged": is_merged,
                    }

        output_data[sheet_name] = sheet_data

    return output_data


def _extract_detailed_data_openpyxl(file_path: str):
    """Extract detailed cell data from .xlsx/.xlsm files using openpyxl."""
    # 1. Load workbook for VALUES
    wb_values = load_workbook(file_path, data_only=True)
    # 2. Load workbook for FORMULAS
    wb_formulas = load_workbook(file_path, data_only=False)

    output_data = {}

    for sheet_name in wb_values.sheetnames:
        sheet_values = wb_values[sheet_name]
        sheet_formulas = wb_formulas[sheet_name]

        sheet_data = {}

        # --- MERGED CELL LOGIC START ---
        # Create a lookup dictionary: {(row, col): (master_row, master_col)}
        # This maps every cell in a merged range to its top-left "master" cell.
        merged_lookup = {}

        # sheet_values.merged_cells.ranges gives a list of all merged ranges (e.g., "A1:B2")
        for merge_range in sheet_values.merged_cells.ranges:
            min_col, min_row, max_col, max_row = (
                merge_range.min_col,
                merge_range.min_row,
                merge_range.max_col,
                merge_range.max_row,
            )
            master_coord = (min_row, min_col)

            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    merged_lookup[(r, c)] = master_coord
        # --- MERGED CELL LOGIC END ---

        for row in sheet_values.iter_rows():
            for cell in row:
                coord = cell.coordinate
                r, c = cell.row, cell.column

                # Check if this cell is part of a merge
                if (r, c) in merged_lookup:
                    # It is merged. Get the master cell coordinates
                    mr, mc = merged_lookup[(r, c)]

                    # Retrieve value from the MASTER cell in the value-workbook
                    val = sheet_values.cell(row=mr, column=mc).value

                    # Retrieve formula from the MASTER cell in the formula-workbook
                    raw_formula_cell = sheet_formulas.cell(row=mr, column=mc)
                    formula = raw_formula_cell.value

                    is_merged = True
                else:
                    # Standard cell
                    val = cell.value

                    # Direct lookup for formula since coordinates match
                    raw_formula_cell = sheet_formulas[coord]
                    formula = raw_formula_cell.value

                    is_merged = False

                # Formula Detection Logic
                has_formula = False
                formula_str = None

                if isinstance(formula, str) and formula.startswith("="):
                    has_formula = True
                    formula_str = formula

                # Decide whether to save the cell
                # We save if it has a value, has a formula, OR is part of a merge (even if visually empty, it implies structure)
                if val is not None or has_formula or is_merged:
                    sheet_data[coord] = {
                        "value": val,
                        "formula": formula_str,
                        "type": "formula" if has_formula else "static",
                        "is_merged": is_merged,
                    }

        output_data[sheet_name] = sheet_data

    return output_data


def excel_to_detailed_json(file_path, output_json_path, log_emit):
    """Convert Excel file to JSON with formulas, values, and merged cell handling."""
    try:
        log_emit(f"Processing: {os.path.basename(file_path)}")

        ext = Path(file_path).suffix.lower()
        if ext == ".xls":
            output_data = _extract_detailed_data_xls(file_path, log_emit)
        else:
            output_data = _extract_detailed_data_openpyxl(file_path)

        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, indent=4, default=str)

        log_emit(f"Saved: {os.path.basename(output_json_path)}")
        return True

    except Exception as e:
        import traceback
        traceback.print_exc() # Print full error to console for debugging
        log_emit(f"Error processing {os.path.basename(file_path)}: {e}")
        return False

def process_files(files: List[str], log_emit) -> tuple[str, int, int]:
    """Process multiple Excel files and convert them to JSON."""
    success_count = 0
    fail_count = 0
    last_output = ""

    for file_path in files:
        try:
            # Generate output filename with timestamp in the same folder
            file_dir = os.path.dirname(file_path)
            file_name = os.path.splitext(os.path.basename(file_path))[0]
            timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            output_name = f"{file_name}_{timestamp}.json"
            output_path = os.path.join(file_dir, output_name)

            # Convert Excel to JSON
            if excel_to_detailed_json(file_path, output_path, log_emit):
                success_count += 1
                last_output = output_path
            else:
                fail_count += 1

        except Exception as e:
            log_emit(f"Failed to process {os.path.basename(file_path)}: {e}")
            fail_count += 1

    return last_output, success_count, fail_count


class MainWidget(QWidget):
    log_message = Signal(str)
    processing_done = Signal(int, int, str)

    def __init__(self):
        super().__init__()
        self.setObjectName("xl_to_json_widget")
        self._build_ui()
        self._connect_signals()

    def _build_ui(self):
        # Description label
        self.desc_label = QLabel("", self)
        self.desc_label.setWordWrap(True)
        self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.desc_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.desc_label.setStyleSheet(
            "color: #dcdcdc; background: transparent; padding: 6px; "
            "border: 1px solid #3a3a3a; border-radius: 6px;"
        )
        self.set_long_description("")

        # Buttons
        self.select_btn = PrimaryPushButton("Select Excel Files", self)
        self.run_btn = PrimaryPushButton("Convert to JSON", self)

        # Labels
        self.files_label = QLabel("Selected files", self)
        self.files_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.files_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        self.logs_label = QLabel("Process logs", self)
        self.logs_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.logs_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        # Text boxes
        shared_style = (
            "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        self.files_box = QTextEdit(self)
        self.files_box.setReadOnly(True)
        self.files_box.setPlaceholderText("Selected Excel files will appear here")
        self.files_box.setStyleSheet(shared_style)

        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText("Live process log will appear here")
        self.log_box.setStyleSheet(shared_style)

        # Layout construction
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)

        # Row 0: Description
        main_layout.addWidget(self.desc_label, 1)

        # Row 1: Select button
        row1_layout = QHBoxLayout()
        row1_layout.addStretch(1)
        row1_layout.addWidget(self.select_btn, 1)
        row1_layout.addStretch(1)
        main_layout.addLayout(row1_layout, 0)

        # Row 2: Run button
        row2_layout = QHBoxLayout()
        row2_layout.addStretch(1)
        row2_layout.addWidget(self.run_btn, 1)
        row2_layout.addStretch(1)
        main_layout.addLayout(row2_layout, 0)

        # Row 3: Labels for files and logs
        row3_layout = QHBoxLayout()
        row3_layout.addWidget(self.files_label, 1)
        row3_layout.addWidget(self.logs_label, 1)
        main_layout.addLayout(row3_layout, 0)

        # Row 4: Files and logs text boxes
        row4_layout = QHBoxLayout()
        row4_layout.addWidget(self.files_box, 1)
        row4_layout.addWidget(self.log_box, 1)
        main_layout.addLayout(row4_layout, 4)

    def set_long_description(self, text: str):
        """Set description from description.md."""
        clean = (text or "").strip()
        if clean:
            self.desc_label.setText(clean)
            self.desc_label.show()
        else:
            self.desc_label.clear()
            self.desc_label.hide()

    def _connect_signals(self):
        """Connect UI signals to handlers."""
        self.select_btn.clicked.connect(self.select_files)
        self.run_btn.clicked.connect(self.run_process)
        self.log_message.connect(self.append_log)
        self.processing_done.connect(self.on_processing_done)

    def select_files(self):
        """Open file dialog to select Excel files."""
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select Excel Files",
            "",
            "Excel Files (*.xlsx *.xlsm *.xls)"
        )
        if files:
            self.files_box.setPlainText("\n".join(files))
        else:
            self.files_box.clear()

    def _selected_files(self) -> List[str]:
        """Get list of selected files from text box."""
        text = self.files_box.toPlainText().strip()
        if not text:
            return []
        return [line.strip() for line in text.split("\n") if line.strip()]

    def run_process(self):
        """Start the conversion process in a background thread."""
        files = self._selected_files()
        if not files:
            MessageBox("Warning", "Please select Excel files to convert.", self).exec()
            return

        self.log_box.clear()
        self.log_message.emit("Conversion process started...")
        self.log_message.emit(f"Files to process: {len(files)}")
        self.log_message.emit("Output files will be saved in the same folder with timestamp suffix.")
        self.log_message.emit("")

        self.run_btn.setEnabled(False)
        self.select_btn.setEnabled(False)

        def worker():
            try:
                last_file, ok, fail = process_files(files, self.log_message.emit)
                self.processing_done.emit(ok, fail, last_file)
            except Exception as e:
                self.log_message.emit(f"CRITICAL ERROR: {e}")
                self.processing_done.emit(0, 0, "")

        threading.Thread(target=worker, daemon=True).start()

    def append_log(self, text: str):
        """Append text to log box."""
        self.log_box.append(text)
        self.log_box.ensureCursorVisible()

    def on_processing_done(self, ok: int, fail: int, last_file: str):
        """Handle completion of processing."""
        self.log_message.emit("")
        self.log_message.emit(f"Conversion complete: {ok} succeeded, {fail} failed")

        if last_file:
            self.log_message.emit(f"Last output: {last_file}")

        self.run_btn.setEnabled(True)
        self.select_btn.setEnabled(True)

        # Show completion dialog
        title = "Conversion complete" if fail == 0 else "Conversion finished with issues"
        lines = [f"Success: {ok}", f"Failed: {fail}"]
        if last_file:
            lines.append(f"Last file: {os.path.basename(last_file)}")

        msg = MessageBox(title, "\n".join(lines), self)
        msg.yesButton.setText("OK")
        msg.cancelButton.hide()
        msg.exec()


def get_widget():
    """Return the main widget for the pycro station."""
    return MainWidget()
