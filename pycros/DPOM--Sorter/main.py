import os
import threading
from datetime import datetime, date, timedelta
from typing import List, Tuple, Any, Optional, Dict
from collections import defaultdict, OrderedDict
from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QVBoxLayout,
    QLabel,
    QTextEdit,
    QWidget,
)
from qfluentwidgets import PrimaryPushButton, MessageBox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

# =============================================================================
# CONSTANTS & COLOR DEFINITIONS
# =============================================================================
# Excel Colors matching VBA constants
COLOR_WHITE = "FFFFFF"
COLOR_BLACK = "FF000000"
COLOR_CYAN = "00FFFF"
COLOR_PINK = "C3C3C3"  # cPink = 12830955
COLOR_YELLOW = "FFFEFEB8"
COLOR_GREEN = "92D050"  # cGreen = 8252325
COLOR_DARK_GREEN = "506E20"  # cDGreen = 5287936
COLOR_MAGENTA = "800080"  # cDMagenta = 8388736

# Destination countries that require separating each Material (incl. colorway)
SPLIT_MATERIAL_BY_DEST_COUNTRY = {"CANADA", "INDONESIA", "MEXICO"}

# Column indices in raw data (0-based)
COL_VENDOR = 0
COL_SEASON = 1
COL_YEAR = 2
COL_STYLE = 3
COL_SILHOUETTE = 4
COL_CATEGORY = 5
COL_PO = 6
COL_TRADING_PO = 7
COL_PO_LINE = 8
COL_DOC_DATE = 9
COL_OGAC = 10
COL_PUR_GROUP_CODE = 11
COL_PUR_GROUP_NAME = 12
COL_PLANT = 13
COL_DOC_TYPE = 14
COL_DOC_TYPE_DESC = 15
COL_TRANS = 16
COL_SHIP_NO = 17
COL_SHIP_NAME = 18
COL_COUNTRY = 19
COL_AFS = 20
COL_CAT_DESC = 21
COL_SUB_CAT = 22
COL_SIZE_DESC = 23
COL_SIZE_QTY = 24
COL_TOTAL_QTY = 25
COL_FOB = 26

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================
def normalize_size(size_val: Any) -> str:
    """Standardizes size strings for sorting/comparison - matches VBA NormalizeSize"""
    if size_val is None:
        return ""
    s = str(size_val).strip().upper().replace("-", "")

    # Replace from most specific to least
    replacements = [
        ("XXXXXL", "6XL"), ("XXXXL", "5XL"), ("XXXL", "3XL"), ("XXL", "2XL"),
        ("XXXXXS", "6XS"), ("XXXXS", "5XS"), ("XXXS", "4XS"), ("XXXS", "3XS"), ("XXS", "2XS")
    ]
    for old, new in replacements:
        s = s.replace(old, new)
    return s


def format_date_val(val) -> str:
    """Formats date objects to MM/DD/YYYY string"""
    if isinstance(val, (datetime, date)):
        return val.strftime("%m/%d/%Y")
    if val and isinstance(val, str):
        try:
            # Try parsing DD/MM/YYYY format from input
            if "/" in val:
                parts = val.split("/")
                if len(parts) == 3:
                    dt = datetime(int(parts[2]), int(parts[1]), int(parts[0]))
                    return dt.strftime("%m/%d/%Y")
        except:
            pass
    return str(val) if val else ""


def parse_ogac_date(val) -> Optional[datetime]:
    """Parse OGAC date from Excel dates or DD/MM/YYYY strings."""
    if not val:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    try:
        if "/" in str(val):
            parts = str(val).split("/")
            if len(parts) == 3:
                # VBA treats parts[0] as month and parts[1] as day
                # even though input is DD/MM/YYYY
                # This creates incorrect but consistent sorting
                day = int(parts[0])
                month = int(parts[1])
                year = int(parts[2])
                # For output display, we need MM/DD/YYYY format
                # So we swap day/month to get the "correct" date for display
                return datetime(year, month, day)
    except:
        pass
    return None


def normalize_country(val: Any) -> str:
    """Normalize destination country for comparisons"""
    if val is None:
        return ""
    return str(val).strip().upper()


def format_afs_category(val: Any, width: int = 5) -> str:
    """Format AFS Category as a zero-padded numeric string (default width=5)."""
    if val is None:
        return ""

    if isinstance(val, bool):
        return str(val)

    try:
        if isinstance(val, (int, float)):
            return str(int(val)).zfill(width)

        s = str(val).strip()
        if not s:
            return ""

        if s.isdigit():
            return s.zfill(width)

        f = float(s)
        i = int(f)
        if abs(f - i) < 1e-9:
            return str(i).zfill(width)
    except Exception:
        pass

    return str(val).strip()


def format_ship_to_customer_number(val: Any) -> str:
    """Format Ship To Customer Number; uses '#' when empty (matches Excel output)."""
    if val is None:
        return "#"
    s = str(val).strip()
    return s if s else "#"


def ensure_unique_path(path: str) -> str:
    """Return a non-existing path by appending ' (n)' before the extension if needed."""
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    n = 1
    while True:
        candidate = f"{base} ({n}){ext}"
        if not os.path.exists(candidate):
            return candidate
        n += 1


def year_sort_value(val: Any) -> int:
    """Best-effort year parsing for sorting (handles numbers/strings like '2025'/'2025.0')."""
    if val in (None, ""):
        return 0
    if isinstance(val, bool):
        return 0
    try:
        if isinstance(val, (int, float)):
            return int(val)
        s = str(val).strip()
        return int(float(s)) if s else 0
    except Exception:
        return 0


# =============================================================================
# MAIN PROCESSING LOGIC
# =============================================================================
class ProcessingLogic:
    def __init__(self, log_signal):
        self.log = log_signal
        self.blind_buy_map = {}  # Map full style code -> BBJ
        self.size_order = []     # List of normalized sizes in order

    def load_references(self, file_path):
        """Load Size order and Blind Buy mappings from reference Excel"""
        self.log.emit(f"Loading reference file: {os.path.basename(file_path)}")
        try:
            wb = load_workbook(file_path, data_only=True)

            # Load Size Order
            if "Size" in wb.sheetnames:
                ws_size = wb["Size"]
                for row in ws_size.iter_rows(min_row=2, max_col=1, values_only=True):
                    if row[0]:
                        self.size_order.append(normalize_size(row[0]))

            # Load Blind Buy mappings
            if "Blind Buy" in wb.sheetnames:
                ws_bb = wb["Blind Buy"]
                for row in ws_bb.iter_rows(min_row=2, max_col=2, values_only=True):
                    if row[0]:
                        sc = str(row[0]).strip().upper()
                        bbj = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                        self.blind_buy_map[sc] = bbj

            wb.close()
            self.log.emit(f"  Loaded {len(self.size_order)} sizes and {len(self.blind_buy_map)} blind buy entries")
        except Exception as e:
            raise Exception(f"Error loading references: {str(e)}")

    def process_dpom_file(self, input_path, output_path):
        """Main processing function - replicates VBA SortRecord + FirstPass + SecondPass + AddNewColumn"""
        self.log.emit(f"Processing: {os.path.basename(input_path)}")

        # Step 1: Load raw data
        wb = load_workbook(input_path, data_only=True)
        ws = wb.active

        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        data_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                data_rows.append(list(row))

        if not data_rows:
            self.log.emit("  Empty file, skipping.")
            return

        # Step 2: Pivot data (group by all columns except size-specific ones)
        grouped_data = self._pivot_data(data_rows)

        # Step 3: Get all unique sizes and sort them
        all_sizes = set()
        for info in grouped_data.values():
            all_sizes.update(info['sizes'].keys())

        sorted_sizes = sorted(list(all_sizes),
                              key=lambda x: self.size_order.index(x) if x in self.size_order else 99999)

        # Step 4: Create sorted pivoted rows
        pivoted_rows = []
        for key, info in grouped_data.items():
            raw = info['raw']

            # Parse OGAC date for sorting - replicate VBA's logic exactly
            # VBA treats DD/MM/YYYY as MM/DD/YYYY, creating sort key YYYYMMDD
            ogac_dt = parse_ogac_date(raw[COL_OGAC])
            if ogac_dt:
                # VBA's sort key: takes first part as MM, second as DD
                # Input 01/10/2025 (1 Oct) → treated as month=01, day=10 → sort key "20250110"
                # Input 24/09/2025 (24 Sep) → treated as month=24, day=09 → sort key "20252409"
                # We need to create the same wrong sort key for consistency
                ogac_raw = str(raw[COL_OGAC])
                if "/" in ogac_raw:
                    parts = ogac_raw.split("/")
                    if len(parts) == 3:
                        mm = parts[0].zfill(2)  # First part treated as month
                        dd = parts[1].zfill(2)  # Second part treated as day
                        yy = parts[2]
                        ogac_str = f"{yy}{mm}{dd}"  # Wrong sort key matching VBA
                    else:
                        ogac_str = "99999999"
                else:
                    ogac_str = "99999999"
            else:
                ogac_str = "99999999"

            # Get season rank
            seas = str(raw[COL_SEASON]).upper()
            season_rank = 9
            if "SP" in seas:
                season_rank = 1
            elif "SU" in seas:
                season_rank = 2
            elif "FA" in seas:
                season_rank = 3
            elif "HO" in seas:
                season_rank = 4

            # Extract style components
            style_full = str(raw[COL_STYLE])
            style_head = style_full[:6] if len(style_full) >= 6 else style_full
            style_cw = style_full[-3:] if "-" in style_full else ""

            obj = {
                'raw': raw,
                'sizes': info['sizes'],
                'total_qty': info['total_qty'],
                'fob': info['fob'],
                'fob_by_size': info.get('fob_by_size', {}),
                'style_head': style_head,
                'style_cw': style_cw,
                'style_full': style_full,
                'year': raw[COL_YEAR],
                'season_rank': season_rank,
                'ogac_dt': ogac_dt,
                'ogac_str': ogac_str,
                'po': str(raw[COL_PO]),
                'po_line': str(raw[COL_PO_LINE]),
                'country': str(raw[COL_COUNTRY]),
                'ship_no': format_ship_to_customer_number(raw[COL_SHIP_NO]),
                'afs': format_afs_category(raw[COL_AFS])
            }
            pivoted_rows.append(obj)

        # Step 5: Sort records (Year + Planning Season first)
        pivoted_rows.sort(key=lambda x: (
            year_sort_value(x.get('year')),
            x['season_rank'],
            x['style_head'],
            x['ogac_dt'] or datetime.max,
            x['po'],
            x['style_cw'],
            x['po_line'],
            x['country'],
            x['ship_no']
        ))

        # Step 6: Build final output structure with correct grouping
        final_rows = self._build_output_structure(pivoted_rows, sorted_sizes)

        # Step 7: Write to Excel with formatting
        self._write_excel_output(final_rows, sorted_sizes, output_path, data_rows)

        wb.close()
        self.log.emit(f"  Completed: {os.path.basename(output_path)}")

    def _pivot_data(self, data_rows):
        """Pivot data by grouping rows with same attributes (except size columns)"""
        grouped = {}

        for row in data_rows:
            # Create key from all columns except size-specific ones.
            # NOTE: FOB can vary by size; it must NOT create a separate group.
            key_list = list(row)
            key_list[COL_SIZE_DESC] = None
            key_list[COL_SIZE_QTY] = None
            key_list[COL_TOTAL_QTY] = None
            key_list[COL_FOB] = None
            key = tuple(key_list)

            if key not in grouped:
                grouped[key] = {
                    'raw': row,
                    'sizes': {},
                    'fob_by_size': {},
                    'total_qty': 0,
                    'fob': row[COL_FOB]
                }

            sz = normalize_size(row[COL_SIZE_DESC])
            qty = row[COL_SIZE_QTY] if isinstance(row[COL_SIZE_QTY], (int, float)) else 0

            prev_qty = grouped[key]['sizes'].get(sz, 0)
            new_qty = prev_qty + qty
            grouped[key]['sizes'][sz] = new_qty
            grouped[key]['total_qty'] += qty

            # Track FOB per size (weighted average if same size appears multiple times).
            if qty:
                fob_val = row[COL_FOB]
                try:
                    fob = float(fob_val) if fob_val not in (None, "") else 0.0
                except Exception:
                    fob = 0.0

                prev_fob = grouped[key]['fob_by_size'].get(sz)
                if prev_fob is None or prev_qty <= 0:
                    grouped[key]['fob_by_size'][sz] = fob
                else:
                    grouped[key]['fob_by_size'][sz] = ((prev_fob * prev_qty) + (fob * qty)) / new_qty

        return grouped

    def _build_output_structure(self, pivoted_rows, sorted_sizes):
        """Build final output structure - each PO gets its own Total OGAC Qty (matching Total PO Qty)"""
        final_rows = []

        # Group by Style (for selected countries, keep full Material incl. colorway)
        style_groups = defaultdict(list)
        style_keep_colorway = {}
        for row in pivoted_rows:
            country_norm = normalize_country(row.get('country'))
            keep_colorway = country_norm in SPLIT_MATERIAL_BY_DEST_COUNTRY
            style_key = row['style_full'] if keep_colorway else row['style_head']
            style_groups[style_key].append(row)
            style_keep_colorway[style_key] = style_keep_colorway.get(style_key, False) or keep_colorway

        def style_group_sort_key(k: str) -> tuple:
            """Preserve SortRecord ordering across style groups (esp. split materials)."""
            items = style_groups.get(k, [])
            if not items:
                return (0, 9, "", datetime.max, "", "", "", "", "", k)

            first = items[0]
            return (
                year_sort_value(first.get('year')),
                first.get('season_rank', 9),
                first.get('style_head', ""),
                first.get('ogac_dt') or datetime.max,
                first.get('po', ""),
                first.get('style_cw', ""),
                first.get('po_line', ""),
                first.get('country', ""),
                first.get('ship_no', ""),
                k,
            )

        for style_key in sorted(style_groups.keys(), key=style_group_sort_key):
            style_items = style_groups[style_key]
            keep_colorway = style_keep_colorway.get(style_key, False)

            # Accumulate style totals
            style_totals = defaultdict(int)
            style_money = 0

            # Group by Planning Year + Planning Season + OGAC date within style.
            # NOTE: We preserve the existing SortRecord ordering (year -> season -> OGAC),
            #       so we must not re-sort OGAC groups purely by OGAC date here.
            ogac_groups = defaultdict(list)
            for item in style_items:
                ogac_dt = item.get('ogac_dt')
                ogac_date = ogac_dt.date() if isinstance(ogac_dt, datetime) else None
                ogac_group_key = (item.get('year') or 0, item.get('season_rank') or 9, ogac_date)
                ogac_groups[ogac_group_key].append(item)

            ogac_keys = list(ogac_groups.keys())
            ogac_none_keys = [k for k in ogac_keys if k[2] is None]
            if ogac_none_keys:
                ogac_keys = [k for k in ogac_keys if k[2] is not None] + ogac_none_keys

            for ogac_group_key in ogac_keys:
                ogac_items = ogac_groups[ogac_group_key]

                # Accumulate OGAC totals (across POs under this style+OGAC)
                ogac_totals = defaultdict(int)

                # Group by PO within OGAC (each PO is independent)
                po_groups = defaultdict(list)
                split_ship_to_idx = 0
                for item in ogac_items:
                    ship_to = item.get('ship_no', '#')
                    if ship_to != "#":
                        # If Ship To has an explicit value (not '#'), do not group it with any other line
                        po_key = (item['country'], item['po'], item['afs'], ship_to, split_ship_to_idx)
                        split_ship_to_idx += 1
                    else:
                        po_key = (item['country'], item['po'], item['afs'], ship_to)
                    po_groups[po_key].append(item)

                for _, po_items in po_groups.items():
                    # Check for blind buy
                    is_blind_buy = False
                    bb_job = ""
                    for item in po_items:
                        if item['style_full'].upper() in self.blind_buy_map:
                            is_blind_buy = True
                            bb_job = self.blind_buy_map[item['style_full'].upper()]
                            break

                    # Accumulate PO totals
                    po_totals = defaultdict(int)

                    # Add individual item rows
                    for item in po_items:
                        final_rows.append({
                            'type': 'ITEM',
                            'data': item,
                            'label': 'Total Item Qty',
                            'sizes': item['sizes'],
                            'total_qty': item['total_qty'],
                            'keep_colorway': keep_colorway,
                        })

                        for sz, qty in item.get('sizes', {}).items():
                            if not qty:
                                continue
                            po_totals[sz] += qty

                    top_fob_val = po_items[0].get('fob', 0)
                    try:
                        top_fob = float(top_fob_val) if top_fob_val not in (None, "") else 0.0
                    except Exception:
                        top_fob = 0.0

                    po_fob_by_size = {}
                    po_money = top_fob * sum(po_totals.values())

                    # Total PO Qty row
                    final_rows.append({
                        'type': 'TOTAL_PO',
                        'data': po_items[0],
                        'sizes': po_totals,
                        'total_qty': sum(po_totals.values()),
                        'blind_buy': is_blind_buy,
                        'bb_job': bb_job,
                        'keep_colorway': keep_colorway,
                    })

                    # Net Unit Price row (PO level)
                    final_rows.append({
                        'type': 'MONEY_PO',
                        'label': 'Net Unit Price',
                        'data': po_items[0],
                        'fob': top_fob,
                        'fob_by_size': po_fob_by_size,
                        'sizes': po_totals,
                        'total_money': po_money,
                        'keep_colorway': keep_colorway,
                    })

                    # Trading Co Net Unit Price row (PO level)
                    final_rows.append({
                        'type': 'MONEY_PO',
                        'label': 'Trading Co Net Unit Price',
                        'data': po_items[0],
                        'fob': top_fob,
                        'fob_by_size': po_fob_by_size,
                        'sizes': po_totals,
                        'total_money': po_money,
                        'keep_colorway': keep_colorway,
                    })

                    # Black separator after PO money rows
                    final_rows.append({'type': 'SEP_BLACK'})

                    # Roll up to OGAC + Style level
                    for sz, qty in po_totals.items():
                        ogac_totals[sz] += qty
                        style_totals[sz] += qty
                    style_money += po_money

                # Total OGAC Qty row (per OGAC date)
                final_rows.append({
                    'type': 'TOTAL_OGAC',
                    'data': ogac_items[0],
                    'sizes': ogac_totals,
                    'total_qty': sum(ogac_totals.values()),
                    'keep_colorway': keep_colorway,
                })

                # Pink separator after OGAC
                final_rows.append({'type': 'SEP_PINK'})

            # Total Style Qty row
            final_rows.append({
                'type': 'TOTAL_STYLE',
                'data': style_items[0],
                'sizes': style_totals,
                'total_qty': sum(style_totals.values()),
                'keep_colorway': keep_colorway,
            })

            # Net Unit Price row (Style level)
            final_rows.append({
                'type': 'MONEY_STYLE',
                'label': 'Net Unit Price',
                'data': style_items[0],
                'total_money': style_money,
                'keep_colorway': keep_colorway,
            })

            # Trading Co Net Unit Price row (Style level)
            final_rows.append({
                'type': 'MONEY_STYLE',
                'label': 'Trading Co Net Unit Price',
                'data': style_items[0],
                'total_money': style_money,
                'keep_colorway': keep_colorway,
            })

            # Dark green separator after Style
            final_rows.append({'type': 'SEP_DARK_GREEN'})

        return final_rows

    def _write_excel_output(self, final_rows, sorted_sizes, output_path, raw_data):
        """Write final output to Excel with proper formatting"""
        wb = Workbook()
        ws = wb.active

        # Define fills
        fill_white = PatternFill("solid", fgColor=COLOR_WHITE)
        fill_yellow = PatternFill("solid", fgColor=COLOR_YELLOW)
        fill_pink = PatternFill("solid", fgColor=COLOR_PINK)
        fill_green = PatternFill("solid", fgColor=COLOR_GREEN)
        fill_dark_green = PatternFill("solid", fgColor=COLOR_DARK_GREEN)
        fill_black = PatternFill("solid", fgColor=COLOR_BLACK)

        font_bold = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # Row 1: PROCESSED marker
        ws['A1'] = "PROCESSED"

        # Build headers - matching VBA output exactly
        out_headers = [
            "Vendor", "Planning Season", "Year", "Material",
            "Job Number", "Product Type", "VNFOB", "Destination", "Blind Buy Job #",
            "Silhouette Description", "Global Category Core Focus Description",
            "PO Number", "Trading Co PO Number", "PO Line Item Number",
            "Document Date", "Estimate BusWeekDate", "OGAC Date",
            "Purchase Group Code", "Purchase Group Name", "Plant",
            "Buy Group", "Doc Type Description", "Mode",
            "",  # Empty column 24 (used for labels in data)
            "Ship To Customer Number",
            "",  # Empty column 26
            "Customer Name",
            "Destination Country", "AFS Category", "Category Description",
            "Sub Category Size Value"
        ]
        out_headers.extend(sorted_sizes)
        out_headers.append("TOTAL")
        out_headers.append("")  # Trailing empty column 46

        # Row 2: Overall Result header AND grand total
        ws.cell(2, len(out_headers) - 1, "Overall Result")

        # Row 3: Column headers
        for c, h in enumerate(out_headers, 1):
            ws.cell(3, c, h)

        # Row 2 only has "Overall Result" text, no grand total calculation
        # (VBA doesn't calculate grand total)

        # Write data rows starting at row 4
        curr_row = 4
        GIPT = 10  # Days to subtract for Estimate BusWeekDate

        for row_obj in final_rows:
            row_type = row_obj['type']

            # Handle separators
            if row_type.startswith('SEP'):
                fill = fill_black
                if row_type == 'SEP_PINK':
                    fill = fill_pink
                elif row_type == 'SEP_DARK_GREEN':
                    fill = fill_dark_green

                for c in range(1, len(out_headers) + 1):
                    ws.cell(curr_row, c).fill = fill
                ws.row_dimensions[curr_row].height = 5
                curr_row += 1
                continue

            # Get data
            raw = row_obj['data']['raw']

            # Determine style code (strip colorway for total rows)
            style_val = raw[COL_STYLE]
            strip_colorway = (
                row_type in ['TOTAL_PO', 'MONEY_PO']
                or (
                    row_type in ['TOTAL_STYLE', 'TOTAL_OGAC', 'MONEY_STYLE']
                    and not row_obj.get('keep_colorway')
                )
            )
            if strip_colorway and "-" in str(style_val):
                style_val = str(style_val).split("-")[0]

            # Column 1-3: Vendor, Season, Year
            ws.cell(curr_row, 1, raw[COL_VENDOR])
            ws.cell(curr_row, 2, raw[COL_SEASON])
            ws.cell(curr_row, 3, raw[COL_YEAR])

            # Column 4: Material (Style)
            ws.cell(curr_row, 4, style_val)

            # Column 5: Job Number (empty for now)

            # Column 6: Product Type (empty)

            # Column 7: VNFOB
            if style_val:
                ws.cell(curr_row, 7, "N")

            # Column 8: Destination (empty)

            # Column 9: Blind Buy Job #
            if row_type == 'TOTAL_PO' and row_obj.get('blind_buy'):
                ws.cell(curr_row, 9, row_obj.get('bb_job', ''))

            # Column 10-11: Silhouette, Category
            ws.cell(curr_row, 10, raw[COL_SILHOUETTE])
            ws.cell(curr_row, 11, raw[COL_CATEGORY])

            # Column 12-14: PO Number, Trading Co PO, PO Line
            ws.cell(curr_row, 12, raw[COL_PO])
            tpo = raw[COL_TRADING_PO]
            if not tpo and style_val:
                tpo = "-"
            ws.cell(curr_row, 13, tpo)
            ws.cell(curr_row, 14, raw[COL_PO_LINE])

            # Column 15: Document Date
            ws.cell(curr_row, 15, format_date_val(raw[COL_DOC_DATE]))

            # Column 16: Estimate BusWeekDate (calculated from OGAC)
            ogac_dt = row_obj['data'].get('ogac_dt')
            if ogac_dt:
                calc_dt = ogac_dt - timedelta(days=GIPT)
                # Get previous Monday
                bus_dt = calc_dt - timedelta(days=calc_dt.weekday())
                ws.cell(curr_row, 16, bus_dt.strftime("%m/%d/%Y"))

            # Column 17: OGAC Date
            ws.cell(curr_row, 17, format_date_val(raw[COL_OGAC]))

            # Column 18-20: Purchase Group Code, Name, Plant
            ws.cell(curr_row, 18, raw[COL_PUR_GROUP_CODE])
            ws.cell(curr_row, 19, raw[COL_PUR_GROUP_NAME])
            ws.cell(curr_row, 20, raw[COL_PLANT])

            # Column 21-23: Buy Group, Doc Type Desc, Mode
            ws.cell(curr_row, 21, raw[COL_DOC_TYPE])
            ws.cell(curr_row, 22, raw[COL_DOC_TYPE_DESC])
            ws.cell(curr_row, 23, raw[COL_TRANS])

            # Column 24: Label column (Total Item Qty, Total PO Qty, etc.)
            if row_type == 'ITEM':
                ws.cell(curr_row, 24, "Total Item Qty")
            elif row_type == 'TOTAL_PO':
                ws.cell(curr_row, 24, "Total PO Qty")
            elif row_type == 'TOTAL_OGAC':
                ws.cell(curr_row, 24, "Total OGAC Qty")
            elif row_type == 'TOTAL_STYLE':
                ws.cell(curr_row, 24, "Total Style Qty")
            elif row_type == 'MONEY_PO' or row_type == 'MONEY_STYLE':
                # Money rows show the label
                ws.cell(curr_row, 24, row_obj.get('label', ''))

            # Column 25: Ship To Customer Number
            if row_type in ['TOTAL_PO', 'TOTAL_OGAC', 'TOTAL_STYLE']:
                # Total rows: empty
                ws.cell(curr_row, 25, "")
            elif row_type == 'ITEM' or row_type == 'MONEY_PO' or row_type == 'MONEY_STYLE':
                # Item and money rows: show customer number or "#"
                ws.cell(curr_row, 25, format_ship_to_customer_number(raw[COL_SHIP_NO]))

            # Column 26: Empty column

            # Column 27: Customer Name
            if row_type in ['TOTAL_PO', 'TOTAL_OGAC', 'TOTAL_STYLE']:
                # Total rows: empty
                ws.cell(curr_row, 27, "")
            elif row_type == 'ITEM' or row_type == 'MONEY_PO' or row_type == 'MONEY_STYLE':
                # Item and money rows: show customer name or "#"
                cn = raw[COL_SHIP_NAME]
                if not cn:
                    cn = "#"
                ws.cell(curr_row, 27, cn)

            # Column 28-31: Country, AFS, Category Desc, Sub Category
            if row_type in ['TOTAL_PO', 'TOTAL_OGAC', 'TOTAL_STYLE']:
                # Total rows (except Item rows): All location/category fields empty
                ws.cell(curr_row, 28, "")
                ws.cell(curr_row, 29, "")
                ws.cell(curr_row, 30, "")
                ws.cell(curr_row, 31, "")
            else:
                # Item and Money rows: Show full info
                ws.cell(curr_row, 28, raw[COL_COUNTRY])
                # AFS Category: Keep leading zeros for Item rows, strip for others
                afs_val = format_afs_category(raw[COL_AFS])
                ws.cell(curr_row, 29, afs_val)
                ws.cell(curr_row, 30, raw[COL_CAT_DESC])
                ws.cell(curr_row, 31, raw[COL_SUB_CAT])

            # Size columns start at column 32
            col_off = 32
            if row_type == 'MONEY_PO':
                # For money rows at PO level: show FOB in each size column that had qty
                fob_by_size = row_obj.get('fob_by_size', {}) or {}
                default_fob = row_obj.get('fob', 0)
                for i, sz in enumerate(sorted_sizes):
                    qty = row_obj.get('sizes', {}).get(sz, 0)
                    if qty > 0:
                        ws.cell(curr_row, col_off + i, fob_by_size.get(sz, default_fob))
            elif row_type == 'MONEY_STYLE':
                # For money rows at Style level: leave size columns empty
                pass
            elif row_type == 'ITEM':
                # For item rows: show quantities from the item's sizes
                sizes_dict = row_obj.get('sizes', {})
                for i, sz in enumerate(sorted_sizes):
                    val = sizes_dict.get(sz, 0)
                    if val > 0:
                        ws.cell(curr_row, col_off + i, val)
            else:
                # For total rows: show quantities
                sizes_dict = row_obj.get('sizes', {})
                for i, sz in enumerate(sorted_sizes):
                    val = sizes_dict.get(sz, 0)
                    if val > 0:
                        ws.cell(curr_row, col_off + i, val)

            # TOTAL column (column 45 - second-to-last)
            total_col_idx = len(out_headers) - 1
            if row_type in ['ITEM', 'TOTAL_PO', 'TOTAL_OGAC', 'TOTAL_STYLE']:
                ws.cell(curr_row, total_col_idx, row_obj.get('total_qty', 0))
            # Money values go in the trailing empty column (column 46)
            elif row_type in ['MONEY_PO', 'MONEY_STYLE']:
                ws.cell(curr_row, len(out_headers), row_obj.get('total_money', 0))

            # Apply colors and borders
            fill = None
            if row_type in ['TOTAL_PO', 'MONEY_PO']:
                fill = fill_yellow
            elif row_type == 'TOTAL_OGAC':
                fill = fill_pink
            elif row_type in ['TOTAL_STYLE', 'MONEY_STYLE']:
                fill = fill_green

            if fill:
                for c in range(1, len(out_headers) + 1):
                    cell = ws.cell(curr_row, c)
                    cell.fill = fill
                    cell.font = font_bold

            # Add border to TOTAL column
            ws.cell(curr_row, total_col_idx).border = thin_border

            curr_row += 1

        # Add final summary row (VBA line 401-527)
        # This row sums all "Total Item Qty" rows
        final_row = curr_row
        ws.cell(final_row, 1, "Overall Result")
        ws.cell(final_row, 24, "Total Item Qty")

        # Calculate grand totals for each size column
        size_grand_totals = defaultdict(int)
        grand_total_qty = 0
        for row_obj in final_rows:
            if row_obj['type'] == 'ITEM':
                for sz in sorted_sizes:
                    qty = row_obj.get('sizes', {}).get(sz, 0)
                    size_grand_totals[sz] += qty
                grand_total_qty += row_obj.get('total_qty', 0)

        # Write size grand totals
        col_off = 32
        for i, sz in enumerate(sorted_sizes):
            val = size_grand_totals.get(sz, 0)
            if val > 0:
                ws.cell(final_row, col_off + i, val)

        # Write total grand total
        ws.cell(final_row, len(out_headers) - 1, grand_total_qty)

        curr_row += 1

        # Auto-fit columns
        for col in range(1, len(out_headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12

        # Save
        wb.save(output_path)
        wb.close()


# =============================================================================
# GUI
# =============================================================================
class MainWidget(QWidget):
    log_message = Signal(str)
    processing_done = Signal(int, int, str)

    def __init__(self):
        super().__init__()
        self.setObjectName("dpom_sorter_widget")
        self._build_ui()
        self._connect_signals()
        self.logic = ProcessingLogic(self.log_message)

    def _build_ui(self):
        self.desc_label = QLabel("DPOM Sorter (Python - VBA Exact Replication)", self)
        self.desc_label.setStyleSheet("color: #dcdcdc; font-weight: bold; padding: 5px;")

        self.select_btn2 = PrimaryPushButton("Select Reference Excel (Size/Blind Buy)", self)
        self.select_btn = PrimaryPushButton("Select Raw DPOM Files", self)
        self.run_btn = PrimaryPushButton("Run Process", self)
        self.run_btn.setEnabled(False)

        self.files_box2 = QTextEdit(self)
        self.files_box2.setReadOnly(True)
        self.files_box2.setMaximumHeight(40)
        self.files_box2.setPlaceholderText("Reference file path...")
        self.files_box2.setStyleSheet("background: #1f1f1f; color: #d0d0d0; border: 1px solid #3a3a3a; border-radius: 4px;")

        self.files_box = QTextEdit(self)
        self.files_box.setReadOnly(True)
        self.files_box.setPlaceholderText("Input files...")
        self.files_box.setStyleSheet("background: #1f1f1f; color: #d0d0d0; border: 1px solid #3a3a3a; border-radius: 4px;")

        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setStyleSheet("background: #1f1f1f; color: #d0d0d0; border: 1px solid #3a3a3a; border-radius: 4px;")

        layout = QVBoxLayout(self)
        layout.addWidget(self.desc_label)

        h1 = QHBoxLayout()
        h1.addWidget(self.select_btn2)
        h1.addWidget(self.files_box2)
        layout.addLayout(h1)

        layout.addWidget(self.select_btn)
        layout.addWidget(self.files_box)
        layout.addWidget(self.run_btn)
        layout.addWidget(self.log_box)

    def _connect_signals(self):
        self.select_btn2.clicked.connect(self.select_ref_file)
        self.select_btn.clicked.connect(self.select_input_files)
        self.run_btn.clicked.connect(self.run_process)
        self.log_message.connect(self.append_log)
        self.processing_done.connect(self.on_done)

    def select_ref_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Reference Excel", "", "Excel Files (*.xlsx *.xlsm)")
        if path:
            self.files_box2.setText(path)
            self.check_ready()

    def select_input_files(self):
        paths, _ = QFileDialog.getOpenFileNames(self, "Select DPOM Files", "", "Excel/CSV Files (*.xlsx *.csv)")
        if paths:
            self.files_box.setText("\n".join(paths))
            self.check_ready()

    def check_ready(self):
        if self.files_box2.toPlainText() and self.files_box.toPlainText():
            self.run_btn.setEnabled(True)

    def run_process(self):
        ref_file = self.files_box2.toPlainText()
        inputs = self.files_box.toPlainText().split('\n')

        self.run_btn.setEnabled(False)
        self.log_box.clear()
        self.log_message.emit("Starting process...")

        def worker():
            try:
                self.logic.load_references(ref_file)
                success_count = 0
                run_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                for inp in inputs:
                    if not inp.strip():
                        continue
                    out = f"{os.path.splitext(inp)[0]}_processed_{run_timestamp}.xlsx"
                    out = ensure_unique_path(out)
                    self.logic.process_dpom_file(inp, out)
                    success_count += 1
                self.processing_done.emit(success_count, 0, f"Processing Complete!\n{success_count} file(s) processed successfully.")
            except Exception as e:
                self.log_message.emit(f"Critical Error: {str(e)}")
                self.processing_done.emit(0, 1, f"Failed: {str(e)}")

        threading.Thread(target=worker, daemon=True).start()

    def append_log(self, text):
        self.log_box.append(text)

    def on_done(self, ok, fail, msg):
        self.run_btn.setEnabled(True)
        MessageBox("Status", msg, self).exec()


def get_widget():
    return MainWidget()


if __name__ == "__main__":
    from PySide6.QtWidgets import QApplication
    import sys
    app = QApplication(sys.argv)
    w = MainWidget()
    w.show()
    sys.exit(app.exec())
