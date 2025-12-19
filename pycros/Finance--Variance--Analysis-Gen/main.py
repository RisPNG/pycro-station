#!/usr/bin/env python3
from __future__ import annotations

import re
import threading
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

import xlrd
from openpyxl import load_workbook

# GUI Imports (for Pycro Station)
from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QVBoxLayout,
    QLabel,
    QTextEdit,
    QWidget,
    QSizePolicy,
)
from qfluentwidgets import PrimaryPushButton, MessageBox, ComboBox


# ===== Unified Workbook Wrapper for .xls and .xlsx =====

class UnifiedWorksheet:
    """Wrapper to provide a unified interface for xlrd and openpyxl worksheets."""

    def __init__(self, ws, is_xlrd: bool = False):
        self._ws = ws
        self._is_xlrd = is_xlrd

    @property
    def max_row(self) -> int:
        if self._is_xlrd:
            return self._ws.nrows
        return self._ws.max_row or 0

    @property
    def max_column(self) -> int:
        if self._is_xlrd:
            return self._ws.ncols
        return self._ws.max_column or 0

    def cell_value(self, row: int, col: int) -> Any:
        """Get cell value. Row and col are 1-indexed (like openpyxl)."""
        if self._is_xlrd:
            # xlrd uses 0-indexed
            try:
                return self._ws.cell_value(row - 1, col - 1)
            except IndexError:
                return None
        else:
            return self._ws.cell(row=row, column=col).value

    def cell_displayed_value(self, row: int, col: int) -> str:
        """Get displayed value as string. Row and col are 1-indexed."""
        val = self.cell_value(row, col)
        if val is None:
            return ""
        return str(val)


class UnifiedWorkbook:
    """Wrapper to provide a unified interface for xlrd and openpyxl workbooks."""

    def __init__(self, wb, is_xlrd: bool = False):
        self._wb = wb
        self._is_xlrd = is_xlrd

    @property
    def sheetnames(self) -> List[str]:
        if self._is_xlrd:
            return self._wb.sheet_names()
        return self._wb.sheetnames

    def __contains__(self, name: str) -> bool:
        return name in self.sheetnames

    def __getitem__(self, name: str) -> UnifiedWorksheet:
        if self._is_xlrd:
            return UnifiedWorksheet(self._wb.sheet_by_name(name), is_xlrd=True)
        return UnifiedWorksheet(self._wb[name], is_xlrd=False)

    def get_sheet(self, name: str) -> Optional[UnifiedWorksheet]:
        if name not in self.sheetnames:
            return None
        return self[name]


def load_workbook_unified(filepath: str, data_only: bool = True) -> UnifiedWorkbook:
    """Load a workbook using the appropriate library based on file extension."""
    ext = Path(filepath).suffix.lower()
    if ext == ".xls":
        wb = xlrd.open_workbook(filepath)
        return UnifiedWorkbook(wb, is_xlrd=True)
    else:
        wb = load_workbook(filepath, data_only=data_only)
        return UnifiedWorkbook(wb, is_xlrd=False)


# ===== Constants =====
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_FULL = {
    "Jan": "JAN", "Feb": "FEB", "Mar": "MAR", "Apr": "APR",
    "May": "MAY", "Jun": "JUN", "Jul": "JUL", "Aug": "AUG",
    "Sep": "SEP", "Oct": "OCT", "Nov": "NOV", "Dec": "DEC"
}
MONTH_TO_NUM = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12
}

# Items to exclude for Higher General Administrative Expenses
GAE_EXCLUDE = ["Bonus", "Performance Incentives", "Charity & Donations", "Write-Off Fixed Asset", "Office Rental"]


def _emit(log_emit, text: str):
    if callable(log_emit):
        try:
            log_emit(text)
            return
        except Exception:
            pass
    print(text)


def get_cell_displayed_value(cell) -> str:
    """Get the displayed value of an openpyxl cell as string."""
    if cell.value is None:
        return ""
    return str(cell.value)


def get_cell_raw_value(cell) -> Any:
    """Get the raw value of an openpyxl cell (formula or value)."""
    return cell.value


def parse_short_date(s: str) -> Tuple[str, int]:
    """Parse short date like Nov'25 -> ('NOV', 2025)."""
    match = re.match(r"([A-Za-z]{3})'?(\d{2})", s.strip())
    if match:
        month_abbr = match.group(1).capitalize()
        year_short = int(match.group(2))
        year_full = 2000 + year_short if year_short < 100 else year_short
        return MONTH_FULL.get(month_abbr, month_abbr.upper()), year_full
    return "", 0


def get_fy_for_month_year(month: str, year: int) -> int:
    """
    Get fiscal year for a given month and calendar year.
    FY runs from May to April. So:
    - May 2025 to Apr 2026 = FY26
    """
    month_num = MONTH_TO_NUM.get(month.capitalize(), 0)
    if month_num >= 5:  # May-Dec: FY is next calendar year
        return year + 1
    else:  # Jan-Apr: FY is current calendar year
        return year


def get_quarter_months(quarter: int, fy: int) -> List[Tuple[str, int]]:
    """
    Get the months for a quarter in a fiscal year.
    FY26: Q1=MAY-JUL 2025, Q2=AUG-OCT 2025, Q3=NOV-DEC 2025+JAN 2026, Q4=FEB-APR 2026
    """
    base_year = fy - 1  # FY26 starts in calendar 2025

    if quarter == 1:
        return [("MAY", base_year), ("JUN", base_year), ("JUL", base_year)]
    elif quarter == 2:
        return [("AUG", base_year), ("SEP", base_year), ("OCT", base_year)]
    elif quarter == 3:
        return [("NOV", base_year), ("DEC", base_year), ("JAN", fy)]
    elif quarter == 4:
        return [("FEB", fy), ("MAR", fy), ("APR", fy)]
    return []


def get_ytd_months_count(selected_month: str, selected_year: int) -> int:
    """
    Get the number of months from start of FY to the selected month.
    FY starts in May.
    """
    month_num = MONTH_TO_NUM.get(selected_month.capitalize(), 0)
    if month_num >= 5:  # May-Dec
        return month_num - 4  # May=1, Jun=2, ..., Dec=8
    else:  # Jan-Apr
        return month_num + 8  # Jan=9, Feb=10, Mar=11, Apr=12


def build_ytd_header(months_count: int, month: str, year: int) -> str:
    """Build YTD header like '07 MONTHS TO NOV 2025'."""
    return f"{months_count:02d} MONTHS TO {MONTH_FULL.get(month.capitalize(), month.upper())} {year}"


def parse_analytical_bracket(bracket_content: str, selected_month: str, selected_year: int) -> Dict:
    """
    Parse the bracket content from Analytical Review header.
    Returns dict with comparison info.
    """
    result = {
        "type": None,
        "left": [],   # List of (month, year) or quarter info
        "right": [],
        "is_ytd": False,
        "left_header": None,
        "right_header": None,
    }

    content = bracket_content.strip()

    # Check for YTD pattern: (YTD FY26 Vs YTD FY25) or (YTD FY26 Vs YTD FY24)
    ytd_match = re.match(r"YTD\s+FY(\d{2})\s+Vs\s+YTD\s+FY(\d{2})", content, re.IGNORECASE)
    if ytd_match:
        result["type"] = "ytd"
        result["is_ytd"] = True
        fy_left = 2000 + int(ytd_match.group(1))
        fy_right = 2000 + int(ytd_match.group(2))

        months_count = get_ytd_months_count(selected_month, selected_year)

        # For left (current), use selected month/year
        result["left_header"] = build_ytd_header(months_count, selected_month, selected_year)

        # For right, calculate the year difference
        year_diff = fy_left - fy_right
        right_year = selected_year - year_diff
        result["right_header"] = build_ytd_header(months_count, selected_month, right_year)

        return result

    # Check for Quarter pattern: (Q2-FY26 Vs Q2-FY25) or (Q2-FY26 Vs Q1-FY26)
    quarter_match = re.match(r"Q(\d)-FY(\d{2})\s+Vs\s+Q(\d)-FY(\d{2})", content, re.IGNORECASE)
    if quarter_match:
        result["type"] = "quarter"
        q_left = int(quarter_match.group(1))
        fy_left = 2000 + int(quarter_match.group(2))
        q_right = int(quarter_match.group(3))
        fy_right = 2000 + int(quarter_match.group(4))

        result["left"] = get_quarter_months(q_left, fy_left)
        result["right"] = get_quarter_months(q_right, fy_right)
        return result

    # Check for combined months pattern: (Sep'25 & Aug'25 Vs Sep'24 & Aug'24)
    combined_match = re.match(
        r"([A-Za-z]{3})'?(\d{2})\s*&\s*([A-Za-z]{3})'?(\d{2})\s+Vs\s+([A-Za-z]{3})'?(\d{2})\s*&\s*([A-Za-z]{3})'?(\d{2})",
        content, re.IGNORECASE
    )
    if combined_match:
        result["type"] = "combined"
        m1, y1 = parse_short_date(f"{combined_match.group(1)}'{combined_match.group(2)}")
        m2, y2 = parse_short_date(f"{combined_match.group(3)}'{combined_match.group(4)}")
        m3, y3 = parse_short_date(f"{combined_match.group(5)}'{combined_match.group(6)}")
        m4, y4 = parse_short_date(f"{combined_match.group(7)}'{combined_match.group(8)}")

        result["left"] = [(m1, y1), (m2, y2)]
        result["right"] = [(m3, y3), (m4, y4)]
        return result

    # Check for simple month comparison: (Nov'25 Vs Oct'25)
    simple_match = re.match(r"([A-Za-z]{3})'?(\d{2})\s+Vs\s+([A-Za-z]{3})'?(\d{2})", content, re.IGNORECASE)
    if simple_match:
        result["type"] = "month"
        m1, y1 = parse_short_date(f"{simple_match.group(1)}'{simple_match.group(2)}")
        m2, y2 = parse_short_date(f"{simple_match.group(3)}'{simple_match.group(4)}")

        result["left"] = [(m1, y1)]
        result["right"] = [(m2, y2)]
        return result

    return result


def build_month_header(month: str, year: int) -> str:
    """Build month header like 'NOV 2025'."""
    return f"{month} {year}"


def load_ma_workbooks(ma_files: List[str], log_emit=None) -> List[Tuple[str, UnifiedWorkbook]]:
    """Load all MA workbooks and return list of (filename, workbook)."""
    workbooks = []
    for fpath in ma_files:
        try:
            wb = load_workbook_unified(fpath, data_only=True)
            workbooks.append((Path(fpath).name, wb))
            _emit(log_emit, f"[OK] Loaded MA file: {Path(fpath).name}")
        except Exception as e:
            _emit(log_emit, f"[ERROR] Failed to load {fpath}: {e}")
    return workbooks


def get_values_for_production_overhead(ma_workbooks: List[Tuple[str, UnifiedWorkbook]],
                                        left_headers: List[str], right_headers: List[str],
                                        search_cols: List[str] = ["G", "H"],
                                        log_emit=None) -> List[Tuple[str, float, float]]:
    """
    Get production overhead values from Page 5 and Page 6.
    Returns list of (description, left_sum, right_sum).
    """
    from openpyxl.utils import column_index_from_string

    results = {}  # description -> [left_values, right_values]

    for sheet_name in ["Page 5", "Page 6"]:
        for fname, wb in ma_workbooks:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]

            # Find column for each header
            for is_left, headers in [(True, left_headers), (False, right_headers)]:
                for header in headers:
                    col_idx = None
                    header_row = None

                    for col_letter in search_cols:
                        c_idx = column_index_from_string(col_letter)
                        for row in range(1, min(20, ws.max_row + 1)):
                            cell_val = ws.cell_displayed_value(row, c_idx).strip().upper()
                            if header.upper() in cell_val:
                                col_idx = c_idx
                                header_row = row
                                break
                        if col_idx:
                            break

                    if not col_idx:
                        continue

                    # Find PRODUCTION OVERHEAD section
                    po_start_row = None
                    for row in range(header_row + 1, ws.max_row + 1):
                        col_a_val = ws.cell_displayed_value(row, 1).strip().upper()
                        if "PRODUCTION OVERHEAD" in col_a_val:
                            po_start_row = row
                            break

                    if not po_start_row:
                        continue

                    # Get values under PRODUCTION OVERHEAD
                    for row in range(po_start_row + 1, ws.max_row + 1):
                        col_a_val = ws.cell_displayed_value(row, 1).strip()
                        cell_val = ws.cell_value(row, col_idx)

                        # Stop at next major section
                        if col_a_val and not col_a_val.startswith(" ") and "TOTAL" in col_a_val.upper():
                            break

                        if cell_val is not None and cell_val != "" and col_a_val:
                            try:
                                numeric_val = float(cell_val) if not isinstance(cell_val, (int, float)) else cell_val
                                key = col_a_val.strip()
                                if key not in results:
                                    results[key] = [0.0, 0.0]
                                if is_left:
                                    results[key][0] += numeric_val
                                else:
                                    results[key][1] += numeric_val
                            except (ValueError, TypeError):
                                pass

    return [(k, v[0], v[1]) for k, v in results.items() if v[0] != 0 or v[1] != 0]


def get_values_for_page8_section(ma_workbooks: List[Tuple[str, UnifiedWorkbook]],
                                  left_headers: List[str], right_headers: List[str],
                                  section_letter: str,
                                  search_cols: List[str] = ["G", "H"],
                                  log_emit=None) -> List[Tuple[str, float, float]]:
    """
    Get values from Page 8 for a specific section (A or B).
    Returns list of (description, left_sum, right_sum).
    """
    from openpyxl.utils import column_index_from_string

    results = {}
    next_section = "B" if section_letter == "A" else None

    for fname, wb in ma_workbooks:
        if "Page 8" not in wb.sheetnames:
            continue
        ws = wb["Page 8"]

        for is_left, headers in [(True, left_headers), (False, right_headers)]:
            for header in headers:
                col_idx = None
                header_row = None

                for col_letter in search_cols:
                    c_idx = column_index_from_string(col_letter)
                    for row in range(1, min(20, ws.max_row + 1)):
                        cell_val = ws.cell_displayed_value(row, c_idx).strip().upper()
                        if header.upper() in cell_val:
                            col_idx = c_idx
                            header_row = row
                            break
                    if col_idx:
                        break

                if not col_idx:
                    continue

                # Find section start (row starting with section_letter)
                section_start = None
                section_end = None

                for row in range(header_row + 1, ws.max_row + 1):
                    col_a_val = ws.cell_displayed_value(row, 1).strip()
                    if col_a_val.upper().startswith(section_letter):
                        section_start = row
                    elif section_start and next_section and col_a_val.upper().startswith(next_section):
                        section_end = row
                        break

                if not section_start:
                    continue

                end_row = section_end if section_end else ws.max_row + 1

                for row in range(section_start + 1, end_row):
                    col_b_val = ws.cell_displayed_value(row, 2).strip()
                    cell_val = ws.cell_value(row, col_idx)

                    if cell_val is not None and cell_val != "" and col_b_val and not col_b_val.upper().startswith("TOTAL"):
                        try:
                            numeric_val = float(cell_val) if not isinstance(cell_val, (int, float)) else cell_val
                            key = col_b_val.strip()
                            if key not in results:
                                results[key] = [0.0, 0.0]
                            if is_left:
                                results[key][0] += numeric_val
                            else:
                                results[key][1] += numeric_val
                        except (ValueError, TypeError):
                            pass

    return [(k, v[0], v[1]) for k, v in results.items() if v[0] != 0 or v[1] != 0]


def get_values_for_page7(ma_workbooks: List[Tuple[str, UnifiedWorkbook]],
                          left_headers: List[str], right_headers: List[str],
                          search_cols: List[str] = ["G", "H"],
                          log_emit=None) -> List[Tuple[str, float, float]]:
    """
    Get values from Page 7 for General Administrative Expenses.
    Returns list of (description, left_sum, right_sum).
    """
    from openpyxl.utils import column_index_from_string

    results = {}

    for fname, wb in ma_workbooks:
        if "Page 7" not in wb.sheetnames:
            continue
        ws = wb["Page 7"]

        for is_left, headers in [(True, left_headers), (False, right_headers)]:
            for header in headers:
                col_idx = None
                header_row = None

                for col_letter in search_cols:
                    c_idx = column_index_from_string(col_letter)
                    for row in range(1, min(20, ws.max_row + 1)):
                        cell_val = ws.cell_displayed_value(row, c_idx).strip().upper()
                        if header.upper() in cell_val:
                            col_idx = c_idx
                            header_row = row
                            break
                    if col_idx:
                        break

                if not col_idx:
                    continue

                for row in range(header_row + 1, ws.max_row + 1):
                    col_b_val = ws.cell_displayed_value(row, 2).strip()
                    cell_val = ws.cell_value(row, col_idx)

                    if cell_val is not None and cell_val != "" and col_b_val:
                        # Skip excluded items
                        skip = False
                        for excl in GAE_EXCLUDE:
                            if excl.upper() in col_b_val.upper():
                                skip = True
                                break
                        if skip:
                            continue

                        try:
                            numeric_val = float(cell_val) if not isinstance(cell_val, (int, float)) else cell_val
                            key = col_b_val.strip()
                            if key not in results:
                                results[key] = [0.0, 0.0]
                            if is_left:
                                results[key][0] += numeric_val
                            else:
                                results[key][1] += numeric_val
                        except (ValueError, TypeError):
                            pass

    return [(k, v[0], v[1]) for k, v in results.items() if v[0] != 0 or v[1] != 0]


def get_values_for_page9(ma_workbooks: List[Tuple[str, UnifiedWorkbook]],
                          left_headers: List[str], right_headers: List[str],
                          search_cols: List[str] = ["G", "H"],
                          log_emit=None) -> List[Tuple[str, float, float]]:
    """
    Get values from Page 9 for Other Income (section B).
    Returns list of (description, left_sum, right_sum).
    """
    from openpyxl.utils import column_index_from_string

    results = {}

    for fname, wb in ma_workbooks:
        if "Page 9" not in wb.sheetnames:
            continue
        ws = wb["Page 9"]

        for is_left, headers in [(True, left_headers), (False, right_headers)]:
            for header in headers:
                col_idx = None
                header_row = None

                for col_letter in search_cols:
                    c_idx = column_index_from_string(col_letter)
                    for row in range(1, min(20, ws.max_row + 1)):
                        cell_val = ws.cell_displayed_value(row, c_idx).strip().upper()
                        if header.upper() in cell_val:
                            col_idx = c_idx
                            header_row = row
                            break
                    if col_idx:
                        break

                if not col_idx:
                    continue

                # Find section B start
                section_start = None
                for row in range(header_row + 1, ws.max_row + 1):
                    col_a_val = ws.cell_displayed_value(row, 1).strip()
                    if col_a_val.upper().startswith("B"):
                        section_start = row
                        break

                if not section_start:
                    continue

                for row in range(section_start + 1, ws.max_row + 1):
                    col_a_val = ws.cell_displayed_value(row, 1).strip()
                    col_b_val = ws.cell_displayed_value(row, 2).strip()
                    cell_val = ws.cell_value(row, col_idx)

                    # Stop if we hit next section
                    if col_a_val and col_a_val.upper().startswith("C"):
                        break

                    if cell_val is not None and cell_val != "" and col_b_val and not col_b_val.upper().startswith("TOTAL"):
                        try:
                            numeric_val = float(cell_val) if not isinstance(cell_val, (int, float)) else cell_val
                            key = col_b_val.strip()
                            if key not in results:
                                results[key] = [0.0, 0.0]
                            if is_left:
                                results[key][0] += numeric_val
                            else:
                                results[key][1] += numeric_val
                        except (ValueError, TypeError):
                            pass

    return [(k, v[0], v[1]) for k, v in results.items() if v[0] != 0 or v[1] != 0]


def find_extreme_value(values: List[Tuple[str, float, float]], find_lowest: bool) -> Optional[str]:
    """
    Find the description with the lowest/highest difference (left - right).
    """
    if not values:
        return None

    diffs = [(desc, left - right) for desc, left, right in values]

    if find_lowest:
        sorted_diffs = sorted(diffs, key=lambda x: x[1])
    else:
        sorted_diffs = sorted(diffs, key=lambda x: x[1], reverse=True)

    for desc, diff in sorted_diffs:
        if desc.strip():
            return desc

    return None


def build_headers_from_comparison(comparison: Dict) -> Tuple[List[str], List[str], List[str]]:
    """
    Build header strings from comparison info.
    Returns (left_headers, right_headers, search_cols).
    """
    search_cols = ["G", "H"]

    if comparison["type"] == "ytd":
        search_cols = ["I", "J"]
        return [comparison["left_header"]], [comparison["right_header"]], search_cols

    left_headers = []
    right_headers = []

    for month, year in comparison.get("left", []):
        left_headers.append(build_month_header(month, year))

    for month, year in comparison.get("right", []):
        right_headers.append(build_month_header(month, year))

    return left_headers, right_headers, search_cols


def process_variance_report(
    report_path: str,
    ma_files: List[str],
    selected_month: str,
    selected_year: int,
    log_emit=None,
) -> str:
    """
    Main processing function.
    """
    # Build sheet name like "Nov'25"
    year_short = selected_year % 100
    sheet_name = f"{selected_month}'{year_short:02d}"

    _emit(log_emit, f"Looking for worksheet: {sheet_name}")

    # Load the report workbook
    report_wb = load_workbook(report_path, data_only=False)  # Keep formulas for modification
    report_wb_data = load_workbook(report_path, data_only=True)  # For reading displayed values

    if sheet_name not in report_wb.sheetnames:
        # Try variations
        found = False
        for sn in report_wb.sheetnames:
            if sn.upper().replace("'", "").replace(" ", "") == sheet_name.upper().replace("'", "").replace(" ", ""):
                sheet_name = sn
                found = True
                break
        if not found:
            raise ValueError(f"Worksheet '{sheet_name}' not found. Available: {report_wb.sheetnames}")

    _emit(log_emit, f"Found worksheet: {sheet_name}")

    ws = report_wb[sheet_name]
    ws_data = report_wb_data[sheet_name]

    # Load MA workbooks
    ma_workbooks = load_ma_workbooks(ma_files, log_emit)
    if not ma_workbooks:
        raise ValueError("No MA files could be loaded.")

    # Find all Analytical Review cells
    analytical_reviews = []
    for row in range(1, ws.max_row + 1):
        for col in [1, 2]:  # Column A and B
            cell_data = ws_data.cell(row=row, column=col)
            cell_val = get_cell_displayed_value(cell_data).strip()
            if cell_val.upper().startswith("ANALYTICAL REVIEW"):
                analytical_reviews.append((row, col, cell_val))
                _emit(log_emit, f"Found Analytical Review at row {row}: {cell_val[:50]}...")

    if not analytical_reviews:
        raise ValueError("No 'Analytical Review' cells found in column A or B.")

    _emit(log_emit, f"\nFound {len(analytical_reviews)} Analytical Review sections")

    # Process each Analytical Review section
    replacements_made = 0

    for i, (ar_row, ar_col, ar_text) in enumerate(analytical_reviews):
        # Extract bracket content
        bracket_match = re.search(r"\(([^)]+)\)", ar_text)
        if not bracket_match:
            _emit(log_emit, f"[SKIP] No bracket content found in: {ar_text[:50]}")
            continue

        bracket_content = bracket_match.group(1)
        _emit(log_emit, f"\nProcessing: ({bracket_content})")

        # Parse the comparison
        comparison = parse_analytical_bracket(bracket_content, selected_month, selected_year)
        if not comparison["type"]:
            _emit(log_emit, f"[SKIP] Could not parse comparison: {bracket_content}")
            continue

        # Build headers
        left_headers, right_headers, search_cols = build_headers_from_comparison(comparison)
        _emit(log_emit, f"  Left headers: {left_headers}")
        _emit(log_emit, f"  Right headers: {right_headers}")
        _emit(log_emit, f"  Search columns: {search_cols}")

        # Determine the range to search (until next Analytical Review or end)
        end_row = analytical_reviews[i + 1][0] if i + 1 < len(analytical_reviews) else ws.max_row + 1

        # Find Description cells in this section
        for row in range(ar_row + 1, end_row):
            for col in [1, 2]:  # Column A and B
                cell = ws.cell(row=row, column=col)
                cell_data = ws_data.cell(row=row, column=col)
                displayed_val = get_cell_displayed_value(cell_data).strip()
                raw_val = get_cell_raw_value(cell)

                if "Description" not in displayed_val:
                    continue

                _emit(log_emit, f"\n  Found Description cell at row {row}: {displayed_val[:60]}...")

                # Determine what to look for based on cell content
                displayed_upper = displayed_val.upper()
                find_lowest = "LOWER" in displayed_upper
                find_highest = "HIGHER" in displayed_upper

                replacement_value = None

                if "PRODUCTION OVERHEAD" in displayed_upper:
                    _emit(log_emit, "    -> Processing Production Overhead (Page 5 & 6)")
                    values = get_values_for_production_overhead(
                        ma_workbooks, left_headers, right_headers, search_cols, log_emit
                    )
                    if values:
                        replacement_value = find_extreme_value(values, find_lowest)
                        _emit(log_emit, f"    -> Found {len(values)} items, selected: {replacement_value}")

                elif "PURCHASE RELATED COSTS" in displayed_upper:
                    _emit(log_emit, "    -> Processing Purchase Related Costs (Page 8, Section A)")
                    values = get_values_for_page8_section(
                        ma_workbooks, left_headers, right_headers, "A", search_cols, log_emit
                    )
                    if values:
                        replacement_value = find_extreme_value(values, find_lowest)
                        _emit(log_emit, f"    -> Found {len(values)} items, selected: {replacement_value}")

                elif "SALES RELATED COSTS" in displayed_upper:
                    _emit(log_emit, "    -> Processing Sales Related Costs (Page 8, Section B)")
                    values = get_values_for_page8_section(
                        ma_workbooks, left_headers, right_headers, "B", search_cols, log_emit
                    )
                    if values:
                        replacement_value = find_extreme_value(values, find_lowest)
                        _emit(log_emit, f"    -> Found {len(values)} items, selected: {replacement_value}")

                elif "GENERAL ADMINISTRATIVE EXPENSES" in displayed_upper:
                    _emit(log_emit, "    -> Processing General Administrative Expenses (Page 7)")
                    values = get_values_for_page7(
                        ma_workbooks, left_headers, right_headers, search_cols, log_emit
                    )
                    if values:
                        replacement_value = find_extreme_value(values, not find_lowest)  # Higher = not lowest
                        _emit(log_emit, f"    -> Found {len(values)} items, selected: {replacement_value}")

                elif "OTHER INCOME" in displayed_upper:
                    _emit(log_emit, "    -> Processing Other Income (Page 9, Section B)")
                    values = get_values_for_page9(
                        ma_workbooks, left_headers, right_headers, search_cols, log_emit
                    )
                    if values:
                        # Determine lowest or highest based on text before Description
                        replacement_value = find_extreme_value(values, find_lowest if find_lowest else not find_highest)
                        _emit(log_emit, f"    -> Found {len(values)} items, selected: {replacement_value}")

                # Replace Description in the formula/value
                if replacement_value:
                    if isinstance(raw_val, str):
                        new_val = raw_val.replace("Description", replacement_value)
                        cell.value = new_val
                        replacements_made += 1
                        _emit(log_emit, f"    -> Replaced 'Description' with '{replacement_value}'")
                    else:
                        _emit(log_emit, f"    -> Cell is not a string/formula, cannot replace")
                else:
                    _emit(log_emit, f"    -> No replacement value found")

    # Save the modified workbook
    output_path = Path(report_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_path.parent / f"{output_path.stem}_modified_{timestamp}{output_path.suffix}"

    report_wb.save(output_file)

    _emit(log_emit, f"\n{'='*50}")
    _emit(log_emit, f"Processing complete!")
    _emit(log_emit, f"Total replacements made: {replacements_made}")
    _emit(log_emit, f"Output saved to: {output_file}")

    return str(output_file)


# --- UI Class (Pycro Station) ---

class MainWidget(QWidget):
    log_message = Signal(str)
    processing_done = Signal(str)

    def __init__(self):
        super().__init__()
        self.setObjectName("varanalysis_desc_widget")
        self._build_ui()
        self._connect_signals()

    def _build_ui(self):
        self.desc_label = QLabel("", self)
        self.desc_label.setWordWrap(True)
        self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.desc_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.desc_label.setStyleSheet(
            "color: #dcdcdc; background: transparent; padding: 6px; "
            "border: 1px solid #3a3a3a; border-radius: 6px;"
        )
        self.set_long_description("")

        # File input 1: Variance Analysis Report
        self.select_report_btn = PrimaryPushButton("Select Variance Analysis Report Excel File", self)

        # Dropdowns row
        self.month_label = QLabel("Month:", self)
        self.month_label.setStyleSheet("color: #dcdcdc; background: transparent;")

        self.month_combo = ComboBox(self)
        self.month_combo.addItems(MONTHS)
        self.month_combo.setCurrentIndex(0)  # Default to Nov

        self.year_label = QLabel("Year:", self)
        self.year_label.setStyleSheet("color: #dcdcdc; background: transparent;")

        self.year_combo = ComboBox(self)
        current_year = datetime.now().year
        years = [str(y) for y in range(1975, current_year + 6)]
        self.year_combo.addItems(years)
        self.year_combo.setCurrentText(str(current_year))

        # File input 2: MA files (multi-select)
        self.select_ma_btn = PrimaryPushButton("Select Management Account Files", self)

        # Run button
        self.run_btn = PrimaryPushButton("Process", self)

        # Report file display
        self.report_label = QLabel("Variance Analysis Report", self)
        self.report_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.report_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        # MA files display
        self.ma_label = QLabel("Management Account Files", self)
        self.ma_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.ma_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        self.logs_label = QLabel("Process logs", self)
        self.logs_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.logs_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        shared_style = (
            "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        self.report_box = QTextEdit(self)
        self.report_box.setReadOnly(True)
        self.report_box.setPlaceholderText("Selected report file will appear here")
        self.report_box.setStyleSheet(shared_style)
        self.report_box.setMaximumHeight(60)

        self.ma_box = QTextEdit(self)
        self.ma_box.setReadOnly(True)
        self.ma_box.setPlaceholderText("Selected MA files will appear here")
        self.ma_box.setStyleSheet(shared_style)

        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText("Live process log will appear here")
        self.log_box.setStyleSheet(shared_style)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)

        main_layout.addWidget(self.desc_label, 0)

        # Report file button
        row1 = QHBoxLayout()
        row1.addStretch(1)
        row1.addWidget(self.select_report_btn, 2)
        row1.addStretch(1)
        main_layout.addLayout(row1, 0)

        # Dropdowns row
        row2 = QHBoxLayout()
        row2.addStretch(1)
        row2.addWidget(self.month_label)
        row2.addWidget(self.month_combo)
        row2.addSpacing(20)
        row2.addWidget(self.year_label)
        row2.addWidget(self.year_combo)
        row2.addStretch(1)
        main_layout.addLayout(row2, 0)

        # MA files button
        row3 = QHBoxLayout()
        row3.addStretch(1)
        row3.addWidget(self.select_ma_btn, 2)
        row3.addStretch(1)
        main_layout.addLayout(row3, 0)

        # Run button
        row4 = QHBoxLayout()
        row4.addStretch(1)
        row4.addWidget(self.run_btn, 1)
        row4.addStretch(1)
        main_layout.addLayout(row4, 0)

        # File displays labels
        row5 = QHBoxLayout()
        row5.addWidget(self.report_label, 1)
        row5.addWidget(self.ma_label, 1)
        main_layout.addLayout(row5, 0)

        # File displays
        row6 = QHBoxLayout()
        row6.addWidget(self.report_box, 1)
        row6.addWidget(self.ma_box, 1)
        main_layout.addLayout(row6, 1)

        # Log label and box
        main_layout.addWidget(self.logs_label, 0)
        main_layout.addWidget(self.log_box, 3)

    def set_long_description(self, text: str):
        clean = (text or "").strip()
        if clean:
            self.desc_label.setText(clean)
            self.desc_label.show()
        else:
            self.desc_label.clear()
            self.desc_label.hide()

    def _connect_signals(self):
        self.select_report_btn.clicked.connect(self.select_report_file)
        self.select_ma_btn.clicked.connect(self.select_ma_files)
        self.run_btn.clicked.connect(self.run_process)
        self.log_message.connect(self.append_log)
        self.processing_done.connect(self.on_done)

    def select_report_file(self):
        file, _ = QFileDialog.getOpenFileName(
            self,
            "Select Variance Analysis Report Excel File",
            "",
            "Excel Files (*.xlsx *.xlsm *.xls)",
        )
        if file:
            self.report_box.setPlainText(file)
        else:
            self.report_box.clear()

    def select_ma_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select Management Account Files",
            "",
            "Excel Files (*.xlsx *.xlsm *.xls)",
        )
        if files:
            self.ma_box.setPlainText("\n".join(files))
        else:
            self.ma_box.clear()

    def _get_report_file(self) -> str:
        return self.report_box.toPlainText().strip()

    def _get_ma_files(self) -> List[str]:
        text = self.ma_box.toPlainText().strip()
        if not text:
            return []
        return [line.strip() for line in text.split("\n") if line.strip()]

    def run_process(self):
        report_file = self._get_report_file()
        ma_files = self._get_ma_files()

        if not report_file:
            MessageBox("Warning", "Please select a Variance Analysis Report file.", self).exec()
            return

        if not ma_files:
            MessageBox("Warning", "Please select Management Account files.", self).exec()
            return

        selected_month = self.month_combo.currentText()
        selected_year = int(self.year_combo.currentText())

        self.log_box.clear()
        self.log_message.emit("Processing started...")
        self.log_message.emit(f"Report file: {Path(report_file).name}")
        self.log_message.emit(f"MA files: {len(ma_files)}")
        self.log_message.emit(f"Selected period: {selected_month} {selected_year}")
        self.log_message.emit("")

        self.run_btn.setEnabled(False)
        self.select_report_btn.setEnabled(False)
        self.select_ma_btn.setEnabled(False)

        def worker():
            try:
                out_path = process_variance_report(
                    report_file, ma_files, selected_month, selected_year,
                    log_emit=self.log_message.emit
                )
                self.processing_done.emit(out_path)
            except Exception as e:
                self.log_message.emit(f"CRITICAL ERROR: {e}")
                import traceback
                self.log_message.emit(traceback.format_exc())
                self.processing_done.emit("")

        threading.Thread(target=worker, daemon=True).start()

    def append_log(self, text: str):
        self.log_box.append(text)
        self.log_box.ensureCursorVisible()

    def on_done(self, out_path: str):
        self.run_btn.setEnabled(True)
        self.select_report_btn.setEnabled(True)
        self.select_ma_btn.setEnabled(True)

        if out_path:
            title = "Processing complete"
            msg_text = f"Output saved to:\n{Path(out_path).name}"
        else:
            title = "Processing failed"
            msg_text = "An error occurred. Check the logs for details."

        msg = MessageBox(title, msg_text, self)
        msg.yesButton.setText("OK")
        msg.cancelButton.hide()
        msg.exec()


def get_widget():
    return MainWidget()
