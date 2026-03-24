#!/usr/bin/env python3
from __future__ import annotations

import json
import threading
import time
from datetime import date, timedelta
from pathlib import Path
from typing import Callable
from urllib.request import Request, urlopen

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QDate, Qt, Signal
from PySide6.QtWidgets import (
    QFileDialog,
    QDateEdit,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QSizePolicy,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)
from qfluentwidgets import MessageBox, PrimaryPushButton

API_URL = "https://www.vietcombank.com.vn/api/exchangerates?date={}"
APP_MIN_DATE = date(2000, 1, 1)
REQUEST_DELAY_SECONDS = 0.2
USER_AGENT = "Mozilla/5.0 (pycro-station)"
MAJOR_CURRENCIES = ["USD", "EUR", "GBP", "JPY", "SGD", "AUD", "CNY", "THB", "KRW", "MYR"]


def _emit(log_emit: Callable[[str], None] | None, text: str):
    if callable(log_emit):
        try:
            log_emit(text)
            return
        except Exception:
            pass
    print(text)


def ensure_unique_path(path: Path) -> Path:
    if not path.exists():
        return path

    counter = 1
    while True:
        candidate = path.with_name(f"{path.stem} ({counter}){path.suffix}")
        if not candidate.exists():
            return candidate
        counter += 1


def make_output_filename(start_date: date, end_date: date) -> str:
    return f"VCB_Exchange_Rates_{start_date:%Y%m%d}_to_{end_date:%Y%m%d}.xlsx"


def parse_rate_number(value):
    if value is None:
        return None

    text = str(value).strip().replace(",", "")
    if not text or text in {"0", "0.0", "0.00"}:
        return None

    try:
        number = float(text)
    except Exception:
        return None

    return number if number > 0 else None


def fetch_rates(start_date: date, end_date: date, log_emit: Callable[[str], None] | None = None):
    all_records = []
    no_data_count = 0
    error_count = 0
    total_days = (end_date - start_date).days + 1

    current = start_date
    day_index = 1
    while current <= end_date:
        iso_date = current.isoformat()
        try:
            req = Request(API_URL.format(iso_date), headers={"User-Agent": USER_AGENT})
            with urlopen(req, timeout=15) as resp:
                body = json.loads(resp.read())

            rows = body.get("Data") or []
            count = int(body.get("Count") or 0)
            if count > 0 and isinstance(rows, list):
                for item in rows:
                    all_records.append(
                        {
                            "date": iso_date,
                            "currency_code": item.get("currencyCode", ""),
                            "currency_name": item.get("currencyName", ""),
                            "buy_cash": item.get("cash"),
                            "buy_transfer": item.get("transfer"),
                            "sell": item.get("sell"),
                        }
                    )
                _emit(log_emit, f"[{day_index}/{total_days}] OK: {iso_date} ({count} currencies)")
            else:
                no_data_count += 1
                _emit(log_emit, f"[{day_index}/{total_days}] NO DATA: {iso_date}")
        except Exception as exc:
            error_count += 1
            _emit(log_emit, f"[{day_index}/{total_days}] ERROR: {iso_date} - {exc}")

        current += timedelta(days=1)
        day_index += 1
        if current <= end_date:
            time.sleep(REQUEST_DELAY_SECONDS)

    return all_records, no_data_count, error_count


def build_workbook(records):
    currencies = sorted({record["currency_code"] for record in records if record.get("currency_code")})
    dates = sorted({record["date"] for record in records if record.get("date")})
    lookup = {(record["date"], record["currency_code"]): record for record in records}

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_fill = PatternFill("solid", fgColor="F2F7FB")

    def write_headers(ws, headers):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def write_rows(ws, rows, numeric_columns):
        for row_idx, row_values in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if col_idx in numeric_columns:
                    cell.value = parse_rate_number(value)
                    cell.number_format = "#,##0.00"
                else:
                    cell.value = value
                cell.font = data_font
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

    ws_all = wb.active
    ws_all.title = "All Rates"
    all_headers = ["Date", "Currency Code", "Currency Name", "Cash Buying", "Telegraphic Buying", "Selling"]
    write_headers(ws_all, all_headers)
    ws_all.column_dimensions["A"].width = 12
    ws_all.column_dimensions["B"].width = 14
    ws_all.column_dimensions["C"].width = 24
    ws_all.column_dimensions["D"].width = 16
    ws_all.column_dimensions["E"].width = 20
    ws_all.column_dimensions["F"].width = 14
    all_rows = [
        [
            record["date"],
            record["currency_code"],
            record["currency_name"],
            record["buy_cash"],
            record["buy_transfer"],
            record["sell"],
        ]
        for record in sorted(records, key=lambda item: (item["date"], item["currency_code"]))
    ]
    write_rows(ws_all, all_rows, numeric_columns={4, 5, 6})
    ws_all.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}{max(1, len(all_rows) + 1)}"
    ws_all.freeze_panes = "A2"

    for currency_code in MAJOR_CURRENCIES:
        if currency_code not in currencies:
            continue

        ws = wb.create_sheet(title=f"{currency_code}_VND")
        headers = ["Date", "Cash Buying", "Telegraphic Buying", "Selling"]
        write_headers(ws, headers)
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 14

        rows = []
        for rate_date in dates:
            record = lookup.get((rate_date, currency_code))
            if record:
                rows.append([rate_date, record["buy_cash"], record["buy_transfer"], record["sell"]])

        write_rows(ws, rows, numeric_columns={2, 3, 4})
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"
        ws.freeze_panes = "A2"

    return wb, len(dates), len(currencies), len(records)


def generate_workbook(start_date: date, end_date: date, output_dir: Path, log_emit: Callable[[str], None] | None = None):
    total_days = (end_date - start_date).days + 1
    _emit(log_emit, f"Fetching VCB exchange rates from {start_date.isoformat()} to {end_date.isoformat()} ({total_days} day(s))")

    records, no_data_count, error_count = fetch_rates(start_date, end_date, log_emit=log_emit)
    if not records:
        if error_count > 0:
            raise RuntimeError("No exchange-rate data could be fetched for the selected range.")
        raise ValueError("No exchange-rate data returned for the selected range.")

    _emit(log_emit, "")
    _emit(log_emit, f"Fetched {len(records)} records across {total_days} day(s)")

    wb, dates_with_data, currency_count, record_count = build_workbook(records)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = ensure_unique_path(output_dir / make_output_filename(start_date, end_date))
    wb.save(output_path)

    _emit(log_emit, f"Saved workbook: {output_path}")
    if no_data_count:
        _emit(log_emit, f"Dates with no data: {no_data_count}")
    if error_count:
        _emit(log_emit, f"Dates with request errors: {error_count}")

    return str(output_path), total_days, dates_with_data, currency_count, record_count


class MainWidget(QWidget):
    log_message = Signal(str)
    processing_done = Signal(bool, str, int, int, int, int)

    def __init__(self):
        super().__init__()
        self.setObjectName("vcb_exchange_rates_widget")
        self.output_dir = self._default_output_dir()
        self.last_output_path = ""

        self._build_ui()
        self._connect_signals()
        self._refresh_summary()

    def _build_ui(self):
        self.desc_label = QLabel("", self)
        self.desc_label.setWordWrap(True)
        self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.desc_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.desc_label.setStyleSheet(
            "color: #dcdcdc; background: transparent; padding: 6px; "
            "border: 1px solid #3a3a3a; border-radius: 6px;"
        )
        self.set_long_description("")

        self.select_output_btn = PrimaryPushButton("Select Output Folder", self)
        self.run_btn = PrimaryPushButton("Fetch VCB Rates", self)

        self.start_label = QLabel("Start date", self)
        self.end_label = QLabel("End date", self)
        for label in (self.start_label, self.end_label):
            label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        date_style = (
            "QDateEdit{background: #1f1f1f; color: #d0d0d0; border: 1px solid #3a3a3a; "
            "border-radius: 6px; padding: 4px 8px; min-height: 22px;}"
        )

        today = date.today()
        default_start = max(APP_MIN_DATE, today - timedelta(days=30))

        self.start_date_edit = QDateEdit(self)
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.start_date_edit.setMinimumDate(QDate(APP_MIN_DATE.year, APP_MIN_DATE.month, APP_MIN_DATE.day))
        self.start_date_edit.setMaximumDate(QDate(today.year, today.month, today.day))
        self.start_date_edit.setDate(QDate(default_start.year, default_start.month, default_start.day))
        self.start_date_edit.setStyleSheet(date_style)

        self.end_date_edit = QDateEdit(self)
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.end_date_edit.setMinimumDate(QDate(APP_MIN_DATE.year, APP_MIN_DATE.month, APP_MIN_DATE.day))
        self.end_date_edit.setMaximumDate(QDate(today.year, today.month, today.day))
        self.end_date_edit.setDate(QDate(today.year, today.month, today.day))
        self.end_date_edit.setStyleSheet(date_style)

        self.summary_label = QLabel("Selection summary", self)
        self.summary_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.summary_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        self.logs_label = QLabel("Process logs", self)
        self.logs_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.logs_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        box_style = (
            "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        self.summary_box = QTextEdit(self)
        self.summary_box.setReadOnly(True)
        self.summary_box.setPlaceholderText("Selected date range and output details will appear here")
        self.summary_box.setStyleSheet(box_style)

        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText("Live process log will appear here")
        self.log_box.setStyleSheet(box_style)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)
        main_layout.addWidget(self.desc_label)

        output_row = QHBoxLayout()
        output_row.addStretch(1)
        output_row.addWidget(self.select_output_btn)
        output_row.addStretch(1)
        main_layout.addLayout(output_row)

        date_grid = QGridLayout()
        date_grid.setColumnStretch(1, 1)
        date_grid.setColumnStretch(3, 1)
        date_grid.addWidget(self.start_label, 0, 0, Qt.AlignLeft)
        date_grid.addWidget(self.start_date_edit, 0, 1)
        date_grid.addWidget(self.end_label, 0, 2, Qt.AlignLeft)
        date_grid.addWidget(self.end_date_edit, 0, 3)
        main_layout.addLayout(date_grid)

        run_row = QHBoxLayout()
        run_row.addStretch(1)
        run_row.addWidget(self.run_btn)
        run_row.addStretch(1)
        main_layout.addLayout(run_row)

        labels_row = QHBoxLayout()
        labels_row.addWidget(self.summary_label, 1)
        labels_row.addWidget(self.logs_label, 1)
        main_layout.addLayout(labels_row)

        boxes_row = QHBoxLayout()
        boxes_row.addWidget(self.summary_box, 1)
        boxes_row.addWidget(self.log_box, 1)
        main_layout.addLayout(boxes_row, 2)

    def set_long_description(self, text: str):
        clean = (text or "").strip()
        if clean:
            self.desc_label.setText(clean)
            self.desc_label.show()
        else:
            self.desc_label.clear()
            self.desc_label.hide()

    def _connect_signals(self):
        self.select_output_btn.clicked.connect(self.select_output_folder)
        self.run_btn.clicked.connect(self.run_process)
        self.start_date_edit.dateChanged.connect(self._on_date_changed)
        self.end_date_edit.dateChanged.connect(self._on_date_changed)
        self.log_message.connect(self.append_log)
        self.processing_done.connect(self.on_processing_done)

    def _default_output_dir(self) -> Path:
        downloads_dir = Path.home() / "Downloads"
        if downloads_dir.is_dir():
            return downloads_dir
        return Path.home()

    def _python_date(self, edit: QDateEdit) -> date:
        qdate = edit.date()
        return date(qdate.year(), qdate.month(), qdate.day())

    def _on_date_changed(self, _value=None):
        start_date = self._python_date(self.start_date_edit)
        end_date = self._python_date(self.end_date_edit)

        if start_date > end_date:
            sender = self.sender()
            if sender is self.start_date_edit:
                self.end_date_edit.setDate(self.start_date_edit.date())
            else:
                self.start_date_edit.setDate(self.end_date_edit.date())

        self._refresh_summary()

    def _refresh_summary(self):
        start_date = self._python_date(self.start_date_edit)
        end_date = self._python_date(self.end_date_edit)
        span_days = (end_date - start_date).days + 1
        filename = make_output_filename(start_date, end_date)

        lines = [
            f"Start date: {start_date.isoformat()}",
            f"End date: {end_date.isoformat()}",
            f"Inclusive range: {span_days} day(s)",
            f"Output folder: {self.output_dir}",
            f"Output file: {filename}",
        ]
        if self.last_output_path:
            lines.append("")
            lines.append(f"Last output: {self.last_output_path}")

        self.summary_box.setPlainText("\n".join(lines))

    def select_output_folder(self):
        selected = QFileDialog.getExistingDirectory(self, "Select Output Folder", str(self.output_dir))
        if selected:
            self.output_dir = Path(selected)
            self._refresh_summary()

    def run_process(self):
        start_date = self._python_date(self.start_date_edit)
        end_date = self._python_date(self.end_date_edit)

        if start_date > end_date:
            MessageBox("Warning", "Start date cannot be later than end date.", self).exec()
            return

        self.log_box.clear()
        self.log_message.emit("Process started...")
        self.log_message.emit(f"Output folder: {self.output_dir}")
        self.log_message.emit("")

        self._set_running_state(True)

        def worker():
            try:
                output_path, total_days, dates_with_data, currency_count, record_count = generate_workbook(
                    start_date,
                    end_date,
                    self.output_dir,
                    log_emit=self.log_message.emit,
                )
                self.processing_done.emit(True, output_path, total_days, dates_with_data, currency_count, record_count)
            except Exception as exc:
                self.log_message.emit(f"CRITICAL ERROR: {exc}")
                self.processing_done.emit(False, "", 0, 0, 0, 0)

        threading.Thread(target=worker, daemon=True).start()

    def _set_running_state(self, running: bool):
        enabled = not running
        self.select_output_btn.setEnabled(enabled)
        self.run_btn.setEnabled(enabled)
        self.start_date_edit.setEnabled(enabled)
        self.end_date_edit.setEnabled(enabled)

    def append_log(self, text: str):
        self.log_box.append(text)
        self.log_box.ensureCursorVisible()

    def on_processing_done(self, success: bool, output_path: str, total_days: int, dates_with_data: int, currency_count: int, record_count: int):
        self._set_running_state(False)

        if success and output_path:
            self.last_output_path = output_path
            self._refresh_summary()
            self.log_message.emit("")
            self.log_message.emit(f"Completed: {dates_with_data} date(s) with data, {currency_count} currencies, {record_count} records")

            lines = [
                f"Range length: {total_days} day(s)",
                f"Dates with data: {dates_with_data}",
                f"Currencies: {currency_count}",
                f"Records: {record_count}",
                f"Output: {Path(output_path).name}",
            ]
            msg = MessageBox("VCB rates fetched", "\n".join(lines), self)
        else:
            msg = MessageBox("VCB fetch finished with issues", "No output workbook was created. Check the log for details.", self)

        msg.yesButton.setText("OK")
        msg.cancelButton.hide()
        msg.exec()


def get_widget():
    return MainWidget()
