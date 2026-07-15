from __future__ import annotations

import os
import re
import tempfile
import threading
import zipfile
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable, Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, range_boundaries
from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QSizePolicy,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)
from qfluentwidgets import LineEdit, MessageBox, PrimaryPushButton


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
HEADER_SCAN_ROWS = 15
FIRST_QUOTE_LABEL = "1st quote"
OUTPUT_SUFFIX = " - Quote Updated"


@dataclass(frozen=True)
class ProcessResult:
    output_path: str
    sheet_name: str
    original_quote_column: int
    quote_file_quote_column: int
    copied_columns: int
    copied_rows: int


class MainWidget(QWidget):
    log_message = Signal(str)
    processing_done = Signal(bool, str, str)

    def __init__(self):
        super().__init__()
        self.setObjectName("cm_quotation_updater_widget")
        self.original_path = ""
        self.quote_path = ""

        self._build_ui()
        self._connect_signals()

    def _build_ui(self):
        self.desc_label = QLabel("", self)
        self.desc_label.setWordWrap(True)
        self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.desc_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.desc_label.setStyleSheet(
            "color: #dcdcdc; background: transparent; padding: 6px; "
            "border: 1px solid #3a3a3a; border-radius: 6px;"
        )
        self.set_long_description("")

        self.original_label = QLabel("Original file", self)
        self.original_label.setStyleSheet("color: #dcdcdc; background: transparent;")
        self.original_field = LineEdit(self)
        self.original_field.setReadOnly(True)
        self.original_field.setPlaceholderText(
            "Workbook that contains the additional calculation columns"
        )
        self.original_button = PrimaryPushButton("Select Original File", self)
        self.original_button.setFixedWidth(180)

        self.quote_label = QLabel("Quote file", self)
        self.quote_label.setStyleSheet("color: #dcdcdc; background: transparent;")
        self.quote_field = LineEdit(self)
        self.quote_field.setReadOnly(True)
        self.quote_field.setPlaceholderText(
            "Workbook that contains the quotation history to copy"
        )
        self.quote_button = PrimaryPushButton("Select Quote File", self)
        self.quote_button.setFixedWidth(180)

        self.run_button = PrimaryPushButton("Update Quotation", self)
        self.run_button.setFixedWidth(220)

        self.logs_label = QLabel("Process logs", self)
        self.logs_label.setStyleSheet(
            "color: #dcdcdc; background: transparent; padding-left: 2px;"
        )
        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText("Processing details will appear here")
        self.log_box.setStyleSheet(
            "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)
        main_layout.addWidget(self.desc_label)

        original_row = QHBoxLayout()
        original_row.setSpacing(10)
        self.original_label.setFixedWidth(90)
        original_row.addWidget(self.original_label)
        original_row.addWidget(self.original_field, 1)
        original_row.addWidget(self.original_button)
        main_layout.addLayout(original_row)

        quote_row = QHBoxLayout()
        quote_row.setSpacing(10)
        self.quote_label.setFixedWidth(90)
        quote_row.addWidget(self.quote_label)
        quote_row.addWidget(self.quote_field, 1)
        quote_row.addWidget(self.quote_button)
        main_layout.addLayout(quote_row)

        run_row = QHBoxLayout()
        run_row.addStretch(1)
        run_row.addWidget(self.run_button)
        run_row.addStretch(1)
        main_layout.addLayout(run_row)

        main_layout.addWidget(self.logs_label)
        main_layout.addWidget(self.log_box, 1)

    def set_long_description(self, text: str):
        clean = (text or "").strip()
        if clean:
            self.desc_label.setText(clean)
            self.desc_label.show()
        else:
            self.desc_label.clear()
            self.desc_label.hide()

    def _connect_signals(self):
        self.original_button.clicked.connect(self.select_original_file)
        self.quote_button.clicked.connect(self.select_quote_file)
        self.run_button.clicked.connect(self.run_process)
        self.log_message.connect(self.append_log)
        self.processing_done.connect(self.on_processing_done)

    def select_original_file(self):
        path = self._select_excel_file("Select Original CM Quotation File")
        if path:
            self.original_path = path
            self.original_field.setText(path)

    def select_quote_file(self):
        path = self._select_excel_file("Select Quote File")
        if path:
            self.quote_path = path
            self.quote_field.setText(path)

    def _select_excel_file(self, title: str) -> str:
        path, _ = QFileDialog.getOpenFileName(
            self,
            title,
            "",
            "Excel Workbooks (*.xlsx *.xlsm)",
        )
        return path or ""

    def run_process(self):
        try:
            original_path, quote_path = validate_input_paths(
                self.original_path, self.quote_path
            )
        except Exception as exc:
            MessageBox("Cannot start", str(exc), self).exec()
            return

        self.log_box.clear()
        self._set_controls_enabled(False)
        self.log_message.emit("Process started.")

        def worker():
            try:
                result = update_quotation_workbook(
                    original_path,
                    quote_path,
                    self.log_message.emit,
                )
                summary = (
                    f"Updated sheet: {result.sheet_name}\n"
                    f"Copied columns: {result.copied_columns}\n"
                    f"Copied rows: {result.copied_rows}\n"
                    f"Output: {result.output_path}"
                )
                self.processing_done.emit(True, summary, result.output_path)
            except Exception as exc:
                self.log_message.emit(f"ERROR: {exc}")
                self.processing_done.emit(False, str(exc), "")

        threading.Thread(target=worker, daemon=True).start()

    def append_log(self, text: str):
        self.log_box.append(text)
        self.log_box.ensureCursorVisible()

    def on_processing_done(self, success: bool, message: str, output_path: str):
        self._set_controls_enabled(True)
        if success:
            self.log_message.emit(f"Output workbook saved to: {output_path}")
            title = "Quotation updated"
        else:
            title = "Update failed"

        dialog = MessageBox(title, message, self)
        dialog.yesButton.setText("OK")
        dialog.cancelButton.hide()
        dialog.exec()

    def _set_controls_enabled(self, enabled: bool):
        self.original_button.setEnabled(enabled)
        self.quote_button.setEnabled(enabled)
        self.run_button.setEnabled(enabled)


def get_widget():
    return MainWidget()


def validate_input_paths(original_path: str, quote_path: str) -> tuple[str, str]:
    original = _validate_excel_path(original_path, "Original file")
    quote = _validate_excel_path(quote_path, "Quote file")

    if os.path.normcase(os.path.abspath(original)) == os.path.normcase(
        os.path.abspath(quote)
    ):
        raise ValueError("The original file and quote file must be different files.")

    return original, quote


def _validate_excel_path(path: str, label: str) -> str:
    clean = (path or "").strip()
    if not clean:
        raise ValueError(f"Select the {label.lower()}.")
    if not os.path.isfile(clean):
        raise ValueError(f"{label} does not exist: {clean}")
    if Path(clean).suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"{label} must be an .xlsx or .xlsm workbook.")
    return os.path.abspath(clean)


def update_quotation_workbook(
    original_path: str,
    quote_path: str,
    log_emit: Optional[Callable[[str], None]] = None,
) -> ProcessResult:
    log = log_emit or (lambda _message: None)
    original_path, quote_path = validate_input_paths(original_path, quote_path)

    log(f"Original file: {original_path}")
    log(f"Quote file: {quote_path}")

    original_wb = None
    quote_wb = None
    try:
        original_wb = _load_workbook_for_editing(original_path)
        quote_wb = _load_workbook_for_editing(quote_path)

        original_ws, original_start = _find_sheet_and_quote_start(original_wb)
        quote_ws, quote_start = _find_sheet_and_quote_start(quote_wb)

        log(
            "Detected original quotation section at "
            f"{original_ws.title}!{get_column_letter(original_start)}."
        )
        log(
            "Detected quote-file quotation section at "
            f"{quote_ws.title}!{get_column_letter(quote_start)}."
        )

        if original_start <= quote_start:
            raise ValueError(
                "The selected original file does not appear to contain additional "
                "columns before the quotation section. Check that the two inputs are "
                "assigned correctly."
            )

        quote_end = _find_quote_section_end(quote_ws, quote_start)
        original_end = _find_quote_section_end(original_ws, original_start)
        source_max_row = quote_ws.max_row
        copied_columns = quote_end - quote_start + 1
        target_end = original_start + copied_columns - 1

        _log_style_alignment(original_ws, quote_ws, log)
        _replace_quote_section(
            original_ws=original_ws,
            quote_ws=quote_ws,
            original_start=original_start,
            original_end=original_end,
            quote_start=quote_start,
            quote_end=quote_end,
            source_max_row=source_max_row,
            log=log,
        )

        _update_auto_filter(
            original_ws,
            quote_ws,
            target_end=target_end,
            log=log,
        )
        _update_print_area_right_edge(original_ws, target_end)

        output_path = _proposed_output_path(original_path)
        _save_workbook_atomically(
            original_wb,
            original_path=original_path,
            output_path=output_path,
            log=log,
        )
        _verify_output(
            output_path,
            sheet_name=original_ws.title,
            expected_quote_column=original_start,
        )

        log(
            f"Preserved columns A:{get_column_letter(original_start - 1)} and copied "
            f"{get_column_letter(quote_start)}:{get_column_letter(quote_end)} from "
            f"the quote file into {get_column_letter(original_start)}:"
            f"{get_column_letter(target_end)}."
        )
        log("Process completed successfully.")

        return ProcessResult(
            output_path=output_path,
            sheet_name=original_ws.title,
            original_quote_column=original_start,
            quote_file_quote_column=quote_start,
            copied_columns=copied_columns,
            copied_rows=source_max_row,
        )
    finally:
        if original_wb is not None:
            original_wb.close()
        if quote_wb is not None:
            quote_wb.close()


def _load_workbook_for_editing(path: str):
    keep_vba = Path(path).suffix.lower() == ".xlsm"
    return load_workbook(
        filename=path,
        data_only=False,
        keep_vba=keep_vba,
        keep_links=True,
        rich_text=True,
    )


def _find_sheet_and_quote_start(workbook):
    candidates = []
    active = workbook.active
    if active is not None:
        candidates.append(active)
    candidates.extend(ws for ws in workbook.worksheets if ws is not active)

    for worksheet in candidates:
        quote_column = _find_first_quote_column(worksheet)
        if quote_column is not None:
            return worksheet, quote_column

    raise ValueError(
        f'Could not find a "{FIRST_QUOTE_LABEL}" header in the first '
        f"{HEADER_SCAN_ROWS} rows of any worksheet."
    )


def _find_first_quote_column(worksheet) -> Optional[int]:
    max_row = min(max(worksheet.max_row, 1), HEADER_SCAN_ROWS)
    for row in worksheet.iter_rows(
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=max(worksheet.max_column, 1),
    ):
        for cell in row:
            if _normalize_header(cell.value) == FIRST_QUOTE_LABEL:
                return cell.column
    return None


def _normalize_header(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").strip().casefold()
    return re.sub(r"\s+", " ", text)


def _find_quote_section_end(worksheet, quote_start: int) -> int:
    remark_column = None
    max_row = min(max(worksheet.max_row, 1), HEADER_SCAN_ROWS)
    for row in worksheet.iter_rows(
        min_row=1,
        max_row=max_row,
        min_col=quote_start,
        max_col=max(worksheet.max_column, quote_start),
    ):
        for cell in row:
            if _normalize_header(cell.value) == "remark":
                remark_column = max(remark_column or 0, cell.column)

    return remark_column or worksheet.max_column


def _replace_quote_section(
    *,
    original_ws,
    quote_ws,
    original_start: int,
    original_end: int,
    quote_start: int,
    quote_end: int,
    source_max_row: int,
    log: Callable[[str], None],
):
    source_width = quote_end - quote_start + 1
    target_end = original_start + source_width - 1

    _warn_about_preserved_formulas(
        original_ws,
        original_start=original_start,
        original_end=original_end,
        log=log,
    )

    _remove_merges_intersecting_columns(original_ws, original_start, original_end)
    _remove_column_dimensions(original_ws, original_start, original_end)

    delete_count = max(0, original_end - original_start + 1)
    if delete_count:
        original_ws.delete_cols(original_start, delete_count)

    _copy_column_dimensions(
        quote_ws,
        original_ws,
        source_start=quote_start,
        source_end=quote_end,
        target_start=original_start,
    )
    _copy_cells(
        quote_ws,
        original_ws,
        source_start=quote_start,
        source_end=quote_end,
        target_start=original_start,
        max_row=source_max_row,
    )
    _copy_merged_ranges(
        quote_ws,
        original_ws,
        source_start=quote_start,
        source_end=quote_end,
        target_start=original_start,
    )

    log(
        f"Rebuilt quotation section through column {get_column_letter(target_end)}."
    )


def _copy_cells(
    source_ws,
    target_ws,
    *,
    source_start: int,
    source_end: int,
    target_start: int,
    max_row: int,
):
    column_offset = target_start - source_start

    for row_number in range(1, max_row + 1):
        for source_column in range(source_start, source_end + 1):
            source_cell = source_ws.cell(row=row_number, column=source_column)
            if isinstance(source_cell, MergedCell):
                continue

            target_column = source_column + column_offset
            target_cell = target_ws.cell(row=row_number, column=target_column)

            value = source_cell.value
            if source_cell.data_type == "f" and isinstance(value, str):
                try:
                    value = Translator(
                        value,
                        origin=source_cell.coordinate,
                    ).translate_formula(target_cell.coordinate)
                except Exception:
                    pass

            target_cell.value = value
            target_cell.data_type = source_cell.data_type
            if source_cell.has_style:
                _copy_style(source_cell, target_cell)
            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)


def _copy_style(source, target):
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def _copy_column_dimensions(
    source_ws,
    target_ws,
    *,
    source_start: int,
    source_end: int,
    target_start: int,
):
    for source_column in range(source_start, source_end + 1):
        source_letter = get_column_letter(source_column)
        target_letter = get_column_letter(
            target_start + (source_column - source_start)
        )
        source_dimension = source_ws.column_dimensions[source_letter]
        target_dimension = target_ws.column_dimensions[target_letter]

        target_dimension.width = source_dimension.width
        target_dimension.hidden = source_dimension.hidden
        target_dimension.bestFit = source_dimension.bestFit
        target_dimension.outlineLevel = source_dimension.outlineLevel
        target_dimension.collapsed = source_dimension.collapsed
        if getattr(source_dimension, "has_style", False):
            _copy_style(source_dimension, target_dimension)


def _copy_merged_ranges(
    source_ws,
    target_ws,
    *,
    source_start: int,
    source_end: int,
    target_start: int,
):
    column_offset = target_start - source_start
    source_ranges = list(source_ws.merged_cells.ranges)

    for merged_range in source_ranges:
        if merged_range.max_col < source_start or merged_range.min_col > source_end:
            continue
        if merged_range.min_col < source_start or merged_range.max_col > source_end:
            continue

        target_ws.merge_cells(
            start_row=merged_range.min_row,
            start_column=merged_range.min_col + column_offset,
            end_row=merged_range.max_row,
            end_column=merged_range.max_col + column_offset,
        )


def _remove_merges_intersecting_columns(
    worksheet,
    start_column: int,
    end_column: int,
):
    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.max_col >= start_column and merged_range.min_col <= end_column:
            worksheet.unmerge_cells(str(merged_range))


def _remove_column_dimensions(worksheet, start_column: int, end_column: int):
    for column_number in range(start_column, end_column + 1):
        worksheet.column_dimensions.pop(get_column_letter(column_number), None)


def _warn_about_preserved_formulas(
    worksheet,
    *,
    original_start: int,
    original_end: int,
    log: Callable[[str], None],
):
    reference_pattern = re.compile(
        r"(?<![A-Z0-9_])\$?([A-Z]{1,3})\$?\d+",
        re.IGNORECASE,
    )

    hits = []
    for row in worksheet.iter_rows(
        min_row=1,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=max(1, original_start - 1),
    ):
        for cell in row:
            value = cell.value
            if cell.data_type == "f" and isinstance(value, str):
                referenced_columns = [
                    _column_letters_to_number(match.group(1))
                    for match in reference_pattern.finditer(value)
                ]
                if any(
                    original_start <= column <= original_end
                    for column in referenced_columns
                ):
                    hits.append(cell.coordinate)
                    if len(hits) >= 5:
                        break
        if len(hits) >= 5:
            break

    if hits:
        log(
            "Warning: preserved formulas may refer to the replaced quotation section "
            f"({', '.join(hits)}). Excel should be used to review their results."
        )


def _column_letters_to_number(letters: str) -> int:
    result = 0
    for character in letters.upper():
        result = result * 26 + (ord(character) - ord("A") + 1)
    return result


def _update_auto_filter(
    target_ws,
    source_ws,
    *,
    target_end: int,
    log: Callable[[str], None],
):
    source_ref = source_ws.auto_filter.ref
    target_ref = target_ws.auto_filter.ref
    chosen_ref = source_ref or target_ref
    if not chosen_ref:
        return

    try:
        source_min_col, min_row, _source_max_col, max_row = range_boundaries(
            chosen_ref
        )
        if target_ref:
            target_min_col, _target_min_row, _target_max_col, _target_max_row = (
                range_boundaries(target_ref)
            )
            source_min_col = target_min_col
        target_ws.auto_filter.ref = (
            f"{get_column_letter(source_min_col)}{min_row}:"
            f"{get_column_letter(target_end)}{max_row}"
        )
        log(f"Updated filter range to {target_ws.auto_filter.ref}.")
    except Exception:
        pass


def _update_print_area_right_edge(worksheet, target_end: int):
    print_area = worksheet.print_area
    if not print_area:
        return

    raw_areas: Iterable[str]
    if hasattr(print_area, "ranges"):
        raw_areas = [str(area) for area in print_area.ranges]
    else:
        raw_areas = [part.strip() for part in str(print_area).split(",")]

    updated_ranges = []
    for raw_area in raw_areas:
        coordinate = raw_area.split("!", 1)[-1].replace("$", "")
        min_col, min_row, _max_col, max_row = range_boundaries(coordinate)
        updated_ranges.append(
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(target_end)}{max_row}"
        )

    if updated_ranges:
        worksheet.print_area = updated_ranges


def _log_style_alignment(original_ws, quote_ws, log: Callable[[str], None]):
    original_styles = _style_values(original_ws)
    quote_styles = _style_values(quote_ws)
    if not original_styles or not quote_styles:
        return

    original_set = set(original_styles)
    quote_set = set(quote_styles)
    missing_in_original = sorted(quote_set - original_set)
    missing_in_quote = sorted(original_set - quote_set)

    if missing_in_original or missing_in_quote:
        log(
            "Warning: the Style lists are not identical. The quotation section is "
            "copied by row position, so verify rows if either workbook was sorted or "
            "had styles inserted/removed."
        )
        if missing_in_original:
            log(
                "Styles only in quote file: "
                + ", ".join(missing_in_original[:8])
                + (" ..." if len(missing_in_original) > 8 else "")
            )
        if missing_in_quote:
            log(
                "Styles only in original file: "
                + ", ".join(missing_in_quote[:8])
                + (" ..." if len(missing_in_quote) > 8 else "")
            )
    elif original_styles != quote_styles:
        log(
            "Warning: both files contain the same Style values but in a different "
            "order. The quotation section is copied by row position; verify the output."
        )
    else:
        log("Style rows align between the two workbooks.")


def _style_values(worksheet) -> list[str]:
    style_column = _find_header_column(worksheet, "style")
    if style_column is None:
        return []

    values = []
    for row_number in range(1, worksheet.max_row + 1):
        value = worksheet.cell(row=row_number, column=style_column).value
        normalized = str(value).strip() if value is not None else ""
        if not normalized or _normalize_header(normalized) == "style":
            continue
        values.append(normalized)
    return values


def _find_header_column(worksheet, header: str) -> Optional[int]:
    normalized_header = _normalize_header(header)
    for row in worksheet.iter_rows(
        min_row=1,
        max_row=min(max(worksheet.max_row, 1), HEADER_SCAN_ROWS),
        min_col=1,
        max_col=max(worksheet.max_column, 1),
    ):
        for cell in row:
            if _normalize_header(cell.value) == normalized_header:
                return cell.column
    return None


def _proposed_output_path(original_path: str) -> str:
    source = Path(original_path)
    candidate = source.with_name(f"{source.stem}{OUTPUT_SUFFIX}{source.suffix}")
    if not candidate.exists():
        return str(candidate)

    counter = 1
    while True:
        candidate = source.with_name(
            f"{source.stem}{OUTPUT_SUFFIX} ({counter}){source.suffix}"
        )
        if not candidate.exists():
            return str(candidate)
        counter += 1


def _save_workbook_atomically(
    workbook,
    *,
    original_path: str,
    output_path: str,
    log: Callable[[str], None],
):
    output_directory = os.path.dirname(output_path) or os.getcwd()
    suffix = Path(output_path).suffix

    with tempfile.NamedTemporaryFile(
        prefix=".pycro_cm_quote_",
        suffix=suffix,
        dir=output_directory,
        delete=False,
    ) as handle:
        temporary_path = handle.name

    patched_path = temporary_path + ".patched"
    try:
        workbook.save(temporary_path)
        if _restore_original_drawing_parts(
            original_path,
            temporary_path,
            patched_path,
        ):
            os.replace(patched_path, temporary_path)
            log("Preserved embedded drawing and media parts from the original file.")
        os.replace(temporary_path, output_path)
    except Exception:
        for path in (temporary_path, patched_path):
            try:
                if os.path.exists(path):
                    os.remove(path)
            except OSError:
                pass
        raise


def _restore_original_drawing_parts(
    original_path: str,
    saved_path: str,
    patched_path: str,
) -> bool:
    drawing_prefixes = ("xl/drawings/", "xl/media/")

    with zipfile.ZipFile(original_path, "r") as original_zip:
        original_parts = {
            name: original_zip.read(name)
            for name in original_zip.namelist()
            if name.startswith(drawing_prefixes)
        }

    if not original_parts:
        return False

    with zipfile.ZipFile(saved_path, "r") as saved_zip:
        saved_names = saved_zip.namelist()
        with zipfile.ZipFile(
            patched_path,
            "w",
            compression=zipfile.ZIP_DEFLATED,
        ) as patched_zip:
            for item in saved_zip.infolist():
                if item.filename in original_parts:
                    continue
                patched_zip.writestr(item, saved_zip.read(item.filename))

            for name, data in original_parts.items():
                patched_zip.writestr(name, data)

    return bool(saved_names)


def _verify_output(
    output_path: str,
    *,
    sheet_name: str,
    expected_quote_column: int,
):
    workbook = _load_workbook_for_editing(output_path)
    try:
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError("The updated worksheet is missing from the saved file.")
        worksheet = workbook[sheet_name]
        actual_quote_column = _find_first_quote_column(worksheet)
        if actual_quote_column != expected_quote_column:
            actual = (
                get_column_letter(actual_quote_column)
                if actual_quote_column is not None
                else "not found"
            )
            raise RuntimeError(
                "The saved workbook failed verification: expected the quotation "
                f"section at {get_column_letter(expected_quote_column)}, found {actual}."
            )
    finally:
        workbook.close()
