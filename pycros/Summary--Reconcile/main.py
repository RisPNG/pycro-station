from __future__ import annotations

import os
import re
import threading
import zipfile
from defusedxml import ElementTree as ET
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Callable, DefaultDict, Dict, Iterable, List, Mapping, MutableMapping, Optional, Sequence, Tuple

from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QFileDialog,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QSizePolicy,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)
from qfluentwidgets import MessageBox, PrimaryPushButton

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


APP_NAME = "Summary Reconcile"
PYCRO_VERSION = "1.2.0"

ROLE_BSD = "bsd"
ROLE_SHIPMENT = "shipment"
ROLE_LOCAL = "local"
ROLE_VN = "vn"

ROLE_LABELS = {
    ROLE_BSD: "Order Control",
    ROLE_SHIPMENT: "Shipment Forecast",
    ROLE_LOCAL: "Weekly Export Local",
    ROLE_VN: "Weekly Export VN",
}

MONTH_NAMES = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]

AmountPair = List[float]
JobMap = DefaultDict[str, AmountPair]
MonthMaps = Dict[str, JobMap]
LogFn = Callable[[str], None]


@dataclass(frozen=True)
class SourceLine:
    full_job: str
    job: str
    qty: float
    amount: float
    invoice: str
    po: str
    source: str
    row_number: int


ShipmentLines = Dict[str, List[SourceLine]]


@dataclass
class ProcessingOptions:
    amount_tolerance: float = 0.01
    quantity_tolerance: float = 0.5


@dataclass
class ProcessingResult:
    output_path: str
    reconciliation_months: List[str]
    bds_counts: Dict[str, int]
    ann_counts: Dict[str, int]
    weekly_sheets: List[str]
    unmatched_weekly_jobs: List[str]
    warnings: List[str]


class MainWidget(QWidget):
    log_message = Signal(str)
    range_detected = Signal(str)
    processing_done = Signal(bool, str, str)

    def __init__(self):
        super().__init__()
        self.setObjectName("summary_reconcile_widget")
        self._paths: Dict[str, QLineEdit] = {}
        self._build_ui()
        self._connect_signals()

    def _build_ui(self):
        self.desc_label = QLabel("", self)
        self.desc_label.setWordWrap(True)
        self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.desc_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.desc_label.setStyleSheet(
            "color: #dcdcdc; background: transparent; padding: 8px; "
            "border: 1px solid #3a3a3a; border-radius: 6px;"
        )
        self.set_long_description("")

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)
        main_layout.addWidget(self.desc_label)

        source_title = QLabel("Source workbooks", self)
        source_title.setStyleSheet("color: #dcdcdc; font-size: 16px; font-weight: 600;")
        main_layout.addWidget(source_title)

        source_grid = QGridLayout()
        source_grid.setHorizontalSpacing(10)
        source_grid.setVerticalSpacing(8)
        for row, role in enumerate((ROLE_BSD, ROLE_SHIPMENT, ROLE_LOCAL, ROLE_VN)):
            label = QLabel(ROLE_LABELS[role], self)
            label.setStyleSheet("color: #dcdcdc;")
            field = QLineEdit(self)
            field.setReadOnly(True)
            field.setPlaceholderText("Select an .xlsx workbook")
            field.setStyleSheet(
                "QLineEdit {background: #1f1f1f; color: #d0d0d0; "
                "border: 1px solid #3a3a3a; border-radius: 6px; padding: 6px;}"
            )
            button = PrimaryPushButton("Browse", self)
            button.clicked.connect(lambda _checked=False, r=role: self._browse_source(r))
            self._paths[role] = field
            source_grid.addWidget(label, row, 0)
            source_grid.addWidget(field, row, 1)
            source_grid.addWidget(button, row, 2)
        source_grid.setColumnStretch(1, 1)
        main_layout.addLayout(source_grid)

        options_title = QLabel("Report settings", self)
        options_title.setStyleSheet("color: #dcdcdc; font-size: 16px; font-weight: 600;")
        main_layout.addWidget(options_title)

        options_grid = QGridLayout()
        options_grid.setHorizontalSpacing(10)
        options_grid.setVerticalSpacing(8)

        self.range_field = QLineEdit(self)
        self.range_field.setReadOnly(True)
        self.range_field.setText("Auto-detected from Ann Forecast during processing")
        self.range_field.setStyleSheet(
            "QLineEdit {background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px; padding: 6px;}"
        )
        options_grid.addWidget(QLabel("Reconciliation months", self), 0, 0)
        options_grid.addWidget(self.range_field, 0, 1, 1, 2)

        self.output_field = QLineEdit(self)
        self.output_field.setReadOnly(True)
        self.output_field.setPlaceholderText("Defaults to the BSD workbook folder")
        self.output_field.setStyleSheet(
            "QLineEdit {background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px; padding: 6px;}"
        )
        output_button = PrimaryPushButton("Browse", self)
        output_button.clicked.connect(self._browse_output_folder)
        options_grid.addWidget(QLabel("Output folder", self), 1, 0)
        options_grid.addWidget(self.output_field, 1, 1)
        options_grid.addWidget(output_button, 1, 2)
        options_grid.setColumnStretch(1, 1)
        main_layout.addLayout(options_grid)

        action_row = QHBoxLayout()
        action_row.addStretch(1)
        self.run_btn = PrimaryPushButton("Build Reconciliation", self)
        self.run_btn.setMinimumWidth(240)
        action_row.addWidget(self.run_btn)
        action_row.addStretch(1)
        main_layout.addLayout(action_row)

        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText("Process log")
        self.log_box.setStyleSheet(
            "QTextEdit {background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )
        main_layout.addWidget(self.log_box, 1)

    def set_long_description(self, text: str):
        clean = (text or "").strip()
        if clean:
            self.desc_label.setText(clean)
            self.desc_label.show()
        else:
            self.desc_label.clear()
            self.desc_label.hide()

    def _connect_signals(self):
        self.run_btn.clicked.connect(self.run_process)
        self.log_message.connect(self._append_log)
        self.range_detected.connect(self.range_field.setText)
        self.processing_done.connect(self._on_processing_done)

    def _browse_source(self, role: str):
        path, _ = QFileDialog.getOpenFileName(
            self,
            f"Select {ROLE_LABELS[role]}",
            "",
            "Excel workbooks (*.xlsx *.xlsm);;All files (*)",
        )
        if path:
            self._paths[role].setText(path)
            if role == ROLE_BSD and not self.output_field.text().strip():
                self.output_field.setText(os.path.dirname(path))

    def _browse_output_folder(self):
        path = QFileDialog.getExistingDirectory(self, "Select output folder")
        if path:
            self.output_field.setText(path)

    def run_process(self):
        paths = {role: field.text().strip() for role, field in self._paths.items()}
        missing = [ROLE_LABELS[role] for role, path in paths.items() if not path]
        if missing:
            MessageBox("Missing files", "Select: " + ", ".join(missing), self).exec()
            return

        output_dir = self.output_field.text().strip() or os.path.dirname(paths[ROLE_BSD])
        options = ProcessingOptions()

        self.log_box.clear()
        self.range_field.setText("Detecting from Ann Forecast...")
        self._set_controls_enabled(False)
        self.log_message.emit("Process starts")

        def worker():
            try:
                result = process_reconciliation(
                    paths,
                    output_dir,
                    options,
                    self.log_message.emit,
                    self.range_detected.emit,
                )
                self.processing_done.emit(True, result.output_path, "")
            except Exception as exc:
                self.log_message.emit(f"ERROR: {exc}")
                self.processing_done.emit(False, "", str(exc))

        threading.Thread(target=worker, daemon=True).start()

    def _set_controls_enabled(self, enabled: bool):
        self.run_btn.setEnabled(enabled)

    def _append_log(self, text: str):
        self.log_box.append(text)
        self.log_box.ensureCursorVisible()

    def _on_processing_done(self, ok: bool, output_path: str, error: str):
        self._set_controls_enabled(True)
        if ok:
            self.log_message.emit(f"Output workbook saved to: {output_path}")
            box = MessageBox(
                "Processing complete",
                f"Created:\n{output_path}",
                self,
            )
        else:
            box = MessageBox("Processing failed", error or "See the process log.", self)
        box.yesButton.setText("OK")
        box.cancelButton.hide()
        box.exec()


def get_widget():
    return MainWidget()


def process_reconciliation(
    paths: Mapping[str, str],
    output_dir: str,
    options: ProcessingOptions,
    log: Optional[LogFn] = None,
    range_callback: Optional[Callable[[str], None]] = None,
) -> ProcessingResult:
    log = log or (lambda _message: None)
    range_callback = range_callback or (lambda _message: None)
    _validate_input_paths(paths)
    _validate_workbook_signatures(paths, log)

    shipment_lines, _shipment_counts, months, shipment_warnings = _read_shipment_forecast(
        paths[ROLE_SHIPMENT],
        log,
    )
    range_text = (
        f"{_month_display(months[0])} to {_month_display(months[-1])} "
        f"({len(months)} month{'s' if len(months) != 1 else ''})"
    )
    range_callback(range_text)
    log(f"Reconciliation months detected from Ann Forecast: {range_text}")

    bds, supplements, bds_counts = _read_bds(paths[ROLE_BSD], months, log)

    canonical_jobs = set()
    for month in months:
        canonical_jobs.update(bds[month].keys())
        canonical_jobs.update(line.job for line in shipment_lines[month])

    weekly_lines, weekly_sheets, unmatched_weekly_jobs, weekly_warnings = _read_weekly_actuals(
        paths[ROLE_LOCAL],
        paths[ROLE_VN],
        months[0],
        canonical_jobs,
        log,
    )

    ann, build_warnings = _build_ann_maps(
        months,
        shipment_lines,
        supplements,
        bds,
        weekly_lines,
        log,
    )
    ann_counts = {month: len(ann[month]) for month in months}
    warnings = list(shipment_warnings) + list(weekly_warnings) + list(build_warnings)

    output_dir = output_dir or os.path.dirname(paths[ROLE_BSD])
    os.makedirs(output_dir, exist_ok=True)
    output_path, generated_at = _timestamped_output_path(output_dir)

    log("Writing result workbook")
    _write_result_workbook(
        output_path=output_path,
        paths=paths,
        options=options,
        months=months,
        bds=bds,
        ann=ann,
        weekly_sheets=weekly_sheets,
        unmatched_weekly_jobs=unmatched_weekly_jobs,
        warnings=warnings,
        generated_at=generated_at,
    )
    log("Workbook creation completed")

    return ProcessingResult(
        output_path=output_path,
        reconciliation_months=list(months),
        bds_counts=bds_counts,
        ann_counts=ann_counts,
        weekly_sheets=weekly_sheets,
        unmatched_weekly_jobs=unmatched_weekly_jobs,
        warnings=warnings,
    )


def _validate_input_paths(paths: Mapping[str, str]):
    for role in (ROLE_BSD, ROLE_SHIPMENT, ROLE_LOCAL, ROLE_VN):
        path = paths.get(role, "")
        if not path:
            raise ValueError(f"Missing {ROLE_LABELS[role]} workbook.")
        if not os.path.isfile(path):
            raise FileNotFoundError(f"{ROLE_LABELS[role]} not found: {path}")
        if Path(path).suffix.lower() not in (".xlsx", ".xlsm"):
            raise ValueError(f"{ROLE_LABELS[role]} must be an .xlsx or .xlsm file.")


def _validate_workbook_signatures(paths: Mapping[str, str], log: LogFn):
    """Validate workbook roles without loading the large workbooks twice.

    Reading workbook metadata directly from the XLSX package avoids retaining a
    second openpyxl parser for the large BSD and VN files before processing.
    The data readers still validate the expected header cells while they read.
    """
    sheet_names = {
        role: _xlsx_sheet_names(path)
        for role, path in paths.items()
    }
    if "pcp2012" not in sheet_names[ROLE_BSD]:
        raise ValueError(
            f"{ROLE_LABELS[ROLE_BSD]} does not contain the expected sheet 'pcp2012'."
        )
    if "SHIPMENTS" not in sheet_names[ROLE_SHIPMENT]:
        raise ValueError(
            f"{ROLE_LABELS[ROLE_SHIPMENT]} does not contain the expected sheet 'SHIPMENTS'."
        )
    for role in (ROLE_LOCAL, ROLE_VN):
        if not any(_parse_weekly_sheet_month(name) for name in sheet_names[role]):
            raise ValueError(f"{ROLE_LABELS[role]} has no recognisable weekly sheets.")
    log("All four workbook layouts were validated")


def _xlsx_sheet_names(path: str) -> List[str]:
    try:
        with zipfile.ZipFile(path) as archive:
            root = ET.fromstring(archive.read("xl/workbook.xml"))
    except (OSError, KeyError, zipfile.BadZipFile, ET.ParseError) as exc:
        raise ValueError(f"Cannot read workbook metadata: {path}") from exc

    namespace = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    return [
        sheet.attrib.get("name", "")
        for sheet in root.findall("m:sheets/m:sheet", namespace)
        if sheet.attrib.get("name")
    ]


def _new_job_map() -> JobMap:
    return defaultdict(lambda: [0.0, 0.0])


def _new_month_maps(months: Sequence[str]) -> MonthMaps:
    return {month: _new_job_map() for month in months}


def _add_pair(target: MutableMapping[str, AmountPair], job: str, qty: float, amount: float):
    pair = target[job]
    pair[0] += qty
    pair[1] += amount


def _read_bds(
    path: str,
    months: Sequence[str],
    log: LogFn,
) -> Tuple[MonthMaps, MonthMaps, Dict[str, int]]:
    log(f"Reading BSD: {os.path.basename(path)}")
    bds = _new_month_maps(months)
    supplements = _new_month_maps(months)

    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        ws = wb["pcp2012"]
        for row in ws.iter_rows(min_row=11, max_col=110, values_only=True):
            job_type = _text(_at(row, 106))
            source_group = _text(_at(row, 41)).upper()  # AO / MCO source group
            if job_type.upper() != "B" or source_group == "SIE_VN":
                continue

            job = _normalise_job(_at(row, 11))
            month = _month_key_from_value(_at(row, 33))  # AG / GAC date
            qty = _number(_at(row, 42))                  # AP / FG
            amount = _number(_at(row, 47))               # AU / Amount
            fty_location = _text(_at(row, 110))          # DF / FtyLoc

            if not job or month not in bds or qty is None or amount is None:
                continue
            _add_pair(bds[month], job, qty, amount)
            if not fty_location:
                _add_pair(supplements[month], job, qty, amount)
    finally:
        wb.close()

    counts = {month: len(bds[month]) for month in months}
    for month in months:
        qty, amount = _map_totals(bds[month])
        log(
            f"BDS {_month_display(month)}: {counts[month]:,} jobs, "
            f"qty {qty:,.0f}, amount {amount:,.2f}"
        )
    return bds, supplements, counts


def _read_shipment_forecast(
    path: str,
    log: LogFn,
) -> Tuple[ShipmentLines, Dict[str, int], List[str], List[str]]:
    """Read Ann Forecast lines and derive the contiguous reconciliation range.

    Exact duplicate source rows are ignored. For each line, actual quantity and
    actual amount are used when both are available; otherwise the planned
    quantity and amount are used. Full job codes are retained for exact weekly
    matching, while reconciliation output remains grouped by the base job code.
    """
    log(f"Reading shipment forecast: {os.path.basename(path)}")
    discovered: Dict[str, List[SourceLine]] = defaultdict(list)
    ignored_missing_month = 0
    duplicate_rows = 0
    duplicate_examples: List[str] = []
    seen_rows = set()

    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        ws = wb["SHIPMENTS"]
        for row_number, row in enumerate(
            ws.iter_rows(min_row=6, max_col=40, values_only=True),
            start=6,
        ):
            full_job = _normalise_job_code(_at(row, 4))
            if not full_job:
                continue
            job = _job_group(full_job)

            planned_qty = _number(_at(row, 5))
            planned_amount = _number(_at(row, 14))
            actual_qty = _number(_at(row, 26))
            actual_amount = _number(_at(row, 27))
            qty, amount = _actual_or_planned(
                actual_qty,
                actual_amount,
                planned_qty,
                planned_amount,
            )
            if qty is None or amount is None:
                continue

            month = _month_key_from_value(_at(row, 20))  # T / PLAN EX-FTY
            if not month:
                ignored_missing_month += 1
                continue

            fingerprint = _row_fingerprint(row)
            if fingerprint in seen_rows:
                duplicate_rows += 1
                if len(duplicate_examples) < 20:
                    duplicate_examples.append(
                        f"SHIPMENTS row {row_number}: {_text(_at(row, 4)) or '(blank job)'}"
                    )
                continue
            seen_rows.add(fingerprint)

            discovered[month].append(
                SourceLine(
                    full_job=full_job,
                    job=job,
                    qty=qty,
                    amount=amount,
                    invoice=_normalise_identifier(_at(row, 6)),
                    po=_normalise_identifier(_at(row, 7)),
                    source="Shipment Forecast",
                    row_number=row_number,
                )
            )
    finally:
        wb.close()

    populated_months = sorted(discovered)
    if not populated_months:
        raise ValueError(
            "Ann Forecast has no usable SHIPMENTS rows with job, quantity, "
            "amount, and PLAN EX-FTY month."
        )

    months = _month_range(populated_months[0], populated_months[-1])
    shipment: ShipmentLines = {month: [] for month in months}
    for month, lines in discovered.items():
        shipment[month] = lines

    warnings: List[str] = []
    missing_months = [month for month in months if month not in discovered]
    if missing_months:
        warning = (
            "Ann Forecast has no usable rows in intermediate month(s): "
            + ", ".join(_month_display(month) for month in missing_months)
        )
        warnings.append(warning)
        log(warning)
    if ignored_missing_month:
        warning = (
            f"Ignored {ignored_missing_month:,} usable Ann row(s) without a "
            "recognisable PLAN EX-FTY month"
        )
        warnings.append(warning)
        log(warning)
    if duplicate_rows:
        warning = f"Ignored {duplicate_rows:,} exact duplicate Ann Forecast row(s)."
        warnings.append(warning)
        log(warning)
        warnings.extend(duplicate_examples)

    counts = {month: len(shipment[month]) for month in months}
    for month in months:
        grouped = _aggregate_lines(shipment[month])
        qty, amount = _map_totals(grouped)
        log(
            f"Shipment {_month_display(month)}: {counts[month]:,} unique lines, "
            f"{len(grouped):,} jobs, qty {qty:,.0f}, amount {amount:,.2f}"
        )
    return shipment, counts, months, warnings


def _read_weekly_actuals(
    local_path: str,
    vn_path: str,
    current_month: str,
    canonical_jobs: Iterable[str],
    log: LogFn,
) -> Tuple[List[SourceLine], List[str], List[str], List[str]]:
    """Read weekly actuals without correcting or fuzzy-matching job numbers."""
    log(f"Reading weekly actual shipments for {_month_display(current_month)}")
    lines: List[SourceLine] = []
    used_sheets: List[str] = []
    warnings: List[str] = []
    ignored_summary_rows = 0

    configs = [
        # path, label, job, planned qty, planned amount, actual qty, actual amount, invoice, PO, first row
        (local_path, "Local", 2, 3, 9, 17, 18, 4, 5, 7),
        (vn_path, "VN", 4, 5, 14, 27, 28, 6, 7, 6),
    ]
    for (
        path,
        label,
        job_col,
        planned_qty_col,
        planned_amount_col,
        actual_qty_col,
        actual_amount_col,
        invoice_col,
        po_col,
        start_row,
    ) in configs:
        wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        try:
            selected = [
                name for name in wb.sheetnames
                if _parse_weekly_sheet_month(name) == current_month
            ]
            if not selected:
                warning = (
                    f"No {_month_display(current_month)} weekly sheet found in {label} workbook."
                )
                warnings.append(warning)
                log(warning)
                continue
            selected.sort(key=_weekly_sheet_sort_key)
            for sheet_name in selected:
                ws = wb[sheet_name]
                max_col = max(
                    job_col,
                    planned_qty_col,
                    planned_amount_col,
                    actual_qty_col,
                    actual_amount_col,
                    invoice_col,
                    po_col,
                )
                for row_number, row in enumerate(
                    ws.iter_rows(min_row=start_row, max_col=max_col, values_only=True),
                    start=start_row,
                ):
                    raw_job = _text(_at(row, job_col))
                    if _is_weekly_summary_job(raw_job):
                        ignored_summary_rows += 1
                        continue
                    full_job = _normalise_job_code(raw_job)
                    if not full_job:
                        continue
                    job = _job_group(full_job)
                    qty, amount = _actual_or_planned(
                        _number(_at(row, actual_qty_col)),
                        _number(_at(row, actual_amount_col)),
                        _number(_at(row, planned_qty_col)),
                        _number(_at(row, planned_amount_col)),
                    )
                    if qty is None or amount is None:
                        continue
                    if abs(qty) < 1e-12 and abs(amount) < 1e-12:
                        continue
                    lines.append(
                        SourceLine(
                            full_job=full_job,
                            job=job,
                            qty=qty,
                            amount=amount,
                            invoice=_normalise_identifier(_at(row, invoice_col)),
                            po=_normalise_identifier(_at(row, po_col)),
                            source=f"{label}: {sheet_name}",
                            row_number=row_number,
                        )
                    )
                used_sheets.append(f"{label}: {sheet_name}")
        finally:
            wb.close()

    canonical = set(canonical_jobs)
    unmatched_weekly_jobs = sorted(
        {
            line.job
            for line in lines
            if line.job not in canonical and line.job not in {"SAMPLE", "SAMPLES"}
        }
    )
    for job in unmatched_weekly_jobs:
        warning = (
            f"Weekly job '{job}' has no exact job match in BDS or Shipment Forecast. "
            "No automatic correction was applied."
        )
        warnings.append(warning)
        log(warning)

    if ignored_summary_rows:
        log(f"Ignored {ignored_summary_rows:,} TOTALWEEK summary row(s)")

    grouped = _aggregate_lines(lines)
    qty, amount = _map_totals(grouped)
    log(
        f"Weekly actual: {len(lines):,} lines / {len(grouped):,} jobs from "
        f"{len(used_sheets)} sheets, qty {qty:,.0f}, amount {amount:,.2f}"
    )
    return lines, used_sheets, unmatched_weekly_jobs, warnings


def _build_ann_maps(
    months: Sequence[str],
    shipment: ShipmentLines,
    supplements: MonthMaps,
    bds: MonthMaps,
    current_actual: Sequence[SourceLine],
    log: LogFn,
) -> Tuple[MonthMaps, List[str]]:
    """Build Ann using confirmed source precedence rules.

    The first two detected months use Shipment Forecast. The first month also
    uses weekly actuals, replacing only the exact forecast line already shipped
    and retaining all remaining forecast lines. From the third month onward,
    Ann follows BDS exactly.
    """
    ann = _new_month_maps(months)
    warnings: List[str] = []

    for index, month in enumerate(months):
        if index >= 2:
            for job, pair in bds[month].items():
                ann[month][job] = pair.copy()
            log(f"Ann {_month_display(month)} follows BDS (month {index + 1} onward)")
            continue

        remaining = list(shipment[month])
        weekly_for_month: Sequence[SourceLine] = current_actual if index == 0 else ()
        if weekly_for_month:
            remaining, match_warnings = _remove_shipped_forecast_lines(
                remaining,
                weekly_for_month,
            )
            warnings.extend(match_warnings)

        for line in remaining:
            _add_pair(ann[month], line.job, line.qty, line.amount)
        for line in weekly_for_month:
            _add_pair(ann[month], line.job, line.qty, line.amount)

        # The Shipment Forecast is VN-oriented. Preserve the established local
        # supplement rule only for the two forecast-driven months.
        for job, pair in supplements[month].items():
            if job not in ann[month]:
                ann[month][job] = pair.copy()

    for warning in warnings:
        log(warning)
    return ann, warnings


def _remove_shipped_forecast_lines(
    forecast_lines: Sequence[SourceLine],
    weekly_lines: Sequence[SourceLine],
) -> Tuple[List[SourceLine], List[str]]:
    remaining = list(forecast_lines)
    warnings: List[str] = []

    for weekly in weekly_lines:
        candidates = [
            index
            for index, forecast in enumerate(remaining)
            if _same_source_line_identity(forecast, weekly)
        ]
        if candidates:
            # Matching is exact on the full job number and shipment identifiers.
            # When repeated identical lines exist, consume one forecast occurrence
            # for each weekly occurrence, preserving multiset quantities.
            remaining.pop(candidates[0])
    return remaining, warnings


def _same_source_line_identity(forecast: SourceLine, weekly: SourceLine) -> bool:
    if forecast.full_job != weekly.full_job:
        return False

    invoice_match = bool(forecast.invoice and weekly.invoice and forecast.invoice == weekly.invoice)
    po_match = bool(forecast.po and weekly.po and forecast.po == weekly.po)
    qty_match = abs(forecast.qty - weekly.qty) <= 0.5
    amount_match = abs(forecast.amount - weekly.amount) <= 0.01

    # Exact job plus a shared invoice or PO identifies the same shipment line.
    # Quantity and amount may legitimately change when planned values become
    # actual values. If neither identifier exists, require an exact value match.
    if invoice_match or po_match:
        return True
    if not forecast.invoice and not weekly.invoice and not forecast.po and not weekly.po:
        return qty_match and amount_match
    return False


def _write_result_workbook(
    output_path: str,
    paths: Mapping[str, str],
    options: ProcessingOptions,
    months: Sequence[str],
    bds: MonthMaps,
    ann: MonthMaps,
    weekly_sheets: Sequence[str],
    unmatched_weekly_jobs: Sequence[str],
    warnings: Sequence[str],
    generated_at: datetime,
):
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    reconciliations: Dict[str, List[Tuple[str, float, float, float, float, str]]] = {}
    for month in months:
        reconciliations[month] = _build_reconciliation_rows(
            month,
            months,
            bds,
            ann,
            options.quantity_tolerance,
            options.amount_tolerance,
        )

    _write_summary_sheet(wb, generated_at, months, bds, ann, reconciliations)
    for month in months:
        _write_recon_sheet(wb, month, generated_at, reconciliations[month])
    _write_audit_sheet(
        wb,
        paths,
        months,
        bds,
        ann,
        weekly_sheets,
        unmatched_weekly_jobs,
        warnings,
        generated_at,
    )

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass
    wb.active = 0
    wb.save(output_path)


def _build_reconciliation_rows(
    month: str,
    months: Sequence[str],
    bds: MonthMaps,
    ann: MonthMaps,
    qty_tolerance: float,
    amount_tolerance: float,
) -> List[Tuple[str, float, float, float, float, str]]:
    jobs = sorted(set(bds[month]) | set(ann[month]))
    rows = []
    for job in jobs:
        bq, ba = bds[month].get(job, [0.0, 0.0])
        aq, aa = ann[month].get(job, [0.0, 0.0])
        remark = _classify_variance(
            job,
            month,
            months,
            bds,
            ann,
            qty_tolerance,
            amount_tolerance,
        )
        rows.append((job, bq, ba, aq, aa, remark))
    return rows


def _classify_variance(
    job: str,
    month: str,
    months: Sequence[str],
    bds: MonthMaps,
    ann: MonthMaps,
    qty_tolerance: float,
    amount_tolerance: float,
) -> str:
    bq, ba = bds[month].get(job, [0.0, 0.0])
    aq, aa = ann[month].get(job, [0.0, 0.0])
    dq = bq - aq
    da = ba - aa
    if abs(dq) <= qty_tolerance and abs(da) <= amount_tolerance:
        return ""

    index = months.index(month)
    signal = dq if abs(dq) > qty_tolerance else da
    prev_signal = None
    next_signal = None
    if index > 0:
        pbq, pba = bds[months[index - 1]].get(job, [0.0, 0.0])
        paq, paa = ann[months[index - 1]].get(job, [0.0, 0.0])
        prev_signal = (pbq - paq) if abs(pbq - paq) > qty_tolerance else (pba - paa)
    if index + 1 < len(months):
        nbq, nba = bds[months[index + 1]].get(job, [0.0, 0.0])
        naq, naa = ann[months[index + 1]].get(job, [0.0, 0.0])
        next_signal = (nbq - naq) if abs(nbq - naq) > qty_tolerance else (nba - naa)

    if signal > 0 and next_signal is not None and next_signal < 0:
        return f"Delay ship fr {_month_display(month)} to {_month_display(months[index + 1])}"
    if signal < 0 and next_signal is not None and next_signal > 0:
        return f"Early ship fr {_month_display(months[index + 1])} to {_month_display(month)}"
    if signal < 0 and prev_signal is not None and prev_signal > 0:
        return f"Delay ship fr {_month_display(months[index - 1])} to {_month_display(month)}"
    if signal > 0 and prev_signal is not None and prev_signal < 0:
        return f"Early ship fr {_month_display(month)} to {_month_display(months[index - 1])}"

    upper_job = job.upper()
    if "SAMPLE" in upper_job:
        return "Sample"
    if abs(bq) <= qty_tolerance and abs(ba) <= amount_tolerance:
        return "Demand Pull"
    if abs(aq) <= qty_tolerance and abs(aa) <= amount_tolerance:
        return "Missing"
    if abs(dq) <= qty_tolerance and abs(da) > amount_tolerance:
        return "Price Discrepancy"
    if abs(dq) > qty_tolerance:
        return "Short Ship" if dq > 0 else "Quantity Variance"
    return "Amount Variance"


def _write_summary_sheet(
    wb: Workbook,
    generated_at: datetime,
    months: Sequence[str],
    bds: MonthMaps,
    ann: MonthMaps,
    reconciliations: Mapping[str, Sequence[Tuple[str, float, float, float, float, str]]],
):
    fiscal_year = _fiscal_year_for_month(months[0])
    ws = wb.create_sheet(f"Summary FYE {fiscal_year}")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B6"

    title_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="5B9BD5")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="B7C9D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("B1:E1")
    ws["B1"] = f"Summary Reconcile - FYE {fiscal_year}"
    ws["B1"].fill = title_fill
    ws["B1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["B1"].alignment = Alignment(horizontal="center")
    ws["B3"] = "Generated"
    ws["C3"] = generated_at
    ws["C3"].number_format = "yyyy-mm-dd hh:mm:ss"
    ws["D3"] = f"Generated {generated_at.strftime('%d %b %Y %H:%M:%S')}"

    standard_reasons = sorted(
        {
            row_data[5]
            for month in months
            for row_data in reconciliations[month]
            if row_data[5] and row_data[5] != "Fx Adjustment"
        }
    )
    if "Price Discrepancy" not in standard_reasons:
        standard_reasons.append("Price Discrepancy")
        standard_reasons.sort()
    price_index = standard_reasons.index("Price Discrepancy")
    reasons = standard_reasons[: price_index + 1] + ["Fx Adjustment"] + standard_reasons[price_index + 1 :]

    start_col = 7
    variance_end_col = start_col + len(months)
    reason_start_row = 6
    reason_end_row = reason_start_row + len(reasons) - 1
    total_variance_row = reason_end_row + 1

    ws.merge_cells(
        start_row=3,
        start_column=start_col,
        end_row=3,
        end_column=variance_end_col,
    )
    ws.cell(row=3, column=start_col, value="Sales Forecast Variance Analysis")
    ws.cell(row=3, column=start_col).fill = title_fill
    ws.cell(row=3, column=start_col).font = white_font
    ws.cell(row=3, column=start_col).alignment = Alignment(horizontal="center")

    variance_headers = ["Variance reason"] + [_month_display(m) for m in months]
    for offset, value in enumerate(variance_headers):
        cell = ws.cell(row=5, column=start_col + offset, value=value)
        cell.fill = header_fill
        cell.font = white_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for index, reason in enumerate(reasons, start=reason_start_row):
        reason_cell = ws.cell(row=index, column=start_col, value=reason)
        if reason == "Fx Adjustment":
            reason_cell.fill = input_fill
            reason_cell.font = bold_font
        for month_offset, month in enumerate(months, start=1):
            col = start_col + month_offset
            cell = ws.cell(row=index, column=col)
            if reason == "Fx Adjustment":
                cell.value = 0.0
                cell.fill = input_fill
                cell.font = bold_font
            else:
                recon_name = f"Recon {_month_display(month)}".replace("'", "''")
                last_recon_row = 3 + len(reconciliations[month])
                cell.value = (
                    f"=SUMIF('{recon_name}'!$H$4:$H${last_recon_row},"
                    f"${get_column_letter(start_col)}{index},"
                    f"'{recon_name}'!$G$4:$G${last_recon_row})"
                )
        for col in range(start_col, variance_end_col + 1):
            ws.cell(row=index, column=col).border = border

    ws.cell(row=total_variance_row, column=start_col, value="Total")
    for month_offset, _month in enumerate(months, start=1):
        col = start_col + month_offset
        letter = get_column_letter(col)
        ws.cell(
            row=total_variance_row,
            column=col,
            value=f"=SUM({letter}{reason_start_row}:{letter}{reason_end_row})",
        )
    for col in range(start_col, variance_end_col + 1):
        cell = ws.cell(row=total_variance_row, column=col)
        cell.fill = section_fill
        cell.font = bold_font
        cell.border = border

    headers = ["Month", "BDS", "Ann Report", "BDS - Ann"]
    for col, value in enumerate(headers, start=2):
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    row = 6
    for month_offset, key in enumerate(months, start=1):
        ws.cell(row=row, column=2, value=_month_display(key))
        recon_name = f"Recon {_month_display(key)}".replace("'", "''")
        variance_total_col = get_column_letter(start_col + month_offset)
        ws.cell(row=row, column=3, value=f"='{recon_name}'!$C$1")
        ws.cell(row=row, column=5, value=f"={variance_total_col}${total_variance_row}")
        ws.cell(row=row, column=4, value=f"=C{row}-E{row}")
        for col in range(2, 6):
            ws.cell(row=row, column=col).border = border
        row += 1

    subtotal_row = row
    ws.cell(row=subtotal_row, column=2, value="Grand Total")
    ws.cell(row=subtotal_row, column=3, value=f"=SUM(C6:C{subtotal_row - 1})")
    ws.cell(row=subtotal_row, column=4, value=f"=SUM(D6:D{subtotal_row - 1})")
    ws.cell(row=subtotal_row, column=5, value=f"=SUM(E6:E{subtotal_row - 1})")
    for col in range(2, 6):
        cell = ws.cell(row=subtotal_row, column=col)
        cell.fill = section_fill
        cell.font = bold_font
        cell.border = border

    for col in (3, 4, 5):
        ws.column_dimensions[get_column_letter(col)].width = 18
        for cell in ws[get_column_letter(col)]:
            cell.number_format = '#,##0.00;[Red]-#,##0.00'
    for col in range(start_col + 1, variance_end_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
        for cell in ws[get_column_letter(col)]:
            cell.number_format = '#,##0.00;[Red]-#,##0.00'
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions[get_column_letter(start_col)].width = 44


def _write_recon_sheet(
    wb: Workbook,
    month: str,
    generated_at: datetime,
    rows: Sequence[Tuple[str, float, float, float, float, str]],
):
    ws = wb.create_sheet(f"Recon {_month_display(month)}")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    header_fill = PatternFill("solid", fgColor="5B9BD5")
    group_fill = PatternFill("solid", fgColor="D9EAF7")
    total_fill = PatternFill("solid", fgColor="E2F0D9")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="C8D4DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "Totals"
    for col in range(2, 8):
        letter = get_column_letter(col)
        ws.cell(row=1, column=col, value=f"=SUBTOTAL(9,{letter}4:{letter}{3 + len(rows)})")
        ws.cell(row=1, column=col).fill = total_fill
        ws.cell(row=1, column=col).font = bold_font
        ws.cell(row=1, column=col).number_format = '#,##0.00;[Red]-#,##0.00'
    ws["A1"].fill = total_fill
    ws["A1"].font = bold_font

    label = generated_at.strftime("%d%m%y")
    ws.merge_cells("B2:C2")
    ws.merge_cells("D2:E2")
    ws["B2"] = f"BDS {label}"
    ws["D2"] = f"Ann {label}"
    for cell in (ws["B2"], ws["D2"]):
        cell.fill = group_fill
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center")

    headers = [
        "Row Labels", "Sum of Qty", "Sum of Amt", "Sum of Qty", "Sum of Amt",
        "Total Sum of Qty", "Total Sum of Amt", "Remark",
    ]
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=value)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_index, (job, bq, ba, aq, aa, remark) in enumerate(rows, start=4):
        values = [job, bq or None, ba or None, aq or None, aa or None]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row_index, column=col, value=value)
        ws.cell(row=row_index, column=6, value=f"=B{row_index}-D{row_index}")
        ws.cell(row=row_index, column=7, value=f"=C{row_index}-E{row_index}")
        ws.cell(row=row_index, column=8, value=remark or None)
        for col in range(1, 9):
            ws.cell(row=row_index, column=col).border = border

    last_row = 3 + len(rows)
    ws.auto_filter.ref = f"A3:H{last_row}"
    ws.auto_filter.add_filter_column(7, [])

    for col in (2, 4, 6):
        for cell in ws[get_column_letter(col)][3:]:
            cell.number_format = '#,##0;[Red]-#,##0'
    for col in (3, 5, 7):
        for cell in ws[get_column_letter(col)][3:]:
            cell.number_format = '#,##0.00;[Red]-#,##0.00'

    if last_row >= 4:
        red_fill = PatternFill("solid", fgColor="FCE4D6")
        ws.conditional_formatting.add(
            f"F4:G{last_row}",
            CellIsRule(operator="notEqual", formula=["0"], fill=red_fill),
        )

    widths = {"A": 22, "B": 15, "C": 17, "D": 15, "E": 17, "F": 18, "G": 20, "H": 42}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[3].height = 32


def _write_audit_sheet(
    wb: Workbook,
    paths: Mapping[str, str],
    months: Sequence[str],
    bds: MonthMaps,
    ann: MonthMaps,
    weekly_sheets: Sequence[str],
    unmatched_weekly_jobs: Sequence[str],
    warnings: Sequence[str],
    generated_at: datetime,
):
    ws = wb.create_sheet("Processing Log")
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="1F4E78")
    sub_fill = PatternFill("solid", fgColor="D9EAF7")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    ws["A1"] = "Summary Reconcile Processing Log"
    ws["A1"].fill = header_fill
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=15)
    ws.merge_cells("A1:D1")

    rows = [
        ("Pycro version", PYCRO_VERSION),
        ("Generated", generated_at),
        ("Detected Ann start month", _month_display(months[0])),
        ("Detected Ann end month", _month_display(months[-1])),
        ("Detected reconciliation months", len(months)),
    ]
    row = 3
    for key, value in rows:
        ws.cell(row=row, column=1, value=key).font = bold_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Input workbooks").fill = sub_fill
    ws.cell(row=row, column=1).font = bold_font
    row += 1
    for role in (ROLE_BSD, ROLE_SHIPMENT, ROLE_LOCAL, ROLE_VN):
        ws.cell(row=row, column=1, value=ROLE_LABELS[role])
        ws.cell(row=row, column=2, value=os.path.abspath(paths[role]))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Monthly totals").fill = sub_fill
    ws.cell(row=row, column=1).font = bold_font
    row += 1
    for col, value in enumerate(["Month", "BDS Qty", "BDS Amt", "Ann Qty", "Ann Amt", "Variance"], start=1):
        ws.cell(row=row, column=col, value=value).font = bold_font
    row += 1
    for month in months:
        bq, ba = _map_totals(bds[month])
        aq, aa = _map_totals(ann[month])
        for col, value in enumerate([_month_display(month), bq, ba, aq, aa, ba - aa], start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Weekly sheets used").fill = sub_fill
    ws.cell(row=row, column=1).font = bold_font
    row += 1
    for item in weekly_sheets:
        ws.cell(row=row, column=1, value=item)
        row += 1

    if unmatched_weekly_jobs:
        row += 1
        ws.cell(row=row, column=1, value="Unmatched weekly job numbers").fill = sub_fill
        ws.cell(row=row, column=1).font = bold_font
        row += 1
        ws.cell(row=row, column=1, value="Exact job number").font = bold_font
        ws.cell(row=row, column=2, value="Action required").font = bold_font
        row += 1
        for job in unmatched_weekly_jobs:
            ws.cell(row=row, column=1, value=job)
            ws.cell(
                row=row,
                column=2,
                value="Correct the source workbook. The Pycro does not guess or merge job numbers.",
            )
            row += 1

    if warnings:
        row += 1
        ws.cell(row=row, column=1, value="Warnings / assumptions").fill = sub_fill
        ws.cell(row=row, column=1).font = bold_font
        row += 1
        for warning in warnings:
            ws.cell(row=row, column=1, value=warning)
            row += 1

    row += 1
    ws.cell(row=row, column=1, value="Processing rules").fill = sub_fill
    ws.cell(row=row, column=1).font = bold_font
    rules = [
        "BDS: pcp2012, Jobtype B, excluding source group SIE_VN, grouped by GAC date and base job number.",
        "Reconciliation range: earliest through latest usable Ann Forecast PLAN EX-FTY month, including empty calendar months between them.",
        "No fuzzy job correction: job numbers must match exactly. Unmatched weekly jobs remain separate and are reported for source correction.",
        "Weekly summary rows beginning TOTALWEEK are ignored; SAMPLES is included when it has usable quantity and amount.",
        "Actual quantity and amount are used when both are available; otherwise planned quantity and amount are used.",
        "Exact duplicate Ann Forecast rows are counted once.",
        "First two detected months: use Shipment Forecast, supplemented by unassigned/local BDS jobs absent from the forecast.",
        "First detected month: weekly actuals replace only the exact forecast shipment line already actual; remaining forecast lines are retained.",
        "Third detected month onward: Ann follows BDS exactly.",
        "Movement remarks are inferred by matching opposite-sign job variances across adjacent months.",
        "The fiscal summary reserves an editable Fx Adjustment row directly below Price Discrepancy. Its month values start at zero and are summed with all other variance reasons. Enter the signed variance shown in the approved reconciliation; a negative value increases Ann and reduces BDS - Ann.",
    ]
    row += 1
    for rule in rules:
        ws.cell(row=row, column=1, value=rule)
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 90
    for col in ("C", "D", "E", "F"):
        ws.column_dimensions[col].width = 18
    for cells in ws.iter_rows():
        for cell in cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.sheet_state = "hidden"


def _variance_totals_by_reason(
    rows: Sequence[Tuple[str, float, float, float, float, str]],
) -> Dict[str, float]:
    totals: DefaultDict[str, float] = defaultdict(float)
    for _job, _bq, ba, _aq, aa, remark in rows:
        if remark:
            totals[remark] += ba - aa
    return dict(totals)


def _map_totals(data: Mapping[str, AmountPair]) -> Tuple[float, float]:
    return (
        sum(pair[0] for pair in data.values()),
        sum(pair[1] for pair in data.values()),
    )


def _timestamped_output_path(output_dir: str) -> Tuple[str, datetime]:
    """Return an unused timestamp-based output path and its report timestamp.

    The filename always follows Summary_Reconcile_yyyymmdd_hhmmss.xlsx. If a
    file already exists for the current second, the timestamp is advanced until
    an unused name is found rather than adding a non-standard suffix.
    """
    generated_at = datetime.now().replace(microsecond=0)
    while True:
        filename = f"Summary_Reconcile_{generated_at.strftime('%Y%m%d_%H%M%S')}.xlsx"
        candidate = os.path.join(output_dir, filename)
        if not os.path.exists(candidate):
            return candidate, generated_at
        generated_at += timedelta(seconds=1)


def _at(row: Sequence[object], column_1_based: int):
    index = column_1_based - 1
    return row[index] if 0 <= index < len(row) else None


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _number(value: object) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _normalise_job_code(value: object) -> str:
    """Validate a job code without guessing or correcting its characters.

    Case, outer whitespace, and Unicode dash presentation are standardised. A
    malformed code is rejected rather than repaired. Valid but different codes
    such as AM06008MS and AM060008MS remain distinct.
    """
    text = _text(value).upper().replace("–", "-").replace("—", "-")
    if not text:
        return ""
    if text in {"SAMPLE", "SAMPLES"}:
        return text
    if len(text) < 6 or not any(char.isdigit() for char in text):
        return ""
    if not re.fullmatch(r"[A-Z0-9]+(?:-[A-Z0-9]+)*", text):
        return ""
    return text


def _job_group(full_job: str) -> str:
    if full_job in {"SAMPLE", "SAMPLES"}:
        return full_job
    return full_job.split("-", 1)[0]


def _normalise_job(value: object) -> str:
    full_job = _normalise_job_code(value)
    return _job_group(full_job) if full_job else ""


def _normalise_identifier(value: object) -> str:
    return re.sub(r"\s+", "", _text(value).upper())


def _is_weekly_summary_job(value: object) -> bool:
    compact = re.sub(r"[^A-Z0-9]", "", _text(value).upper())
    return compact.startswith("TOTALWEEK")


def _actual_or_planned(
    actual_qty: Optional[float],
    actual_amount: Optional[float],
    planned_qty: Optional[float],
    planned_amount: Optional[float],
) -> Tuple[Optional[float], Optional[float]]:
    actual_available = (
        actual_qty is not None
        and actual_amount is not None
        and abs(actual_qty) > 1e-12
        and abs(actual_amount) > 1e-12
    )
    if actual_available:
        return actual_qty, actual_amount
    return planned_qty, planned_amount


def _row_fingerprint(row: Sequence[object]) -> Tuple[object, ...]:
    values: List[object] = []
    for value in row:
        if isinstance(value, datetime):
            values.append(("datetime", value.isoformat()))
        elif isinstance(value, date):
            values.append(("date", value.isoformat()))
        elif isinstance(value, float):
            values.append(("number", round(value, 10)))
        elif value is None:
            values.append(None)
        else:
            values.append(("text", str(value).strip()))
    return tuple(values)


def _aggregate_lines(lines: Sequence[SourceLine]) -> JobMap:
    result = _new_job_map()
    for line in lines:
        _add_pair(result, line.job, line.qty, line.amount)
    return result


def _month_key_from_value(value: object) -> Optional[str]:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, date):
        return value.strftime("%Y-%m")
    if isinstance(value, (int, float)):
        try:
            converted = datetime(1899, 12, 30) + timedelta(days=float(value))
            return converted.strftime("%Y-%m")
        except (OverflowError, ValueError):
            return None
    text = _text(value)
    match = re.search(r"(20\d{2})[-/](0?[1-9]|1[0-2])", text)
    if match:
        return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"
    return None


def _month_range(start_month: str, end_month: str) -> List[str]:
    start_year, start_number = (int(part) for part in start_month.split("-"))
    end_year, end_number = (int(part) for part in end_month.split("-"))
    start_index = start_year * 12 + start_number - 1
    end_index = end_year * 12 + end_number - 1
    if end_index < start_index:
        raise ValueError("Ann Forecast end month is earlier than its start month.")
    result = []
    for absolute in range(start_index, end_index + 1):
        year, month_index = divmod(absolute, 12)
        result.append(f"{year:04d}-{month_index + 1:02d}")
    return result


def _fiscal_year_for_month(month_key: str) -> int:
    year, month = (int(part) for part in month_key.split("-"))
    return year + 1 if month >= 5 else year


def _month_key(value: date) -> str:
    return value.strftime("%Y-%m")


def _month_display(month_key: str) -> str:
    year, month = (int(part) for part in month_key.split("-"))
    return f"{MONTH_NAMES[month - 1]}'{str(year)[2:]}"


def _add_months(value: date, offset: int) -> date:
    absolute = value.year * 12 + (value.month - 1) + offset
    year, month_index = divmod(absolute, 12)
    month = month_index + 1
    return date(year, month, min(value.day, _days_in_month(year, month)))


def _days_in_month(year: int, month: int) -> int:
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)
    return (next_month - date(year, month, 1)).days


def _parse_weekly_sheet_month(sheet_name: str) -> Optional[str]:
    clean = sheet_name.strip()
    match = re.match(r"^([A-Za-z]{3})'(\d{2})\s+W[kK]\s*(\d+)", clean)
    if not match:
        return None
    month_text, year_text, _week = match.groups()
    try:
        month = [name.lower() for name in MONTH_NAMES].index(month_text.lower()) + 1
    except ValueError:
        return None
    return f"20{year_text}-{month:02d}"


def _weekly_sheet_sort_key(sheet_name: str) -> Tuple[int, str]:
    match = re.search(r"W[kK]\s*(\d+)", sheet_name)
    return (int(match.group(1)) if match else 999, sheet_name)
