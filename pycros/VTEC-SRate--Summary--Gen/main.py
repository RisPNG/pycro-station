#!/usr/bin/env python3
from __future__ import annotations

import argparse
import os
import re
import threading
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

try:
    from PySide6.QtCore import Qt, Signal
    from PySide6.QtWidgets import (
        QFileDialog,
        QHBoxLayout,
        QLabel,
        QSizePolicy,
        QTextEdit,
        QVBoxLayout,
        QWidget,
    )
    from qfluentwidgets import MessageBox, PrimaryPushButton

    GUI_AVAILABLE = True
except Exception:
    GUI_AVAILABLE = False


SCAN_MAX_COL = 64
HEADER_SCAN_ROWS = 10
VND_NUMBER_FORMAT = "#,##0"
USD_NUMBER_FORMAT = "#,###.00"
GENERAL_RATE_FORMAT = "#,##0.00"
DATE_NUMBER_FORMAT = "yyyy/mm/dd"
DATA_ROW_HEIGHT = 20
BLACK = "000000"
WHITE = "FFFFFF"
HEADER_FILL = "1F4E78"
STANDARD_RATE_FILL = "DDEBF7"
STATUS_MATCH_FILL = "C6EFCE"
STATUS_MATCH_FONT = "006100"
STATUS_MISMATCH_FILL = "FFC7CE"
STATUS_MISMATCH_FONT = "9C0006"
STATUS_NEUTRAL_FILL = "FFF2CC"
STATUS_NEUTRAL_FONT = "7F6000"
MONTH_ABBR = {
    1: "JAN",
    2: "FEB",
    3: "MAR",
    4: "APR",
    5: "MAY",
    6: "JUN",
    7: "JUL",
    8: "AUG",
    9: "SEP",
    10: "OCT",
    11: "NOV",
    12: "DEC",
}


@dataclass(frozen=True)
class ColumnSpec:
    key: str
    title: str
    required_header: bool
    matcher: Callable[[str], bool]


@dataclass
class SourceRow:
    source_file: Path
    sheet_name: str
    row_number: int
    payment_to_supplier_date: date
    vat_invoice_date: date
    values: Dict[str, object]


@dataclass(frozen=True)
class SRateRange:
    source_file: Path
    sheet_name: str
    row_number: int
    from_date: date
    to_date: date
    standard_rate: float


@dataclass
class AuditLog:
    folders_without_xlsx: List[str] = field(default_factory=list)
    multi_sheet_files: List[str] = field(default_factory=list)
    source_file_errors: List[str] = field(default_factory=list)
    missing_required_headers: List[str] = field(default_factory=list)
    file_level_row_issues: List[str] = field(default_factory=list)
    invalid_payment_to_supplier_rows: List[str] = field(default_factory=list)
    srate_file_errors: List[str] = field(default_factory=list)
    srate_missing_headers: List[str] = field(default_factory=list)
    invalid_srate_rows: List[str] = field(default_factory=list)
    duplicate_srate_ranges: List[str] = field(default_factory=list)
    missing_standard_rates: Dict[str, List[str]] = field(default_factory=lambda: defaultdict(list))
    ambiguous_standard_rate_matches: Dict[str, List[str]] = field(default_factory=lambda: defaultdict(list))


@dataclass
class ProcessResult:
    workbook_path: Path
    log_path: Path
    source_files_found: int
    source_files_imported: int
    source_files_skipped: int
    rows_imported: int
    rows_ignored_non_date_payment_to_supplier: int
    month_sheet_count: int


def _emit(log_emit: Callable[[str], None] | None, text: str):
    if callable(log_emit):
        try:
            log_emit(text)
            return
        except Exception:
            pass
    print(text)


def ensure_unique_path(path: Path) -> Path:
    if not path.exists():
        return path

    counter = 1
    while True:
        candidate = path.with_name(f"{path.stem} ({counter}){path.suffix}")
        if not candidate.exists():
            return candidate
        counter += 1


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text.upper()


def starts_with_header(actual: str, expected: str) -> bool:
    if not actual or not expected:
        return False
    if actual == expected:
        return True
    if not actual.startswith(expected):
        return False
    if len(actual) == len(expected):
        return True
    next_char = actual[len(expected)]
    return not next_char.isalnum()


def match_po_amount(norm_header: str, currency_suffix: str) -> bool:
    return norm_header.startswith("PO AMOUNT") and norm_header.endswith(currency_suffix)


def match_srate_header(norm_header: str) -> bool:
    return (
        starts_with_header(norm_header, "S RATE")
        or starts_with_header(norm_header, "SRATE")
        or starts_with_header(norm_header, "STANDARD RATE")
    )


SOURCE_COLUMNS: List[ColumnSpec] = [
    ColumnSpec("no", "No", False, lambda s: starts_with_header(s, "NO")),
    ColumnSpec("supplier_name", "Supplier Name", False, lambda s: starts_with_header(s, "SUPPLIER NAME")),
    ColumnSpec("payment_term", "Payment Term", False, lambda s: starts_with_header(s, "PAYMENT TERM")),
    ColumnSpec(
        "fabric_accessories",
        "Fabric / Accessories",
        False,
        lambda s: starts_with_header(s, "FABRIC / ACCESSORIES"),
    ),
    ColumnSpec("order_type", "Order Type", False, lambda s: starts_with_header(s, "ORDER TYPE")),
    ColumnSpec("vat_invoice_date", "VAT Invoice Date", True, lambda s: starts_with_header(s, "VAT INVOICE DATE")),
    ColumnSpec("vat_invoice_number", "VAT Invoice Number", False, lambda s: starts_with_header(s, "VAT INVOICE NUMBER")),
    ColumnSpec(
        "other_references_number",
        "Other References Number",
        False,
        lambda s: starts_with_header(s, "OTHER REFERENCES NUMBER"),
    ),
    ColumnSpec("po_amount_before_vat_vnd", "PO Amount BEFORE VAT (VND)", True, lambda s: match_po_amount(s, "(VND)")),
    ColumnSpec(
        "surcharge_other_vnd",
        "Surcharge / Other (VND)",
        False,
        lambda s: starts_with_header(s, "SURCHARGE / OTHER (VND)"),
    ),
    ColumnSpec("total_vnd", "Total (VND)", False, lambda s: starts_with_header(s, "TOTAL (VND)")),
    ColumnSpec("currency_rate", "Currency Rate", False, lambda s: starts_with_header(s, "CURRENCY RATE")),
    ColumnSpec("po_amount_before_vat_usd", "PO Amount BEFORE VAT (USD)", True, lambda s: match_po_amount(s, "(USD)")),
    ColumnSpec(
        "surcharge_other_usd",
        "Surcharge / Other (USD)",
        False,
        lambda s: starts_with_header(s, "SURCHARGE / OTHER (USD)"),
    ),
    ColumnSpec("total_usd", "Total (USD)", False, lambda s: starts_with_header(s, "TOTAL (USD)")),
    ColumnSpec("purchaser", "Purchaser", False, lambda s: starts_with_header(s, "PURCHASER")),
    ColumnSpec("from", "From", False, lambda s: starts_with_header(s, "FROM")),
    ColumnSpec("to", "To", False, lambda s: starts_with_header(s, "TO")),
    ColumnSpec(
        "sig_due_date",
        "SIG Due Date",
        False,
        lambda s: starts_with_header(s, "SIG DUE DATE") or starts_with_header(s, "DUE DATE"),
    ),
    ColumnSpec("payment_to_vtec", "Payment to VTEC", False, lambda s: starts_with_header(s, "PAYMENT TO VTEC")),
    ColumnSpec(
        "payment_to_supplier",
        "Payment to Supplier",
        True,
        lambda s: starts_with_header(s, "PAYMENT TO SUPPLIER"),
    ),
]
SOURCE_COLUMN_BY_KEY = {spec.key: spec for spec in SOURCE_COLUMNS}
MANDATORY_SOURCE_HEADER_KEYS = (
    "vat_invoice_date",
    "po_amount_before_vat_vnd",
    "po_amount_before_vat_usd",
    "payment_to_supplier",
)


def _build_output_layout() -> tuple[List[str], Dict[str, int], int, int]:
    headers: List[str] = []
    source_col_map: Dict[str, int] = {}
    standard_rate_col = 0
    status_col = 0

    for spec in SOURCE_COLUMNS:
        source_col_map[spec.key] = len(headers) + 1
        headers.append(spec.title)
        if spec.key == "currency_rate":
            standard_rate_col = len(headers) + 1
            headers.append("Standard Rate")
            status_col = len(headers) + 1
            headers.append("Status")

    return headers, source_col_map, standard_rate_col, status_col


OUTPUT_HEADERS, SOURCE_OUTPUT_COL_BY_KEY, OUTPUT_COL_STANDARD_RATE, OUTPUT_COL_STATUS = _build_output_layout()
TOTAL_OUTPUT_COLUMNS = len(OUTPUT_HEADERS)

VND_KEYS = {"po_amount_before_vat_vnd", "surcharge_other_vnd", "total_vnd"}
USD_KEYS = {"po_amount_before_vat_usd", "surcharge_other_usd", "total_usd"}
DATE_KEYS = {"vat_invoice_date", "sig_due_date", "payment_to_vtec", "payment_to_supplier"}

COLUMN_WIDTHS = {
    "no": 8,
    "supplier_name": 34,
    "payment_term": 18,
    "fabric_accessories": 16,
    "order_type": 14,
    "vat_invoice_date": 14,
    "vat_invoice_number": 18,
    "other_references_number": 22,
    "po_amount_before_vat_vnd": 18,
    "surcharge_other_vnd": 16,
    "total_vnd": 16,
    "currency_rate": 14,
    "standard_rate": 14,
    "status": 16,
    "po_amount_before_vat_usd": 18,
    "surcharge_other_usd": 16,
    "total_usd": 16,
    "purchaser": 16,
    "from": 12,
    "to": 12,
    "sig_due_date": 14,
    "payment_to_vtec": 14,
    "payment_to_supplier": 17,
}


def is_blank(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def trim_trailing_blank(values: Iterable[object]) -> List[object]:
    trimmed = list(values)
    while trimmed and is_blank(trimmed[-1]):
        trimmed.pop()
    return trimmed


def row_value(row_values: List[object], one_based_index: Optional[int]) -> object:
    if not one_based_index:
        return None
    zero_based_index = one_based_index - 1
    if zero_based_index < 0 or zero_based_index >= len(row_values):
        return None
    return row_values[zero_based_index]


def parse_number(value: object) -> Optional[float]:
    if is_blank(value) or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"
    try:
        return float(text)
    except Exception:
        return None


def parse_date_value(value: object, excel_epoch) -> Optional[date]:
    if is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            excel_dt = from_excel(value, excel_epoch)
            if isinstance(excel_dt, datetime):
                return excel_dt.date()
            if isinstance(excel_dt, date):
                return excel_dt
        except Exception:
            return None

    text = re.sub(r"\s+", " ", str(value).strip())
    if not text:
        return None

    formats = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y.%m.%d",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%d.%m.%Y",
        "%d.%m.%y",
        "%d %b %Y",
        "%d %B %Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def normalize_output_value(key: str, value: object, excel_epoch) -> object:
    if key in DATE_KEYS:
        parsed = parse_date_value(value, excel_epoch)
        return parsed if parsed else value
    if key in VND_KEYS or key in USD_KEYS or key == "currency_rate":
        numeric = parse_number(value)
        return numeric if numeric is not None else value
    return value


def rates_equal(left: Optional[float], right: Optional[float], tolerance: float = 1e-9) -> bool:
    return left is not None and right is not None and abs(left - right) <= tolerance


def month_sheet_name(value: date) -> str:
    return f"{MONTH_ABBR[value.month]}'{value.year % 100:02d}"


def build_output_paths(output_dir: Path) -> tuple[Path, Path]:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    workbook_path = ensure_unique_path(output_dir / f"VTEC_Payment_SRate_Monthly_Combined_{timestamp}.xlsx")
    log_path = workbook_path.with_name(f"{workbook_path.stem}_issues.txt")
    return workbook_path, log_path


def collect_source_files(input_folders: List[Path], audit_log: AuditLog) -> List[Path]:
    collected: List[Path] = []
    seen: set[Path] = set()

    for folder in input_folders:
        if not folder.is_dir():
            audit_log.source_file_errors.append(f"{folder} | folder does not exist or is not accessible")
            continue

        folder_files: List[Path] = []

        def handle_walk_error(exc: OSError):
            error_path = getattr(exc, "filename", None) or folder
            audit_log.source_file_errors.append(f"{error_path} | could not scan subfolder: {exc.strerror or exc}")

        for root, _dirs, files in os.walk(folder, onerror=handle_walk_error):
            root_path = Path(root)
            for file_name in sorted(files, key=str.lower):
                if file_name.startswith("~$") or not file_name.lower().endswith(".xlsx"):
                    continue
                file_path = root_path / file_name
                if file_path.is_file():
                    folder_files.append(file_path)

        folder_files.sort(key=lambda path: str(path.relative_to(folder)).lower())
        if not folder_files:
            audit_log.folders_without_xlsx.append(str(folder))
            continue

        for path in folder_files:
            resolved = path.resolve()
            if resolved not in seen:
                collected.append(path)
                seen.add(resolved)

    return collected


def build_header_map(row_values: List[object]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for col_index, cell_value in enumerate(row_values, start=1):
        norm = normalize_header(cell_value)
        if not norm:
            continue
        for spec in SOURCE_COLUMNS:
            if spec.key in mapping:
                continue
            if spec.matcher(norm):
                mapping[spec.key] = col_index
                break
    return mapping


def find_header_row_from_rows(scanned_rows: List[List[object]]) -> tuple[int, Dict[str, int]] | tuple[None, None]:
    best_row = None
    best_map = None
    best_score = -1

    for row_index, row_values in enumerate(scanned_rows, start=1):
        if not row_values:
            continue
        mapping = build_header_map(row_values)
        if "no" not in mapping:
            continue
        if "supplier_name" not in mapping and "vat_invoice_date" not in mapping:
            continue
        score = len(mapping)
        if score > best_score:
            best_row = row_index
            best_map = mapping
            best_score = score

    if best_row is None or best_map is None:
        return None, None
    return best_row, best_map


def propagated_headers(row_values: List[object]) -> List[str]:
    values: List[str] = []
    current = ""
    for value in row_values:
        norm = normalize_header(value)
        if norm:
            current = norm
        values.append(current)
    return values


def find_srate_header_row_from_rows(scanned_rows: List[List[object]]) -> tuple[int, Dict[str, int]] | tuple[None, None]:
    best_row = None
    best_map = None
    best_score = -1

    for top_index in range(len(scanned_rows) - 1):
        top_row = propagated_headers(scanned_rows[top_index])
        bottom_row = [normalize_header(value) for value in scanned_rows[top_index + 1]]
        mapping: Dict[str, int] = {}
        max_cols = max(len(top_row), len(bottom_row))

        for col_index in range(max_cols):
            parent = top_row[col_index] if col_index < len(top_row) else ""
            child = bottom_row[col_index] if col_index < len(bottom_row) else ""
            combined = " ".join(part for part in (parent, child) if part)

            if starts_with_header(parent, "MATERIAL DELIVERY"):
                if "from" not in mapping and starts_with_header(child, "FROM"):
                    mapping["from"] = col_index + 1
                if "to" not in mapping and starts_with_header(child, "TO"):
                    mapping["to"] = col_index + 1

            if "from" not in mapping and "MATERIAL DELIVERY" in combined and combined.endswith("FROM"):
                mapping["from"] = col_index + 1
            if "to" not in mapping and "MATERIAL DELIVERY" in combined and combined.endswith("TO"):
                mapping["to"] = col_index + 1
            if "s_rate" not in mapping and (
                match_srate_header(child) or match_srate_header(parent) or match_srate_header(combined)
            ):
                mapping["s_rate"] = col_index + 1

        if {"from", "to", "s_rate"} <= set(mapping):
            score = len(mapping)
            if score > best_score:
                best_row = top_index + 2
                best_map = mapping
                best_score = score

    if best_row is not None and best_map is not None:
        return best_row, best_map

    for row_index, row_values in enumerate(scanned_rows, start=1):
        mapping: Dict[str, int] = {}
        for col_index, cell_value in enumerate(row_values, start=1):
            norm = normalize_header(cell_value)
            if not norm:
                continue
            if "from" not in mapping and "MATERIAL DELIVERY" in norm and norm.endswith("FROM"):
                mapping["from"] = col_index
            if "to" not in mapping and "MATERIAL DELIVERY" in norm and norm.endswith("TO"):
                mapping["to"] = col_index
            if "s_rate" not in mapping and match_srate_header(norm):
                mapping["s_rate"] = col_index
        if {"from", "to", "s_rate"} <= set(mapping):
            return row_index, mapping

    return None, None


def aggregate_date_issue(bucket: Dict[str, List[str]], issue_date: date, example: str):
    key = issue_date.isoformat()
    if len(bucket[key]) < 5:
        bucket[key].append(example)


def parse_source_file(
    source_file: Path,
    audit_log: AuditLog,
    log_emit: Callable[[str], None] | None = None,
) -> tuple[List[SourceRow], int]:
    _emit(log_emit, f"Reading source workbook: {source_file}")

    try:
        wb = load_workbook(source_file, read_only=True, data_only=True)
    except Exception as exc:
        audit_log.source_file_errors.append(f"{source_file} | could not open workbook: {exc}")
        return [], 0

    try:
        if len(wb.sheetnames) != 1:
            audit_log.multi_sheet_files.append(f"{source_file} | sheets: {', '.join(wb.sheetnames)}")
            return [], 0

        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        scanned_rows: List[List[object]] = []
        for raw_values in ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row, HEADER_SCAN_ROWS),
            min_col=1,
            max_col=SCAN_MAX_COL,
            values_only=True,
        ):
            scanned_rows.append(trim_trailing_blank(raw_values))

        header_row, header_map = find_header_row_from_rows(scanned_rows)
        if header_row is None or header_map is None:
            audit_log.missing_required_headers.append(
                f"{source_file} | no recognizable header row was found within the first {HEADER_SCAN_ROWS} rows"
            )
            return [], 0

        missing_headers = [
            SOURCE_COLUMN_BY_KEY[key].title
            for key in MANDATORY_SOURCE_HEADER_KEYS
            if key not in header_map
        ]
        if missing_headers:
            audit_log.missing_required_headers.append(
                f"{source_file} | missing required header(s): {', '.join(missing_headers)}"
            )
            return [], 0

        rows: List[SourceRow] = []
        missing_vat_rows: List[int] = []
        missing_payment_rows: List[int] = []
        invalid_payment_rows: List[int] = []

        for row_index, raw_values in enumerate(
            ws.iter_rows(
                min_row=header_row + 1,
                min_col=1,
                max_col=SCAN_MAX_COL,
                values_only=True,
            ),
            start=header_row + 1,
        ):
            row_values = trim_trailing_blank(raw_values)
            no_value = row_value(row_values, header_map.get("no"))
            if is_blank(no_value):
                continue

            payment_to_supplier_raw = row_value(row_values, header_map.get("payment_to_supplier"))
            if is_blank(payment_to_supplier_raw):
                missing_payment_rows.append(row_index)
                continue

            payment_to_supplier_date = parse_date_value(payment_to_supplier_raw, wb.epoch)
            if payment_to_supplier_date is None:
                invalid_payment_rows.append(row_index)
                continue

            vat_invoice_raw = row_value(row_values, header_map.get("vat_invoice_date"))
            vat_invoice_date = parse_date_value(vat_invoice_raw, wb.epoch)
            if vat_invoice_date is None:
                missing_vat_rows.append(row_index)
                continue

            normalized_values: Dict[str, object] = {}
            for spec in SOURCE_COLUMNS:
                normalized_values[spec.key] = normalize_output_value(
                    spec.key,
                    row_value(row_values, header_map.get(spec.key)),
                    wb.epoch,
                )

            normalized_values["vat_invoice_date"] = vat_invoice_date
            normalized_values["payment_to_supplier"] = payment_to_supplier_date

            rows.append(
                SourceRow(
                    source_file=source_file,
                    sheet_name=sheet_name,
                    row_number=row_index,
                    payment_to_supplier_date=payment_to_supplier_date,
                    vat_invoice_date=vat_invoice_date,
                    values=normalized_values,
                )
            )

        if missing_payment_rows or missing_vat_rows:
            parts = [f"{source_file} | skipped whole file"]
            if missing_payment_rows:
                parts.append(
                    "Payment to Supplier missing/invalid at row(s): "
                    + ", ".join(str(row) for row in missing_payment_rows)
                )
            if missing_vat_rows:
                parts.append(
                    "VAT Invoice Date missing/invalid at row(s): "
                    + ", ".join(str(row) for row in missing_vat_rows)
                )
            audit_log.file_level_row_issues.append(" | ".join(parts))
            return [], 0

        if invalid_payment_rows:
            audit_log.invalid_payment_to_supplier_rows.append(
                f"{source_file} | row(s) ignored because Payment to Supplier was not a date: "
                + ", ".join(str(row) for row in invalid_payment_rows)
            )

        _emit(log_emit, f"Imported {len(rows)} row(s) from {source_file.name}")
        return rows, len(invalid_payment_rows)
    finally:
        wb.close()


def load_srate_ranges(
    srate_files: List[Path],
    audit_log: AuditLog,
    log_emit: Callable[[str], None] | None = None,
) -> List[SRateRange]:
    ranges: List[SRateRange] = []
    seen_ranges: Dict[tuple[date, date], SRateRange] = {}

    for srate_file in srate_files:
        _emit(log_emit, f"Loading S Rate workbook: {srate_file}")
        try:
            wb = load_workbook(srate_file, read_only=True, data_only=True)
        except Exception as exc:
            audit_log.srate_file_errors.append(f"{srate_file} | could not open workbook: {exc}")
            continue

        try:
            matched_sheet = False
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                scanned_rows: List[List[object]] = []
                for raw_values in ws.iter_rows(
                    min_row=1,
                    max_row=min(ws.max_row, HEADER_SCAN_ROWS),
                    min_col=1,
                    max_col=SCAN_MAX_COL,
                    values_only=True,
                ):
                    scanned_rows.append(trim_trailing_blank(raw_values))

                header_row, header_map = find_srate_header_row_from_rows(scanned_rows)
                if header_row is None or header_map is None:
                    continue

                matched_sheet = True
                for row_index, raw_values in enumerate(
                    ws.iter_rows(
                        min_row=header_row + 1,
                        min_col=1,
                        max_col=SCAN_MAX_COL,
                        values_only=True,
                    ),
                    start=header_row + 1,
                ):
                    row_values = trim_trailing_blank(raw_values)
                    if not row_values:
                        continue

                    from_raw = row_value(row_values, header_map.get("from"))
                    to_raw = row_value(row_values, header_map.get("to"))
                    srate_raw = row_value(row_values, header_map.get("s_rate"))

                    if is_blank(from_raw) and is_blank(to_raw) and is_blank(srate_raw):
                        continue

                    from_date = parse_date_value(from_raw, wb.epoch)
                    to_date = parse_date_value(to_raw, wb.epoch)
                    standard_rate = parse_number(srate_raw)

                    if from_date is None or to_date is None or standard_rate is None:
                        audit_log.invalid_srate_rows.append(
                            f"{srate_file} | {sheet_name} | row {row_index}: invalid Material Delivery From/To or S Rate"
                        )
                        continue
                    if to_date < from_date:
                        audit_log.invalid_srate_rows.append(
                            f"{srate_file} | {sheet_name} | row {row_index}: Material Delivery To is earlier than From"
                        )
                        continue

                    range_key = (from_date, to_date)
                    existing = seen_ranges.get(range_key)
                    if existing and not rates_equal(existing.standard_rate, standard_rate):
                        audit_log.duplicate_srate_ranges.append(
                            f"{srate_file} | {sheet_name} | row {row_index}: "
                            f"{from_date.isoformat()} to {to_date.isoformat()} duplicates "
                            f"{existing.source_file.name} row {existing.row_number} "
                            f"({existing.standard_rate} kept, {standard_rate} ignored)"
                        )
                        continue
                    if existing:
                        continue

                    entry = SRateRange(
                        source_file=srate_file,
                        sheet_name=sheet_name,
                        row_number=row_index,
                        from_date=from_date,
                        to_date=to_date,
                        standard_rate=standard_rate,
                    )
                    seen_ranges[range_key] = entry
                    ranges.append(entry)
                break

            if not matched_sheet:
                audit_log.srate_missing_headers.append(
                    f"{srate_file} | no sheet with Material Delivery -> From/To and S Rate headers was found"
                )
        finally:
            wb.close()

    ranges.sort(
        key=lambda item: (
            item.from_date,
            item.to_date,
            str(item.source_file).lower(),
            item.row_number,
        )
    )
    return ranges


def resolve_standard_rate(
    lookup_date: date,
    srate_ranges: List[SRateRange],
    cache: Dict[date, tuple[Optional[SRateRange], int]],
) -> tuple[Optional[SRateRange], int]:
    cached = cache.get(lookup_date)
    if cached is not None:
        return cached

    matches = [item for item in srate_ranges if item.from_date <= lookup_date <= item.to_date]
    matches.sort(
        key=lambda item: (
            (item.to_date - item.from_date).days,
            item.from_date,
            str(item.source_file).lower(),
            item.row_number,
        )
    )
    resolved = (matches[0] if matches else None, len(matches))
    cache[lookup_date] = resolved
    return resolved


def apply_header_style(cell):
    thin = Side(style="thin", color=BLACK)
    cell.font = Font(bold=True, color=WHITE, name="Arial", size=10)
    cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def apply_data_style(cell, fill_color: str = WHITE, font_color: str = BLACK, bold: bool = False):
    thin = Side(style="thin", color="BFBFBF")
    cell.font = Font(name="Arial", size=10, color=font_color, bold=bold)
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def set_number_format(cell, key: str):
    if key in DATE_KEYS and isinstance(cell.value, (date, datetime)):
        cell.number_format = DATE_NUMBER_FORMAT
    elif key in VND_KEYS:
        cell.number_format = VND_NUMBER_FORMAT
    elif key in USD_KEYS:
        cell.number_format = USD_NUMBER_FORMAT
    elif key == "currency_rate" and isinstance(cell.value, (int, float)):
        cell.number_format = GENERAL_RATE_FORMAT


def setup_month_sheet(ws):
    for col_index, header in enumerate(OUTPUT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        apply_header_style(cell)

    for spec in SOURCE_COLUMNS:
        ws.column_dimensions[get_column_letter(SOURCE_OUTPUT_COL_BY_KEY[spec.key])].width = COLUMN_WIDTHS.get(spec.key, 14)
    ws.column_dimensions[get_column_letter(OUTPUT_COL_STANDARD_RATE)].width = COLUMN_WIDTHS["standard_rate"]
    ws.column_dimensions[get_column_letter(OUTPUT_COL_STATUS)].width = COLUMN_WIDTHS["status"]

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 28


def status_details(currency_rate: Optional[float], standard_rate: Optional[float]) -> tuple[str, str, str]:
    if standard_rate is None:
        return "NO S RATE", STATUS_NEUTRAL_FILL, STATUS_NEUTRAL_FONT
    if currency_rate is None:
        return "NO CURRENCY RATE", STATUS_MISMATCH_FILL, STATUS_MISMATCH_FONT
    if rates_equal(currency_rate, standard_rate):
        return "MATCH", STATUS_MATCH_FILL, STATUS_MATCH_FONT
    return "MISMATCH", STATUS_MISMATCH_FILL, STATUS_MISMATCH_FONT


def write_data_row(
    ws,
    output_row_index: int,
    source_row: SourceRow,
    srate_ranges: List[SRateRange],
    srate_cache: Dict[date, tuple[Optional[SRateRange], int]],
    audit_log: AuditLog,
):
    for spec in SOURCE_COLUMNS:
        col_index = SOURCE_OUTPUT_COL_BY_KEY[spec.key]
        value = source_row.values.get(spec.key)
        cell = ws.cell(row=output_row_index, column=col_index, value=value)
        apply_data_style(cell)
        set_number_format(cell, spec.key)
        if spec.key in {"supplier_name", "payment_term", "other_references_number"}:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    matched_range, match_count = resolve_standard_rate(source_row.vat_invoice_date, srate_ranges, srate_cache)
    if matched_range is None:
        aggregate_date_issue(
            audit_log.missing_standard_rates,
            source_row.vat_invoice_date,
            f"{source_row.source_file.name} row {source_row.row_number}",
        )
    elif match_count > 1:
        aggregate_date_issue(
            audit_log.ambiguous_standard_rate_matches,
            source_row.vat_invoice_date,
            f"{source_row.source_file.name} row {source_row.row_number}",
        )

    standard_rate = matched_range.standard_rate if matched_range else None
    currency_rate = parse_number(source_row.values.get("currency_rate"))
    status_text, status_fill, status_font_color = status_details(currency_rate, standard_rate)

    standard_rate_cell = ws.cell(row=output_row_index, column=OUTPUT_COL_STANDARD_RATE, value=standard_rate)
    apply_data_style(standard_rate_cell, fill_color=STANDARD_RATE_FILL)
    standard_rate_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    if isinstance(standard_rate, (int, float)):
        standard_rate_cell.number_format = GENERAL_RATE_FORMAT

    status_cell = ws.cell(row=output_row_index, column=OUTPUT_COL_STATUS, value=status_text)
    apply_data_style(status_cell, fill_color=status_fill, font_color=status_font_color, bold=True)
    status_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    currency_rate_cell = ws.cell(row=output_row_index, column=SOURCE_OUTPUT_COL_BY_KEY["currency_rate"])
    currency_rate_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    if isinstance(currency_rate_cell.value, (int, float)):
        currency_rate_cell.number_format = GENERAL_RATE_FORMAT

    for key in (
        "po_amount_before_vat_vnd",
        "surcharge_other_vnd",
        "total_vnd",
        "po_amount_before_vat_usd",
        "surcharge_other_usd",
        "total_usd",
    ):
        cell = ws.cell(row=output_row_index, column=SOURCE_OUTPUT_COL_BY_KEY[key])
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

    ws.row_dimensions[output_row_index].height = DATA_ROW_HEIGHT


def render_audit_log(
    audit_log: AuditLog,
    input_folders: List[Path],
    srate_files: List[Path],
    result: ProcessResult,
) -> str:
    lines = [
        "VTEC Payment Monthly S Rate Combiner - Audit Log",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "Selections",
        "----------",
        f"Input folders: {len(input_folders)}",
    ]
    lines.extend(f"- {folder}" for folder in input_folders)
    lines.append(f"S Rate files: {len(srate_files)}")
    lines.extend(f"- {path}" for path in srate_files)
    lines.extend(
        [
            "",
            "Summary",
            "-------",
            f"Source files found: {result.source_files_found}",
            f"Source files imported: {result.source_files_imported}",
            f"Source files skipped: {result.source_files_skipped}",
            f"Rows imported: {result.rows_imported}",
            f"Rows ignored because Payment to Supplier was not a date: {result.rows_ignored_non_date_payment_to_supplier}",
            f"Monthly sheets created: {result.month_sheet_count}",
            f"Workbook output: {result.workbook_path}",
            f"Audit log: {result.log_path}",
        ]
    )

    sections = [
        ("Folders with no .xlsx files in them or their subfolders", audit_log.folders_without_xlsx),
        ("Source files skipped because they have more than 1 sheet", audit_log.multi_sheet_files),
        ("Source file read/open errors", audit_log.source_file_errors),
        ("Source files skipped because required headers were missing", audit_log.missing_required_headers),
        ("Source files skipped because mandatory date cells were missing or invalid", audit_log.file_level_row_issues),
        ("Rows ignored because Payment to Supplier was not a date", audit_log.invalid_payment_to_supplier_rows),
        ("S Rate file errors", audit_log.srate_file_errors),
        ("S Rate files skipped because required headers were not found", audit_log.srate_missing_headers),
        ("S Rate rows ignored because From/To/S Rate was invalid", audit_log.invalid_srate_rows),
        ("Duplicate S Rate ranges ignored", audit_log.duplicate_srate_ranges),
    ]

    for title, items in sections:
        lines.extend(["", title, "-" * len(title)])
        if items:
            lines.extend(f"- {item}" for item in items)
        else:
            lines.append("- None")

    def append_date_section(title: str, mapping: Dict[str, List[str]], suffix: str):
        lines.extend(["", title, "-" * len(title)])
        if not mapping:
            lines.append("- None")
            return
        for issue_date in sorted(mapping.keys()):
            examples = "; ".join(mapping[issue_date])
            lines.append(
                f"- {issue_date}: {len(mapping[issue_date])} example(s) {suffix}"
                + (f" | {examples}" if examples else "")
            )

    append_date_section(
        "VAT Invoice dates with no matching S Rate range",
        audit_log.missing_standard_rates,
        "without a matching S Rate range",
    )
    append_date_section(
        "VAT Invoice dates with multiple matching S Rate ranges",
        audit_log.ambiguous_standard_rate_matches,
        "with multiple matching S Rate ranges",
    )
    return "\n".join(lines) + "\n"


def process_files(
    input_folders: List[Path],
    srate_files: List[Path],
    output_dir: Optional[Path] = None,
    log_emit: Callable[[str], None] | None = None,
) -> ProcessResult:
    if not input_folders:
        raise ValueError("Please select at least one input folder.")
    if not srate_files:
        raise ValueError("Please select at least one S Rate workbook.")

    audit_log = AuditLog()
    output_root = output_dir if output_dir else input_folders[0].parent
    output_root.mkdir(parents=True, exist_ok=True)
    workbook_path, log_path = build_output_paths(output_root)

    _emit(log_emit, "Scanning selected input folders recursively for .xlsx files...")
    source_files = collect_source_files(input_folders, audit_log)
    if not source_files:
        raise ValueError("No .xlsx files were found in the selected input folders or their subfolders.")

    srate_ranges = load_srate_ranges(srate_files, audit_log, log_emit=log_emit)
    if not srate_ranges:
        raise ValueError("No valid S Rate rows could be read from the selected S Rate file(s).")

    all_rows: List[SourceRow] = []
    imported_files = 0
    skipped_files = 0
    rows_ignored_non_date_payment_to_supplier = 0

    for source_file in source_files:
        parsed_rows, ignored_rows = parse_source_file(source_file, audit_log, log_emit=log_emit)
        rows_ignored_non_date_payment_to_supplier += ignored_rows

        if parsed_rows:
            imported_files += 1
            all_rows.extend(parsed_rows)
        else:
            skipped_files += 1

    if not all_rows:
        skipped_files = max(skipped_files, len(source_files))
        result = ProcessResult(
            workbook_path=workbook_path,
            log_path=log_path,
            source_files_found=len(source_files),
            source_files_imported=imported_files,
            source_files_skipped=skipped_files,
            rows_imported=0,
            rows_ignored_non_date_payment_to_supplier=rows_ignored_non_date_payment_to_supplier,
            month_sheet_count=0,
        )
        log_path.write_text(render_audit_log(audit_log, input_folders, srate_files, result), encoding="utf-8")
        raise ValueError(f"No valid rows were imported. Audit log written to: {log_path}")

    grouped_rows: Dict[tuple[int, int], List[SourceRow]] = defaultdict(list)
    for source_row in all_rows:
        grouped_rows[(source_row.payment_to_supplier_date.year, source_row.payment_to_supplier_date.month)].append(source_row)

    wb = Workbook()
    sorted_group_keys = sorted(grouped_rows.keys())
    srate_cache: Dict[date, tuple[Optional[SRateRange], int]] = {}

    for group_index, group_key in enumerate(sorted_group_keys):
        year, month = group_key
        ws = wb.active if group_index == 0 else wb.create_sheet()
        ws.title = month_sheet_name(date(year, month, 1))
        setup_month_sheet(ws)

        sorted_rows = sorted(
            grouped_rows[group_key],
            key=lambda item: (
                item.payment_to_supplier_date,
                item.vat_invoice_date,
                str(item.values.get("supplier_name") or "").upper(),
                item.source_file.name.upper(),
                item.row_number,
            ),
        )
        for output_row_index, source_row in enumerate(sorted_rows, start=2):
            write_data_row(ws, output_row_index, source_row, srate_ranges, srate_cache, audit_log)

    wb.save(workbook_path)

    result = ProcessResult(
        workbook_path=workbook_path,
        log_path=log_path,
        source_files_found=len(source_files),
        source_files_imported=imported_files,
        source_files_skipped=skipped_files,
        rows_imported=len(all_rows),
        rows_ignored_non_date_payment_to_supplier=rows_ignored_non_date_payment_to_supplier,
        month_sheet_count=len(sorted_group_keys),
    )
    log_path.write_text(render_audit_log(audit_log, input_folders, srate_files, result), encoding="utf-8")
    return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Combine VTEC payment summary workbooks into monthly worksheets using S Rate date ranges."
    )
    parser.add_argument(
        "folders",
        nargs="+",
        help="Input folders containing .xlsx payment-summary files. Subfolders are scanned automatically.",
    )
    parser.add_argument(
        "--srate",
        nargs="+",
        required=True,
        help="One or more S Rate workbook paths.",
    )
    parser.add_argument(
        "--output-dir",
        help="Output folder. Defaults to the first selected input folder's parent.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    result = process_files(
        input_folders=[Path(path) for path in args.folders],
        srate_files=[Path(path) for path in args.srate],
        output_dir=Path(args.output_dir) if args.output_dir else None,
        log_emit=print,
    )
    print(f"Workbook: {result.workbook_path}")
    print(f"Audit log: {result.log_path}")
    print(f"Rows imported: {result.rows_imported}")
    print(f"Monthly sheets: {result.month_sheet_count}")


if __name__ == "__main__":
    main()


if GUI_AVAILABLE:
    class MainWidget(QWidget):
        log_message = Signal(str)
        processing_done = Signal(bool, str, str, int, int, int, int)

        def __init__(self):
            super().__init__()
            self.setObjectName("vtec_payment_srate_monthly_combiner_widget")
            self.input_folders: List[Path] = []
            self.srate_files: List[Path] = []
            self.output_dir: Path = Path.home()
            self.output_dir_locked = False

            self._build_ui()
            self._connect_signals()
            self._refresh_summary()

        def _build_ui(self):
            self.desc_label = QLabel("", self)
            self.desc_label.setWordWrap(True)
            self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
            self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
            self.desc_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
            self.desc_label.setStyleSheet(
                "color: #dcdcdc; background: transparent; padding: 6px; "
                "border: 1px solid #3a3a3a; border-radius: 6px;"
            )
            self.set_long_description("")

            self.select_folders_btn = PrimaryPushButton("Add Input Folder", self)
            self.clear_folders_btn = PrimaryPushButton("Clear Input Folders", self)
            self.select_srate_btn = PrimaryPushButton("Select S Rate File(s)", self)
            self.select_output_btn = PrimaryPushButton("Select Output Folder", self)
            self.run_btn = PrimaryPushButton("Generate Monthly Workbook", self)

            self.summary_label = QLabel("Selections", self)
            self.summary_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

            self.logs_label = QLabel("Process logs", self)
            self.logs_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

            box_style = (
                "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
                "border: 1px solid #3a3a3a; border-radius: 6px;}"
            )

            self.summary_box = QTextEdit(self)
            self.summary_box.setReadOnly(True)
            self.summary_box.setPlaceholderText("Input folders, S Rate files, and output details will appear here")
            self.summary_box.setStyleSheet(box_style)

            self.log_box = QTextEdit(self)
            self.log_box.setReadOnly(True)
            self.log_box.setPlaceholderText("Live process log will appear here")
            self.log_box.setStyleSheet(box_style)

            layout = QVBoxLayout(self)
            layout.setContentsMargins(16, 16, 16, 16)
            layout.setSpacing(12)
            layout.addWidget(self.desc_label)

            row1 = QHBoxLayout()
            row1.addWidget(self.select_folders_btn)
            row1.addWidget(self.clear_folders_btn)
            row1.addWidget(self.select_srate_btn)
            row1.addWidget(self.select_output_btn)
            layout.addLayout(row1)

            row2 = QHBoxLayout()
            row2.addStretch(1)
            row2.addWidget(self.run_btn)
            row2.addStretch(1)
            layout.addLayout(row2)

            row3 = QHBoxLayout()
            row3.addWidget(self.summary_label, 1)
            row3.addWidget(self.logs_label, 1)
            layout.addLayout(row3)

            row4 = QHBoxLayout()
            row4.addWidget(self.summary_box, 1)
            row4.addWidget(self.log_box, 1)
            layout.addLayout(row4, 3)

        def _connect_signals(self):
            self.select_folders_btn.clicked.connect(self.select_input_folders)
            self.clear_folders_btn.clicked.connect(self.clear_input_folders)
            self.select_srate_btn.clicked.connect(self.select_srate_files)
            self.select_output_btn.clicked.connect(self.select_output_folder)
            self.run_btn.clicked.connect(self.run_process)
            self.log_message.connect(self.append_log)
            self.processing_done.connect(self.on_processing_done)

        def set_long_description(self, text: str):
            clean = (text or "").strip()
            if clean:
                self.desc_label.setText(clean)
                self.desc_label.show()
            else:
                self.desc_label.clear()
                self.desc_label.hide()

        def _dedupe_paths(self, paths: Iterable[Path]) -> List[Path]:
            deduped: List[Path] = []
            seen: set[Path] = set()
            for path in paths:
                resolved = path.resolve()
                if resolved not in seen:
                    deduped.append(path)
                    seen.add(resolved)
            return deduped

        def _refresh_default_output_dir(self):
            if not self.output_dir_locked and self.input_folders:
                self.output_dir = self.input_folders[0].parent

        def _refresh_summary(self):
            self._refresh_default_output_dir()

            lines = [
                f"Input folders selected: {len(self.input_folders)}",
                "Input scan: recursive (includes subfolders)",
            ]
            if self.input_folders:
                lines.extend(f"- {path}" for path in self.input_folders)
            else:
                lines.append("- None")

            lines.append("")
            lines.append(f"S Rate files selected: {len(self.srate_files)}")
            if self.srate_files:
                lines.extend(f"- {path}" for path in self.srate_files)
            else:
                lines.append("- None")

            lines.extend(
                [
                    "",
                    f"Output folder: {self.output_dir}",
                    "Output workbook name: auto-generated with timestamp",
                    "Audit log name: matches workbook name with _issues.txt suffix",
                ]
            )

            self.summary_box.setPlainText("\n".join(lines))

        def _select_input_folder(self) -> Optional[Path]:
            start_dir = self.input_folders[-1] if self.input_folders else self.output_dir
            selected = QFileDialog.getExistingDirectory(self, "Select Input Folder", str(start_dir))
            if selected:
                return Path(selected)
            return None

        def select_input_folders(self):
            selected = self._select_input_folder()
            if not selected:
                return
            self.input_folders = self._dedupe_paths([*self.input_folders, selected])
            self._refresh_summary()

        def clear_input_folders(self):
            self.input_folders = []
            if not self.output_dir_locked:
                self.output_dir = Path.home()
            self._refresh_summary()

        def select_srate_files(self):
            files, _ = QFileDialog.getOpenFileNames(
                self,
                "Select S Rate Files",
                "",
                "Excel Files (*.xlsx)",
            )
            if files:
                self.srate_files = self._dedupe_paths(Path(path) for path in files)
            else:
                self.srate_files = []
            self._refresh_summary()

        def select_output_folder(self):
            selected = QFileDialog.getExistingDirectory(self, "Select Output Folder", str(self.output_dir))
            if selected:
                self.output_dir = Path(selected)
                self.output_dir_locked = True
                self._refresh_summary()

        def _set_running_state(self, running: bool):
            enabled = not running
            self.select_folders_btn.setEnabled(enabled)
            self.clear_folders_btn.setEnabled(enabled)
            self.select_srate_btn.setEnabled(enabled)
            self.select_output_btn.setEnabled(enabled)
            self.run_btn.setEnabled(enabled)

        def run_process(self):
            if not self.input_folders:
                MessageBox("Warning", "Please add at least one input folder.", self).exec()
                return
            if not self.srate_files:
                MessageBox("Warning", "Please select at least one S Rate workbook.", self).exec()
                return

            self.log_box.clear()
            self.log_message.emit("Process started...")
            self.log_message.emit(f"Input folders: {len(self.input_folders)}")
            self.log_message.emit(f"S Rate files: {len(self.srate_files)}")
            self.log_message.emit(f"Output folder: {self.output_dir}")
            self.log_message.emit("")
            self._set_running_state(True)

            def worker():
                try:
                    result = process_files(
                        input_folders=self.input_folders,
                        srate_files=self.srate_files,
                        output_dir=self.output_dir,
                        log_emit=self.log_message.emit,
                    )
                    self.processing_done.emit(
                        True,
                        str(result.workbook_path),
                        str(result.log_path),
                        result.source_files_imported,
                        result.source_files_skipped,
                        result.rows_imported,
                        result.month_sheet_count,
                    )
                except Exception as exc:
                    self.log_message.emit(f"CRITICAL ERROR: {exc}")
                    self.processing_done.emit(False, "", "", 0, 0, 0, 0)

            threading.Thread(target=worker, daemon=True).start()

        def append_log(self, text: str):
            self.log_box.append(text)
            self.log_box.ensureCursorVisible()

        def on_processing_done(
            self,
            success: bool,
            workbook_path: str,
            log_path: str,
            imported_files: int,
            skipped_files: int,
            rows_imported: int,
            month_sheet_count: int,
        ):
            self._set_running_state(False)

            if success:
                self.log_message.emit("")
                self.log_message.emit(f"Workbook saved: {workbook_path}")
                self.log_message.emit(f"Audit log saved: {log_path}")
                self.log_message.emit(
                    f"Completed: {imported_files} imported file(s), {skipped_files} skipped file(s), "
                    f"{rows_imported} row(s), {month_sheet_count} monthly sheet(s)"
                )

                msg = MessageBox(
                    "Monthly workbook created",
                    "\n".join(
                        [
                            f"Imported files: {imported_files}",
                            f"Skipped files: {skipped_files}",
                            f"Rows imported: {rows_imported}",
                            f"Monthly sheets: {month_sheet_count}",
                            f"Workbook: {Path(workbook_path).name}",
                            f"Audit log: {Path(log_path).name}",
                        ]
                    ),
                    self,
                )
            else:
                msg = MessageBox(
                    "Monthly workbook finished with issues",
                    "No workbook was created. Check the process log for details.",
                    self,
                )

            msg.yesButton.setText("OK")
            msg.cancelButton.hide()
            msg.exec()


    def get_widget():
        return MainWidget()


else:
    def get_widget():
        raise RuntimeError("PySide6 and qfluentwidgets are required to use this pycro inside Pycro Station.")
