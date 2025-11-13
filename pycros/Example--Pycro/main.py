import threading
from PySide6.QtCore import Signal
from PySide6.QtWidgets import QWidget, QVBoxLayout, QTextEdit, QFileDialog
from qfluentwidgets import PrimaryPushButton, PushButton, MessageBox
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


class MainWidget(QWidget):
    log_message = Signal(str)
    processing_done = Signal(int, int)
    def __init__(self):
        super().__init__()
        self.setObjectName('test_macro_one_widget')

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(12)

        self.select_btn = PrimaryPushButton('Select Files...', self)
        self.select_btn.clicked.connect(self.select_files)

        self.files_box = QTextEdit(self)
        self.files_box.setReadOnly(True)
        self.files_box.setPlaceholderText('Selected files will appear here')
        self.files_box.setStyleSheet('QTextEdit{background: #2a2a2a; color: white; border: 1px solid #3a3a3a; border-radius: 6px;}')

        # Live process log below the files box
        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText('Live process log will appear here')
        self.log_box.setStyleSheet('QTextEdit{background: #1f1f1f; color: #d0d0d0; border: 1px solid #3a3a3a; border-radius: 6px;}')

        self.run_btn = PrimaryPushButton('Run', self)
        self.run_btn.clicked.connect(self.run_process)

        layout.addWidget(self.select_btn)
        layout.addWidget(self.files_box, 1)
        layout.addWidget(self.log_box, 1)
        layout.addWidget(self.run_btn)

        # connect signals
        self.log_message.connect(self.append_log)
        self.processing_done.connect(self.on_processing_done)

    def select_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, 'Select files')
        if files:
            self.files_box.setPlainText('\n'.join(files))
        else:
            self.files_box.clear()

    def _selected_files(self):
        text = self.files_box.toPlainText().strip()
        if not text:
            return []
        return [line for line in text.split('\n') if line.strip()]

    def run_process(self):
        files = self._selected_files()
        if not files:
            MessageBox('No files', 'Please select one or more Excel files first.', self).exec()
            return
        self.log_box.clear()
        self.log_message.emit(f'Starting processing of {len(files)} file(s)...')
        self.run_btn.setEnabled(False)
        self.select_btn.setEnabled(False)

        def worker():
            ok, fail = 0, 0
            for path in files:
                try:
                    self.log_message.emit(f'Opening: {path}')
                    wb = load_workbook(path)
                    ws = wb.active
                    c = ws.cell(row=1, column=1)
                    c.value = 'test complete'
                    c.font = Font(color='FFFF0000')  # red text
                    c.fill = PatternFill(fill_type='solid', start_color='FF000000', end_color='FF000000')  # black bg
                    wb.save(path)
                    self.log_message.emit(f'Updated A1 and saved: {path}')
                    ok += 1
                except Exception as e:
                    self.log_message.emit(f'ERROR processing {path}: {e}')
                    fail += 1
            # notify UI thread
            self.processing_done.emit(ok, fail)

        threading.Thread(target=worker, daemon=True).start()

    def append_log(self, text: str):
        self.log_box.append(text)
        self.log_box.ensureCursorVisible()

    def on_processing_done(self, ok: int, fail: int):
        # log completion and re-enable controls after dialog dismissed
        self.log_message.emit(f'Completed: {ok} success, {fail} failed.')
        self.run_btn.setEnabled(True)
        self.select_btn.setEnabled(True)


def get_widget():
    return MainWidget()
