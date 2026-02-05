import os
import re
import threading
from copy import copy
from datetime import datetime
from typing import Callable, Dict, List, Optional, Tuple, Any

from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QSizePolicy,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)
from qfluentwidgets import MessageBox, PrimaryPushButton

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


GRAY_HEX = "FFB2B2B2"
GRAY_FILL = PatternFill(fill_type="solid", start_color=GRAY_HEX, end_color=GRAY_HEX)

HEADER_MARKER = "embellishment cost"
GAIN_VALUES = {"gain", "gain/pc"}

EMBELLISHMENT_COMPONENTS = {"(emb", "printing", "apc", "pad", "laser", "bonding)"}


class MainWidget(QWidget):
    log_message = Signal(str)
    processing_done = Signal(int, int)

    def __init__(self):
        super().__init__()
        self.setObjectName("costing_data_transformer_widget")
        self._build_ui()
        self._connect_signals()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(12)

        self.desc_label = QLabel("", self)
        self.desc_label.setWordWrap(True)
        self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.desc_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.desc_label.setStyleSheet(
            "color: #dcdcdc; background: transparent; padding: 6px; "
            "border: 1px solid #3a3a3a; border-radius: 6px;"
        )
        self.set_long_description("")

        self.select_data_btn = PrimaryPushButton("Select Data Excel Files", self)
        self.select_lookup_btn = PrimaryPushButton("Select Lookup Table File", self)
        self.run_btn = PrimaryPushButton("Run", self)

        row_btns = QHBoxLayout()
        row_btns.addWidget(self.select_data_btn, 1)
        row_btns.addWidget(self.select_lookup_btn, 1)
        row_btns.addWidget(self.run_btn, 1)

        self.files_box = QTextEdit(self)
        self.files_box.setReadOnly(True)
        self.files_box.setPlaceholderText("Selected data files will appear here")
        self.files_box.setStyleSheet(
            "QTextEdit{background: #2a2a2a; color: white; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        self.lookup_box = QTextEdit(self)
        self.lookup_box.setReadOnly(True)
        self.lookup_box.setPlaceholderText("Selected lookup table file will appear here")
        self.lookup_box.setMaximumHeight(60)
        self.lookup_box.setStyleSheet(
            "QTextEdit{background: #2a2a2a; color: white; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText("Live process log will appear here")
        self.log_box.setStyleSheet(
            "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        layout.addWidget(self.desc_label, 0)
        layout.addLayout(row_btns, 0)
        layout.addWidget(QLabel("Data Files:"), 0)
        layout.addWidget(self.files_box, 1)
        layout.addWidget(QLabel("Lookup Table File:"), 0)
        layout.addWidget(self.lookup_box, 0)
        layout.addWidget(QLabel("Process Log:"), 0)
        layout.addWidget(self.log_box, 2)

    def _connect_signals(self):
        self.select_data_btn.clicked.connect(self.select_data_files)
        self.select_lookup_btn.clicked.connect(self.select_lookup_file)
        self.run_btn.clicked.connect(self.run_process)
        self.log_message.connect(self.append_log)
        self.processing_done.connect(self.on_processing_done)

    def set_long_description(self, text: str):
        clean = (text or "").strip()
        if clean:
            self.desc_label.setText(clean)
            self.desc_label.show()
        else:
            self.desc_label.clear()
            self.desc_label.hide()

    def select_data_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "Select Data Excel files", "", "Excel Workbook (*.xlsx)"
        )
        if files:
            self.files_box.setPlainText("\n".join(files))
        else:
            self.files_box.clear()

    def select_lookup_file(self):
        file, _ = QFileDialog.getOpenFileName(
            self, "Select Lookup Table Excel file", "", "Excel Workbook (*.xlsx)"
        )
        if file:
            self.lookup_box.setPlainText(file)
        else:
            self.lookup_box.clear()

    def _selected_data_files(self) -> List[str]:
        text = self.files_box.toPlainText().strip()
        if not text:
            return []
        return [line for line in text.split("\n") if line.strip()]

    def _selected_lookup_file(self) -> Optional[str]:
        text = self.lookup_box.toPlainText().strip()
        return text if text else None

    def run_process(self):
        files = self._selected_data_files()
        lookup_file = self._selected_lookup_file()

        if not files:
            MessageBox("No files", "Please select one or more data .xlsx files first.", self).exec()
            return
        if not lookup_file:
            MessageBox("No lookup", "Please select a lookup table .xlsx file.", self).exec()
            return

        self.log_box.clear()
        self.log_message.emit(f"Starting processing of {len(files)} file(s)...")
        self.run_btn.setEnabled(False)
        self.select_data_btn.setEnabled(False)
        self.select_lookup_btn.setEnabled(False)

        def worker():
            ok, fail = 0, 0
            try:
                lookup_data = load_lookup_tables(lookup_file, self.log_message.emit)
            except Exception as e:
                self.log_message.emit(f"ERROR loading lookup tables: {e}")
                self.processing_done.emit(0, len(files))
                return

            for path in files:
                try:
                    self.log_message.emit(f"Opening: {path}")
                    success, out_path = process_file(path, lookup_data, self.log_message.emit)
                    if success:
                        self.log_message.emit(f"Saved: {out_path}")
                        ok += 1
                    else:
                        fail += 1
                except Exception as e:
                    self.log_message.emit(f"ERROR processing {path}: {e}")
                    fail += 1
            self.processing_done.emit(ok, fail)

        threading.Thread(target=worker, daemon=True).start()

    def append_log(self, text: str):
        self.log_box.append(text)
        self.log_box.ensureCursorVisible()

    def on_processing_done(self, ok: int, fail: int):
        self.log_message.emit(f"Completed: {ok} success, {fail} failed.")
        self.run_btn.setEnabled(True)
        self.select_data_btn.setEnabled(True)
        self.select_lookup_btn.setEnabled(True)
        title = "Processing complete" if fail == 0 else "Processing finished with issues"
        lines = [f"Success: {ok}", f"Failed: {fail}", "", "Outputs are saved next to the input file(s)."]
        msg = MessageBox(title, "\n".join(lines), self)
        msg.yesButton.setText("OK")
        msg.cancelButton.hide()
        msg.exec()


def get_widget():
    return MainWidget()


def load_lookup_tables(lookup_path: str, log: Callable[[str], None]) -> Dict[str, Any]:
    """Load lookup tables from the lookup file."""
    log(f"Loading lookup tables from: {lookup_path}")
    wb = load_workbook(lookup_path, data_only=True)

    lookup_data = {
        "prod_type": {},
        "prod_cat": {},
        "pco2175_sheets": [],
        "pid103_sheets": [],
    }

    for ws in wb.worksheets:
        sheet_name_lower = ws.title.lower()

        if sheet_name_lower == "prod type":
            log(f"  Loading Prod Type lookup from sheet: {ws.title}")
            for row in ws.iter_rows(min_row=2):
                if len(row) >= 3:
                    key = _norm(row[1].value)
                    val = _norm(row[2].value)
                    if key:
                        lookup_data["prod_type"][key.lower()] = val

        elif sheet_name_lower == "prod cat":
            log(f"  Loading Prod Cat lookup from sheet: {ws.title}")
            for row in ws.iter_rows(min_row=2):
                if len(row) >= 2:
                    key = _norm(row[0].value)
                    val = _norm(row[1].value)
                    if key:
                        lookup_data["prod_cat"][key.lower()] = val

        elif sheet_name_lower.startswith("pco2175"):
            log(f"  Loading pco2175 data from sheet: {ws.title}")
            pco_data = _load_pco2175_sheet(ws)
            lookup_data["pco2175_sheets"].append(pco_data)

        elif sheet_name_lower.startswith("pid103"):
            log(f"  Loading pid103 data from sheet: {ws.title}")
            pid_data = _load_pid103_sheet(ws, log)
            lookup_data["pid103_sheets"].append(pid_data)

    log(f"  Prod Type entries: {len(lookup_data['prod_type'])}")
    log(f"  Prod Cat entries: {len(lookup_data['prod_cat'])}")
    log(f"  pco2175 sheets: {len(lookup_data['pco2175_sheets'])}")
    log(f"  pid103 sheets: {len(lookup_data['pid103_sheets'])}")

    # Show sample pid103 entries for debugging
    for i, pid_data in enumerate(lookup_data["pid103_sheets"]):
        entries = list(pid_data["style_to_total"].items())[:5]
        if entries:
            log(f"  pid103 sheet {i+1} sample entries: {entries}")

    return lookup_data


def _load_pco2175_sheet(ws: Worksheet) -> Dict[str, Any]:
    """Load pco2175 sheet data - find SEASON&STYLE column and extract & WR and PPG CODE."""
    data = {"season_style_to_wr": {}, "season_style_to_ppg": {}}

    header_row = None
    season_style_col = None
    wr_col = None
    ppg_col = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=50), start=1):
        for cell in row:
            val = _norm(cell.value).lower()
            if val == "season&style":
                header_row = row_idx
                season_style_col = cell.column
                break
        if header_row:
            break

    if header_row is None or season_style_col is None:
        return data

    for cell in ws[header_row]:
        val = _norm(cell.value).lower()
        if val == "& wr":
            wr_col = cell.column
        elif val == "ppg code":
            ppg_col = cell.column

    for row in ws.iter_rows(min_row=header_row + 1):
        season_style_val = None
        wr_val = None
        ppg_val = None

        for cell in row:
            if cell.column == season_style_col:
                season_style_val = _norm(cell.value)
            elif cell.column == wr_col:
                wr_val = cell.value
            elif cell.column == ppg_col:
                ppg_val = _norm(cell.value)

        if season_style_val:
            key = season_style_val.lower()
            if wr_val is not None:
                data["season_style_to_wr"][key] = wr_val
            if ppg_val:
                data["season_style_to_ppg"][key] = ppg_val

    return data


def _load_pid103_sheet(ws: Worksheet, log: Callable[[str], None]) -> Dict[str, Any]:
    """Load pid103 sheet data - find Style column and TOTAL column."""
    data = {"style_to_total": {}}

    header_row = None
    style_col = None
    total_col = None

    # First pass: find both Style and TOTAL columns anywhere in first 50 rows
    # Search up to column 30 (AD) minimum, or max_column if larger
    max_search_col = max(30, ws.max_column + 1)
    log(f"    Sheet dimensions: max_row={ws.max_row}, max_column={ws.max_column}")
    log(f"    Searching rows 1-50, columns 1-{max_search_col-1}")

    # Debug: show first 5 rows structure
    for debug_row in range(1, min(6, ws.max_row + 1)):
        cells_preview = []
        for debug_col in range(1, min(25, max_search_col)):
            val = ws.cell(row=debug_row, column=debug_col).value
            if val is not None:
                cells_preview.append(f"{get_column_letter(debug_col)}:{repr(str(val)[:20])}")
        if cells_preview:
            log(f"    Row {debug_row}: {', '.join(cells_preview[:8])}...")

    # First pass: find Style column
    for row_idx in range(1, min(51, ws.max_row + 1)):
        for col_idx in range(1, max_search_col):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = _norm(cell.value).lower()

            # Find Style column - match "style" exactly or starting with "style"
            if style_col is None and val and (val == "style" or val.startswith("style")):
                header_row = row_idx
                style_col = col_idx
                log(f"    Found Style column at row {row_idx}, col {get_column_letter(col_idx)}: '{cell.value}'")
                break
        if style_col:
            break

    # Second pass: find TOTAL column anywhere in first 10 rows (may be in different row than Style)
    for row_idx in range(1, min(11, ws.max_row + 1)):
        for col_idx in range(1, max_search_col):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = _norm(cell.value).lower()

            if total_col is None and val and "total" in val:
                total_col = col_idx
                log(f"    Found TOTAL column at row {row_idx}, col {get_column_letter(col_idx)}: '{cell.value}'")
                break
        if total_col:
            break

    if header_row is None:
        log(f"    WARNING: No header row found in pid103 sheet")
        return data

    if style_col is None:
        log(f"    WARNING: No 'Style' column found in pid103 sheet")
        return data

    if total_col is None:
        log(f"    WARNING: No 'TOTAL' column found in pid103 sheet")
        # List all headers found for debugging
        headers_found = []
        for col_idx in range(1, max_search_col):
            cell_val = ws.cell(row=header_row, column=col_idx).value
            if cell_val:
                headers_found.append(f"{get_column_letter(col_idx)}:{_norm(cell_val)}")
        log(f"    Headers in row {header_row}: {headers_found}")
        # Show columns T-W (20-23) in rows 1-10 to help find TOTAL
        log(f"    Debug: Checking columns T-W in rows 1-10:")
        for r in range(1, 11):
            cols_tw = []
            for c in range(20, 24):  # T=20, U=21, V=22, W=23
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    cols_tw.append(f"{get_column_letter(c)}:{repr(str(val)[:30])}")
            if cols_tw:
                log(f"      Row {r}: {', '.join(cols_tw)}")
        return data

    log(f"    Header row={header_row}, Style col={get_column_letter(style_col)}, TOTAL col={get_column_letter(total_col)}")

    # Load data rows
    count = 0
    rows_to_load = ws.max_row - header_row
    log(f"    Loading data from row {header_row + 1} to {ws.max_row} ({rows_to_load} potential rows)")

    # Debug: show first 5 data rows
    for debug_row in range(header_row + 1, min(header_row + 6, ws.max_row + 1)):
        style_sample = ws.cell(row=debug_row, column=style_col).value
        total_sample = ws.cell(row=debug_row, column=total_col).value
        log(f"    Sample row {debug_row}: Style='{style_sample}', TOTAL='{total_sample}'")

    for row_idx in range(header_row + 1, ws.max_row + 1):
        style_val = _norm(ws.cell(row=row_idx, column=style_col).value)
        total_val = ws.cell(row=row_idx, column=total_col).value

        if style_val:
            data["style_to_total"][style_val.lower()] = total_val
            count += 1

    log(f"    Loaded {count} style entries from pid103 sheet")

    # Show first few entries for verification
    if count > 0:
        samples = list(data["style_to_total"].items())[:5]
        log(f"    First entries: {samples}")

    return data


def process_file(
    path: str, lookup_data: Dict[str, Any], log: Callable[[str], None]
) -> Tuple[bool, str]:
    """Process a single data file."""
    if not path.lower().endswith(".xlsx"):
        log("Skipped (not .xlsx).")
        return False, ""

    wb = load_workbook(path)
    for ws in wb.worksheets:
        _process_sheet(ws, lookup_data, log)

    out_path = _build_output_path(path)
    wb.save(out_path)
    return True, out_path


def _process_sheet(ws: Worksheet, lookup_data: Dict[str, Any], log: Callable[[str], None]) -> None:
    """Process a single worksheet."""
    log(f"  Sheet: {ws.title}")

    header_row = _find_header_row(ws)
    if header_row is None:
        log(f"    - No 'Embellishment Cost' header found, skipping sheet.")
        return

    log(f"    - Found header at row {header_row}")

    season_value = _find_season_value(ws)
    log(f"    - Season value: {season_value or 'NOT FOUND'}")

    headers = _get_headers(ws, header_row)
    log(f"    - Found {len(headers)} header columns")

    _format_header_row(ws, header_row, len(headers))

    _delete_gain_rows(ws, header_row, headers, log)

    _insert_new_columns(ws, header_row, headers, lookup_data, season_value, log)

    _autofit_columns(ws, header_row, log)


def _autofit_columns(ws: Worksheet, _header_row: int, log: Callable[[str], None]) -> None:
    """Autofit columns from B to the end based on content width."""
    COL_B = 2
    max_col = ws.max_column
    max_row = ws.max_row

    for col_idx in range(COL_B, max_col + 1):
        max_length = 0
        for row_idx in range(1, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_length:
                    max_length = cell_len

        if max_length > 0:
            adjusted_width = min(max(max_length + 2, 8), 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    log(f"    - Autofit columns B to {get_column_letter(max_col)}")


def _find_header_row(ws: Worksheet) -> Optional[int]:
    """Find the row containing 'Embellishment Cost' header."""
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if _norm(cell.value).lower() == HEADER_MARKER:
                return row_idx
    return None


def _find_season_value(ws: Worksheet) -> Optional[str]:
    """Find 'Season :' cell and extract the season value (next 5 characters)."""
    for row in ws.iter_rows():
        for cell in row:
            val = _norm(cell.value)
            if val.lower().startswith("season :") or val.lower().startswith("season:"):
                match = re.search(r'season\s*:\s*(.{1,5})', val, re.IGNORECASE)
                if match:
                    return match.group(1).strip()
    return None


def _get_headers(ws: Worksheet, header_row: int) -> Dict[str, int]:
    """Get mapping of header names (lowercase) to column indices."""
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        val = _norm(cell.value)
        if val:
            headers[val.lower()] = col_idx
    return headers


def _format_header_row(ws: Worksheet, header_row: int, max_col: int) -> None:
    """Bold the header row and set background to #b2b2b2."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col)
        _make_bold(cell)
        _apply_gray_fill(cell)


def _delete_gain_rows(
    ws: Worksheet, header_row: int, headers: Dict[str, int], log: Callable[[str], None]
) -> None:
    """Delete rows where column I (9) has 'Gain' or 'Gain/pc', and empty rows."""
    col_i = 9
    max_col = ws.max_column

    gain_rows = []
    empty_rows = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_i)
        val = _norm(cell.value).lower()

        if val in GAIN_VALUES:
            gain_rows.append(row_idx)
        else:
            is_empty = True
            for col_idx in range(1, max_col + 1):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None and str(cell_val).strip() != "":
                    is_empty = False
                    break
            if is_empty:
                empty_rows.append(row_idx)

    rows_to_delete = sorted(set(gain_rows + empty_rows), reverse=True)

    for row_idx in rows_to_delete:
        ws.delete_rows(row_idx)

    if gain_rows:
        log(f"    - Deleted {len(gain_rows)} rows with Gain/Gain/pc values")
    if empty_rows:
        log(f"    - Deleted {len(empty_rows)} empty rows")


def _insert_new_columns(
    ws: Worksheet,
    header_row: int,
    headers: Dict[str, int],
    lookup_data: Dict[str, Any],
    season_value: Optional[str],
    log: Callable[[str], None],
) -> None:
    """Insert all new columns with proper values."""

    new_cols = {}

    def insert_column_after(after_col: str, new_header: str, track_key: str, calc_func: Callable) -> Optional[int]:
        """Helper to insert a column after a specified column."""
        if after_col not in headers and after_col not in new_cols:
            return None

        base_col = headers.get(after_col) or new_cols.get(after_col)
        insert_at = base_col + 1
        ws.insert_cols(insert_at)

        for old_key in list(headers.keys()):
            if headers[old_key] >= insert_at:
                headers[old_key] += 1
        for old_key in list(new_cols.keys()):
            if new_cols[old_key] >= insert_at:
                new_cols[old_key] += 1

        ws.cell(row=header_row, column=insert_at).value = new_header
        _make_bold(ws.cell(row=header_row, column=insert_at))
        _apply_gray_fill(ws.cell(row=header_row, column=insert_at))

        new_cols[track_key] = insert_at
        headers[new_header.lower()] = insert_at

        for row_idx in range(header_row + 1, ws.max_row + 1):
            row_data = _get_row_data(ws, row_idx, headers)
            row_data.update({k: ws.cell(row=row_idx, column=v).value for k, v in new_cols.items()})
            value = calc_func(row_idx, row_data)
            ws.cell(row=row_idx, column=insert_at).value = value

        log(f"    - Inserted column '{new_header}' after '{after_col}'")
        return insert_at

    if "buyer" in headers:
        insert_column_after("buyer", "Season", "season",
            lambda r, d: season_value or "")

    if "job number" in headers:
        insert_column_after("job number", "Job Short", "job_short",
            lambda r, d: _calc_job_short(r, d))

    if "p.type" in headers:
        insert_column_after("p.type", "P.Type Group", "p_type_group",
            lambda r, d: _lookup_prod_type(r, d, headers, lookup_data))

    if "p.cat" in headers:
        insert_column_after("p.cat", "P.Cat Group", "p_cat_group",
            lambda r, d: _lookup_prod_cat(r, d, headers, lookup_data))

    if "style #" in headers:
        insert_column_after("style #", "Season+Style", "season_style",
            lambda r, d: _calc_season_style(r, d, headers, season_value))

    if "season_style" in new_cols:
        insert_column_after("season_style", "TS&WR", "ts_wr",
            lambda r, d: _lookup_ts_wr(r, d, headers, lookup_data, season_value))

    if "embellishment cost" in headers:
        insert_column_after("embellishment cost", "Checking", "checking_emb",
            lambda r, d: _calc_emb_checking(r, d, headers))

    if "gross cm" in headers:
        insert_column_after("gross cm", "Checking", "checking_gross",
            lambda r, d: _calc_gross_cm_checking(r, d, headers))

    if "net cm" in headers:
        insert_column_after("net cm", "Checking", "checking_net",
            lambda r, d: _calc_net_cm_checking(r, d, headers))

    if "ntu(min)" in headers:
        insert_column_after("ntu(min)", "pid103", "pid103",
            lambda r, d: _lookup_pid103(r, d, headers, lookup_data))

    if "pid103" in new_cols:
        insert_column_after("pid103", "Checking", "checking_ntu",
            lambda r, d: _calc_ntu_checking(r, d, headers, new_cols))

    if "checking_ntu" in new_cols:
        insert_column_after("checking_ntu", "Total NTU", "total_ntu",
            lambda r, d: _calc_total_ntu(r, d, headers, new_cols))

    if "total_ntu" in new_cols:
        insert_column_after("total_ntu", "Status", "status",
            lambda r, d: _calc_status(r, d, headers))

    if "line construct" in headers:
        insert_column_after("line construct", "PPG", "ppg",
            lambda r, d: _lookup_ppg(r, d, headers, lookup_data, season_value))


def _get_row_data(ws: Worksheet, row_idx: int, headers: Dict[str, int]) -> Dict[str, Any]:
    """Get all cell values for a row as a dict."""
    data = {}
    for header_name, col_idx in headers.items():
        data[header_name] = ws.cell(row=row_idx, column=col_idx).value
    return data


def _calc_job_short(row_idx: int, row_data: Dict[str, Any]) -> str:
    """Get first 5 characters of Job Number."""
    job_num = _norm(row_data.get("job number", ""))
    return job_num[:5] if job_num else ""


def _lookup_prod_type(
    row_idx: int, row_data: Dict[str, Any], headers: Dict[str, int], lookup_data: Dict[str, Any]
) -> str:
    """Look up P.Type Group from Prod Type lookup table."""
    p_type = _norm(row_data.get("p.type", "")).lower()
    return lookup_data["prod_type"].get(p_type, "")


def _lookup_prod_cat(
    row_idx: int, row_data: Dict[str, Any], headers: Dict[str, int], lookup_data: Dict[str, Any]
) -> str:
    """Look up P.Cat Group from Prod Cat lookup table."""
    p_cat = _norm(row_data.get("p.cat", "")).lower()
    return lookup_data["prod_cat"].get(p_cat, "")


def _calc_season_style(
    row_idx: int, row_data: Dict[str, Any], headers: Dict[str, int], season_value: Optional[str]
) -> str:
    """Calculate Season+Style = <Season><Style #>."""
    style = _norm(row_data.get("style #", ""))
    season = season_value or ""
    return f"{season}{style}" if style else ""


def _lookup_ts_wr(
    row_idx: int,
    row_data: Dict[str, Any],
    headers: Dict[str, int],
    lookup_data: Dict[str, Any],
    season_value: Optional[str],
) -> Any:
    """Look up TS&WR from pco2175 sheet based on Season+Style."""
    season_style = _norm(row_data.get("season+style", "")).lower()
    if not season_style:
        style = _norm(row_data.get("style #", ""))
        season = season_value or ""
        season_style = f"{season}{style}".lower()

    for pco_data in lookup_data["pco2175_sheets"]:
        if season_style in pco_data["season_style_to_wr"]:
            return pco_data["season_style_to_wr"][season_style]
    return ""


def _calc_emb_checking(row_idx: int, _row_data: Dict[str, Any], headers: Dict[str, int]) -> str:
    """Return formula: ROUND(Embellishment Cost - (Emb + Printing + APC + PAD + Laser + Bonding), 3)."""
    emb_col = headers.get("embellishment cost")
    if not emb_col:
        return ""

    emb_ref = f"{get_column_letter(emb_col)}{row_idx}"

    component_refs = []
    for comp in EMBELLISHMENT_COMPONENTS:
        for header_name, col_idx in headers.items():
            if comp in header_name.lower():
                component_refs.append(f"{get_column_letter(col_idx)}{row_idx}")

    if not component_refs:
        return f"=ROUND({emb_ref},3)"

    return f"=ROUND({emb_ref}-" + "-".join(component_refs) + ",3)"


def _calc_gross_cm_checking(row_idx: int, _row_data: Dict[str, Any], headers: Dict[str, int]) -> str:
    """Return formula: (Gross CM - Sales USD - Material Cost - Embellishment Cost - Other Cost)."""
    gross_col = headers.get("gross cm")
    if not gross_col:
        return ""

    gross_ref = f"{get_column_letter(gross_col)}{row_idx}"

    other_refs = []
    for col_name in ["sales usd", "material cost", "embellishment cost", "other cost"]:
        col_idx = headers.get(col_name)
        if col_idx:
            other_refs.append(f"{get_column_letter(col_idx)}{row_idx}")

    if not other_refs:
        return f"={gross_ref}"

    return f"={gross_ref}-(" + "-".join(other_refs) + ")"


def _calc_net_cm_checking(row_idx: int, _row_data: Dict[str, Any], headers: Dict[str, int]) -> str:
    """Return formula: Gross CM - /PO-VTEC CM - Net CM."""
    gross_col = headers.get("gross cm")
    if not gross_col:
        return ""

    gross_ref = f"{get_column_letter(gross_col)}{row_idx}"

    po_vtec_col = None
    for header_name, col_idx in headers.items():
        if "/po-vtec cm" in header_name.lower() or "po-vtec cm" in header_name.lower():
            po_vtec_col = col_idx
            break

    net_col = headers.get("net cm")

    formula = f"={gross_ref}"
    if po_vtec_col:
        formula += f"-{get_column_letter(po_vtec_col)}{row_idx}"
    if net_col:
        formula += f"-{get_column_letter(net_col)}{row_idx}"

    return formula


def _lookup_pid103(
    row_idx: int, row_data: Dict[str, Any], headers: Dict[str, int], lookup_data: Dict[str, Any]
) -> Any:
    """Look up pid103 TOTAL based on Style #."""
    raw_style = row_data.get("style #", "")
    style = _norm(raw_style).lower()
    if not style:
        return ""

    for pid_data in lookup_data["pid103_sheets"]:
        if style in pid_data["style_to_total"]:
            return pid_data["style_to_total"][style]

    # Debug: log first few misses to help identify format differences
    if row_idx <= 10 and lookup_data["pid103_sheets"]:
        available = list(lookup_data["pid103_sheets"][0]["style_to_total"].keys())[:3]
        print(f"DEBUG row {row_idx}: Looking for '{style}' (raw: '{raw_style}'), available samples: {available}")

    return ""


def _calc_ntu_checking(row_idx: int, _row_data: Dict[str, Any], headers: Dict[str, int], new_cols: Dict[str, int]) -> str:
    """Return formula: NTU(min) - pid103."""
    ntu_col = headers.get("ntu(min)")
    pid103_col = new_cols.get("pid103")

    if not ntu_col or not pid103_col:
        return ""

    ntu_ref = f"{get_column_letter(ntu_col)}{row_idx}"
    pid103_ref = f"{get_column_letter(pid103_col)}{row_idx}"

    return f"={ntu_ref}-{pid103_ref}"


def _calc_total_ntu(row_idx: int, _row_data: Dict[str, Any], headers: Dict[str, int], new_cols: Dict[str, int]) -> str:
    """Return formula: pid103 * quantity."""
    pid103_col = new_cols.get("pid103")
    if not pid103_col:
        return ""

    pid103_ref = f"{get_column_letter(pid103_col)}{row_idx}"

    # Find quantity column
    qty_col = None
    for header_name, col_idx in headers.items():
        if "quantity" in header_name.lower() or "qty" in header_name.lower():
            qty_col = col_idx
            break

    if not qty_col:
        return f"={pid103_ref}"

    qty_ref = f"{get_column_letter(qty_col)}{row_idx}"
    return f"={pid103_ref}*{qty_ref}"


def _calc_status(row_idx: int, row_data: Dict[str, Any], headers: Dict[str, int]) -> str:
    """
    Calculate Status:
    - If /PO-VTEC CM is 0 and PO/Quo is 'Q' or 'PO', put 'P'
    - Otherwise if not 0 for Q or PO, put 'C'
    """
    po_vtec_cm = 0
    for header_name in headers.keys():
        if "/po-vtec cm" in header_name.lower() or "po-vtec cm" in header_name.lower():
            po_vtec_cm = _to_number(row_data.get(header_name, 0))
            break

    po_quo = ""
    for header_name in headers.keys():
        if "po/quo" in header_name.lower():
            po_quo = _norm(row_data.get(header_name, "")).upper()
            break

    if po_quo in ("Q", "PO"):
        if po_vtec_cm == 0:
            return "P"
        else:
            return "C"
    return ""


def _lookup_ppg(
    row_idx: int,
    row_data: Dict[str, Any],
    headers: Dict[str, int],
    lookup_data: Dict[str, Any],
    season_value: Optional[str],
) -> str:
    """Look up PPG CODE from pco2175 sheet based on Season+Style."""
    season_style = _norm(row_data.get("season+style", "")).lower()
    if not season_style:
        style = _norm(row_data.get("style #", ""))
        season = season_value or ""
        season_style = f"{season}{style}".lower()

    for pco_data in lookup_data["pco2175_sheets"]:
        if season_style in pco_data["season_style_to_ppg"]:
            return pco_data["season_style_to_ppg"][season_style]
    return ""


def _norm(value: Any) -> str:
    """Normalize a cell value to string - trim and collapse whitespace."""
    if value is None:
        return ""
    s = str(value).strip()
    # Collapse multiple whitespace to single space and handle special whitespace chars
    return " ".join(s.split())


def _to_number(value: Any) -> float:
    """Convert a value to a number, defaulting to 0."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (ValueError, TypeError):
        return 0.0


def _make_bold(cell) -> None:
    """Make a cell bold."""
    try:
        font = copy(cell.font) if cell.font is not None else Font()
        font.bold = True
        cell.font = font
    except Exception:
        pass


def _apply_gray_fill(cell) -> None:
    """Apply gray background fill to a cell."""
    try:
        cell.fill = GRAY_FILL
    except Exception:
        pass


def _build_output_path(src_path: str) -> str:
    """Build output file path with timestamp."""
    base_dir = os.path.dirname(src_path) or os.getcwd()
    base_name = os.path.basename(src_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_name = f"{timestamp}_transformed_{base_name}"
    candidate = os.path.join(base_dir, out_name)

    if not os.path.exists(candidate):
        return candidate

    name, ext = os.path.splitext(out_name)
    counter = 1
    while True:
        candidate = os.path.join(base_dir, f"{name} ({counter}){ext}")
        if not os.path.exists(candidate):
            return candidate
        counter += 1
