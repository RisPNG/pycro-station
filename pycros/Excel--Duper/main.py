import os
import threading
import warnings
from datetime import datetime
from typing import List, Tuple, Any, Optional

from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QVBoxLayout,
    QLabel,
    QTextEdit,
    QWidget,
    QSizePolicy
)
from qfluentwidgets import PrimaryPushButton, MessageBox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

try:
    import xlrd  # for legacy excel
except Exception:
    xlrd = None

class MainWidget(QWidget):
    log_message = Signal(str)
    processing_done = Signal(int, int, str)

    def __init__(self):
        super().__init__()
        self.setObjectName("excel_duper_widget")

        self._build_ui()
        self._connect_signals()

    # UI
    def _build_ui(self):
        # DEFINE WIDGETS

        # Description
        self.desc_label = QLabel("", self)
        self.desc_label.setWordWrap(True)
        self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.desc_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.desc_label.setStyleSheet(
            "color: #dcdcdc; background: transparent; padding: 6px; "
            "border: 1px solid #3a3a3a; border-radius: 6px;"
        )
        self.set_long_description("")

        # Buttons
        self.select_btn = PrimaryPushButton("Select Excel Files", self)
        self.run_btn = PrimaryPushButton("Run", self)

        # Labels
        self.files_label = QLabel("Selected files", self)
        self.files_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.files_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")
        self.logs_label = QLabel("Process logs", self)
        self.logs_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.logs_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        # TextBoxes
        self.files_box = QTextEdit(self)
        self.files_box.setReadOnly(True)
        self.files_box.setPlaceholderText("Selected files will appear here")
        self.files_box.setStyleSheet(
            "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText("Live process log will appear here")
        self.log_box.setStyleSheet(
            "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        # CONSTRUCT LAYOUTS

        # Main vertical layout
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)

        # Row 0
        main_layout.addWidget(self.desc_label, 1)  # Row 0: fixed height

        # Row 1: Select button layout (3 columns with button in middle)
        row1_layout = QHBoxLayout()
        row1_layout.addStretch(1)  # Left spacer
        row1_layout.addWidget(self.select_btn, 1)  # Button in middle
        row1_layout.addStretch(1)  # Right spacer
        main_layout.addLayout(row1_layout, 0)      # Row 1: fixed height

        # Row 2: Run button layout (3 columns with button in middle)
        row2_layout = QHBoxLayout()
        row2_layout.addStretch(1)  # Left spacer
        row2_layout.addWidget(self.run_btn, 1)  # Button in middle
        row2_layout.addStretch(1)  # Right spacer
        main_layout.addLayout(row2_layout, 0)      # Row 2: fixed height

        # Row 3: Labels for files and logs
        row3_layout = QHBoxLayout()
        row3_layout.addWidget(self.files_label, 1)
        row3_layout.addWidget(self.logs_label, 1)
        main_layout.addLayout(row3_layout, 0)

        # Row 4: Files and logs layout
        row4_layout = QHBoxLayout()
        row4_layout.addWidget(self.files_box, 1)
        row4_layout.addWidget(self.log_box, 1)
        main_layout.addLayout(row4_layout, 4)      # Row 4: grows to fill space


    # For displaying long description from description.md
    def set_long_description(self, text: str):
        clean = (text or "").strip()
        if clean:
            self.desc_label.setText(clean)
            self.desc_label.show()
        else:
            self.desc_label.clear()
            self.desc_label.hide()

    # Middleman from UI to Functions
    def _connect_signals(self):
        self.select_btn.clicked.connect(self.select_files)
        self.run_btn.clicked.connect(self.run_process)
        self.log_message.connect(self.append_log)
        self.processing_done.connect(self.on_processing_done)

    # Functions
    def select_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, "Select")
        if files:
            self.files_box.setPlainText("\n".join(files))
        else:
            self.files_box.clear()

    def _selected_files(self) -> List[str]:
        text = self.files_box.toPlainText().strip()
        if not text:
            return []
        return [line for line in text.split("\n") if line.strip()]

    def run_process(self):
        files = self._selected_files()
        if not files:
            MessageBox("Warning", "Nothing to process.", self).exec()
            return

        self.log_box.clear()
        self.log_message.emit(f"Process starts")
        self.run_btn.setEnabled(False)
        self.select_btn.setEnabled(False)

        def worker():
            ok, fail, out_path = 0, 0, ""
            try:
                out_path, ok, fail = process_files(files, self.log_message.emit)
            except Exception as e:
                self.log_message.emit(f"ERROR: {e}")
            self.processing_done.emit(ok, fail, out_path)

        threading.Thread(target=worker, daemon=True).start()

    def append_log(self, text: str):
        self.log_box.append(text)
        self.log_box.ensureCursorVisible()

    def on_processing_done(self, ok: int, fail: int, out_path: str):
        if out_path:
            self.log_message.emit(f"Output workbook saved to: {out_path}")
        self.log_message.emit(f"Completed: {ok} success, {fail} failed.")
        self.run_btn.setEnabled(True)
        self.select_btn.setEnabled(True)
        title = "Processing complete" if fail == 0 else "Processing finished with issues"
        lines = [f"Success: {ok}", f"Failed: {fail}"]
        if out_path:
            lines.append(f"Output: {os.path.basename(out_path)}")
        else:
            lines.append("Output: (see logs)")
        msg = MessageBox(title, "\n".join(lines), self)
        msg.yesButton.setText("OK")
        msg.cancelButton.hide()
        msg.exec()

def get_widget():
    return MainWidget()

# Pycro Main Process

def process_files(file_paths: List[str], log_emit) -> Tuple[str, int, int]:
    """
    Duplicate each selected workbook, trimming trailing empty rows/columns
    on every worksheet while preserving data, formulas, and styles.

    Returns (out_path_for_ui, success_count, fail_count).
    out_path_for_ui is only populated when exactly one file is processed
    successfully so the UI message stays meaningful.
    """
    if log_emit is None:
        def _noop(_msg: str):
            pass
        log_emit = _noop

    total = len(file_paths)
    success = 0
    fail = 0
    last_output_path = ""

    log_emit("Process Begin...")

    for index, path in enumerate(file_paths, start=1):
        label = f"({index}/{total}) {os.path.basename(path)}"
        try:
            if not os.path.isfile(path):
                raise FileNotFoundError(f"File not found: {path}")

            ext = os.path.splitext(path)[1].lower()
            if ext == ".xls":
                raise ValueError("Legacy .xls files are not supported. Please convert to .xlsx before using this pycro.")

            size_bytes = os.path.getsize(path)
            keep_vba = ext in (".xlsm", ".xltm", ".xlam")
            log_emit(
                f"{label} - Opening workbook (size={size_bytes} bytes, ext={ext}, keep_vba={keep_vba})..."
            )

            open_started = datetime.now()
            with warnings.catch_warnings():
                warnings.filterwarnings(
                    "ignore",
                    message="wmf image format is not supported so the image is being dropped",
                    category=UserWarning,
                    module="openpyxl.reader.drawings",
                )
                wb = load_workbook(path, data_only=False, keep_vba=keep_vba)
            open_elapsed = (datetime.now() - open_started).total_seconds()
            log_emit(
                f"{label} - Workbook opened in {open_elapsed:.2f}s with {len(wb.worksheets)} sheet(s)."
            )

            sheet_total = len(wb.worksheets)
            for sheet_idx, ws in enumerate(wb.worksheets, start=1):
                log_emit(f"{label} - Scanning sheet {sheet_idx}/{sheet_total}: {ws.title}")
                max_used_row = 0
                max_used_col = 0

                # Scan for cells that actually contain a value or formula.
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            if cell.row > max_used_row:
                                max_used_row = cell.row
                            if cell.column > max_used_col:
                                max_used_col = cell.column

                # Ensure merged ranges containing a value are fully kept.
                if ws.merged_cells.ranges and (max_used_row or max_used_col):
                    for cell_range in ws.merged_cells.ranges:
                        top_left = ws.cell(row=cell_range.min_row, column=cell_range.min_col)
                        if top_left.value is not None:
                            if cell_range.max_row > max_used_row:
                                max_used_row = cell_range.max_row
                            if cell_range.max_col > max_used_col:
                                max_used_col = cell_range.max_col

                # If the sheet is completely empty, leave it as-is.
                if max_used_row == 0 and max_used_col == 0:
                    log_emit(f"{label} - Sheet '{ws.title}' is empty, skipping trim.")
                    continue

                # Remove trailing completely empty rows/columns.
                if ws.max_row > max_used_row:
                    ws.delete_rows(max_used_row + 1, ws.max_row - max_used_row)
                if ws.max_column > max_used_col:
                    ws.delete_cols(max_used_col + 1, ws.max_column - max_used_col)

                log_emit(
                    f"{label} - Sheet '{ws.title}' trimmed to max_row={max_used_row}, max_col={max_used_col}."
                )

            # Build output path: YYYYMMDD_HHMMSS_duped_originalfilename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_dir = os.path.dirname(path) or os.getcwd()
            base_name = os.path.basename(path)
            out_name = f"{timestamp}_duped_{base_name}"
            out_path = os.path.join(base_dir, out_name)

            # Avoid accidental overwrite if a file with the same name already exists.
            if os.path.exists(out_path):
                counter = 1
                name, ext_out = os.path.splitext(out_name)
                while True:
                    candidate = os.path.join(base_dir, f"{name} ({counter}){ext_out}")
                    if not os.path.exists(candidate):
                        out_path = candidate
                        break
                    counter += 1

            wb.save(out_path)

            success += 1
            last_output_path = out_path
            log_emit(f"{label} - Output workbook saved to: {out_path}")
        except Exception as exc:
            fail += 1
            log_emit(f"{label} - Error: {exc}")

    log_emit("Process Completed")

    ui_out_path = last_output_path if success == 1 else ""
    return ui_out_path, success, fail
