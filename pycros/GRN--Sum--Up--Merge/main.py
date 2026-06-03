#!/usr/bin/env python3
from __future__ import annotations

import argparse
import threading
from copy import copy
from datetime import date, datetime, time
from pathlib import Path
from typing import Callable, Iterable, Optional, TypedDict

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

try:
    from PySide6.QtCore import Qt, Signal
    from PySide6.QtWidgets import QFileDialog, QHBoxLayout, QLabel, QSizePolicy, QTextEdit, QVBoxLayout, QWidget
    from qfluentwidgets import MessageBox, PrimaryPushButton

    GUI_AVAILABLE = True
except Exception:
    Qt = Signal = QFileDialog = QHBoxLayout = QLabel = QSizePolicy = QTextEdit = QVBoxLayout = QWidget = None
    MessageBox = PrimaryPushButton = None
    GUI_AVAILABLE = False


SHEET_NAME = "GRN Sum"
REFERENCE_DIR = Path("/home/faris/Downloads/GRN Sum Up Merge")
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
DOC_DATE_COL = 3
WHSE_COL = 12


class RowDimensionStyle(TypedDict):
    height: Optional[float]
    hidden: bool
    outline_level: int
    collapsed: bool


class ColumnDimensionStyle(TypedDict):
    letter: str
    width: Optional[float]
    hidden: bool
    best_fit: bool
    outline_level: int
    collapsed: bool


class CellStyle(TypedDict, total=False):
    font: object
    fill: object
    border: object
    alignment: object
    number_format: str
    protection: object


class VisualTemplate(TypedDict):
    path: Path
    header_styles: list[CellStyle]
    data_styles: list[CellStyle]
    column_dimensions: list[ColumnDimensionStyle]
    header_row_dimension: RowDimensionStyle
    data_row_dimension: RowDimensionStyle
    sheet_format: object
    sheet_properties: object
    page_margins: object
    page_setup: object
    print_options: object
    views: object
    freeze_panes: object
    auto_filter_ref: Optional[str]


def _emit(log_emit: Optional[Callable[[str], None]], text: str) -> None:
    if callable(log_emit):
        try:
            log_emit(text)
            return
        except Exception:
            pass
    print(text)


def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip().upper()


def is_blank(value) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def is_blank_row(row: Iterable[object]) -> bool:
    return all(is_blank(value) for value in row)


def is_lock_file(path: Path) -> bool:
    name = path.name
    return name.startswith("~$") or name.startswith(".~lock") or name.endswith("#")


def iter_input_files(items: Iterable[str]) -> list[Path]:
    files: list[Path] = []
    for item in items:
        path = Path(item).expanduser()
        if path.is_dir():
            for candidate in sorted(path.iterdir()):
                if candidate.is_file():
                    files.append(candidate)
        elif path.is_file():
            files.append(path)

    seen: set[Path] = set()
    out: list[Path] = []
    for path in files:
        resolved = path.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        if is_lock_file(path):
            continue
        if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            continue
        out.append(path)
    return out


def ensure_unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    base = path.with_suffix("")
    ext = path.suffix
    n = 1
    while True:
        candidate = Path(f"{base} ({n}){ext}")
        if not candidate.exists():
            return candidate
        n += 1


def proposed_output_path(files: list[Path]) -> Path:
    base_dir = files[0].parent if files else Path.cwd()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return ensure_unique_path(base_dir / f"GRN-Sum-Merged-{timestamp}.xlsx")


def trim_row(row: Iterable[object], width: int) -> list[object]:
    values = list(row)
    if len(values) < width:
        values.extend([None] * (width - len(values)))
    return values[:width]


def headers_match(left: list[object], right: list[object]) -> bool:
    if len(left) != len(right):
        return False
    return [normalize_text(value) for value in left] == [normalize_text(value) for value in right]


def validate_header(header: list[object], log_emit=None) -> None:
    doc_header = normalize_text(header[DOC_DATE_COL - 1]) if len(header) >= DOC_DATE_COL else ""
    whse_header = normalize_text(header[WHSE_COL - 1]) if len(header) >= WHSE_COL else ""
    if doc_header != "DOC DATE":
        _emit(log_emit, f"[WARN] Column C header is '{header[DOC_DATE_COL - 1] if len(header) >= DOC_DATE_COL else ''}', expected 'Doc Date'.")
    if whse_header != "WHSE":
        _emit(log_emit, f"[WARN] Column L header is '{header[WHSE_COL - 1] if len(header) >= WHSE_COL else ''}', expected 'Whse'.")


def capture_cell_style(cell) -> CellStyle:
    style: CellStyle = {}
    if cell.has_style:
        style["font"] = copy(cell.font)
        style["fill"] = copy(cell.fill)
        style["border"] = copy(cell.border)
        style["alignment"] = copy(cell.alignment)
        style["number_format"] = cell.number_format
        style["protection"] = copy(cell.protection)
    return style


def apply_cell_style(cell, style: CellStyle) -> None:
    if "font" in style:
        cell.font = style["font"]
    if "fill" in style:
        cell.fill = style["fill"]
    if "border" in style:
        cell.border = style["border"]
    if "alignment" in style:
        cell.alignment = style["alignment"]
    if "number_format" in style:
        cell.number_format = style["number_format"]
    if "protection" in style:
        cell.protection = style["protection"]


def make_styled_cell(worksheet, value, style: CellStyle):
    cell = WriteOnlyCell(worksheet, value=value)
    apply_cell_style(cell, style)
    return cell


def capture_column_dimension(worksheet, column_index: int) -> ColumnDimensionStyle:
    letter = get_column_letter(column_index)
    dimension = worksheet.column_dimensions[letter]
    return {
        "letter": letter,
        "width": dimension.width,
        "hidden": bool(dimension.hidden),
        "best_fit": bool(getattr(dimension, "bestFit", False)),
        "outline_level": int(getattr(dimension, "outlineLevel", 0) or 0),
        "collapsed": bool(dimension.collapsed),
    }


def apply_column_dimension(worksheet, style: ColumnDimensionStyle) -> None:
    dimension = worksheet.column_dimensions[style["letter"]]
    dimension.width = style["width"]
    dimension.hidden = style["hidden"]
    dimension.bestFit = style["best_fit"]
    dimension.outlineLevel = style["outline_level"]
    dimension.collapsed = style["collapsed"]


def capture_row_dimension(worksheet, row_index: int) -> RowDimensionStyle:
    dimension = worksheet.row_dimensions[row_index]
    return {
        "height": dimension.height,
        "hidden": bool(dimension.hidden),
        "outline_level": int(getattr(dimension, "outlineLevel", 0) or 0),
        "collapsed": bool(dimension.collapsed),
    }


def row_dimension_has_visual(style: RowDimensionStyle) -> bool:
    return (
        style["height"] is not None
        or style["hidden"]
        or style["outline_level"] != 0
        or style["collapsed"]
    )


def apply_row_dimension(worksheet, row_index: int, style: RowDimensionStyle) -> None:
    dimension = worksheet.row_dimensions[row_index]
    if style["height"] is not None:
        dimension.height = style["height"]
    dimension.hidden = style["hidden"]
    dimension.outlineLevel = style["outline_level"]
    dimension.collapsed = style["collapsed"]


def load_visual_template(template_path: Path, width: int, log_emit=None) -> VisualTemplate:
    _emit(log_emit, f"[STYLE] Using visual template from: {template_path.name}")
    workbook = load_workbook(template_path, read_only=False, data_only=False)
    try:
        worksheet = workbook[SHEET_NAME]
        return {
            "path": template_path,
            "header_styles": [capture_cell_style(worksheet.cell(1, column_index)) for column_index in range(1, width + 1)],
            "data_styles": [capture_cell_style(worksheet.cell(2, column_index)) for column_index in range(1, width + 1)],
            "column_dimensions": [capture_column_dimension(worksheet, column_index) for column_index in range(1, width + 1)],
            "header_row_dimension": capture_row_dimension(worksheet, 1),
            "data_row_dimension": capture_row_dimension(worksheet, 2),
            "sheet_format": copy(worksheet.sheet_format),
            "sheet_properties": copy(worksheet.sheet_properties),
            "page_margins": copy(worksheet.page_margins),
            "page_setup": copy(worksheet.page_setup),
            "print_options": copy(worksheet.print_options),
            "views": copy(worksheet.views),
            "freeze_panes": worksheet.freeze_panes,
            "auto_filter_ref": worksheet.auto_filter.ref,
        }
    finally:
        workbook.close()


def apply_visual_template(worksheet, template: VisualTemplate, width: int) -> None:
    for attr in ("sheet_format", "sheet_properties", "page_margins", "page_setup", "print_options", "views"):
        try:
            setattr(worksheet, attr, copy(template[attr]))
        except Exception:
            pass

    for column_style in template["column_dimensions"]:
        apply_column_dimension(worksheet, column_style)

    apply_row_dimension(worksheet, 1, template["header_row_dimension"])
    worksheet.freeze_panes = template["freeze_panes"]
    worksheet.auto_filter.ref = template["auto_filter_ref"] or f"A1:{get_column_letter(width)}1"


def parse_doc_date(value) -> tuple[int, datetime]:
    """
    Sort dates ascending. Missing/unparseable values go after valid dates inside
    their non-VN/VN group while preserving original order as a later key.
    """
    if value is None or value == "":
        return (1, datetime.max)
    if isinstance(value, datetime):
        return (0, value)
    if isinstance(value, date):
        return (0, datetime.combine(value, time.min))
    if isinstance(value, bool):
        return (1, datetime.max)
    if isinstance(value, (int, float)):
        try:
            parsed = from_excel(value)
            if isinstance(parsed, datetime):
                return (0, parsed)
            if isinstance(parsed, date):
                return (0, datetime.combine(parsed, time.min))
        except Exception:
            return (1, datetime.max)

    text = str(value).strip()
    if not text:
        return (1, datetime.max)

    formats = (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%m-%d-%Y",
        "%d-%b-%Y",
        "%d-%b-%y",
        "%d %b %Y",
        "%d %b %y",
        "%Y %b %d",
        "%Y %B %d",
    )
    for fmt in formats:
        try:
            return (0, datetime.strptime(text, fmt))
        except Exception:
            continue
    try:
        return (0, datetime.fromisoformat(text))
    except Exception:
        return (1, datetime.max)


def whse_is_vn(row: list[object]) -> bool:
    if len(row) < WHSE_COL:
        return False
    return normalize_text(row[WHSE_COL - 1]) == "VN"


def sort_and_split_rows(rows: list[tuple[int, list[object]]]) -> tuple[list[list[object]], list[list[object]]]:
    non_vn: list[tuple[int, list[object]]] = []
    vn: list[tuple[int, list[object]]] = []
    for seq, row in rows:
        if whse_is_vn(row):
            vn.append((seq, row))
        else:
            non_vn.append((seq, row))

    non_vn.sort(
        key=lambda item: (
            *parse_doc_date(item[1][DOC_DATE_COL - 1] if len(item[1]) >= DOC_DATE_COL else None),
            item[0],
        )
    )
    vn.sort(
        key=lambda item: (
            *parse_doc_date(item[1][DOC_DATE_COL - 1] if len(item[1]) >= DOC_DATE_COL else None),
            item[0],
        )
    )
    return [row for _, row in non_vn], [row for _, row in vn]


def collect_rows(files: list[Path], log_emit=None) -> tuple[list[object], list[tuple[int, list[object]]], Path, int, int]:
    header: Optional[list[object]] = None
    template_path: Optional[Path] = None
    rows: list[tuple[int, list[object]]] = []
    ok_files = 0
    fail_files = 0
    row_sequence = 0

    for path in files:
        _emit(log_emit, f"[OPEN] {path.name}")
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            fail_files += 1
            _emit(log_emit, f"[FAIL] {path.name}: could not open workbook ({exc})")
            continue

        try:
            if SHEET_NAME not in workbook.sheetnames:
                fail_files += 1
                _emit(log_emit, f"[SKIP] {path.name}: sheet '{SHEET_NAME}' not found.")
                continue

            worksheet = workbook[SHEET_NAME]
            iterator = worksheet.iter_rows(values_only=True)
            source_header = next(iterator, None)
            if source_header is None or is_blank_row(source_header):
                fail_files += 1
                _emit(log_emit, f"[SKIP] {path.name}: '{SHEET_NAME}' has no header row.")
                continue

            if header is None:
                header = list(source_header)
                template_path = path
                validate_header(header, log_emit)
            else:
                source_header = trim_row(source_header, len(header))
                if not headers_match(source_header, header):
                    _emit(log_emit, f"[WARN] {path.name}: header differs from first workbook; rows will be copied by column position.")

            width = len(header)
            appended = 0
            for source_row in iterator:
                row = trim_row(source_row, width)
                if is_blank_row(row):
                    continue
                rows.append((row_sequence, row))
                row_sequence += 1
                appended += 1

            ok_files += 1
            _emit(log_emit, f"[OK] {path.name}: appended {appended} row(s)")
        except Exception as exc:
            fail_files += 1
            _emit(log_emit, f"[FAIL] {path.name}: {exc}")
        finally:
            workbook.close()

    if header is None:
        raise ValueError(f"No usable '{SHEET_NAME}' sheets found.")
    if template_path is None:
        raise ValueError(f"No visual template found for '{SHEET_NAME}'.")

    return header, rows, template_path, ok_files, fail_files


def _write_sheet(worksheet, header: list[object], rows: list[list[object]], visual_template: VisualTemplate, width: int) -> None:
    header_cells = [
        make_styled_cell(worksheet, value, visual_template["header_styles"][column_index])
        for column_index, value in enumerate(header)
    ]
    worksheet.append(header_cells)

    data_row_has_visual = row_dimension_has_visual(visual_template["data_row_dimension"])
    data_styles = visual_template["data_styles"]
    # Pre-extract style attributes to avoid dict lookups in the hot loop
    fonts = [s.get("font") for s in data_styles]
    fills = [s.get("fill") for s in data_styles]
    borders = [s.get("border") for s in data_styles]
    alignments = [s.get("alignment") for s in data_styles]
    number_formats = [s.get("number_format") for s in data_styles]
    protections = [s.get("protection") for s in data_styles]

    for row_index, row in enumerate(rows, start=2):
        if data_row_has_visual:
            apply_row_dimension(worksheet, row_index, visual_template["data_row_dimension"])
        styled_row: list[object] = []
        for column_index in range(width):
            cell = WriteOnlyCell(worksheet, value=row[column_index] if column_index < len(row) else None)
            if fonts[column_index] is not None:
                cell.font = fonts[column_index]
            if fills[column_index] is not None:
                cell.fill = fills[column_index]
            if borders[column_index] is not None:
                cell.border = borders[column_index]
            if alignments[column_index] is not None:
                cell.alignment = alignments[column_index]
            if number_formats[column_index] is not None:
                cell.number_format = number_formats[column_index]
            if protections[column_index] is not None:
                cell.protection = protections[column_index]
            styled_row.append(cell)
        worksheet.append(styled_row)


def write_output(path: Path, header: list[object], non_vn_rows: list[list[object]], vn_rows: list[list[object]], visual_template: VisualTemplate, log_emit=None) -> None:
    workbook = Workbook(write_only=True)
    width = len(header)

    non_vn_sheet = workbook.create_sheet(SHEET_NAME)
    apply_visual_template(non_vn_sheet, visual_template, width)
    _write_sheet(non_vn_sheet, header, non_vn_rows, visual_template, width)
    _emit(log_emit, f"[SHEET] {SHEET_NAME}: {len(non_vn_rows)} row(s)")

    vn_sheet = workbook.create_sheet(f"{SHEET_NAME} VN")
    apply_visual_template(vn_sheet, visual_template, width)
    _write_sheet(vn_sheet, header, vn_rows, visual_template, width)
    _emit(log_emit, f"[SHEET] {SHEET_NAME} VN: {len(vn_rows)} row(s)")

    workbook.save(path)
    _emit(log_emit, f"[DONE] Output saved: {path}")


def process_files(
    items: list[str],
    log_emit=None,
    output: Optional[Path] = None,
) -> tuple[str, int, int, int]:
    files = iter_input_files(items)
    if not files:
        raise ValueError("No supported Excel files selected (.xlsx/.xlsm/.xltx/.xltm).")

    _emit(log_emit, f"Files to process: {len(files)}")
    header, indexed_rows, template_path, ok_files, fail_files = collect_rows(files, log_emit=log_emit)
    _emit(log_emit, f"Rows collected: {len(indexed_rows)}")
    _emit(log_emit, "Sorting rows: non-VN by Doc Date, VN by Doc Date (separate sheets).")
    non_vn_rows, vn_rows = sort_and_split_rows(indexed_rows)

    out_path = ensure_unique_path(output) if output else proposed_output_path(files)
    visual_template = load_visual_template(template_path, len(header), log_emit=log_emit)
    write_output(out_path, header, non_vn_rows, vn_rows, visual_template, log_emit=log_emit)
    return str(out_path), ok_files, fail_files, len(non_vn_rows) + len(vn_rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Merge selected workbooks' 'GRN Sum' sheets into one sorted workbook.")
    parser.add_argument("paths", nargs="+", help="Input Excel files or folders.")
    parser.add_argument("-o", "--output", help="Output .xlsx path. Defaults beside the first input file.")
    args = parser.parse_args()

    output = Path(args.output).expanduser() if args.output else None
    out_path, ok, fail, total_rows = process_files(args.paths, output=output)
    print(f"\nDone. Success: {ok}, Failed: {fail}, Rows: {total_rows}")
    print(f"Wrote merged file to: {out_path}")


if GUI_AVAILABLE:

    class MainWidget(QWidget):
        log_message = Signal(str)
        processing_done = Signal(int, int, int, str)

        def __init__(self):
            super().__init__()
            self.setObjectName("grn_sum_up_merge_widget")
            self._build_ui()
            self._connect_signals()

        def _build_ui(self):
            self.desc_label = QLabel("", self)
            self.desc_label.setWordWrap(True)
            self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
            self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
            self.desc_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
            self.desc_label.setStyleSheet(
                "color: #dcdcdc; background: transparent; padding: 6px; "
                "border: 1px solid #3a3a3a; border-radius: 6px;"
            )
            self.set_long_description("")

            self.select_btn = PrimaryPushButton("Select GRN Sum Files", self)
            self.run_btn = PrimaryPushButton("Merge", self)

            self.files_label = QLabel("Selected files", self)
            self.files_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.files_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

            self.logs_label = QLabel("Process logs", self)
            self.logs_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.logs_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

            shared_style = (
                "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
                "border: 1px solid #3a3a3a; border-radius: 6px;}"
            )

            self.files_box = QTextEdit(self)
            self.files_box.setReadOnly(True)
            self.files_box.setPlaceholderText("Selected Excel files will appear here")
            self.files_box.setStyleSheet(shared_style)

            self.log_box = QTextEdit(self)
            self.log_box.setReadOnly(True)
            self.log_box.setPlaceholderText("Live process log will appear here")
            self.log_box.setStyleSheet(shared_style)

            main_layout = QVBoxLayout(self)
            main_layout.setContentsMargins(16, 16, 16, 16)
            main_layout.setSpacing(12)

            main_layout.addWidget(self.desc_label, 1)

            select_row = QHBoxLayout()
            select_row.addStretch(1)
            select_row.addWidget(self.select_btn, 1)
            select_row.addStretch(1)
            main_layout.addLayout(select_row, 0)

            run_row = QHBoxLayout()
            run_row.addStretch(1)
            run_row.addWidget(self.run_btn, 1)
            run_row.addStretch(1)
            main_layout.addLayout(run_row, 0)

            label_row = QHBoxLayout()
            label_row.addWidget(self.files_label, 1)
            label_row.addWidget(self.logs_label, 1)
            main_layout.addLayout(label_row, 0)

            box_row = QHBoxLayout()
            box_row.addWidget(self.files_box, 1)
            box_row.addWidget(self.log_box, 1)
            main_layout.addLayout(box_row, 4)

        def set_long_description(self, text: str):
            clean = (text or "").strip()
            if clean:
                self.desc_label.setText(clean)
                self.desc_label.show()
            else:
                self.desc_label.clear()
                self.desc_label.hide()

        def _connect_signals(self):
            self.select_btn.clicked.connect(self.select_files)
            self.run_btn.clicked.connect(self.run_process)
            self.log_message.connect(self.append_log)
            self.processing_done.connect(self.on_done)

        def select_files(self):
            start_dir = str(REFERENCE_DIR) if REFERENCE_DIR.exists() else ""
            files, _ = QFileDialog.getOpenFileNames(
                self,
                "Select GRN Sum Files",
                start_dir,
                "Excel Files (*.xlsx *.xlsm *.xltx *.xltm)",
            )
            if files:
                self.files_box.setPlainText("\n".join(files))
            else:
                self.files_box.clear()

        def _selected_files(self) -> list[str]:
            text = self.files_box.toPlainText().strip()
            if not text:
                return []
            return [line.strip() for line in text.splitlines() if line.strip()]

        def run_process(self):
            inputs = self._selected_files()
            if not inputs:
                MessageBox("Warning", "Please select Excel files to merge.", self).exec()
                return

            self.log_box.clear()
            self.log_message.emit("GRN Sum merge started...")
            self.log_message.emit("Output will be saved beside the first selected file with a timestamp suffix.")
            self.log_message.emit("")
            self.run_btn.setEnabled(False)
            self.select_btn.setEnabled(False)

            def worker():
                try:
                    out_path, ok, fail, total_rows = process_files(inputs, log_emit=self.log_message.emit)
                    self.processing_done.emit(ok, fail, total_rows, out_path)
                except Exception as exc:
                    self.log_message.emit(f"CRITICAL ERROR: {exc}")
                    self.processing_done.emit(0, 0, 0, "")

            threading.Thread(target=worker, daemon=True).start()

        def append_log(self, text: str):
            self.log_box.append(text)
            self.log_box.ensureCursorVisible()

        def on_done(self, ok: int, fail: int, total_rows: int, out_path: str):
            self.log_message.emit("")
            self.log_message.emit(f"Merge complete: {ok} succeeded, {fail} failed, {total_rows} row(s) written")
            if out_path:
                self.log_message.emit(f"Output: {out_path}")

            self.run_btn.setEnabled(True)
            self.select_btn.setEnabled(True)

            title = "Merge complete" if fail == 0 else "Merge finished with issues"
            lines = [f"Success: {ok}", f"Failed: {fail}", f"Rows written: {total_rows}"]
            if out_path:
                lines.append(f"Output: {out_path}")
            msg = MessageBox(title, "\n".join(lines), self)
            msg.yesButton.setText("OK")
            msg.cancelButton.hide()
            msg.exec()

else:

    class MainWidget:
        def __init__(self):
            raise RuntimeError("PySide6 and qfluentwidgets are required to launch this pycro in Pycro Station.")


def get_widget():
    return MainWidget()


if __name__ == "__main__":
    main()
