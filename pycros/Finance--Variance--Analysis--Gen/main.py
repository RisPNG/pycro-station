#!/usr/bin/env python3
from __future__ import annotations

import re
import threading
import math
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

import xlrd
from openpyxl import load_workbook, Workbook

# GUI Imports (for Pycro Station)
from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QVBoxLayout,
    QLabel,
    QTextEdit,
    QWidget,
    QSizePolicy,
)
from qfluentwidgets import PrimaryPushButton, MessageBox, ComboBox


# ===== Unified Workbook Wrapper for .xls and .xlsx =====

class MissingRequiredHeadersError(ValueError):
    def __init__(self, missing_by_sheet: Dict[str, List[str]]):
        self.missing_by_sheet = missing_by_sheet
        super().__init__(self.user_message())

    def user_message(self) -> str:
        lines = [
            "Cannot continue because the following required header(s) were not found in the selected Management Account file(s):",
            "",
        ]
        for sheet_name in sorted(self.missing_by_sheet.keys()):
            headers = ", ".join(self.missing_by_sheet[sheet_name])
            lines.append(f"{sheet_name}: {headers}")
        return "\n".join(lines)


class UnifiedWorksheet:
    """Wrapper to provide a unified interface for xlrd and openpyxl worksheets."""

    def __init__(self, ws, is_xlrd: bool = False):
        self._ws = ws
        self._is_xlrd = is_xlrd

    @property
    def max_row(self) -> int:
        if self._is_xlrd:
            return self._ws.nrows
        return self._ws.max_row or 0

    @property
    def max_column(self) -> int:
        if self._is_xlrd:
            return self._ws.ncols
        return self._ws.max_column or 0

    def cell_value(self, row: int, col: int) -> Any:
        """Get cell value. Row and col are 1-indexed (like openpyxl)."""
        if self._is_xlrd:
            # xlrd uses 0-indexed
            try:
                return self._ws.cell_value(row - 1, col - 1)
            except IndexError:
                return None
        else:
            return self._ws.cell(row=row, column=col).value

    def cell_displayed_value(self, row: int, col: int) -> str:
        """Get displayed value as string. Row and col are 1-indexed."""
        val = self.cell_value(row, col)
        if val is None:
            return ""
        return str(val)


class UnifiedWorkbook:
    """Wrapper to provide a unified interface for xlrd and openpyxl workbooks."""

    def __init__(self, wb, is_xlrd: bool = False):
        self._wb = wb
        self._is_xlrd = is_xlrd

    @property
    def sheetnames(self) -> List[str]:
        if self._is_xlrd:
            return self._wb.sheet_names()
        return self._wb.sheetnames

    def __contains__(self, name: str) -> bool:
        return name in self.sheetnames

    def __getitem__(self, name: str) -> UnifiedWorksheet:
        if self._is_xlrd:
            return UnifiedWorksheet(self._wb.sheet_by_name(name), is_xlrd=True)
        return UnifiedWorksheet(self._wb[name], is_xlrd=False)

    def get_sheet(self, name: str) -> Optional[UnifiedWorksheet]:
        if name not in self.sheetnames:
            return None
        return self[name]


def load_workbook_unified(filepath: str, data_only: bool = True) -> UnifiedWorkbook:
    """Load a workbook using the appropriate library based on file extension."""
    ext = Path(filepath).suffix.lower()
    if ext == ".xls":
        wb = xlrd.open_workbook(filepath)
        return UnifiedWorkbook(wb, is_xlrd=True)
    else:
        wb = load_workbook(filepath, data_only=data_only)
        return UnifiedWorkbook(wb, is_xlrd=False)


# ===== Constants =====
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_FULL = {
    "Jan": "JAN", "Feb": "FEB", "Mar": "MAR", "Apr": "APR",
    "May": "MAY", "Jun": "JUN", "Jul": "JUL", "Aug": "AUG",
    "Sep": "SEP", "Oct": "OCT", "Nov": "NOV", "Dec": "DEC"
}
MONTH_TO_NUM = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12
}

# Items to exclude for Higher General Administrative Expenses
GAE_EXCLUDE = ["Bonus", "Performance Incentives", "Charity & Donations", "Write-Off Fixed Asset", "Office Rental"]

# Page 7 (General Administrative Expenses) payroll-related items to display as a single group.
PAGE7_PAYROLL_ITEMS = {
    "Salaries, Wages & Related Cost",
    "Allowance",
    "Bonus",
    "Performance Incentives",
    "E.P.F.",
    "EIS (SIP)",
    "Gratuity Payment",
    "Overtime",
    "Red Packets",
    "Socso",
}
PAGE7_PAYROLL_ITEMS_NORM = {item.strip().casefold() for item in PAGE7_PAYROLL_ITEMS}


def _emit(log_emit, text: str):
    if callable(log_emit):
        try:
            log_emit(text)
            return
        except Exception:
            pass
    print(text)


def header_matches_cell(cell_text: Any, header: str) -> bool:
    """
    Return True if a MA header cell matches the desired header.
    Supports whitespace variants like 'NOV 2025' vs 'NOV2025'.
    """
    if cell_text is None or header is None:
        return False

    cell_upper = str(cell_text).strip().upper()
    header_upper = str(header).strip().upper()

    if not cell_upper or not header_upper:
        return False

    if header_upper in cell_upper:
        return True

    cell_compact = re.sub(r"\s+", "", cell_upper)
    header_compact = re.sub(r"\s+", "", header_upper)
    return header_compact in cell_compact


def get_cell_displayed_value(cell) -> str:
    """Get the displayed value of an openpyxl cell as string."""
    if cell.value is None:
        return ""
    value = cell.value

    if isinstance(value, str):
        return value

    # Some openpyxl versions can return rich-text wrapper objects (e.g. Text).
    # Prefer extracting their plain content when available.
    if hasattr(value, "content"):
        try:
            content = value.content
            if content is not None:
                return str(content)
        except Exception:
            pass

    if hasattr(value, "plain"):
        try:
            plain = value.plain
            if plain is not None:
                return str(plain)
        except Exception:
            pass

    return str(value)


def get_cell_raw_value(cell) -> Any:
    """Get the raw value of an openpyxl cell (formula or value)."""
    return cell.value


def _is_excel_row_hidden(ws, row: int) -> bool:
    if row not in ws.row_dimensions:
        return False

    dim = ws.row_dimensions[row]
    if bool(getattr(dim, "hidden", False)):
        return True

    height = getattr(dim, "height", None)
    if isinstance(height, (int, float)) and height <= 0:
        return True

    if bool(getattr(dim, "zeroHeight", False)):
        return True

    return False


def _analytical_review_group_requires_actions(ws, ws_data, start_row: int, end_row_exclusive: int) -> bool:
    start_row = max(1, start_row)
    end_row_exclusive = max(start_row, end_row_exclusive)

    for row in range(start_row, end_row_exclusive):
        for col in (1, 2):
            displayed_val = get_cell_displayed_value(ws_data.cell(row=row, column=col)).strip()
            if "Description" in displayed_val:
                return True

    # Page 9B '*' summary processing is independent of "Description" placeholders.
    return _find_star_cell_row_in_column_b(ws, ws_data, start_row, end_row_exclusive) is not None


def _normalize_excel_text(value: str) -> str:
    """
    Normalize text read from Excel:
    - convert NBSP to space
    - remove zero-width spaces
    - collapse whitespace
    """
    if not value:
        return ""
    value = value.replace("\u00a0", " ").replace("\u200b", "")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def _extract_excel_constant_string_formula(value: str) -> Optional[str]:
    """
    If an Excel cell contains a formula that is *exactly* a single string literal
    like ="Analytical Review (Nov'25 Vs Oct'25)", return that literal.
    """
    if not value:
        return None
    match = re.fullmatch(r'\s*=\s*"((?:[^"]|"")*)"\s*', value)
    if not match:
        return None
    return match.group(1).replace('""', '"')


def _split_excel_concat_expression(expr: str) -> Optional[List[str]]:
    """
    Split an Excel string concatenation expression into parts, using & as the separator.
    Only supports splitting on & outside of quoted string literals.
    """
    parts: List[str] = []
    buf: List[str] = []
    in_quotes = False
    i = 0

    while i < len(expr):
        ch = expr[i]
        if ch == '"':
            if in_quotes:
                # Escaped quote within a string literal is represented as ""
                if i + 1 < len(expr) and expr[i + 1] == '"':
                    buf.append('""')
                    i += 2
                    continue
                in_quotes = False
                buf.append('"')
                i += 1
                continue
            in_quotes = True
            buf.append('"')
            i += 1
            continue

        if ch == "&" and not in_quotes:
            parts.append("".join(buf).strip())
            buf = []
            i += 1
            continue

        buf.append(ch)
        i += 1

    if in_quotes:
        return None

    if buf:
        parts.append("".join(buf).strip())

    return parts


def _parse_excel_string_literal(token: str) -> Optional[str]:
    token = token.strip()
    if len(token) < 2 or not token.startswith('"') or not token.endswith('"'):
        return None
    inner = token[1:-1]
    return inner.replace('""', '"')


def _parse_excel_cell_reference(token: str) -> Optional[Tuple[Optional[str], str]]:
    """
    Parse an Excel cell reference like:
      - AW5
      - $AW$5
      - Sheet1!AW5
      - 'Nov''25'!AW5
    Returns (sheet_name_or_None, coordinate_without_dollars).
    """
    m = re.fullmatch(
        r"\s*(?:(?:'((?:[^']|'')+)'|([A-Za-z0-9_]+))!)?(\$?[A-Za-z]{1,3}\$?\d+)\s*",
        token,
    )
    if not m:
        return None

    quoted_sheet = m.group(1)
    unquoted_sheet = m.group(2)
    sheet_name: Optional[str]
    if quoted_sheet is not None:
        sheet_name = quoted_sheet.replace("''", "'")
    else:
        sheet_name = unquoted_sheet

    coord = m.group(3).replace("$", "")
    return sheet_name, coord


def _read_cell_text_best_effort(ws_cell, ws_data_cell, *, depth: int, max_depth: int) -> Optional[str]:
    displayed = _normalize_excel_text(get_cell_displayed_value(ws_data_cell))
    if displayed:
        return displayed

    raw = ws_cell.value
    if raw is None:
        return None

    if isinstance(raw, str):
        constant = _extract_excel_constant_string_formula(raw)
        if constant:
            constant_text = _normalize_excel_text(constant)
            return constant_text or None

        if depth < max_depth:
            evaluated = _evaluate_excel_concat_formula(
                raw,
                ws_cell.parent,
                ws_data_cell.parent,
                depth=depth + 1,
                max_depth=max_depth,
            )
            if evaluated:
                evaluated_text = _normalize_excel_text(evaluated)
                return evaluated_text or None

        # Only use raw text when it's not a formula.
        if not raw.lstrip().startswith("="):
            raw_text = _normalize_excel_text(raw)
            return raw_text or None

        return None

    raw_text = _normalize_excel_text(str(raw))
    return raw_text or None


def _evaluate_excel_concat_formula(
    formula: str,
    ws,
    ws_data,
    *,
    depth: int,
    max_depth: int,
) -> Optional[str]:
    """
    Best-effort evaluation of a very small subset of Excel formulas:
      - string concatenation using & with string literals and cell references

    This is intended for cases where openpyxl cannot provide cached formula results but
    the formula is effectively just building a text label (e.g. Analytical Review header).
    """
    if not formula:
        return None

    f = formula.strip()
    if not f.startswith("="):
        return None

    expr = f[1:].strip()
    if "&" not in expr:
        return None

    parts = _split_excel_concat_expression(expr)
    if not parts:
        return None

    out: List[str] = []
    for part in parts:
        if not part:
            continue

        literal = _parse_excel_string_literal(part)
        if literal is not None:
            out.append(literal)
            continue

        ref = _parse_excel_cell_reference(part)
        if ref is None:
            return None

        sheet_name, coord = ref
        ws_ref = ws
        ws_data_ref = ws_data

        if sheet_name:
            try:
                ws_ref = ws.parent[sheet_name]
                ws_data_ref = ws_data.parent[sheet_name]
            except Exception:
                return None

        ws_cell = ws_ref[coord]
        ws_data_cell = ws_data_ref[coord]
        text = _read_cell_text_best_effort(ws_cell, ws_data_cell, depth=depth, max_depth=max_depth)
        if text is None:
            return None

        out.append(text)

    return "".join(out)


def _cell_text_candidates(ws_cell, ws_data_cell) -> List[str]:
    """
    Return possible textual representations for a report cell by checking both:
    - the data_only workbook (displayed value, when cached)
    - the formula workbook (raw value / formula text)

    Note: avoid returning raw formula strings (e.g. =... with concatenations) because the
    downstream logic expects final displayed text (for parsing brackets, etc).
    """
    candidates: List[str] = []

    displayed = _normalize_excel_text(get_cell_displayed_value(ws_data_cell))
    if displayed:
        candidates.append(displayed)

    raw = ws_cell.value
    if raw is not None:
        if isinstance(raw, str):
            constant = _extract_excel_constant_string_formula(raw)
            if constant:
                constant_text = _normalize_excel_text(constant)
                if constant_text:
                    candidates.insert(0, constant_text)

            evaluated = _evaluate_excel_concat_formula(raw, ws_cell.parent, ws_data_cell.parent, depth=0, max_depth=2)
            if evaluated:
                evaluated_text = _normalize_excel_text(evaluated)
                if evaluated_text:
                    candidates.insert(0, evaluated_text)

            if not raw.lstrip().startswith("="):
                raw_text = _normalize_excel_text(raw)
                if raw_text:
                    candidates.append(raw_text)
        else:
            raw_text = _normalize_excel_text(str(raw))
            if raw_text:
                candidates.append(raw_text)

    # De-dupe while preserving order
    seen = set()
    unique: List[str] = []
    for text in candidates:
        if text in seen:
            continue
        seen.add(text)
        unique.append(text)
    return unique


def parse_short_date(s: str) -> Tuple[str, int]:
    """Parse short date like Nov'25 -> ('NOV', 2025)."""
    match = re.match(r"([A-Za-z]{3})'?(\d{2})", s.strip())
    if match:
        month_abbr = match.group(1).capitalize()
        year_short = int(match.group(2))
        year_full = 2000 + year_short if year_short < 100 else year_short
        return MONTH_FULL.get(month_abbr, month_abbr.upper()), year_full
    return "", 0


def get_fy_for_month_year(month: str, year: int) -> int:
    """
    Get fiscal year for a given month and calendar year.
    FY runs from May to April. So:
    - May 2025 to Apr 2026 = FY26
    """
    month_num = MONTH_TO_NUM.get(month.capitalize(), 0)
    if month_num >= 5:  # May-Dec: FY is next calendar year
        return year + 1
    else:  # Jan-Apr: FY is current calendar year
        return year


def get_quarter_months(quarter: int, fy: int) -> List[Tuple[str, int]]:
    """
    Get the months for a quarter in a fiscal year.
    FY26: Q1=MAY-JUL 2025, Q2=AUG-OCT 2025, Q3=NOV-DEC 2025+JAN 2026, Q4=FEB-APR 2026
    """
    base_year = fy - 1  # FY26 starts in calendar 2025

    if quarter == 1:
        return [("MAY", base_year), ("JUN", base_year), ("JUL", base_year)]
    elif quarter == 2:
        return [("AUG", base_year), ("SEP", base_year), ("OCT", base_year)]
    elif quarter == 3:
        return [("NOV", base_year), ("DEC", base_year), ("JAN", fy)]
    elif quarter == 4:
        return [("FEB", fy), ("MAR", fy), ("APR", fy)]
    return []


def get_ytd_months_count(selected_month: str, selected_year: int) -> int:
    """
    Get the number of months from start of FY to the selected month.
    FY starts in May.
    """
    month_num = MONTH_TO_NUM.get(selected_month.capitalize(), 0)
    if month_num >= 5:  # May-Dec
        return month_num - 4  # May=1, Jun=2, ..., Dec=8
    else:  # Jan-Apr
        return month_num + 8  # Jan=9, Feb=10, Mar=11, Apr=12


def build_ytd_header(months_count: int, month: str, year: int) -> str:
    """Build YTD header like '07 MONTHS TO NOV 2025'."""
    return f"{months_count:02d} MONTHS TO {MONTH_FULL.get(month.capitalize(), month.upper())} {year}"


def parse_analytical_bracket(bracket_content: str, selected_month: str, selected_year: int) -> Dict:
    """
    Parse the bracket content from Analytical Review header.
    Returns dict with comparison info.
    """
    result = {
        "type": None,
        "left": [],   # List of (month, year) or quarter info
        "right": [],
        "is_ytd": False,
        "left_header": None,
        "right_header": None,
    }

    content = bracket_content.strip()

    # Check for YTD pattern: (YTD FY26 Vs YTD FY25) or (YTD FY26 Vs YTD FY24)
    ytd_match = re.match(r"YTD\s+FY(\d{2})\s+Vs\s+YTD\s+FY(\d{2})", content, re.IGNORECASE)
    if ytd_match:
        result["type"] = "ytd"
        result["is_ytd"] = True
        fy_left = 2000 + int(ytd_match.group(1))
        fy_right = 2000 + int(ytd_match.group(2))

        months_count = get_ytd_months_count(selected_month, selected_year)

        # For left (current), use selected month/year
        result["left_header"] = build_ytd_header(months_count, selected_month, selected_year)

        # For right, calculate the year difference
        year_diff = fy_left - fy_right
        right_year = selected_year - year_diff
        result["right_header"] = build_ytd_header(months_count, selected_month, right_year)

        return result

    # Check for Quarter pattern: (Q2-FY26 Vs Q2-FY25) or (Q2-FY26 Vs Q1-FY26)
    quarter_match = re.match(r"Q(\d)-FY(\d{2})\s+Vs\s+Q(\d)-FY(\d{2})", content, re.IGNORECASE)
    if quarter_match:
        result["type"] = "quarter"
        q_left = int(quarter_match.group(1))
        fy_left = 2000 + int(quarter_match.group(2))
        q_right = int(quarter_match.group(3))
        fy_right = 2000 + int(quarter_match.group(4))

        result["left"] = get_quarter_months(q_left, fy_left)
        result["right"] = get_quarter_months(q_right, fy_right)
        return result

    # Check for combined months pattern: (Sep'25 & Aug'25 Vs Sep'24 & Aug'24)
    combined_match = re.match(
        r"([A-Za-z]{3})'?(\d{2})\s*&\s*([A-Za-z]{3})'?(\d{2})\s+Vs\s+([A-Za-z]{3})'?(\d{2})\s*&\s*([A-Za-z]{3})'?(\d{2})",
        content, re.IGNORECASE
    )
    if combined_match:
        result["type"] = "combined"
        m1, y1 = parse_short_date(f"{combined_match.group(1)}'{combined_match.group(2)}")
        m2, y2 = parse_short_date(f"{combined_match.group(3)}'{combined_match.group(4)}")
        m3, y3 = parse_short_date(f"{combined_match.group(5)}'{combined_match.group(6)}")
        m4, y4 = parse_short_date(f"{combined_match.group(7)}'{combined_match.group(8)}")

        result["left"] = [(m1, y1), (m2, y2)]
        result["right"] = [(m3, y3), (m4, y4)]
        return result

    # Check for simple month comparison: (Nov'25 Vs Oct'25)
    simple_match = re.match(r"([A-Za-z]{3})'?(\d{2})\s+Vs\s+([A-Za-z]{3})'?(\d{2})", content, re.IGNORECASE)
    if simple_match:
        result["type"] = "month"
        m1, y1 = parse_short_date(f"{simple_match.group(1)}'{simple_match.group(2)}")
        m2, y2 = parse_short_date(f"{simple_match.group(3)}'{simple_match.group(4)}")

        result["left"] = [(m1, y1)]
        result["right"] = [(m2, y2)]
        return result

    return result


def build_month_header(month: str, year: int) -> str:
    """Build month header like 'NOV 2025'."""
    return f"{month} {year}"


def load_ma_workbooks(ma_files: List[str], log_emit=None) -> List[Tuple[str, UnifiedWorkbook]]:
    """Load all MA workbooks and return list of (filename, workbook)."""
    workbooks = []
    for fpath in ma_files:
        try:
            wb = load_workbook_unified(fpath, data_only=True)
            workbooks.append((Path(fpath).name, wb))
            _emit(log_emit, f"[OK] Loaded MA file: {Path(fpath).name}")
        except Exception as e:
            _emit(log_emit, f"[ERROR] Failed to load {fpath}: {e}")
    return workbooks


def get_values_for_production_overhead(ma_workbooks: List[Tuple[str, UnifiedWorkbook]],
                                        left_headers: List[str], right_headers: List[str],
                                        search_cols: List[str] = ["G", "H"],
                                        log_emit=None) -> List[Tuple[str, float, float]]:
    """
    Get production overhead values from Page 5 and Page 6.
    Returns list of (description, left_sum, right_sum).
    Only reads each header from ONE source file to avoid double-counting.
    """
    from openpyxl.utils import column_index_from_string

    results = {}  # description -> [left_value, right_value]

    # Track which headers have been processed per sheet to avoid double-counting
    processed_headers = {
        "Page 5": {"left": set(), "right": set()},
        "Page 6": {"left": set(), "right": set()}
    }

    for sheet_name in ["Page 5", "Page 6"]:
        sheet_items_count = 0
        for fname, wb in ma_workbooks:
            if sheet_name not in wb.sheetnames:
                _emit(log_emit, f"      [{sheet_name}] Sheet not found in {fname}")
                continue
            ws = wb[sheet_name]
            _emit(log_emit, f"      [{sheet_name}] Processing from {fname}")

            # Find column for each header
            for is_left, headers in [(True, left_headers), (False, right_headers)]:
                side_key = "left" if is_left else "right"

                for header in headers:
                    # Skip if this header was already processed from another file for this sheet
                    if header in processed_headers[sheet_name][side_key]:
                        continue

                    col_idx = None
                    header_row = None

                    for col_letter in search_cols:
                        c_idx = column_index_from_string(col_letter)
                        for row in range(1, min(50, ws.max_row + 1)):
                            cell_val = ws.cell_displayed_value(row, c_idx)
                            if header_matches_cell(cell_val, header):
                                col_idx = c_idx
                                header_row = row
                                _emit(log_emit, f"      [{sheet_name}] Found header '{header}' at row {row}, col {col_letter}")
                                break
                        if col_idx:
                            break

                    if not col_idx:
                        _emit(log_emit, f"      [{sheet_name}] Header '{header}' not found")
                        continue

                    # Mark this header as processed for this sheet
                    processed_headers[sheet_name][side_key].add(header)

                    # Check if page title (before header row) contains "PRODUCTION OVERHEAD"
                    # This handles Page 6 where A2 = "COST OF PRODUCTION - PRODUCTION OVERHEADS"
                    is_page_level_overhead = False
                    for row in range(1, header_row):
                        col_a_val = ws.cell_displayed_value(row, 1).strip().upper()
                        if "PRODUCTION OVERHEAD" in col_a_val:
                            is_page_level_overhead = True
                            _emit(log_emit, f"      [{sheet_name}] Page-level overhead detected at row {row}: {col_a_val[:50]}")
                            break

                    if is_page_level_overhead:
                        # Page 6 style: entire page is production overhead, start after header row
                        # Skip the currency row (RM) by starting from header_row + 2
                        data_start_row = header_row + 2
                        _emit(log_emit, f"      [{sheet_name}] Reading items from row {data_start_row} (page-level)")
                    else:
                        # Page 5 style: find PRODUCTION OVERHEAD section marker
                        po_start_row = None
                        for row in range(header_row + 1, ws.max_row + 1):
                            col_a_val = ws.cell_displayed_value(row, 1).strip().upper()
                            if "PRODUCTION OVERHEAD" in col_a_val:
                                po_start_row = row
                                _emit(log_emit, f"      [{sheet_name}] Found section marker at row {row}: {col_a_val[:50]}")
                                break

                        if not po_start_row:
                            _emit(log_emit, f"      [{sheet_name}] No PRODUCTION OVERHEAD section found after header row")
                            continue
                        data_start_row = po_start_row + 1
                        _emit(log_emit, f"      [{sheet_name}] Reading items from row {data_start_row} (section)")

                    # Get values under PRODUCTION OVERHEAD
                    items_found = 0
                    for row in range(data_start_row, ws.max_row + 1):
                        col_a_val = ws.cell_displayed_value(row, 1).strip()
                        cell_val = ws.cell_value(row, col_idx)

                        # Stop at next major section (but not at Sub-Total carry-forward rows)
                        col_a_upper = col_a_val.upper()
                        if col_a_val and not col_a_val.startswith(" ") and "TOTAL" in col_a_upper:
                            # Skip "Sub-Total" rows - these are carry-forwards, not section endings
                            if "SUB-TOTAL" not in col_a_upper and "SUB TOTAL" not in col_a_upper and "SUBTOTAL" not in col_a_upper:
                                _emit(log_emit, f"      [{sheet_name}] Stopped at TOTAL row {row}: {col_a_val[:30]}")
                                break

                        if cell_val is not None and cell_val != "" and col_a_val:
                            try:
                                numeric_val = float(cell_val) if not isinstance(cell_val, (int, float)) else cell_val
                                key = col_a_val.strip()
                                if key not in results:
                                    results[key] = [0.0, 0.0]
                                if is_left:
                                    results[key][0] += numeric_val
                                else:
                                    results[key][1] += numeric_val
                                items_found += 1
                            except (ValueError, TypeError):
                                pass

                    sheet_items_count += items_found
                    _emit(log_emit, f"      [{sheet_name}] Found {items_found} items for header '{header}'")

        _emit(log_emit, f"      [{sheet_name}] Total items from this sheet: {sheet_items_count}")

    return [(k, v[0], v[1]) for k, v in results.items() if v[0] != 0 or v[1] != 0]


def get_values_for_page8_section(ma_workbooks: List[Tuple[str, UnifiedWorkbook]],
                                  left_headers: List[str], right_headers: List[str],
                                  section_letter: str,
                                  search_cols: List[str] = ["G", "H"],
                                  log_emit=None) -> List[Tuple[str, float, float]]:
    """
    Get values from Page 8 for a specific section (A or B).
    Returns list of (description, left_sum, right_sum).
    Only reads each header from ONE source file to avoid double-counting.
    """
    from openpyxl.utils import column_index_from_string

    results = {}  # description -> [left_value, right_value]
    next_section = "B" if section_letter == "A" else None

    # Track which headers have been processed to avoid double-counting
    processed_headers = {"left": set(), "right": set()}

    for fname, wb in ma_workbooks:
        if "Page 8" not in wb.sheetnames:
            continue
        ws = wb["Page 8"]

        for is_left, headers in [(True, left_headers), (False, right_headers)]:
            side_key = "left" if is_left else "right"

            for header in headers:
                # Skip if this header was already processed from another file
                if header in processed_headers[side_key]:
                    continue

                col_idx = None
                header_row = None

                for col_letter in search_cols:
                    c_idx = column_index_from_string(col_letter)
                    for row in range(1, min(20, ws.max_row + 1)):
                        cell_val = ws.cell_displayed_value(row, c_idx)
                        if header_matches_cell(cell_val, header):
                            col_idx = c_idx
                            header_row = row
                            break
                    if col_idx:
                        break

                if not col_idx:
                    continue

                # Mark this header as processed
                processed_headers[side_key].add(header)

                # Find section start (row starting with section_letter)
                section_start = None
                section_end = None

                for row in range(header_row + 1, ws.max_row + 1):
                    col_a_val = ws.cell_displayed_value(row, 1).strip()
                    if col_a_val.upper().startswith(section_letter):
                        section_start = row
                    elif section_start and next_section and col_a_val.upper().startswith(next_section):
                        section_end = row
                        break

                if not section_start:
                    continue

                end_row = section_end if section_end else ws.max_row + 1

                for row in range(section_start + 1, end_row):
                    col_b_val = ws.cell_displayed_value(row, 2).strip()
                    cell_val = ws.cell_value(row, col_idx)

                    if cell_val is not None and cell_val != "" and col_b_val and not col_b_val.upper().startswith("TOTAL"):
                        try:
                            numeric_val = float(cell_val) if not isinstance(cell_val, (int, float)) else cell_val
                            key = col_b_val.strip()
                            if key not in results:
                                results[key] = [0.0, 0.0]
                            if is_left:
                                results[key][0] += numeric_val
                            else:
                                results[key][1] += numeric_val
                        except (ValueError, TypeError):
                            pass

    return [(k, v[0], v[1]) for k, v in results.items() if v[0] != 0 or v[1] != 0]


def get_values_for_page7(ma_workbooks: List[Tuple[str, UnifiedWorkbook]],
                          left_headers: List[str], right_headers: List[str],
                          search_cols: List[str] = ["G", "H"],
                          log_emit=None) -> List[Tuple[str, float, float]]:
    """
    Get values from Page 7 for General Administrative Expenses.
    Returns list of (description, left_sum, right_sum).
    Only reads each header from ONE source file to avoid double-counting.
    """
    from openpyxl.utils import column_index_from_string

    results = {}

    # Track which headers have been processed to avoid double-counting
    processed_headers = {"left": set(), "right": set()}

    items_count = 0
    for fname, wb in ma_workbooks:
        if "Page 7" not in wb.sheetnames:
            _emit(log_emit, f"      [Page 7] Sheet not found in {fname}")
            continue
        ws = wb["Page 7"]
        _emit(log_emit, f"      [Page 7] Processing from {fname}, max_row={ws.max_row}")

        for is_left, headers in [(True, left_headers), (False, right_headers)]:
            side_key = "left" if is_left else "right"

            for header in headers:
                # Skip if this header was already processed from another file
                if header in processed_headers[side_key]:
                    continue

                col_idx = None
                header_row = None

                for col_letter in search_cols:
                    c_idx = column_index_from_string(col_letter)
                    for row in range(1, min(20, ws.max_row + 1)):
                        cell_val = ws.cell_displayed_value(row, c_idx)
                        if header_matches_cell(cell_val, header):
                            col_idx = c_idx
                            header_row = row
                            _emit(log_emit, f"      [Page 7] Found header '{header}' at row {row}, col {col_letter}")
                            break
                    if col_idx:
                        break

                if not col_idx:
                    _emit(log_emit, f"      [Page 7] Header '{header}' not found")
                    continue

                # Mark this header as processed
                processed_headers[side_key].add(header)

                header_items = 0
                for row in range(header_row + 1, ws.max_row + 1):
                    col_b_val = ws.cell_displayed_value(row, 2).strip()
                    cell_val = ws.cell_value(row, col_idx)

                    if cell_val is not None and cell_val != "" and col_b_val:
                        # Skip excluded items
                        skip = False
                        for excl in GAE_EXCLUDE:
                            if excl.upper() in col_b_val.upper():
                                skip = True
                                break
                        if skip:
                            continue

                        try:
                            numeric_val = float(cell_val) if not isinstance(cell_val, (int, float)) else cell_val
                            key = col_b_val.strip()
                            if key not in results:
                                results[key] = [0.0, 0.0]
                            if is_left:
                                results[key][0] += numeric_val
                            else:
                                results[key][1] += numeric_val
                            header_items += 1
                        except (ValueError, TypeError):
                            pass

                _emit(log_emit, f"      [Page 7] Found {header_items} items for header '{header}'")
                items_count += header_items

    _emit(log_emit, f"      [Page 7] Total unique items collected: {len(results)}")

    # Log top 5 by absolute difference
    sorted_results = sorted(results.items(), key=lambda x: abs(x[1][0] - x[1][1]), reverse=True)[:5]
    for desc, (left, right) in sorted_results:
        _emit(log_emit, f"      [Page 7] Top item: {desc[:30]} = {left:.2f} - {right:.2f} = {left-right:.2f}")

    return [(k, v[0], v[1]) for k, v in results.items() if v[0] != 0 or v[1] != 0]


def get_values_for_page9(ma_workbooks: List[Tuple[str, UnifiedWorkbook]],
                          left_headers: List[str], right_headers: List[str],
                          search_cols: List[str] = ["G", "H"],
                          log_emit=None) -> List[Tuple[str, float, float]]:
    """
    Get values from Page 9 for Other Income (section B).
    Returns list of (description, left_sum, right_sum).
    Only reads each header from ONE source file to avoid double-counting.
    """
    from openpyxl.utils import column_index_from_string

    results = {}

    # Track which headers have been processed to avoid double-counting
    processed_headers = {"left": set(), "right": set()}

    for fname, wb in ma_workbooks:
        if "Page 9" not in wb.sheetnames:
            continue
        ws = wb["Page 9"]

        for is_left, headers in [(True, left_headers), (False, right_headers)]:
            side_key = "left" if is_left else "right"

            for header in headers:
                # Skip if this header was already processed from another file
                if header in processed_headers[side_key]:
                    continue

                col_idx = None
                header_row = None

                for col_letter in search_cols:
                    c_idx = column_index_from_string(col_letter)
                    for row in range(1, min(20, ws.max_row + 1)):
                        cell_val = ws.cell_displayed_value(row, c_idx)
                        if header_matches_cell(cell_val, header):
                            col_idx = c_idx
                            header_row = row
                            break
                    if col_idx:
                        break

                if not col_idx:
                    continue

                # Mark this header as processed
                processed_headers[side_key].add(header)

                # Find section B start
                section_start = None
                for row in range(header_row + 1, ws.max_row + 1):
                    col_a_val = ws.cell_displayed_value(row, 1).strip()
                    if col_a_val.upper().startswith("B"):
                        section_start = row
                        break

                if not section_start:
                    continue

                skip_rows_remaining = 0
                for row in range(section_start + 1, ws.max_row + 1):
                    col_a_val = ws.cell_displayed_value(row, 1).strip()
                    col_b_val = ws.cell_displayed_value(row, 2).strip()
                    cell_val = ws.cell_value(row, col_idx)

                    # Stop if we hit next section
                    if col_a_val and col_a_val.upper().startswith("C"):
                        break

                    # Ignore any bullet/comment rows (and the row directly below them).
                    if skip_rows_remaining > 0:
                        skip_rows_remaining -= 1
                        continue
                    if col_b_val.startswith("*"):
                        skip_rows_remaining = 1
                        continue

                    if cell_val is not None and cell_val != "" and col_b_val and not col_b_val.upper().startswith("TOTAL"):
                        # Skip items starting with "GAIN/(LOSS)"
                        if col_b_val.upper().startswith("GAIN/(LOSS) ON FOREX"):
                            continue

                        try:
                            numeric_val = float(cell_val) if not isinstance(cell_val, (int, float)) else cell_val
                            key = col_b_val.strip()
                            if key not in results:
                                results[key] = [0.0, 0.0]
                            if is_left:
                                results[key][0] += numeric_val
                            else:
                                results[key][1] += numeric_val
                        except (ValueError, TypeError):
                            pass

    return [(k, v[0], v[1]) for k, v in results.items() if v[0] != 0 or v[1] != 0]


def find_extreme_value(values: List[Tuple[str, float, float]], find_lowest: bool) -> Optional[str]:
    """
    Find the description with the lowest/highest difference (left - right).
    """
    if not values:
        return None

    diffs = [(desc, left - right) for desc, left, right in values]

    if find_lowest:
        sorted_diffs = sorted(diffs, key=lambda x: x[1])
    else:
        sorted_diffs = sorted(diffs, key=lambda x: x[1], reverse=True)

    for desc, diff in sorted_diffs:
        if desc.strip():
            return desc

    return None


def _round_half_up(value: float) -> int:
    if math.isnan(value) or math.isinf(value):
        return 0
    return int(math.floor(value + 0.5))


def format_rm_variance_amount(amount: float) -> str:
    """
    Format an RM variance amount using:
    - Thousands and above: rounded to nearest thousand, shown as '<n>k' (e.g., 928352 -> 928k)
    - Hundreds and below: shown as integer (e.g., 123)
    """
    amt = abs(float(amount))
    if amt < 1000:
        return str(_round_half_up(amt))
    return f"{_round_half_up(amt / 1000)}k"


def build_page9b_variance_summary(values: List[Tuple[str, float, float]]) -> Optional[str]:
    """
    Build a single-cell bullet summary for Page 9 Section B (Other Income),
    comparing x - y (left - right) for each line item.
    """
    phrases: List[str] = []

    for desc, left, right in values:
        desc_clean = str(desc).strip()
        if not desc_clean:
            continue
        if desc_clean.upper().startswith("GAIN/(LOSS) ON FOREX"):
            continue

        diff = float(left) - float(right)
        if abs(diff) < 1e-9:
            continue

        direction = "Higher" if diff > 0 else "Lower"
        amount_str = format_rm_variance_amount(diff)
        phrases.append(f"{direction} {desc_clean} by RM{amount_str}")

    if not phrases:
        return None

    return "*" + ", ".join(phrases)


def _find_star_cell_row_in_column_b(
    ws,
    ws_data,
    start_row: int,
    end_row_exclusive: int,
    *,
    anchor_row: Optional[int] = None,
    max_distance: int = 25,
) -> Optional[int]:
    start_row = max(1, start_row)
    end_row_exclusive = max(1, end_row_exclusive)

    def _cell_startswith_star(row: int) -> bool:
        ws_cell = ws.cell(row=row, column=2)
        ws_data_cell = ws_data.cell(row=row, column=2)
        for candidate in _cell_text_candidates(ws_cell, ws_data_cell):
            if not candidate:
                continue
            text = candidate.lstrip()
            if text.startswith("*"):
                return True
        return False

    if anchor_row is None:
        for row in range(start_row, end_row_exclusive):
            if _cell_startswith_star(row):
                return row
        return None

    anchor_row = max(start_row, min(anchor_row, end_row_exclusive - 1))
    window_start = max(start_row, anchor_row - max_distance)
    window_end_exclusive = min(end_row_exclusive, anchor_row + max_distance + 1)

    star_rows: List[int] = []
    for row in range(window_start, window_end_exclusive):
        if _cell_startswith_star(row):
            star_rows.append(row)

    if not star_rows:
        return None
    return min(star_rows, key=lambda r: abs(r - anchor_row))


def _write_page9b_summary_to_star_cell(
    ws,
    *,
    star_row: int,
    summary: Optional[str],
    values_overrides: Dict[Tuple[int, int], Any],
    log_emit=None,
) -> bool:
    """
    Write the provided summary into the given '*' placeholder cell (column B), and clear the row below.
    """
    ws.cell(row=star_row, column=2).value = None
    values_overrides[(star_row, 2)] = None

    if summary:
        ws.cell(row=star_row, column=2).value = summary
        values_overrides[(star_row, 2)] = summary

    if star_row + 1 <= ws.max_row:
        ws.cell(row=star_row + 1, column=2).value = None
        values_overrides[(star_row + 1, 2)] = None

    _emit(log_emit, f"    -> Updated Page 9B summary at row {star_row}")
    return True


def build_headers_from_comparison(comparison: Dict) -> Tuple[List[str], List[str], List[str]]:
    """
    Build header strings from comparison info.
    Returns (left_headers, right_headers, search_cols).
    """
    search_cols = ["G", "H"]

    if comparison["type"] == "ytd":
        search_cols = ["I", "J"]
        return [comparison["left_header"]], [comparison["right_header"]], search_cols

    left_headers = []
    right_headers = []

    for month, year in comparison.get("left", []):
        left_headers.append(build_month_header(month, year))

    for month, year in comparison.get("right", []):
        right_headers.append(build_month_header(month, year))

    return left_headers, right_headers, search_cols


def _ma_sheet_contains_header(
    ws: UnifiedWorksheet,
    header: str,
    search_cols: List[str],
    max_search_rows: int,
) -> bool:
    from openpyxl.utils import column_index_from_string

    max_row = min(max_search_rows, ws.max_row or 0)
    if max_row <= 0:
        return False

    for col_letter in search_cols:
        col_idx = column_index_from_string(col_letter)
        for row in range(1, max_row + 1):
            cell_val = ws.cell_displayed_value(row, col_idx)
            if header_matches_cell(cell_val, header):
                return True
    return False


def _collect_required_ma_header_requirements(
    ws_data,
    analytical_reviews: List[Tuple[int, int, str]],
    selected_month: str,
    selected_year: int,
    log_emit=None,
) -> Dict[Tuple[str, Tuple[str, ...], int], set[str]]:
    requirements: Dict[Tuple[str, Tuple[str, ...], int], set[str]] = {}

    for i, (ar_row, _ar_col, ar_text) in enumerate(analytical_reviews):
        bracket_match = re.search(r"\(([^)]+)\)", ar_text)
        if not bracket_match:
            continue

        bracket_content = bracket_match.group(1)
        comparison = parse_analytical_bracket(bracket_content, selected_month, selected_year)
        if not comparison["type"]:
            continue

        left_headers, right_headers, search_cols = build_headers_from_comparison(comparison)
        headers = [*left_headers, *right_headers]
        if not headers:
            continue

        end_row = analytical_reviews[i + 1][0] if i + 1 < len(analytical_reviews) else ws_data.max_row + 1

        required_sheets: set[str] = set()
        for row in range(ar_row + 1, end_row):
            for col in [1, 2]:  # Column A and B
                cell_val = get_cell_displayed_value(ws_data.cell(row=row, column=col)).strip()
                if "Description" not in cell_val:
                    continue

                displayed_upper = cell_val.upper()
                if "PRODUCTION OVERHEAD" in displayed_upper:
                    required_sheets.update({"Page 5", "Page 6"})
                elif "PURCHASE RELATED COSTS" in displayed_upper or "SALES RELATED COSTS" in displayed_upper:
                    required_sheets.add("Page 8")
                elif "GENERAL ADMINISTRATIVE EXPENSES" in displayed_upper:
                    required_sheets.add("Page 7")
                elif "OTHER INCOME" in displayed_upper:
                    required_sheets.add("Page 9")

        for sheet_name in required_sheets:
            max_search_rows = 50 if sheet_name in {"Page 5", "Page 6"} else 20
            key = (sheet_name, tuple(search_cols), max_search_rows)
            requirements.setdefault(key, set()).update(headers)

    return requirements


def _collect_strict_ma_header_requirements(
    analytical_reviews: List[Tuple[int, int, str]],
    selected_month: str,
    selected_year: int,
    log_emit=None,
) -> Dict[Tuple[str, Tuple[str, ...], int], set[str]]:
    """
    Strict mode: require that every Analytical Review section's left/right headers exist
    in the MA workbook (for the relevant search columns) across the standard MA pages.

    This intentionally does not depend on whether any 'Description' placeholders remain
    in the report (e.g. already-processed reports still validate inputs).
    """
    requirements: Dict[Tuple[str, Tuple[str, ...], int], set[str]] = {}

    page_max_rows: Dict[str, int] = {
        "Page 5": 50,
        "Page 6": 50,
        "Page 7": 20,
        "Page 8": 20,
        "Page 9": 20,
    }

    for _ar_row, _ar_col, ar_text in analytical_reviews:
        bracket_match = re.search(r"\(([^)]+)\)", ar_text)
        if not bracket_match:
            continue

        bracket_content = bracket_match.group(1)
        comparison = parse_analytical_bracket(bracket_content, selected_month, selected_year)
        if not comparison["type"]:
            continue

        left_headers, right_headers, search_cols = build_headers_from_comparison(comparison)
        headers = [*left_headers, *right_headers]
        if not headers:
            continue

        for sheet_name, max_search_rows in page_max_rows.items():
            key = (sheet_name, tuple(search_cols), max_search_rows)
            requirements.setdefault(key, set()).update(headers)

    return requirements


def _validate_required_ma_headers(
    ma_workbooks: List[Tuple[str, UnifiedWorkbook]],
    requirements: Dict[Tuple[str, Tuple[str, ...], int], set[str]],
    log_emit=None,
) -> None:
    if not requirements:
        return

    missing_by_sheet: Dict[str, set[str]] = {}

    for (sheet_name, search_cols, max_search_rows), headers in requirements.items():
        for header in sorted(headers):
            header_found = False
            for fname, wb in ma_workbooks:
                if sheet_name not in wb.sheetnames:
                    continue

                ws = wb[sheet_name]
                if _ma_sheet_contains_header(ws, header, list(search_cols), max_search_rows):
                    header_found = True
                    break

            if not header_found:
                missing_by_sheet.setdefault(sheet_name, set()).add(header)
                _emit(log_emit, f"[MISSING HEADER] {sheet_name}: {header}")

    if missing_by_sheet:
        missing_sorted = {k: sorted(v) for k, v in missing_by_sheet.items()}
        raise MissingRequiredHeadersError(missing_sorted)


def _export_values_only_workbook_for_sheet(
    ws_data,
    values_overrides: Dict[Tuple[int, int], Any],
    sheet_name: str,
    output_file: Path,
    log_emit=None,
) -> str:
    from copy import copy as copy_style

    max_row = ws_data.max_row or 0
    max_col = ws_data.max_column or 0

    values_wb = Workbook()
    ws_out = values_wb.active
    ws_out.title = sheet_name

    ws_out.freeze_panes = ws_data.freeze_panes
    ws_out.sheet_view.showGridLines = ws_data.sheet_view.showGridLines

    for col_key, dim in ws_data.column_dimensions.items():
        out_dim = ws_out.column_dimensions[col_key]
        out_dim.width = dim.width
        out_dim.hidden = dim.hidden
        out_dim.outlineLevel = dim.outlineLevel
        out_dim.collapsed = dim.collapsed

    for row_key, dim in ws_data.row_dimensions.items():
        out_dim = ws_out.row_dimensions[row_key]
        out_dim.height = dim.height
        out_dim.hidden = dim.hidden
        out_dim.outlineLevel = dim.outlineLevel
        out_dim.collapsed = dim.collapsed

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            src_cell = ws_data.cell(row=row, column=col)
            dst_cell = ws_out.cell(row=row, column=col)

            dst_cell.value = values_overrides.get((row, col), src_cell.value)

            if src_cell.has_style:
                dst_cell.font = copy_style(src_cell.font)
                dst_cell.fill = copy_style(src_cell.fill)
                dst_cell.border = copy_style(src_cell.border)
                dst_cell.alignment = copy_style(src_cell.alignment)
                dst_cell.number_format = src_cell.number_format
                dst_cell.protection = copy_style(src_cell.protection)
                if src_cell.comment:
                    dst_cell.comment = copy_style(src_cell.comment)

            if src_cell.hyperlink:
                dst_cell.hyperlink = copy_style(src_cell.hyperlink)

    for merged_range in ws_data.merged_cells.ranges:
        ws_out.merge_cells(str(merged_range))

    values_wb.save(output_file)
    _emit(log_emit, f"Values-only sheet saved to: {output_file}")
    return str(output_file)


def process_variance_report(
    report_path: str,
    ma_files: List[str],
    selected_month: str,
    selected_year: int,
    log_emit=None,
) -> Tuple[str, str]:
    """
    Main processing function.
    """
    # Build sheet name like "Nov'25"
    year_short = selected_year % 100
    sheet_name = f"{selected_month}'{year_short:02d}"

    _emit(log_emit, f"Looking for worksheet: {sheet_name}")

    # Load the report workbook
    report_wb = load_workbook(report_path, data_only=False)  # Keep formulas for modification
    report_wb_data = load_workbook(report_path, data_only=True)  # For reading displayed values

    if sheet_name not in report_wb.sheetnames:
        # Try variations
        found = False
        for sn in report_wb.sheetnames:
            if sn.upper().replace("'", "").replace(" ", "") == sheet_name.upper().replace("'", "").replace(" ", ""):
                sheet_name = sn
                found = True
                break
        if not found:
            raise ValueError(f"Worksheet '{sheet_name}' not found. Available: {report_wb.sheetnames}")

    _emit(log_emit, f"Found worksheet: {sheet_name}")

    ws = report_wb[sheet_name]
    ws_data = report_wb_data[sheet_name]

    # Load MA workbooks
    ma_workbooks = load_ma_workbooks(ma_files, log_emit)
    if not ma_workbooks:
        raise ValueError("No MA files could be loaded.")

    # Find all Analytical Review cells
    analytical_reviews: List[Tuple[int, int, str]] = []
    for row in range(1, ws.max_row + 1):
        for col in (1, 2):  # Column A and B
            ws_cell = ws.cell(row=row, column=col)
            ws_data_cell = ws_data.cell(row=row, column=col)

            matched_text: Optional[str] = None
            for candidate in _cell_text_candidates(ws_cell, ws_data_cell):
                if candidate.upper().startswith("ANALYTICAL REVIEW"):
                    matched_text = candidate
                    break

            if matched_text:
                analytical_reviews.append((row, col, matched_text))
                hidden_tag = " (hidden)" if _is_excel_row_hidden(ws, row) else ""
                _emit(log_emit, f"Found Analytical Review at row {row}{hidden_tag}: {matched_text[:50]}...")

    if not analytical_reviews:
        try:
            import openpyxl as _openpyxl  # type: ignore

            openpyxl_version = getattr(_openpyxl, "__version__", "unknown")
        except Exception:
            openpyxl_version = "unknown"

        sample_lines: List[str] = []
        max_sample_rows = min(ws.max_row or 0, 300)
        for row in range(1, max_sample_rows + 1):
            for col in (1, 2):
                ws_cell = ws.cell(row=row, column=col)
                ws_data_cell = ws_data.cell(row=row, column=col)
                displayed = _normalize_excel_text(get_cell_displayed_value(ws_data_cell))
                raw = ws_cell.value
                raw_text = _normalize_excel_text(str(raw)) if raw is not None else ""

                if displayed or raw_text:
                    sample_lines.append(f"  R{row}C{col}: data_only={displayed!r} raw={raw_text!r}")

                if len(sample_lines) >= 10:
                    break
            if len(sample_lines) >= 10:
                break

        message_lines = [
            "No 'Analytical Review' cells found in column A or B.",
            f"(openpyxl {openpyxl_version})",
            "",
            "This usually happens when the 'Analytical Review' headers are formulas but the workbook has no cached formula results.",
            "openpyxl cannot calculate formulas, so it may see empty/placeholder values.",
            "Fix: open the report in Excel, set Calculation to Automatic, press Calculate Now, save the file, then rerun.",
        ]

        if sample_lines:
            message_lines.extend(["", "Sample values read from columns A/B:"])
            message_lines.extend(sample_lines)

        raise ValueError("\n".join(message_lines))

    _emit(log_emit, f"\nFound {len(analytical_reviews)} Analytical Review sections")

    # Validate required MA headers before modifying anything in the report workbook.
    hidden_analytical_rows = {row for row, _col, _text in analytical_reviews if _is_excel_row_hidden(ws, row)}
    analytical_reviews_for_validation: List[Tuple[int, int, str]] = []
    for i, (ar_row, ar_col, ar_text) in enumerate(analytical_reviews):
        if ar_row in hidden_analytical_rows:
            continue

        end_row = analytical_reviews[i + 1][0] if i + 1 < len(analytical_reviews) else ws.max_row + 1
        if not _analytical_review_group_requires_actions(ws, ws_data, ar_row + 1, end_row):
            continue

        analytical_reviews_for_validation.append((ar_row, ar_col, ar_text))
    requirements = _collect_strict_ma_header_requirements(
        analytical_reviews_for_validation, selected_month, selected_year, log_emit
    )
    _validate_required_ma_headers(ma_workbooks, requirements, log_emit)

    # Process each Analytical Review section
    replacements_made = 0
    values_overrides: Dict[Tuple[int, int], Any] = {}

    for i, (ar_row, ar_col, ar_text) in enumerate(analytical_reviews):
        if ar_row in hidden_analytical_rows:
            _emit(
                log_emit,
                f"\n[SKIP] Analytical Review row {ar_row} is hidden; leaving this group untouched.",
            )
            continue

        # Extract bracket content
        bracket_match = re.search(r"\(([^)]+)\)", ar_text)
        if not bracket_match:
            _emit(log_emit, f"[SKIP] No bracket content found in: {ar_text[:50]}")
            continue

        bracket_content = bracket_match.group(1)
        _emit(log_emit, f"\nProcessing: ({bracket_content})")

        # Parse the comparison
        comparison = parse_analytical_bracket(bracket_content, selected_month, selected_year)
        if not comparison["type"]:
            _emit(log_emit, f"[SKIP] Could not parse comparison: {bracket_content}")
            continue

        # Build headers
        left_headers, right_headers, search_cols = build_headers_from_comparison(comparison)
        _emit(log_emit, f"  Left headers: {left_headers}")
        _emit(log_emit, f"  Right headers: {right_headers}")
        _emit(log_emit, f"  Search columns: {search_cols}")

        # Determine the range to search (until next Analytical Review or end)
        end_row = analytical_reviews[i + 1][0] if i + 1 < len(analytical_reviews) else ws.max_row + 1
        page9_values_cache: Optional[List[Tuple[str, float, float]]] = None

        # Page 9B '*' summary updates are driven by any '*' bullet cell in column B.
        star_row_in_section = _find_star_cell_row_in_column_b(ws, ws_data, ar_row + 1, end_row)
        if star_row_in_section:
            if page9_values_cache is None:
                page9_values_cache = get_values_for_page9(
                    ma_workbooks, left_headers, right_headers, search_cols, log_emit
                )

            summary = build_page9b_variance_summary(page9_values_cache)
            _write_page9b_summary_to_star_cell(
                ws=ws,
                star_row=star_row_in_section,
                summary=summary,
                values_overrides=values_overrides,
                log_emit=log_emit,
            )
            if not summary:
                _emit(log_emit, "    -> No Page 9B variance summary generated; cleared '*' cell")

        # Find Description cells in this section
        for row in range(ar_row + 1, end_row):
            for col in [1, 2]:  # Column A and B
                cell = ws.cell(row=row, column=col)
                cell_data = ws_data.cell(row=row, column=col)
                displayed_val_raw = get_cell_displayed_value(cell_data)
                displayed_val = displayed_val_raw.strip()
                raw_val = get_cell_raw_value(cell)

                if "Description" not in displayed_val:
                    continue

                _emit(log_emit, f"\n  Found Description cell at row {row}: {displayed_val[:60]}...")

                # Determine what to look for based on cell content
                displayed_upper = displayed_val.upper()
                find_lowest = "LOWER" in displayed_upper
                find_highest = "HIGHER" in displayed_upper

                replacement_value = None

                if "PRODUCTION OVERHEAD" in displayed_upper:
                    _emit(log_emit, "    -> Processing Production Overhead (Page 5 & 6)")
                    values = get_values_for_production_overhead(
                        ma_workbooks, left_headers, right_headers, search_cols, log_emit
                    )
                    if values:
                        replacement_value = find_extreme_value(values, find_lowest)
                        _emit(log_emit, f"    -> Found {len(values)} items, selected: {replacement_value}")

                elif "PURCHASE RELATED COSTS" in displayed_upper:
                    _emit(log_emit, "    -> Processing Purchase Related Costs (Page 8, Section A)")
                    values = get_values_for_page8_section(
                        ma_workbooks, left_headers, right_headers, "A", search_cols, log_emit
                    )
                    if values:
                        replacement_value = find_extreme_value(values, find_lowest)
                        _emit(log_emit, f"    -> Found {len(values)} items, selected: {replacement_value}")

                elif "SALES RELATED COSTS" in displayed_upper:
                    _emit(log_emit, "    -> Processing Sales Related Costs (Page 8, Section B)")
                    values = get_values_for_page8_section(
                        ma_workbooks, left_headers, right_headers, "B", search_cols, log_emit
                    )
                    if values:
                        replacement_value = find_extreme_value(values, find_lowest)
                        _emit(log_emit, f"    -> Found {len(values)} items, selected: {replacement_value}")

                elif "GENERAL ADMINISTRATIVE EXPENSES" in displayed_upper:
                    _emit(log_emit, "    -> Processing General Administrative Expenses (Page 7)")
                    values = get_values_for_page7(
                        ma_workbooks, left_headers, right_headers, search_cols, log_emit
                    )
                    if values:
                        replacement_value = find_extreme_value(values, find_lowest)
                        if (
                            replacement_value
                            and replacement_value.strip().casefold() in PAGE7_PAYROLL_ITEMS_NORM
                        ):
                            replacement_value = "Payroll & Related Costs"
                        _emit(log_emit, f"    -> Found {len(values)} items, selected: {replacement_value}")

                elif "OTHER INCOME" in displayed_upper:
                    _emit(log_emit, "    -> Processing Other Income (Page 9, Section B)")
                    if page9_values_cache is None:
                        page9_values_cache = get_values_for_page9(
                            ma_workbooks, left_headers, right_headers, search_cols, log_emit
                        )
                    values = page9_values_cache

                    if values:
                        replacement_value = find_extreme_value(values, find_lowest)
                        _emit(log_emit, f"    -> Found {len(values)} items, selected: {replacement_value}")

                # Replace Description in the formula/value
                if replacement_value:
                    if isinstance(raw_val, str):
                        new_val = raw_val.replace("Description", replacement_value)
                        cell.value = new_val
                        values_overrides[(row, col)] = displayed_val_raw.replace("Description", replacement_value)
                        replacements_made += 1
                        _emit(log_emit, f"    -> Replaced 'Description' with '{replacement_value}'")
                    else:
                        _emit(log_emit, f"    -> Cell is not a string/formula, cannot replace")
                else:
                    _emit(log_emit, f"    -> No replacement value found")

    # Save the modified workbook
    output_path = Path(report_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_path.parent / f"{output_path.stem}_modified_{timestamp}{output_path.suffix}"

    report_wb.save(output_file)

    safe_sheet = re.sub(r"[^A-Za-z0-9]+", "_", sheet_name).strip("_") or "sheet"
    values_only_file = output_path.parent / f"{output_path.stem}_{safe_sheet}_values_{timestamp}.xlsx"
    values_only_out = _export_values_only_workbook_for_sheet(
        ws_data=ws_data,
        values_overrides=values_overrides,
        sheet_name=sheet_name,
        output_file=values_only_file,
        log_emit=log_emit,
    )

    _emit(log_emit, f"\n{'='*50}")
    _emit(log_emit, f"Processing complete!")
    _emit(log_emit, f"Total replacements made: {replacements_made}")
    _emit(log_emit, f"Output saved to: {output_file}")

    return str(output_file), values_only_out


# --- UI Class (Pycro Station) ---

class MainWidget(QWidget):
    log_message = Signal(str)
    processing_done = Signal(str, str)

    def __init__(self):
        super().__init__()
        self.setObjectName("varanalysis_desc_widget")
        self._last_error: Optional[Exception] = None
        self._build_ui()
        self._connect_signals()

    def _build_ui(self):
        self.desc_label = QLabel("", self)
        self.desc_label.setWordWrap(True)
        self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.desc_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.desc_label.setStyleSheet(
            "color: #dcdcdc; background: transparent; padding: 6px; "
            "border: 1px solid #3a3a3a; border-radius: 6px;"
        )
        self.set_long_description("")

        # File input 1: Variance Analysis Report
        self.select_report_btn = PrimaryPushButton("Select Variance Analysis Report Excel File", self)

        # Dropdowns row
        self.month_label = QLabel("Month:", self)
        self.month_label.setStyleSheet("color: #dcdcdc; background: transparent;")

        self.month_combo = ComboBox(self)
        self.month_combo.addItems(MONTHS)
        self.month_combo.setCurrentIndex(0)  # Default to Nov

        self.year_label = QLabel("Year:", self)
        self.year_label.setStyleSheet("color: #dcdcdc; background: transparent;")

        self.year_combo = ComboBox(self)
        current_year = datetime.now().year
        years = [str(y) for y in range(1975, current_year + 6)]
        self.year_combo.addItems(years)
        self.year_combo.setCurrentText(str(current_year))

        # File input 2: MA files (multi-select)
        self.select_ma_btn = PrimaryPushButton("Select Management Account Files", self)

        # Run button
        self.run_btn = PrimaryPushButton("Process", self)

        # Report file display
        self.report_label = QLabel("Variance Analysis Report", self)
        self.report_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.report_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        # MA files display
        self.ma_label = QLabel("Management Account Files", self)
        self.ma_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.ma_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        self.logs_label = QLabel("Process logs", self)
        self.logs_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.logs_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        shared_style = (
            "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        self.report_box = QTextEdit(self)
        self.report_box.setReadOnly(True)
        self.report_box.setPlaceholderText("Selected report file will appear here")
        self.report_box.setStyleSheet(shared_style)
        self.report_box.setMaximumHeight(60)

        self.ma_box = QTextEdit(self)
        self.ma_box.setReadOnly(True)
        self.ma_box.setPlaceholderText("Selected MA files will appear here")
        self.ma_box.setStyleSheet(shared_style)

        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText("Live process log will appear here")
        self.log_box.setStyleSheet(shared_style)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)

        main_layout.addWidget(self.desc_label, 0)

        # Report file button
        row1 = QHBoxLayout()
        row1.addStretch(1)
        row1.addWidget(self.select_report_btn, 2)
        row1.addStretch(1)
        main_layout.addLayout(row1, 0)

        # Dropdowns row
        row2 = QHBoxLayout()
        row2.addStretch(1)
        row2.addWidget(self.month_label)
        row2.addWidget(self.month_combo)
        row2.addSpacing(20)
        row2.addWidget(self.year_label)
        row2.addWidget(self.year_combo)
        row2.addStretch(1)
        main_layout.addLayout(row2, 0)

        # MA files button
        row3 = QHBoxLayout()
        row3.addStretch(1)
        row3.addWidget(self.select_ma_btn, 2)
        row3.addStretch(1)
        main_layout.addLayout(row3, 0)

        # Run button
        row4 = QHBoxLayout()
        row4.addStretch(1)
        row4.addWidget(self.run_btn, 1)
        row4.addStretch(1)
        main_layout.addLayout(row4, 0)

        # File displays labels
        row5 = QHBoxLayout()
        row5.addWidget(self.report_label, 1)
        row5.addWidget(self.ma_label, 1)
        main_layout.addLayout(row5, 0)

        # File displays
        row6 = QHBoxLayout()
        row6.addWidget(self.report_box, 1)
        row6.addWidget(self.ma_box, 1)
        main_layout.addLayout(row6, 1)

        # Log label and box
        main_layout.addWidget(self.logs_label, 0)
        main_layout.addWidget(self.log_box, 3)

    def set_long_description(self, text: str):
        clean = (text or "").strip()
        if clean:
            self.desc_label.setText(clean)
            self.desc_label.show()
        else:
            self.desc_label.clear()
            self.desc_label.hide()

    def _connect_signals(self):
        self.select_report_btn.clicked.connect(self.select_report_file)
        self.select_ma_btn.clicked.connect(self.select_ma_files)
        self.run_btn.clicked.connect(self.run_process)
        self.log_message.connect(self.append_log)
        self.processing_done.connect(self.on_done)

    def select_report_file(self):
        file, _ = QFileDialog.getOpenFileName(
            self,
            "Select Variance Analysis Report Excel File",
            "",
            "Excel Files (*.xlsx *.xlsm *.xls)",
        )
        if file:
            self.report_box.setPlainText(file)
        else:
            self.report_box.clear()

    def select_ma_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select Management Account Files",
            "",
            "Excel Files (*.xlsx *.xlsm *.xls)",
        )
        if files:
            self.ma_box.setPlainText("\n".join(files))
        else:
            self.ma_box.clear()

    def _get_report_file(self) -> str:
        return self.report_box.toPlainText().strip()

    def _get_ma_files(self) -> List[str]:
        text = self.ma_box.toPlainText().strip()
        if not text:
            return []
        return [line.strip() for line in text.split("\n") if line.strip()]

    def run_process(self):
        report_file = self._get_report_file()
        ma_files = self._get_ma_files()
        self._last_error = None

        if not report_file:
            MessageBox("Warning", "Please select a Variance Analysis Report file.", self).exec()
            return

        if not ma_files:
            MessageBox("Warning", "Please select Management Account files.", self).exec()
            return

        selected_month = self.month_combo.currentText()
        selected_year = int(self.year_combo.currentText())

        self.log_box.clear()
        self.log_message.emit("Processing started...")
        self.log_message.emit(f"Report file: {Path(report_file).name}")
        self.log_message.emit(f"MA files: {len(ma_files)}")
        self.log_message.emit(f"Selected period: {selected_month} {selected_year}")
        self.log_message.emit("")

        self.run_btn.setEnabled(False)
        self.select_report_btn.setEnabled(False)
        self.select_ma_btn.setEnabled(False)

        def worker():
            try:
                out_report_path, out_values_path = process_variance_report(
                    report_file, ma_files, selected_month, selected_year,
                    log_emit=self.log_message.emit
                )
                self.processing_done.emit(out_report_path, out_values_path)
            except MissingRequiredHeadersError as e:
                self._last_error = e
                self.log_message.emit("ABORTED: Missing required header(s) in Management Account file(s).")
                self.log_message.emit(e.user_message())
                self.processing_done.emit("", "")
            except Exception as e:
                self._last_error = e
                self.log_message.emit(f"CRITICAL ERROR: {e}")
                import traceback
                self.log_message.emit(traceback.format_exc())
                self.processing_done.emit("", "")

        threading.Thread(target=worker, daemon=True).start()

    def append_log(self, text: str):
        self.log_box.append(text)
        self.log_box.ensureCursorVisible()

    def on_done(self, out_path: str, values_only_path: str):
        self.run_btn.setEnabled(True)
        self.select_report_btn.setEnabled(True)
        self.select_ma_btn.setEnabled(True)

        if out_path:
            title = "Processing complete"
            lines = [f"Full report:\n{Path(out_path).name}"]
            if values_only_path:
                lines.append(f"Values-only sheet:\n{Path(values_only_path).name}")
            msg_text = "\n\n".join(lines)
        else:
            if isinstance(self._last_error, MissingRequiredHeadersError):
                title = "Missing required headers"
                msg_text = self._last_error.user_message()
            else:
                title = "Processing failed"
                msg_text = "An error occurred. Check the logs for details."

        msg = MessageBox(title, msg_text, self)
        msg.yesButton.setText("OK")
        msg.cancelButton.hide()
        msg.exec()


def get_widget():
    return MainWidget()
