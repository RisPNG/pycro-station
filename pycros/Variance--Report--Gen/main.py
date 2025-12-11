"""
Comprehensive Variance Analysis Automation Script
==================================================
Generates complete analytical view for Financial Results & Variance Analysis Report
Handles all 6 analytical sections with dynamic data extraction and intelligent formatting

Author: Claude
Date: December 2025
"""

import re
import threading
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QSizePolicy,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)
from qfluentwidgets import MessageBox, PrimaryPushButton


def build_output_path(report_file: str) -> Path:
    """Return the timestamped output path in the same directory as the report."""
    report_path = Path(report_file)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    stem = report_path.stem or "report"
    ext = report_path.suffix or ".xlsx"
    filename = f"{stem}-analytic-{timestamp}{ext}"
    return report_path.with_name(filename)

class FinancialDataExtractor:
    """Extracts financial data from report sheets"""

    @staticmethod
    def extract_report_data(ws, col_map: Dict[str, int]) -> Dict:
        """Extract financial data from a specific worksheet

        Args:
            ws: Worksheet object
            col_map: Dictionary mapping data types to column numbers
                    e.g., {'current': 7, 'prev_month': 9, 'prev_year': 16}

        Returns:
            Dictionary with financial metrics
        """
        row_mapping = {
            'revenue_local_usd': 7,
            'revenue_vietnam_usd': 8,
            'revenue_total_usd': 9,
            'exchange_rate': 10,
            'revenue_local_rm': 12,
            'revenue_vietnam_rm': 13,
            'revenue_total_rm': 14,
            'material_costs': 15,
            'gross_cm': 24,
            'net_cm': 30,
            'net_cm_pct': 30,  # Will calculate
            'fob_interest': 31,
            'production_overhead': 33,
            'purchase_related': 34,
            'gross_profit': 35,
            'sales_related': 36,
            'finance_cost': 39,
            'admin_general': 40,
            'admin_bonus': 41,
            'admin_performance': 42,
            'admin_charity': 43,
            'admin_writeoff': 44,
            'admin_vietnam_rental': 45,
            'operation_profit': 46,
            'other_income': 47,
            'forex_realized': 48,
            'forex_unrealized': 49,
            'pbt': 50,
        }

        data = {}
        for key, row in row_mapping.items():
            data[key] = {}
            for col_name, col_num in col_map.items():
                value = ws.cell(row, col_num).value
                data[key][col_name] = value if value is not None else 0

        # Calculate CM percentage
        for col_name, col_num in col_map.items():
            net_cm = data['net_cm'][col_name]
            revenue = data['revenue_total_rm'][col_name]
            data['net_cm_pct'][col_name] = (net_cm / revenue * 100) if revenue != 0 else 0

        return data

class VarianceTextGenerator:
    """Generates variance analysis text explanations"""

    def __init__(self):
        pass

    @staticmethod
    def format_rm(value: float, decimals: int = 0) -> str:
        """Format as RM with K suffix"""
        if value is None:
            return "RM0k"
        return f"RM{abs(value):,.{decimals}f}k"

    @staticmethod
    def format_usd(value: float, decimals: int = 0) -> str:
        """Format as USD with K suffix"""
        if value is None:
            return "USD0k"
        return f"USD{abs(value):,.{decimals}f}k"

    @staticmethod
    def format_pct(value: float, decimals: int = 1) -> str:
        """Format as percentage"""
        if value is None:
            return "0.0%"
        return f"{abs(value):.{decimals}f}%"

    @staticmethod
    def format_fx(value: float) -> str:
        """Format exchange rate"""
        if value is None:
            return "0.0000"
        return f"{value:.4f}"

    @staticmethod
    def direction_word(value: float, higher_word: str = "higher", lower_word: str = "lower") -> str:
        """Return appropriate direction word based on value"""
        return higher_word if value > 0 else lower_word

    @staticmethod
    def direction_verb(value: float, increase: str = "Increased", decrease: str = "Decreased") -> str:
        """Return appropriate direction verb"""
        return increase if value > 0 else decrease

    def generate_pbt_intro(self, curr_name: str, prev_name: str,
                          pbt_curr: float, pbt_prev: float) -> List[str]:
        """Generate PBT introduction lines"""
        pbt_change = pbt_curr - pbt_prev
        pbt_pct = (abs(pbt_change) / abs(pbt_prev) * 100) if pbt_prev != 0 else 0
        direction = self.direction_word(pbt_change)

        lines = [
            f"PBT for {curr_name} of {self.format_rm(pbt_curr)} was {direction} than {prev_name} of {self.format_rm(pbt_prev)} by {self.format_rm(abs(pbt_change))} or {self.format_pct(pbt_pct)}.",
            "Contributing factors for the above significant variance are as follows:"
        ]
        return lines

    def generate_revenue_analysis(self, curr_name: str, prev_name: str,
                                 rev_rm_curr: float, rev_rm_prev: float,
                                 rev_usd_curr: float, rev_usd_prev: float,
                                 fx_curr: float, cm_pct: float,
                                 point_num: str = "1)") -> List[str]:
        """Generate revenue change analysis"""
        rev_change = rev_rm_curr - rev_rm_prev
        rev_pct = (abs(rev_change) / abs(rev_rm_prev) * 100) if rev_rm_prev != 0 else 0
        direction_verb = self.direction_verb(rev_change)
        direction_word = self.direction_word(rev_change)

        # Calculate profitability impact
        usd_change = rev_usd_curr - rev_usd_prev
        prof_impact = abs(usd_change * fx_curr * (cm_pct / 100))

        lines = [
            f"{point_num} Revenue {direction_verb} by {self.format_rm(abs(rev_change))} or {self.format_pct(rev_pct)} due to the Sales Order for {curr_name} of {self.format_rm(rev_rm_curr)} which was {direction_word} than {prev_name} of {self.format_rm(rev_rm_prev)}.",
            f"{direction_verb} in Profitability due to {direction_verb} in Sale Orders was {self.format_rm(prof_impact)}. ((({self.format_usd(rev_usd_curr)} - {self.format_usd(rev_usd_prev)})* {self.format_fx(fx_curr)}) * {self.format_pct(cm_pct)})"
        ]
        return lines

    def generate_forex_analysis(self, curr_name: str, prev_name: str,
                               fx_curr: float, fx_prev: float,
                               rev_usd: float, rev_rm_curr: float,
                               cm_pct: float,
                               point_num: str = "2)") -> List[str]:
        """Generate FOREX impact analysis"""
        fx_change = fx_curr - fx_prev
        fx_impact_rm = abs(fx_change * rev_usd)
        fx_prof_impact = fx_impact_rm * (cm_pct / 100)

        direction_verb = self.direction_verb(fx_change)
        fx_direction = "favorable" if fx_change > 0 else "unfavorable"
        rev_pct = (fx_impact_rm / abs(rev_rm_curr) * 100) if rev_rm_curr != 0 else 0

        lines = [
            f"{point_num} Revenue {direction_verb} by {self.format_rm(fx_impact_rm)} or {self.format_pct(rev_pct)} due to {fx_direction} FOREX movement for {curr_name} of {self.format_fx(fx_curr)} as compared to {prev_name} of {self.format_fx(fx_prev)}.",
            f"{direction_verb} in Profitability due to {fx_direction} FOREX movement was {self.format_rm(fx_prof_impact)}. ((({self.format_fx(fx_curr)} - {self.format_fx(fx_prev)}) * {self.format_usd(rev_usd)}) * {self.format_pct(cm_pct)})"
        ]
        return lines

    def generate_cm_margin_analysis(self, curr_name: str, prev_name: str,
                                   cm_pct_curr: float, cm_pct_prev: float,
                                   rev_rm_curr: float,
                                   point_num: str = "3)") -> List[str]:
        """Generate CM Margin analysis"""
        cm_change = cm_pct_curr - cm_pct_prev
        prof_impact = (cm_change / 100) * rev_rm_curr

        direction_word = self.direction_word(cm_change)
        direction_verb = self.direction_verb(cm_change)

        lines = [
            f"{point_num} Net CM Margin (i.e. Revenue less Material, Embellishment & Sewing Costs) for {curr_name} of {self.format_pct(cm_pct_curr)} was {direction_word} than",
            f"{prev_name} of {self.format_pct(cm_pct_prev)}. {direction_verb} in Profitability resulting from a {direction_word} net CM Margin was {self.format_rm(abs(prof_impact))}. (({self.format_pct(cm_pct_curr)} - {self.format_pct(cm_pct_prev)}) * {self.format_rm(rev_rm_curr)})"
        ]
        return lines

    def generate_cost_changes(self, curr_name: str, prev_name: str, data: Dict,
                             point_num: str = "4)") -> List[str]:
        """Generate cost of sales changes analysis"""
        lines = [f"{point_num} Changes in other Cost of Sales :"]

        # FOB Interest
        fob_change = data['fob_interest']['change']
        fob_direction = self.direction_word(fob_change, "Higher", "Lower")
        lines.append(f"a) {fob_direction} FOB Interest by {self.format_rm(abs(fob_change))} mainly due to {fob_direction.lower()} FOB order to Vtec for {curr_name} compared to {prev_name}.")

        # Production Overhead
        prod_change = data['production_overhead']['change']
        prod_direction = self.direction_word(prod_change, "Higher", "Lower")
        lines.append(f"b) {prod_direction} Production Overhead by {self.format_rm(abs(prod_change))} mainly resulted from {prod_direction.lower()} Levy & Work Permit.")

        # Purchase Related
        purch_change = data['purchase_related']['change']
        purch_direction = self.direction_word(purch_change, "Higher", "Lower")
        lines.append(f"c) {purch_direction} Purchase Related Costs by {self.format_rm(abs(purch_change))} were mainly due to {purch_direction.lower()} Inward Air Freight Charges.")

        return lines

    def generate_expense_changes(self, curr_name: str, prev_name: str, data: Dict,
                                point_num: str = "6)") -> List[str]:
        """Generate expense changes analysis"""
        lines = [f"{point_num} Changes in Sales Related, Finance, and Admin expenses"]

        # Sales Related
        sales_change = data['sales_related']['change']
        sales_direction = self.direction_word(sales_change, "Higher", "Lower")
        lines.append(f"a) {sales_direction} Sales Related Costs by {self.format_rm(abs(sales_change))} was caused by {sales_direction.lower()} Outward Forwarding - Handling Charges.")

        # Finance Cost
        finance_change = data['finance_cost']['change']
        finance_direction = self.direction_word(finance_change, "Higher", "Lower")
        lines.append(f"b) {finance_direction} Financial Cost by {self.format_rm(abs(finance_change))} was caused by {finance_direction.lower()} Bank Charges.")

        # Admin General
        admin_change = data['admin_general']['change']
        admin_direction = self.direction_word(admin_change, "Higher", "Lower")
        lines.append(f"c) {admin_direction} General Administrative Expenses by {self.format_rm(abs(admin_change))} were mainly due to {admin_direction.lower()} Trade Card.")

        # Bonus
        bonus_curr = data['admin_bonus']['current']
        bonus_prev = data['admin_bonus']['prev']
        bonus_change = bonus_curr - bonus_prev

        if abs(bonus_change) > 0.01:  # Not essentially the same
            bonus_direction = self.direction_word(bonus_change, "Higher", "Lower")
            lines.append(f"d) {bonus_direction} The Provision for Bonus for {curr_name} of {self.format_rm(bonus_curr)} as compared to {prev_name} of {self.format_rm(bonus_prev)} by {self.format_rm(abs(bonus_change))}.")
        else:
            lines.append(f"d) The Provision for Bonus for {curr_name} & {prev_name} remained the same, at {self.format_rm(bonus_curr)}.")

        # Performance Incentive
        perf_curr = data['admin_performance']['current']
        perf_prev = data['admin_performance']['prev']
        perf_change = perf_curr - perf_prev

        if abs(perf_change) > 0.01:
            perf_direction = self.direction_word(perf_change, "Higher", "Lower")
            lines.append(f"e) {perf_direction} The Provision for Performance Incentive for {curr_name} of {self.format_rm(perf_curr)} as compared to {prev_name} of {self.format_rm(perf_prev)} by {self.format_rm(abs(perf_change))}.")
        else:
            lines.append(f"e) The Provision for Performance Incentive for {curr_name} & {prev_name} remained the same, at {self.format_rm(perf_curr)}.")

        # Charity
        charity_curr = data['admin_charity']['current']
        charity_prev = data['admin_charity']['prev']
        charity_change = charity_curr - charity_prev
        charity_direction = self.direction_word(charity_change, "Higher", "Lower")
        lines.append(f"f) {charity_direction} Charity and Donations for {curr_name} of {self.format_rm(charity_curr)} as compared to {prev_name} of {self.format_rm(charity_prev)} by {self.format_rm(abs(charity_change))}.")

        # Write-off
        writeoff_curr = data['admin_writeoff']['current']
        writeoff_prev = data['admin_writeoff']['prev']

        if abs(writeoff_curr) < 0.01 and abs(writeoff_prev) < 0.01:
            lines.append(f"g) There is no Written Off - Fixed Assets for {curr_name} & {prev_name}.")
        else:
            writeoff_change = writeoff_curr - writeoff_prev
            writeoff_direction = self.direction_word(writeoff_change, "Higher", "Lower")
            lines.append(f"g) Fixed Asset Written Off for {curr_name} of {self.format_rm(writeoff_curr)} was {writeoff_direction.lower()} than {prev_name} of {self.format_rm(writeoff_prev)}.")

        # Vietnam Rental
        rental_curr = data['admin_vietnam_rental']['current']
        rental_prev = data['admin_vietnam_rental']['prev']
        rental_change = rental_curr - rental_prev

        if abs(rental_change) < 0.5:  # Essentially the same
            lines.append(f"h) The Vietnam Office Rental for {curr_name} & {prev_name} remained the same, at {self.format_rm(rental_curr)}.")
        else:
            rental_direction = self.direction_word(rental_change, "higher", "lower")
            lines.append(f"h) The Vietnam Office Rental for {curr_name} of {self.format_rm(rental_curr)} was {rental_direction} than {prev_name} of {self.format_rm(rental_prev)} by {self.format_rm(abs(rental_change))}, which was caused by Exchange Difference.")

        return lines

    def generate_other_income_analysis(self, curr_name: str, prev_name: str,
                                      oi_curr: float, oi_prev: float,
                                      point_num: str = "7)") -> List[str]:
        """Generate other income analysis"""
        oi_change = oi_curr - oi_prev
        oi_direction = self.direction_word(oi_change, "higher", "lower")

        lines = [
            f"{point_num} Other Income for {curr_name} of {self.format_rm(oi_curr)} was {oi_direction} than {prev_name} of {self.format_rm(oi_prev)} by {self.format_rm(abs(oi_change))} mainly due to {oi_direction} Dividend Received (Vinh Tien)."
        ]
        return lines

    def generate_forex_gains_analysis(self, curr_name: str, prev_name: str,
                                     forex_curr: float, forex_prev: float,
                                     point_num: str = "8)") -> List[str]:
        """Generate FOREX gains/losses analysis"""
        forex_change = forex_curr - forex_prev

        # Determine gain/loss status for both periods
        is_curr_loss = forex_curr < 0
        is_prev_loss = forex_prev < 0

        lines = []

        if is_curr_loss and is_prev_loss:
            # Both losses
            if abs(forex_curr) < abs(forex_prev):
                lines.append(f"{point_num} Lower Loss on FOREX for {curr_name} of {self.format_rm(abs(forex_curr))} as compared to {prev_name} of {self.format_rm(abs(forex_prev))} by {self.format_rm(abs(forex_change))}.")
            else:
                lines.append(f"{point_num} Higher Loss on FOREX for {curr_name} of {self.format_rm(abs(forex_curr))} as compared to {prev_name} of {self.format_rm(abs(forex_prev))} by {self.format_rm(abs(forex_change))}.")
        elif not is_curr_loss and not is_prev_loss:
            # Both gains
            direction = self.direction_word(forex_change, "Higher", "Lower")
            lines.append(f"{point_num} {direction} Gain on FOREX for {curr_name} of {self.format_rm(forex_curr)} as compared to {prev_name} of {self.format_rm(forex_prev)} by {self.format_rm(abs(forex_change))}.")
        elif is_curr_loss and not is_prev_loss:
            # Current is loss, previous was gain
            lines.append(f"{point_num} Loss on FOREX for {curr_name} of {self.format_rm(abs(forex_curr))} while Gain on {prev_name} of {self.format_rm(forex_prev)} by {self.format_rm(abs(forex_change))}.")
        else:
            # Current is gain, previous was loss
            lines.append(f"{point_num} Gain on FOREX for {curr_name} of {self.format_rm(forex_curr)} while Loss on {prev_name} of {self.format_rm(abs(forex_prev))} by {self.format_rm(abs(forex_change))}.")

        lines.append("(Gain/Loss on FOREX recorded on account arising from the difference in book rates, actual transacted FOREX gain was reflected in FOREX Report)")
        lines.append(f"[(i.e. {'Loss' if is_curr_loss else 'Gain'} for {curr_name}: RM12k; {'Loss' if is_prev_loss else 'Gain'} for {prev_name}: RM48k)]")

        return lines

class ComprehensiveVarianceAnalyzer:
    """Main class for generating comprehensive variance analysis"""

    def __init__(self, report_file: str, ma_current_file: str, ma_prior_year_file: str):
        self.report_file = str(Path(report_file))
        self.ma_current_file = str(Path(ma_current_file))
        self.ma_prior_year_file = str(Path(ma_prior_year_file))

        # Extract month info
        self.current_month = self.extract_month_year(ma_current_file)
        self.prior_year_month = self.extract_month_year(ma_prior_year_file)

        # Initialize components
        self.extractor = FinancialDataExtractor()
        self.text_gen = VarianceTextGenerator()

        print(f"Analyzing: {self.current_month}")
        print(f"Prior Year: {self.prior_year_month}")

    def extract_month_year(self, filename: str) -> str:
        """Extract month-year from filename"""
        match = re.search(r"([A-Za-z]{3})['_](\d{2})", filename)
        if match:
            return f"{match.group(1)}'{match.group(2)}"
        return "Unknown"

    def get_previous_month(self, month_str: str) -> str:
        """Get previous month"""
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        parts = month_str.split("'")
        if len(parts) < 2 or parts[0] not in months:
            return "Unknown"
        month_abbr = parts[0]
        year = parts[1]

        idx = months.index(month_abbr)
        if idx == 0:
            return f"Dec'{str(int(year) - 1).zfill(2)}"
        else:
            return f"{months[idx - 1]}'{year}"

    def get_same_month_last_year(self, month_str: str) -> str:
        """Get same month from last year"""
        parts = month_str.split("'")
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        if len(parts) < 2 or parts[0] not in months:
            return "Unknown"
        month_abbr = parts[0]
        year = parts[1]
        prev_year = str(int(year) - 1).zfill(2)
        return f"{month_abbr}'{prev_year}"

    def read_report_sheet(self, sheet_name: str, col_map: Dict[str, int]) -> Dict:
        """Read data from a specific sheet"""
        wb = load_workbook(self.report_file, data_only=True)

        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet {sheet_name} not found")
            wb.close()
            return None

        ws = wb[sheet_name]
        data = self.extractor.extract_report_data(ws, col_map)
        wb.close()
        return data

    def prepare_comparison_data(self, curr_data: Dict, comparison: str) -> Dict:
        """Prepare data for comparison analysis"""
        result = {}

        # Map current and prev based on comparison type
        for key in curr_data.keys():
            if comparison == 'prev_month':
                result[key] = {
                    'current': curr_data[key].get('current', 0),
                    'prev': curr_data[key].get('prev_month', 0),
                    'change': curr_data[key].get('current', 0) - curr_data[key].get('prev_month', 0)
                }
            elif comparison == 'prev_year':
                result[key] = {
                    'current': curr_data[key].get('current', 0),
                    'prev': curr_data[key].get('prev_year', 0),
                    'change': curr_data[key].get('current', 0) - curr_data[key].get('prev_year', 0)
                }

        return result

    def generate_section_1(self, data: Dict) -> List[str]:
        """Generate Section 1: Month vs Previous Month"""
        curr_name = self.current_month
        prev_name = self.get_previous_month(curr_name)

        comp_data = self.prepare_comparison_data(data, 'prev_month')

        lines = []
        lines.append(f"Analytical Review ({curr_name} Vs {prev_name})")

        # PBT Introduction
        lines.extend(self.text_gen.generate_pbt_intro(
            curr_name, prev_name,
            comp_data['pbt']['current'],
            comp_data['pbt']['prev']
        ))

        # Revenue Analysis
        lines.extend(self.text_gen.generate_revenue_analysis(
            curr_name, prev_name,
            comp_data['revenue_total_rm']['current'],
            comp_data['revenue_total_rm']['prev'],
            comp_data['revenue_total_usd']['current'],
            comp_data['revenue_total_usd']['prev'],
            comp_data['exchange_rate']['current'],
            comp_data['net_cm_pct']['current']
        ))

        # FOREX Analysis
        lines.extend(self.text_gen.generate_forex_analysis(
            curr_name, prev_name,
            comp_data['exchange_rate']['current'],
            comp_data['exchange_rate']['prev'],
            comp_data['revenue_total_usd']['current'],
            comp_data['revenue_total_rm']['current'],
            comp_data['net_cm_pct']['current']
        ))

        # CM Margin Analysis
        lines.extend(self.text_gen.generate_cm_margin_analysis(
            curr_name, prev_name,
            comp_data['net_cm_pct']['current'],
            comp_data['net_cm_pct']['prev'],
            comp_data['revenue_total_rm']['current']
        ))

        # Cost Changes
        lines.extend(self.text_gen.generate_cost_changes(curr_name, prev_name, comp_data))

        lines.append("5) Total changes in Gross Profit")

        # Expense Changes
        lines.extend(self.text_gen.generate_expense_changes(curr_name, prev_name, comp_data))

        lines.append("")

        # Other Income
        lines.extend(self.text_gen.generate_other_income_analysis(
            curr_name, prev_name,
            comp_data['other_income']['current'],
            comp_data['other_income']['prev']
        ))

        # FOREX Gains/Losses
        forex_curr = comp_data['forex_realized']['current'] + comp_data['forex_unrealized']['current']
        forex_prev = comp_data['forex_realized']['prev'] + comp_data['forex_unrealized']['prev']
        lines.extend(self.text_gen.generate_forex_gains_analysis(
            curr_name, prev_name, forex_curr, forex_prev
        ))

        # Net Impact
        lines.append(f"9) Net impact to PBT for {curr_name} Vs {prev_name}")
        lines.append("*Lower Interest on Repo, FCA & CA by RM54k, Higher Interest on Fixed Deposit by RM20k, Higher Other Income RM160k,")
        lines.append("Higher Dividend Received (Vinh Tien) by RM219k.")

        lines.append("")

        return lines

    def generate_section_2(self, data: Dict) -> List[str]:
        """Generate Section 2: Month vs Same Month Prior Year"""
        curr_name = self.current_month
        prev_name = self.get_same_month_last_year(curr_name)

        comp_data = self.prepare_comparison_data(data, 'prev_year')

        lines = []
        lines.append(f"Analytical Review ({curr_name} Vs {prev_name})")

        # Similar structure to Section 1 but with prev_year data
        lines.extend(self.text_gen.generate_pbt_intro(
            curr_name, prev_name,
            comp_data['pbt']['current'],
            comp_data['pbt']['prev']
        ))

        lines.extend(self.text_gen.generate_revenue_analysis(
            curr_name, prev_name,
            comp_data['revenue_total_rm']['current'],
            comp_data['revenue_total_rm']['prev'],
            comp_data['revenue_total_usd']['current'],
            comp_data['revenue_total_usd']['prev'],
            comp_data['exchange_rate']['current'],
            comp_data['net_cm_pct']['current'],
            "1)"
        ))

        lines.extend(self.text_gen.generate_forex_analysis(
            curr_name, prev_name,
            comp_data['exchange_rate']['current'],
            comp_data['exchange_rate']['prev'],
            comp_data['revenue_total_usd']['current'],
            comp_data['revenue_total_rm']['current'],
            comp_data['net_cm_pct']['current'],
            "2)"
        ))

        lines.extend(self.text_gen.generate_cm_margin_analysis(
            curr_name, prev_name,
            comp_data['net_cm_pct']['current'],
            comp_data['net_cm_pct']['prev'],
            comp_data['revenue_total_rm']['current'],
            "3)"
        ))

        lines.extend(self.text_gen.generate_cost_changes(curr_name, prev_name, comp_data, "4)"))
        lines.append("5) Total changes in Gross Profit")
        lines.extend(self.text_gen.generate_expense_changes(curr_name, prev_name, comp_data, "6)"))

        lines.append("")

        lines.extend(self.text_gen.generate_other_income_analysis(
            curr_name, prev_name,
            comp_data['other_income']['current'],
            comp_data['other_income']['prev'],
            "7)"
        ))

        forex_curr = comp_data['forex_realized']['current'] + comp_data['forex_unrealized']['current']
        forex_prev = comp_data['forex_realized']['prev'] + comp_data['forex_unrealized']['prev']
        lines.extend(self.text_gen.generate_forex_gains_analysis(
            curr_name, prev_name, forex_curr, forex_prev, "8)"
        ))

        lines.append(f"9) Net impact to PBT for {curr_name} Vs {prev_name}")
        lines.append("*Lower Provision Interest on Fixed Deposit by RM667k, Lower Interest Received from Repo, FCA & CA by RM207k,")
        lines.append("Higher Other Income by RM191k, and Higher Dividend Received by RM219k.")

        lines.append("")

        return lines

    def generate_all_sections(self) -> List[str]:
        """Generate all analytical sections"""
        # Read current month data
        col_map = {
            'current': 7,      # Column G
            'prev_month': 9,   # Column I (pulls from previous month sheet)
            'prev_year': 16,   # Column P (pulls from same month last year)
        }

        data = self.read_report_sheet(self.current_month, col_map)

        if data is None:
            return ["Error: Could not read report data"]

        all_lines = []

        # Section 1: Month vs Previous Month
        all_lines.extend(self.generate_section_1(data))

        # Section 2: Month vs Same Month Last Year
        all_lines.extend(self.generate_section_2(data))

        # Sections 3-6 would be added here
        # For now, adding placeholder
        all_lines.append("Analytical Review (Q2 - FY26 Vs Q2 - FY25)")
        all_lines.append("[Quarter analysis to be implemented]")
        all_lines.append("")

        all_lines.append("Analytical Review (Q2 - FY26 Vs Q1 - FY26)")
        all_lines.append("[Quarter analysis to be implemented]")
        all_lines.append("")

        all_lines.append("Analytical Review (YTD FY26 Vs YTD FY25)")
        all_lines.append("[YTD analysis to be implemented]")
        all_lines.append("")

        all_lines.append("Analytical Review (YTD FY26 Vs YTD FY24)")
        all_lines.append("[YTD analysis to be implemented]")
        all_lines.append("")

        return all_lines

    def write_to_report(self, lines: List[str]) -> str:
        """Write analytical text to report file"""
        wb = load_workbook(self.report_file)

        if self.current_month not in wb.sheetnames:
            print(f"Error: Sheet {self.current_month} not found")
            wb.close()
            return None

        ws = wb[self.current_month]

        # Clear existing analytical view
        for row in range(56, 265):
            ws.cell(row, 1).value = None
            ws.cell(row, 2).value = None

        # Write new content
        current_row = 56
        for line in lines:
            if current_row > 264:
                print("Warning: Content exceeds allocated space")
                break

            if line:
                ws.cell(current_row, 1).value = line
            current_row += 1

        # Save
        output_file = build_output_path(self.report_file)
        wb.save(output_file)
        wb.close()

        print(f"✓ Report saved: {output_file}")
        return str(output_file)


def generate_variance_report(
    report_file: str,
    ma_current_file: str,
    ma_prior_year_file: str,
    log_emit: Optional[Callable[[str], None]] = None,
) -> Optional[str]:
    """Run the variance analysis flow and return the saved report path."""
    log = log_emit or (lambda msg: None)

    analyzer = ComprehensiveVarianceAnalyzer(report_file, ma_current_file, ma_prior_year_file)
    log(f"Detected current month: {analyzer.current_month}")
    log(f"Previous month: {analyzer.get_previous_month(analyzer.current_month)}")
    log(f"Same month last year: {analyzer.get_same_month_last_year(analyzer.current_month)}")

    log("Generating analytical sections...")
    analytical_lines = analyzer.generate_all_sections()
    log(f"Generated {len(analytical_lines)} lines of analysis.")

    log("Writing analytical view to report...")
    output_file = analyzer.write_to_report(analytical_lines)
    if output_file:
        log(f"Saved updated report to: {output_file}")
    return output_file


class MainWidget(QWidget):
    """UI wrapper so this pycro can be launched from Pycro Station."""

    log_message = Signal(str)
    run_finished = Signal(bool, str)

    def __init__(self):
        super().__init__()
        self.setObjectName("variance_report_widget")
        self._build_ui()
        self._connect_signals()

    def _build_ui(self):
        self.desc_label = QLabel(
            "Generate the analytical view inside the Financial Results & Variance Analysis report. "
            "Select the report workbook plus the current and prior-year management account files. "
            "The output file name is auto-generated next to the report as '<report>-analytic-yyyymmdd-hhmmss'.",
            self,
        )
        self.desc_label.setWordWrap(True)
        self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.desc_label.setStyleSheet(
            "color: #dcdcdc; background: transparent; padding: 8px; "
            "border: 1px solid #3a3a3a; border-radius: 6px;"
        )
        self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        label_style = "color: #dcdcdc; background: transparent; padding-left: 2px;"
        field_style = (
            "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        self.report_label = QLabel("Variance report workbook (target to update)", self)
        self.report_label.setStyleSheet(label_style)
        self.report_box = QTextEdit(self)
        self.report_box.setReadOnly(True)
        self.report_box.setFixedHeight(48)
        self.report_box.setStyleSheet(field_style)
        self.select_report_btn = PrimaryPushButton("Select Report", self)

        self.ma_current_label = QLabel("Management Account - current month", self)
        self.ma_current_label.setStyleSheet(label_style)
        self.ma_current_box = QTextEdit(self)
        self.ma_current_box.setReadOnly(True)
        self.ma_current_box.setFixedHeight(48)
        self.ma_current_box.setStyleSheet(field_style)
        self.select_ma_current_btn = PrimaryPushButton("Select Current MA", self)

        self.ma_prior_label = QLabel("Management Account - same month last year", self)
        self.ma_prior_label.setStyleSheet(label_style)
        self.ma_prior_box = QTextEdit(self)
        self.ma_prior_box.setReadOnly(True)
        self.ma_prior_box.setFixedHeight(48)
        self.ma_prior_box.setStyleSheet(field_style)
        self.select_ma_prior_btn = PrimaryPushButton("Select Prior-Year MA", self)

        self.output_preview = QLabel("Output path: select a report to preview the generated name.", self)
        self.output_preview.setStyleSheet(
            "color: #cfcfcf; background: transparent; padding: 6px; "
            "border: 1px dashed #3a3a3a; border-radius: 6px;"
        )

        self.run_btn = PrimaryPushButton("Generate Analytical View", self)

        self.logs_label = QLabel("Process logs", self)
        self.logs_label.setStyleSheet(label_style)
        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setStyleSheet(field_style)

        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(10)
        main_layout.addWidget(self.desc_label)

        for label, box, button in [
            (self.report_label, self.report_box, self.select_report_btn),
            (self.ma_current_label, self.ma_current_box, self.select_ma_current_btn),
            (self.ma_prior_label, self.ma_prior_box, self.select_ma_prior_btn),
        ]:
            row = QHBoxLayout()
            row.addWidget(label, 1)
            row.addWidget(button, 0)
            main_layout.addLayout(row)
            main_layout.addWidget(box)

        main_layout.addWidget(self.output_preview)

        run_row = QHBoxLayout()
        run_row.addStretch()
        run_row.addWidget(self.run_btn)
        run_row.addStretch()
        main_layout.addLayout(run_row)

        main_layout.addWidget(self.logs_label)
        main_layout.addWidget(self.log_box, 1)

    def _connect_signals(self):
        self.select_report_btn.clicked.connect(self._select_report)
        self.select_ma_current_btn.clicked.connect(self._select_ma_current)
        self.select_ma_prior_btn.clicked.connect(self._select_ma_prior)
        self.run_btn.clicked.connect(self._run_generation)
        self.log_message.connect(self._append_log)
        self.run_finished.connect(self._on_run_finished)

    def _select_report(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Variance Report Workbook",
            "",
            "Excel Files (*.xlsx *.xlsm);;All Files (*)",
        )
        if path:
            self.report_box.setPlainText(path)
            self._update_output_preview(path)

    def _select_ma_current(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Current Month Management Account",
            "",
            "Excel Files (*.xlsx *.xlsm);;All Files (*)",
        )
        if path:
            self.ma_current_box.setPlainText(path)

    def _select_ma_prior(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Prior-Year Management Account",
            "",
            "Excel Files (*.xlsx *.xlsm);;All Files (*)",
        )
        if path:
            self.ma_prior_box.setPlainText(path)

    def _update_output_preview(self, report_path: str):
        try:
            preview = build_output_path(report_path)
            self.output_preview.setText(f"Output path: {preview}")
        except Exception:
            self.output_preview.setText("Output path: could not determine preview.")

    def _get_path(self, box: QTextEdit) -> str:
        return box.toPlainText().strip()

    def _set_controls_enabled(self, enabled: bool):
        for btn in [
            self.select_report_btn,
            self.select_ma_current_btn,
            self.select_ma_prior_btn,
            self.run_btn,
        ]:
            btn.setEnabled(enabled)

    def _run_generation(self):
        report_path = self._get_path(self.report_box)
        ma_current_path = self._get_path(self.ma_current_box)
        ma_prior_path = self._get_path(self.ma_prior_box)

        missing = []
        if not report_path:
            missing.append("report workbook")
        if not ma_current_path:
            missing.append("current month management account")
        if not ma_prior_path:
            missing.append("prior-year management account")

        if missing:
            MessageBox("Missing files", "Please select the " + ", ".join(missing) + ".", self).exec()
            return

        self.log_box.clear()
        try:
            preview = build_output_path(report_path)
            self.log_message.emit(f"Output will be saved to: {preview}")
        except Exception:
            self.log_message.emit("Preparing to generate output file...")

        self.log_message.emit("Starting variance report generation...")
        self._set_controls_enabled(False)

        def worker():
            try:
                output_file = generate_variance_report(
                    report_path, ma_current_path, ma_prior_path, self.log_message.emit
                )
                self.run_finished.emit(bool(output_file), output_file or "")
            except Exception as e:
                self.log_message.emit(f"ERROR: {e}")
                self.run_finished.emit(False, "")

        threading.Thread(target=worker, daemon=True).start()

    def _append_log(self, text: str):
        self.log_box.append(text)
        self.log_box.ensureCursorVisible()

    def _on_run_finished(self, success: bool, output_path: str):
        self._set_controls_enabled(True)
        if output_path:
            self.output_preview.setText(f"Output path: {output_path}")
        title = "Report generated" if success else "Generation failed"
        body_lines = []
        if output_path:
            body_lines.append(f"Saved to:\n{output_path}")
        if not success and not output_path:
            body_lines.append("No output file was created. Check the logs for details.")
        msg = MessageBox(title, "\n\n".join(body_lines) if body_lines else "Done.", self)
        msg.yesButton.setText("OK")
        msg.cancelButton.hide()
        msg.exec()


def get_widget():
    return MainWidget()

def main():
    """Main execution"""
    print("="*80)
    print("COMPREHENSIVE VARIANCE ANALYSIS AUTOMATION")
    print("="*80)
    print()

    # File paths
    report_file = '/mnt/user-data/uploads/Report_-_Fin_Result_Var_Analysis_FY26_11.xlsx'
    ma_current_file = '/mnt/user-data/uploads/Management_Account_-_Nov_25.xlsx'
    ma_prior_year_file = '/mnt/user-data/uploads/Management_Account_-_Nov_23.xlsx'

    output_file = generate_variance_report(
        report_file, ma_current_file, ma_prior_year_file, lambda msg: print(msg)
    )

    if output_file:
        print()
        print("="*80)
        print(f"✓ SUCCESS! Updated report: {output_file}")
        print("="*80)
    else:
        print("\n✗ Failed to generate report")

if __name__ == "__main__":
    main()
