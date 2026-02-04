import os
import re
import sys
import threading
import types
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Callable, Iterable, Optional

from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import QFileDialog, QHBoxLayout, QLabel, QSizePolicy, QTextEdit, QVBoxLayout, QWidget
from qfluentwidgets import ComboBox, MessageBox, PrimaryPushButton

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils.datetime import from_excel
from openpyxl.utils import get_column_letter


MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# Default 2026 MY public holidays (editable in UI). These are commonly used national holidays;
# some Malaysia holidays are state-specific, so users can add/remove as needed.
DEFAULT_MY_HOLIDAYS_2026 = [
    "2026-01-01",  # New Year's Day (state holiday in some MY states)
    "2026-02-17",  # Chinese New Year
    "2026-02-18",  # Chinese New Year (Day 2)
    "2026-03-21",  # Hari Raya Puasa
    "2026-03-22",  # Hari Raya Puasa (Day 2)
    "2026-05-01",  # Labour Day
    "2026-05-27",  # Hari Raya Qurban
    "2026-05-28",  # Hari Raya Qurban (Day 2)
    "2026-05-31",  # Wesak Day
    "2026-06-01",  # Agong's Birthday
    "2026-06-17",  # Awal Muharram
    "2026-08-25",  # Maulidur Rasul
    "2026-08-31",  # National Day
    "2026-09-16",  # Malaysia Day
    "2026-11-08",  # Deepavali
    "2026-12-25",  # Christmas Day
]


def _emit(log_emit: Optional[Callable[[str], None]], msg: str) -> None:
    if log_emit:
        log_emit(msg)


def ensure_olefile_available(log_emit=None) -> None:
    """
    Pycro Station's embedded venv can end up with a broken `olefile` namespace package
    (no MAGIC/OleFileIO). Pillow's OLE-based plugins then crash during format detection.
    We patch `sys.modules['olefile']` with a minimal compatible shim to prevent crashes.
    """
    try:
        import olefile as _olefile  # type: ignore
        if hasattr(_olefile, "MAGIC") and hasattr(_olefile, "OleFileIO"):
            return
    except Exception:
        pass

    shim = types.ModuleType("olefile")
    shim.__version__ = "0.0-shim"
    shim.MAGIC = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"

    class OleFileIO:  # noqa: N801 (match upstream name)
        def __init__(self, *args, **kwargs):
            raise OSError("olefile is not installed (shim active)")

    def isOleFile(*args, **kwargs):  # noqa: N802 (match upstream name)
        return False

    shim.OleFileIO = OleFileIO
    shim.isOleFile = isOleFile

    sys.modules["olefile"] = shim
    _emit(log_emit, "[ENV] Patched broken 'olefile' module (Pillow image detection).")


def _norm_str(val) -> str:
    return str(val).strip() if val is not None else ""


def normalize_invoice(val) -> str:
    if val is None:
        return ""
    if isinstance(val, bool):
        return ""
    if isinstance(val, int):
        return str(val).strip()
    if isinstance(val, float):
        if abs(val - int(val)) < 1e-9:
            return str(int(val)).strip()
        return str(val).strip()
    s = str(val).strip()
    if not s:
        return ""
    if re.fullmatch(r"\d+\.0", s):
        try:
            return str(int(float(s))).strip()
        except Exception:
            return s
    return s


def parse_money(val) -> Optional[float]:
    if val is None or val == "":
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        try:
            return float(val)
        except Exception:
            return None
    s = str(val).strip()
    if not s:
        return None
    if s.startswith("="):
        return None
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    s = s.replace("$", "").replace(",", "").strip()
    # Handle values like "24172.90 USD" by extracting the first number.
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return None
    try:
        f = float(m.group(0))
        return -f if neg else f
    except Exception:
        return None


def parse_date_any(val, *, fallback_year: Optional[int] = None) -> Optional[date]:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        try:
            return from_excel(val).date()
        except Exception:
            return None
    s = str(val).strip()
    if not s:
        return None
    fmts = [
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%b-%Y",
        "%d-%b-%y",
        "%d-%b",
        "%d %b %Y",
        "%d %b %y",
    ]
    for fmt in fmts:
        try:
            dt = datetime.strptime(s, fmt)
            if fmt in ("%d-%b",) and fallback_year is not None:
                dt = dt.replace(year=int(fallback_year))
            return dt.date()
        except Exception:
            continue
    return None


def ensure_unique_path(path: str) -> str:
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    n = 1
    while True:
        candidate = f"{base} ({n}){ext}"
        if not os.path.exists(candidate):
            return candidate
        n += 1


def format_sheet_name(month_abbrev: str, year: int, week: int) -> str:
    yy = str(int(year))[-2:]
    m = month_abbrev.strip()
    return f"{m}'{yy} Wk {int(week)}"


def parse_holiday_lines(text: str, log_emit=None) -> set[date]:
    holidays: set[date] = set()
    for raw in (text or "").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        d = parse_date_any(line)
        if not d:
            _emit(log_emit, f"[HOLIDAYS] Skipped invalid date: {line}")
            continue
        holidays.add(d)
    return holidays


def next_business_day(d: date, holidays: set[date]) -> date:
    candidate = d + timedelta(days=1)
    while candidate.weekday() >= 5 or candidate in holidays:
        candidate += timedelta(days=1)
    return candidate


def normalize_ref_no(val) -> str:
    """
    Normalize a Payment/Ref number so it can be matched between:
    - Export Bill column J values like 'TC-508801542519'
    - FEAC chart 'Ref. No.' values like 508801542519
    """
    s = normalize_invoice(val)
    if not s:
        return ""
    s = s.strip()
    if s.upper().startswith("TC-"):
        s = s[3:].strip()
    return s


def _find_cell_with_value(ws, needle: str, *, max_rows: int = 80, max_cols: int = 30) -> Optional[tuple[int, int]]:
    target = needle.strip().upper()
    mr = min(ws.max_row or 0, max_rows)
    mc = min(ws.max_column or 0, max_cols)
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            if str(v).strip().upper() == target:
                return (r, c)
    return None


def read_feac_ref_order(path: str, *, log_emit=None) -> dict[str, int]:
    """
    Reads FEAC chart and returns an order map for Ref numbers.
    For each sheet (in workbook order): find cell 'Ref. No.' and read values under it (same column),
    appending to one continuous list.
    """
    _emit(log_emit, f"[FEAC] Opening: {path}")
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        order: dict[str, int] = {}
        seq = 0
        for ws in wb.worksheets:
            pos = _find_cell_with_value(ws, "Ref. No.")
            if not pos:
                _emit(log_emit, f"[FEAC] '{ws.title}': 'Ref. No.' not found (skipped).")
                continue
            header_r, header_c = pos
            blanks_in_a_row = 0
            started = False
            for r in range(header_r + 1, (ws.max_row or header_r) + 1):
                v = ws.cell(r, header_c).value
                if v is None or str(v).strip() == "":
                    blanks_in_a_row += 1
                    if started and blanks_in_a_row >= 30:
                        break
                    continue
                blanks_in_a_row = 0
                started = True
                ref = normalize_ref_no(v)
                if not ref:
                    continue
                if ref not in order:
                    order[ref] = seq
                    seq += 1
        _emit(log_emit, f"[FEAC] Loaded {len(order)} Ref. No. item(s).")
        return order
    finally:
        wb.close()


@dataclass
class _CellSnap:
    value: object
    style: object
    hyperlink: object
    comment: object


@dataclass
class _RowSnap:
    origin_row: int
    cells: list[_CellSnap]
    height: Optional[float]


def _is_formula_cell(cell) -> bool:
    v = cell.value
    if getattr(cell, "data_type", None) == "f":
        return True
    return isinstance(v, str) and v.strip().startswith("=")


def _row_is_blank(cells: Iterable) -> bool:
    for c in cells:
        v = c.value
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        return False
    return True


def _snapshot_row(cells: Iterable, row_idx: int, ws) -> _RowSnap:
    snaps: list[_CellSnap] = []
    for c in cells:
        snaps.append(_CellSnap(value=c.value, style=getattr(c, "_style", None), hyperlink=c.hyperlink, comment=c.comment))
    h = ws.row_dimensions[row_idx].height
    return _RowSnap(origin_row=row_idx, cells=snaps, height=h)


def _apply_row_snapshot(ws, dst_row: int, snap: _RowSnap) -> None:
    if snap.height is not None:
        ws.row_dimensions[dst_row].height = snap.height
    for col_idx, cs in enumerate(snap.cells, start=1):
        dst = ws.cell(dst_row, col_idx)
        v = cs.value
        if isinstance(v, str) and v.startswith("="):
            origin = f"{get_column_letter(col_idx)}{snap.origin_row}"
            dest = f"{get_column_letter(col_idx)}{dst_row}"
            try:
                v = Translator(v, origin=origin).translate_formula(dest)
            except Exception:
                pass
        dst.value = v
        if cs.style is not None:
            dst._style = cs.style
        if cs.hyperlink:
            dst.hyperlink = cs.hyperlink
        if cs.comment:
            dst.comment = cs.comment


def _apply_blank_row(ws, dst_row: int, template: Optional[_RowSnap], max_col: int) -> None:
    if template and template.height is not None:
        ws.row_dimensions[dst_row].height = template.height
    for col_idx in range(1, max_col + 1):
        dst = ws.cell(dst_row, col_idx)
        dst.value = None
        if template and col_idx - 1 < len(template.cells):
            st = template.cells[col_idx - 1].style
            if st is not None:
                dst._style = st


def _apply_total_row(
    ws,
    dst_row: int,
    template: Optional[_RowSnap],
    max_col: int,
    group_start_row: int,
    group_end_row: int,
) -> None:
    if template and template.height is not None:
        ws.row_dimensions[dst_row].height = template.height
    for col_idx in range(1, max_col + 1):
        dst = ws.cell(dst_row, col_idx)
        dst.value = None
        if template and col_idx - 1 < len(template.cells):
            st = template.cells[col_idx - 1].style
            if st is not None:
                dst._style = st

    # Keep any label in column A from template (if present)
    if template and template.cells:
        v = template.cells[0].value
        if v is not None and not (isinstance(v, str) and v.strip().startswith("=")):
            ws.cell(dst_row, 1).value = v

    if group_end_row >= group_start_row:
        ws.cell(dst_row, 2).value = f"=SUM(B{group_start_row}:B{group_end_row})"
    else:
        ws.cell(dst_row, 2).value = "=0"


def _find_first_data_row_export_bill(ws, *, amount_col: int = 2) -> Optional[int]:
    max_row = ws.max_row or 0
    for r in range(1, max_row + 1):
        inv = normalize_invoice(ws.cell(r, 1).value)
        if not inv:
            continue
        if parse_money(ws.cell(r, amount_col).value) is None:
            continue
        return r
    return None


def regroup_and_sort_export_bill_sheet(ws, ref_order: dict[str, int], *, log_emit=None) -> None:
    """
    Rebuilds the grouped area so that:
    - Rows with Value Date (col E) are grouped by that date, ascending
    - Within each date group, rows are sorted by FEAC order of column J (Payment No)
    - Rows with no Value Date stay in the last group (\"no tradecard\") with its own total
    """
    start = _find_first_data_row_export_bill(ws)
    end = find_last_total_row(ws)
    if start is None or end is None or end < start:
        _emit(log_emit, f"[SORT] {ws.title}: unable to locate data region (skipped).")
        return

    max_col = ws.max_column or 0
    if max_col <= 0:
        return

    total_template: Optional[_RowSnap] = None
    blank_template: Optional[_RowSnap] = None
    date_groups: dict[date, list[_RowSnap]] = {}
    unmatched: list[_RowSnap] = []

    for row_idx, row_cells in enumerate(
        ws.iter_rows(min_row=start, max_row=end, min_col=1, max_col=max_col),
        start=start,
    ):
        b = row_cells[1] if len(row_cells) >= 2 else ws.cell(row_idx, 2)
        if _is_formula_cell(b):
            if total_template is None:
                total_template = _snapshot_row(row_cells, row_idx, ws)
            continue

        if _row_is_blank(row_cells):
            if blank_template is None:
                blank_template = _snapshot_row(row_cells, row_idx, ws)
            continue

        inv = normalize_invoice(row_cells[0].value if row_cells else ws.cell(row_idx, 1).value)
        if not inv:
            continue

        snap = _snapshot_row(row_cells, row_idx, ws)

        e_val = row_cells[4].value if len(row_cells) >= 5 else ws.cell(row_idx, 5).value
        e_date = parse_date_any(e_val)
        if e_date:
            date_groups.setdefault(e_date, []).append(snap)
        else:
            unmatched.append(snap)

    if not date_groups and not unmatched:
        _emit(log_emit, f"[SORT] {ws.title}: no data rows detected (skipped).")
        return

    def sort_key(s: _RowSnap) -> tuple[int, int]:
        j_val = s.cells[9].value if len(s.cells) >= 10 else None
        ref = normalize_ref_no(j_val)
        return (ref_order.get(ref, 10**9), s.origin_row)

    out_rows: list[tuple[str, object]] = []
    cursor = start

    for idx, d in enumerate(sorted(date_groups.keys())):
        group = sorted(date_groups[d], key=sort_key)
        group_start = cursor
        for s in group:
            out_rows.append(("data", s))
            cursor += 1
        group_end = cursor - 1
        out_rows.append(("total", (group_start, group_end)))
        cursor += 1
        if idx != len(date_groups) - 1 or unmatched:
            out_rows.append(("blank", None))
            cursor += 1

    # Unmatched (no Value Date) group at the end
    if unmatched:
        unmatched_start = cursor
        for s in unmatched:
            out_rows.append(("data", s))
            cursor += 1
        unmatched_end = cursor - 1
        out_rows.append(("total", (unmatched_start, unmatched_end)))
        cursor += 1

    # Rewrite region
    orig_count = end - start + 1
    new_count = len(out_rows)
    ws.delete_rows(start, orig_count)
    ws.insert_rows(start, new_count)

    write_row = start
    for kind, payload in out_rows:
        if kind == "data":
            _apply_row_snapshot(ws, write_row, payload)  # type: ignore[arg-type]
        elif kind == "blank":
            _apply_blank_row(ws, write_row, blank_template, max_col)
        elif kind == "total":
            group_start, group_end = payload  # type: ignore[misc]
            _apply_total_row(ws, write_row, total_template, max_col, int(group_start), int(group_end))
        write_row += 1

    _emit(log_emit, f"[SORT] {ws.title}: regrouped {sum(len(v) for v in date_groups.values())} matched row(s), {len(unmatched)} unmatched.")

def find_header_row_by_value(ws, needle: str, *, max_scan_rows: Optional[int] = None) -> Optional[int]:
    target = needle.strip().upper()
    max_row = ws.max_row or 0
    if max_scan_rows is not None:
        max_row = min(max_row, int(max_scan_rows))
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, values_only=True), start=1):
        for v in row:
            if v is None:
                continue
            if str(v).strip().upper() == target:
                return r_idx
    return None


def header_indices(header_row: Iterable, name: str) -> list[int]:
    target = name.strip().upper()
    out: list[int] = []
    for idx, v in enumerate(header_row):
        if v is None:
            continue
        if str(v).strip().upper() == target:
            out.append(idx)
    return out


def header_index_prefer_exact(header_row: Iterable, exact_text: str) -> Optional[int]:
    """Find a header column index, preferring an exact (case-sensitive) match, else case-insensitive."""
    exact = exact_text.strip()
    for idx, v in enumerate(header_row):
        if v is None:
            continue
        if str(v).strip() == exact:
            return idx
    matches = header_indices(header_row, exact_text)
    return matches[0] if matches else None


@dataclass(frozen=True)
class ExportRecord:
    dest_sheet: str
    invoice: str
    amount: float
    exfty: date
    lead_days: int


@dataclass(frozen=True)
class AmountMismatch:
    trade_card_file: str
    invoice: str
    export_bill_sheet: str
    export_bill_row: int
    trade_card_amount: float
    export_bill_amount: float


def read_vn_records(path: str, sheet_name: str, *, log_emit=None) -> list[ExportRecord]:
    _emit(log_emit, f"[VN] Opening: {path}")
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"[VN] Sheet not found: {sheet_name}")
        ws = wb[sheet_name]
        header_row_idx = find_header_row_by_value(ws, "JOB NO.")
        if not header_row_idx:
            raise ValueError("[VN] Header row not found (cell value 'JOB NO.').")
        header_row = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))

        idx_buyer = header_indices(header_row, "BUYER")
        idx_exfty = header_indices(header_row, "ACT. EX-FTY")
        idx_inv = header_indices(header_row, "INV #")
        idx_amt = header_indices(header_row, "ACT. AMOUNT")
        payterm_i = header_index_prefer_exact(header_row, "PAYMENT TERM")

        if not idx_buyer or not idx_exfty or not idx_inv or not idx_amt or payterm_i is None:
            raise ValueError(
                "[VN] Missing one or more headers: BUYER, PAYMENT TERM, ACT. EX-FTY, INV #, ACT. AMOUNT"
            )

        buyer_i = idx_buyer[0]
        exfty_i = idx_exfty[0]
        inv_i = idx_inv[0]
        amt_i = idx_amt[0]

        out: list[ExportRecord] = []
        rows_with_inv = 0
        skipped_term = 0
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            inv = normalize_invoice(row[inv_i] if inv_i < len(row) else None)
            if not inv:
                continue
            rows_with_inv += 1

            term_val = _norm_str(row[payterm_i] if payterm_i < len(row) else None)
            term_norm = re.sub(r"\s+", "", term_val.strip().upper())
            if term_norm != "BYT/C":
                skipped_term += 1
                continue

            buyer = _norm_str(row[buyer_i] if buyer_i < len(row) else None).strip().upper()
            if buyer == "NK":
                dest = "NK"
            elif buyer == "PG":
                dest = "Patagonia"
            else:
                _emit(log_emit, f"[VN] Skipped invoice {inv}: unsupported BUYER='{buyer}'")
                continue

            exfty = parse_date_any(row[exfty_i] if exfty_i < len(row) else None)
            amt = parse_money(row[amt_i] if amt_i < len(row) else None)
            if not exfty:
                _emit(log_emit, f"[VN] Skipped invoice {inv}: missing/invalid ACT. EX-FTY")
                continue
            if amt is None:
                _emit(log_emit, f"[VN] Skipped invoice {inv}: missing/invalid ACT. AMOUNT")
                continue

            out.append(ExportRecord(dest_sheet=dest, invoice=inv, amount=float(amt), exfty=exfty, lead_days=45))
        _emit(log_emit, f"[VN] Found {len(out)} row(s) with INV #. (Skipped {skipped_term}/{rows_with_inv} not BY T/C)")
        return out
    finally:
        wb.close()


def read_local_records(path: str, sheet_name: str, *, log_emit=None) -> list[ExportRecord]:
    _emit(log_emit, f"[LOCAL] Opening: {path}")
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"[LOCAL] Sheet not found: {sheet_name}")
        ws = wb[sheet_name]
        header_row_idx = find_header_row_by_value(ws, "JOB NO.")
        if not header_row_idx:
            raise ValueError("[LOCAL] Header row not found (cell value 'JOB NO.').")
        header_row = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))

        idx_exfty_all = header_indices(header_row, "EX-FTY")
        idx_amt_all = header_indices(header_row, "AMOUNT")
        idx_inv = header_indices(header_row, "INV #")
        term_i = header_index_prefer_exact(header_row, "TERM")

        if len(idx_exfty_all) < 2 or len(idx_amt_all) < 2 or not idx_inv or term_i is None:
            raise ValueError("[LOCAL] Missing headers: INV #, TERM, and 2nd EX-FTY + 2nd AMOUNT")

        exfty_i = idx_exfty_all[1]
        amt_i = idx_amt_all[1]
        inv_i = idx_inv[0]

        out: list[ExportRecord] = []
        skipped_term = 0
        rows_with_inv = 0
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            inv = normalize_invoice(row[inv_i] if inv_i < len(row) else None)
            if not inv:
                continue
            rows_with_inv += 1

            term_val = _norm_str(row[term_i] if term_i < len(row) else None)
            term_norm = re.sub(r"\s+", "", term_val.strip().upper())
            if term_norm != "BYTC":
                skipped_term += 1
                continue

            exfty = parse_date_any(row[exfty_i] if exfty_i < len(row) else None)
            amt = parse_money(row[amt_i] if amt_i < len(row) else None)
            if not exfty:
                _emit(log_emit, f"[LOCAL] Skipped invoice {inv}: missing/invalid EX-FTY")
                continue
            if amt is None:
                _emit(log_emit, f"[LOCAL] Skipped invoice {inv}: missing/invalid AMOUNT")
                continue
            out.append(ExportRecord(dest_sheet="NK Local Export", invoice=inv, amount=float(amt), exfty=exfty, lead_days=21))

        _emit(
            log_emit,
            f"[LOCAL] Found {len(out)} row(s) with INV #. (Skipped {skipped_term}/{rows_with_inv} not BY TC)",
        )
        return out
    finally:
        wb.close()


def scan_existing_invoices(wb) -> set[str]:
    seen: set[str] = set()
    for ws in wb.worksheets:
        max_row = ws.max_row or 0
        for r in range(1, max_row + 1):
            inv = normalize_invoice(ws.cell(r, 1).value)
            if not inv:
                continue
            inv_u = inv.strip().upper()
            if inv_u in {"INV #", "INV#", "INVOICE", "INVOICE #"}:
                continue
            seen.add(inv)
    return seen


def find_last_total_row(ws, *, amount_col: int = 2) -> Optional[int]:
    max_row = ws.max_row or 0
    # Prefer a SUM formula (matches the Export Bill "total row" behavior).
    for r in range(max_row, 0, -1):
        v = ws.cell(r, amount_col).value
        if isinstance(v, str):
            s = v.strip()
            if s.startswith("=") and "SUM" in s.upper():
                return r
    for r in range(max_row, 0, -1):
        cell = ws.cell(r, amount_col)
        v = cell.value
        if isinstance(v, str) and v.strip().startswith("="):
            return r
        if getattr(cell, "data_type", None) == "f":
            return r
    return None


def find_last_group_start(ws, total_row: int, *, key_cols: tuple[int, int] = (1, 2)) -> int:
    r = total_row - 1
    while r > 1:
        a = ws.cell(r, key_cols[0]).value
        b = ws.cell(r, key_cols[1]).value
        if (a is None or str(a).strip() == "") and (b is None or str(b).strip() == ""):
            return r + 1
        r -= 1
    return 2


def copy_row_style(ws, src_row: int, dst_row: int, *, min_col: int = 1, max_col: int = 12) -> None:
    from copy import copy as copy_style

    try:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    except Exception:
        pass

    for col in range(min_col, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if not src.has_style:
            continue
        dst.font = copy_style(src.font)
        dst.fill = copy_style(src.fill)
        dst.border = copy_style(src.border)
        dst.alignment = copy_style(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy_style(src.protection)


def insert_export_bill_record(ws, rec: ExportRecord, *, log_emit=None) -> int:
    total_row = find_last_total_row(ws)
    if total_row is None:
        total_row = (ws.max_row or 0) + 1
    ws.insert_rows(total_row, 1)

    new_row = total_row
    template_row = max(new_row - 1, 1)
    copy_row_style(ws, template_row, new_row, min_col=1, max_col=12)

    ws.cell(new_row, 1).value = rec.invoice

    c_amt = ws.cell(new_row, 2)
    c_amt.value = float(rec.amount)

    c_date = ws.cell(new_row, 3)
    c_date.value = rec.exfty

    c_d = ws.cell(new_row, 4)
    c_d.value = f"=C{new_row}+{int(rec.lead_days)}"

    _emit(log_emit, f"[EXPORT BILL] Inserted {rec.invoice} into '{ws.title}' at row {new_row}.")
    return new_row


def recalc_last_group_total(ws, *, amount_col: int = 2, log_emit=None) -> None:
    total_row = find_last_total_row(ws, amount_col=amount_col)
    if total_row is None:
        return
    start = find_last_group_start(ws, total_row, key_cols=(1, amount_col))
    end = total_row - 1
    if end < start:
        return
    col_letter = "B" if amount_col == 2 else None
    if col_letter is None:
        return
    ws.cell(total_row, amount_col).value = f"=SUM({col_letter}{start}:{col_letter}{end})"
    _emit(log_emit, f"[EXPORT BILL] Recalculated '{ws.title}' total row {total_row}: SUM({col_letter}{start}:{col_letter}{end}).")


def build_invoice_index(wb) -> dict[str, tuple[str, int]]:
    idx: dict[str, tuple[str, int]] = {}
    for ws in wb.worksheets:
        max_row = ws.max_row or 0
        for r in range(1, max_row + 1):
            inv = normalize_invoice(ws.cell(r, 1).value)
            if not inv:
                continue
            inv_u = inv.strip().upper()
            if inv_u in {"INV #", "INV#", "INVOICE", "INVOICE #"}:
                continue
            if inv not in idx:
                idx[inv] = (ws.title, r)
    return idx


def find_value_right_of_label(ws, label: str) -> Optional[object]:
    target = label.strip().upper()
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            if str(v).strip().upper() == target:
                for cc in range(c + 1, max_col + 1):
                    vv = ws.cell(r, cc).value
                    if vv is not None and str(vv).strip() != "":
                        return vv
                return None
    return None


def find_payment_ref(ws) -> Optional[str]:
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    pat = re.compile(r"^Payment\s*-\s*(.+)$", re.IGNORECASE)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            m = pat.match(str(v).strip())
            if m:
                return m.group(1).strip()
    return None


def read_trade_card(path: str, *, log_emit=None) -> tuple[date, Optional[str], list[tuple[str, Optional[float]]]]:
    _emit(log_emit, f"[TRADE CARD] Opening: {path}")
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if not wb.worksheets:
            raise ValueError("[TRADE CARD] No sheets found.")

        ws1 = wb.worksheets[0]
        value_date_raw = find_value_right_of_label(ws1, "Value Date")
        value_date = parse_date_any(value_date_raw)
        if not value_date:
            raise ValueError("[TRADE CARD] 'Value Date' not found or invalid.")

        payment_num = find_payment_ref(ws1)
        if payment_num:
            _emit(log_emit, f"[TRADE CARD] Payment ref found: {payment_num}")
        else:
            _emit(log_emit, "[TRADE CARD] Payment ref not found (cell 'Payment - ...').")

        if len(wb.worksheets) < 3:
            raise ValueError("[TRADE CARD] Third sheet not found (expected invoice/amount listing).")
        ws3 = wb.worksheets[2]

        entries: list[tuple[str, Optional[float]]] = []
        for row in ws3.iter_rows(min_row=2, values_only=True):
            inv = normalize_invoice(row[0] if len(row) > 0 else None)
            if not inv:
                continue
            amt = parse_money(row[1] if len(row) > 1 else None)
            entries.append((inv, amt))

        _emit(log_emit, f"[TRADE CARD] Found {len(entries)} invoice row(s).")
        return value_date, payment_num, entries
    finally:
        wb.close()


def update_export_bill_from_trade_card(
    export_wb,
    value_date: date,
    payment_num: Optional[str],
    entries: list[tuple[str, Optional[float]]],
    holidays: set[date],
    invoice_index: Optional[dict[str, tuple[str, int]]] = None,
    mismatches: Optional[list[AmountMismatch]] = None,
    trade_card_file: str = "",
    *,
    log_emit=None,
) -> tuple[int, int, int]:
    if invoice_index is None:
        invoice_index = build_invoice_index(export_wb)
    updated = 0
    missing = 0
    mismatched = 0

    pay_value = f"TC-{payment_num}" if payment_num else None
    next_bd = next_business_day(value_date, holidays)

    for inv, tc_amt in entries:
        loc = invoice_index.get(inv)
        if not loc:
            _emit(log_emit, f"[TRADE CARD] Invoice not found in Export Bill: {inv}")
            missing += 1
            continue
        sheet_name, row = loc
        ws = export_wb[sheet_name]

        eb_amt = parse_money(ws.cell(row, 2).value)
        if tc_amt is not None and eb_amt is not None:
            if abs(float(tc_amt) - float(eb_amt)) > 0.01:
                _emit(log_emit, f"[AMOUNT MISMATCH] {inv} ({sheet_name} row {row}): TradeCard={tc_amt} vs ExportBill={eb_amt}")
                mismatched += 1
                if mismatches is not None:
                    mismatches.append(
                        AmountMismatch(
                            trade_card_file=trade_card_file,
                            invoice=inv,
                            export_bill_sheet=sheet_name,
                            export_bill_row=row,
                            trade_card_amount=float(tc_amt),
                            export_bill_amount=float(eb_amt),
                        )
                    )

        for col in (5, 6):
            c = ws.cell(row, col)
            c.value = value_date

        c_g = ws.cell(row, 7)
        c_g.value = next_bd

        if pay_value:
            ws.cell(row, 10).value = pay_value

        ws.cell(row, 12).value = f"=D{row}-G{row}"
        updated += 1

    return updated, missing, mismatched


def write_mismatch_log_txt(output_xlsx_path: str, export_wb, mismatches: list[AmountMismatch], *, log_emit=None) -> str:
    if not mismatches:
        return ""

    base, _ext = os.path.splitext(output_xlsx_path)
    out_path = ensure_unique_path(f"{base}_amount_mismatches.txt")

    # Build final invoice->row map per sheet (after sorting/regrouping).
    sheet_inv_row: dict[tuple[str, str], int] = {}
    try:
        for ws in export_wb.worksheets:
            max_row = ws.max_row or 0
            for r in range(1, max_row + 1):
                inv = normalize_invoice(ws.cell(r, 1).value)
                if not inv:
                    continue
                inv_u = inv.strip().upper()
                if inv_u in {"INV #", "INV#", "INVOICE", "INVOICE #", "INVOICE NO", "INVOICE NO."}:
                    continue
                key = (ws.title, inv)
                if key not in sheet_inv_row:
                    sheet_inv_row[key] = r
    except Exception:
        sheet_inv_row = {}

    try:
        with open(out_path, "w", encoding="utf-8") as f:
            f.write("Export Bill Sorter - Amount Mismatches\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Output workbook: {output_xlsx_path}\n")
            f.write("\n")
            f.write("Columns: trade_card_file | invoice | sheet | row | trade_card_amount | export_bill_amount\n")
            for m in mismatches:
                row = sheet_inv_row.get((m.export_bill_sheet, m.invoice), m.export_bill_row)
                f.write(
                    f"{m.trade_card_file} | {m.invoice} | {m.export_bill_sheet} | {row} | "
                    f"{m.trade_card_amount} | {m.export_bill_amount}\n"
                )
        _emit(log_emit, f"[MISMATCH LOG] Saved: {out_path}")
        return out_path
    except Exception as e:
        _emit(log_emit, f"[MISMATCH LOG] Failed to write mismatch log: {e}")
        return ""


class ProcessingLogic:
    def __init__(self, log_emit=None):
        self.log_emit = log_emit

    def run(
        self,
        vn_weekly_paths: list[str],
        local_weekly_paths: list[str],
        export_bill_path: str,
        trade_card_paths: list[str],
        feac_chart_path: str,
        *,
        vn_year: int,
        vn_month_abbrev: str,
        vn_week: int,
        local_year: int,
        local_month_abbrev: str,
        local_week: int,
        holidays: set[date],
    ) -> tuple[str, int, int, int, int, int, str]:
        ensure_olefile_available(self.log_emit)
        vn_sheet_name = format_sheet_name(vn_month_abbrev, vn_year, vn_week)
        local_sheet_name = format_sheet_name(local_month_abbrev, local_year, local_week)
        _emit(self.log_emit, f"[INPUT] VN sheet name: {vn_sheet_name}")
        _emit(self.log_emit, f"[INPUT] Local sheet name: {local_sheet_name}")

        records: list[ExportRecord] = []
        for p in vn_weekly_paths:
            records.extend(read_vn_records(p, vn_sheet_name, log_emit=self.log_emit))
        for p in local_weekly_paths:
            records.extend(read_local_records(p, local_sheet_name, log_emit=self.log_emit))

        ref_order = read_feac_ref_order(feac_chart_path, log_emit=self.log_emit)

        keep_vba = export_bill_path.lower().endswith(".xlsm")
        _emit(self.log_emit, f"[EXPORT BILL] Opening: {export_bill_path}")
        export_wb = load_workbook(export_bill_path, keep_vba=keep_vba)
        try:
            existing = scan_existing_invoices(export_wb)
            inserted = 0
            skipped_existing = 0

            touched_sheets: set[str] = set()
            for rec in records:
                if rec.dest_sheet not in export_wb.sheetnames:
                    raise ValueError(f"[EXPORT BILL] Sheet not found: {rec.dest_sheet}")
                if rec.invoice in existing:
                    skipped_existing += 1
                    _emit(self.log_emit, f"[EXPORT BILL] Skipped existing invoice: {rec.invoice}")
                    continue
                ws = export_wb[rec.dest_sheet]
                insert_export_bill_record(ws, rec, log_emit=self.log_emit)
                existing.add(rec.invoice)
                inserted += 1
                touched_sheets.add(rec.dest_sheet)

            for sheet in sorted(touched_sheets):
                recalc_last_group_total(export_wb[sheet], log_emit=self.log_emit)

            invoice_index = build_invoice_index(export_wb)
            updated_total = 0
            missing_total = 0
            mismatched_total = 0
            mismatch_rows: list[AmountMismatch] = []
            for tc_path in trade_card_paths:
                value_date, payment_num, tc_entries = read_trade_card(tc_path, log_emit=self.log_emit)
                updated, missing, mismatched = update_export_bill_from_trade_card(
                    export_wb,
                    value_date,
                    payment_num,
                    tc_entries,
                    holidays,
                    invoice_index,
                    mismatch_rows,
                    os.path.basename(tc_path),
                    log_emit=self.log_emit,
                )
                updated_total += updated
                missing_total += missing
                mismatched_total += mismatched

            # Regroup/sort rows by Value Date (col E) then FEAC order of Payment No (col J)
            for ws in export_wb.worksheets:
                regroup_and_sort_export_bill_sheet(ws, ref_order, log_emit=self.log_emit)

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            base, ext = os.path.splitext(export_bill_path)
            safe_vn = re.sub(r"[^A-Za-z0-9_-]+", "_", vn_sheet_name)
            safe_local = re.sub(r"[^A-Za-z0-9_-]+", "_", local_sheet_name)
            out_path = ensure_unique_path(f"{base}_{safe_vn}_{safe_local}_{ts}{ext}")
            _emit(self.log_emit, f"[SAVE] Writing output: {out_path}")
            export_wb.save(out_path)
            mismatch_log_path = write_mismatch_log_txt(out_path, export_wb, mismatch_rows, log_emit=self.log_emit)
            return out_path, inserted, skipped_existing, updated_total, missing_total, mismatched_total, mismatch_log_path
        finally:
            export_wb.close()


class MainWidget(QWidget):
    log_message = Signal(str)
    processing_done = Signal(bool, str)

    def __init__(self):
        super().__init__()
        self.setObjectName("export_bill_sorter_widget")
        self._build_ui()
        self._connect()
        self.logic = ProcessingLogic(self.log_message.emit)

    def _build_ui(self):
        self.desc_label = QLabel("Export Bill Sorter (VN/Local Weekly Export + Trade Card)", self)
        self.desc_label.setWordWrap(True)
        self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.desc_label.setStyleSheet("color: #dcdcdc; font-weight: bold; padding: 6px;")

        shared_box_style = (
            "QTextEdit{background:#1f1f1f; color:#d0d0d0; border:1px solid #3a3a3a; border-radius:6px;}"
        )
        label_style = "color:#dcdcdc; background:transparent;"

        self.vn_btn = PrimaryPushButton("Select VN Weekly Export Chart(s)", self)
        self.vn_box = QTextEdit(self)
        self.vn_box.setReadOnly(True)
        self.vn_box.setMaximumHeight(80)
        self.vn_box.setStyleSheet(shared_box_style)

        self.local_btn = PrimaryPushButton("Select Local Weekly Export Chart(s)", self)
        self.local_box = QTextEdit(self)
        self.local_box.setReadOnly(True)
        self.local_box.setMaximumHeight(80)
        self.local_box.setStyleSheet(shared_box_style)

        self.export_btn = PrimaryPushButton("Select Export Bill.xlsx", self)
        self.export_box = QTextEdit(self)
        self.export_box.setReadOnly(True)
        self.export_box.setMaximumHeight(42)
        self.export_box.setStyleSheet(shared_box_style)

        self.trade_btn = PrimaryPushButton("Select Trade Card(s)", self)
        self.trade_box = QTextEdit(self)
        self.trade_box.setReadOnly(True)
        self.trade_box.setMaximumHeight(80)
        self.trade_box.setStyleSheet(shared_box_style)

        self.feac_btn = PrimaryPushButton("Select Foreign Exchange Administrative Control Chart", self)
        self.feac_box = QTextEdit(self)
        self.feac_box.setReadOnly(True)
        self.feac_box.setMaximumHeight(42)
        self.feac_box.setStyleSheet(shared_box_style)

        # Inputs rows: VN and Local Year / Month / Week (separate)
        years: list[str] = []
        current_year = datetime.now().year
        years = [str(y) for y in range(1995, current_year + 6)]

        self.vn_sheet_label = QLabel("VN Sheet:", self)
        self.vn_sheet_label.setStyleSheet(label_style)
        self.vn_year_label = QLabel("Year:", self)
        self.vn_year_label.setStyleSheet(label_style)
        self.vn_year_combo = ComboBox(self)
        self.vn_year_combo.addItems(years)
        self.vn_year_combo.setCurrentText(str(current_year))
        self.vn_month_label = QLabel("Month:", self)
        self.vn_month_label.setStyleSheet(label_style)
        self.vn_month_combo = ComboBox(self)
        self.vn_month_combo.addItems(MONTHS)
        self.vn_month_combo.setCurrentIndex(max(0, datetime.now().month - 1))
        self.vn_week_label = QLabel("Week:", self)
        self.vn_week_label.setStyleSheet(label_style)
        self.vn_week_combo = ComboBox(self)
        self.vn_week_combo.addItems([str(i) for i in range(1, 7)])
        self.vn_week_combo.setCurrentIndex(0)

        self.local_sheet_label = QLabel("Local Sheet:", self)
        self.local_sheet_label.setStyleSheet(label_style)
        self.local_year_label = QLabel("Year:", self)
        self.local_year_label.setStyleSheet(label_style)
        self.local_year_combo = ComboBox(self)
        self.local_year_combo.addItems(years)
        self.local_year_combo.setCurrentText(str(current_year))
        self.local_month_label = QLabel("Month:", self)
        self.local_month_label.setStyleSheet(label_style)
        self.local_month_combo = ComboBox(self)
        self.local_month_combo.addItems(MONTHS)
        self.local_month_combo.setCurrentIndex(max(0, datetime.now().month - 1))
        self.local_week_label = QLabel("Week:", self)
        self.local_week_label.setStyleSheet(label_style)
        self.local_week_combo = ComboBox(self)
        self.local_week_combo.addItems([str(i) for i in range(1, 7)])
        self.local_week_combo.setCurrentIndex(0)

        self.holidays_label = QLabel("MY Public Holidays (one per line, YYYY-MM-DD)", self)
        self.holidays_label.setStyleSheet(label_style)
        self.holidays_box = QTextEdit(self)
        self.holidays_box.setStyleSheet(shared_box_style)
        self.holidays_box.setPlaceholderText("2026-01-01\n2026-02-17\n...")
        self.holidays_box.setPlainText("\n".join(DEFAULT_MY_HOLIDAYS_2026))
        self.holidays_box.setMaximumHeight(120)

        self.run_btn = PrimaryPushButton("Run", self)
        self.run_btn.setEnabled(False)

        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setStyleSheet(shared_box_style)
        self.log_box.setPlaceholderText("Process logs…")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        layout.addWidget(self.desc_label)

        def add_row(btn, box):
            row = QHBoxLayout()
            row.addWidget(btn, 2)
            row.addWidget(box, 3)
            layout.addLayout(row)

        add_row(self.vn_btn, self.vn_box)
        vn_inputs_row = QHBoxLayout()
        vn_inputs_row.addWidget(self.vn_sheet_label)
        vn_inputs_row.addSpacing(8)
        vn_inputs_row.addWidget(self.vn_year_label)
        vn_inputs_row.addWidget(self.vn_year_combo)
        vn_inputs_row.addSpacing(10)
        vn_inputs_row.addWidget(self.vn_month_label)
        vn_inputs_row.addWidget(self.vn_month_combo)
        vn_inputs_row.addSpacing(10)
        vn_inputs_row.addWidget(self.vn_week_label)
        vn_inputs_row.addWidget(self.vn_week_combo)
        vn_inputs_row.addStretch(1)
        layout.addLayout(vn_inputs_row)

        add_row(self.local_btn, self.local_box)
        local_inputs_row = QHBoxLayout()
        local_inputs_row.addWidget(self.local_sheet_label)
        local_inputs_row.addSpacing(8)
        local_inputs_row.addWidget(self.local_year_label)
        local_inputs_row.addWidget(self.local_year_combo)
        local_inputs_row.addSpacing(10)
        local_inputs_row.addWidget(self.local_month_label)
        local_inputs_row.addWidget(self.local_month_combo)
        local_inputs_row.addSpacing(10)
        local_inputs_row.addWidget(self.local_week_label)
        local_inputs_row.addWidget(self.local_week_combo)
        local_inputs_row.addStretch(1)
        layout.addLayout(local_inputs_row)

        add_row(self.export_btn, self.export_box)
        add_row(self.trade_btn, self.trade_box)
        add_row(self.feac_btn, self.feac_box)

        layout.addWidget(self.holidays_label)
        layout.addWidget(self.holidays_box)
        layout.addWidget(self.run_btn)
        layout.addWidget(self.log_box, 1)

    def _connect(self):
        self.vn_btn.clicked.connect(lambda: self._pick_files(self.vn_box, "Select VN Weekly Export Chart(s)"))
        self.local_btn.clicked.connect(lambda: self._pick_files(self.local_box, "Select Local Weekly Export Chart(s)"))
        self.export_btn.clicked.connect(lambda: self._pick_file(self.export_box, "Select Export Bill"))
        self.trade_btn.clicked.connect(lambda: self._pick_files(self.trade_box, "Select Trade Card(s)"))
        self.feac_btn.clicked.connect(lambda: self._pick_file(self.feac_box, "Select Foreign Exchange Administrative Control Chart"))

        self.log_message.connect(self._append_log)
        self.processing_done.connect(self._on_done)

        for box in (self.vn_box, self.local_box, self.export_box, self.trade_box, self.feac_box):
            box.textChanged.connect(self._check_ready)

        self.run_btn.clicked.connect(self._run)

    def _pick_file(self, target_box: QTextEdit, title: str):
        path, _ = QFileDialog.getOpenFileName(self, title, "", "Excel Files (*.xlsx *.xlsm)")
        if path:
            target_box.setText(path)
        self._check_ready()

    def _pick_files(self, target_box: QTextEdit, title: str):
        paths, _ = QFileDialog.getOpenFileNames(self, title, "", "Excel Files (*.xlsx *.xlsm)")
        if paths:
            target_box.setText("\n".join(paths))
        self._check_ready()

    def _check_ready(self):
        def has_any_line(box: QTextEdit) -> bool:
            return any(line.strip() for line in (box.toPlainText() or "").splitlines())

        ready = all(has_any_line(b) for b in (self.vn_box, self.local_box, self.export_box, self.trade_box, self.feac_box))
        self.run_btn.setEnabled(bool(ready))

    def _append_log(self, text: str):
        self.log_box.append(text)
        self.log_box.ensureCursorVisible()

    def _set_enabled_inputs(self, enabled: bool):
        for w in (
            self.vn_btn,
            self.local_btn,
            self.export_btn,
            self.trade_btn,
            self.feac_btn,
            self.vn_year_combo,
            self.vn_month_combo,
            self.vn_week_combo,
            self.local_year_combo,
            self.local_month_combo,
            self.local_week_combo,
            self.holidays_box,
            self.vn_box,
            self.local_box,
            self.export_box,
            self.trade_box,
            self.feac_box,
        ):
            w.setEnabled(enabled)
        self.run_btn.setEnabled(enabled and self.run_btn.isEnabled())

    def _run(self):
        def split_lines(text: str) -> list[str]:
            return [line.strip() for line in (text or "").splitlines() if line.strip()]

        vn_paths = split_lines(self.vn_box.toPlainText())
        local_paths = split_lines(self.local_box.toPlainText())
        export_lines = split_lines(self.export_box.toPlainText())
        trade_paths = split_lines(self.trade_box.toPlainText())
        feac_lines = split_lines(self.feac_box.toPlainText())

        if len(export_lines) != 1:
            MessageBox("Invalid Export Bill", "Please select exactly one Export Bill file.", self).exec()
            return
        export_bill = export_lines[0]

        if len(feac_lines) != 1:
            MessageBox("Invalid FEAC chart", "Please select exactly one Foreign Exchange Administrative Control Chart file.", self).exec()
            return
        feac_chart = feac_lines[0]

        for paths, label in (
            (vn_paths, "VN Weekly Export Chart(s)"),
            (local_paths, "Local Weekly Export Chart(s)"),
            ([export_bill], "Export Bill"),
            (trade_paths, "Trade Card(s)"),
            ([feac_chart], "Foreign Exchange Administrative Control Chart"),
        ):
            if not paths:
                MessageBox("Missing file", f"Please select at least one file for: {label}", self).exec()
                return
            for p in paths:
                if not os.path.isfile(p):
                    MessageBox("Missing file", f"Invalid file for {label}:\n{p}", self).exec()
                    return

        try:
            vn_year = int(self.vn_year_combo.currentText())
            vn_month = self.vn_month_combo.currentText()
            vn_week = int(self.vn_week_combo.currentText())

            local_year = int(self.local_year_combo.currentText())
            local_month = self.local_month_combo.currentText()
            local_week = int(self.local_week_combo.currentText())
        except Exception:
            MessageBox("Invalid input", "Please check VN/Local Year/Month/Week inputs.", self).exec()
            return

        self.log_box.clear()
        self.log_message.emit(
            f"Starting… (VN files: {len(vn_paths)}, Local files: {len(local_paths)}, Trade cards: {len(trade_paths)})"
        )
        self._set_enabled_inputs(False)
        self.run_btn.setEnabled(False)

        def worker():
            try:
                holidays = parse_holiday_lines(self.holidays_box.toPlainText(), log_emit=self.log_message.emit)
                out_path, inserted, skipped, updated, missing, mismatched, mismatch_log_path = self.logic.run(
                    vn_paths,
                    local_paths,
                    export_bill,
                    trade_paths,
                    feac_chart,
                    vn_year=vn_year,
                    vn_month_abbrev=vn_month,
                    vn_week=vn_week,
                    local_year=local_year,
                    local_month_abbrev=local_month,
                    local_week=local_week,
                    holidays=holidays,
                )
                msg = (
                    "Done!\n\n"
                    + f"Output: {out_path}\n"
                    + (f"Mismatch log: {mismatch_log_path}\n" if mismatch_log_path else "")
                    + f"Inserted new invoices: {inserted}\n"
                    + f"Skipped existing invoices: {skipped}\n"
                    + f"Trade card updated invoices: {updated}\n"
                    + f"Trade card missing invoices: {missing}\n"
                    + f"Amount mismatches: {mismatched}"
                )
                self.processing_done.emit(True, msg)
            except Exception as e:
                self.log_message.emit(f"ERROR: {e}")
                self.processing_done.emit(False, str(e))

        threading.Thread(target=worker, daemon=True).start()

    def _on_done(self, ok: bool, message: str):
        self._set_enabled_inputs(True)
        self._check_ready()
        title = "Success" if ok else "Failed"
        MessageBox(title, message, self).exec()


def get_widget():
    return MainWidget()


if __name__ == "__main__":
    from PySide6.QtWidgets import QApplication
    import sys

    app = QApplication(sys.argv)
    w = MainWidget()
    w.show()
    sys.exit(app.exec())
