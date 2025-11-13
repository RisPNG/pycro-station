import threading
from typing import List, Tuple, Dict, Any
from PySide6.QtCore import Signal
from PySide6.QtWidgets import QWidget, QVBoxLayout, QTextEdit, QFileDialog
from qfluentwidgets import PrimaryPushButton, MessageBox
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import column_index_from_string, get_column_letter
import os

try:
    import xlrd  # for legacy .xls
except Exception:
    xlrd = None


class MainWidget(QWidget):
    log_message = Signal(str)
    processing_done = Signal(int, int)

    def __init__(self):
        super().__init__()
        self.setObjectName('test_macro_one_widget')
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(12)

        self.select_btn = PrimaryPushButton('Select Files...', self)
        self.select_btn.clicked.connect(self.select_files)

        self.files_box = QTextEdit(self)
        self.files_box.setReadOnly(True)
        self.files_box.setPlaceholderText('Selected files will appear here')
        self.files_box.setStyleSheet('QTextEdit{background: #2a2a2a; color: white; border: 1px solid #3a3a3a; border-radius: 6px;}')

        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText('Live process log will appear here')
        self.log_box.setStyleSheet('QTextEdit{background: #1f1f1f; color: #d0d0d0; border: 1px solid #3a3a3a; border-radius: 6px;}')

        self.run_btn = PrimaryPushButton('Run', self)
        self.run_btn.clicked.connect(self.run_process)

        layout.addWidget(self.select_btn)
        layout.addWidget(self.files_box, 1)
        layout.addWidget(self.log_box, 1)
        layout.addWidget(self.run_btn)

        self.log_message.connect(self.append_log)
        self.processing_done.connect(self.on_processing_done)

    def select_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, 'Select files')
        if files:
            self.files_box.setPlainText('\n'.join(files))
        else:
            self.files_box.clear()

    def _selected_files(self):
        text = self.files_box.toPlainText().strip()
        if not text:
            return []
        return [line for line in text.split('\n') if line.strip()]

    def run_process(self):
        files = self._selected_files()
        if not files:
            MessageBox('No files', 'Please select one or more Excel files first.', self).exec()
            return

        self.log_box.clear()
        self.log_message.emit(f'Starting processing of {len(files)} file(s)...')
        self.run_btn.setEnabled(False)
        self.select_btn.setEnabled(False)

        def worker():
            ok, fail = 0, 0
            for path in files:
                try:
                    self.log_message.emit(f'Opening: {path}')
                    success, out_path = _safe_process_with_output(path, self.log_message.emit)
                    if success:
                        if out_path and out_path != path:
                            self.log_message.emit(f'Processed and saved as: {out_path}')
                        else:
                            self.log_message.emit(f'Processed and saved: {path}')
                        ok += 1
                    else:
                        pass
                except Exception as e:
                    self.log_message.emit(f'ERROR processing {path}: {e}')
                    fail += 1
            self.processing_done.emit(ok, fail)

        threading.Thread(target=worker, daemon=True).start()

    def append_log(self, text: str):
        self.log_box.append(text)
        self.log_box.ensureCursorVisible()

    def on_processing_done(self, ok: int, fail: int):
        self.log_message.emit(f'Completed: {ok} success, {fail} failed.')
        self.run_btn.setEnabled(True)
        self.select_btn.setEnabled(True)


def get_widget():
    return MainWidget()


# ========================
# FAL SORTER (VBA parity)
# ========================

def _clean_text(input_text: str) -> str:
    """Remove various whitespace characters including spaces, line breaks, carriage returns, tabs"""
    if input_text is None:
        return ''
    s = str(input_text)
    s = s.replace('\n', '').replace('\r', '')
    s = s.replace('\t', ' ')
    s = s.replace('\xa0', ' ')
    s = s.strip()
    while '  ' in s:
        s = s.replace('  ', ' ')
    return s


def _is_serial_row(row_vals: List[Any], last_col: int) -> bool:
    """Check if row is a serial number row (Column A contains 'Serial Number' and all others empty)"""
    a = row_vals[0]
    if a is None:
        return False
    if 'serial number' in str(a).lower():
        for j in range(1, last_col):
            if row_vals[j] not in (None, ''):
                return False
        return True
    return False


def _extract_serial(s: str) -> str:
    """Extract serial number from string (text after ':')"""
    if s is None:
        return ''
    txt = str(s)
    p = txt.find(':')
    if p >= 0:
        return txt[p + 1:].strip()
    return txt.strip()


def _is_report_total_row(row_vals: List[Any]) -> bool:
    """Check if row is a report total row"""
    t = str(row_vals[0]).strip().lower() if row_vals and row_vals[0] is not None else ''
    return ('total' in t) or ('report total' in t)


def _auto_fit_columns(ws: Worksheet):
    """Auto-fit columns and freeze top row"""
    dims: Dict[int, int] = {}
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    for r in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for c_idx, cell in enumerate(r, start=1):
            v = cell.value
            if v is None:
                l = 0
            else:
                s = str(v)
                l = len(s) + (1 if isinstance(v, (int, float)) else 0)
            if l:
                dims[c_idx] = max(dims.get(c_idx, 0), l)

    for c_idx, w in dims.items():
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(w + 2, 8), 60)

    ws.freeze_panes = 'A2'


def _delete_columns(ws: Worksheet):
    """Delete columns R, O, J, I, H, G, E in correct order to avoid shifting issues"""
    letters = ['R', 'O', 'J', 'I', 'H', 'G', 'E']
    indices = []

    for col_letter in letters:
        try:
            idx = column_index_from_string(col_letter)
            if idx <= (ws.max_column or 0):
                indices.append(idx)
        except Exception:
            pass

    indices.sort(reverse=True)
    for idx in indices:
        try:
            ws.delete_cols(idx, 1)
        except Exception:
            pass


def _format_column_L(ws: Worksheet):
    """Left align column L and set number format to prevent scientific notation"""
    max_row = ws.max_row or 1
    col_idx = column_index_from_string('L')
    for r in range(1, max_row + 1):
        c = ws.cell(row=r, column=col_idx)
        c.alignment = Alignment(horizontal='left', vertical='top')
        c.number_format = '0'


def _read_source(ws: Worksheet, last_row: int, last_col: int) -> List[List[Any]]:
    """Read all source data into array"""
    data: List[List[Any]] = []
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_col, values_only=True):
        data.append(list(row))
    if not data:
        data = [[None] * last_col]
    return data


def _copy_row(source: List[List[Any]], src_row: int, last_col: int) -> List[Any]:
    """Copy a single row from source data"""
    row = [None] * last_col
    for j in range(last_col):
        row[j] = source[src_row][j] if j < len(source[src_row]) else None
    return row


def _process_sheet(source_data: List[List[Any]], target_ws: Worksheet, last_row: int, last_col: int, sheet_type: str):
    """Process and filter sheet data based on sheet type (PRAI/VIETNAM/ALL)"""
    if target_ws.max_row and target_ws.max_row > 1:
        try:
            target_ws.delete_rows(2, target_ws.max_row - 1)
        except Exception:
            pass

    filtered: List[List[Any]] = [[None for _ in range(last_col)] for _ in range(max(1, last_row))]

    if source_data:
        for j in range(last_col):
            filtered[0][j] = source_data[0][j] if j < len(source_data[0]) else None

    filtered_count = 1

    for i in range(1, last_row):
        included = False
        col_e = source_data[i][4] if last_col >= 5 and i < len(source_data) else None

        if col_e not in (None, ''):
            s = str(col_e)
            if len(s) >= 2:
                first_two = s[:2]
                if sheet_type == 'VIETNAM':
                    included = (first_two == '03')
                elif sheet_type == 'PRAI':
                    included = (first_two != '03')
                elif sheet_type == 'ALL':
                    included = True

        if included:
            if i < len(source_data):
                filtered[filtered_count] = _copy_row(source_data, i, last_col)
                filtered_count += 1

            if (i + 1) < last_row and (i + 1) < len(source_data):
                a1 = source_data[i + 1][0]
                if a1 not in (None, '') and ('serial number' in str(a1).lower()):
                    filtered[filtered_count] = _copy_row(source_data, i + 1, last_col)
                    filtered_count += 1

    if filtered_count > 1:
        _categorize_and_write_data(target_ws, filtered, filtered_count, last_col)
    else:
        for j in range(last_col):
            hv = filtered[0][j]
            target_ws.cell(row=1, column=j + 1, value=_clean_text(str(hv)) if hv not in (None, '') else hv)


def _format_dates_in_worksheet(ws: Worksheet):
    """Format all date cells to dd/mm/yyyy format"""
    from datetime import datetime
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            # Check if cell contains a date
            if isinstance(cell.value, datetime):
                cell.number_format = 'DD/MM/YYYY'


def _categorize_and_write_data(ws: Worksheet, data_array: List[List[Any]], data_rows: int, last_col: int):
    """Categorize data and write to worksheet with proper formatting"""
    categories: List[Tuple[str, str]] = [
        ('FL', 'Freehold Land'),
        ('LL', 'Leasehold Land'),
        ('BD', 'Building'),
        ('RN', 'Renovation'),
        ('EI', 'Electricity Item'),
        ('PM', 'Plant and Machinery'),
        ('FF', 'Furniture and Fitting'),
        ('CD', 'Computer Equipment'),
        ('OE', 'Other Equipment'),
        ('CQ', 'Canteen Equipment'),
        ('MV', 'Motor Vehicle'),
        ('UC', 'Unknown Category'),
    ]

    known_count = 11
    category_data: List[List[int]] = [[] for _ in range(12)]
    serial_map: Dict[int, str] = {}
    prev_data_row = 0

    for i in range(1, data_rows):
        row = data_array[i]

        if _is_report_total_row(row):
            continue

        if _is_serial_row(row, last_col):
            if prev_data_row > 0:
                serial_map[prev_data_row] = _extract_serial(str(row[0]))
            continue

        category_index = 0
        col_d = row[3] if last_col >= 4 else None

        if col_d not in (None, ''):
            d = str(col_d)
            if len(d) >= 4:
                code = d[2:4]
                if code == 'CE':
                    code = 'CD'

                found = 0
                for j in range(known_count):
                    if categories[j][0] == code:
                        category_index = j
                        found = 1
                        break

                if not found:
                    category_index = 11
            else:
                category_index = 11
        else:
            category_index = 11

        category_data[category_index].append(i)
        prev_data_row = i

    out_serial_col = last_col + 2
    current_row = 1

    # Write header
    ws.cell(row=current_row, column=1, value='Asset ID')
    ws.cell(row=current_row, column=2, value='Description')

    for j in range(1, last_col):
        hv = data_array[0][j] if 0 < len(data_array) and j < len(data_array[0]) else None
        ws.cell(row=current_row, column=j + 2, value=_clean_text(str(hv)) if hv not in (None, '') else hv)

    ws.cell(row=current_row, column=out_serial_col, value='Serial Number')
    current_row += 1

    grand_totals: Dict[int, float] = {k: 0.0 for k in range(10, 18)}
    sum_label_col = 10
    sum_start_out_col = 11
    sum_end_out_col = min(last_col + 1, 18)

    for i, cat_rows in enumerate(category_data, start=1):
        if not cat_rows:
            continue

        # Category header
        ws.cell(row=current_row, column=1, value=categories[i - 1][1])
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

        subtotals: Dict[int, float] = {k: 0.0 for k in range(10, 18)}

        for src_idx in cat_rows:
            raw_a = str(data_array[src_idx][0]) if data_array[src_idx][0] is not None else ''
            p = raw_a.find('/')

            if p >= 0:
                asset_id = raw_a[:p].strip()
                desc_text = raw_a[p + 1:].strip()
            else:
                asset_id = raw_a.strip()
                desc_text = ''

            ws.cell(row=current_row, column=1, value=asset_id)
            ws.cell(row=current_row, column=2, value=desc_text)

            for j in range(1, last_col):
                ws.cell(row=current_row, column=j + 2, value=data_array[src_idx][j])

            if src_idx in serial_map:
                ws.cell(row=current_row, column=out_serial_col, value=serial_map[src_idx])

            for k in range(10, 18):
                if k == 10:
                    continue
                if k - 1 < last_col:
                    v = data_array[src_idx][k - 1]
                    if isinstance(v, (int, float)):
                        subtotals[k] += float(v)

            current_row += 1

        # Subtotal row
        ws.cell(row=current_row, column=sum_label_col, value='Subtotal:')
        ws.cell(row=current_row, column=sum_label_col).font = Font(bold=True)

        for k in range(10, 18):
            if k == 10:
                continue
            if (k + 1) <= sum_end_out_col and subtotals[k] != 0:
                c = ws.cell(row=current_row, column=k + 1, value=subtotals[k])
                c.number_format = '#,##0.00'

        # Subtotal formatting
        for c in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=sum_label_col, max_col=sum_end_out_col):
            for cell in c:
                cell.font = Font(bold=True)
                cell.border = Border(top=Side(style='thin'))

        for k in range(10, 18):
            if k != 10:
                grand_totals[k] += subtotals[k]

        current_row += 2

    # Grand Total row
    ws.cell(row=current_row, column=sum_label_col, value='Grand Total:')
    ws.cell(row=current_row, column=sum_label_col).font = Font(bold=True)

    for k in range(10, 18):
        if k == 10:
            continue
        if (k + 1) <= sum_end_out_col and grand_totals[k] != 0:
            c = ws.cell(row=current_row, column=k + 1, value=grand_totals[k])
            c.number_format = '#,##0.00'

    # Grand total formatting
    for c in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=sum_label_col, max_col=sum_end_out_col):
        for cell in c:
            cell.font = Font(bold=True)
            cell.border = Border(top=Side(style='double'), bottom=Side(style='thin'))

    # Apply number formatting to all cells in numeric range
    last_out_row = current_row
    if sum_start_out_col <= sum_end_out_col:
        for r in range(2, last_out_row + 1):
            for c in range(sum_start_out_col, sum_end_out_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.number_format = '#,##0.00'

    # Format all dates in the worksheet to dd/mm/yyyy
    _format_dates_in_worksheet(ws)


def _copy_after(wb, ws_to_move: Worksheet, after_ws: Worksheet):
    """Move worksheet to be placed right after another worksheet"""
    try:
        sheets = wb.worksheets
        tgt_idx = sheets.index(after_ws) + 1
        sheets.remove(ws_to_move)
        sheets.insert(tgt_idx, ws_to_move)
    except Exception:
        pass


def _sheet_names(wb) -> List[str]:
    """Get list of all sheet names in workbook"""
    return [ws.title for ws in wb.worksheets]


def _first_sheet(wb) -> Worksheet:
    """Get first worksheet in workbook"""
    return wb.worksheets[0]


def _max_used(ws: Worksheet) -> Tuple[int, int]:
    """Get maximum used row and column"""
    return ws.max_row or 1, ws.max_column or 1


def _has_processed_sheets(wb) -> bool:
    """Check if workbook already has PRAI/VIETNAM/ALL sheets"""
    names = set(_sheet_names(wb))
    return any(n in names for n in ('PRAI', 'VIETNAM', 'ALL'))


def _prepare_target_sheets(wb, source_ws: Worksheet) -> Tuple[Worksheet, Worksheet, Worksheet]:
    """Create three new sheets by copying the source sheet"""
    prai = wb.copy_worksheet(source_ws)
    prai.title = 'PRAI'
    _copy_after(wb, prai, source_ws)

    vn = wb.copy_worksheet(source_ws)
    vn.title = 'VIETNAM'
    _copy_after(wb, vn, prai)

    all_ws = wb.copy_worksheet(source_ws)
    all_ws.title = 'ALL'
    _copy_after(wb, all_ws, vn)

    return prai, vn, all_ws


def _format_all_three(prai: Worksheet, vn: Worksheet, all_ws: Worksheet):
    """Format all three processed sheets"""
    for ws in (prai, vn, all_ws):
        _auto_fit_columns(ws)
        _delete_columns(ws)
        _format_column_L(ws)


def _process_master(wb):
    """Main processing function equivalent to VBA MasterStart"""
    if _has_processed_sheets(wb):
        raise RuntimeError('This file has already been processed. Delete sheets PRAI/VIETNAM/ALL to re-run.')

    source_ws = _first_sheet(wb)
    prai_ws, vn_ws, all_ws = _prepare_target_sheets(wb, source_ws)

    last_row, last_col = _max_used(source_ws)
    source_data = _read_source(source_ws, last_row, last_col)

    _process_sheet(source_data, prai_ws, last_row, last_col, 'PRAI')
    _process_sheet(source_data, vn_ws, last_row, last_col, 'VIETNAM')
    _process_sheet(source_data, all_ws, last_row, last_col, 'ALL')

    _format_all_three(prai_ws, vn_ws, all_ws)


def _safe_save(wb, path: str):
    """Save workbook to file"""
    wb.save(path)


def _is_xlsm(path: str) -> bool:
    """Check if file is .xlsm format"""
    return path.lower().endswith('.xlsm')


def _load_workbook_any(path: str):
    """Load workbook, preserving VBA if .xlsm"""
    if _is_xlsm(path):
        return load_workbook(path, keep_vba=True, data_only=False)
    return load_workbook(path, data_only=False)


def _is_xls(path: str) -> bool:
    """Check if file is legacy .xls format"""
    return path.lower().endswith('.xls')


def _proposed_xlsx_path(xls_path: str) -> str:
    """Generate output path for converted .xls file"""
    root, _ = os.path.splitext(xls_path)
    candidate = root + '.xlsx'
    if not os.path.exists(candidate):
        return candidate

    n = 1
    while True:
        candidate = f"{root} (converted {n}).xlsx"
        if not os.path.exists(candidate):
            return candidate
        n += 1


def _xlrd_cell_value(book, cell):
    """Map xlrd cell to Python types that openpyxl accepts"""
    from xlrd import XL_CELL_DATE, xldate_as_datetime, XL_CELL_BOOLEAN

    if cell.ctype == XL_CELL_DATE:
        try:
            return xldate_as_datetime(cell.value, book.datemode)
        except Exception:
            return cell.value

    if cell.ctype == XL_CELL_BOOLEAN:
        return bool(cell.value)

    return cell.value


def _load_xls_as_openpyxl_workbook(path: str):
    """Load legacy .xls file and convert to openpyxl workbook"""
    if xlrd is None:
        raise RuntimeError('xlrd is required to read .xls files. Please add xlrd to requirements.')

    book = xlrd.open_workbook(path, formatting_info=False)
    sh = book.sheet_by_index(0)

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sh.name if sh.name else 'Sheet1'

    for r in range(sh.nrows):
        row = sh.row(r)
        for c, cell in enumerate(row):
            ws.cell(row=r + 1, column=c + 1, value=_xlrd_cell_value(book, cell))

    return wb


def _sheet_presence_summary(wb) -> str:
    """Get summary of sheets in workbook"""
    names = ', '.join(_sheet_names(wb))
    return f'sheets=[{names}]'


def _warn_processed(log):
    """Log warning about already processed file"""
    _log_safe(log, 'File already processed (PRAI/VIETNAM/ALL present). Skipping.')


def _open_wb_for_processing(path: str):
    """Open workbook for processing, handling both .xls and .xlsx/.xlsm"""
    if _is_xls(path):
        return _load_xls_as_openpyxl_workbook(path)
    return _load_workbook_any(path)


def _log_safe(log_emit, msg: str):
    """Safely emit log message"""
    try:
        log_emit(msg)
    except Exception:
        pass


def _get_sheet(wb, name: str) -> Worksheet:
    """Get worksheet by name"""
    for s in wb.worksheets:
        if s.title == name:
            return s
    raise KeyError(name)


def _safe_process_with_output(path: str, log) -> Tuple[bool, str]:
    """Process file and return success status and output path"""
    is_xls = _is_xls(path)
    wb = _open_wb_for_processing(path)

    _log_safe(log, _sheet_presence_summary(wb))

    if _has_processed_sheets(wb):
        _warn_processed(log)
        return False, ''

    _process_master(wb)

    out_path = path
    if is_xls:
        out_path = _proposed_xlsx_path(path)
        _log_safe(log, f'.xls detected; saving output as: {out_path}')

    _safe_save(wb, out_path)
    _log_safe(log, f'Processing complete.')

    return True, out_path