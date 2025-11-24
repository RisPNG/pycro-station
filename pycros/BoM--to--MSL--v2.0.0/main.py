import os
import threading
from datetime import datetime
from typing import List, Tuple, Any, Optional

from PySide6.QtCore import Signal
from PySide6.QtWidgets import QWidget, QVBoxLayout, QTextEdit, QFileDialog
from qfluentwidgets import PrimaryPushButton, MessageBox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

try:
    import xlrd  # optional, for legacy .xls input
except Exception:
    xlrd = None

# -----------------------------
# Constants (mirroring the VBA)
# -----------------------------
NORMAL_COL_COUNT = 12

STYLE_LABEL = "STYLE#:"
SEASON_LABEL = "SEASON:"

STYLE_LENGTH = 6
SEASON_LENGTH = 4
ITEM_LENGTH = 7

S_COL_SEQ = 1
S_COL_DESC = 4
S_COL_UOM = 8
S_COL_CW = 9

S_UOM_VALUE = "UOM"

R_START_ROW = 1
R_START_COL = 0

R_COL_SEASON = 7
R_COL_SEQ = 8
R_COL_STYLE_NUMBER = 9
R_COL_STYLE_NAME = 10
R_COL_STYLE_CW = 11
R_COL_VENDOR_CODE = 12
R_COL_VENDOR_NAME = 13
R_COL_VENDOR_MCO = 14
R_COL_ITEM = 15
R_COL_DESC = 16
R_COL_COLOR_CODE = 17
R_COL_COLOR_NAME = 18
R_COL_REMARKS = 24
R_COL_END = 25

HEADER_COLOR_1 = "FF333333"
HEADER_COLOR_2 = "FF729FCF"
HEADER_COLOR_3 = "FFFF4000"
HEADER_COLOR_4 = "FFFF7B59"


def _clone_alignment(align: Optional[Alignment], **updates) -> Alignment:
    """Copy an Alignment, ignoring fields that may not exist in older openpyxl versions."""
    base = align or Alignment()
    data = {
        "horizontal": getattr(base, "horizontal", None),
        "vertical": getattr(base, "vertical", None),
        "text_rotation": getattr(base, "text_rotation", None),
        "wrap_text": getattr(base, "wrap_text", None),
        "shrink_to_fit": getattr(base, "shrink_to_fit", None),
        "indent": getattr(base, "indent", None),
        "justify_last_line": getattr(base, "justify_last_line", None),
        "reading_order": getattr(base, "reading_order", None),
    }
    data.update(updates)
    clean = {k: v for k, v in data.items() if v is not None}
    return Alignment(**clean)


class MainWidget(QWidget):
    log_message = Signal(str)
    processing_done = Signal(int, int, str)

    def __init__(self):
        super().__init__()
        self.setObjectName("bom_to_msl_widget")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(12)

        self.select_btn = PrimaryPushButton("Select BoM files...", self)
        self.select_btn.clicked.connect(self.select_files)

        self.files_box = QTextEdit(self)
        self.files_box.setReadOnly(True)
        self.files_box.setPlaceholderText("Selected files will appear here")
        self.files_box.setStyleSheet("QTextEdit{background: #2a2a2a; color: white; border: 1px solid #3a3a3a; border-radius: 6px;}")

        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText("Live process log will appear here")
        self.log_box.setStyleSheet("QTextEdit{background: #1f1f1f; color: #d0d0d0; border: 1px solid #3a3a3a; border-radius: 6px;}")

        self.run_btn = PrimaryPushButton("Run", self)
        self.run_btn.clicked.connect(self.run_process)

        layout.addWidget(self.select_btn)
        layout.addWidget(self.files_box, 1)
        layout.addWidget(self.log_box, 1)
        layout.addWidget(self.run_btn)

        self.log_message.connect(self.append_log)
        self.processing_done.connect(self.on_processing_done)

    def select_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, "Select BoM files")
        if files:
            self.files_box.setPlainText("\n".join(files))
        else:
            self.files_box.clear()

    def _selected_files(self) -> List[str]:
        text = self.files_box.toPlainText().strip()
        if not text:
            return []
        return [line for line in text.split("\n") if line.strip()]

    def run_process(self):
        files = self._selected_files()
        if not files:
            MessageBox("No files", "Please select one or more Excel files first.", self).exec()
            return

        self.log_box.clear()
        self.log_message.emit(f"Starting processing of {len(files)} file(s)...")
        self.run_btn.setEnabled(False)
        self.select_btn.setEnabled(False)

        def worker():
            ok, fail, out_path = 0, 0, ""
            try:
                out_path, ok, fail = process_files(files, self.log_message.emit)
            except Exception as e:
                self.log_message.emit(f"ERROR: {e}")
            self.processing_done.emit(ok, fail, out_path)

        threading.Thread(target=worker, daemon=True).start()

    def append_log(self, text: str):
        self.log_box.append(text)
        self.log_box.ensureCursorVisible()

    def on_processing_done(self, ok: int, fail: int, out_path: str):
        if out_path:
            self.log_message.emit(f"Output workbook saved to: {out_path}")
        self.log_message.emit(f"Completed: {ok} success, {fail} failed.")
        self.run_btn.setEnabled(True)
        self.select_btn.setEnabled(True)


def get_widget():
    return MainWidget()


# -----------------------------
# Processing implementation
# -----------------------------

def process_files(file_paths: List[str], log_emit) -> Tuple[str, int, int]:
    processor = BoMToMSLProcessor(log_emit)
    output_path = processor.process(file_paths)
    return output_path, processor.success_files, processor.fail_files


class BoMToMSLProcessor:
    def __init__(self, log_emit):
        self.log_emit = log_emit or (lambda _x: None)
        self.wb_result = Workbook()
        self.ws_result = self.wb_result.active
        self.ws_result.title = "RESULT"
        self.ws_log = self.wb_result.create_sheet("LOG")
        self.log_count = 0
        self.result_count = 0
        self.error_flag = False
        self.success_files = 0
        self.fail_files = 0
        self.header_titles: List[str] = []

        self._print_header()
        self._adjust_columns()

    # ---- helpers mirroring VBA ----
    def log(self, msg: str):
        stamp = f"[{datetime.now()}] {msg}"
        self.log_count += 1
        self.ws_log.cell(row=self.log_count, column=1, value=stamp)
        try:
            self.log_emit(msg)
        except Exception:
            pass

    def process(self, file_paths: List[str]) -> str:
        self.log("Process Begin...")
        total = len(file_paths)
        for idx, path in enumerate(file_paths, start=1):
            label = f"({idx}/{total}) {os.path.basename(path)}"
            try:
                self.log(f"{label} - Start Processing...")
                imported = self._process_style_sheet(path)
                self.success_files += 1
                self.log(f"{label} - End Processing... Total {imported} inserted.")
            except Exception as exc:
                self.error_flag = True
                self.fail_files += 1
                self.log(f"{label} - Error : {exc}")

        self._change_formatting()
        self._apply_layout_shrink()
        self._auto_fit_selected_columns()
        self._auto_fit_headers()
        self.log("Process Completed")

        out_path = self._proposed_output_path(file_paths)
        self.wb_result.save(out_path)
        return out_path

    def _process_style_sheet(self, path: str) -> int:
        cw_row = 0
        imported_rows = 0

        temp_file = os.path.basename(path)
        wb, ws = self._load_sheet(path)
        try:
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0

            if max_col != NORMAL_COL_COUNT:
                raise ValueError(f"{temp_file} - Unusual column size detected. (Default : {NORMAL_COL_COUNT})")

            style_season = self._get_season_style(ws)
            if not style_season:
                raise ValueError(f"{temp_file} - Error: Unable to get Season/Style")

            season, style = style_season.split("|", 1)

            for i in range(1, max_row + 1):
                seq = ""
                supplier = ""
                item = ""
                desc = ""
                cw = ""
                color = ""

                uom_val = ws.cell(row=i, column=S_COL_UOM).value
                if self._match_uom(uom_val):
                    cw_row = i

                seq_cell = ws.cell(row=i, column=S_COL_SEQ).value
                if self._has_numeric(seq_cell) and cw_row > 0:
                    seq = self._value_to_str(seq_cell)
                    desc_source = self._value_to_str(ws.cell(row=i, column=S_COL_DESC).value)
                    supplier = self._get_supplier(desc_source)
                    item = self._get_item(desc_source)
                    desc = self._get_desc(desc_source)

                    for j in range(S_COL_CW, max_col + 1):
                        cw_header = self._value_to_str(ws.cell(row=cw_row, column=j).value)
                        if not cw_header:
                            continue
                        cw = cw_header

                        color_raw = self._value_to_str(ws.cell(row=i, column=j).value)
                        color = self._format_color(color_raw)

                        if supplier.startswith("YKK"):
                            color = color.replace("/", "\n")
                            color = self._process_zipper(color)
                        else:
                            color = color.replace("/", " / ")

                        self._insert_data(season, style, seq, supplier, item, desc, cw, color)
                        imported_rows += 1
        finally:
            try:
                wb.close()
            except Exception:
                pass
        return imported_rows

    def _print_header(self):
        header = [
            "DATE RECEIVED", "MAT NO", "ORDER TYPE", "SAMPLE TYPE", "MATERIAL OETC", "CATEGORY",
            "SEASON", "SEQ", "STYLE NUMBER", "STYLE NAME", "STYLE CW", "VENDOR CODE", "VENDOR NAME", "VENDOR MCO",
            "ITEM", "DESCRIPTION", "COLOR CODE", "COLOR NAME", "SIZE", "QUALITY", "SIZE MATRIX/COO",
            "QTY", "UOM", "REMARKS", "REQUESTER",
        ]
        self.header_titles = header
        for idx, text in enumerate(header, start=1):
            self.ws_result.cell(row=R_START_ROW, column=R_START_COL + idx, value=text)

    def _adjust_columns(self):
        widths = [
            11, 11, 11, 11, 11,
            7.5, 7.5, 5, 7.5, 11, 8.5, 8.5, 38, 8.5,
            8.5, 38, 8.5, 27,
            11, 11, 11, 11, 11, 11, 11,
        ]
        colors = [
            HEADER_COLOR_1, HEADER_COLOR_2, HEADER_COLOR_3, HEADER_COLOR_3, HEADER_COLOR_3,
            HEADER_COLOR_2, HEADER_COLOR_3, HEADER_COLOR_1, HEADER_COLOR_3, HEADER_COLOR_4,
            HEADER_COLOR_4, HEADER_COLOR_3, HEADER_COLOR_2, HEADER_COLOR_3, HEADER_COLOR_3,
            HEADER_COLOR_2, HEADER_COLOR_3, HEADER_COLOR_4,
            HEADER_COLOR_2, HEADER_COLOR_2, HEADER_COLOR_2, HEADER_COLOR_3, HEADER_COLOR_3,
            HEADER_COLOR_2, HEADER_COLOR_2,
        ]
        thin = Side(style="thin", color="FF000000")

        self.ws_result.row_dimensions[R_START_ROW].height = 30

        for idx, width in enumerate(widths, start=1):
            cell = self.ws_result.cell(row=R_START_ROW, column=R_START_COL + idx)
            cell.fill = PatternFill("solid", fgColor=colors[idx - 1])
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font = Font(name="Calibri", sz=11, color="FFFFFFFF")
            col_letter = get_column_letter(R_START_COL + idx)
            self.ws_result.column_dimensions[col_letter].width = width

    def _insert_data(self, season: str, style: str, seq: str, supplier: str, item: str, desc: str, cw: str, color: str):
        vendor_code, vendor_name, vendor_mco = self._split_supplier(supplier)

        self.result_count += 1
        row = R_START_ROW + self.result_count

        self.ws_result.cell(row=row, column=R_COL_SEASON, value=season)
        self.ws_result.cell(row=row, column=R_COL_SEQ, value=seq)
        self.ws_result.cell(row=row, column=R_COL_STYLE_NUMBER, value=style)
        self.ws_result.cell(row=row, column=R_COL_STYLE_CW, value=cw)
        self.ws_result.cell(row=row, column=R_COL_VENDOR_CODE, value=vendor_code)
        self.ws_result.cell(row=row, column=R_COL_VENDOR_NAME, value=vendor_name)
        self.ws_result.cell(row=row, column=R_COL_VENDOR_MCO, value=vendor_mco)
        item_cell = self.ws_result.cell(row=row, column=R_COL_ITEM, value=item)
        item_cell.number_format = "@"
        self.ws_result.cell(row=row, column=R_COL_DESC, value=desc)
        self.ws_result.cell(row=row, column=R_COL_COLOR_NAME, value=color)

    def _split_supplier(self, supplier: str) -> Tuple[str, str, str]:
        supplier = self._normalize_supplier(supplier)
        supplier = self._strip_nike_tag(supplier)

        vendor_code = ""
        vendor_name = ""
        vendor_mco = ""

        if not supplier or supplier.strip() == "-":
            return vendor_code, vendor_name, vendor_mco

        first_dash = self._find_meaningful_dash(supplier, 0)

        if first_dash < 0:
            vendor_name = supplier.strip()
            return vendor_code, vendor_name, vendor_mco

        before = supplier[:first_dash].strip()

        if len(before) < 8:
            vendor_code = before
            second_dash = self._find_meaningful_dash(supplier, first_dash + 1)

            if second_dash < 0:
                vendor_name = supplier[first_dash + 1:].strip()
            else:
                vendor_name = supplier[first_dash + 1:second_dash].strip()
                vendor_mco = supplier[second_dash + 1:].strip()
        else:
            vendor_name = before
            vendor_mco = supplier[first_dash + 1:].strip()

        vendor_name = self._strip_nike_tag(vendor_name)
        vendor_mco = self._strip_nike_tag(vendor_mco)

        return vendor_code, vendor_name, vendor_mco

    def _find_meaningful_dash(self, text: str, start: int) -> int:
        for idx in range(start, len(text)):
            ch = text[idx]
            if self._is_dash_char(ch):
                prev_ch = text[idx - 1] if idx > 0 else ""
                next_ch = text[idx + 1] if (idx + 1) < len(text) else ""

                if (self._is_space_char(prev_ch) or idx == 0) or (self._is_space_char(next_ch) or idx == len(text) - 1):
                    return idx

                if not (self._is_alphanumeric(prev_ch) and self._is_alphanumeric(next_ch)):
                    return idx
        return -1

    def _change_formatting(self):
        ws = self.ws_result
        max_row = ws.max_row or 0

        for row in range(2, max_row + 1):
            # Season formatting
            season_val = self._value_to_str(ws.cell(row=row, column=R_COL_SEASON).value)
            if season_val:
                ws.cell(row=row, column=R_COL_SEASON, value=f"{season_val[:2]}'{season_val[-2:]}")

            # Vendor name cleanup
            vendor_val = self._value_to_str(ws.cell(row=row, column=R_COL_VENDOR_NAME).value)
            if vendor_val:
                vendor_val = vendor_val.replace(",", "")
                vendor_val = vendor_val.replace("NIKE-APPROVED VENDOR", "")
                ws.cell(row=row, column=R_COL_VENDOR_NAME, value=vendor_val.strip())

            # Description cleanup
            desc_val = self._value_to_str(ws.cell(row=row, column=R_COL_DESC).value)
            if desc_val:
                ws.cell(row=row, column=R_COL_DESC, value=desc_val.replace(",", "").strip())

            # Style CW formatting
            cw_val = self._value_to_str(ws.cell(row=row, column=R_COL_STYLE_CW).value)
            if cw_val:
                first = cw_val[0]
                if first in ("*", "@"):
                    cw_val = "#" + cw_val[1:].strip()
                elif "#" not in cw_val:
                    cw_val = "#" + cw_val

                cw_val = cw_val.strip()
                if cw_val.startswith("#") and len(cw_val) < 4:
                    cw_val = "#0" + cw_val.replace("#", "")
                ws.cell(row=row, column=R_COL_STYLE_CW, value=cw_val)

            # Color name + remarks
            color_val = self._value_to_str(ws.cell(row=row, column=R_COL_COLOR_NAME).value)
            if color_val and "DTM" in color_val:
                ws.cell(row=row, column=R_COL_COLOR_NAME, value=color_val.replace("DTM", "").strip())
                ws.cell(row=row, column=R_COL_REMARKS, value="DTM")

        # Remove wrap text from all populated cells
        max_col = ws.max_column or 0
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.alignment = _clone_alignment(cell.alignment, wrap_text=False)

    def _apply_layout_shrink(self):
        ws = self.ws_result
        last_row = ws.max_row or 0
        if last_row < 2:
            return

        shrink_cols = [R_COL_VENDOR_NAME, R_COL_DESC, R_COL_COLOR_NAME, R_COL_STYLE_NAME]
        for col_idx in shrink_cols:
            col_letter = get_column_letter(col_idx)
            for cell in ws[col_letter]:
                cell.alignment = _clone_alignment(cell.alignment, wrap_text=False, shrink_to_fit=True)

        # Row height and vertical alignment
        for row_idx in range(2, last_row + 1):
            ws.row_dimensions[row_idx].height = 18
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.alignment = _clone_alignment(cell.alignment, vertical="center")

        # Left-align long text columns
        left_cols = [R_COL_VENDOR_NAME, R_COL_DESC, R_COL_COLOR_NAME]
        for col_idx in left_cols:
            col_letter = get_column_letter(col_idx)
            for cell in ws[col_letter]:
                cell.alignment = _clone_alignment(cell.alignment, horizontal="left")

    def _auto_fit_selected_columns(self):
        ws = self.ws_result
        cols = [
            1, 3, 4, 20,
            R_COL_SEASON,
            R_COL_SEQ,
            R_COL_STYLE_NUMBER,
            R_COL_STYLE_NAME,
            R_COL_STYLE_CW,
            R_COL_VENDOR_CODE,
            R_COL_VENDOR_NAME,
            R_COL_VENDOR_MCO,
            R_COL_ITEM,
            R_COL_COLOR_CODE,
            R_COL_COLOR_NAME,
        ]

        max_row = ws.max_row or 0
        for col_idx in cols:
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[col_letter][:max_row]:
                val = cell.value
                if val is None:
                    continue
                length = len(str(val))
                if length > max_len:
                    max_len = length
            if max_len:
                ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    def _auto_fit_headers(self):
        """Ensure header columns are at least wide enough for their header text."""
        ws = self.ws_result
        if not self.header_titles:
            return
        for idx, title in enumerate(self.header_titles, start=1):
            if not title:
                continue
            col_letter = get_column_letter(R_START_COL + idx)
            current_width = ws.column_dimensions[col_letter].width or 0
            # Give headers extra breathing room beyond the raw text length
            needed = min(len(title) + 6, 70)
            if needed > current_width:
                ws.column_dimensions[col_letter].width = needed

    # ---- parsing helpers ----
    def _get_season_style(self, ws) -> str:
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        for i in range(1, max_row + 1):
            for j in range(1, max_col + 1):
                cell_val = self._value_to_str(ws.cell(row=i, column=j).value).upper()
                cell_val = cell_val.replace(" ", "")
                if STYLE_LABEL in cell_val and SEASON_LABEL in cell_val:
                    season = self._get_season(cell_val)
                    style = self._get_style(cell_val)
                    if season and style:
                        return f"{season}|{style}"
        return ""

    def _process_zipper(self, source: str) -> str:
        lines = source.splitlines()
        if len(lines) < 5:
            return source

        codes: List[str] = []
        for idx in range(5):
            code = (lines[idx] if idx < len(lines) else "")[:3]
            if not code.strip():
                continue
            if code not in codes:
                codes.append(code)

        result_parts: List[str] = []
        for code in codes:
            if code.strip().upper() == "NOCOLR":
                continue
            segments: List[str] = []
            for j in range(5):
                line_val = (lines[j] if j < len(lines) else "")[:3]
                if code == line_val:
                    if j == 0:
                        segments.append("TP")
                    elif j == 1:
                        segments.append("TH")
                    elif j == 2:
                        segments.append("PL")
                    elif j == 3:
                        segments.append("SL")
                    elif j == 4:
                        segments.append("SP")
            if segments:
                text = "/".join(segments)
                if " " in code:
                    text += f":#{code[:3]} - "
                else:
                    text += f":#{code} - "
                result_parts.append(text)

        result = "".join(result_parts)
        if len(lines) >= 6:
            if result:
                result = result + "\n" + lines[5]
            else:
                result = "\n" + lines[5]

        return result if result else source

    def _get_supplier(self, source: str) -> str:
        if "***" not in source or not source.strip():
            return "-"
        parts = source.split("***")
        return parts[0].strip().upper()

    def _get_item(self, source: str) -> str:
        if "***" not in source or not source.strip():
            return "-"
        parts = source.split("***")
        tail = parts[1].strip()
        return tail[1:1 + ITEM_LENGTH].upper()

    def _get_desc(self, source: str) -> str:
        if "***" not in source or not source.strip():
            return "-"
        parts = source.split("***")
        tail = parts[1].strip()
        return tail[ITEM_LENGTH + 1:].upper()

    def _format_color(self, source: str) -> str:
        if not source:
            return source
        result = list(source)
        for idx, ch in enumerate(result):
            if ch == " " and idx >= 4:
                snippet = "".join(result[idx - 4:idx]).strip()
                if len(snippet) > 3:
                    result[idx] = "/"
        return "".join(result)

    def _get_style(self, source: str) -> str:
        pos = source.find(STYLE_LABEL)
        if pos < 0:
            return ""
        start = pos + len(STYLE_LABEL)
        return source[start:start + STYLE_LENGTH]

    def _get_season(self, source: str) -> str:
        pos = source.find(SEASON_LABEL)
        if pos < 0:
            return ""
        start = pos + len(SEASON_LABEL)
        return source[start:start + SEASON_LENGTH]

    # ---- low level utilities ----
    def _value_to_str(self, val: Any) -> str:
        if val is None:
            return ""
        return str(val).strip()

    def _has_numeric(self, val: Any) -> bool:
        if val is None:
            return False
        if isinstance(val, bool):
            return False
        try:
            float(str(val).strip())
            return True
        except Exception:
            return False

    def _match_uom(self, val: Any) -> bool:
        if val is None:
            return False
        return str(val).strip().upper() == S_UOM_VALUE.upper()

    def _normalize_supplier(self, s: str) -> str:
        if s is None:
            return ""
        # Dashes
        dash_chars = [chr(8211), chr(8212), chr(8213), chr(8208), chr(8210), chr(8722), chr(173)]
        for d in dash_chars:
            s = s.replace(d, "-")
        # Spaces
        space_chars = [chr(160), chr(8239), chr(8201), chr(8194), chr(8195)]
        for sp in space_chars:
            s = s.replace(sp, " ")
        while "  " in s:
            s = s.replace("  ", " ")
        return s.strip()

    def _strip_nike_tag(self, s: str) -> str:
        if s is None:
            return ""
        s = s.replace("NIKE-APPROVED VENDOR", "")
        s = s.replace(", ,", ",")
        s = s.replace(" -  - ", " - ")
        s = s.strip()
        while s.endswith("-") or s.endswith(",") or s.endswith(" "):
            s = s[:-1].strip()
        return s

    def _is_dash_char(self, ch: str) -> bool:
        if not ch:
            return False
        code = ord(ch)
        return ch == "-" or code in (8211, 8212, 8213, 8208, 8210, 8722, 173)

    def _is_space_char(self, ch: str) -> bool:
        if not ch:
            return False
        code = ord(ch)
        return ch == " " or code in (160, 8239, 8201, 8194, 8195)

    def _is_alphanumeric(self, ch: str) -> bool:
        if not ch:
            return False
        return ch.isalnum()

    def _proposed_output_path(self, inputs: List[str]) -> str:
        base_dir = os.path.dirname(inputs[0]) if inputs else os.getcwd()
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = f"{stamp}_MSL.xlsx"
        candidate = os.path.join(base_dir, base_name)
        if not os.path.exists(candidate):
            return candidate
        n = 1
        while True:
            candidate = os.path.join(base_dir, f"{stamp}_MSL ({n}).xlsx")
            if not os.path.exists(candidate):
                return candidate
            n += 1

    def _load_sheet(self, path: str):
        wb = self._open_workbook(path)
        ws = wb.active
        return wb, ws

    def _open_workbook(self, path: str):
        if path.lower().endswith(".xls"):
            return self._load_xls(path)
        return load_workbook(path, data_only=True)

    def _load_xls(self, path: str):
        if xlrd is None:
            raise RuntimeError("xlrd is required to read .xls files. Please install xlrd.")
        book = xlrd.open_workbook(path, formatting_info=False)
        sh = book.sheet_by_index(0)
        wb = Workbook()
        ws = wb.active
        ws.title = sh.name if sh.name else "Sheet1"
        for r in range(sh.nrows):
            row = sh.row(r)
            for c, cell in enumerate(row):
                ws.cell(row=r + 1, column=c + 1, value=self._xlrd_cell_value(book, cell))
        return wb

    def _xlrd_cell_value(self, book, cell):
        from xlrd import XL_CELL_DATE, xldate_as_datetime, XL_CELL_BOOLEAN
        if cell.ctype == XL_CELL_DATE:
            try:
                return xldate_as_datetime(cell.value, book.datemode)
            except Exception:
                return cell.value
        if cell.ctype == XL_CELL_BOOLEAN:
            return bool(cell.value)
        return cell.value
