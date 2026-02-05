#!/usr/bin/env python3
from __future__ import annotations

import re
import threading
import os
import platform
from contextlib import nullcontext
from datetime import datetime
from html import unescape
from html.parser import HTMLParser
from pathlib import Path
from typing import Callable, List, Tuple, TYPE_CHECKING

from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QComboBox,
    QFileDialog,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QProgressBar,
    QSpinBox,
    QTextEdit,
    QVBoxLayout,
    QWidget,
    QSizePolicy,
)
from qfluentwidgets import PrimaryPushButton, MessageBox

# Lazy imports for heavy ML dependencies
if TYPE_CHECKING:
    import pandas as pd
    import pypdfium2 as pdfium
    import torch
    from PIL import Image
    from transformers import LightOnOcrForConditionalGeneration, LightOnOcrProcessor


def check_dependencies() -> Tuple[bool, str]:
    """
    Check if required dependencies are installed.

    Returns:
        Tuple of (all_ok, error_message)
    """
    missing = []

    try:
        import torch
    except ImportError:
        missing.append("torch")

    try:
        import pandas
    except ImportError:
        missing.append("pandas")

    try:
        import pypdfium2
    except ImportError:
        missing.append("pypdfium2")

    try:
        from PIL import Image
    except ImportError:
        missing.append("pillow")

    try:
        import openpyxl
    except ImportError:
        missing.append("openpyxl")

    try:
        from transformers import LightOnOcrForConditionalGeneration, LightOnOcrProcessor
    except ImportError as e:
        if "No module named 'transformers'" in str(e):
            missing.append("transformers (from git)")
        else:
            missing.append(f"transformers with LightOnOCR support: {e}")

    if missing:
        msg = "Missing dependencies:\n" + "\n".join(f"  - {m}" for m in missing)
        msg += "\n\nInstall with:\n  pip install torch pandas pypdfium2 pillow openpyxl accelerate defusedxml"
        msg += "\n  pip install git+https://github.com/huggingface/transformers"
        return False, msg

    return True, ""


def _emit(log_emit: Callable[[str], None] | None, text: str):
    """Log helper that tolerates missing/failed callbacks."""
    if callable(log_emit):
        try:
            log_emit(text)
            return
        except Exception:
            pass
    print(text)


def ensure_unique_path(path: Path) -> Path:
    """Return a non-existing path by appending (n) before the extension if needed."""
    if not path.exists():
        return path
    stem = path.stem
    ext = path.suffix
    parent = path.parent
    counter = 1
    while True:
        candidate = parent / f"{stem} ({counter}){ext}"
        if not candidate.exists():
            return candidate
        counter += 1


def get_total_memory_gb() -> float | None:
    """Best-effort total physical RAM in GiB."""
    try:
        import psutil  # type: ignore

        return float(psutil.virtual_memory().total) / (1024**3)
    except Exception:
        pass

    # Unix-like (Linux/macOS)
    try:
        page_size = os.sysconf("SC_PAGE_SIZE")
        phys_pages = os.sysconf("SC_PHYS_PAGES")
        return (page_size * phys_pages) / (1024**3)
    except Exception:
        pass

    # Windows fallback
    if platform.system() == "Windows":
        try:
            import ctypes

            class _MEMORYSTATUSEX(ctypes.Structure):
                _fields_ = [
                    ("dwLength", ctypes.c_uint32),
                    ("dwMemoryLoad", ctypes.c_uint32),
                    ("ullTotalPhys", ctypes.c_uint64),
                    ("ullAvailPhys", ctypes.c_uint64),
                    ("ullTotalPageFile", ctypes.c_uint64),
                    ("ullAvailPageFile", ctypes.c_uint64),
                    ("ullTotalVirtual", ctypes.c_uint64),
                    ("ullAvailVirtual", ctypes.c_uint64),
                    ("ullAvailExtendedVirtual", ctypes.c_uint64),
                ]

            stat = _MEMORYSTATUSEX()
            stat.dwLength = ctypes.sizeof(_MEMORYSTATUSEX)
            if ctypes.windll.kernel32.GlobalMemoryStatusEx(ctypes.byref(stat)) == 0:
                return None
            return float(stat.ullTotalPhys) / (1024**3)
        except Exception:
            return None

    return None


def _recommended_cpu_threads() -> int:
    cpu = os.cpu_count() or 1
    return min(8, max(1, cpu - 2))


def _device_kind(device) -> str:
    s = str(device or "cpu").lower()
    return s.split(":", 1)[0]


def _recommended_defaults_for_device(device, total_ram_gb: float | None) -> dict:
    low_ram = total_ram_gb is None or total_ram_gb <= 10.0
    kind = _device_kind(device)

    if kind == "cpu":
        return {
            "max_tokens": 8192,
            "render_size": 1280,
            "cpu_threads": _recommended_cpu_threads(),
        }

    # GPU-ish backends (cuda/mps/directml/privateuseone/etc.)
    return {
        "max_tokens": 8192,
        "render_size": 1280,
        "cpu_threads": _recommended_cpu_threads(),
    }


def detect_gpu_info(total_ram_gb: float | None = None) -> dict:
    """
    Detect available GPU hardware and return info dict.

    Returns dict with:
        - devices: list of (device, display_name, dtype_str, is_discrete)
        - recommended: index of recommended device
    """
    devices = []
    recommended_idx = 0

    try:
        import torch
    except ImportError:
        # torch not installed, return CPU-only option
        devices.append(("cpu", "CPU - (torch not installed)", "float32", False))
        return {"devices": devices, "recommended": 0}

    # Check CUDA (NVIDIA discrete GPUs)
    if torch.cuda.is_available():
        for i in range(torch.cuda.device_count()):
            props = torch.cuda.get_device_properties(i)
            vram_gb = props.total_memory / (1024**3)
            name = f"CUDA:{i} - {props.name} ({vram_gb:.1f}GB)"
            # Use bfloat16 for CUDA (optimal for transformer models)
            devices.append((f"cuda:{i}", name, "bfloat16", True))

    # Check MPS (Apple Silicon)
    if torch.backends.mps.is_available():
        name = "MPS - Apple Silicon GPU"
        # Use float32 for MPS (bfloat16 support is limited)
        devices.append(("mps", name, "float32", True))

    # Check DirectML (Windows integrated/discrete GPUs; optional)
    if platform.system() == "Windows":
        try:
            import torch_directml  # type: ignore

            dml_device = torch_directml.device()
            devices.append((dml_device, "DirectML - Windows GPU (experimental)", "float16", False))
        except Exception:
            pass

    # Always add CPU option
    cpu_name = "CPU - System Processor"
    # Check if this is likely integrated graphics only
    if not devices:
        cpu_name = "CPU - No discrete GPU detected (slower)"
    devices.append(("cpu", f"{cpu_name} (bfloat16, low RAM)", "bfloat16", False))
    devices.append(("cpu", f"{cpu_name} (float32, compatibility)", "float32", False))

    # Recommend first discrete GPU if available, otherwise CPU
    for i, (_, _, _, is_discrete) in enumerate(devices):
        if is_discrete:
            recommended_idx = i
            break
    else:
        # CPU-only (or integrated/experimental backends): default to CPU bfloat16 for lower RAM usage.
        for i, (dev, _, dtype, _) in enumerate(devices):
            if _device_kind(dev) == "cpu" and dtype == "bfloat16":
                recommended_idx = i
                break

    return {"devices": devices, "recommended": recommended_idx}


def render_pdf_page(pdf, page_num: int, target_size: int = 1540):
    """
    Render a PDF page to PIL Image at target longest dimension.

    Args:
        pdf: Open PdfDocument
        page_num: 0-indexed page number
        target_size: Target size for longest dimension (default 1540px per model recommendation)

    Returns:
        PIL Image of the rendered page
    """
    page = pdf[page_num]

    # Get page dimensions at 72 DPI (PDF standard)
    width, height = page.get_size()

    # Calculate scale to achieve target longest dimension
    longest = max(width, height)
    scale = target_size / longest

    # Render at calculated scale
    bitmap = page.render(scale=scale)
    pil_image = bitmap.to_pil()
    for obj in (bitmap, page):
        try:
            close = getattr(obj, "close", None)
            if callable(close):
                close()
        except Exception:
            pass

    return pil_image


def parse_ocr_to_dataframe(ocr_text: str):
    """
    Parse OCR output text into a DataFrame.

    Attempts to detect table structures in the OCR output.
    Falls back to line-by-line text if no table structure found.
    """
    import pandas as pd

    lines = ocr_text.strip().split("\n")

    if not lines:
        return pd.DataFrame({"Text": [""]})

    # Try to detect markdown table format
    table_lines = []
    in_table = False

    for line in lines:
        # Check for markdown table separator (e.g., |---|---|)
        if re.match(r"^\s*\|[-:\s|]+\|\s*$", line):
            in_table = True
            continue

        # Check for table row (starts and ends with |)
        if line.strip().startswith("|") and line.strip().endswith("|"):
            cells = [cell.strip() for cell in line.strip()[1:-1].split("|")]
            table_lines.append(cells)
            in_table = True
        elif in_table and not line.strip():
            # Empty line ends table
            break

    if table_lines and len(table_lines) > 1:
        # Use first row as headers
        headers = table_lines[0]
        data = table_lines[1:]

        # Ensure all rows have same number of columns
        max_cols = max(len(row) for row in table_lines)
        headers = headers + [""] * (max_cols - len(headers))
        data = [row + [""] * (max_cols - len(row)) for row in data]

        return pd.DataFrame(data, columns=headers)

    # Try to detect tab-separated or multi-space separated values
    potential_rows = []
    for line in lines:
        if "\t" in line:
            cells = line.split("\t")
        elif "  " in line:  # Two or more spaces as delimiter
            cells = re.split(r"\s{2,}", line.strip())
        else:
            cells = [line]
        potential_rows.append(cells)

    if potential_rows:
        # Check if most rows have consistent column count
        col_counts = [len(row) for row in potential_rows]
        most_common_count = max(set(col_counts), key=col_counts.count)

        if most_common_count > 1:
            # Normalize rows to most common column count
            normalized = []
            for row in potential_rows:
                if len(row) < most_common_count:
                    row = row + [""] * (most_common_count - len(row))
                elif len(row) > most_common_count:
                    row = row[:most_common_count]
                normalized.append(row)

            return pd.DataFrame(normalized[1:], columns=normalized[0]) if len(normalized) > 1 else pd.DataFrame(normalized)

    # Fallback: single column with all text
    return pd.DataFrame({"Text": lines})


def _strip_markdown_line(line: str) -> str:
    line = line.strip()
    if not line:
        return ""

    # Remove markdown headings ("#", "##", etc.)
    line = re.sub(r"^\s{0,3}#{1,6}\s*", "", line)

    # Remove common emphasis markers
    line = re.sub(r"\*\*(.+?)\*\*", r"\1", line)
    line = re.sub(r"__(.+?)__", r"\1", line)
    line = re.sub(r"\*(.+?)\*", r"\1", line)
    line = re.sub(r"_(.+?)_", r"\1", line)

    return line.strip()


def _extract_html_tables(text: str) -> Tuple[str, List[str]]:
    """
    Extract <table>...</table> blocks from OCR text.

    Returns (text_without_tables, tables_html).
    Works even if the last table is truncated (missing </table>).
    """
    tables: List[str] = []
    kept_lines: List[str] = []

    in_table = False
    buf: List[str] = []

    for raw in (text or "").splitlines():
        line = raw.rstrip("\n")
        lower = line.lower()

        if not in_table and "<table" in lower:
            in_table = True
            buf = [line]
            if "</table>" in lower:
                in_table = False
                tables.append("\n".join(buf))
                buf = []
            continue

        if in_table:
            buf.append(line)
            if "</table>" in lower:
                in_table = False
                tables.append("\n".join(buf))
                buf = []
            continue

        kept_lines.append(line)

    if in_table and buf:
        tables.append("\n".join(buf))

    return "\n".join(kept_lines), tables


def _extract_markdown_tables(text: str) -> Tuple[str, List[List[str]]]:
    """
    Extract GitHub/markdown pipe tables from OCR text.

    Returns (text_without_tables, tables_as_lines).
    """
    lines = (text or "").splitlines()
    kept: List[str] = []
    tables: List[List[str]] = []

    pipe_row = re.compile(r"^\s*\|.*\|\s*$")
    sep_row = re.compile(r"^\s*\|[-:\s|]+\|\s*$")

    i = 0
    while i < len(lines):
        line = lines[i]
        if pipe_row.match(line) and i + 1 < len(lines) and sep_row.match(lines[i + 1]):
            table_lines = [line]
            i += 2  # skip separator row
            while i < len(lines) and pipe_row.match(lines[i]):
                table_lines.append(lines[i])
                i += 1
            tables.append(table_lines)
            continue

        kept.append(line)
        i += 1

    return "\n".join(kept), tables


def _parse_markdown_table(table_lines: List[str]) -> List[List[str]]:
    if not table_lines:
        return []

    rows: List[List[str]] = []
    for line in table_lines:
        line = line.strip()
        if not (line.startswith("|") and line.endswith("|")):
            continue
        cells = [c.strip() for c in line[1:-1].split("|")]
        rows.append(cells)

    if not rows:
        return []

    max_cols = max((len(r) for r in rows), default=0)
    return [r + [""] * (max_cols - len(r)) for r in rows]


class _HTMLTableParser(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=False)
        self.rows: List[List[str]] = []
        self._current_row: List[str] = []
        self._current_cell: List[str] = []
        self._in_cell = False

    def handle_starttag(self, tag, attrs):
        tag = (tag or "").lower()
        if tag == "tr":
            self._current_row = []
        elif tag in ("td", "th"):
            self._in_cell = True
            self._current_cell = []

    def handle_endtag(self, tag):
        tag = (tag or "").lower()
        if tag in ("td", "th"):
            cell = unescape("".join(self._current_cell)).strip()
            self._current_row.append(cell)
            self._current_cell = []
            self._in_cell = False
        elif tag == "tr":
            if any(c.strip() for c in self._current_row):
                self.rows.append(self._current_row)
            self._current_row = []

    def handle_data(self, data):
        if self._in_cell:
            self._current_cell.append(data)

    def handle_entityref(self, name):  # pragma: no cover
        if self._in_cell:
            self._current_cell.append(f"&{name};")

    def handle_charref(self, name):  # pragma: no cover
        if self._in_cell:
            self._current_cell.append(f"&#{name};")


def _parse_html_table(table_html: str) -> List[List[str]]:
    if not table_html:
        return []

    parser = _HTMLTableParser()
    try:
        parser.feed(table_html)
        parser.close()
    except Exception:
        # Be forgiving: return whatever we could parse.
        pass

    rows = list(parser.rows)

    # Flush truncated final cell/row if any.
    if parser._current_cell:
        cell = unescape("".join(parser._current_cell)).strip()
        parser._current_row.append(cell)
        parser._current_cell = []
    if parser._current_row and any(c.strip() for c in parser._current_row):
        rows.append(parser._current_row)

    if not rows:
        return []

    max_cols = max((len(r) for r in rows), default=0)
    padded = [r + [""] * (max_cols - len(r)) for r in rows]
    return padded


def _split_key_value(line: str) -> Tuple[str, str] | None:
    """
    Heuristic split of a line into (key, value).
    Returns None if the line looks like a heading/paragraph.
    """
    if not line:
        return None

    # Skip page markers like "1/4"
    if re.fullmatch(r"\s*\d+\s*/\s*\d+\s*", line):
        return None

    known_keys = (
        "Payer(s)",
        "Payer",
        "Payment Provider",
        "Payment Reference",
        "Account Number",
        "Account Name",
        "Paid Through",
        "Status",
        "Payment Method",
        "Value Date",
        "Fees (USD)",
        "Fees(USD)",
        "Fee Type",
        "Invoice Fee",
    )

    for key in sorted(known_keys, key=len, reverse=True):
        if line.startswith(key + " "):
            value = line[len(key) :].strip()
            if value:
                return key, value

    # Split on colon
    if ":" in line and not line.strip().startswith("http"):
        left, right = line.split(":", 1)
        if left.strip() and right.strip():
            return left.strip(), right.strip()

    # Split if the "key" ends with ')' (common for labels like "Payer(s)")
    m = re.match(r"^(.+?\))\s+(.+)$", line)
    if m:
        key = m.group(1).strip()
        value = m.group(2).strip()
        if key and value:
            return key, value

    return None


def _ocr_text_to_rows(text: str) -> List[List[str]]:
    rows: List[List[str]] = []
    if not text:
        return rows

    cleaned_lines = []
    for raw in text.splitlines():
        line = _strip_markdown_line(raw)
        if not line:
            continue
        if re.fullmatch(r"\d+\s*/\s*\d+", line):
            continue
        cleaned_lines.append(line)

    i = 0
    while i < len(cleaned_lines):
        line = cleaned_lines[i]

        # If the first two lines look like (timestamp, title), place them on one row.
        if (
            i == 0
            and i + 1 < len(cleaned_lines)
            and re.fullmatch(r"\d{1,2}/\d{1,2}/\d{2},\s*\d{1,2}:\d{2}\s*[AP]M", line)
        ):
            rows.append([line, cleaned_lines[i + 1]])
            i += 2
            continue

        kv = _split_key_value(line)
        if kv:
            rows.append([kv[0], kv[1]])
        else:
            rows.append([line])
        i += 1

    return rows


def _write_rows(ws, rows: List[List[str]], start_row: int = 1, start_col: int = 1):
    r = start_row
    for row in rows:
        c = start_col
        for val in row:
            if val is not None and str(val) != "":
                ws.cell(row=r, column=c, value=val)
            c += 1
        r += 1


def _normalize_cell_text(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip().lower())


def _is_likely_header_row(row: List[str]) -> bool:
    if not row:
        return False

    joined = " ".join(_normalize_cell_text(c) for c in row if c and c.strip())
    if not joined:
        return False

    keywords = (
        "reference number",
        "invoice number",
        "amount (in invoice currency)",
        "settlement amount",
        "fee type",
        "contract",
        "payer",
        "purchase order",
        "transaction fees",
    )
    if any(k in joined for k in keywords):
        return True

    alpha_cells = sum(any(ch.isalpha() for ch in (c or "")) for c in row)
    digit_cells = sum(any(ch.isdigit() for ch in (c or "")) for c in row)

    # Mostly alphabetic header-like row, not mostly numbers.
    if alpha_cells >= max(2, len(row) // 2) and digit_cells <= max(1, len(row) // 3):
        return True

    return False


def _find_header_row_index(rows: List[List[str]]) -> int | None:
    for idx, row in enumerate(rows[:5]):  # only scan the top section
        if _is_likely_header_row(row):
            return idx
    return None


def _pad_row(row: List[str], target_cols: int) -> List[str]:
    if len(row) >= target_cols:
        return row
    return row + [""] * (target_cols - len(row))


def _build_workbook_from_pages(page_texts: List[str]):
    from openpyxl import Workbook

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    metadata_rows: List[List[str]] = []
    table_accs: List[dict] = []

    for page_idx, ocr_text in enumerate(page_texts, start=1):
        text_wo_tables, tables = _extract_html_tables(ocr_text or "")
        text_wo_tables, md_tables = _extract_markdown_tables(text_wo_tables)

        page_rows = _ocr_text_to_rows(text_wo_tables)

        # Light dedupe: repeated headers at the top of subsequent pages.
        if page_idx > 1 and page_rows:
            for idx, row in enumerate(page_rows):
                if idx < 3 and row in metadata_rows:
                    continue
                metadata_rows.append(row)
        else:
            metadata_rows.extend(page_rows)

        for t in tables:
            parsed = _parse_html_table(t)
            if not parsed:
                continue

            header_idx = _find_header_row_index(parsed)
            if header_idx is not None:
                header_row = parsed[header_idx]
                sig = tuple(_normalize_cell_text(c) for c in header_row)

                data_rows = parsed[header_idx + 1 :]
                target_cols = len(header_row)
                data_rows = [_pad_row(r, target_cols) for r in data_rows]

                existing = next((acc for acc in table_accs if acc.get("sig") == sig), None)
                if existing is None:
                    table_accs.append({"sig": sig, "rows": [header_row] + data_rows})
                else:
                    existing_cols = max((len(r) for r in existing["rows"]), default=0)
                    target_cols = max(existing_cols, target_cols)
                    existing["rows"] = [_pad_row(r, target_cols) for r in existing["rows"]]
                    existing["rows"].extend(_pad_row(r, target_cols) for r in data_rows)
            else:
                # No header: treat as continuation of the most recent table.
                if table_accs:
                    target_cols = max((len(r) for r in table_accs[-1]["rows"]), default=0)
                    if target_cols == 0:
                        table_accs[-1]["rows"].extend(parsed)
                    else:
                        table_accs[-1]["rows"].extend(_pad_row(r, target_cols) for r in parsed)
                else:
                    table_accs.append({"sig": None, "rows": parsed})

        for md in md_tables:
            parsed = _parse_markdown_table(md)
            if not parsed:
                continue

            header_idx = _find_header_row_index(parsed)
            if header_idx is not None:
                header_row = parsed[header_idx]
                sig = tuple(_normalize_cell_text(c) for c in header_row)
                data_rows = parsed[header_idx + 1 :]
                target_cols = len(header_row)
                data_rows = [_pad_row(r, target_cols) for r in data_rows]

                existing = next((acc for acc in table_accs if acc.get("sig") == sig), None)
                if existing is None:
                    table_accs.append({"sig": sig, "rows": [header_row] + data_rows})
                else:
                    existing_cols = max((len(r) for r in existing["rows"]), default=0)
                    target_cols = max(existing_cols, target_cols)
                    existing["rows"] = [_pad_row(r, target_cols) for r in existing["rows"]]
                    existing["rows"].extend(_pad_row(r, target_cols) for r in data_rows)
            else:
                if table_accs:
                    target_cols = max((len(r) for r in table_accs[-1]["rows"]), default=0)
                    if target_cols == 0:
                        table_accs[-1]["rows"].extend(parsed)
                    else:
                        table_accs[-1]["rows"].extend(_pad_row(r, target_cols) for r in parsed)
                else:
                    table_accs.append({"sig": None, "rows": parsed})

    # If there are no tables, still create a single sheet with the metadata.
    ws1 = wb.create_sheet("Table_1")
    _write_rows(ws1, metadata_rows, start_row=1, start_col=1)

    write_row = len(metadata_rows) + 2 if metadata_rows else 1
    if table_accs:
        _write_rows(ws1, table_accs[0]["rows"], start_row=write_row, start_col=1)

        for idx, acc in enumerate(table_accs[1:], start=2):
            ws = wb.create_sheet(f"Table_{idx}")
            _write_rows(ws, acc["rows"], start_row=1, start_col=1)

    return wb


class OCRModel:
    """Singleton-ish class to manage model loading and inference."""

    _instance = None
    _model = None
    _processor = None
    _device = None
    _dtype = None
    _cpu_threads = None

    _MODEL_ID = "lightonai/LightOnOCR-2-1B"

    @classmethod
    def load(cls, device, dtype_str: str, log_emit=None, cpu_threads: int | None = None) -> "OCRModel":
        """Load or return cached model."""
        import torch
        from transformers import LightOnOcrForConditionalGeneration, LightOnOcrProcessor

        # Convert dtype string to torch dtype
        dtype_map = {
            "bfloat16": torch.bfloat16,
            "float16": torch.float16,
            "float32": torch.float32,
        }
        dtype = dtype_map.get(dtype_str, torch.float32)

        if cls._model is None or cls._device != device or cls._dtype != dtype:
            cls.unload()

            # Configure CPU threading (best-effort; no-op on GPU)
            if _device_kind(device) == "cpu":
                desired_threads = cpu_threads or _recommended_cpu_threads()
                if cls._cpu_threads != desired_threads:
                    try:
                        torch.set_num_threads(int(desired_threads))
                    except Exception:
                        pass
                    try:
                        torch.set_num_interop_threads(1)
                    except Exception:
                        pass
                    cls._cpu_threads = desired_threads

            _emit(log_emit, f"Loading LightOnOCR-2-1B on {device}...")

            model_kwargs = {"torch_dtype": dtype}
            # Reduce peak RAM during load if accelerate is available.
            try:
                import accelerate  # type: ignore  # noqa: F401

                model_kwargs["low_cpu_mem_usage"] = True
            except Exception:
                pass

            cls._processor = LightOnOcrProcessor.from_pretrained(cls._MODEL_ID)
            try:
                cls._model = LightOnOcrForConditionalGeneration.from_pretrained(cls._MODEL_ID, **model_kwargs).to(device)
            except TypeError:
                # Be tolerant of older/newer transformers signatures.
                model_kwargs.pop("low_cpu_mem_usage", None)
                cls._model = LightOnOcrForConditionalGeneration.from_pretrained(cls._MODEL_ID, **model_kwargs).to(device)
            except Exception as e:
                # If reduced precision fails on CPU, fall back to float32 compatibility mode.
                if _device_kind(device) == "cpu" and dtype != torch.float32:
                    _emit(
                        log_emit,
                        f"Load failed with dtype={dtype_str} on CPU; retrying float32 ({type(e).__name__}: {e})",
                    )
                    model_kwargs["torch_dtype"] = torch.float32
                    model_kwargs.pop("low_cpu_mem_usage", None)
                    cls._model = LightOnOcrForConditionalGeneration.from_pretrained(cls._MODEL_ID, **model_kwargs).to(device)
                    dtype = torch.float32
                else:
                    raise
            cls._model.eval()
            cls._device = device
            cls._dtype = dtype

            _emit(log_emit, "Model loaded successfully.")

        return cls

    @classmethod
    def run_ocr(cls, image, max_tokens: int = 8192) -> str:
        """Run OCR on a single image."""
        if cls._model is None:
            raise RuntimeError("Model not loaded. Call load() first.")

        import torch

        # Prepare conversation format for the model
        conversation = [
            {
                "role": "user",
                "content": [{"type": "image", "image": image}]
            }
        ]

        inputs = cls._processor.apply_chat_template(
            conversation,
            add_generation_prompt=True,
            tokenize=True,
            return_dict=True,
            return_tensors="pt",
        )

        # Move inputs to device with appropriate dtype
        inputs = {
            k: v.to(device=cls._device, dtype=cls._dtype) if v.is_floating_point() else v.to(cls._device)
            for k, v in inputs.items()
        }

        device_kind = _device_kind(cls._device)
        autocast_dtype_ok = cls._dtype in (torch.bfloat16, torch.float16)
        autocast_cm = (
            torch.autocast(device_type=device_kind, dtype=cls._dtype)
            if hasattr(torch, "autocast") and autocast_dtype_ok and device_kind in ("cpu", "cuda")
            else nullcontext()
        )

        with torch.inference_mode(), autocast_cm:
            output_ids = cls._model.generate(**inputs, max_new_tokens=max_tokens)
        generated_ids = output_ids[0, inputs["input_ids"].shape[1]:]
        output_text = cls._processor.decode(generated_ids, skip_special_tokens=True)

        return output_text

    @classmethod
    def unload(cls):
        """Unload model to free memory."""
        if cls._model is not None:
            del cls._model
            del cls._processor
            cls._model = None
            cls._processor = None
            cls._device = None
            cls._dtype = None

            # Clear CUDA cache if available
            try:
                import torch
                if torch.cuda.is_available():
                    torch.cuda.empty_cache()
            except ImportError:
                pass


def process_pdf(
    pdf_path: Path,
    device,
    dtype_str: str,
    max_tokens: int,
    render_size: int,
    cpu_threads: int | None,
    log_emit=None,
    progress_emit=None,
) -> Path:
    """
    Process a single PDF and return the output Excel path.

    Args:
        pdf_path: Path to input PDF
        device: Torch device string
        dtype_str: Dtype string for model (e.g. "bfloat16", "float32")
        max_tokens: Max tokens for OCR generation
        log_emit: Callback for log messages
        progress_emit: Callback for progress updates (current, total)

    Returns:
        Path to output Excel file
    """
    import pypdfium2 as pdfium

    _emit(log_emit, f"Processing: {pdf_path.name}")

    # Load model
    OCRModel.load(device, dtype_str, log_emit, cpu_threads=cpu_threads)

    # Open PDF and get page count
    pdf = pdfium.PdfDocument(pdf_path)
    total_pages = len(pdf)
    _emit(log_emit, f"PDF has {total_pages} page(s)")
    page_texts: List[str] = []

    if progress_emit:
        progress_emit(0, total_pages)

    for page_num in range(total_pages):
        _emit(log_emit, f"Processing page {page_num + 1}/{total_pages}...")

        if progress_emit:
            progress_emit(page_num + 1, total_pages)

        # Render page to image
        image = render_pdf_page(pdf, page_num, target_size=render_size)

        # Run OCR
        try:
            ocr_text = OCRModel.run_ocr(image, max_tokens)
        finally:
            try:
                image.close()
            except Exception:
                pass
        _emit(log_emit, f"Page {page_num + 1} OCR complete ({len(ocr_text)} chars)")

        page_texts.append(ocr_text)

    if progress_emit:
        progress_emit(total_pages, total_pages)

    try:
        close = getattr(pdf, "close", None)
        if callable(close):
            close()
    except Exception:
        pass

    # Generate output path
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_name = f"{pdf_path.stem}__ocr__{ts}.xlsx"
    out_path = ensure_unique_path(pdf_path.with_name(out_name))

    # Write to Excel (structured tables + key/value rows)
    wb = _build_workbook_from_pages(page_texts)
    wb.save(out_path)
    _emit(log_emit, f"Saved: {out_path.name}")

    return out_path


def process_files(
    files: List[str],
    device,
    dtype_str: str,
    max_tokens: int,
    render_size: int,
    cpu_threads: int | None,
    log_emit=None,
    progress_emit=None,
) -> Tuple[int, int, List[Path]]:
    """Process all files, returning (ok_count, fail_count, output_paths)."""
    ok = 0
    fail = 0
    outputs: List[Path] = []

    for raw in files:
        path = Path(raw)
        _emit(log_emit, f"[START] {path.name}")
        try:
            out = process_pdf(
                path,
                device,
                dtype_str,
                max_tokens,
                render_size,
                cpu_threads,
                log_emit,
                progress_emit,
            )
            outputs.append(out)
            ok += 1
            _emit(log_emit, f"[DONE] {path.name}")
        except Exception as e:
            fail += 1
            _emit(log_emit, f"[FAIL] {path.name}: {e}")

    return ok, fail, outputs


class MainWidget(QWidget):
    log_message = Signal(str)
    progress_update = Signal(int, int)
    processing_done = Signal(int, int, list)

    def __init__(self):
        super().__init__()
        self.setObjectName("lightonocr2_pdf_to_excel_widget")
        self._ram_gb = get_total_memory_gb()
        self._gpu_info = detect_gpu_info(total_ram_gb=self._ram_gb)
        self._build_ui()
        self._connect_signals()

    def _build_ui(self):
        # Description
        self.desc_label = QLabel("", self)
        self.desc_label.setWordWrap(True)
        self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.desc_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.desc_label.setStyleSheet(
            "color: #dcdcdc; background: transparent; padding: 6px; "
            "border: 1px solid #3a3a3a; border-radius: 6px;"
        )
        self._load_long_description()

        # Controls
        self.select_btn = PrimaryPushButton("Select PDF Files", self)
        self.run_btn = PrimaryPushButton("Run OCR", self)
        self.unload_btn = PrimaryPushButton("Unload Model", self)
        self.unload_btn.setToolTip("Free memory by unloading the model")

        # Device selection
        self.device_combo = QComboBox(self)
        for device_str, display_name, dtype, is_discrete in self._gpu_info["devices"]:
            self.device_combo.addItem(display_name, userData=(device_str, dtype))
        self.device_combo.setCurrentIndex(self._gpu_info["recommended"])
        self.device_label = QLabel("Compute Device", self)
        self.device_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        # Max tokens
        self.tokens_spin = QSpinBox(self)
        self.tokens_spin.setMinimum(256)
        self.tokens_spin.setMaximum(8192)
        self.tokens_spin.setSingleStep(256)
        self.tokens_label = QLabel("Max tokens per page", self)
        self.tokens_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        # Render resolution (longest side)
        self.render_spin = QSpinBox(self)
        self.render_spin.setMinimum(768)
        self.render_spin.setMaximum(2048)
        self.render_spin.setSingleStep(128)
        self.render_label = QLabel("Render size (px)", self)
        self.render_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        # CPU threads
        self.threads_spin = QSpinBox(self)
        self.threads_spin.setMinimum(1)
        self.threads_spin.setMaximum(max(1, os.cpu_count() or 1))
        self.threads_spin.setSingleStep(1)
        self.threads_label = QLabel("CPU threads", self)
        self.threads_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        # Apply hardware-aware defaults
        device_str, _dtype = self.device_combo.currentData()
        defaults = _recommended_defaults_for_device(device_str, self._ram_gb)
        self.tokens_spin.setValue(int(defaults["max_tokens"]))
        self.render_spin.setValue(int(defaults["render_size"]))
        self.threads_spin.setValue(int(defaults["cpu_threads"]))

        # Progress bar
        self.progress_bar = QProgressBar(self)
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat("Ready")

        # Selected files / log
        self.files_label = QLabel("Selected files", self)
        self.files_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.files_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")
        self.logs_label = QLabel("Process logs", self)
        self.logs_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.logs_label.setStyleSheet("color: #dcdcdc; background: transparent; padding-left: 2px;")

        self.files_box = QTextEdit(self)
        self.files_box.setReadOnly(True)
        self.files_box.setPlaceholderText("Selected PDF files will appear here")
        self.files_box.setStyleSheet(
            "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        self.log_box = QTextEdit(self)
        self.log_box.setReadOnly(True)
        self.log_box.setPlaceholderText("Live process log will appear here")
        self.log_box.setStyleSheet(
            "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
            "border: 1px solid #3a3a3a; border-radius: 6px;}"
        )

        # Layouts
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)

        main_layout.addWidget(self.desc_label)

        btn_row = QHBoxLayout()
        btn_row.addStretch(1)
        btn_row.addWidget(self.select_btn)
        btn_row.addStretch(1)
        main_layout.addLayout(btn_row)

        grid = QGridLayout()
        grid.setColumnStretch(1, 1)
        grid.addWidget(self.device_label, 0, 0, Qt.AlignLeft)
        grid.addWidget(self.device_combo, 0, 1, Qt.AlignLeft)
        grid.addWidget(self.tokens_label, 1, 0, Qt.AlignLeft)
        grid.addWidget(self.tokens_spin, 1, 1, Qt.AlignLeft)
        grid.addWidget(self.render_label, 2, 0, Qt.AlignLeft)
        grid.addWidget(self.render_spin, 2, 1, Qt.AlignLeft)
        grid.addWidget(self.threads_label, 3, 0, Qt.AlignLeft)
        grid.addWidget(self.threads_spin, 3, 1, Qt.AlignLeft)
        main_layout.addLayout(grid)

        run_row = QHBoxLayout()
        run_row.addStretch(1)
        run_row.addWidget(self.run_btn)
        run_row.addWidget(self.unload_btn)
        run_row.addStretch(1)
        main_layout.addLayout(run_row)

        main_layout.addWidget(self.progress_bar)

        labels_row = QHBoxLayout()
        labels_row.addWidget(self.files_label, 1)
        labels_row.addWidget(self.logs_label, 1)
        main_layout.addLayout(labels_row)

        boxes_row = QHBoxLayout()
        boxes_row.addWidget(self.files_box, 1)
        boxes_row.addWidget(self.log_box, 1)
        main_layout.addLayout(boxes_row)

    def _connect_signals(self):
        self.select_btn.clicked.connect(self.select_files)
        self.run_btn.clicked.connect(self.run_process)
        self.unload_btn.clicked.connect(self.unload_model)
        self.log_message.connect(self.append_log)
        self.progress_update.connect(self.update_progress)
        self.processing_done.connect(self.on_processing_done)

    def _load_long_description(self):
        md_path = Path(__file__).with_name("description.md")
        if not md_path.exists():
            return
        try:
            lines = md_path.read_text(encoding="utf-8").splitlines()
        except Exception:
            return
        body: List[str] = []
        for ln in lines:
            if ln.startswith(">"):
                continue
            body.append(ln)
        text = "\n".join(body).strip()
        if text:
            self.desc_label.setText(text)

    def select_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select PDF files",
            "",
            "PDF Files (*.pdf);;All Files (*)",
        )
        if files:
            self.files_box.setPlainText("\n".join(files))
        else:
            self.files_box.clear()

    def _selected_files(self) -> List[str]:
        text = self.files_box.toPlainText().strip()
        if not text:
            return []
        return [line for line in text.split("\n") if line.strip()]

    def run_process(self):
        files = self._selected_files()
        if not files:
            MessageBox("Warning", "No PDF files selected.", self).exec()
            return

        # Check dependencies before running
        deps_ok, deps_msg = check_dependencies()
        if not deps_ok:
            MessageBox("Missing Dependencies", deps_msg, self).exec()
            return

        # Get device configuration
        device_str, dtype_str = self.device_combo.currentData()
        max_tokens = int(self.tokens_spin.value())
        render_size = int(self.render_spin.value())
        cpu_threads = int(self.threads_spin.value())

        self.log_box.clear()
        self.log_message.emit(f"Starting OCR for {len(files)} file(s)...")
        self.log_message.emit(f"Using device: {self.device_combo.currentText()}")
        if self._ram_gb:
            self.log_message.emit(f"Detected RAM: {self._ram_gb:.1f} GB")
        self.log_message.emit(
            f"Settings: render={render_size}px, max_tokens={max_tokens}, cpu_threads={cpu_threads}"
        )

        self._set_controls_enabled(False)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("Starting...")

        def worker():
            try:
                ok, fail, outputs = process_files(
                    files,
                    device_str,
                    dtype_str,
                    max_tokens,
                    render_size,
                    cpu_threads,
                    log_emit=self.log_message.emit,
                    progress_emit=self.progress_update.emit,
                )
            except Exception as e:
                ok, fail, outputs = 0, len(files), []
                self.log_message.emit(f"Fatal error: {e}")
            self.processing_done.emit(ok, fail, outputs)

        threading.Thread(target=worker, daemon=True).start()

    def unload_model(self):
        """Unload model to free GPU memory."""
        OCRModel.unload()
        self.log_message.emit("Model unloaded. GPU memory freed.")

    def _set_controls_enabled(self, enabled: bool):
        self.run_btn.setEnabled(enabled)
        self.select_btn.setEnabled(enabled)
        self.device_combo.setEnabled(enabled)
        self.tokens_spin.setEnabled(enabled)
        self.render_spin.setEnabled(enabled)
        self.threads_spin.setEnabled(enabled)
        self.unload_btn.setEnabled(enabled)

    def append_log(self, text: str):
        self.log_box.append(text)
        self.log_box.ensureCursorVisible()

    def update_progress(self, current: int, total: int):
        if total > 0:
            percent = int((current / total) * 100)
            self.progress_bar.setValue(percent)
            self.progress_bar.setFormat(f"Page {current}/{total}")

    def on_processing_done(self, ok: int, fail: int, outputs: list):
        if outputs:
            self.log_message.emit("Generated files:")
            for p in outputs:
                self.log_message.emit(f" - {p}")
        self.log_message.emit(f"Completed: {ok} succeeded, {fail} failed.")

        self.progress_bar.setValue(100)
        self.progress_bar.setFormat("Done")
        self._set_controls_enabled(True)
        title = "Processing complete" if fail == 0 else "Processing finished with issues"
        lines = [f"Success: {ok}", f"Failed: {fail}"]
        if outputs:
            try:
                sample = outputs[:5]
                lines.append("")
                lines.append("Outputs:")
                for p in sample:
                    lines.append(f"- {os.path.basename(str(p))}")
                remaining = len(outputs) - len(sample)
                if remaining > 0:
                    lines.append(f"... and {remaining} more")
            except Exception:
                pass
        msg = MessageBox(title, "\n".join(lines), self)
        msg.yesButton.setText("OK")
        msg.cancelButton.hide()
        msg.exec()


def get_widget():
    return MainWidget()
