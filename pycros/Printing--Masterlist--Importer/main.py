#!/usr/bin/env python3
from __future__ import annotations

import argparse
import math
import os
import posixpath
import re
import threading
import zipfile
from copy import deepcopy
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional, Tuple
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils.datetime import time_to_days, timedelta_to_days, to_excel

try:
    from PySide6.QtCore import Qt, Signal
    from PySide6.QtWidgets import (
        QFileDialog,
        QHBoxLayout,
        QLabel,
        QSizePolicy,
        QTextEdit,
        QVBoxLayout,
        QWidget,
    )
    from qfluentwidgets import MessageBox, PrimaryPushButton
except Exception:
    Qt = None
    Signal = None
    QFileDialog = None
    QHBoxLayout = None
    QLabel = None
    QSizePolicy = None
    QTextEdit = None
    QVBoxLayout = None
    QWidget = object
    MessageBox = None
    PrimaryPushButton = None


NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
NS_CT = "http://schemas.openxmlformats.org/package/2006/content-types"
NS_RICH = "http://schemas.microsoft.com/office/spreadsheetml/2017/richdata"
NS_RICH_REL = "http://schemas.microsoft.com/office/spreadsheetml/2022/richvaluerel"
NS_XML = "http://www.w3.org/XML/1998/namespace"

IMAGE_REL_TYPE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"
MAIN = f"{{{NS_MAIN}}}"
REL = f"{{{NS_REL}}}"
PKG_REL = f"{{{NS_PKG_REL}}}"
RICH = f"{{{NS_RICH}}}"
RICH_REL = f"{{{NS_RICH_REL}}}"
XML_SPACE = f"{{{NS_XML}}}space"

HEADER_STYLE = "STYLE NO."
HEADER_LOGO_ARTWORK = "LOGO ARTWORK"
ERROR_LITERALS = {
    "#NULL!",
    "#DIV/0!",
    "#VALUE!",
    "#REF!",
    "#NAME?",
    "#NUM!",
    "#N/A",
    "#GETTING_DATA",
}

NS_REGISTRY = {
    "": NS_MAIN,
    "r": NS_REL,
    "mc": "http://schemas.openxmlformats.org/markup-compatibility/2006",
    "x14ac": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac",
    "x15": "http://schemas.microsoft.com/office/spreadsheetml/2010/11/main",
    "x16r2": "http://schemas.microsoft.com/office/spreadsheetml/2015/02/main",
    "xr": "http://schemas.microsoft.com/office/spreadsheetml/2014/revision",
    "xr2": "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2",
    "xr3": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3",
    "xr6": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision6",
    "xr10": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision10",
    "xlrd": NS_RICH,
}
for _prefix, _uri in NS_REGISTRY.items():
    ET.register_namespace(_prefix, _uri)


def _emit(log_emit, text: str):
    if callable(log_emit):
        try:
            log_emit(text)
            return
        except Exception:
            pass
    print(text)


def norm_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").replace("_x000a_", " ")
    return re.sub(r"\s+", " ", text).strip().upper()


def has_value(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, float) and math.isnan(value):
        return False
    return str(value).strip() != ""


def clean_xml_text(value: str) -> str:
    return "".join(
        ch
        for ch in value
        if ch in "\t\n\r" or ord(ch) >= 0x20
    )


def _with_excel_xml_declaration(body: str) -> bytes:
    return ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n' + body).encode("UTF-8")


def _restore_ignorable_namespace_declarations(text: str) -> str:
    match = re.search(r"<(?!\?)([^>\s]+)([^>]*)>", text)
    if not match:
        return text

    root_start = match.group(0)
    ignorable_match = re.search(r"\bmc:Ignorable=\"([^\"]+)\"", root_start)
    if not ignorable_match:
        return text

    missing: List[str] = []
    for prefix in ignorable_match.group(1).split():
        uri = NS_REGISTRY.get(prefix)
        if uri and f"xmlns:{prefix}=" not in root_start:
            missing.append(f' xmlns:{prefix}="{uri}"')
    if not missing:
        return text

    insert_at = match.end() - (2 if root_start.endswith("/>") else 1)
    return text[:insert_at] + "".join(missing) + text[insert_at:]


def _force_default_namespace(text: str, uri: str, root_name: str, child_names: Iterable[str]) -> str:
    prefix_match = re.search(
        rf"<([A-Za-z_][\w.-]*):{re.escape(root_name)}\s+xmlns:\1=\"{re.escape(uri)}\"",
        text,
    )
    if not prefix_match:
        return text

    prefix = prefix_match.group(1)
    text = text.replace(f"<{prefix}:{root_name} xmlns:{prefix}=\"{uri}\"", f"<{root_name} xmlns=\"{uri}\"", 1)
    text = text.replace(f"</{prefix}:{root_name}>", f"</{root_name}>")
    for child_name in child_names:
        text = text.replace(f"<{prefix}:{child_name}", f"<{child_name}")
        text = text.replace(f"</{prefix}:{child_name}>", f"</{child_name}>")
    return text


def xml_bytes(root: ET.Element) -> bytes:
    body = ET.tostring(root, encoding="unicode")
    body = _restore_ignorable_namespace_declarations(body)

    if root.tag == f"{PKG_REL}Relationships":
        body = _force_default_namespace(body, NS_PKG_REL, "Relationships", ["Relationship"])
    elif root.tag == f"{RICH_REL}richValueRels":
        body = _force_default_namespace(body, NS_RICH_REL, "richValueRels", ["rel"])
    elif root.tag == f"{RICH}rvData":
        body = _force_default_namespace(body, NS_RICH, "rvData", ["rv", "v"])
    elif root.tag == f"{{{NS_CT}}}Types":
        body = _force_default_namespace(body, NS_CT, "Types", ["Default", "Override"])

    return _with_excel_xml_declaration(body)


def parse_ref(ref: str) -> Tuple[int, int]:
    match = re.match(r"^([A-Z]+)(\d+)$", ref)
    if not match:
        raise ValueError(f"Invalid cell reference: {ref}")
    return int(match.group(2)), column_index_from_string(match.group(1))


def split_range(ref: str) -> Tuple[str, int, str, int]:
    start, end = ref.split(":", 1)
    start_col = re.match(r"^([A-Z]+)", start).group(1)
    start_row = int(re.search(r"(\d+)$", start).group(1))
    end_col = re.match(r"^([A-Z]+)", end).group(1)
    end_row = int(re.search(r"(\d+)$", end).group(1))
    return start_col, start_row, end_col, end_row


def rebuild_range(ref: str, end_row: int) -> str:
    start_col, start_row, end_col, old_end_row = split_range(ref)
    return f"{start_col}{start_row}:{end_col}{max(old_end_row, end_row)}"


def normalize_target(target: str, base_dir: str = "xl") -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(base_dir, target))


def worksheet_xml_path(xlsx_path: Path, sheet_name: str) -> str:
    with zipfile.ZipFile(xlsx_path) as archive:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

    rels = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels_root.findall(f"{PKG_REL}Relationship")
    }
    for sheet in workbook_root.findall(f"{MAIN}sheets/{MAIN}sheet"):
        if sheet.attrib.get("name") == sheet_name:
            rel_id = sheet.attrib.get(f"{REL}id")
            target = rels.get(rel_id)
            if not target:
                raise ValueError(f"Worksheet relationship missing for sheet: {sheet_name}")
            return normalize_target(target, "xl")
    raise ValueError(f"Sheet not found: {sheet_name}")


def find_header_row_and_map(ws, required_header: str = HEADER_STYLE) -> Tuple[int, Dict[str, int]]:
    best_row = 0
    best_map: Dict[str, int] = {}
    best_score = -1
    max_scan_row = min(ws.max_row or 0, 25)
    for row_idx, row_values in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_row, values_only=True),
        start=1,
    ):
        col_map: Dict[str, int] = {}
        for col, value in enumerate(row_values, start=1):
            key = norm_header(value)
            if key and key not in col_map:
                col_map[key] = col
        if norm_header(required_header) not in col_map:
            continue
        score = len(col_map)
        if score > best_score:
            best_row = row_idx
            best_map = col_map
            best_score = score
    if not best_row:
        raise ValueError(f"Header row with '{required_header}' was not found in sheet '{ws.title}'.")
    return best_row, best_map


def first_blank_style_row(ws, header_row: int, style_col: int) -> int:
    for row_idx, (value,) in enumerate(
        ws.iter_rows(
            min_row=header_row + 1,
            max_row=ws.max_row or header_row,
            min_col=style_col,
            max_col=style_col,
            values_only=True,
        ),
        start=header_row + 1,
    ):
        if not has_value(value):
            return row_idx
    return (ws.max_row or header_row) + 1


def load_source_rows(source_path: Path, log_emit=None) -> Tuple[str, int, Dict[str, int], List[int], Dict[Tuple[int, int], object]]:
    wb = load_workbook(source_path, data_only=True, read_only=True)
    try:
        for ws in wb.worksheets:
            try:
                header_row, header_map = find_header_row_and_map(ws)
                break
            except ValueError:
                continue
        else:
            raise ValueError("No worksheet with a STYLE NO. header was found.")

        style_col = header_map[norm_header(HEADER_STYLE)]
        wanted_cols = sorted(set(header_map.values()))
        wanted_col_set = set(wanted_cols)
        rows: List[int] = []
        values: Dict[Tuple[int, int], object] = {}
        for row_idx, row_values in enumerate(
            ws.iter_rows(
                min_row=header_row + 1,
                max_row=ws.max_row or header_row,
                max_col=max(wanted_cols) if wanted_cols else ws.max_column,
                values_only=True,
            ),
            start=header_row + 1,
        ):
            style_value = row_values[style_col - 1] if style_col <= len(row_values) else None
            if not has_value(style_value):
                continue
            rows.append(row_idx)
            for col in wanted_col_set:
                values[(row_idx, col)] = row_values[col - 1] if col <= len(row_values) else None

        _emit(log_emit, f"{source_path.name}: found {len(rows)} source row(s) in '{ws.title}'.")
        return ws.title, header_row, header_map, rows, values
    finally:
        wb.close()


def load_master_layout(master_path: Path) -> Tuple[str, int, Dict[str, int], int]:
    wb = load_workbook(master_path, data_only=True, read_only=True)
    try:
        for ws in wb.worksheets:
            try:
                header_row, header_map = find_header_row_and_map(ws)
                break
            except ValueError:
                continue
        else:
            raise ValueError("No worksheet with a STYLE NO. header was found in master workbook.")
        style_col = header_map[norm_header(HEADER_STYLE)]
        append_row = first_blank_style_row(ws, header_row, style_col)
        return ws.title, header_row, header_map, append_row
    finally:
        wb.close()


class SourceImageMap:
    def __init__(self, source_path: Path, sheet_name: str):
        self.source_path = source_path
        self.sheet_path = worksheet_xml_path(source_path, sheet_name)
        self._cells: Dict[Tuple[int, int], Tuple[bytes, str, ET.Element]] = {}
        self._load()

    def get(self, row: int, col: int) -> Optional[Tuple[bytes, str, ET.Element]]:
        return self._cells.get((row, col))

    def _load(self):
        with zipfile.ZipFile(self.source_path) as archive:
            names = set(archive.namelist())
            required = {
                "xl/metadata.xml",
                "xl/richData/rdrichvalue.xml",
                "xl/richData/richValueRel.xml",
                "xl/richData/_rels/richValueRel.xml.rels",
                self.sheet_path,
            }
            if not required.issubset(names):
                return

            sheet_root = ET.fromstring(archive.read(self.sheet_path))
            metadata_root = ET.fromstring(archive.read("xl/metadata.xml"))
            rv_root = ET.fromstring(archive.read("xl/richData/rdrichvalue.xml"))
            rich_rel_root = ET.fromstring(archive.read("xl/richData/richValueRel.xml"))
            rels_root = ET.fromstring(archive.read("xl/richData/_rels/richValueRel.xml.rels"))

            value_metadata = metadata_root.find(f"{MAIN}valueMetadata")
            if value_metadata is None:
                return
            vm_to_rv: Dict[int, int] = {}
            for index, bk in enumerate(value_metadata.findall(f"{MAIN}bk"), start=1):
                rc = bk.find(f"{MAIN}rc")
                if rc is None:
                    continue
                try:
                    vm_to_rv[index] = int(rc.attrib.get("v", ""))
                except ValueError:
                    continue

            rv_entries = rv_root.findall(f"{RICH}rv")
            rich_rels = rich_rel_root.findall(f"{RICH_REL}rel")
            rel_targets = {
                rel.attrib.get("Id"): rel.attrib.get("Target")
                for rel in rels_root.findall(f"{PKG_REL}Relationship")
                if rel.attrib.get("Type") == IMAGE_REL_TYPE
            }

            for cell in sheet_root.findall(f".//{MAIN}c"):
                vm_text = cell.attrib.get("vm")
                ref = cell.attrib.get("r")
                if not vm_text or not ref:
                    continue
                try:
                    vm_index = int(vm_text)
                    rv_index = vm_to_rv[vm_index]
                    rv = rv_entries[rv_index]
                    rel_index = int(rv.findall(f"{RICH}v")[0].text)
                    rel_id = rich_rels[rel_index].attrib.get(f"{REL}id")
                    target = rel_targets[rel_id]
                except Exception:
                    continue

                image_path = normalize_target(target, "xl/richData")
                if image_path not in names:
                    continue
                row, col = parse_ref(ref)
                ext = Path(image_path).suffix.lower() or ".png"
                self._cells[(row, col)] = (archive.read(image_path), ext, deepcopy(rv))


class PackageEditor:
    def __init__(self, master_path: Path, sheet_name: str):
        self.master_path = master_path
        self.sheet_name = sheet_name
        self.parts: Dict[str, bytes] = {}
        self.order: List[str] = []
        self.new_parts: List[str] = []
        self.sheet_path = worksheet_xml_path(master_path, sheet_name)
        self.sheet_root: ET.Element
        self.sheet_data: ET.Element
        self.table_paths: List[str] = []
        self.workbook_root: Optional[ET.Element] = None
        self.max_written_row = 0
        self._wrap_style_cache: Dict[int, int] = {}
        self._load()

    def _load(self):
        with zipfile.ZipFile(self.master_path) as archive:
            for info in archive.infolist():
                if info.filename not in self.parts:
                    self.order.append(info.filename)
                self.parts[info.filename] = archive.read(info.filename)

        self.sheet_root = ET.fromstring(self.parts[self.sheet_path])
        self.sheet_data = self.sheet_root.find(f"{MAIN}sheetData")
        if self.sheet_data is None:
            raise ValueError("Master worksheet has no sheetData.")
        self.table_paths = sorted(name for name in self.parts if name.startswith("xl/tables/") and name.endswith(".xml"))
        if "xl/workbook.xml" in self.parts:
            self.workbook_root = ET.fromstring(self.parts["xl/workbook.xml"])

    def existing_row_numbers(self) -> List[int]:
        rows = []
        for row in self.sheet_data.findall(f"{MAIN}row"):
            try:
                rows.append(int(row.attrib["r"]))
            except Exception:
                continue
        return rows

    def _row_map(self) -> Dict[int, ET.Element]:
        return {
            int(row.attrib["r"]): row
            for row in self.sheet_data.findall(f"{MAIN}row")
            if row.attrib.get("r", "").isdigit()
        }

    def template_row(self, preferred_row: int) -> ET.Element:
        rows = self._row_map()
        if preferred_row in rows:
            return rows[preferred_row]
        row_numbers = sorted(rows)
        if not row_numbers:
            raise ValueError("Master worksheet has no rows to use as a template.")
        lower = [row for row in row_numbers if row < preferred_row]
        return rows[lower[-1] if lower else row_numbers[-1]]

    def ensure_row(self, row_num: int, preferred_template_row: int) -> ET.Element:
        rows = self._row_map()
        if row_num in rows:
            self.max_written_row = max(self.max_written_row, row_num)
            return rows[row_num]

        row = deepcopy(self.template_row(preferred_template_row))
        row.attrib["r"] = str(row_num)
        for cell in row.findall(f"{MAIN}c"):
            old_ref = cell.attrib.get("r")
            if old_ref:
                _, col = parse_ref(old_ref)
                cell.attrib["r"] = f"{get_column_letter(col)}{row_num}"
            self.clear_cell(cell)

        children = list(self.sheet_data)
        insert_at = len(children)
        for idx, child in enumerate(children):
            if child.tag != f"{MAIN}row":
                continue
            try:
                child_row = int(child.attrib.get("r", "0"))
            except ValueError:
                continue
            if child_row > row_num:
                insert_at = idx
                break
        self.sheet_data.insert(insert_at, row)
        self.max_written_row = max(self.max_written_row, row_num)
        return row

    def ensure_cell(self, row: ET.Element, row_num: int, col_num: int, preferred_template_row: int) -> ET.Element:
        target_ref = f"{get_column_letter(col_num)}{row_num}"
        for cell in row.findall(f"{MAIN}c"):
            if cell.attrib.get("r") == target_ref:
                return cell

        template = self.template_row(preferred_template_row)
        template_attrs: Dict[str, str] = {}
        for cell in template.findall(f"{MAIN}c"):
            ref = cell.attrib.get("r")
            if ref and parse_ref(ref)[1] == col_num:
                template_attrs = {
                    key: value
                    for key, value in cell.attrib.items()
                    if key not in {"r", "t", "vm", "cm"}
                }
                break

        new_cell = ET.Element(f"{MAIN}c", {"r": target_ref, **template_attrs})
        cells = row.findall(f"{MAIN}c")
        insert_at = len(row)
        cell_positions = [list(row).index(cell) for cell in cells]
        for idx, cell in enumerate(cells):
            ref = cell.attrib.get("r")
            if ref and parse_ref(ref)[1] > col_num:
                insert_at = cell_positions[idx]
                break
        row.insert(insert_at, new_cell)
        return new_cell

    @staticmethod
    def clear_cell(cell: ET.Element):
        for child in list(cell):
            cell.remove(child)
        for attr in ("t", "vm", "cm"):
            cell.attrib.pop(attr, None)

    def set_cell_value(self, row_num: int, col_num: int, value: object, preferred_template_row: int):
        row = self.ensure_row(row_num, preferred_template_row)
        cell = self.ensure_cell(row, row_num, col_num, preferred_template_row)
        self.clear_cell(cell)

        if not has_value(value):
            return

        if isinstance(value, bool):
            cell.attrib["t"] = "b"
            ET.SubElement(cell, f"{MAIN}v").text = "1" if value else "0"
        elif isinstance(value, datetime):
            ET.SubElement(cell, f"{MAIN}v").text = str(to_excel(value))
        elif isinstance(value, date):
            ET.SubElement(cell, f"{MAIN}v").text = str(to_excel(value))
        elif isinstance(value, time):
            ET.SubElement(cell, f"{MAIN}v").text = str(time_to_days(value))
        elif isinstance(value, timedelta):
            ET.SubElement(cell, f"{MAIN}v").text = str(timedelta_to_days(value))
        elif isinstance(value, (int, float)) and not isinstance(value, bool):
            if isinstance(value, float) and math.isnan(value):
                return
            text = str(value)
            if isinstance(value, float) and value.is_integer():
                text = str(int(value))
            ET.SubElement(cell, f"{MAIN}v").text = text
        else:
            text = clean_xml_text(str(value))
            if text.upper() in ERROR_LITERALS:
                cell.attrib["t"] = "e"
                ET.SubElement(cell, f"{MAIN}v").text = text.upper()
                return
            cell.attrib["t"] = "inlineStr"
            inline = ET.SubElement(cell, f"{MAIN}is")
            text_el = ET.SubElement(inline, f"{MAIN}t")
            text_el.text = text
            if text != text.strip() or "\n" in text or "\r" in text:
                text_el.attrib[XML_SPACE] = "preserve"
            self.apply_wrap_text(cell)

    def set_image_cell(
        self,
        row_num: int,
        col_num: int,
        image_bytes: bytes,
        image_ext: str,
        source_rv: ET.Element,
        preferred_template_row: int,
    ):
        vm_index = self.append_rich_image(image_bytes, image_ext, source_rv)
        row = self.ensure_row(row_num, preferred_template_row)
        cell = self.ensure_cell(row, row_num, col_num, preferred_template_row)
        self.clear_cell(cell)
        cell.attrib["t"] = "e"
        cell.attrib["vm"] = str(vm_index)
        ET.SubElement(cell, f"{MAIN}v").text = "#VALUE!"

    def apply_wrap_text(self, cell: ET.Element):
        style_text = cell.attrib.get("s", "0")
        try:
            style_idx = int(style_text)
        except ValueError:
            style_idx = 0
        wrap_style_idx = self.wrap_style_index(style_idx)
        cell.attrib["s"] = str(wrap_style_idx)

    def wrap_style_index(self, style_idx: int) -> int:
        if style_idx in self._wrap_style_cache:
            return self._wrap_style_cache[style_idx]
        if "xl/styles.xml" not in self.parts:
            return style_idx

        styles_root = ET.fromstring(self.parts["xl/styles.xml"])
        cell_xfs = styles_root.find(f"{MAIN}cellXfs")
        if cell_xfs is None:
            return style_idx

        xfs = list(cell_xfs.findall(f"{MAIN}xf"))
        if style_idx < 0 or style_idx >= len(xfs):
            return style_idx

        original = xfs[style_idx]
        alignment = original.find(f"{MAIN}alignment")
        if alignment is not None and alignment.attrib.get("wrapText") in {"1", "true", "True"}:
            self._wrap_style_cache[style_idx] = style_idx
            return style_idx

        new_xf = deepcopy(original)
        new_alignment = new_xf.find(f"{MAIN}alignment")
        if new_alignment is None:
            new_alignment = ET.Element(f"{MAIN}alignment")
            insert_at = len(list(new_xf))
            for idx, child in enumerate(list(new_xf)):
                if child.tag in {f"{MAIN}protection", f"{MAIN}extLst"}:
                    insert_at = idx
                    break
            new_xf.insert(insert_at, new_alignment)
        new_alignment.attrib["wrapText"] = "1"
        new_xf.attrib["applyAlignment"] = "1"

        new_style_idx = len(xfs)
        cell_xfs.append(new_xf)
        cell_xfs.attrib["count"] = str(new_style_idx + 1)
        self.parts["xl/styles.xml"] = xml_bytes(styles_root)
        self._wrap_style_cache[style_idx] = new_style_idx
        return new_style_idx

    def append_rich_image(self, image_bytes: bytes, image_ext: str, source_rv: ET.Element) -> int:
        required = [
            "xl/metadata.xml",
            "xl/richData/rdrichvalue.xml",
            "xl/richData/richValueRel.xml",
            "xl/richData/_rels/richValueRel.xml.rels",
        ]
        missing = [name for name in required if name not in self.parts]
        if missing:
            raise ValueError(
                "Master workbook has no in-cell image richData structure; missing "
                + ", ".join(missing)
            )

        self.ensure_image_content_type(image_ext)

        media_name = self.next_media_name(image_ext)
        self.parts[media_name] = image_bytes
        self.new_parts.append(media_name)

        rich_rel_root = ET.fromstring(self.parts["xl/richData/richValueRel.xml"])
        rels_root = ET.fromstring(self.parts["xl/richData/_rels/richValueRel.xml.rels"])
        rv_root = ET.fromstring(self.parts["xl/richData/rdrichvalue.xml"])
        metadata_root = ET.fromstring(self.parts["xl/metadata.xml"])

        rel_ids = [
            rel.attrib.get("Id", "")
            for rel in rels_root.findall(f"{PKG_REL}Relationship")
        ]
        next_id_num = 1
        for rel_id in rel_ids:
            match = re.match(r"rId(\d+)$", rel_id)
            if match:
                next_id_num = max(next_id_num, int(match.group(1)) + 1)
        rel_id = f"rId{next_id_num}"

        rich_rel_count = len(rich_rel_root.findall(f"{RICH_REL}rel"))
        ET.SubElement(rich_rel_root, f"{RICH_REL}rel", {f"{REL}id": rel_id})
        ET.SubElement(
            rels_root,
            f"{PKG_REL}Relationship",
            {
                "Id": rel_id,
                "Type": IMAGE_REL_TYPE,
                "Target": f"../media/{Path(media_name).name}",
            },
        )

        new_rv_index = len(rv_root.findall(f"{RICH}rv"))
        new_rv = deepcopy(source_rv)
        values = new_rv.findall(f"{RICH}v")
        if not values:
            new_rv = ET.Element(f"{RICH}rv", {"s": source_rv.attrib.get("s", "0")})
            values = [ET.SubElement(new_rv, f"{RICH}v"), ET.SubElement(new_rv, f"{RICH}v")]
            values[1].text = "5"
        values[0].text = str(rich_rel_count)
        rv_root.append(new_rv)
        rv_root.attrib["count"] = str(new_rv_index + 1)

        future_metadata = metadata_root.find(f"{MAIN}futureMetadata[@name='XLRICHVALUE']")
        value_metadata = metadata_root.find(f"{MAIN}valueMetadata")
        if future_metadata is None or value_metadata is None:
            raise ValueError("Master metadata.xml is missing XLRICHVALUE metadata.")

        bk = ET.SubElement(future_metadata, f"{MAIN}bk")
        ext_lst = ET.SubElement(bk, f"{MAIN}extLst")
        ext = ET.SubElement(ext_lst, f"{MAIN}ext", {"uri": "{3e2802c4-a4d2-4d8b-9148-e3be6c30e623}"})
        ET.SubElement(ext, f"{RICH}rvb", {"i": str(new_rv_index)})
        future_metadata.attrib["count"] = str(len(future_metadata.findall(f"{MAIN}bk")))

        vm_zero_based = len(value_metadata.findall(f"{MAIN}bk"))
        value_bk = ET.SubElement(value_metadata, f"{MAIN}bk")
        ET.SubElement(value_bk, f"{MAIN}rc", {"t": "1", "v": str(new_rv_index)})
        value_metadata.attrib["count"] = str(vm_zero_based + 1)

        self.parts["xl/richData/richValueRel.xml"] = xml_bytes(rich_rel_root)
        self.parts["xl/richData/_rels/richValueRel.xml.rels"] = xml_bytes(rels_root)
        self.parts["xl/richData/rdrichvalue.xml"] = xml_bytes(rv_root)
        self.parts["xl/metadata.xml"] = xml_bytes(metadata_root)
        return vm_zero_based + 1

    def ensure_image_content_type(self, image_ext: str):
        ext = image_ext.lower().lstrip(".")
        content_type = {
            "png": "image/png",
            "jpg": "image/jpeg",
            "jpeg": "image/jpeg",
            "gif": "image/gif",
            "bmp": "image/bmp",
            "webp": "image/webp",
        }.get(ext, f"image/{ext}")

        root = ET.fromstring(self.parts["[Content_Types].xml"])
        for default in root.findall(f"{{{NS_CT}}}Default"):
            if default.attrib.get("Extension", "").lower() == ext:
                return
        ET.SubElement(root, f"{{{NS_CT}}}Default", {"Extension": ext, "ContentType": content_type})
        self.parts["[Content_Types].xml"] = xml_bytes(root)

    def next_media_name(self, image_ext: str) -> str:
        ext = image_ext.lower()
        if not ext.startswith("."):
            ext = f".{ext}"
        used = set(self.parts)
        highest = 0
        for name in used:
            match = re.match(r"^xl/media/image(\d+)\.[A-Za-z0-9]+$", name)
            if match:
                highest = max(highest, int(match.group(1)))
        n = highest + 1
        while True:
            candidate = f"xl/media/image{n}{ext}"
            if candidate not in used:
                return candidate
            n += 1

    def update_ranges(self):
        if self.max_written_row <= 0:
            return

        dimension = self.sheet_root.find(f"{MAIN}dimension")
        if dimension is not None and ":" in dimension.attrib.get("ref", ""):
            dimension.attrib["ref"] = rebuild_range(dimension.attrib["ref"], self.max_written_row)
        self.parts[self.sheet_path] = xml_bytes(self.sheet_root)

        for table_path in self.table_paths:
            root = ET.fromstring(self.parts[table_path])
            ref = root.attrib.get("ref", "")
            if ":" in ref:
                root.attrib["ref"] = rebuild_range(ref, self.max_written_row)
            auto_filter = root.find(f"{MAIN}autoFilter")
            if auto_filter is not None and ":" in auto_filter.attrib.get("ref", ""):
                auto_filter.attrib["ref"] = rebuild_range(auto_filter.attrib["ref"], self.max_written_row)
            self.parts[table_path] = xml_bytes(root)

        if self.workbook_root is not None:
            changed = False
            for defined_name in self.workbook_root.findall(f"{MAIN}definedNames/{MAIN}definedName"):
                text = defined_name.text or ""
                if self.sheet_name not in text or ":" not in text:
                    continue
                new_text = re.sub(
                    r"(\$[A-Z]+\$\d+:\$[A-Z]+\$)(\d+)",
                    lambda m: f"{m.group(1)}{max(int(m.group(2)), self.max_written_row)}",
                    text,
                )
                if new_text != text:
                    defined_name.text = new_text
                    changed = True
            if changed:
                self.parts["xl/workbook.xml"] = xml_bytes(self.workbook_root)

    def write(self, output_path: Path):
        self.update_ranges()
        written = set()
        with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for name in self.order:
                if name in self.parts and name not in written:
                    archive.writestr(name, self.parts[name])
                    written.add(name)
            for name in sorted(self.parts):
                if name not in written:
                    archive.writestr(name, self.parts[name])
                    written.add(name)


def unique_output_path(master_path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    candidate = master_path.with_name(f"{timestamp}_imported_{master_path.name}")
    if not candidate.exists():
        return candidate
    stem = candidate.stem
    suffix = candidate.suffix
    for n in range(1, 1000):
        next_candidate = candidate.with_name(f"{stem} ({n}){suffix}")
        if not next_candidate.exists():
            return next_candidate
    raise FileExistsError("Could not generate a unique output file name.")


def process_import(source_paths: Iterable[str], master_path: str, output_path: Optional[str] = None, log_emit=None) -> Tuple[str, int]:
    sources = [Path(path) for path in source_paths if str(path).strip()]
    master = Path(master_path)
    if not sources:
        raise ValueError("No A source files selected.")
    if not master.exists():
        raise FileNotFoundError(f"Master B file not found: {master}")
    for source in sources:
        if not source.exists():
            raise FileNotFoundError(f"A source file not found: {source}")

    sheet_name, master_header_row, master_headers, append_row = load_master_layout(master)
    master_sheet_path = worksheet_xml_path(master, sheet_name)
    output = Path(output_path) if output_path else unique_output_path(master)
    editor = PackageEditor(master, sheet_name)
    _emit(log_emit, f"Master: {master.name}")
    _emit(log_emit, f"Sheet: {sheet_name} ({master_sheet_path})")
    _emit(log_emit, f"Append starts at row {append_row} based on empty STYLE NO.")

    source_payloads = []
    total_rows = 0
    for source in sources:
        sheet, header_row, source_headers, rows, values = load_source_rows(source, log_emit=log_emit)
        common = {
            source_col: master_headers[key]
            for key, source_col in source_headers.items()
            if key in master_headers
        }
        missing = [key for key in source_headers if key not in master_headers]
        if missing:
            _emit(log_emit, f"{source.name}: skipped {len(missing)} header(s) missing in master: {', '.join(missing)}")
        if norm_header(HEADER_STYLE) not in source_headers:
            raise ValueError(f"{source.name}: missing STYLE NO. header.")
        source_payloads.append((source, sheet, source_headers, rows, values, common))
        total_rows += len(rows)

    if total_rows == 0:
        raise ValueError("No source rows with STYLE NO. filled.")

    logo_key = norm_header(HEADER_LOGO_ARTWORK)
    target_row = append_row
    imported = 0
    image_count = 0
    template_row = append_row

    for source, sheet, source_headers, rows, values, common in source_payloads:
        images = SourceImageMap(source, sheet)
        logo_source_col = source_headers.get(logo_key)
        logo_target_col = master_headers.get(logo_key)

        for source_row in rows:
            for source_col, target_col in common.items():
                if source_col == logo_source_col and logo_target_col:
                    image = images.get(source_row, source_col)
                    if image:
                        img_bytes, img_ext, rv = image
                        editor.set_image_cell(target_row, target_col, img_bytes, img_ext, rv, template_row)
                        image_count += 1
                        continue
                editor.set_cell_value(
                    target_row,
                    target_col,
                    values.get((source_row, source_col)),
                    template_row,
                )
            _emit(log_emit, f"{source.name}: row {source_row} -> master row {target_row}")
            imported += 1
            target_row += 1

    editor.write(output)
    _emit(log_emit, f"Imported rows: {imported}")
    _emit(log_emit, f"Transferred in-cell images: {image_count}")
    _emit(log_emit, f"Output: {output}")
    return str(output), imported


def main() -> None:
    parser = argparse.ArgumentParser(description="Import BDD Strike Off Request rows into Print Strike Off Tracker.")
    parser.add_argument("--master", "-m", required=True, help="B master tracker workbook")
    parser.add_argument("--output", "-o", help="Output workbook path")
    parser.add_argument("sources", nargs="+", help="A request workbook(s)")
    args = parser.parse_args()

    try:
        output, count = process_import(args.sources, args.master, args.output)
        print(f"Done. Imported {count} row(s).")
        print(f"Output: {output}")
    except Exception as exc:
        print(f"Error: {exc}")
        raise SystemExit(1)


if __name__ == "__main__":
    main()


if Signal is not None:
    class MainWidget(QWidget):
        log_message = Signal(str)
        processing_done = Signal(int, str, str)

        def __init__(self):
            super().__init__()
            self.setObjectName("printing_masterlist_importer_widget")
            self._build_ui()
            self._connect_signals()

        def _build_ui(self):
            layout = QVBoxLayout(self)
            layout.setContentsMargins(12, 12, 12, 12)
            layout.setSpacing(12)

            self.desc_label = QLabel("", self)
            self.desc_label.setWordWrap(True)
            self.desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
            self.desc_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
            self.desc_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
            self.desc_label.setStyleSheet(
                "color: #dcdcdc; background: transparent; padding: 6px; "
                "border: 1px solid #3a3a3a; border-radius: 6px;"
            )
            self.set_long_description("")

            self.select_sources_btn = PrimaryPushButton("Select A Request Files", self)
            self.select_master_btn = PrimaryPushButton("Select B Masterlist", self)
            self.run_btn = PrimaryPushButton("Run Import", self)

            button_row = QHBoxLayout()
            button_row.addWidget(self.select_sources_btn, 1)
            button_row.addWidget(self.select_master_btn, 1)
            button_row.addWidget(self.run_btn, 1)

            self.sources_box = QTextEdit(self)
            self.sources_box.setReadOnly(True)
            self.sources_box.setPlaceholderText("Selected A files will appear here")
            self.sources_box.setStyleSheet(
                "QTextEdit{background: #2a2a2a; color: white; "
                "border: 1px solid #3a3a3a; border-radius: 6px;}"
            )

            self.master_box = QTextEdit(self)
            self.master_box.setReadOnly(True)
            self.master_box.setMaximumHeight(60)
            self.master_box.setPlaceholderText("Selected B masterlist file will appear here")
            self.master_box.setStyleSheet(
                "QTextEdit{background: #2a2a2a; color: white; "
                "border: 1px solid #3a3a3a; border-radius: 6px;}"
            )

            self.log_box = QTextEdit(self)
            self.log_box.setReadOnly(True)
            self.log_box.setPlaceholderText("Live process log will appear here")
            self.log_box.setStyleSheet(
                "QTextEdit{background: #1f1f1f; color: #d0d0d0; "
                "border: 1px solid #3a3a3a; border-radius: 6px;}"
            )

            label_sources = QLabel("A Request Files:", self)
            label_sources.setStyleSheet("color: #dcdcdc; background: transparent;")
            label_master = QLabel("B Masterlist:", self)
            label_master.setStyleSheet("color: #dcdcdc; background: transparent;")
            label_logs = QLabel("Process Log:", self)
            label_logs.setStyleSheet("color: #dcdcdc; background: transparent;")

            layout.addWidget(self.desc_label, 0)
            layout.addLayout(button_row, 0)
            layout.addWidget(label_sources, 0)
            layout.addWidget(self.sources_box, 1)
            layout.addWidget(label_master, 0)
            layout.addWidget(self.master_box, 0)
            layout.addWidget(label_logs, 0)
            layout.addWidget(self.log_box, 2)

        def _connect_signals(self):
            self.select_sources_btn.clicked.connect(self.select_sources)
            self.select_master_btn.clicked.connect(self.select_master)
            self.run_btn.clicked.connect(self.run_process)
            self.log_message.connect(self.append_log)
            self.processing_done.connect(self.on_processing_done)

        def set_long_description(self, text: str):
            clean = (text or "").strip()
            if clean:
                self.desc_label.setText(clean)
                self.desc_label.show()
            else:
                self.desc_label.clear()
                self.desc_label.hide()

        def select_sources(self):
            files, _ = QFileDialog.getOpenFileNames(
                self,
                "Select A Request Excel files",
                "",
                "Excel Workbooks (*.xlsx *.xlsm)",
            )
            self.sources_box.setPlainText("\n".join(files) if files else "")

        def select_master(self):
            file, _ = QFileDialog.getOpenFileName(
                self,
                "Select B Masterlist Excel file",
                "",
                "Excel Workbooks (*.xlsx *.xlsm)",
            )
            self.master_box.setPlainText(file or "")

        def selected_sources(self) -> List[str]:
            text = self.sources_box.toPlainText().strip()
            return [line for line in text.splitlines() if line.strip()]

        def selected_master(self) -> str:
            return self.master_box.toPlainText().strip()

        def run_process(self):
            sources = self.selected_sources()
            master = self.selected_master()
            if not sources:
                MessageBox("No A files", "Please select one or more A request files.", self).exec()
                return
            if not master:
                MessageBox("No B file", "Please select one B masterlist file.", self).exec()
                return

            self.log_box.clear()
            self.run_btn.setEnabled(False)
            self.select_sources_btn.setEnabled(False)
            self.select_master_btn.setEnabled(False)
            self.log_message.emit(f"Starting import for {len(sources)} A file(s)...")

            def worker():
                try:
                    output, count = process_import(sources, master, log_emit=self.log_message.emit)
                    self.processing_done.emit(count, output, "")
                except Exception as exc:
                    self.log_message.emit(f"ERROR: {exc}")
                    self.processing_done.emit(0, "", str(exc))

            threading.Thread(target=worker, daemon=True).start()

        def append_log(self, text: str):
            self.log_box.append(text)
            self.log_box.ensureCursorVisible()

        def on_processing_done(self, count: int, output: str, error: str):
            self.run_btn.setEnabled(True)
            self.select_sources_btn.setEnabled(True)
            self.select_master_btn.setEnabled(True)
            if error:
                msg = MessageBox("Import failed", error, self)
            else:
                self.log_message.emit(f"Completed: {count} row(s) imported.")
                msg = MessageBox("Import complete", f"Imported rows: {count}\nOutput: {os.path.basename(output)}", self)
            msg.yesButton.setText("OK")
            msg.cancelButton.hide()
            msg.exec()


    def get_widget():
        return MainWidget()
else:
    def get_widget():
        raise RuntimeError("PySide6 is required to load this pycro widget.")
